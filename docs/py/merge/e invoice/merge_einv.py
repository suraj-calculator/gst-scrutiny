"""
merge_einv.py
Drop this file (+ gst_merge_common.py) into a folder containing any number of
E-Invoice GSTR-1 auto-populated workbooks and run:

    python merge_einv.py

It will detect every E-Invoice file in the folder (by content, not filename),
GROUP them by Tax Period (month) - since a single month's data can now be
split across 2 (or more) files - merge each month's files together into one
continuous block, sort the months chronologically, and produce
EINV_Merged.xlsx with the same 4 data sheets (Read me sheet dropped).
Missing months are simply skipped, no error.
"""
from openpyxl import Workbook, load_workbook
from gst_merge_common import (
    find_xlsx_files, detect_file_type, einv_period_to_key,
    sheet_max_data_row, write_separator, HEADER_FONT, find_sheet,
)

DATA_SHEETS = ["b2b, sez, de", "cdnr", "cdnur", "exp"]
HEADER_ROWS = 4  # title, blank, section title, column headers


def read_meta(wb):
    ws = wb["Read me"]
    return {
        "fy": ws["C4"].value,
        "tax_period": ws["C5"].value,
        "gstin": ws["C6"].value,
        "legal_name": ws["C7"].value,
        "date_updated_till": ws["C9"].value,
    }


def main(folder="."):
    files = find_xlsx_files(folder)
    einv_files = [f for f in files if detect_file_type(f) == "EINV"]
    if not einv_files:
        print("No E-Invoice files found in this folder.")
        return

    records = []
    for f in einv_files:
        wb = load_workbook(f, data_only=True)
        meta = read_meta(wb)
        key = einv_period_to_key(meta["tax_period"])
        records.append({"path": f, "wb": wb, "meta": meta, "key": key})

    # Group by month (same Tax Period). A month can now have more than one
    # file (data split across parts) - all files sharing a key get merged
    # together, in the same month-block, in the final output.
    groups = {}
    for r in records:
        groups.setdefault(r["key"], []).append(r)

    # Stable, predictable order within a month: sort by filename.
    for key in groups:
        groups[key].sort(key=lambda r: r["path"])

    sorted_keys = sorted(groups.keys())

    print("Merge order (E-Invoice), grouped by month:")
    for k in sorted_keys:
        group = groups[k]
        tag = f" [{len(group)} files - will be merged into one block]" if len(group) > 1 else ""
        print(f"  Tax Period {group[0]['meta']['tax_period']} (FY {group[0]['meta']['fy']}){tag}")
        for r in group:
            print(f"      {r['path']}")

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    first_wb = records[0]["wb"]

    for sheet_name in DATA_SHEETS:
        first_ws = find_sheet(first_wb, sheet_name)
        n_cols = first_ws.max_column
        ws_out = wb_out.create_sheet(title=sheet_name)

        for r in range(1, HEADER_ROWS + 1):
            for c in range(1, n_cols + 1):
                src_cell = first_ws.cell(row=r, column=c)
                dst_cell = ws_out.cell(row=r, column=c, value=src_cell.value)
                if r == HEADER_ROWS:
                    dst_cell.font = HEADER_FONT

        current_row = HEADER_ROWS + 1

        for k in sorted_keys:
            group = groups[k]
            meta = group[0]["meta"]
            part_note = f"  |  ({len(group)} files merged)" if len(group) > 1 else ""
            sep_text = (
                f"Financial Year: {meta['fy']}  |  Tax Period: {meta['tax_period']}  |  "
                f"Date Updated till: {meta['date_updated_till']}{part_note}"
            )
            write_separator(ws_out, current_row, sep_text, n_cols)
            current_row += 1

            # Write every file belonging to this month, one after another,
            # under the single separator above - so the whole month reads
            # as one continuous block regardless of how many files it came from.
            for rec in group:
                ws_src = find_sheet(rec["wb"], sheet_name)
                last_row = sheet_max_data_row(ws_src, HEADER_ROWS + 1)
                for r in range(HEADER_ROWS + 1, last_row + 1):
                    for c in range(1, n_cols + 1):
                        ws_out.cell(row=current_row, column=c, value=ws_src.cell(row=r, column=c).value)
                    current_row += 1

        for col_letter, dim in first_ws.column_dimensions.items():
            if dim.width:
                ws_out.column_dimensions[col_letter].width = dim.width

    out_path = "EINV_Merged.xlsx"
    wb_out.save(out_path)
    print(f"\nSaved: {out_path}")
    print(f"Months merged: {len(sorted_keys)}")


if __name__ == "__main__":
    main()
