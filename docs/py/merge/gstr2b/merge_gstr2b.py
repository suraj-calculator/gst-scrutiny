"""
merge_gstr2b.py  (v4)
Drop this file (+ gst_merge_common.py) into a folder containing any number of
GSTR-2B workbooks - monthly, quarterly, or a mix across years - and run:

    python merge_gstr2b.py

Files are detected by content (sheet-name signature), sorted chronologically
by Tax Period (quarterly labels like "Apr-Jun" sort by their start month), and
merged into GSTR2B_Merged.xlsx.

v2: the old script assumed every GSTR-2B file has the same 11 sheets, using
the FIRST file's sheet list for all of them. In practice the GST portal's
"Summary" excel (files ending in _summary.xlsx) only contains a subset of
sheets - it drops several of the line-item detail sheets. Mixing a detailed
file and a summary file in the same folder made the old script crash with a
KeyError the moment it reached a detail sheet for a summary-only period.
This version builds the sheet list from the UNION of every input file, and
for each sheet only pulls in the records that actually contain it, printing
a note for periods that are missing it. If you actually need a detail sheet
for a period that only has a _summary.xlsx, the summary file never had that
data - you'd need to re-download the DETAILED excel for that period.

v3: the portal has also added sheets that didn't exist when this script was
first written (ECO, ECOA, IMPGA, IMPGSEZA, ITC Rejected, six "(Rejected)"
sheets - the IMS rollout). Header-row count is DETECTED per sheet from its
own merged cells (the portal merges header cells across rows for grouped
columns, but never merges data-row cells, so the highest row touched by a
merge near the top of a sheet is exactly the last header row), instead of a
hardcoded per-sheet-name list that goes stale every time the portal adds a
sheet type.

v4: read_meta() now goes through the same defensive machinery as
merge_gstr3b.py (gst_merge_common.robust_read_meta) instead of trusting
fixed C4:C9 cells outright. GSTR-3B has already shown the portal will shift
a filing template's header layout mid-year (once by row, once by column) -
the 'Read me' sheet here uses the exact same label + merged-value layout,
so it's exposed to the same risk even though it hasn't broken yet. This also
means a file whose 'Read me' can't be parsed is skipped with a diagnostic
dump instead of crashing the whole batch.

Note on 'ITC Available' / 'ITC not available' / 'ITC Rejected': these three
sheets have a fixed report layout. For a monthly filer the column headers
are the same every month, but for a quarterly filer they show the actual
month names of that period (e.g. April/May/June for a Q1 file) - since
those can change from file to file, each period's own header is kept with
its block instead of reusing a single shared header.
"""
from openpyxl import Workbook, load_workbook
from gst_merge_common import (
    find_xlsx_files, detect_file_type, fy_start_year, month_key,
    sheet_max_data_row, write_separator, HEADER_FONT, copy_sheet_full, warn_duplicates,
    robust_read_meta, looks_like_fy, looks_like_tax_period,
)

# sheets whose header (incl. month labels) must be repeated for every period
REPEAT_HEADER_SHEETS = {"ITC Available", "ITC not available", "ITC Rejected"}
TITLE_ROWS = 4  # rows 1-4 are the generic title/section banner, same for all periods

# Sheet types seen and verified as of this script version - used only to
# flag genuinely new sheet types the portal introduces later, so you get a
# heads-up rather than the tool silently doing its best guess with no notice.
KNOWN_SHEETS = {
    "B2B", "B2BA", "B2B-CDNR", "B2B-CDNRA", "ISD", "ISDA", "IMPG", "IMPGA",
    "IMPGSEZ", "IMPGSEZA", "ECO", "ECOA",
    "B2B(Rejected)", "B2BA(Rejected)", "B2B-CDNR(Rejected)", "B2B-CDNRA(Rejected)",
    "ECO(Rejected)", "ECOA(Rejected)",
    "ITC Available", "ITC not available", "ITC Rejected",
}

FIXED_CELLS = {
    "fy": "C4",
    "tax_period": "C5",
    "gstin": "C6",
    "legal_name": "C7",
    "generated_on": "C9",
}

LABEL_FALLBACKS = {
    "fy": ["Financial Year", "Year", "FY"],
    "tax_period": ["Tax Period"],
    "gstin": ["GSTIN"],
    "legal_name": ["Legal Name", "Legal name of the registered person"],
    "generated_on": ["Date of generation", "Date and Time of Generation", "Generated on"],
}

VALIDATORS = {
    "fy": looks_like_fy,
    "tax_period": looks_like_tax_period,
}


def detect_header_rows(ws, default=6, max_scan_row=10):
    """How many rows at the top of this sheet are header, not data. The
    portal merges header cells across rows for grouped columns (e.g. an
    'Invoice Details' label spanning two rows) but never merges data-row
    cells in these per-record listings - so the highest row touched by any
    merge near the top of the sheet is exactly the last header row."""
    candidates = [rng.max_row for rng in ws.merged_cells.ranges if rng.min_row <= max_scan_row]
    return max(candidates) if candidates else default


def read_meta(wb, path):
    return robust_read_meta(wb["Read me"], FIXED_CELLS, LABEL_FALLBACKS, path, VALIDATORS)


def sep_text(meta):
    return (
        f"Financial Year: {meta['fy']}  |  Tax Period: {meta['tax_period']}  |  "
        f"Date of Generation: {meta['generated_on']}"
    )


def merge_static_header_sheet(wb_out, sheet_name, records_with_sheet):
    first_ws = records_with_sheet[0]["wb"][sheet_name]
    header_rows = detect_header_rows(first_ws)
    n_cols = first_ws.max_column
    ws_out = wb_out.create_sheet(title=sheet_name[:31])

    for r in range(1, header_rows + 1):
        for c in range(1, n_cols + 1):
            src_cell = first_ws.cell(row=r, column=c)
            dst_cell = ws_out.cell(row=r, column=c, value=src_cell.value)
            if r >= header_rows - 1:
                dst_cell.font = HEADER_FONT

    current_row = header_rows + 1
    for rec in records_with_sheet:
        write_separator(ws_out, current_row, sep_text(rec["meta"]), n_cols)
        current_row += 1

        ws_src = rec["wb"][sheet_name]
        n_cols_src = ws_src.max_column
        last_row = sheet_max_data_row(ws_src, header_rows + 1)
        for r in range(header_rows + 1, last_row + 1):
            for c in range(1, n_cols_src + 1):
                ws_out.cell(row=current_row, column=c, value=ws_src.cell(row=r, column=c).value)
            current_row += 1

    for col_letter, dim in first_ws.column_dimensions.items():
        if dim.width:
            ws_out.column_dimensions[col_letter].width = dim.width


def merge_repeat_header_sheet(wb_out, sheet_name, records_with_sheet):
    first_ws = records_with_sheet[0]["wb"][sheet_name]
    n_cols = first_ws.max_column
    ws_out = wb_out.create_sheet(title=sheet_name[:31])

    for r in range(1, TITLE_ROWS + 1):
        for c in range(1, n_cols + 1):
            ws_out.cell(row=r, column=c, value=first_ws.cell(row=r, column=c).value)

    current_row = TITLE_ROWS + 1
    for rec in records_with_sheet:
        write_separator(ws_out, current_row, sep_text(rec["meta"]), n_cols)
        current_row += 1

        ws_src = rec["wb"][sheet_name]
        n_cols_src = ws_src.max_column
        for r in range(TITLE_ROWS + 1, ws_src.max_row + 1):
            for c in range(1, n_cols_src + 1):
                ws_out.cell(row=current_row, column=c, value=ws_src.cell(row=r, column=c).value)
            current_row += 1

    for col_letter, dim in first_ws.column_dimensions.items():
        if dim.width:
            ws_out.column_dimensions[col_letter].width = dim.width


def main(folder="."):
    files = find_xlsx_files(folder)
    gstr2b_files = [f for f in files if detect_file_type(f) == "GSTR2B"]
    if not gstr2b_files:
        print("No GSTR-2B files found in this folder.")
        return

    records = []
    skipped = []
    for f in gstr2b_files:
        wb = load_workbook(f, data_only=True)
        meta = read_meta(wb, f)
        if meta is None:
            skipped.append(f)
            continue
        key = (fy_start_year(meta["fy"]), month_key(meta["tax_period"]))
        records.append({"path": f, "wb": wb, "meta": meta, "key": key})

    if not records:
        print("\nNo GSTR-2B file could be read - see the dump(s) above.")
        return

    records.sort(key=lambda r: r["key"])

    print("Merge order (GSTR-2B):")
    for r in records:
        print(f"  {r['path']}  ->  FY {r['meta']['fy']}, Tax Period {r['meta']['tax_period']}")
    if skipped:
        print(f"\n{len(skipped)} file(s) skipped - see dump(s) above:")
        for f in skipped:
            print(f"  {f}")

    warn_duplicates(records)

    # Union of every sheet seen across ALL files - not just records[0]'s sheets,
    # since a _summary.xlsx period has fewer sheets than a detailed period.
    all_sheet_names = []
    for r in records:
        for name in r["wb"].sheetnames:
            if name != "Read me" and name not in all_sheet_names:
                all_sheet_names.append(name)

    new_sheets = [s for s in all_sheet_names if s not in KNOWN_SHEETS]
    if new_sheets:
        print(f"\nHeads up: {len(new_sheets)} sheet type(s) not seen before - "
              f"header rows were auto-detected from merged cells, worth a quick "
              f"look at the merged output to confirm they line up:")
        for s in new_sheets:
            print(f"  {s}")

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    copy_sheet_full(records[0]["wb"]["Read me"], wb_out, "Read me")

    for sheet_name in all_sheet_names:
        records_with_sheet = [r for r in records if sheet_name in r["wb"].sheetnames]
        records_missing = [r for r in records if sheet_name not in r["wb"].sheetnames]

        if records_missing:
            print(f"\nNote: '{sheet_name}' is missing from {len(records_missing)} file(s) "
                  f"(likely a _summary.xlsx period) - merging without them:")
            for r in records_missing:
                print(f"  {r['path']}  (Tax Period {r['meta']['tax_period']})")

        if not records_with_sheet:
            continue

        if sheet_name in REPEAT_HEADER_SHEETS:
            merge_repeat_header_sheet(wb_out, sheet_name, records_with_sheet)
        else:
            merge_static_header_sheet(wb_out, sheet_name, records_with_sheet)

    out_path = "GSTR2B_Merged.xlsx"
    wb_out.save(out_path)
    print(f"\nSaved: {out_path}")


if __name__ == "__main__":
    main()
