"""
Core utilities shared by the GSTR-9, GSTR-9C and BO-Profile PDF -> Excel converters.

Design principle (per user's explicit requirement):
  A worksheet must contain the COMPLETE data for one logical heading/section of the
  source PDF. A new heading always starts a brand-new worksheet. Sections are NEVER
  split mid-table across two worksheets (which was the bug in the earlier conversion,
  e.g. GSTR-9C item 9's table was cut between "Table 3" and "Table 4").

Sections are detected using the PDF's own structural markers (item numbers, "Pt."
labels, or known field/column-header names for the BO Profile report) rather than
page boundaries or fixed row counts -- so the split is never "random".
"""
import re
import pdfplumber
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------------
# Cell cleaning
# ----------------------------------------------------------------------------

# The source PDFs carry a diagonal watermark ("FINAL" on the GSTR forms, the
# report-generator's name/stamp on the BO Profile report). A single glyph of
# that watermark sometimes bleeds into a table cell's bounding box and gets
# extracted as a short, meaningless leading fragment + newline, e.g.
# "N\nReasons for ..." or "j\nC". This strips exactly that artifact: a leading
# run of 1-3 letters immediately followed by a newline and then the real
# content (which starts with an uppercase letter, a digit, or punctuation).
_WATERMARK_PREFIX_RE = re.compile(r'^[A-Za-z]\n(?=\S)')
_WATERMARK_DASH_RE = re.compile(r'^[A-Za-z] -$')

# Repeating page letterhead/footer used in the BO Profile report.
_LETTERHEAD_RE = re.compile(r'GSTIN:\s*\S+.*TradeName', re.S)
_FOOTER_RE = re.compile(r'This is system generated report', re.I)


def clean_cell(v, is_id_col=False):
    """Clean a single extracted PDF cell: strip watermark artifacts, collapse
    newlines/whitespace, return None for empty cells."""
    if v is None:
        return None
    s = str(v)
    stripped = _WATERMARK_PREFIX_RE.sub('', s)
    prefix_was_stripped = (stripped != s)
    s = stripped.replace('\n', ' ')
    s = re.sub(r'\s+', ' ', s).strip()
    if not s:
        return None
    if _WATERMARK_DASH_RE.match(s):  # e.g. stray "N -" -> the real value is just "-"
        s = '-'
    # A lone single letter that was NEVER part of a multi-line cell (i.e. the
    # raw PDF cell was just that one character on its own, in a non-ID
    # column) is a leftover watermark glyph rather than real data. A letter
    # that survived a prefix-strip (e.g. "j\nC" -> "C", a real account-type
    # code) is left alone.
    if not is_id_col and len(s) == 1 and s.isalpha() and not prefix_was_stripped:
        return None
    return s


def is_letterhead_or_footer(row):
    joined = ' '.join(c for c in row if c) if row else ''
    return bool(_LETTERHEAD_RE.search(joined) or _FOOTER_RE.search(joined))


def clean_row(row):
    return [clean_cell(c, is_id_col=(i == 0)) for i, c in enumerate(row)]


# ----------------------------------------------------------------------------
# PDF extraction
# ----------------------------------------------------------------------------

def extract_verification_rows(pdf_path):
    """The 'Verification of registered person' block at the end of GSTR-9 /
    GSTR-9C is free-floating text (no table border), so it never shows up via
    find_tables(). Pull it directly from the page text instead."""
    lines = []
    found = False
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ''
            for raw_line in text.split('\n'):
                line = raw_line.strip()
                if not line:
                    continue
                if line.startswith('Verification'):
                    found = True
                    lines = []  # keep the LAST occurrence only
                if found:
                    if len(line) == 1 and line in 'FINAL':
                        continue  # stray watermark glyph on its own line
                    lines.append(line)
    if not lines:
        return []

    consumed = set()
    fields = []
    date_re = re.compile(r'(\d{2}-\d{2}-\d{4})')

    for idx, line in enumerate(lines[1:], start=1):
        if idx in consumed:
            continue
        matched = False
        if line.startswith('Date:') or line.startswith('Date '):
            m = date_re.search(line)
            fields.append(('Date', m.group(1) if m else line.split(':', 1)[-1].strip()))
            matched = True
        if re.search(r'Name of (?:the )?Authori[sz]ed? Signatory', line, re.I):
            if idx + 1 < len(lines):
                fields.append(('Name of the Authorised Signatory', lines[idx + 1]))
                consumed.add(idx + 1)
            matched = True
        if line.rstrip(':').replace(' ', '') in ('Designation/Status', 'Designation'):
            if idx + 1 < len(lines):
                fields.append(('Designation/Status', lines[idx + 1]))
                consumed.add(idx + 1)
            matched = True
        if matched:
            consumed.add(idx)

    para = [l for i, l in enumerate(lines[1:], start=1) if i not in consumed]
    rows = [['Declaration', ' '.join(para)]]
    rows.extend([k, v] for k, v in fields)
    return rows


_FY_RE = re.compile(r'^\d{4}-\d{2}$')


def extract_flat_rows(pdf_path, drop_letterhead=False):
    """Extract every table row from every page, in reading order, as a flat
    list of cleaned rows. Also applies a narrow "orphan fragment" merge: when a
    table gets cut by a page break, a short leftover fragment (e.g. "(DRC-07)")
    sometimes lands alone as the first row of the next page's table. Such a row
    (blank first cell, <=2 non-empty cells) is merged back into the previous row
    rather than kept as a separate row.
    """
    rows = []  # list of ((page_no, table_idx), row)
    with pdfplumber.open(pdf_path) as pdf:
        for page_no, page in enumerate(pdf.pages):
            for table_idx, t in enumerate(page.find_tables()):
                for raw in t.extract():
                    row = clean_row(raw)
                    if all(c is None for c in row):
                        continue
                    if drop_letterhead and is_letterhead_or_footer(raw):
                        continue
                    if row[0] is None and len(row) > 1 and row[1] and _FY_RE.match(row[1]):
                        row = row[1:]
                    non_empty = sum(1 for c in row if c)
                    if rows and (not row[0]) and non_empty <= 2:
                        # merge fragment into previous row
                        prev = rows[-1][1]
                        for i, c in enumerate(row):
                            if c and i < len(prev):
                                prev[i] = (prev[i] + ' ' + c).strip() if prev[i] else c
                        continue
                    rows.append(((page_no, table_idx), row))

    # A running page header, or occasionally a whole short table, sometimes
    # repeats verbatim - drop that repeat. The one case where an identical
    # block is NOT a repeat-artifact is two distinct tables sitting
    # side-by-side on the same page (e.g. a GSTR-3B mini-table next to an
    # identically-shaped GSTR-1 mini-table): same page, different table
    # object. Everything else that repeats a prior block verbatim - whether
    # across a page break or within the same table - is a duplicate.
    deduped = []
    i = 0
    while i < len(rows):
        skipped = False
        cur_key = rows[i][0]
        for block_len in (3, 2, 1):
            if len(deduped) >= block_len and i + block_len <= len(rows):
                block = [r for _, r in rows[i:i + block_len]]
                prev_block = [r for _, r in deduped[-block_len:]]
                prev_keys = {k for k, _ in deduped[-block_len:]}
                same_page_diff_table = (
                    len(prev_keys) == 1
                    and next(iter(prev_keys))[0] == cur_key[0]
                    and next(iter(prev_keys))[1] != cur_key[1]
                )
                if block == prev_block and not same_page_diff_table:
                    i += block_len
                    skipped = True
                    break
        if not skipped:
            deduped.append(rows[i])
            i += 1
    return [r for _, r in deduped]


# ----------------------------------------------------------------------------
# Numeric formatting helpers
# ----------------------------------------------------------------------------

_NUM_RE = re.compile(r'^-?[\d,]+\.?\d*$')
_PCT_RE = re.compile(r'^-?[\d,]+\.?\d*%$')

INDIAN_NUM_FMT = '#,##,##0.00;[RED]-#,##,##0.00;"-"'
INDIAN_INT_FMT = '#,##,##0;[RED]-#,##,##0;"-"'
PCT_FMT = '0.00%'


def to_cell_value(text):
    """Return (value, number_format_or_None) for a cleaned text cell.

    Long bare digit strings with no comma/decimal (phone numbers, VAT TINs,
    YYYYMM period codes, account numbers) are identifiers, not amounts - keep
    them as text so leading zeros and exact digit strings survive. Genuine
    monetary/count figures (which in this report always carry a comma or a
    decimal point, or are short counts) get converted to real numbers so the
    sheet is usable for calculations.
    """
    if text is None:
        return None, None
    t = text.strip()
    if t in ('-', ''):
        return None, None
    if _PCT_RE.match(t):
        try:
            return float(t.replace(',', '').rstrip('%')) / 100.0, PCT_FMT
        except ValueError:
            return t, None
    if _NUM_RE.match(t) and any(ch.isdigit() for ch in t):
        has_separator = (',' in t) or ('.' in t)
        bare_digits = t.replace(',', '').replace('.', '').replace('-', '')
        if has_separator or len(bare_digits) <= 5:
            try:
                val = float(t.replace(',', ''))
                fmt = INDIAN_NUM_FMT if '.' in t else INDIAN_INT_FMT
                return val, fmt
            except ValueError:
                return t, None
        return t, None  # long bare-digit string -> identifier, keep as text
    return t, None


# ----------------------------------------------------------------------------
# Excel writing
# ----------------------------------------------------------------------------

TITLE_FILL = PatternFill('solid', fgColor='1F4E5F')
TITLE_FONT = Font(name='Arial', bold=True, size=12, color='FFFFFF')
NOTE_FONT = Font(name='Arial', italic=True, size=9, color='595959')
HEADER_FILL = PatternFill('solid', fgColor='D9E6EC')
HEADER_FONT = Font(name='Arial', bold=True, size=10)
BODY_FONT = Font(name='Arial', size=10)
ID_FONT = Font(name='Arial', bold=True, size=10)
THIN = Side(style='thin', color='B7C6CC')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

_INVALID_SHEETCHARS = re.compile(r'[:\\/?*\[\]]')


def safe_sheet_name(name, used):
    n = _INVALID_SHEETCHARS.sub('-', name)[:31].strip()
    if not n:
        n = 'Sheet'
    base = n
    i = 2
    while n in used:
        suffix = f' ({i})'
        n = (base[: 31 - len(suffix)] + suffix)
        i += 1
    used.add(n)
    return n


def _looks_like_header_row(row):
    """Heuristic: a row of short, label-ish text with no numeric data -> treat
    as a bold column-header row (e.g. 'Sr. No | Description | ... | Amount')."""
    texts = [c for c in row if c]
    if not texts:
        return False
    numericish = sum(1 for c in texts if _NUM_RE.match(c) or _PCT_RE.match(c))
    return numericish == 0 and len(texts) >= 2


def write_section(wb, used_names, title, note, rows, first_col_is_id=True):
    """Write one logical section as its own worksheet."""
    ws = wb.create_sheet(safe_sheet_name(title, used_names))
    max_cols = max((len(r) for r in rows), default=2)
    max_cols = max(max_cols, 2)

    r = 1
    ws.cell(row=r, column=1, value=title)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_cols)
    c = ws.cell(row=r, column=1)
    c.font = TITLE_FONT
    c.fill = TITLE_FILL
    c.alignment = Alignment(vertical='center', horizontal='left', indent=1)
    ws.row_dimensions[r].height = 22
    r += 1

    if note:
        ws.cell(row=r, column=1, value=note)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_cols)
        nc = ws.cell(row=r, column=1)
        nc.font = NOTE_FONT
        nc.alignment = Alignment(horizontal='left', indent=1)
        r += 1

    r += 1  # blank spacer row
    header_row_idx = None
    col_widths = [12] * max_cols

    for row in rows:
        padded = list(row) + [None] * (max_cols - len(row))
        is_header = _looks_like_header_row(padded)
        for col_idx, cell_text in enumerate(padded, start=1):
            value, numfmt = to_cell_value(cell_text)
            cell = ws.cell(row=r, column=col_idx, value=value)
            cell.border = BORDER
            if is_header:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                cell.font = ID_FONT if (col_idx == 1 and first_col_is_id) else BODY_FONT
                if numfmt:
                    cell.number_format = numfmt
                    cell.alignment = Alignment(horizontal='right', vertical='top')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            text_len = len(str(value)) if value is not None else 0
            col_widths[col_idx - 1] = min(60, max(col_widths[col_idx - 1], text_len + 2))
        if is_header and header_row_idx is None:
            header_row_idx = r
        r += 1

    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(10, min(w, 55))

    ws.freeze_panes = f'A{(header_row_idx or 4) + 1}'
    ws.sheet_view.showGridLines = False
    return ws
