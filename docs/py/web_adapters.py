"""
Web adapters — the only new Python in this project's web port.

Every function here bridges "bytes the browser handed us" to the existing,
unmodified scripts in ../ewb, ../merge, ../pdf, ../extract_align, ../core —
by materialising those bytes into a real folder (in Pyodide's virtual
filesystem in the browser, or a plain temp folder under CPython for local
testing), calling the existing script's functions directly, and reading
whatever they wrote back out as bytes to hand back.

Every function here is written to run unchanged under plain CPython (for
fast local testing — see test_web_adapters.py) and under Pyodide in the
browser (see app.js). Nothing here is Pyodide-specific.
"""

import glob
import importlib
import os
import sys

_HERE = os.path.dirname(os.path.abspath(__file__))
# Only "ewb" and "core" are safe to leave permanently on sys.path: every
# module name in them is unique. "merge" is deliberately NOT added here —
# its 5 subfolders (gstr1/gstr2a/gstr2b/gstr3b/e invoice) each ship their
# own same-named gst_merge_common.py with real content differences between
# copies (see forms merger/README.md). Adding more than one of those
# folders to sys.path at once would let the first-imported copy silently
# shadow the others for the rest of the session. _use_merge_subfolder()
# below adds exactly one at a time and clears the import cache around it.
for _sub in ("ewb", "core"):
    _p = os.path.join(_HERE, _sub)
    if _p not in sys.path:
        sys.path.insert(0, _p)
for _sub in ("extract_align/extractor", "extract_align/alligner"):
    _p = os.path.join(_HERE, *_sub.split("/"))
    if _p not in sys.path:
        sys.path.insert(0, _p)

_MERGE_ROOT = os.path.join(_HERE, "merge")


def _read_bytes(path):
    with open(path, "rb") as f:
        return f.read()


def _use_merge_subfolder(subfolder, module_name):
    """
    Make exactly one forms-merger subfolder importable, guaranteeing a
    fresh import of both `module_name` and its sibling `gst_merge_common`
    even if a different subfolder's copies of those same two module names
    are still cached in sys.modules from an earlier call this session.
    Returns the imported module.
    """
    for mod in ("gst_merge_common", module_name):
        sys.modules.pop(mod, None)
    for p in list(sys.path):
        if p == _MERGE_ROOT or os.path.dirname(p) == _MERGE_ROOT:
            sys.path.remove(p)
    sys.path.insert(0, os.path.join(_MERGE_ROOT, subfolder))
    return importlib.import_module(module_name)


def process_ewb(inward_files, outward_files, work_dir):
    """
    inward_files / outward_files: list of (filename, bytes) tuples — the raw
    EWB_MIS_Report_Excel (N).xls downloads, one direction each.
    work_dir: an empty folder to do the work in (caller creates/cleans it).

    Returns a dict:
        {
          "inward":  {"rows": int, "output_name": str, "output_bytes": bytes} | None,
          "outward": {"rows": int, "output_name": str, "output_bytes": bytes} | None,
          "log": [str, ...],
        }
    "inward"/"outward" are None if that direction had no files.
    """
    import convert_ewb_files
    import auto_ewb_merger

    log = []
    inward_dir = os.path.join(work_dir, auto_ewb_merger.INWARD_FOLDER)
    outward_dir = os.path.join(work_dir, auto_ewb_merger.OUTWARD_FOLDER)
    os.makedirs(inward_dir, exist_ok=True)
    os.makedirs(outward_dir, exist_ok=True)

    for name, data in inward_files:
        with open(os.path.join(inward_dir, name), "wb") as f:
            f.write(data)
    for name, data in outward_files:
        with open(os.path.join(outward_dir, name), "wb") as f:
            f.write(data)

    prev_cwd = os.getcwd()
    os.chdir(work_dir)
    try:
        # Convert every raw .xls in both folders to real .xlsx first
        # (mirrors convert_ewb_files.main(), minus its input()/print noise).
        for folder in (auto_ewb_merger.INWARD_FOLDER, auto_ewb_merger.OUTWARD_FOLDER):
            for path in glob.glob(os.path.join(folder, "*.xls")):
                ok = convert_ewb_files.convert_file(path)
                log.append(f"convert {os.path.basename(path)}: {'ok' if ok else 'FAILED'}")

        auto_ewb_merger.merge_all_files()

        result = {"inward": None, "outward": None, "log": log}
        merge_dir = os.path.join(work_dir, auto_ewb_merger.MERGE_FOLDER)

        inward_out = os.path.join(merge_dir, "inward_eway_bill_merged.xlsx")
        if os.path.exists(inward_out):
            import openpyxl
            wb = openpyxl.load_workbook(inward_out, read_only=True)
            rows = wb.active.max_row - 1  # minus header
            wb.close()
            result["inward"] = {
                "rows": rows,
                "output_name": "inward_eway_bill_merged.xlsx",
                "output_bytes": _read_bytes(inward_out),
            }

        outward_out = os.path.join(merge_dir, "outward_eway_bill_merged.xlsx")
        if os.path.exists(outward_out):
            import openpyxl
            wb = openpyxl.load_workbook(outward_out, read_only=True)
            rows = wb.active.max_row - 1
            wb.close()
            result["outward"] = {
                "rows": rows,
                "output_name": "outward_eway_bill_merged.xlsx",
                "output_bytes": _read_bytes(outward_out),
            }

        return result
    finally:
        os.chdir(prev_cwd)


# kind -> (subfolder under docs/py/merge, module name, output filename)
MERGE_KINDS = {
    "gstr1": ("gstr1", "merge_gstr1", "GSTR1_Merged.xlsx"),
    "gstr2a": ("gstr2a", "merge_r2a", "R2A_Merged.xlsx"),
    "gstr2b": ("gstr2b", "merge_gstr2b", "GSTR2B_Merged.xlsx"),
    "gstr3b": ("gstr3b", "merge_gstr3b", "GSTR3B_Merged.xlsx"),
    "einv": ("e invoice", "merge_einv", "EINV_Merged.xlsx"),
}


def process_merge(kind, files, work_dir):
    """
    Runs one of the 5 forms-merger scripts (gstr1/gstr2a/gstr2b/gstr3b/einv)
    against a batch of uploaded per-period workbooks.

    files: list of (filename, bytes) — the raw per-period .xlsx exports.
    work_dir: an empty folder to do the work in (caller creates/cleans it).

    Returns {"output_name": str, "output_bytes": bytes} or None if the
    script found no files of its own type among what was uploaded (each
    script detects its type by workbook content, not filename, so this can
    happen even with files present if none of them actually match).
    """
    if kind not in MERGE_KINDS:
        raise ValueError(f"unknown merge kind: {kind!r}")
    subfolder, module_name, out_name = MERGE_KINDS[kind]
    module = _use_merge_subfolder(subfolder, module_name)

    os.makedirs(work_dir, exist_ok=True)
    for name, data in files:
        with open(os.path.join(work_dir, name), "wb") as f:
            f.write(data)

    prev_cwd = os.getcwd()
    os.chdir(work_dir)
    try:
        module.main(".")
        if not os.path.exists(out_name):
            return None
        return {"output_name": out_name, "output_bytes": _read_bytes(out_name)}
    finally:
        os.chdir(prev_cwd)


def process_gstr3b(files, work_dir):
    """
    GSTR-3B's real pipeline is two steps: extract the Excel files bundled
    inside each portal .zip download, then merge them. Accepts a mix of
    GSTR3B_<GSTIN>_<MMYYYY>.zip bundles and/or already-extracted .xlsx
    files (useful for local testing against fixtures either way).
    """
    import run as extractor_run  # extract_align/extractor/run.py

    zip_dir = os.path.join(work_dir, "zips")
    os.makedirs(zip_dir, exist_ok=True)
    direct_xlsx = []
    for name, data in files:
        if name.lower().endswith(".zip"):
            with open(os.path.join(zip_dir, name), "wb") as f:
                f.write(data)
        else:
            direct_xlsx.append((name, data))

    extractor_run.extract_excel_from_zips(zip_dir)
    extracted_dir = os.path.join(zip_dir, extractor_run.DEST_FOLDER_NAME)
    extracted = []
    if os.path.isdir(extracted_dir):
        for fname in os.listdir(extracted_dir):
            extracted.append((fname, _read_bytes(os.path.join(extracted_dir, fname))))

    all_xlsx = direct_xlsx + extracted
    return process_merge("gstr3b", all_xlsx, os.path.join(work_dir, "merge"))


def process_gstr2b(files, work_dir):
    """
    GSTR-2B's real pipeline is two steps: align every uploaded workbook to
    the same set of worksheets (files out/alligner, mutates in place), then
    merge (forms merger/gstr2b).
    """
    import complete_workbooks  # extract_align/alligner/complete_workbooks.py

    os.makedirs(work_dir, exist_ok=True)
    paths = []
    for name, data in files:
        p = os.path.join(work_dir, name)
        with open(p, "wb") as f:
            f.write(data)
        paths.append(p)

    master_order, reference_path = complete_workbooks.get_master_sheet_order(paths, None)
    for p in paths:
        complete_workbooks.complete_workbook(
            p, master_order, reference_path, complete_workbooks.DEFAULT_COPY_SHEETS
        )

    aligned = [(os.path.basename(p), _read_bytes(p)) for p in paths]
    return process_merge("gstr2b", aligned, os.path.join(work_dir, "merge"))


def _render_bs_pl_module(bs_pl_data):
    """
    master_build.py does a plain `import bs_pl_input` and reads its
    BS_PL_DATA module attribute (see main gst tool/bs_pl_input.py's own
    docstring) — there's no function call to hand data to directly. This
    renders a real bs_pl_input.py source file from the dict the browser's
    form collected, so that import picks it up. bs_pl_data values must
    already be numbers/None, not strings — the caller (app.js) is
    responsible for parsing the form's text inputs first.
    """
    import json

    lines = ["BS_PL_DATA = " + json.dumps(bs_pl_data, indent=4)]
    # JSON has no None-as-bare-identifier issue (json.dumps already emits
    # `null`), but Python needs `None`, not `null` — swap it back.
    return "\n".join(lines).replace("null", "None") + "\n"


def process_full_scrutiny(files, bs_pl_data, work_dir):
    """
    The final step: assemble every merged/collected file into one folder
    and run master_build.py's actual scrutiny engine against it, exactly
    as the CLI tool would if you dropped these files in a folder yourself.

    files: list of (filename, bytes) — every *_Merged.xlsx / annual EWB /
        ledger CSV / portal export / reference master collected so far.
        Only GSTR-1 and GSTR-3B are genuinely mandatory; master_build.py
        itself raises a clear RuntimeError naming what's missing if not.
    bs_pl_data: dict shaped like BS_PL_DATA (see bs_pl_input.py), or None/
        empty to skip the Balance Sheet/P&L checks (R0-R12) entirely.
    work_dir: an empty folder to do the work in (caller creates/cleans it).

    Returns {"output_name": str, "output_bytes": bytes, "log": str} — log
    is everything master_build.main() printed, for an honest run summary
    (months covered, finding counts, etc.) in the UI, not a fabricated one.
    Raises whatever master_build.main() raises (e.g. missing mandatory
    inputs) — the caller should catch and surface that message as-is, since
    it's already written to be a clear, actionable explanation.
    """
    import contextlib
    import io

    os.makedirs(work_dir, exist_ok=True)
    for name, data in files:
        with open(os.path.join(work_dir, name), "wb") as f:
            f.write(data)

    core_dir = os.path.join(_HERE, "core")
    bs_pl_path = os.path.join(core_dir, "bs_pl_input.py")
    sys.modules.pop("bs_pl_input", None)
    sys.modules.pop("master_build", None)
    if bs_pl_data:
        with open(bs_pl_path, "w") as f:
            f.write(_render_bs_pl_module(bs_pl_data))
    elif not os.path.exists(bs_pl_path):
        # No form data supplied and no template present in this vendored
        # copy: write an empty one so `import bs_pl_input` still succeeds
        # (master_build.py handles an empty/missing BS_PL_DATA gracefully —
        # it just skips R0-R12 — but the import itself must not fail).
        with open(bs_pl_path, "w") as f:
            f.write("BS_PL_DATA = {}\n")

    import master_build

    prev_cwd = os.getcwd()
    os.chdir(work_dir)
    log_buf = io.StringIO()
    try:
        with contextlib.redirect_stdout(log_buf):
            outfile = master_build.main(".")
        return {
            "output_name": outfile,
            "output_bytes": _read_bytes(outfile),
            "log": log_buf.getvalue(),
        }
    finally:
        os.chdir(prev_cwd)


def process_pdf_export(xlsx_bytes, work_dir):
    """Renders the already-built GST_MASTER workbook (see
    process_full_scrutiny's output_bytes) to PDF -- the "Download PDF"
    button's backend. Re-runs no scrutiny logic, just formats the finished
    result. See gst_report_pdf.py for the actual rendering.

    Returns {"output_bytes": bytes}.
    """
    import gst_report_pdf
    return {"output_bytes": gst_report_pdf.render_workbook_pdf(xlsx_bytes, work_dir)}
