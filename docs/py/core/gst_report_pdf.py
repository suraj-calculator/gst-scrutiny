#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Renders an already-built GST_MASTER workbook to a PDF -- one landscape
page-set per sheet, columns truncated to a readable width. This is the
"Download PDF" button's backend: it takes the exact xlsx bytes
master_build.py already produced and does NOT re-run any scrutiny logic.

fpdf2 is a pure-Python PDF library (no C extensions) -- confirmed installable
and usable both under plain CPython and under Pyodide/WASM in the browser via
micropip. A feasibility spike against a real 92-sheet, 127k-row workbook
rendered every sheet in ~1 minute of native CPython wall-clock; budget
several minutes in-browser (WASM, single-threaded, no JIT).

A real Unicode font is required: fpdf2's built-in core fonts (helvetica etc.)
are Latin-1 only and raise on the very first Rupee sign (U+20B9), which is
throughout this tool's output. DejaVu Sans (bundled in assets/, public-domain
licensed) is the one confirmed working choice.
"""
import os
from fpdf import FPDF
import openpyxl

FONT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "DejaVuSans.ttf")
MAX_COLS = 20     # sheets wider than this have extra columns dropped, not wrapped -- keeps cells readable
CELL_CHARS = 28    # per-cell text truncation
ROW_H = 4
FONT_SIZE = 6
TITLE_SIZE = 12

# DejaVu Sans (like fpdf2's other bundled/common fonts) has no colour-emoji
# glyphs -- the severity markers this tool writes into QA Summary / Action
# Required / Reviewed Master Dashboard (confirmed: 4971 cells in the latter
# alone on a real run) would otherwise render as blank/missing-glyph boxes.
# Swapped for a plain-text equivalent instead, so the PDF stays readable.
_EMOJI_MAP = {
    "\U0001F7E2": "[OK]",        # green circle
    "\U0001F534": "[FLAG]",      # red circle
    "\U0001F7E1": "[REVIEW]",    # yellow circle
    "\U0001F7E0": "[WARN]",      # orange circle
    "\U0001F535": "[INFO]",      # blue circle
    "\U0001F6A8": "[CRITICAL]",  # siren
}


def _sanitize(text):
    for glyph, tag in _EMOJI_MAP.items():
        text = text.replace(glyph, tag)
    # Anything else outside the Basic Multilingual Plane (further emoji, etc.)
    # or otherwise unsupported by the bundled font -- dropped rather than
    # left to fpdf2's own missing-glyph fallback, which renders inconsistently.
    # Also drops C0/C1 control characters (0x00-0x1F, 0x7F-0x9F, keeping
    # plain space) -- confirmed real stray bytes in this tool's own text
    # fields (mis-decoded curly-quote/euro-sign artifacts), never meaningful
    # in a table cell.
    return "".join(
        c for c in text
        if ord(c) <= 0xFFFF and not (ord(c) < 0x20 or 0x7F <= ord(c) <= 0x9F)
    )


def _render_sheet(pdf, ws, title):
    pdf.add_page(orientation="L")
    pdf.set_font("dejavu", "", TITLE_SIZE)
    pdf.cell(0, 8, _sanitize(title), new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("dejavu", "", FONT_SIZE)
    col_count = min(ws.max_column or 1, MAX_COLS)
    col_w = (pdf.w - 2 * pdf.l_margin) / col_count
    rows = 0
    for row in ws.iter_rows(values_only=True):
        for val in row[:MAX_COLS]:
            text = "" if val is None else _sanitize(str(val))[:CELL_CHARS]
            pdf.cell(col_w, ROW_H, text, border=0)
        pdf.ln(ROW_H)
        rows += 1
        if pdf.get_y() > pdf.h - 15:
            pdf.add_page(orientation="L")
    return rows


def render_workbook_pdf(xlsx_bytes, work_dir):
    """xlsx_bytes: the already-built GST_MASTER workbook, as bytes.
    work_dir: scratch folder (caller creates/cleans it) -- openpyxl's
    read_only mode needs a real path for a workbook this size, not an
    in-memory buffer.
    Returns the finished PDF as bytes."""
    os.makedirs(work_dir, exist_ok=True)
    src_path = os.path.join(work_dir, "_source.xlsx")
    with open(src_path, "wb") as f:
        f.write(xlsx_bytes)

    wb = openpyxl.load_workbook(src_path, read_only=True, data_only=True)
    pdf = FPDF()
    pdf.add_font("dejavu", "", FONT_PATH)
    for name in wb.sheetnames:
        _render_sheet(pdf, wb[name], name)
    wb.close()

    out_path = os.path.join(work_dir, "_out.pdf")
    pdf.output(out_path)
    with open(out_path, "rb") as f:
        return f.read()
