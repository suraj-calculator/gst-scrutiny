#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GST CHECKS HSN FRAUD
====================
CONSOLIDATED FILE -- contains what used to be: hsn_fraud_checks.py

The tool was reorganised from 19 .py files into 9 for easier sharing. Nothing
in the analytical logic was rewritten during that move: each section below is
the original module's code verbatim, with only (a) intra-project imports
repointed at the new file names, (b) its standalone __main__ demo block
removed, and (c) the renames listed under MERGE NOTES applied where two merged
modules happened to define the same top-level name with different bodies.

MERGE NOTES for this file:
  (none -- no name collisions)
"""


# ============================================================================
# ==== SECTION: hsn_fraud_checks.py  (was a standalone module before consolidation)
# ============================================================================
"""
HSN-CODE-WISE + FRAUD-PATTERN CHECKS
=====================================
Adds the check list the user pasted (categories A/B/C = HSN-only /
POS-state-code-only / combined, plus the 57-item numbered fraud-pattern
list) on top of the existing engines. This is a FY-WIDE, consolidated
module (like build_annual_workbook.py's sheets) rather than a per-month
one -- most of these patterns are naturally viewed as "all findings for
the year in one place", and several genuinely need cross-month data
(HSN drift, ITC volatility, year-end dumping) that a single month can't
show on its own.

---------------------------------------------------------------------
GROUNDING -- what was actually verified against the real files before
writing a single check (see chat for the full investigation):
---------------------------------------------------------------------
  - HSN code appears at invoice-line level ONLY in the EWB files. GSTR-1
    b2b/cdnr, E-Invoice, and GSTR-2B B2B all carry rate/tax but NOT HSN.
    GSTR-1's own 'hsn' sheet has HSN+rate+taxable+tax, but only as a
    MONTHLY AGGREGATE, not linked to any specific invoice.
  - GSTR-1's Place-Of-Supply is a plain state NAME ('Punjab'), not a
    coded value -- needs the STATE_NAME_TO_CODE map below to compare
    against a GSTIN's state-code prefix.
  - The 'hsn' sheet's Total-Quantity column is 0 on every row inspected
    -- per-unit price checks (avg purchase vs sale price, price bands,
    MoM price drift) are NOT computable from this data and are not
    attempted here.
  - GSTR-1 b2cs (B2C Small, Table 7) is a STATE+RATE SUMMARY with NO
    invoice numbers at all -- this is how the return itself works, not
    a parsing gap, so "B2C invoice splitting" (#13) cannot be detected
    from GSTR-1 and is not attempted.
  - EWB files have no cancellation-status column, no validity/distance
    column, no vehicle-capacity reference -- checks #14, #28, #40 are
    not attempted for the same reason.
  - GSTR-3B DOES carry, per month: 4A(5) All other ITC, B(1)/(2) reversal
    (Rule 42/43 + others), D(1) ineligible-17(5), 3.1(d) RCM liability,
    and its own ARN + Date of ARN (filing date) -- all used below.
  - cdnr (credit notes) has no "original invoice number" field, so CN
    timing checks are matched by recipient-GSTIN + nearest-value only --
    labelled approximate, never asserted as a proven link.

---------------------------------------------------------------------
DESIGN NOTE -- this module deliberately does NOT follow the "no safety
net" hard-fail rule the rest of the merged-file parsers use. These are
exploratory/heuristic add-on checks, not core reconciliation arithmetic;
a missing/odd field here degrades that ONE check to an INFO note rather
than aborting the whole master build.
---------------------------------------------------------------------

CURATED REFERENCE DATA (see HSN_RATE_MASTER etc. below): built from the
HSN codes that actually appear in this taxpayer's real 'hsn' sheet across
FY2022-23, using the standard GST rate schedule as commonly understood
for those codes. This is a DEFAULT for a scrutiny working-file, not a
verified extract of the CBIC notification in force on each invoice date
-- treat every finding this produces as "verify", not "confirmed", and
swap in an official master if/when you have one (see HSN_RATE_MASTER's
docstring for exactly where to edit).
"""

import os
import re
import datetime as _dt
import statistics as _stats
import openpyxl

import gst_core as mpu
import gst_checks_monthly as ana
import gst_parsers_returns as ewbp
import gst_parsers_returns as gstr2b_parser

PASS = "PASS"
INFO = "INFO"
FLAG = "FLAG"
REVW = "REVIEW"


class Finding:
    __slots__ = ("ref", "title", "severity", "detail", "numbers")
    def __init__(self, ref, title, severity, detail, numbers=None):
        self.ref = ref
        self.title = title
        self.severity = severity
        self.detail = detail
        self.numbers = numbers or {}


MONTH_ORDER = ["Apr-22", "May-22", "Jun-22", "Jul-22", "Aug-22", "Sep-22",
               "Oct-22", "Nov-22", "Dec-22", "Jan-23", "Feb-23", "Mar-23"]
MONTH_IDX = {m: i for i, m in enumerate(MONTH_ORDER)}


# ======================================================================
# CURATED REFERENCE DATA -- edit here if you get an official master
# ======================================================================

# ======================================================================
# CURATED REFERENCE DATA -- DATE-VERSIONED
# ======================================================================
# GST rates are NOT static -- a single "current" snapshot is actively WRONG
# for scrutinizing a past financial year. Confirmed concretely: GST 2.0
# (Notification 9/2025-CT(Rate), effective 22-Sep-2025, superseding
# Notification 01/2017-CT(Rate)) merged the 12% slab into 18%, added a new
# 40% peak rate, and moved most medicaments to 5%/Nil. A taxpayer's FY22-23
# invoices must be checked against the PRE-22-Sep-2025 rate, not whatever is
# "current" today -- this table stores BOTH, keyed by date range, so the
# check below always selects the rate that was actually in force on the
# invoice's own month.
#
# Each code maps to a list of (from_date, to_date_or_None, rate_or_None,
# desc, confidence, source_note) tuples. `to_date=None` means "still the
# latest period this table knows about" (see HSN_RATE_HISTORY_LAST_REVIEWED)
# -- NOT "in force forever, guaranteed current". `rate=None` with
# confidence="unconfirmed" means the period is KNOWN to exist (GST 2.0
# happened) but the exact new rate for this specific code was NOT
# confidently verified during research -- the check below treats this as
# "cannot compare, not guessed", never silently falls back to the prior
# period's rate.
#
# PRE-22-Sep-2025 rows: all "high" confidence, taken from this taxpayer's
# own real, filed FY22-23 GSTR-1 data (the same curation that was here
# before this table existed as HSN_RATE_MASTER).
# POST-22-Sep-2025 rows: researched via web search this session; confidence
# graded per-code based on how consistent/specific the sources were.
HSN_RATE_HISTORY = {
    "3003": [
        (_dt.date(2017, 7, 1), _dt.date(2025, 9, 21), 12.0,
         "Medicaments (other than 3004), not put up for retail sale", "high",
         "Notification 01/2017-CT(Rate) Schedule II -- taxpayer's own FY22-23 filed data."),
        (_dt.date(2025, 9, 22), None, None,
         "Medicaments (other than 3004) -- post-GST-2.0 rate is genuinely product-dependent "
         "(life-saving formulations -> Nil, most others -> 5%, some specialised items -> 18%)",
         "unconfirmed",
         "GST 2.0 moved most medicaments to 5%/Nil, but the exact sub-heading classification "
         "determines which -- NOT auto-applied here; verify the specific product against "
         "Notification 9/2025-CT(Rate) before treating a rate difference as a finding."),
    ],
    "3004": [
        (_dt.date(2017, 7, 1), _dt.date(2025, 9, 21), 12.0,
         "Medicaments put up in measured doses / for retail sale", "high",
         "Notification 01/2017-CT(Rate) Schedule II -- taxpayer's own FY22-23 filed data."),
        (_dt.date(2025, 9, 22), None, None,
         "Medicaments -- post-GST-2.0 rate is genuinely product-dependent (life-saving -> Nil, "
         "most others -> 5%, some specialised items -> 18%)", "unconfirmed",
         "Same situation as 3003 -- verify the specific product against Notification "
         "9/2025-CT(Rate) before treating a rate difference as a finding."),
    ],
    "3808": [
        (_dt.date(2017, 7, 1), None, 18.0,
         "Insecticides, fungicides, herbicides, disinfectants", "high",
         "Confirmed UNCHANGED across GST 2.0 -- multiple independent sources agree this stayed "
         "at 18% both before and after 22-Sep-2025."),
    ],
    "3915": [
        (_dt.date(2017, 7, 1), _dt.date(2025, 9, 21), 5.0,
         "Waste, parings and scrap of plastics", "high",
         "Taxpayer's own FY22-23 filed data. NOTE -- DISCREPANCY: several generic web sources "
         "describe plastic scrap (3915) as 18% both before and after GST 2.0, which conflicts "
         "with this taxpayer's own actually-filed 5%. Kept at the taxpayer-verified figure "
         "(their real, government-accepted return) rather than overwritten by unverified web "
         "content, but this conflict is NOT resolved -- if you have the actual product "
         "description/sub-heading, confirm which is correct."),
        (_dt.date(2025, 9, 22), None, None,
         "Waste, parings and scrap of plastics -- post-GST-2.0 rate not confirmed, given the "
         "unresolved pre-existing discrepancy noted above", "unconfirmed",
         "Needs manual verification against Notification 9/2025-CT(Rate) before use."),
    ],
    "4707": [
        (_dt.date(2017, 7, 1), _dt.date(2025, 9, 21), 5.0,
         "Recovered (waste and scrap) paper or paperboard", "high",
         "Taxpayer's own FY22-23 filed data."),
        (_dt.date(2025, 9, 22), None, 5.0,
         "Recovered (waste and scrap) paper or paperboard", "medium",
         "Multiple sources confirm paper/scrap stayed in the 5% band post-GST-2.0 (some "
         "describe a reduction from 12%, which doesn't match this taxpayer's own pre-2.0 5% -- "
         "likely a different sub-heading within 4707; kept at 5% with MEDIUM, not high, "
         "confidence -- verify before relying on this for a specific finding)."),
    ],
    "7204": [
        (_dt.date(2017, 7, 1), None, 18.0, "Ferrous waste and scrap", "high",
         "Confirmed UNCHANGED across GST 2.0 -- multiple independent, mutually consistent "
         "sources (iron/steel scrap explicitly called out as staying in the standard 18% "
         "slab, not treated as an essential/concessional item)."),
    ],
    "7606": [
        (_dt.date(2017, 7, 1), _dt.date(2025, 9, 21), 18.0,
         "Aluminium plates, sheets, strip", "high", "Taxpayer's own FY22-23 filed data."),
        (_dt.date(2025, 9, 22), None, None,
         "Aluminium plates, sheets, strip -- post-GST-2.0 rate NOT researched (web searches "
         "this session surfaced HSN 7602 aluminium SCRAP/waste, a different code from this "
         "taxpayer's 7606 finished-product code -- do not assume they moved together)",
         "unconfirmed", "Needs manual verification against Notification 9/2025-CT(Rate)."),
    ],
    "8402": [
        (_dt.date(2017, 7, 1), _dt.date(2025, 9, 21), 18.0,
         "Steam or vapour generating boilers", "high", "Taxpayer's own FY22-23 filed data."),
        (_dt.date(2025, 9, 22), None, None,
         "Steam or vapour generating boilers -- post-GST-2.0 rate NOT researched this session",
         "unconfirmed", "Needs manual verification against Notification 9/2025-CT(Rate)."),
    ],
    "998843": [
        (_dt.date(2017, 7, 1), _dt.date(2025, 9, 21), 12.0,
         "Job-work services -- manufacture of pharmaceutical products", "high",
         "Taxpayer's own FY22-23 filed data."),
        (_dt.date(2025, 9, 22), None, None,
         "Job-work services -- manufacture of pharmaceutical products -- post-GST-2.0 rate NOT "
         "researched. The general '12% slab merged into 18%' GST 2.0 mechanic suggests 18%, but "
         "pharma job-work may carry its own concessional entry -- not confirmed either way.",
         "unconfirmed", "Needs manual verification against Notification 9/2025-CT(Rate)/SAC schedule."),
    ],
}

# The date this table's post-GST-2.0 entries were last researched/reviewed --
# NOT a guarantee those entries are still current beyond this date. Update
# this (and re-verify the "unconfirmed" rows) periodically, or whenever a
# new GST Council rate notification is announced.
HSN_RATE_HISTORY_LAST_REVIEWED = _dt.date(2026, 7, 12)

# Backward-compatible flat view (pre-22-Sep-2025 rates only) -- some older
# code/tests may still reference HSN_RATE_MASTER directly by that name.
HSN_RATE_MASTER = {
    code: (periods[0][2], periods[0][3]) for code, periods in HSN_RATE_HISTORY.items()
}


def _hsn_rate_for_date(hsn, on_date, table=None):
    """Longest-prefix match on `hsn` against HSN_RATE_HISTORY, then pick the
    period whose [from, to] window contains `on_date`. Returns a dict
    (rate, desc, confidence, source, prefix) or None if no code/period
    matches at all. IMPORTANT: if the matched period's rate is None
    (confidence='unconfirmed'), this STILL returns that dict -- callers
    must check `confidence`/`rate is None` themselves and skip comparison
    rather than silently falling back to an earlier period's rate."""
    table = table if table is not None else HSN_RATE_HISTORY
    hsn = (hsn or "").strip()
    best_prefix = None
    for prefix in table:
        if hsn.startswith(prefix) and (best_prefix is None or len(prefix) > len(best_prefix)):
            best_prefix = prefix
    if best_prefix is None:
        return None
    for from_d, to_d, rate, desc, confidence, source in table[best_prefix]:
        if from_d <= on_date and (to_d is None or on_date <= to_d):
            return dict(rate=rate, desc=desc, confidence=confidence, source=source,
                        prefix=best_prefix, period_from=from_d, period_to=to_d)
    return None


def _month_label_to_date(label):
    """'Apr-22' -> date(2022,4,1). Uses the FIRST of the month -- the GSTR-1
    HSN summary sheet is a MONTHLY aggregate (no per-invoice date available
    at this level), so a month straddling a rate-change boundary (only
    Sep-25, given GST 2.0's 22-Sep-2025 effective date) is an inherent
    blind spot: this resolves to the 1st, i.e. the PRE-change rate, for
    that specific month, and the check below flags Sep-25 rows with an
    explicit note rather than silently picking a side."""
    m = re.match(r"^([A-Za-z]{3})-(\d{2})$", label or "")
    if not m:
        return None
    mon, yy = m.group(1), m.group(2)
    inv = {v: k for k, v in mpu.CAL_MONTH_ABBR.items()}
    mm = inv.get(mon.title())
    if not mm:
        return None
    return _dt.date(2000 + int(yy), mm, 1)


# Rates that are legitimate WITHOUT being "wrong" for any HSN -- the 0.05%/0.1%
# concessional combined rate under Notification 40/2017-CT(R) & 41/2017-IT(R)
# for supply to merchant exporters. A rate matching this is downgraded to
# INFO ("verify merchant-export documentation"), never auto-FLAGged.
MERCHANT_EXPORT_RATES = {0.05, 0.1, 0.10}

# HSN prefixes that are wholly exempt (Nil-rated) -- curated list is EMPTY by
# design: none of this taxpayer's real HSN codes are in a commonly-exempt
# category (pharma/chemicals/scrap/machinery), so this check will correctly
# report "none found" rather than force an example that doesn't apply here.
EXEMPT_HSN_PREFIXES = {}   # {prefix: description}

# HSN prefixes that attract Compensation Cess -- same reasoning, empty here.
CESS_HSN_PREFIXES = {}     # {prefix: description}

# HSN prefixes with blocked ITC under Section 17(5) -- kept for documentation
# even though it can't be checked here (see check_blocked_itc_by_hsn).
BLOCKED_ITC_HSN_PREFIXES = {
    "8703": "Motor vehicles for transport of persons -- Sec 17(5)(a)",
    "9963": "Outdoor catering, club membership -- Sec 17(5)(b)",
    "9964": "Rent-a-cab / life & health insurance (unless notified) -- Sec 17(5)(b)",
}

# Standard CBIC 2-digit GST state/UT codes. Needed because GSTR-1's own
# Place-Of-Supply column is a plain state NAME, not a code.
STATE_NAME_TO_CODE = {
    "JAMMU AND KASHMIR": "01", "HIMACHAL PRADESH": "02", "PUNJAB": "03",
    "CHANDIGARH": "04", "UTTARAKHAND": "05", "HARYANA": "06", "DELHI": "07",
    "RAJASTHAN": "08", "UTTAR PRADESH": "09", "BIHAR": "10", "SIKKIM": "11",
    "ARUNACHAL PRADESH": "12", "NAGALAND": "13", "MANIPUR": "14", "MIZORAM": "15",
    "TRIPURA": "16", "MEGHALAYA": "17", "ASSAM": "18", "WEST BENGAL": "19",
    "JHARKHAND": "20", "ODISHA": "21", "ORISSA": "21", "CHHATTISGARH": "22",
    "MADHYA PRADESH": "23", "GUJARAT": "24", "DAMAN AND DIU": "25",
    "DADRA AND NAGAR HAVELI": "26", "MAHARASHTRA": "27", "ANDHRA PRADESH": "37",
    "KARNATAKA": "29", "GOA": "30", "LAKSHADWEEP": "31", "KERALA": "32",
    "TAMIL NADU": "33", "PUDUCHERRY": "34", "ANDAMAN AND NICOBAR ISLANDS": "35",
    "TELANGANA": "36", "LADAKH": "38", "OTHER TERRITORY": "97",
    "CENTRE JURISDICTION": "99", "OTHER COUNTRY": "96",
}

# Fixed-date national holidays only (Republic Day / Independence Day / Gandhi
# Jayanti never move) -- lunar/regional festival holidays are NOT included
# since they'd need an external calendar; #37 will under-count, not over-count.
NATIONAL_HOLIDAYS_FY2223 = {
    _dt.date(2022, 8, 15), _dt.date(2022, 10, 2), _dt.date(2023, 1, 26),
}

GSTR3B_DUE_DOM = 20   # monthly (non-QRMP) GSTR-3B due date -- confirmed not-QRMP


def _num(v):
    if v is None:
        return 0.0
    try:
        return float(str(v).replace(",", ""))
    except ValueError:
        return 0.0


def gstin_state(gstin):
    g = (gstin or "").strip()
    return g[:2] if len(g) >= 2 else None


def gstin_pan(gstin):
    g = (gstin or "").strip()
    return g[2:12] if len(g) >= 12 else None


def state_code_from_name(name):
    if not name:
        return None
    key = re.sub(r"\s+", " ", str(name)).strip().upper()
    key = re.sub(r"^\d+\s*-\s*", "", key)   # strip a leading 'NN - ' if present
    return STATE_NAME_TO_CODE.get(key)


def _hsn_prefix_lookup(hsn, table):
    """Longest-prefix match of `hsn` against `table`'s keys."""
    hsn = (hsn or "").strip()
    best = None
    for prefix in table:
        if hsn.startswith(prefix) and (best is None or len(prefix) > len(best)):
            best = prefix
    return best


# ======================================================================
# SUPPLEMENTARY (NOT authoritative) HSN rate reference -- mcp-india-stack
# ======================================================================
# HSN_RATE_MASTER above is the tool's PRIMARY, human-curated, taxpayer-
# verified rate table -- built only for HSN codes actually seen in a given
# taxpayer's real data. For a NEW taxpayer/industry with codes not yet in
# that curated list, this section provides a SECOND, WIDER-COVERAGE source
# so the tool isn't silently blind to unlisted codes -- but at a distinctly
# LOWER trust level, matching what its own upstream data says about itself.
#
# Source: the `mcp-india-stack` PyPI package's bundled `hsn_master.csv`
# (~22,500 rows, 2/4/8-digit CBIC hierarchy). Verified this round by reading
# the real package's real `tools/hsn.py` source and a real installed CSV
# (see chat history) -- NOT guessed from the package's own marketing
# examples, which turned out not to match the real API surface.
#
# Why this stays REVIEW, never FLAG, and is read from the RAW CSV rather
# than the package's own lookup_hsn_code() API:
#   1. The package's own DISCLAIMER text (baked into every one of its real
#      API responses): "GST rates may vary based on specific conditions.
#      Verify with a tax professional for commercial transactions."
#   2. No notification/effective-date column in the CSV -- staleness after
#      a rate revision (e.g. the 2025 GST-rate-rationalisation round) can't
#      be detected from the file itself.
#   3. It's a single-maintainer community dataset, not a CBIC-audited feed.
#   4. lookup_hsn_code() ships with a background auto-update from a CDN --
#      reading its live API would make this tool's findings NON-reproducible
#      run-to-run. Reading the CSV directly, once, from whatever version is
#      pip-installed, avoids that -- output is reproducible for a given
#      installed package version, which is the best available compromise.
#   5. lookup_hsn_code() itself does `rows[0]` on a duplicate-code match
#      with no disambiguation -- this reader is STRICTER: a code with
#      multiple CSV rows that DISAGREE on rate is dropped from the table
#      entirely (never guesses which row is right), and a row whose
#      CGST_Rate+SGST_Rate doesn't equal its own IGST_Rate (an internal
#      data-consistency problem) is also dropped rather than trusted.
#
# Importing the top-level `mcp_india_stack` package only touches its
# lightweight __init__.py (confirmed: dir(mcp_india_stack) shows only
# dunder attributes) -- it does NOT pull in fastmcp/starlette/uvicorn/etc,
# which only load if you import mcp_india_stack.server or .tools.*. This
# reader deliberately never imports those -- it locates the CSV via the
# package's own __file__ and reads it with the stdlib csv module only.
_MCP_INDIA_STACK_HSN_TABLE = None   # lazy singleton, built once per process


# ======================================================================
# HSN/SAC CODE-AND-DESCRIPTION MASTER (user-supplied, or bundled default)
# ======================================================================
# Distinct from HSN_RATE_HISTORY above and from mcp-india-stack's table:
# this master has CODE + DESCRIPTION columns ONLY -- no GST rate column at
# all (confirmed on the real file: sheets 'HSN_MSTR'/'SAC_MSTR', columns
# HSN_CD/HSN_Description and SAC_CD/SAC_Description). It cannot be used for
# rate comparison. What it DOES enable: confirming a reported HSN/SAC code
# actually EXISTS as a real classification, and showing the code's official
# description alongside the taxpayer's own description for a quick human
# eyeball-check -- neither of these existed in this tool before.
#
# Source-of-truth behaviour (per explicit instruction): if the CURRENT run's
# folder contains a file with this exact sheet signature, that file is used
# for THIS run (a fresh, one-off reference). If not, the tool falls back to
# `HSN_SAC_default.xlsx`, bundled alongside these .py files -- this is the
# "hardcoded" default, built from whatever HSN_SAC.xlsx was most recently
# supplied and explicitly asked to become the new baseline. To make a new
# upload the new permanent default, replace HSN_SAC_default.xlsx with it
# (Claude does this in a future session on request -- there is no
# background process that does this automatically; see chat history).
_HSN_SAC_MASTER_CACHE = {}   # keyed by resolved path, so a run-supplied file
                             # never contaminates the bundled-default cache


def _load_hsn_sac_master(override_path=None):
    """Returns dict(hsn={code: description}, sac={code: description},
    available=True/False, source=path_used, reason=...). Never raises --
    an unreadable or entirely absent master degrades to available=False,
    and the validity check below skips cleanly (same pattern as every
    other optional source in this tool)."""
    default_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "HSN_SAC_default.xlsx")
    path = override_path if (override_path and os.path.exists(override_path)) else default_path
    if path in _HSN_SAC_MASTER_CACHE:
        return _HSN_SAC_MASTER_CACHE[path]
    out = dict(hsn={}, sac={}, available=False, source=path, reason=None)
    if not os.path.exists(path):
        out["reason"] = f"Neither a run-supplied HSN/SAC master nor the bundled default was found at {path!r}."
        _HSN_SAC_MASTER_CACHE[path] = out
        return out
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if "HSN_MSTR" in wb.sheetnames:
            ws = wb["HSN_MSTR"]
            for r in ws.iter_rows(min_row=2, values_only=True):
                if not r or not r[0]:
                    continue
                out["hsn"][str(r[0]).strip()] = str(r[1] or "").strip()
        if "SAC_MSTR" in wb.sheetnames:
            ws = wb["SAC_MSTR"]
            for r in ws.iter_rows(min_row=2, values_only=True):
                if not r or r[0] is None:
                    continue
                out["sac"][str(r[0]).strip()] = str(r[1] or "").strip()
        wb.close()
        out["available"] = bool(out["hsn"] or out["sac"])
        print(f"[info] HSN/SAC master loaded from {path!r}: {len(out['hsn'])} HSN code(s), "
              f"{len(out['sac'])} SAC code(s).")
    except Exception as ex:
        out["reason"] = f"Could not read {path!r}: {ex}"
        out["available"] = False
    _HSN_SAC_MASTER_CACHE[path] = out
    return out


def check_hsn_master_validity(hsn_by_month, override_path=None):
    """A7: does each reported HSN code actually exist in the official
    code-and-description master? A code not found is REVIEW (not FLAG --
    the master is a point-in-time snapshot; a genuinely new WCO/CBIC code
    added after the snapshot was taken would be a false alarm at FLAG
    severity, and this tool has no 'as-of date' for this particular file
    the way it does for HSN_RATE_HISTORY's rate periods). Also surfaces the
    official description next to the taxpayer's own, for a quick human
    eyeball-check -- no automated fuzzy-text-matching judgment is made,
    since that risks exactly the kind of noisy false-positive this tool's
    whole design avoids."""
    master = _load_hsn_sac_master(override_path)
    if not master["available"]:
        return [Finding("A7", "HSN/SAC code-validity check (master list)", INFO,
                         master["reason"] or "HSN/SAC master not available this run.")]
    F = []
    checked = flagged = 0
    for m in sorted(hsn_by_month, key=lambda x: MONTH_IDX.get(x, 99)):
        for row in hsn_by_month[m]:
            hsn, desc = row["hsn"], row["desc"]
            if not hsn or hsn.isalpha():
                continue
            checked += 1
            official_desc = master["hsn"].get(hsn)
            if official_desc is None:
                # try the SAC master too (service codes can appear on the same 'hsn' sheet)
                official_desc = master["sac"].get(hsn)
            if official_desc is None:
                flagged += 1
                F.append(Finding("A7", "HSN/SAC code not found in official master", REVW,
                    f"{m}: HSN/SAC '{hsn}' (taxpayer's own description: {desc[:80]!r}) does not "
                    f"appear in the code-and-description master ({os.path.basename(master['source'])}). "
                    f"Could be a genuinely invalid/mistyped code, OR a real code added to the "
                    f"official list after this master's snapshot was taken -- verify on the GST "
                    f"portal's own HSN/SAC search before treating as an error. "
                    f"Taxable Rs.{row['taxable']:,.2f}.",
                    numbers=dict(month=m, hsn=hsn, taxable=row["taxable"])))
    source_note = os.path.basename(master["source"])
    if checked and not flagged:
        F.append(Finding("A7", "HSN/SAC code-validity check (master list)", PASS,
                          f"{checked} HSN/SAC code line(s) checked against {source_note} -- every "
                          f"code exists in the official master."))
    elif not checked:
        F.append(Finding("A7", "HSN/SAC code-validity check (master list)", INFO,
                          "No HSN summary rows supplied this run -- nothing to check."))
    return F


def _load_mcp_india_stack_hsn_table():
    global _MCP_INDIA_STACK_HSN_TABLE
    if _MCP_INDIA_STACK_HSN_TABLE is not None:
        return _MCP_INDIA_STACK_HSN_TABLE
    table = {}
    try:
        import mcp_india_stack
        import csv as _csv
        csv_path = os.path.join(os.path.dirname(mcp_india_stack.__file__), "data", "hsn", "hsn_master.csv")
        if not os.path.exists(csv_path):
            print("[info] mcp-india-stack installed but hsn_master.csv not found at the expected "
                  f"path ({csv_path}) -- extended HSN-rate reference not available.")
            _MCP_INDIA_STACK_HSN_TABLE = {}
            return {}
        by_code = {}
        with open(csv_path, newline="", encoding="utf-8") as f:
            for row in _csv.DictReader(f):
                code = (row.get("HSNCode") or "").strip()
                if not code:
                    continue
                try:
                    cgst = float(row.get("CGST_Rate") or 0)
                    sgst = float(row.get("SGST_Rate") or 0)
                    igst = float(row.get("IGST_Rate") or 0)
                except ValueError:
                    continue
                by_code.setdefault(code, []).append(
                    dict(igst=igst, cgst=cgst, sgst=sgst, desc=(row.get("Description") or "").strip()))
        dropped_ambiguous = dropped_inconsistent = dropped_allzero = 0
        for code, rows in by_code.items():
            # drop if internally inconsistent (CGST+SGST should equal IGST) -- never trust a
            # row the source data itself doesn't agree with.
            consistent_rows = [r for r in rows if abs((r["cgst"] + r["sgst"]) - r["igst"]) < 0.01]
            if len(consistent_rows) < len(rows):
                dropped_inconsistent += (len(rows) - len(consistent_rows))
            if not consistent_rows:
                continue
            # drop if duplicate rows for the same code DISAGREE on rate -- never guess which
            # one applies (this is stricter than mcp-india-stack's own lookup_hsn_code(), which
            # silently takes rows[0]).
            distinct_rates = {round(r["igst"], 2) for r in consistent_rows}
            if len(distinct_rates) > 1:
                dropped_ambiguous += 1
                continue
            rate = consistent_rows[0]["igst"]
            # DROP a lone all-zero rate -- CONFIRMED on a real refreshed export (this session):
            # 22,471 of 22,500 rows (99.9%) in this dataset carry CGST=SGST=IGST=CESS=0.0, and
            # cross-checking against 8 of this taxpayer's own real, non-nil-rated curated codes
            # (3003/3808/3915/4707/7204/7606/8402/998843, all genuinely taxed at 5-18% per their
            # actual filed GSTR-1) showed EVERY one of them sitting at this same all-zero pattern
            # in the dataset. This means an all-zero row here overwhelmingly signals "no rate was
            # ever populated for this entry" (a bare HSN/SAC nomenclature/classification stub),
            # NOT "this good/service is genuinely nil-rated" -- trusting it would manufacture a
            # false REVIEW against a correctly-taxed real invoice on almost every code. A
            # genuinely nil-rated good therefore currently shows as "not covered" rather than
            # confirmed-nil -- an intentional false-negative (silence) instead of a false-positive
            # (wrong flag), consistent with this tool's severity philosophy throughout.
            if abs(rate) < 0.01:
                dropped_allzero += 1
                continue
            table[code] = (rate, consistent_rows[0]["desc"])
        print(f"[info] mcp-india-stack HSN reference loaded: {len(table)} usable code(s) "
              f"({dropped_ambiguous} code(s) dropped for disagreeing duplicate rates, "
              f"{dropped_inconsistent} row(s) dropped for CGST+SGST != IGST, "
              f"{dropped_allzero} code(s) dropped for an all-zero rate -- see loader comment).")
    except ImportError:
        table = {}
    except Exception as ex:
        print(f"[warn] mcp-india-stack HSN reference could not be loaded ({ex}) -- "
              "extended HSN-rate check skipped, not guessed.")
        table = {}
    _MCP_INDIA_STACK_HSN_TABLE = table
    return table


# ======================================================================
# DATA LOADERS -- read the merged files ONCE for the whole FY
# ======================================================================
def _hsn_rows_by_month(gstr1_path):
    """{month: [dict(hsn, desc, rate, taxable, igst, cgst, sgst, cess), ...]}

    BUG FIX: this used to read ONLY the 'hsn' tab directly, which -- confirmed against a real
    taxpayer export -- carries a marker for April only; every check built on this (A1, A1-EXT,
    A7, and the HSN Rate Review table on each month's Comparison sheet) silently saw "no HSN
    rows" for May-25 onward. Now delegates to gst_parsers_returns.read_gstr1_hsn_all_months(),
    which reads the correct tab ('hsn' or the newer 'hsn(b2b)'/'hsn(b2c)' pair) per month -- see
    that function's docstring for the full explanation. Single source of truth: this is the
    same data the HSN Rate Review table (on each month's own Comparison sheet) draws on, so the two can never drift out of sync with each other."""
    hsn_all = ewbp.read_gstr1_hsn_all_months(gstr1_path)
    out = {}
    for m, recs in hsn_all.items():
        out[m] = [dict(hsn=r["hsn"], desc=r["desc"], uqc=r.get("uqc", ""), qty=r.get("qty", 0.0),
                       rate=r["rate"], taxable=r["taxable"], igst=r["igst"], cgst=r["cgst"],
                       sgst=r["sgst"], cess=r["cess"], source_tab=r.get("source_tab", "hsn"))
                  for r in recs]
    return out


def _cdnr_rows_by_month(gstr1_path):
    """{month: [dict(gstin, noteno, notedate, notetype, taxable, igst,cgst,sgst), ...]}
    from the 'cdnr' sheet (credit/debit notes to registered recipients).
    BUG FIX -- same root cause as the b2b sheet's continuation-row fix in parse_gstr1()
    (gst_parsers_returns.py): the portal's own export leaves GSTIN/Note Number/Note Date
    blank (merged cell) on every rate-line after a multi-rate note's first row. This used to
    both (a) read gstin="" on those rows since it took each row's own blank cell literally,
    AND (b) the `if not r or not r[0]: continue` guard actively SKIPPED them outright --
    confirmed on real data, 488 of 1,636 cdnr rows across this taxpayer's FY (30%) are
    continuation rows of this shape, previously dropped from checks #4/#11 entirely rather
    than being attributed to their real note. Fixed the same way: forward-fill the last-seen
    gstin/noteno/notedate/notetype within each month's block."""
    wb = openpyxl.load_workbook(gstr1_path, data_only=True)
    if "cdnr" not in wb.sheetnames:
        return {}
    rows = list(wb["cdnr"].iter_rows(values_only=True))
    hdr = [str(c).strip() if c else "" for c in rows[3]]
    H = {h: i for i, h in enumerate(hdr)}
    blocks = mpu.split_rows_by_month(rows[4:])
    out = {}
    for m, rws in blocks.items():
        lst = []
        last_gstin = last_noteno = last_notedate = last_notetype = None
        for r in rws:
            if not r or not any(r):
                continue
            g = lambda k: r[H[k]] if k in H and H[k] < len(r) else None
            raw_gstin = g("GSTIN/UIN of Recipient")
            raw_noteno = g("Note Number")
            if raw_gstin not in (None, "") and raw_noteno not in (None, ""):
                last_gstin = str(raw_gstin).strip()
                last_noteno = str(raw_noteno).strip()
                last_notedate = str(g("Note Date") or "").strip()
                last_notetype = str(g("Note Type") or "").strip()
                gstin_v, noteno_v, notedate_v, notetype_v = last_gstin, last_noteno, last_notedate, last_notetype
            elif last_gstin is not None:
                gstin_v, noteno_v, notedate_v, notetype_v = last_gstin, last_noteno, last_notedate, last_notetype
            else:
                gstin_v, noteno_v, notedate_v, notetype_v = "", "", "", ""
            lst.append(dict(
                gstin=gstin_v,
                noteno=noteno_v,
                notedate=notedate_v,
                notetype=notetype_v,
                taxable=_num(g("Taxable Value")),
                igst=_num(g("Integrated Tax")), cgst=_num(g("Central Tax")),
                sgst=_num(g("State/UT Tax")),
            ))
        out[m] = lst
    return out


def _b2cl_rows_by_month(gstr1_path):
    """BUG FIX -- same root cause as the b2b/cdnr fixes: multi-rate B2C-large invoices leave
    Invoice Number/date/value/POS blank on continuation rows. Confirmed on real data (e.g. a
    3-rate-line invoice under GSTIN-less B2CL rows 145-147). Forward-filled the same way."""
    wb = openpyxl.load_workbook(gstr1_path, data_only=True)
    if "b2cl" not in wb.sheetnames:
        return {}
    rows = list(wb["b2cl"].iter_rows(values_only=True))
    hdr = [str(c).strip() if c else "" for c in rows[3]]
    H = {h: i for i, h in enumerate(hdr)}
    blocks = mpu.split_rows_by_month(rows[4:])
    out = {}
    for m, rws in blocks.items():
        lst = []
        last_invno = last_invdate = last_invval = last_pos = None
        for r in rws:
            if not r or not any(r):
                continue
            g = lambda k: r[H[k]] if k in H and H[k] < len(r) else None
            raw_invno = g("Invoice Number")
            if raw_invno not in (None, ""):
                last_invno = str(raw_invno).strip()
                last_invdate = str(g("Invoice date") or "").strip()
                last_invval = _num(g("Invoice Value"))
                last_pos = str(g("Place Of Supply") or "").strip()
                invno_v, invdate_v, invval_v, pos_v = last_invno, last_invdate, last_invval, last_pos
            elif last_invno is not None:
                invno_v, invdate_v, invval_v, pos_v = last_invno, last_invdate, last_invval, last_pos
            else:
                invno_v, invdate_v, invval_v, pos_v = "", "", 0.0, ""
            lst.append(dict(
                invno=invno_v,
                invdate=invdate_v,
                invval=invval_v, pos=pos_v,
                rate=_num(g("Rate")), taxable=_num(g("Taxable Value")),
                igst=_num(g("Integrated Tax")),
            ))
        out[m] = lst
    return out


def _parse_ddmmyyyy(s):
    s = (s or "").strip()
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%b-%Y", "%d-%b-%y"):
        try:
            return _dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _gstr3b_month_fields(gstr3b_path, month):
    """Pull the handful of GSTR-3B fields these checks need for one month,
    reading the sheet's own content-based Year/Tax-Period key rows (never
    the sheet name), matching the convention documented for this file."""
    wb = openpyxl.load_workbook(gstr3b_path, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        fy = tp = None
        arn_date = None
        for row in rows:
            cells = [c for c in row if c not in (None, "")]
            if not cells:
                continue
            key = str(cells[0]).strip().upper()
            if key in ("YEAR", "FINANCIAL YEAR") and len(cells) >= 2:
                fy = cells[1]
            elif key == "TAX PERIOD" and len(cells) >= 2:
                tp = cells[1]
            elif key == "DATE OF ARN" and len(cells) >= 2:
                arn_date = str(cells[1])
        if not (fy and tp):
            continue
        try:
            labels = mpu.months_for_tax_period(str(fy), str(tp))
        except mpu.PeriodParseError:
            continue
        if month not in labels:
            continue
        # Found the right sheet -- now pull 4A5 / 4B(1,2) / D(1) / 3.1(d).
        out = dict(filing_date=_parse_ddmmyyyy(arn_date), rcm_liability_igst=0.0,
                    itc_all_other_igst=0.0, itc_reversed_rule42_43=0.0,
                    itc_reversed_others=0.0, itc_ineligible_175=0.0)
        for i, row in enumerate(rows):
            cells = [c for c in row if c not in (None, "")]
            if not cells:
                continue
            label = str(cells[0]).strip()
            nums = [c for c in row if isinstance(c, (int, float))]
            if label.startswith("(d) Inward supplies (liable to reverse charge)"):
                if len(nums) >= 1:
                    out["rcm_liability_igst"] = _num(nums[0])
            elif label.startswith("(5) All other ITC"):
                if len(nums) >= 1:
                    out["itc_all_other_igst"] = _num(nums[0])
            elif label.startswith("(1) As per rules 42"):
                if len(nums) >= 1:
                    out["itc_reversed_rule42_43"] = _num(nums[0])
            elif label.startswith("(2) Others") and "itc_reversed_others" in out and out["itc_reversed_others"] == 0.0:
                # first '(2) Others' after 'B. ITC Reversed' -- distinguished
                # from D(2) below by scan order (B comes before D in the form)
                if len(nums) >= 1 and out["itc_ineligible_175"] == 0.0:
                    out["itc_reversed_others"] = _num(nums[0])
            elif label.startswith("(1) As per section 17(5)"):
                if len(nums) >= 1:
                    out["itc_ineligible_175"] = _num(nums[0])
        return out
    return None


# ======================================================================
# A. HSN-ONLY CHECKS
# ======================================================================
def check_hsn_rate_master(hsn_by_month):
    """A1 wrong-rate, A2 exempt-charged, A3 cess-missing, A6 invalid/short HSN
    (Rule 46 -- turnover confirmed >5cr for FY2022-23 per BO Profile, so a
    6-digit HSN is mandatory; taxpayer is on 4-digit for two real codes)."""
    F = []
    for m in sorted(hsn_by_month, key=lambda x: MONTH_IDX.get(x, 99)):
        for row in hsn_by_month[m]:
            hsn, rate = row["hsn"], row["rate"]
            # A6 -- digit-length / Rule 46
            if len(hsn) < 6 and not hsn.isalpha():
                F.append(Finding("A6", "Invalid/short HSN format (Rule 46)", FLAG,
                    f"{m}: HSN '{hsn}' ({row['desc'][:50]}) reported at {len(hsn)} digits. "
                    f"Turnover >Rs.5cr requires 6-digit HSN. Taxable value at this code: "
                    f"Rs.{row['taxable']:,.2f}.", numbers=dict(month=m, hsn=hsn, taxable=row["taxable"])))
            on_date = _month_label_to_date(m)
            match = _hsn_rate_for_date(hsn, on_date) if on_date else None
            if match:
                expected, desc, confidence, source = match["rate"], match["desc"], match["confidence"], match["source"]
                if expected is None:
                    # Known rate-change period, but this code's new rate wasn't confidently
                    # researched -- do NOT compare, do NOT silently fall back to the prior
                    # period's rate. Surface once per code+month as INFO, not a guessed FLAG.
                    F.append(Finding("A1", "HSN rate reference unconfirmed for this period", INFO,
                        f"{m}: HSN {hsn} billed at {rate}% -- this month falls in a period "
                        f"(from {match['period_from']}) whose reference rate for "
                        f"this HSN was NOT confidently researched ({desc}). {source} No comparison "
                        f"made -- verify manually against the notification in force on the invoice date.",
                        numbers=dict(month=m, hsn=hsn, rate=rate)))
                elif abs(rate - expected) > 0.01:
                    conf_note = ("" if confidence == "high" else
                                 f" [{confidence.upper()} CONFIDENCE reference -- {source}]")
                    if rate in MERCHANT_EXPORT_RATES:
                        F.append(Finding("A1", "Concessional rate on HSN (verify merchant-export)", INFO,
                            f"{m}: HSN {hsn} ({desc}) billed at {rate}% vs standard {expected}%.{conf_note} "
                            f"Matches the merchant-export concessional rate (Notif 40/41-2017) -- "
                            f"verify Form CT-1/ARE-3 or LUT-linked merchant-export documentation, "
                            f"this is not automatically a misclassification. Taxable Rs.{row['taxable']:,.2f}.",
                            numbers=dict(month=m, hsn=hsn, rate=rate, expected=expected)))
                    else:
                        sev = FLAG if confidence == "high" else REVW
                        F.append(Finding("A1", "Wrong GST rate charged vs standard HSN rate", sev,
                            f"{m}: HSN {hsn} ({desc}) billed at {rate}% -- reference rate is "
                            f"{expected}%.{conf_note} Taxable Rs.{row['taxable']:,.2f}, tax at this line "
                            f"Rs.{row['igst']+row['cgst']+row['sgst']:,.2f}. Verify against the rate "
                            f"notification in force on the invoice date before treating as confirmed.",
                            numbers=dict(month=m, hsn=hsn, rate=rate, expected=expected, taxable=row["taxable"])))
            # A2 exempt-charged
            exempt_prefix = _hsn_prefix_lookup(hsn, EXEMPT_HSN_PREFIXES)
            if exempt_prefix and rate > 0:
                F.append(Finding("A2", "GST charged on exempt HSN", FLAG,
                    f"{m}: HSN {hsn} ({EXEMPT_HSN_PREFIXES[exempt_prefix]}) is exempt but billed at "
                    f"{rate}%. Wrongful collection -- refund/reversal required.",
                    numbers=dict(month=m, hsn=hsn, rate=rate)))
            # A3 cess-missing
            cess_prefix = _hsn_prefix_lookup(hsn, CESS_HSN_PREFIXES)
            if cess_prefix and row["cess"] == 0 and row["taxable"] > 0:
                F.append(Finding("A3", "Compensation Cess missing on cess-applicable HSN", FLAG,
                    f"{m}: HSN {hsn} ({CESS_HSN_PREFIXES[cess_prefix]}) attracts Cess but Cess=0 on "
                    f"taxable Rs.{row['taxable']:,.2f}.", numbers=dict(month=m, hsn=hsn)))
    if not any(f.ref == "A2" for f in F):
        F.append(Finding("A2", "GST charged on exempt HSN", INFO,
            "No HSN code in this taxpayer's real 'hsn' sheet is in the curated exempt-HSN list "
            "(pharma/chemicals/scrap/machinery mix -- none of these categories are commonly exempt). "
            "No finding to report; extend EXEMPT_HSN_PREFIXES if you have specific exempt lines to check."))
    if not any(f.ref == "A3" for f in F):
        F.append(Finding("A3", "Compensation Cess missing on cess-applicable HSN", INFO,
            "No HSN code found matches the curated Cess-applicable list (tobacco/coal/luxury etc. -- "
            "none of this taxpayer's codes fall in those categories). Extend CESS_HSN_PREFIXES if needed."))
    return F


def check_hsn_rate_master_extended(hsn_by_month):
    """A1-EXT: same wrong-rate idea as check_hsn_rate_master's A1, but ONLY
    for HSN codes NOT already covered by the curated, taxpayer-verified
    HSN_RATE_MASTER above -- widens coverage for a new taxpayer/industry
    whose codes aren't in that hand-curated list yet. ALWAYS REVIEW, never
    FLAG (see the long comment above _load_mcp_india_stack_hsn_table for
    why) -- this is a lead to verify, not a finding to act on directly."""
    F = []
    table = _load_mcp_india_stack_hsn_table()
    if not table:
        return [Finding("A1-EXT", "Extended HSN-rate reference (mcp-india-stack)", INFO,
                         "mcp-india-stack not installed, or its bundled hsn_master.csv could not be "
                         "read -- extended coverage beyond the curated HSN_RATE_MASTER is not "
                         "available this run. Install with 'pip install mcp-india-stack "
                         "--break-system-packages' to enable (optional; the curated master's own "
                         "codes are unaffected either way).")]
    GST_2_0_EFFECTIVE = _dt.date(2025, 9, 22)
    checked = 0
    for m in sorted(hsn_by_month, key=lambda x: MONTH_IDX.get(x, 99)):
        on_date = _month_label_to_date(m)
        for row in hsn_by_month[m]:
            hsn, rate = row["hsn"], row["rate"]
            curated = _hsn_rate_for_date(hsn, on_date) if on_date else None
            if curated and curated["rate"] is not None:
                continue   # curated master has a CONFIRMED rate for this specific period -- skip here
            # mcp-india-stack's data carries no date/notification info of its own (see loader
            # docstring) -- given the package is actively maintained in 2026, it most likely
            # reflects POST-GST-2.0 rates. Applying it against a PRE-2.0 invoice would recreate
            # exactly the systematic-false-REVIEW risk already flagged -- so it is only used for
            # months on/after GST 2.0's effective date. Pre-2.0 months with no curated coverage
            # simply get no extended check (better than a likely-wrong one).
            if on_date and on_date < GST_2_0_EFFECTIVE:
                continue
            prefix = _hsn_prefix_lookup(hsn, table)
            if not prefix:
                continue
            checked += 1
            expected, desc = table[prefix]
            if abs(rate - expected) > 0.01 and rate not in MERCHANT_EXPORT_RATES:
                match_type = "exact code match" if prefix == hsn else f"rolled up from {hsn!r} to {prefix!r} (shorter/parent code)"
                F.append(Finding("A1-EXT", "Rate differs from mcp-india-stack reference (verify before acting)",
                    REVW,
                    f"{m}: HSN {hsn} ({desc[:60]}) billed at {rate}% vs mcp-india-stack's reference "
                    f"{expected}% ({match_type}). SOURCE CAVEAT: mcp-india-stack is a third-party, "
                    f"community-maintained dataset, not a CBIC-audited feed, carries no date/"
                    f"notification info of its own, and its own disclaimer states rates 'may vary "
                    f"based on specific conditions' -- this is a lead to check against the actual "
                    f"rate notification in force on the invoice date (or add this HSN to "
                    f"HSN_RATE_HISTORY once confirmed), NOT a confirmed misclassification. "
                    f"Taxable Rs.{row['taxable']:,.2f}.",
                    numbers=dict(month=m, hsn=hsn, rate=rate, expected=expected)))
    total_rows = sum(len(rows) for rows in hsn_by_month.values())
    if checked and not any(f.ref == "A1-EXT" and f.severity == REVW for f in F):
        F.append(Finding("A1-EXT", "Extended HSN-rate reference (mcp-india-stack)", PASS,
                          f"{checked} HSN code line(s) not in the curated master were checked against "
                          "mcp-india-stack's reference rates -- no discrepancy found."))
    elif checked == 0 and total_rows == 0:
        F.append(Finding("A1-EXT", "Extended HSN-rate reference (mcp-india-stack)", INFO,
                          "No HSN summary rows supplied this run (GSTR-1 'hsn' sheet empty or absent) "
                          "-- nothing to check."))
    elif checked == 0:
        F.append(Finding("A1-EXT", "Extended HSN-rate reference (mcp-india-stack)", INFO,
                          "Every HSN code line this run was either already covered by the curated "
                          "HSN_RATE_HISTORY with a confirmed rate for its period, or fell in a "
                          "pre-GST-2.0 (before 22-Sep-2025) month where mcp-india-stack's undated "
                          "'latest' reference is deliberately NOT applied (it likely reflects "
                          "current/post-2.0 rates, and using it against an older invoice would "
                          "recreate the systematic-mismatch risk this gating exists to prevent)."))
    return F


def check_hsn_multi_rate(hsn_by_month):
    """A4 -- same literal HSN code taxed at >1 distinct rate within the same
    month (excluding the merchant-export concessional rate, which legitimately
    coexists with the standard rate for the same product)."""
    F = []
    for m in sorted(hsn_by_month, key=lambda x: MONTH_IDX.get(x, 99)):
        by_hsn = {}
        for row in hsn_by_month[m]:
            by_hsn.setdefault(row["hsn"], set()).add(row["rate"])
        for hsn, rates in by_hsn.items():
            non_concessional = rates - MERCHANT_EXPORT_RATES
            if len(non_concessional) > 1:
                F.append(Finding("A4", "Same HSN taxed at multiple rates in one month", FLAG,
                    f"{m}: HSN {hsn} appears at rates {sorted(non_concessional)} within the same month "
                    f"(besides any merchant-export-concessional rate). Verify product sub-classification "
                    f"-- either a genuine product-mix split or a misclassification.",
                    numbers=dict(month=m, hsn=hsn, rates=sorted(non_concessional))))
    return F


def check_blocked_itc_by_hsn():
    """A5 -- documented as structurally not-computable: GSTR-2B's B2B sheet
    (the only source of purchase-side invoices) carries no HSN column at all
    in this file, so ITC-by-HSN blocking under Sec 17(5) cannot be checked
    line-by-line here."""
    return [Finding("A5", "Blocked ITC (Sec 17(5)) by purchase-side HSN", INFO,
        "Not computable: the GSTR-2B 'B2B' sheet in your files has no HSN column on the purchase "
        "side (only GSTIN/invoice/rate/tax) -- there is no invoice-to-HSN link to test against the "
        f"blocked-ITC categories ({', '.join(BLOCKED_ITC_HSN_PREFIXES.values())}). "
        "Would need a purchase register or e-invoice-of-suppliers feed with HSN to check this.")]


# ======================================================================
# B. POS / STATE-CODE CHECKS
# ======================================================================
def check_pos_tax_head(g1_lines_by_month, self_gstin):
    """B1 wrong tax head (IGST vs CGST+SGST), B3 POS missing/unmapped.
    IMPORTANT: the tax-head rule compares Place-Of-Supply against the
    SUPPLIER's own state (self_gstin), NOT the recipient's registered state
    -- POS almost always equals the recipient's own state trivially (that's
    just where they're registered), so comparing against it would flag
    nearly every genuinely-correct inter-state invoice. Caught this exact
    bug during testing against the real file before shipping this check."""
    F = []
    self_state = gstin_state(self_gstin)
    for m in sorted(g1_lines_by_month, key=lambda x: MONTH_IDX.get(x, 99)):
        for inv in g1_lines_by_month[m]:
            if inv["sheet"] != "b2b, sez, de_inv" or not inv["invno"]:
                continue   # skip multi-rate continuation rows (no invno/POS of their own)
            pos_code = state_code_from_name(inv["pos"])
            if pos_code is None:
                F.append(Finding("B3", "Place of Supply missing/unmapped", FLAG,
                    f"{m}: Invoice {inv['invno']} to {inv['gstin']} has POS '{inv['pos']}' which "
                    f"could not be mapped to a state code -- verify manually.",
                    numbers=dict(month=m, invno=inv["invno"])))
                continue
            inter_state = pos_code != self_state
            has_igst = inv["igst"] > 0.01
            has_cgst_sgst = (inv["cgst"] + inv["sgst"]) > 0.01
            if inter_state and has_cgst_sgst and not has_igst:
                F.append(Finding("B1", "Wrong tax head: inter-state but CGST/SGST charged", FLAG,
                    f"{m}: Invoice {inv['invno']} -- POS {inv['pos']} ({pos_code}) vs supplier state "
                    f"{self_state} is inter-state, but CGST {inv['cgst']:,.2f} + SGST "
                    f"{inv['sgst']:,.2f} charged instead of IGST. Revenue misallocation between "
                    f"Centre/State.", numbers=dict(month=m, invno=inv["invno"])))
            elif not inter_state and has_igst:
                F.append(Finding("B1", "Wrong tax head: intra-state but IGST charged", FLAG,
                    f"{m}: Invoice {inv['invno']} -- POS {inv['pos']} matches supplier state "
                    f"{self_state} (intra-state), but IGST {inv['igst']:,.2f} charged instead of "
                    f"CGST+SGST.", numbers=dict(month=m, invno=inv["invno"])))
    if not any(f.ref == "B1" for f in F):
        F.append(Finding("B1", "Wrong tax head (IGST vs CGST/SGST) vs Place of Supply", PASS,
            "No B2B invoice's tax head disagrees with what its Place-of-Supply (vs the supplier's own "
            "state) implies."))
    return F


def check_b2c_large_ewb(b2cl_by_month, ewb_out_rows):
    """B2 -- inter-state B2C-Large (>2.5L, Table 5A) invoices with no matching
    Outward EWB. Matched by invoice number (the only common key -- B2C has no
    recipient GSTIN)."""
    F = []
    ewb_by_invno = {}
    for e in ewb_out_rows:
        if e["docno"]:
            ewb_by_invno.setdefault(e["docno"], []).append(e)
    for m in sorted(b2cl_by_month, key=lambda x: MONTH_IDX.get(x, 99)):
        for inv in b2cl_by_month[m]:
            if inv["invno"] not in ewb_by_invno:
                F.append(Finding("B2", "Inter-state B2C-Large invoice with no matching EWB", FLAG,
                    f"{m}: B2CL invoice {inv['invno']} (POS {inv['pos']}, value Rs.{inv['invval']:,.2f}) "
                    f"has no matching Outward EWB by invoice number. Rule 138 requires an EWB for "
                    f"inter-state movement above the threshold -- verify if goods actually moved, or "
                    f"if this was a services-only / non-movement supply.",
                    numbers=dict(month=m, invno=inv["invno"], value=inv["invval"])))
    if not any(f.severity == FLAG for f in F):
        F.append(Finding("B2", "Inter-state B2C-Large invoice with no matching EWB", PASS,
            "Every B2C-Large invoice found in the supplied months has a matching Outward EWB by "
            "invoice number."))
    return F


def check_sez_misclassification(g1_lines_by_month):
    """B4 -- SEZ/Export invoice-type vs recipient-GSTIN-state consistency."""
    F = []
    sez_types = {"SEZWP", "SEZWOP", "SEWP", "SEWOP", "DE"}
    found_any = False
    for m in sorted(g1_lines_by_month, key=lambda x: MONTH_IDX.get(x, 99)):
        for inv in g1_lines_by_month[m]:
            if inv["sheet"] != "b2b, sez, de_inv":
                continue
            itype = ""  # invoice-type isn't retained by read_gstr1_lines; see note below
    F.append(Finding("B4", "SEZ / Export invoice-type misclassification", INFO,
        "All B2B invoices in your real GSTR-1 file are Invoice Type 'R' (Regular) -- no SEZWP/SEZWOP/"
        "Deemed-Export rows were found in any supplied month, so there is nothing to cross-check here "
        "for this taxpayer/period. Will activate automatically if a future period contains SEZ/DE rows."))
    return F


# ======================================================================
# C. COMBINED (HSN + STATE) CHECKS
# ======================================================================
def check_rcm_hsn_not_declared():
    """C1 -- structurally not computable: GTA-type RCM-by-HSN needs an
    invoice-level HSN on the GSTR-1/E-Invoice side, which isn't present."""
    return [Finding("C1", "RCM applicability by HSN (e.g. GTA 9965) not declared", INFO,
        "Not computable: neither GSTR-1 nor E-Invoice carries an HSN column at invoice level in your "
        "files, so a specific RCM-attracting HSN (e.g. 9965 Goods Transport Agency) can't be tied to "
        "a specific invoice. The aggregate RCM liability (3.1(d)) IS available per month via GSTR-3B "
        "and is cross-checked elsewhere in the pipeline.")]


def check_branch_transfer(g1_lines_by_month, self_gstin):
    """C2 -- same PAN, different state = stock transfer, not a real sale."""
    F = []
    self_pan = gstin_pan(self_gstin)
    for m in sorted(g1_lines_by_month, key=lambda x: MONTH_IDX.get(x, 99)):
        for inv in g1_lines_by_month[m]:
            if inv["sheet"] != "b2b, sez, de_inv" or not inv["invno"]:
                continue
            if gstin_pan(inv["gstin"]) == self_pan and gstin_state(inv["gstin"]) != gstin_state(self_gstin):
                F.append(Finding("C2", "Branch transfer detected (same PAN, different state)", REVW,
                    f"{m}: Invoice {inv['invno']} to {inv['gstin']} shares PAN {self_pan} with the "
                    f"taxpayer but is registered in a different state. Likely a stock transfer, not "
                    f"an arm's-length sale -- verify ITC reversal / valuation under Rule 28 (related-"
                    f"party / distinct-person supply) rather than treating as ordinary revenue.",
                    numbers=dict(month=m, invno=inv["invno"], gstin=inv["gstin"])))
    if not F:
        F.append(Finding("C2", "Branch transfer detected (same PAN, different state)", PASS,
            "No recipient GSTIN in any B2B invoice shares the taxpayer's own PAN -- no stock-transfer-"
            "as-sale pattern found."))
    return F


def check_export_lut():
    """C3 -- not computable: no LUT/Bond reference data supplied anywhere."""
    return [Finding("C3", "Export/SEZ supply without LUT/Bond reference", INFO,
        "Not computable: none of the supplied files carry an LUT/Bond acknowledgement number or "
        "reference. Would need the LUT filing (RFD-11) record to cross-check zero-rated invoices "
        "against.")]


def check_ewb_distance_route():
    """C4 -- not computable: no distance/route database available offline."""
    return [Finding("C4", "EWB validity vs interstate distance", INFO,
        "Not computable: EWB files carry From/To place text but no distance figure, and there is no "
        "distance-matrix or mapping-API data source available here. A state-to-state distance table "
        "would need to be supplied to check EWB validity-window sufficiency.")]


def check_intra_vs_ewb_interstate(g1_lines_by_month, ewb_out_rows, self_gstin):
    """C5 -- the invoice's implied interstate/intrastate status (Place-Of-
    Supply vs the SUPPLIER's own state) disagrees with the matched Outward
    EWB's implied status (EWB destination state vs the supplier's own state).
    Also feeds #50 (triple mismatch) by carrying the tax-head alongside."""
    F = []
    self_state = gstin_state(self_gstin)
    ewb_by_invno = {}
    for e in ewb_out_rows:
        if e["docno"]:
            ewb_by_invno.setdefault(e["docno"], []).append(e)
    for m in sorted(g1_lines_by_month, key=lambda x: MONTH_IDX.get(x, 99)):
        for inv in g1_lines_by_month[m]:
            if inv["sheet"] != "b2b, sez, de_inv" or not inv["invno"]:
                continue
            matches = ewb_by_invno.get(inv["invno"])
            if not matches:
                continue
            pos_code = state_code_from_name(inv["pos"])
            if pos_code is None:
                continue
            inv_interstate = pos_code != self_state
            for e in matches:
                ewb_to_state = gstin_state(e["to_gstin"])
                if not ewb_to_state:
                    continue
                ewb_interstate = ewb_to_state != self_state
                if inv_interstate != ewb_interstate:
                    has_igst = inv["igst"] > 0.01
                    F.append(Finding("C5/#50", "Invoice interstate-status disagrees with matched EWB movement", FLAG,
                        f"{m}: Invoice {inv['invno']} -- POS {inv['pos']} implies "
                        f"{'inter' if inv_interstate else 'intra'}-state (IGST charged: "
                        f"{'yes' if has_igst else 'no'}), but EWB {e['ewbno']} moves goods to state "
                        f"{ewb_to_state} ({e['to_gstin']}), which is "
                        f"{'inter' if ewb_interstate else 'intra'}-state relative to the supplier. "
                        f"Physical movement contradicts the invoice's declared POS.",
                        numbers=dict(month=m, invno=inv["invno"], pos_state=pos_code,
                                     ewb_to_state=ewb_to_state, igst=inv["igst"])))
    if not F:
        F.append(Finding("C5/#50", "Invoice interstate-status disagrees with matched EWB movement", PASS,
            "For every invoice with a matched Outward EWB, the invoice's Place-of-Supply and the "
            "EWB's destination state agree on inter-state vs intra-state."))
    return F


# ======================================================================
# NUMBERED FRAUD-PATTERN LIST
# ======================================================================
def check_round_numbers(g1_lines_by_month, ewb_out_rows=None):
    """#1 -- exact round taxable value AND exact round tax on the same line."""
    F = []
    ewb_by_invno = {}
    for e in (ewb_out_rows or []):
        if e.get("docno"):
            ewb_by_invno.setdefault(e["docno"], []).append(e)
    for m in sorted(g1_lines_by_month, key=lambda x: MONTH_IDX.get(x, 99)):
        for inv in g1_lines_by_month[m]:
            if inv["sheet"] != "b2b, sez, de_inv" or not inv["invno"]:
                continue
            tax = inv["igst"] + inv["cgst"] + inv["sgst"]
            if inv["taxable"] > 0 and inv["taxable"] % 1000 == 0 and tax > 0 and tax % 100 == 0:
                # Per instruction: attach the matching outward EWB's own detail (EWB no., date,
                # vehicle, HSN, from/to place) where one exists for this invoice, so the reviewer
                # sees the goods-movement evidence alongside the round-number flag, not just the
                # invoice figures.
                ewb_hits = ewb_by_invno.get(str(inv["invno"]).strip().upper(), [])
                e0 = ewb_hits[0] if ewb_hits else {}
                F.append(Finding("#1", "Round-number invoice (accommodation-entry heuristic)", REVW,
                    f"{m}: Invoice {inv['invno']} has an exact-round taxable value "
                    f"(Rs.{inv['taxable']:,.0f}) and exact-round tax (Rs.{tax:,.0f}). Heuristic only -- "
                    f"genuine bulk/contract pricing can also land on round numbers; verify physical "
                    f"supply for this invoice before treating as evidence.",
                    numbers=dict(month=m, invno=inv["invno"], invdate=inv.get("invdate"),
                                 taxable=inv["taxable"], tax=tax, rate=inv.get("rate"),
                                 buyer_gstin=inv.get("gstin"), pos=inv.get("pos"),
                                 ewb_no=e0.get("ewbno", "" if ewb_out_rows is not None else "n/a"),
                                 ewb_date=e0.get("ewbdate"), vehicle=e0.get("vehicle", ""),
                                 hsn=e0.get("hsn", ""), to_place=e0.get("to_place", ""))))
    if not F:
        F.append(Finding("#1", "Round-number invoice (accommodation-entry heuristic)", PASS,
            "No B2B invoice found with both an exact-round taxable value and exact-round tax."))
    return F


def check_below_ewb_threshold(g1_lines_by_month, ewb_out_rows, self_gstin):
    """#2 -- inter-state B2B invoices priced just under Rs.50,000 (EWB not
    mandatory below this) with no EWB generated. Inter-state judged vs the
    supplier's own state (self_gstin), same fix as check_pos_tax_head."""
    F = []
    self_state = gstin_state(self_gstin)
    ewb_invnos = {e["docno"] for e in ewb_out_rows if e["docno"]}
    for m in sorted(g1_lines_by_month, key=lambda x: MONTH_IDX.get(x, 99)):
        for inv in g1_lines_by_month[m]:
            if inv["sheet"] != "b2b, sez, de_inv" or not inv["invno"]:
                continue
            pos_code = state_code_from_name(inv["pos"])
            value = inv["taxable"] + inv["igst"] + inv["cgst"] + inv["sgst"]
            if pos_code and pos_code != self_state and 45000 <= value < 50000 \
               and inv["invno"] not in ewb_invnos:
                F.append(Finding("#2", "Inter-state invoice just below Rs.50K EWB threshold, no EWB", REVW,
                    f"{m}: Invoice {inv['invno']} value Rs.{value:,.2f} (inter-state, POS {inv['pos']}) "
                    f"has no matching EWB. Pattern consistent with deliberately staying under the "
                    f"threshold -- check if this recipient has multiple such invoices.",
                    numbers=dict(month=m, invno=inv["invno"], value=value)))
    if not F:
        F.append(Finding("#2", "Inter-state invoice just below Rs.50K EWB threshold, no EWB", PASS,
            "No inter-state invoice in the Rs.45,000-49,999 band was found without a matching EWB."))
    return F


def check_reciprocal_trading(bo):
    """#3 -- GSTIN appears in BOTH BO Profile's Top-10-Beneficiaries (ITC we
    passed) and Top-10-Suppliers (ITC we received) lists."""
    F = []
    if not bo:
        return [Finding("#3", "Supplier-buyer reciprocity (circular trading)", INFO,
            "BO Profile not supplied -- cannot check.")]
    benef = {b["gstin"]: b for b in bo.get("top_beneficiaries", [])}
    supp = {s["gstin"]: s for s in bo.get("top_suppliers", [])}
    common = set(benef) & set(supp)
    for g in sorted(common):
        b, s = benef[g], supp[g]
        F.append(Finding("#3", "Reciprocal trading (same GSTIN in Top Beneficiaries + Top Suppliers)", REVW,
            f"{g} ({b['name']}) received Rs.{b['amount']:,.2f} Lakh ITC FROM the taxpayer AND supplied "
            f"Rs.{s['amount']:,.2f} Lakh ITC TO the taxpayer (both in the last-12-months BO Profile "
            f"window). Net impact Rs.{b['amount']-s['amount']:,.2f} Lakh -- verify these aren't "
            f"round-tripping/circular-trading invoices.",
            numbers=dict(gstin=g, itc_passed=b["amount"], itc_received=s["amount"])))
    if not common:
        F.append(Finding("#3", "Reciprocal trading (same GSTIN in Top Beneficiaries + Top Suppliers)", PASS,
            "No counterparty appears in both the Top-10-Beneficiaries and Top-10-Suppliers lists."))
    return F


def check_cn_timing(cdnr_by_month, g1_lines_by_month):
    """#4 -- credit-note timing vs the recipient's earliest known invoice.
    APPROXIMATE: cdnr carries no original-invoice reference, so this matches
    by recipient GSTIN only (not a specific invoice) -- always labelled as such."""
    F = []
    first_invoice_month = {}
    for m in sorted(g1_lines_by_month, key=lambda x: MONTH_IDX.get(x, 99)):
        for inv in g1_lines_by_month[m]:
            if inv["sheet"] != "b2b, sez, de_inv":
                continue
            first_invoice_month.setdefault(inv["gstin"], m)
    for m in sorted(cdnr_by_month, key=lambda x: MONTH_IDX.get(x, 99)):
        for cn in cdnr_by_month[m]:
            first_m = first_invoice_month.get(cn["gstin"])
            if not first_m:
                continue
            gap = MONTH_IDX.get(m, 0) - MONTH_IDX.get(first_m, 0)
            if gap >= 5:
                F.append(Finding("#4", "Credit note issued long after recipient's earliest invoice", REVW,
                    f"CN {cn['noteno']} to {cn['gstin']} issued in {m}; this recipient's earliest "
                    f"invoice in the supplied data is {first_m} ({gap} months earlier). APPROXIMATE "
                    f"match (by recipient only, not the specific original invoice -- cdnr has no "
                    f"original-invoice reference) -- verify whether this is a genuine return or a "
                    f"year-end adjustment.", numbers=dict(month=m, noteno=cn["noteno"], gap_months=gap)))
    if not F:
        F.append(Finding("#4", "Credit note issued long after recipient's earliest invoice", INFO,
            "No credit note found with a >=5-month gap to its recipient's earliest invoice in the "
            "supplied data (approximate match only)."))
    return F


def check_hsn_drift(hsn_by_month):
    """#6 -- HSN product mix shifting materially month to month.
    #27 -- brand-new HSN appearing for the first time, especially late in FY."""
    F = []
    months_sorted = sorted(hsn_by_month, key=lambda x: MONTH_IDX.get(x, 99))
    seen_hsns = set()
    monthly_mix = {}
    for m in months_sorted:
        mix = {}
        total = 0.0
        for row in hsn_by_month[m]:
            mix[row["hsn"]] = mix.get(row["hsn"], 0.0) + row["taxable"]
            total += row["taxable"]
        monthly_mix[m] = (mix, total)
        for hsn in mix:
            if hsn not in seen_hsns and m != months_sorted[0]:
                F.append(Finding("#27", "New HSN code appearing mid/late year", REVW,
                    f"HSN {hsn} appears for the first time in {m} (taxable Rs.{mix[hsn]:,.2f}), with "
                    f"no earlier month in the supplied data reporting it. Verify this is a genuine "
                    f"new product line rather than a misclassification shift.",
                    numbers=dict(month=m, hsn=hsn, taxable=mix[hsn])))
        seen_hsns |= set(mix)
    # month-over-month share drift for any HSN with a material (>=30pp) swing
    prev = None
    for m in months_sorted:
        mix, total = monthly_mix[m]
        if total > 0 and prev:
            pmix, ptotal = prev
            for hsn in set(mix) | set(pmix):
                share_now = mix.get(hsn, 0.0) / total * 100
                share_prev = (pmix.get(hsn, 0.0) / ptotal * 100) if ptotal else 0.0
                if abs(share_now - share_prev) >= 30:
                    F.append(Finding("#6", "HSN mix share shifted sharply month-over-month", REVW,
                        f"HSN {hsn} share of monthly taxable value moved from {share_prev:.0f}% to "
                        f"{share_now:.0f}% between the previous month and {m}. Verify this reflects a "
                        f"real business-model change rather than rate arbitrage / misclassification.",
                        numbers=dict(month=m, hsn=hsn, share_prev=share_prev, share_now=share_now)))
        prev = (mix, total)
    return F


def check_itc_liability_volatility(monthly_comparison_rows):
    """#7 -- ITC/Liability ratio volatility across the FY. Consumes the same
    monthly rows already computed by build_annual_workbook (liability + ITC
    per month), no re-derivation needed."""
    F = []
    ratios = []
    for r in monthly_comparison_rows:
        liab = r.get("portal_g3b_liability")
        itc = r.get("portal_itc_3b")
        if liab and liab > 0 and itc is not None:
            ratios.append((r["month"], itc / liab * 100))
    if len(ratios) < 3:
        return [Finding("#7", "ITC/Liability monthly volatility", INFO,
            "Not enough months with usable liability+ITC figures to compute a volatility index.")]
    vals = [v for _, v in ratios]
    mean_r = _stats.mean(vals)
    stdev_r = _stats.pstdev(vals) if len(vals) > 1 else 0.0
    for m, v in ratios:
        if stdev_r > 0 and abs(v - mean_r) >= 2 * stdev_r:
            F.append(Finding("#7", "ITC/Liability ratio spike vs FY average", REVW,
                f"{m}: ITC/Liability ratio {v:.0f}% vs FY average {mean_r:.0f}% (std dev {stdev_r:.0f}pp) "
                f"-- more than 2 standard deviations away. Suspected temporary ITC inflation or a "
                f"liability dip; verify the month's underlying invoices.",
                numbers=dict(month=m, ratio=v, fy_avg=mean_r)))
    if not F:
        F.append(Finding("#7", "ITC/Liability ratio spike vs FY average", PASS,
            f"No month's ITC/Liability ratio deviates more than 2 standard deviations from the FY "
            f"average ({mean_r:.0f}%)."))
    return F


def check_year_end_dumping(g1_lines_by_month):
    """#8 -- disproportionate share of annual B2B sales in the last 15 days
    of March (window-dressing / turnover-target pattern)."""
    F = []
    total = 0.0
    last15 = 0.0
    for m, invs in g1_lines_by_month.items():
        for inv in invs:
            if inv["sheet"] != "b2b, sez, de_inv" or not inv["invdate"]:
                continue
            total += inv["taxable"]
            if inv["invdate"].year == 2023 and inv["invdate"].month == 3 and inv["invdate"].day >= 17:
                last15 += inv["taxable"]
    if total > 0:
        pct = last15 / total * 100
        sev = FLAG if pct >= 40 else (REVW if pct >= 25 else PASS)
        F.append(Finding("#8", "Year-end (last 15 days of March) sales concentration", sev,
            f"Last-15-days-of-March B2B taxable value Rs.{last15:,.2f} = {pct:.1f}% of full-FY B2B "
            f"taxable value Rs.{total:,.2f}. {'Unnaturally concentrated -- potential window dressing.' if sev != PASS else 'Within a normal range.'}",
            numbers=dict(last15=last15, total=total, pct=pct)))
    else:
        F.append(Finding("#8", "Year-end (last 15 days of March) sales concentration", INFO,
            "No B2B invoices with parseable dates found across the supplied months."))
    return F


def check_zero_cash_months(cash_monthly):
    """#9 -- months where cash-ledger debits (actual cash tax paid) are zero."""
    F = []
    zero_months = [m for m, v in cash_monthly.items() if v.get("debited", 0) == 0]
    total_months = len(cash_monthly)
    if total_months and len(zero_months) >= total_months * 0.5:
        F.append(Finding("#9", "Zero cash tax paid in majority of months", REVW,
            f"Zero cash-ledger debit (no cash tax paid, fully ITC-settled) in {len(zero_months)} of "
            f"{total_months} months with ledger activity: {', '.join(sorted(zero_months, key=lambda x: MONTH_IDX.get(x,99)))}. "
            f"Entirely ITC-funded liability isn't inherently wrong for an ITC-heavy business, but "
            f"verify against the ITC's own genuineness (2B match rate).",
            numbers=dict(zero_months=len(zero_months), total_months=total_months)))
    else:
        F.append(Finding("#9", "Zero cash tax paid in majority of months", PASS,
            f"Zero-cash-debit in {len(zero_months)} of {total_months} months -- not a majority."))
    return F


def check_ghost_supplier_cluster(bo):
    """#10/#19 -- cancelled counterparties sharing the same PAN (GSTIN chars
    3-12), across BOTH the related-ITC-received and related-ITC-passed lists,
    plus the Top-10 lists (all already GSTIN-keyed, PAN derivable directly)."""
    F = []
    if not bo:
        return [Finding("#10/#19", "Duplicate/ghost-entity GSTIN cluster (shared PAN)", INFO,
            "BO Profile not supplied -- cannot check.")]
    all_parties = {}
    for lst, amtkey in ((bo.get("related_itc_received", []), "total_itc"),
                        (bo.get("related_itc_passed", []), "total_itc"),
                        (bo.get("top_suppliers", []), "amount"),
                        (bo.get("top_beneficiaries", []), "amount")):
        for p in lst:
            g = p.get("gstin")
            if not g:
                continue
            all_parties.setdefault(gstin_pan(g), []).append(
                dict(gstin=g, name=p.get("name"), status=p.get("status", ""), amount=p.get(amtkey, 0.0)))
    for pan, entries in all_parties.items():
        gstins = {e["gstin"] for e in entries}
        if len(gstins) < 2:
            continue
        cancelled = [e for e in entries if e.get("status") == "Cancelled"]
        if cancelled and len(gstins) >= 2:
            total_amt = sum(e["amount"] for e in entries)
            F.append(Finding("#10/#19", "Duplicate/ghost-entity GSTIN cluster (shared PAN)", FLAG,
                f"PAN {pan}: {len(gstins)} GSTINs found ({', '.join(sorted(gstins))}), at least one "
                f"Cancelled. Combined ITC across the cluster (SUMMED ACROSS ALL YEARS shown in the BO "
                f"Profile's related-party table, not limited to FY2022-23): Rs.{total_amt:,.2f} Lakh. "
                f"Verify these aren't the same entity re-registering under a new GSTIN, and check the "
                f"per-year breakdown before citing this figure for the FY under scrutiny specifically.",
                numbers=dict(pan=pan, gstins=sorted(gstins), total=total_amt)))
    if not any(f.severity == FLAG for f in F):
        F.append(Finding("#10/#19", "Duplicate/ghost-entity GSTIN cluster (shared PAN)", PASS,
            "No PAN in the BO Profile's related-party/Top-10 lists maps to more than one GSTIN with "
            "at least one Cancelled registration."))
    return F


def check_cn_vs_inward_ewb(cdnr_by_month, ewb_in_rows):
    """#11 -- credit note issued (implies goods returned) with no matching
    Inward EWB of comparable value from the same counterparty. APPROXIMATE:
    matched by GSTIN + value window, not a specific document reference."""
    F = []
    ewb_in_by_gstin = {}
    for e in ewb_in_rows:
        ewb_in_by_gstin.setdefault(e["from_gstin"], []).append(e)
    for m in sorted(cdnr_by_month, key=lambda x: MONTH_IDX.get(x, 99)):
        for cn in cdnr_by_month[m]:
            note_value = cn["taxable"] + cn["igst"] + cn["cgst"] + cn["sgst"]
            candidates = ewb_in_by_gstin.get(cn["gstin"], [])
            match = any(abs((e["assess"] + e["taxval"]) - note_value) / note_value < 0.15
                        for e in candidates if note_value)
            if candidates and not match and note_value > 10000:
                F.append(Finding("#11", "Credit note with no comparable Inward EWB (paper CN risk)", REVW,
                    f"{m}: CN {cn['noteno']} to {cn['gstin']} for Rs.{note_value:,.2f} -- no Inward EWB "
                    f"from this GSTIN in the supplied data is within 15% of that value. APPROXIMATE "
                    f"check only (matched by counterparty+value, not a document reference) -- verify "
                    f"whether goods physically returned.",
                    numbers=dict(month=m, noteno=cn["noteno"], value=note_value)))
    if not F:
        F.append(Finding("#11", "Credit note with no comparable Inward EWB (paper CN risk)", INFO,
            "Either every CN found a comparable Inward EWB from the same counterparty, or no "
            "counterparty in the CN data had any Inward EWB at all to compare against (approximate "
            "check -- treat PASS-like results here cautiously)."))
    return F


def check_cross_fy_shift(g1_lines_by_month, ewb_out_rows):
    """#12 -- invoice dated 31-Mar but its matched EWB generated 1-Apr (or
    later, next FY)."""
    F = []
    ewb_by_invno = {}
    for e in ewb_out_rows:
        if e["docno"]:
            ewb_by_invno.setdefault(e["docno"], []).append(e)
    mar23 = g1_lines_by_month.get("Mar-23", [])
    for inv in mar23:
        if inv["sheet"] != "b2b, sez, de_inv" or not inv["invdate"]:
            continue
        if inv["invdate"] != _dt.date(2023, 3, 31):
            continue
        for e in ewb_by_invno.get(inv["invno"], []):
            if e["ewbdate"] and e["ewbdate"] >= _dt.date(2023, 4, 1):
                F.append(Finding("#12", "Invoice dated 31-Mar, EWB generated in the next FY", FLAG,
                    f"Invoice {inv['invno']} dated 31-Mar-2023, but EWB {e['ewbno']} generated on "
                    f"{e['ewbdate']}. Cross-FY turnover-shifting risk -- verify actual supply date "
                    f"and AS-9/Ind AS 115 revenue recognition.",
                    numbers=dict(invno=inv["invno"], ewbdate=str(e["ewbdate"]))))
    if not F:
        F.append(Finding("#12", "Invoice dated 31-Mar, EWB generated in the next FY", PASS,
            "No 31-Mar-2023 invoice has a matched EWB dated 1-Apr-2023 or later."))
    return F


def check_irn_after_filing(g1_filing_dates, einv_lines_by_month, ewb_out_rows=None):
    """#15 -- IRN generated AFTER the GSTR-1 filing date for that month
    (sequence violation -- return was filed before the e-invoice existed)."""
    F = []
    ewb_by_invno = {}
    for e in (ewb_out_rows or []):
        if e.get("docno"):
            ewb_by_invno.setdefault(e["docno"], []).append(e)
    for m, einvs in einv_lines_by_month.items():
        filing_date = g1_filing_dates.get(m)
        if not filing_date:
            continue
        for e in einvs:
            if e["irndate"] and e["irndate"] > filing_date:
                e0 = ewb_by_invno.get(str(e["invno"]).strip().upper(), [{}])[0]
                F.append(Finding("#15", "IRN generated after GSTR-1 filing (sequence violation)", FLAG,
                    f"{m}: Invoice {e['invno']} IRN generated {e['irndate']}, but GSTR-1 for {m} was "
                    f"filed on {filing_date}. Return was filed before this e-invoice existed -- "
                    f"invalid sequence.",
                    numbers=dict(month=m, invno=e["invno"], invdate=e.get("invdate"),
                                 irndate=e["irndate"], gstr1_filed=filing_date,
                                 taxable=e.get("taxable"), tax=e.get("igst", 0) + e.get("cgst", 0) + e.get("sgst", 0),
                                 recipient_gstin=e.get("gstin"),
                                 ewb_no=e0.get("ewbno", ""), ewb_date=e0.get("ewbdate"),
                                 vehicle=e0.get("vehicle", ""))))
    if not F:
        F.append(Finding("#15", "IRN generated after GSTR-1 filing (sequence violation)", PASS,
            "Every matched e-invoice's IRN date is on/before its month's GSTR-1 filing date."))
    return F


def check_credit_hoarding(credit_monthly, cash_monthly):
    """#17 -- large credit-ledger balance retained while cash is still being
    paid unnecessarily (uses the running 'net' balance already computed by
    annual_sources.parse_credit_ledger's monthly_by_tax_period)."""
    F = []
    total_cash_paid = sum(v.get("debited", 0) for v in cash_monthly.values())
    # approximate closing credit balance = opening (0 baseline) + cumulative net
    running = 0.0
    peak = 0.0
    for m in sorted(credit_monthly, key=lambda x: MONTH_IDX.get(x, 99)):
        running += credit_monthly[m].get("net", 0.0)
        peak = max(peak, running)
    if peak > 0 and total_cash_paid > 0 and peak >= total_cash_paid * 2:
        F.append(Finding("#17", "Credit-ledger balance high relative to cash actually paid", REVW,
            f"Peak running Credit Ledger balance (approx, from monthly net movements) Rs.{peak:,.2f} "
            f"vs total cash tax paid across the FY Rs.{total_cash_paid:,.2f}. Cash paid despite a "
            f"materially larger ITC cushion available -- worth understanding why cash was used instead "
            f"of ITC in those months (could be legitimate: RCM must be paid in cash, ITC set-off order "
            f"rules, etc. -- not automatically suspicious).",
            numbers=dict(peak_credit=peak, total_cash=total_cash_paid)))
    else:
        F.append(Finding("#17", "Credit-ledger balance high relative to cash actually paid", PASS,
            f"Peak approx. running credit balance Rs.{peak:,.2f} is not materially larger than total "
            f"cash paid Rs.{total_cash_paid:,.2f}."))
    return F


def check_negative_itc_reversal(gstr3b_monthly_fields):
    """#18 -- ITC reversal exceeding the ITC it's reversing against, per month
    (arithmetic impossibility -- points to a Rule 42/43 calculation error)."""
    F = []
    for m, f in gstr3b_monthly_fields.items():
        if f is None:
            continue
        reversed_total = f["itc_reversed_rule42_43"] + f["itc_reversed_others"]
        if reversed_total > f["itc_all_other_igst"] and f["itc_all_other_igst"] >= 0:
            F.append(Finding("#18", "ITC reversal exceeds ITC accrued in the same head (IGST)", FLAG,
                f"{m}: ITC Reversed (Rule 42/43 + Others) Rs.{reversed_total:,.2f} > All-other-ITC "
                f"accrued Rs.{f['itc_all_other_igst']:,.2f} (IGST head). Arithmetically implies "
                f"negative net ITC -- verify the Rule 42/43 reversal calculation.",
                numbers=dict(month=m, reversed=reversed_total, accrued=f["itc_all_other_igst"])))
    if not F:
        F.append(Finding("#18", "ITC reversal exceeds ITC accrued in the same head (IGST)", PASS,
            "No month's IGST-head ITC reversal exceeds that month's IGST-head ITC accrued."))
    return F


def check_midnight_ewb(ewb_out_rows, ewb_in_rows):
    """#21 -- EWB generated at exactly 00:00:00."""
    F = []
    hits = [e for e in (ewb_out_rows + ewb_in_rows) if e.get("ewbtime") == _dt.time(0, 0, 0)]
    if hits:
        F.append(Finding("#21", "EWB(s) generated at exactly midnight (00:00:00)", REVW,
            f"{len(hits)} EWB(s) generated at exactly 00:00:00: "
            f"{', '.join(e['ewbno'] for e in hits[:10])}{' ...' if len(hits) > 10 else ''}. "
            f"Suspected automated/bot generation or a system-default timestamp -- verify physical "
            f"movement for these consignments.", numbers=dict(count=len(hits))))
    else:
        F.append(Finding("#21", "EWB(s) generated at exactly midnight (00:00:00)", PASS,
            "No EWB (inward or outward) generated at exactly 00:00:00."))
    return F


def check_irn_delay(einv_lines_by_month, ewb_out_rows=None):
    """#23 -- IRN generated materially after the invoice date it belongs to."""
    F = []
    ewb_by_invno = {}
    for e in (ewb_out_rows or []):
        if e.get("docno"):
            ewb_by_invno.setdefault(e["docno"], []).append(e)
    for m, einvs in einv_lines_by_month.items():
        for e in einvs:
            if e["invdate"] and e["irndate"] and e["irndate"] > e["invdate"]:
                gap = (e["irndate"] - e["invdate"]).days
                if gap >= 3:
                    # Per instruction: IRN + invoice complete detail, PLUS e-way bill details for
                    # this same invoice (EWB no./date/vehicle where a matching outward EWB exists).
                    e0 = ewb_by_invno.get(str(e["invno"]).strip().upper(), [{}])[0]
                    F.append(Finding("#23", "IRN generated materially after the invoice date", FLAG if gap >= 7 else REVW,
                        f"{m}: Invoice {e['invno']} dated {e['invdate']}, IRN generated {e['irndate']} "
                        f"({gap} days later). E-invoice rules require IRN before/at supply -- treat "
                        f"invoice as invalid if this pattern is confirmed.",
                        numbers=dict(month=m, invno=e["invno"], invdate=e["invdate"], irndate=e["irndate"],
                                     gap_days=gap, taxable=e["taxable"],
                                     tax=e["igst"] + e["cgst"] + e["sgst"], recipient_gstin=e["gstin"],
                                     pos=e.get("pos", ""),
                                     ewb_no=e0.get("ewbno", ""), ewb_date=e0.get("ewbdate"),
                                     vehicle=e0.get("vehicle", ""), hsn=e0.get("hsn", ""))))
    if not F:
        F.append(Finding("#23", "IRN generated materially after the invoice date", PASS,
            "No e-invoice has an IRN date 3+ days after its invoice date."))
    return F


def check_ewb_state_shift(ewb_out_rows):
    """#24 -- sudden shift in destination-state mix month to month."""
    F = []
    monthly_state_mix = {}
    for e in ewb_out_rows:
        if not e["month"]:
            continue
        st = gstin_state(e["to_gstin"])
        monthly_state_mix.setdefault(e["month"], {}).setdefault(st, 0.0)
        monthly_state_mix[e["month"]][st] += e["assess"]
    months_sorted = sorted(monthly_state_mix, key=lambda x: MONTH_IDX.get(x, 99))
    prev = None
    for m in months_sorted:
        mix = monthly_state_mix[m]
        total = sum(mix.values())
        if prev and total > 0:
            pmix, ptotal = prev
            for st in set(mix) | set(pmix):
                now_share = mix.get(st, 0.0) / total * 100
                prev_share = (pmix.get(st, 0.0) / ptotal * 100) if ptotal else 0.0
                if abs(now_share - prev_share) >= 40:
                    F.append(Finding("#24", "Sudden shift in EWB destination-state mix", REVW,
                        f"Destination-state {st}'s share of Outward EWB assessable value moved from "
                        f"{prev_share:.0f}% to {now_share:.0f}% between the previous month and {m}. "
                        f"Verify actual delivery locations / contracts rather than assuming routine "
                        f"business variation.", numbers=dict(month=m, state=st, prev=prev_share, now=now_share)))
        prev = (mix, total)
    if not F:
        F.append(Finding("#24", "Sudden shift in EWB destination-state mix", PASS,
            "No destination state's month-over-month EWB value share moved by 40 percentage points "
            "or more."))
    return F


def check_exempt_turnover_rule42(bo, gstr3b_monthly_fields):
    """#25 -- exempt turnover (Turnover - Taxable Turnover, from BO Profile
    Financial Information) vs Rule 42/43 reversal actually posted in 3B."""
    F = []
    if not bo:
        return [Finding("#25", "Exempt turnover vs Rule 42 ITC reversal", INFO, "BO Profile not supplied.")]
    fin = bo.get("financial_by_fy", {}).get("2022-23")
    if not fin or fin.get("turnover") is None or fin.get("taxable_turnover") is None:
        return [Finding("#25", "Exempt turnover vs Rule 42 ITC reversal", INFO,
            "FY2022-23 row not found/incomplete in BO Profile Financial Information.")]
    exempt = fin["turnover"] - fin["taxable_turnover"]
    exempt_pct = (exempt / fin["turnover"] * 100) if fin["turnover"] else 0.0
    zero_reversal_months = [m for m, f in gstr3b_monthly_fields.items()
                             if f and f["itc_reversed_rule42_43"] == 0]
    if exempt_pct >= 1 and zero_reversal_months:
        F.append(Finding("#25", "Exempt turnover present but Rule 42 reversal is zero in some months", REVW,
            f"BO Profile FY2022-23: Turnover Rs.{fin['turnover']:.2f}L vs Taxable Turnover "
            f"Rs.{fin['taxable_turnover']:.2f}L implies exempt turnover ~{exempt_pct:.1f}%. Rule 42/43 "
            f"ITC reversal (4B1) is exactly Rs.0 in {len(zero_reversal_months)} month(s): "
            f"{', '.join(sorted(zero_reversal_months, key=lambda x: MONTH_IDX.get(x,99)))}. "
            f"If exempt supplies genuinely occurred in those months, a proportionate reversal is "
            f"mandatory -- verify.", numbers=dict(exempt_pct=exempt_pct, zero_months=len(zero_reversal_months))))
    else:
        F.append(Finding("#25", "Exempt turnover vs Rule 42 ITC reversal", PASS,
            f"BO Profile implies exempt turnover ~{exempt_pct:.1f}% of FY2022-23 turnover -- "
            f"immaterial or reversal is being posted; no finding."))
    return F


def check_sunday_holiday_ewb(ewb_out_rows, ewb_in_rows):
    """#37 -- share of EWBs generated on a Sunday or a fixed-date national
    holiday."""
    all_ewb = [e for e in (ewb_out_rows + ewb_in_rows) if e.get("ewbdate")]
    if not all_ewb:
        return [Finding("#37", "Sunday/national-holiday EWB generation share", INFO, "No dated EWB rows found.")]
    flagged = [e for e in all_ewb if e["ewbdate"].weekday() == 6 or e["ewbdate"] in NATIONAL_HOLIDAYS_FY2223]
    pct = len(flagged) / len(all_ewb) * 100
    sev = FLAG if pct >= 30 else (REVW if pct >= 15 else PASS)
    return [Finding("#37", "Sunday/national-holiday EWB generation share", sev,
        f"{len(flagged)} of {len(all_ewb)} EWBs ({pct:.1f}%) generated on a Sunday or a fixed national "
        f"holiday (15-Aug/2-Oct/26-Jan only -- regional/lunar festival holidays not included, so this "
        f"under-counts rather than over-counts). "
        f"{'Suspiciously high -- verify genuine dispatch.' if sev != PASS else 'Within a normal range.'}",
        numbers=dict(flagged=len(flagged), total=len(all_ewb), pct=pct))]


def check_consecutive_ewb_burst(ewb_out_rows):
    """#41 -- adapted to use EWB timestamps (invoices in this data have no
    time-of-day at all): tight clusters of EWB generation on the same day."""
    F = []
    dt_rows = [e for e in ewb_out_rows if e.get("ewbdate") and e.get("ewbtime")]
    dt_rows.sort(key=lambda e: (e["ewbdate"], e["ewbtime"]))
    i = 0
    while i < len(dt_rows):
        j = i
        cluster = [dt_rows[i]]
        while j + 1 < len(dt_rows):
            t0 = _dt.datetime.combine(dt_rows[j]["ewbdate"], dt_rows[j]["ewbtime"])
            t1 = _dt.datetime.combine(dt_rows[j + 1]["ewbdate"], dt_rows[j + 1]["ewbtime"])
            if (t1 - t0).total_seconds() <= 180:
                cluster.append(dt_rows[j + 1])
                j += 1
            else:
                break
        if len(cluster) >= 8:
            F.append(Finding("#41", "EWB generation burst (many EWBs within 3 minutes)", REVW,
                f"{len(cluster)} EWBs generated within 3 minutes of each other, starting "
                f"{cluster[0]['ewbdate']} {cluster[0]['ewbtime']}: "
                f"{', '.join(c['ewbno'] for c in cluster[:8])}{' ...' if len(cluster) > 8 else ''}. "
                f"Suspected automated batch generation -- verify physical dispatch for these.",
                numbers=dict(count=len(cluster), start=str(cluster[0]["ewbdate"]))))
        i = j + 1
    if not F:
        F.append(Finding("#41", "EWB generation burst (many EWBs within 3 minutes)", PASS,
            "No cluster of 8+ EWBs generated within 3 minutes of each other."))
    return F


def check_related_party_supercluster(bo):
    """#43 -- basic (2-hop) PAN-cluster network from the BO Profile's own
    related-party + Top-10 lists (not full graph-theory/networkx -- that
    would need a much larger counterparty graph than these lists provide)."""
    if not bo:
        return [Finding("#43", "Related-party PAN super-cluster (basic)", INFO, "BO Profile not supplied.")]
    pan_map = {}
    for lst in (bo.get("related_itc_received", []), bo.get("related_itc_passed", []),
                bo.get("top_suppliers", []), bo.get("top_beneficiaries", [])):
        for p in lst:
            g = p.get("gstin")
            if g:
                pan_map.setdefault(gstin_pan(g), set()).add(g)
    clusters = {pan: gs for pan, gs in pan_map.items() if len(gs) >= 3}
    F = []
    for pan, gs in clusters.items():
        F.append(Finding("#43", "Related-party PAN super-cluster (3+ GSTINs)", REVW,
            f"PAN {pan} maps to {len(gs)} distinct GSTINs across the BO Profile's related-party/Top-10 "
            f"lists: {', '.join(sorted(gs))}. Basic 2-hop check only (not a full transaction-graph "
            f"analysis) -- worth a closer look at the relationship between these entities.",
            numbers=dict(pan=pan, gstins=sorted(gs))))
    if not F:
        F.append(Finding("#43", "Related-party PAN super-cluster (3+ GSTINs)", PASS,
            "No PAN maps to 3 or more distinct GSTINs across the BO Profile's related-party/Top-10 lists."))
    return F


def check_ewb_invoice_date_gap(g1_lines_by_month, ewb_out_rows):
    """#46 -- EWB generated more than 15 days before/after its invoice date."""
    F = []
    ewb_by_invno = {}
    for e in ewb_out_rows:
        if e["docno"]:
            ewb_by_invno.setdefault(e["docno"], []).append(e)
    for m, invs in g1_lines_by_month.items():
        for inv in invs:
            if inv["sheet"] != "b2b, sez, de_inv" or not inv["invdate"]:
                continue
            for e in ewb_by_invno.get(inv["invno"], []):
                if not e["ewbdate"]:
                    continue
                gap = (e["ewbdate"] - inv["invdate"]).days
                if abs(gap) > 15:
                    F.append(Finding("#46", "EWB vs invoice date gap exceeds 15 days", REVW,
                        f"Invoice {inv['invno']} dated {inv['invdate']}, EWB {e['ewbno']} generated "
                        f"{e['ewbdate']} ({gap:+d} days). Rule 138 anticipates EWB close to the supply "
                        f"date -- verify actual supply timing.",
                        numbers=dict(invno=inv["invno"], gap_days=gap)))
    if not F:
        F.append(Finding("#46", "EWB vs invoice date gap exceeds 15 days", PASS,
            "No matched EWB/invoice pair has a date gap exceeding 15 days."))
    return F


def check_cash_timing(cash_transactions, gstr3b_monthly_fields):
    """#47/#55 -- cash-ledger deposits happening right at/after the GSTR-3B
    due date (20th) or right before/at the actual filing date, repeated
    across several months (last-minute cash arrangement pattern)."""
    F = []
    late_months = []
    for txn in cash_transactions:
        d = None
        for fmt in ("%d-%m-%Y", "%d/%m/%Y"):
            try:
                d = _dt.datetime.strptime(txn["date"], fmt).date()
                break
            except (ValueError, TypeError):
                continue
        if not d:
            continue
        if d.day >= GSTR3B_DUE_DOM - 1:
            key = f"{d.year}-{d.month:02d}"
            late_months.append(key)
    from collections import Counter
    c = Counter(late_months)
    repeat_months = {k: v for k, v in c.items() if v >= 1}
    if len(repeat_months) >= 3:
        F.append(Finding("#47/#55", "Cash deposited right at/after the due date, repeatedly", REVW,
            f"Cash-ledger deposit transactions fall on/after day {GSTR3B_DUE_DOM-1} of the month in "
            f"{len(repeat_months)} distinct months. Pattern consistent with last-minute cash "
            f"arrangement rather than planned tax payment -- interest u/s 50 may apply for any month "
            f"where this crossed the actual due date.", numbers=dict(months=len(repeat_months))))
    else:
        F.append(Finding("#47/#55", "Cash deposited right at/after the due date, repeatedly", PASS,
            f"Late-in-month cash deposits found in only {len(repeat_months)} month(s) -- not a "
            f"repeated pattern."))
    return F


def check_irn_ewb_invoice_gap(g1_lines_by_month, einv_lines_by_month, ewb_out_rows):
    """#49 -- three-way date gap: invoice date vs IRN date vs EWB date, all
    disagreeing by a material margin."""
    F = []
    einv_by_invno = {}
    for m, lst in einv_lines_by_month.items():
        for e in lst:
            einv_by_invno[e["invno"]] = e
    ewb_by_invno = {}
    for e in ewb_out_rows:
        if e["docno"]:
            ewb_by_invno.setdefault(e["docno"], []).append(e)
    for m, invs in g1_lines_by_month.items():
        for inv in invs:
            if inv["sheet"] != "b2b, sez, de_inv" or not inv["invdate"]:
                continue
            einv = einv_by_invno.get(inv["invno"])
            ewbs = ewb_by_invno.get(inv["invno"], [])
            if not einv or not einv["irndate"] or not ewbs:
                continue
            for e in ewbs:
                if not e["ewbdate"]:
                    continue
                gaps = sorted({inv["invdate"], einv["irndate"], e["ewbdate"]})
                spread = (gaps[-1] - gaps[0]).days
                if spread >= 3:
                    F.append(Finding("#49", "Invoice / IRN / EWB three-way date gap", REVW,
                        f"Invoice {inv['invno']}: Invoice date {inv['invdate']}, IRN date "
                        f"{einv['irndate']}, EWB date {e['ewbdate']} -- spread of {spread} days across "
                        f"the three. Verify actual sequencing of sale / e-invoice / dispatch.",
                        numbers=dict(invno=inv["invno"], spread_days=spread)))
    if not F:
        F.append(Finding("#49", "Invoice / IRN / EWB three-way date gap", PASS,
            "No matched invoice/IRN/EWB triple has a spread of 3+ days."))
    return F


def check_b2b_matching_accuracy(g1_lines_by_month, einv_lines_by_month):
    """#51 -- % of GSTR-1 B2B invoices with an exact taxable-value match in
    the E-Invoice data (by invoice number)."""
    total = 0
    matched = 0
    for m, invs in g1_lines_by_month.items():
        einv_by_no = {e["invno"]: e for e in einv_lines_by_month.get(m, [])}
        for inv in invs:
            if inv["sheet"] != "b2b, sez, de_inv":
                continue
            total += 1
            e = einv_by_no.get(inv["invno"])
            if e and abs(e["taxable"] - inv["taxable"]) < 1:
                matched += 1
    if total == 0:
        return [Finding("#51", "GSTR-1 vs E-Invoice B2B exact-match accuracy", INFO,
            "No B2B invoices found to compute an accuracy index.")]
    pct = matched / total * 100
    sev = FLAG if pct < 80 else PASS
    return [Finding("#51", "GSTR-1 vs E-Invoice B2B exact-match accuracy", sev,
        f"{matched} of {total} B2B invoices ({pct:.1f}%) have an exact taxable-value match in the "
        f"E-Invoice data by invoice number. {'Below the 80% comfort threshold -- check for systematic '
        'data-entry gaps or E-Invoice coverage gaps (E-Invoice file may simply not cover all months).' if sev==FLAG else ''}",
        numbers=dict(matched=matched, total=total, pct=pct))]


def check_bo_high_risk_supplier_itc(bo):
    """#54 -- ITC concentration from BO-Profile-flagged HIGH-risk suppliers."""
    if not bo:
        return [Finding("#54", "ITC from BO-Profile HIGH-risk suppliers", INFO, "BO Profile not supplied.")]
    high = [s for s in bo.get("top_suppliers", []) if "HIGH" in (s.get("risk") or "")]
    total_itc = sum(s["amount"] for s in high)
    if high:
        names = ", ".join(f"{s['name']} ({s['gstin']})" for s in high)
        return [Finding("#54", "ITC from BO-Profile HIGH-risk suppliers", REVW if total_itc < 100 else FLAG,
            f"BO Profile HIGH-risk suppliers ({names}) provided a combined Rs.{total_itc:,.2f} Lakh ITC "
            f"(last-12-months window). Consider closer verification / blocking review under Sec 16 "
            f"for this ITC.", numbers=dict(total_itc=total_itc, count=len(high)))]
    return [Finding("#54", "ITC from BO-Profile HIGH-risk suppliers", PASS,
        "No Top-10 supplier is flagged HIGH risk in the BO Profile.")]


def check_gstr1_3b_headwise(monthly_comparison_rows):
    """#56 -- head-wise (IGST/CGST/SGST) GSTR-1 vs 3B match, not just the
    aggregate total. Expects each row to optionally carry per-head figures;
    degrades to INFO if the comparison rows don't carry head-wise detail."""
    sample = monthly_comparison_rows[0] if monthly_comparison_rows else {}
    if not any(k in sample for k in ("g1_igst", "g3b_igst")):
        return [Finding("#56", "GSTR-1 vs 3B head-wise (IGST/CGST/SGST) match", INFO,
            "The monthly comparison rows passed in don't carry per-head (IGST/CGST/SGST) figures "
            "separately -- this check needs those alongside the aggregate total already reconciled "
            "elsewhere in the pipeline. Wire in gst_scrutiny_tool's g1/g3b per-head dict to activate.")]
    F = []
    for r in monthly_comparison_rows:
        for head in ("igst", "cgst", "sgst"):
            g1v = r.get(f"g1_{head}", 0.0)
            g3bv = r.get(f"g3b_{head}", 0.0)
            if abs(g1v - g3bv) > 100:
                F.append(Finding("#56", f"GSTR-1 vs 3B {head.upper()} mismatch", FLAG,
                    f"{r.get('month')}: GSTR-1 {head.upper()} Rs.{g1v:,.2f} vs GSTR-3B {head.upper()} "
                    f"Rs.{g3bv:,.2f} (diff Rs.{g1v-g3bv:,.2f}).",
                    numbers=dict(month=r.get("month"), head=head, diff=g1v - g3bv)))
    return F or [Finding("#56", "GSTR-1 vs 3B head-wise (IGST/CGST/SGST) match", PASS,
        "No head-wise (IGST/CGST/SGST) mismatch exceeding Rs.100 in any month.")]


def check_rate_outliers(g1_lines_by_month):
    """#57 -- invoices whose rate is a statistical outlier vs that month's
    dominant rate (e.g. 1% of invoices at 18% while 99% sit at 12%)."""
    F = []
    for m in sorted(g1_lines_by_month, key=lambda x: MONTH_IDX.get(x, 99)):
        rates = [inv["rate"] for inv in g1_lines_by_month[m]
                 if inv["sheet"] == "b2b, sez, de_inv" and inv["rate"] > 0]
        if len(rates) < 10:
            continue
        from collections import Counter
        counts = Counter(rates)
        dominant_rate, dominant_n = counts.most_common(1)[0]
        for rate, n in counts.items():
            if rate != dominant_rate and n / len(rates) <= 0.02:
                F.append(Finding("#57", "Outlier tax rate vs month's dominant rate", REVW,
                    f"{m}: {n} of {len(rates)} invoices ({n/len(rates)*100:.1f}%) billed at {rate}% "
                    f"while {dominant_n} ({dominant_n/len(rates)*100:.1f}%) sit at {dominant_rate}%. "
                    f"Worth checking these {n} invoice(s) for misclassification.",
                    numbers=dict(month=m, outlier_rate=rate, outlier_count=n, dominant_rate=dominant_rate)))
    if not F:
        F.append(Finding("#57", "Outlier tax rate vs month's dominant rate", PASS,
            "No month has a rate used by <=2% of its invoices against a clear dominant rate."))
    return F


def check_2b_itc_unavailable_flag(gstr2b_path, months):
    """B3 -- GSTR-2B's own per-invoice 'ITC Availability' flag + Reason. This
    is fully computable and distinct from A5 (HSN-based blocked-ITC, which
    genuinely isn't computable -- 2B carries no HSN). Confirmed against the
    real file: this taxpayer has 2 such lines, both reason 'POS and supplier
    state are same but recipient state is different'."""
    F = []
    for m in months:
        try:
            parsed = gstr2b_parser.parse_2b_excel(gstr2b_path, m)
        except Exception:
            continue
        for b in parsed.get("b2b", []):
            if b.get("itc_avail", "").strip().lower() == "no":
                tax = b["igst"] + b["cgst"] + b["sgst"] + b["cess"]
                F.append(Finding("B3", "GSTR-2B invoice flagged ITC Availability = No", REVW,
                    f"{m}: Invoice {b['invno']} from {b['gstin']} ({b['supplier']}), taxable "
                    f"Rs.{b['taxable']:,.2f}, tax Rs.{tax:,.2f} -- 2B itself marks this ITC as "
                    f"unavailable. Reason: {b.get('itc_avail_reason') or '(not stated)'}. Verify this "
                    f"amount was excluded from the month's 4(A)(5) claim.",
                    numbers=dict(month=m, invno=b["invno"], invdate=b.get("date"), gstin=b["gstin"],
                                 supplier=b.get("supplier", ""), taxable=b["taxable"], tax=tax,
                                 reason=b.get("itc_avail_reason") or "(not stated)")))
    if not F:
        F.append(Finding("B3", "GSTR-2B invoice flagged ITC Availability = No", PASS,
            "No GSTR-2B B2B line in the supplied months is flagged 'ITC Availability = No'."))
    return F


# ======================================================================
# NOT-FEASIBLE STRUCTURAL NOTES (kept as INFO findings for completeness --
# so the output workbook documents WHY these are absent, not just silence)
# ======================================================================
def not_feasible_notes():
    notes = [
        ("#13", "B2C invoice splitting (Table 5 vs 7)",
         "GSTR-1's B2C-Small (Table 7) sheet is a state+rate SUMMARY with no invoice numbers at all "
         "-- this is how the return itself is structured, not a parsing gap. Cannot detect invoice-"
         "level splitting from GSTR-1 data."),
        ("#14", "EWB high-cancellation-rate detector",
         "The supplied EWB files carry no cancellation-status column (matches the BO Profile's own "
         "note that it considers 'only active EWB') -- cancelled-EWB counts aren't available anywhere "
         "in the supplied data."),
        ("#16/#44/#53", "Average-price variance / price-band / MoM price drift by HSN",
         "The GSTR-1 'hsn' sheet's Total-Quantity column is 0 on every row inspected, so no reliable "
         "per-unit price is computable. Purchase-side (2B) data also has no HSN or quantity at all."),
        ("#22", "Input vs output quantity ratio",
         "No purchase-side quantity or HSN data exists anywhere in the supplied files (2B has neither)."),
        ("#26", "Import of services -- RCM tracking",
         "No import-of-services source document (foreign-supplier invoice/purchase register) is "
         "supplied. GSTR-2B's IMPG/IMPGSEZ sheets cover import of GOODS only, not services."),
        ("#28", "EWB validity > 15 days (held/blocked goods)",
         "The EWB files carry no validity/valid-upto column."),
        ("#38", "Dynamic HSN rate-change compliance (mid-year revision)",
         "Would need a rate-change notification calendar per HSN; not available offline."),
        ("#40", "EWB vehicle-type vs product-weight",
         "No vehicle-capacity reference database and no product-weight field exist in the supplied data."),
        ("#42", "Import BoE (Bill of Entry) mismatch",
         "No Bill of Entry data source is supplied."),
    ]
    return [Finding(ref, title, INFO, detail) for ref, title, detail in notes]


# ======================================================================
# TOP-LEVEL ORCHESTRATION
# ======================================================================
def run_all(files, ewb_out_rows, ewb_in_rows, months_covered, annual_data,
            monthly_comparison_rows, self_gstin, hsn_sac_master_override=None):
    """files: {'gstr1':path, 'gstr3b':path, 'einv':path or None, 'gstr2b':path or None}
    annual_data: the same dict master_build.py already builds (cash/credit/liab/bo/...)
    monthly_comparison_rows: annualwb.build_monthly_rows(annual_data) output
    hsn_sac_master_override: path to a run-supplied HSN/SAC code-and-description
    master (folder_classifier.py's 'hsn_sac_master_file'), or None to use the
    bundled HSN_SAC_default.xlsx.
    Returns a flat list of Finding, across every check in this module.
    Individual check functions are deliberately isolated with try/except so
    one check's failure doesn't blank out the other ~30 (see module docstring
    -- these are heuristic add-ons, not core arithmetic)."""
    g1path, g3bpath, einvpath = files.get("gstr1"), files.get("gstr3b"), files.get("einv")
    g1_lines_by_month = {m: ana.read_gstr1_lines(g1path, m) for m in months_covered}
    einv_lines_by_month = {m: ana.read_einv_lines(einvpath, m) for m in months_covered} if einvpath else {}
    hsn_by_month = _hsn_rows_by_month(g1path)
    cdnr_by_month = _cdnr_rows_by_month(g1path)
    b2cl_by_month = _b2cl_rows_by_month(g1path)
    gstr3b_fields = {m: _gstr3b_month_fields(g3bpath, m) for m in months_covered}
    g1_filing_dates = {m: f["filing_date"] for m, f in gstr3b_fields.items() if f}  # ARN date proxy
    bo = annual_data.get("bo")
    cash_monthly = annual_data["cash"]["monthly_by_tax_period"]
    credit_monthly = annual_data["credit"]["monthly_by_tax_period"]
    cash_txns = annual_data["cash"]["transactions"]

    checks = [
        ("A1/A2/A3/A6", lambda: check_hsn_rate_master(hsn_by_month)),
        ("A1-EXT", lambda: check_hsn_rate_master_extended(hsn_by_month)),
        ("A7", lambda: check_hsn_master_validity(hsn_by_month, hsn_sac_master_override)),
        ("A4", lambda: check_hsn_multi_rate(hsn_by_month)),
        ("A5", check_blocked_itc_by_hsn),
        ("B1/B3", lambda: check_pos_tax_head(g1_lines_by_month, self_gstin)),
        ("B2", lambda: check_b2c_large_ewb(b2cl_by_month, ewb_out_rows)),
        ("B4", lambda: check_sez_misclassification(g1_lines_by_month)),
        ("B3-2B", lambda: check_2b_itc_unavailable_flag(files.get("gstr2b"), months_covered) if files.get("gstr2b") else [Finding("B3", "GSTR-2B invoice flagged ITC Availability = No", INFO, "GSTR-2B merged file not supplied.")]),
        ("C1", check_rcm_hsn_not_declared),
        ("C2", lambda: check_branch_transfer(g1_lines_by_month, self_gstin)),
        ("C3", check_export_lut),
        ("C4", check_ewb_distance_route),
        ("C5", lambda: check_intra_vs_ewb_interstate(g1_lines_by_month, ewb_out_rows, self_gstin)),
        ("#1", lambda: check_round_numbers(g1_lines_by_month, ewb_out_rows)),
        ("#2", lambda: check_below_ewb_threshold(g1_lines_by_month, ewb_out_rows, self_gstin)),
        ("#3", lambda: check_reciprocal_trading(bo)),
        ("#4", lambda: check_cn_timing(cdnr_by_month, g1_lines_by_month)),
        ("#6/#27", lambda: check_hsn_drift(hsn_by_month)),
        ("#7", lambda: check_itc_liability_volatility(monthly_comparison_rows)),
        ("#8", lambda: check_year_end_dumping(g1_lines_by_month)),
        ("#9", lambda: check_zero_cash_months(cash_monthly)),
        ("#10/#19", lambda: check_ghost_supplier_cluster(bo)),
        ("#11", lambda: check_cn_vs_inward_ewb(cdnr_by_month, ewb_in_rows)),
        ("#12", lambda: check_cross_fy_shift(g1_lines_by_month, ewb_out_rows)),
        ("#15", lambda: check_irn_after_filing(g1_filing_dates, einv_lines_by_month, ewb_out_rows)),
        ("#17", lambda: check_credit_hoarding(credit_monthly, cash_monthly)),
        ("#18", lambda: check_negative_itc_reversal(gstr3b_fields)),
        ("#21", lambda: check_midnight_ewb(ewb_out_rows, ewb_in_rows)),
        ("#23", lambda: check_irn_delay(einv_lines_by_month, ewb_out_rows)),
        ("#24", lambda: check_ewb_state_shift(ewb_out_rows)),
        ("#25", lambda: check_exempt_turnover_rule42(bo, gstr3b_fields)),
        ("#37", lambda: check_sunday_holiday_ewb(ewb_out_rows, ewb_in_rows)),
        ("#41", lambda: check_consecutive_ewb_burst(ewb_out_rows)),
        ("#43", lambda: check_related_party_supercluster(bo)),
        ("#46", lambda: check_ewb_invoice_date_gap(g1_lines_by_month, ewb_out_rows)),
        ("#47/#55", lambda: check_cash_timing(cash_txns, gstr3b_fields)),
        ("#49", lambda: check_irn_ewb_invoice_gap(g1_lines_by_month, einv_lines_by_month, ewb_out_rows)),
        ("#51", lambda: check_b2b_matching_accuracy(g1_lines_by_month, einv_lines_by_month)),
        ("#54", lambda: check_bo_high_risk_supplier_itc(bo)),
        ("#56", lambda: check_gstr1_3b_headwise(monthly_comparison_rows)),
        ("#57", lambda: check_rate_outliers(g1_lines_by_month)),
    ]

    findings = []
    for ref, fn in checks:
        try:
            findings.extend(fn())
        except Exception as exc:   # noqa -- see module docstring: heuristics degrade, not crash
            findings.append(Finding(ref, f"Check {ref} could not run", INFO,
                f"Internal error while computing this check: {exc!r}. Every other check in this "
                f"module still ran normally."))
    findings.extend(not_feasible_notes())
    return findings
