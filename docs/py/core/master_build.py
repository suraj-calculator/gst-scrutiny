#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MASTER BUILD
============
CONSOLIDATED FILE -- contains what used to be: run_monthly_pipeline.py, master_build.py

The tool was reorganised from 19 .py files into 9 for easier sharing. Nothing
in the analytical logic was rewritten during that move: each section below is
the original module's code verbatim, with only (a) intra-project imports
repointed at the new file names, (b) its standalone __main__ demo block
removed, and (c) the renames listed under MERGE NOTES applied where two merged
modules happened to define the same top-level name with different bodies.

MERGE NOTES for this file:
  (none -- no name collisions)
"""


# ============================================================================
# ==== SECTION: run_monthly_pipeline.py  (was a standalone module before consolidation)
# ============================================================================
"""
RUN ONE MONTH  --  drives the three existing single-month engines
(gst_scrutiny_tool, gst_analysis_checks, gst_eway_recon) for one period,
now reading that period's data out of the MERGED (whole-FY) workbooks
identified by folder_classifier.py, instead of one file per month.

Two things changed from the original per-month-file design:
  1. The merged-workbook paths are the SAME for every month (there is only
     one file per document type) -- only PERIOD_LABEL changes between calls.
     Every parser below (parse_gstr1, parse_gstr3b, parse_einv, 2B, amendments,
     EWB invoice readers) now takes an explicit `month` argument and reads
     only that month's block out of the shared file.
  2. E-Way-Bill input still comes from the whole-FY annual workbooks, filtered
     down to this month via ewb_annual_parser.filter_by_month() -- unchanged
     from before, since those files were never per-month to begin with.

CHANGELOG (this revision):
  - ewb_out_file_supplied / ewb_in_file_supplied now passed through to
    gst_eway_recon.run() so a totally-absent EWB direction produces an
    honest SKIP instead of misleading PASS/REVIEW noise (see that module's
    docstring).
  - Cancelled-e-invoice rows (gst_scrutiny_tool.parse_einv()'s new
    'cancelled' key) are collected here per month and returned in the
    result dict, for the master build to aggregate FY-wide.
  - ARN dates (filing_compliance.py) are looked up per month here (once
    per run via the cache passed in by master_build.py, not re-read from
    disk every month) and the resulting late-fee/interest record is
    returned in the result dict for check #10/#8 to actually use --
    previously always fell through to INFO because GSTR1_FILING_DATE/
    GSTR3B_FILING_DATE were never actually set anywhere in this pipeline.
"""

import gst_parsers_returns as raw
import gst_checks_monthly as ana
import gst_checks_monthly as eway
import gst_parsers_returns as g2b
import gst_parsers_returns as ewbp
import gst_parsers_returns as amd
import gst_checks_forensic as fc


def run_month(month_label, files, ewb_out_annual_rows, ewb_in_annual_rows,
              self_gstin, company_name, ewb_out_file_supplied=True, ewb_in_file_supplied=True,
              gstr1_arn_by_month=None, gstr3b_arn_by_month=None,
              gstr1_is_qrmp=False, gstr3b_is_qrmp=False, annual_turnover=None):
    """files: {'gstr1':path,'gstr3b':path,'einv':path or None,'gstr2b':path or None}
    -- these are the MERGED workbook paths, the same on every call; only
    month_label changes which block gets read out of each of them.

    gstr1_arn_by_month / gstr3b_arn_by_month: pre-computed once per run (not
    per month -- these read the WHOLE merged file each time) by the caller
    via filing_compliance.gstr1_arn_dates_by_month()/gstr3b_arn_dates_by_month(),
    then passed in here so this function stays a pure per-month reader like
    every other parser call in this file."""
    g1path, g3bpath, einvpath, twobpath = (files.get("gstr1"), files.get("gstr3b"),
                                            files.get("einv"), files.get("gstr2b"))

    # ---- set the shared modules up for this period ----
    # (file paths are the same across months; only PERIOD_LABEL actually changes)
    raw.GSTR1_FILE = g1path
    raw.GSTR3B_FILE = g3bpath
    raw.EINV_FILE = einvpath
    raw.GSTR2B_FILE = twobpath
    raw.SELF_GSTIN = self_gstin
    raw.COMPANY_NAME = company_name
    raw.PERIOD_LABEL = month_label
    eway.SELF_GSTIN = self_gstin
    eway.GSTR2B_FILE = twobpath

    # ---- filing compliance: ARN dates + late fee/interest for THIS month ----
    # (fixes the previously-broken/unwired legacy extraction -- see filing_compliance.py)
    compliance = None
    if gstr1_arn_by_month is not None or gstr3b_arn_by_month is not None:
        compliance = fc.month_filing_compliance(
            month_label, gstr1_arn_by_month or {}, gstr3b_arn_by_month or {},
            gstr1_is_qrmp=gstr1_is_qrmp, gstr3b_is_qrmp=gstr3b_is_qrmp, self_gstin=self_gstin,
            annual_turnover=annual_turnover)
        # feed the real filing dates into gst_analysis_checks' checks #8/#10
        # (CONFIG-based, same mechanism the codebase already had -- just actually populated now)
        raw.GSTR1_FILING_DATE = compliance.get("gstr1_filing_date")
        raw.GSTR3B_FILING_DATE = compliance.get("gstr3b_filing_date")

    # ---- pipeline 1: comparison + Sooraj's 14 checks (unchanged engines,
    #      now called with the explicit month) ----
    comparisons, comp_raw = raw.build_comparisons()
    g1 = comp_raw["g1"]; g3b = comp_raw["g3b"]; einv = comp_raw["einv"]; b2b = comp_raw["b2b"]
    g1_lines = ana.read_gstr1_lines(g1path, month_label)
    einv_lines = ana.read_einv_lines(einvpath, month_label) if einvpath else []
    findings14 = ana.run_checks(g1, g3b, einv, b2b, g1_lines, einv_lines)

    # ---- pipeline 2: E-Way-Bill 27-check matrix, fed from the ANNUAL EWB lists
    #      filtered to this month (by EWB date) -- unchanged ----
    ewb_out = ewbp.filter_by_month(ewb_out_annual_rows, month_label)
    ewb_in = ewbp.filter_by_month(ewb_in_annual_rows, month_label)
    g1inv = eway.read_gstr1_invoices(g1path, month_label)
    einv_ew = eway.read_einv_invoices(einvpath, month_label) if einvpath else {}
    b2b_ew = g2b.summary_for_month(twobpath, month_label)
    hsn_rows_month = raw.read_gstr1_hsn_all_months(g1path).get(month_label, [])
    is_services_dom, dom_hsn, dom_share = raw.dominant_hsn_is_services(hsn_rows_month)
    findings27 = eway.run(ewb_out, ewb_in, g1inv, einv_ew, g3b, b2b_ew,
                          ewb_out_file_supplied=ewb_out_file_supplied,
                          ewb_in_file_supplied=ewb_in_file_supplied,
                          is_services_dominant=is_services_dom, dominant_hsn=dom_hsn,
                          dominant_hsn_share=dom_share)

    # ---- amendments + doc-series-integrity (this month's OWN GSTR-1 block only;
    #      cross-month linkage to the ORIGINAL month happens in master_build) ----
    b2ba = amd.parse_b2ba(g1path, month_label)
    cdnra = amd.parse_cdnra(g1path, month_label)
    docs = amd.parse_docs(g1path, month_label)
    actual_invnos = set(k[0] for k in g1.get("lines", {}).keys())
    doc_gap = amd.doc_series_gap_check(docs, actual_invnos)

    # ---- cancelled e-invoices this month (new) ----
    cancelled_this_month = einv.get("cancelled", []) if einv.get("available") else []
    einv_column_found = einv.get("cancel_col_found", False) if einv.get("available") else False
    # BUG FIX: previously there was no way to tell "E-Invoice WAS supplied for this month but
    # had zero cancellations" apart from "E-Invoice not supplied at all" -- both produced an
    # empty `cancelled_this_month` list, and the aggregator downstream only recorded a month
    # key in `cancelled_by_month` when the list was non-empty, so a genuinely clean taxpayer
    # (E-Invoice supplied, zero cancellations every month) collapsed to a totally EMPTY
    # `cancelled_by_month` dict -- indistinguishable from "no E-Invoice file supplied at all",
    # which build_cancelled_einvoice_findings() then reported as "No E-Invoice data supplied --
    # cannot test." even though EINV_Merged.xlsx WAS supplied and WAS parsed. `einv_available`
    # (a plain bool, independent of whether any row happened to be cancelled) fixes this.
    einv_available_this_month = bool(einv.get("available"))

    return dict(
        month=month_label, comparisons=comparisons, comp_raw=comp_raw,
        findings14=findings14, findings27=findings27,
        b2ba=b2ba, cdnra=cdnra, docs=docs, doc_gap=doc_gap,
        compliance=compliance,
        cancelled_einvoices=cancelled_this_month, einv_cancel_col_found=einv_column_found,
        einv_available=einv_available_this_month,
        g1_named_invnos=set(k[0] for k in g1.get("lines", {}).keys() if k[0]),
        meta=dict(
            ewb_out_n=len(ewb_out), ewb_in_n=len(ewb_in),
            twob_src=b2b_ew.get("_source"), twob_file=b2b_ew.get("_file"),
            twob_available=b2b_ew.get("available", True),
            einv_file=einvpath,
        ),
    )


# ============================================================================
# ==== SECTION: master_build.py  (was a standalone module before consolidation)
# ============================================================================
"""
MASTER BUILD  --  the single entry point for the whole tool.

    python master_build.py [folder]

Put every input file (however many months you have) in one folder next to
all the .py files listed in HOW_TO_RUN.md, then run this. It will:
  1. Classify every file in the folder by content signature (folder_classifier.py)
  2. Run the full single-month engine (Comparison + Analysis-14 + EWB-27 +
     Doc-Series-Integrity) for every month that has at least GSTR-1 + GSTR-3B
  3. Run the Phase-1 annual reconciliation (ledgers + portal comparison
     + BO Profile) -- build_annual_workbook.py's logic, reused as-is
  4. Cross-month rectification pairing: match every GSTR-1 amendment-sheet row
     (b2ba/cdnra) found in ANY month against the ORIGINAL invoice/note in
     whichever earlier month first reported it, plus a best-effort DRC-payment
     cross-reference
  5. Write ONE workbook: Master Dashboard first, then per-month sheets, then
     the annual sheets, then the rectification-pairs sheet.

Missing months are handled gracefully -- whatever you have is what gets
analysed; the Dashboard states plainly which months are covered.
"""

import os
import re
import sys
import datetime as _dt
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter

from gst_core import classify_folder
import gst_parsers_returns as ewbp
# [merged] `from run_monthly_pipeline import run_month` -- now defined in this file
import gst_parsers_returns as amd
import gst_report as uni
import gst_parsers_returns as raw

from gst_parsers_dept import (parse_cash_or_liability_ledger, parse_credit_ledger,
                             parse_portal_comparison)
from gst_parsers_dept import parse_bo_profile
import gst_report as annualwb
import gst_checks_hsn_fraud as hfc
import gst_checks_forensic as fc
import gst_parsers_dept as arp
import gst_checks_forensic as fchk
import gst_checks_flow as flow
import gst_core as mpu
import gst_blocked_credit as bcred
import gst_machinery_scan as mscan

RED = PatternFill("solid", fgColor="FFC7CE")
GREEN = PatternFill("solid", fgColor="C6EFCE")
AMBER = PatternFill("solid", fgColor="FFEB9C")
BLUE = PatternFill("solid", fgColor="DDEBF7")
GREY = PatternFill("solid", fgColor="E7E6E6")
HEAD = PatternFill("solid", fgColor="1F3864")
SECT = PatternFill("solid", fgColor="D9E1F2")
TITLEF = Font(bold=True, size=13, color="1F3864")
BORDER = Border(*[Side(style="thin", color="BFBFBF")] * 4)
SEV_FILL = {"FLAG": RED, "MISMATCH": RED, "REVIEW": AMBER, "INFO": BLUE,
            "PASS": GREEN, "MATCH": GREEN, "SKIPPED": GREY, "EXPLAINED": PatternFill("solid", fgColor="D6DCE5")}
SEV_FONT = {"FLAG": Font(bold=True, color="9C0006"), "MISMATCH": Font(bold=True, color="9C0006"),
            "REVIEW": Font(bold=True, color="9C6500"), "INFO": Font(bold=True, color="2F5496"),
            "PASS": Font(bold=True, color="006100"), "MATCH": Font(bold=True, color="006100"),
            "SKIPPED": Font(bold=True, color="808080"), "EXPLAINED": Font(bold=True, color="31708F")}

MONTH_ORDER = ["Apr-22", "May-22", "Jun-22", "Jul-22", "Aug-22", "Sep-22",
               "Oct-22", "Nov-22", "Dec-22", "Jan-23", "Feb-23", "Mar-23"]
# ^ RETAINED for any code path that still imports this name directly, but
# main() below no longer uses it -- see _sort_months_chronologically(), which
# builds the real month order dynamically from whatever months are actually
# present in the data (any number of FYs, not just one).

_MONTH_NUM = {v: k for k, v in mpu.CAL_MONTH_ABBR.items()}


def _month_sort_key(label):
    """'Jan-23' -> a real sortable value (2023, 1). Works across any number
    of years -- this is what makes multi-year runs order correctly without
    a hardcoded 12-month list."""
    m = re.match(r"^([A-Za-z]{3})-(\d{2})$", label) if label else None
    if not m:
        return (9999, 99)
    mon, yy = m.group(1), int(m.group(2))
    return (2000 + yy, _MONTH_NUM.get(mon.title(), 0))


def _sort_months_chronologically(months):
    return sorted(months, key=_month_sort_key)


def _fy_label_for_month(label):
    """'Apr-22'->'2022-23', 'Jan-23'->'2022-23', 'Apr-23'->'2023-24'. Indian
    FY runs Apr-Mar. Used only for grouping/display, not for any parsing."""
    year, mon = _month_sort_key(label)
    if mon >= 4:
        return f"{year}-{str(year+1)[2:]}"
    return f"{year-1}-{str(year)[2:]}"


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def sheet_name(prefix, month, limit=31):
    n = f"{month} {prefix}"
    return n[:limit]


# ======================================================================
# CROSS-MONTH: rectification pairing
# ======================================================================
def build_rectification_pairs(month_results, month_g1_lines, months_covered):
    """month_g1_lines: {month: {(invno,rate): [taxable, total_tax]}} from each month's own B2B.
    months_covered: the ACTUAL chronologically-sorted list of months this run
    covers (any number of FYs) -- replaces the old hardcoded MONTH_ORDER so
    multi-year runs still correctly find "the earlier month" across an FY
    boundary, not just within one hardcoded 12-month window.
    For every amendment row in ANY month, find which earlier month first
    reported the 'original' invoice/note number, and pair them.

    UPDATED (per instruction): now also carries the ORIGINAL invoice's own taxable value and
    total tax (from that earlier month's own B2B line, already available in month_g1_lines --
    no new source needed) alongside the REVISED (amended) values already captured, plus the
    DELTA between them -- previously this only showed revised values with no visible 'what
    actually changed'. CGST/SGST breakdown of the ORIGINAL is not separately available (this
    tool's GSTR-1 line index tracks taxable + total tax per rate-line, not a CGST/SGST split)
    -- stated plainly on the sheet itself rather than guessing a 50/50 split, which would be
    wrong for any interstate invoice."""
    pairs = []
    for res in month_results:
        amend_month = res["month"]
        for row in res["b2ba"]:
            orig_month = None
            orig_taxable = orig_tax = None
            for m in months_covered:
                if m == amend_month:
                    break
                if m in month_g1_lines:
                    match = next((v for k, v in month_g1_lines[m].items() if k[0] == row["orig_invno"]), None)
                    if match is not None:
                        orig_month = m
                        orig_taxable, orig_tax = match[0], match[1]
                        break
            revised_tax = (row.get("igst") or 0.0) + (row.get("cgst") or 0.0) + (row.get("sgst") or 0.0) + (row.get("cess") or 0.0)
            pairs.append(dict(
                kind="B2B Invoice Amendment", gstin=row["gstin"], recipient=row["recipient"],
                original_ref=row["orig_invno"], original_month=orig_month or "NOT FOUND in any earlier month provided",
                revised_ref=row["revised_invno"], amended_in_month=amend_month,
                taxable=row["taxable"], igst=row["igst"], cgst=row["cgst"], sgst=row["sgst"],
                original_taxable=orig_taxable, original_tax=orig_tax,
                revised_taxable=row["taxable"], revised_tax=revised_tax,
                delta_taxable=(row["taxable"] - orig_taxable) if orig_taxable is not None else None,
                delta_tax=(revised_tax - orig_tax) if orig_tax is not None else None,
            ))
        for row in res["cdnra"]:
            revised_tax = (row.get("igst") or 0.0) + (row.get("cgst") or 0.0) + (row.get("sgst") or 0.0) + (row.get("cess") or 0.0)
            pairs.append(dict(
                kind="Credit/Debit Note Amendment", gstin=row["gstin"], recipient="",
                original_ref=row["orig_noteno"], original_month="(note amendments not month-matched -- see original_ref)",
                revised_ref=row["revised_noteno"], amended_in_month=amend_month,
                taxable=row["taxable"], igst=row["igst"], cgst=row["cgst"], sgst=row["sgst"],
                original_taxable=None, original_tax=None,
                revised_taxable=row["taxable"], revised_tax=revised_tax,
                delta_taxable=None, delta_tax=None,
            ))
    return pairs


# ======================================================================
# WRITERS
# ======================================================================
def _extract_amount(numbers):
    """Best-effort rupee magnitude from a Finding's .numbers dict, trying the most likely key
    names in priority order. Returns None (never a guess) if nothing numeric-and-plausible is
    found -- callers must leave the Severity cell blank rather than force a band onto it."""
    if not numbers:
        return None
    for key in ("difference", "diff", "gap", "residual", "delta", "net_delta_taxable",
                "cumulative_diff_taxable", "amount"):
        if key in numbers and isinstance(numbers[key], (int, float)):
            return numbers[key]
    for v in numbers.values():
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            return v
    return None


def _extract_ewb_amount(finding):
    """gst_checks_monthly.F (E-Way-Bill findings) has no .numbers dict -- its numeric detail
    lives in .rows (a header row + data rows). Best-effort: sum the LAST numeric-looking column
    across all data rows (by convention in this module, the rightmost numeric column is a
    rupee/tax figure -- 'assess', 'tax', 'diff', 'consignment ₹', etc.). Returns None (blank
    Severity cell) rather than guessing when .rows has no clear numeric column."""
    rows = getattr(finding, "rows", None)
    if not rows or len(rows) < 2:
        return None
    header, data = rows[0], rows[1:]
    numeric_cols = [i for i in range(len(header))
                    if any(isinstance(r[i], (int, float)) for r in data if i < len(r))]
    if not numeric_cols:
        return None
    col = numeric_cols[-1]
    total = sum(r[col] for r in data if col < len(r) and isinstance(r[col], (int, float)))
    return total if total else None


def _root_id_and_level(pipeline, ref, title, month):
    """(root_id, level) for the dedup/rollup system.

    CONSERVATIVE BY DESIGN (per this session's dedup-fix prompt): only merges rows that are
    MATHEMATICALLY DERIVED from the same source computation -- an exact CGST/SGST tax-rate
    split of one taxable-value gap, a gross/net-of-CN variant within a small tolerance of the
    same invoice set, or an FY rollup that is a literal sum of monthly PARENT rows. Two rows
    are NEVER merged just because they land in the same month or check category, or because
    they happen to cite a shared INPUT number -- confirmed against real data that this specific
    trap exists: F1 ('Purchase vs sales value flow') and F4 ('GSTR-1 vs GSTR-3B outward value')
    both cite the same GSTR-1 net-sales figure as ONE INPUT, but compute two genuinely different
    results from two different second inputs (2B purchases vs 3B liability) -- they stay
    separate PARENTs. (The apparent duplicate in an earlier ad-hoc QA pass was actually a bug in
    THAT script's own amount-extraction, which grabbed the largest number in free text instead
    of each Finding's own structured 'difference' value -- fixed by no longer doing that; see
    _extract_amount()/_extract_ewb_amount(), which already prefer the structured field and were
    never affected by this bug in the live dashboard's own Severity column.)

    level is one of:
      PARENT -- the primary, count-worthy finding. Summed for total-exposure figures.
      VIEW   -- a derived breakdown/near-duplicate of a PARENT. Never counted in totals; kept
                fully visible for drill-down.
      ROLLUP -- an FY-wide sum of monthly PARENT rows. Never counted alongside the PARENT rows
                it sums (would double the exposure); shown for its own FY-level context.

    Any row not explicitly matched here is its own independent PARENT (own root_id) -- the
    default is to NOT merge, so a coincidental second issue in the same month is never hidden.
    """
    # ---- Issue 1 family: GSTR-1-vs-GSTR-3B outward-value gap ----
    if pipeline == "Comparison" and ref == "A" and "Outward taxable value" in title:
        return (f"outward_gap:{month}", "PARENT")
    if pipeline == "Comparison" and ref == "A" and ("Outward CGST" in title or "Outward SGST" in title):
        return (f"outward_gap:{month}", "VIEW")   # exact CGST/SGST split of the taxable-value gap
    if pipeline == "Analysis(14)" and ref == "#2":
        return (f"outward_gap:{month}", "VIEW")   # same gap, tax-figure wording instead of taxable-value
    if pipeline == "Flow/Counterparty" and ref == "F4":
        return ("outward_gap:FY", "ROLLUP")

    # ---- Issue 2 family: IGST ITC-claimed-vs-available gap (monthly-aggregate, CN-netted) ----
    if pipeline == "Comparison" and ref == "D" and "IGST" in title and "NET of CN" in title:
        return (f"igst_itc_gap:{month}", "PARENT")
    if pipeline == "Comparison" and ref == "D" and "IGST" in title and "gross" in title:
        return (f"igst_itc_gap:{month}", "VIEW")   # gross vs net-of-CN, same invoice set, <2% apart
    # F7 is DELIBERATELY NOT merged into igst_itc_gap's ROLLUP: its own title says "invoice
    # level" -- it matches 4(A)(5) against 2B's actual invoice list, a genuinely different and
    # more granular computation than D's monthly-aggregate-CN-netted comparison, not a rollup of
    # D's numbers. Confirmed by the rollup sanity check itself catching a large mismatch when
    # this was first tried (69K aggregated-parent-sum vs 26.1L invoice-level figure) -- exactly
    # the kind of over-merge the conservative "same source computation" rule exists to prevent.
    # Kept as its own independent PARENT/family so it is never silently netted against D.
    if pipeline == "Flow/Counterparty" and ref == "F7":
        return ("igst_itc_gap_invoice_level:FY", "PARENT")
    # D2's IGST reversal (4B2 vs 2B-CN) is DELIBERATELY kept a separate PARENT, not merged into
    # igst_itc_gap above: 4(B)(2) (reversal declared) and 4(A)(5) (ITC claimed) are different
    # return line items, not a mathematical derivation of one another, even though both are
    # symptoms of the same "IGST doesn't net the way CGST/SGST does for this filer" root cause.

    # Everything else: independent PARENT, its own root_id (never merged by default).
    return (f"{pipeline}:{ref}:{title}:{month}", "PARENT")


def write_ewb_pattern_annual(ws, month_results, hsn_findings):
    """NEW sheet: 'EWB Pattern Checks -- Annual Summary'. Consumes the .raw per-occurrence
    records now retained by #27/#13/#15/#17 (gst_checks_monthly.py) and #24
    (gst_checks_hsn_fraud.py) instead of the single narrative line each used to collapse into.
    One row per distinct check-subject, sorted by total value moved (a high-value repeat pair is
    a bigger question mark than a low-value one at the same frequency), with a frequency
    signal distinguishing "fired most months -> likely routine" from "fired in only 1-3 months
    -> worth a closer look". Existing per-month EWB/EWB Detail sheets are unchanged -- this is
    additive, nothing is deleted."""
    ws.cell(1, 1, "EWB PATTERN CHECKS -- ANNUAL SUMMARY").font = TITLEF
    ws.cell(2, 1, "Consolidates checks that recur nearly every month but were only ever visible "
                  "one month-tab at a time -- a check firing 12/12 months is very likely routine "
                  "business behaviour (a regular transporter/supplier); one firing in only 1-3 "
                  "months is the one actually worth a closer look. Sorted by total value moved, "
                  "not just frequency.").font = Font(size=9, italic=True)
    hdr = ["Check Type", "Subject", "Months Fired / 12", "Frequency Signal", "Occurrence Count",
           "E-Way-Bill / Doc Numbers", "Total Value Moved (Rs)", "Avg Days Between Occurrences",
           "Source (month tabs)"]
    r = 4
    for i, h in enumerate(hdr, 1):
        ws.cell(r, i, h)
    _style_header(ws, r, len(hdr))
    r += 1

    def freq_signal(n_months):
        if n_months >= 9:
            return "Likely routine (fired most months)"
        if n_months <= 3:
            return "WORTH A CLOSER LOOK (fired rarely)"
        return "Moderate frequency"

    subjects = {}   # (check_type, subject_key) -> dict(months=set, occ=0, docs=[], value=0.0, dates=[])

    def add(check_type, subject, month, doc, value, date):
        key = (check_type, subject)
        s = subjects.setdefault(key, dict(months=set(), occ=0, docs=[], value=0.0, dates=[]))
        s["months"].add(month); s["occ"] += 1
        if doc:
            s["docs"].append(str(doc))
        s["value"] += value or 0.0
        if date:
            s["dates"].append(date)

    for res in month_results:
        m = res["month"]
        for f in res.get("findings27", []):
            if f.ref == "#27":
                for rec in getattr(f, "raw", []):
                    subj = f"{rec['vehicle']} | {rec['from_gstin']} -> {rec['to_gstin']}"
                    add("#27 Same vehicle, repeated trips", subj, m, rec["docno"], rec["assess"], rec["ewbdate"])
            elif f.ref == "#13":
                for rec in getattr(f, "raw", []):
                    subj = f"{rec['gstin']} ({rec.get('supplier','')})"
                    add("#13 2B inter-state >Rs50k, no inward EWB", subj, m, rec["invno"], rec["consignment"], None)
            elif f.ref == "#15":
                for rec in getattr(f, "raw", []):
                    subj = f"{rec['out_gstin']} <-> {rec['in_gstin']}"
                    add("#15 EWB-Out vs EWB-In (same doc-no)", subj, m, rec["docno"],
                        max(rec["out_assess"], rec["in_assess"]), rec.get("ewbdate"))
            elif f.ref == "#17":
                for rec in getattr(f, "raw", []):
                    subj = rec["invno"]
                    add("#17 Triangulation gap", subj, m, rec["invno"], rec.get("taxable") or 0.0, None)
    for f in hsn_findings:
        if f.ref == "#24":
            n = f.numbers or {}
            subj = f"Destination state {n.get('state')}"
            add("#24 Sudden EWB destination-state shift", subj, n.get("month"), None, 0.0, None)

    rows_out = []
    for (check_type, subject), s in subjects.items():
        n_months = len(s["months"])
        avg_gap = None
        dates = sorted(d for d in s["dates"] if d)
        if len(dates) > 1:
            deltas = [(dates[i+1] - dates[i]).days for i in range(len(dates)-1)]
            avg_gap = sum(deltas) / len(deltas)
        docs_display = ", ".join(sorted(set(s["docs"]))[:15]) + (" ..." if len(set(s["docs"])) > 15 else "")
        rows_out.append((check_type, subject, n_months, freq_signal(n_months), s["occ"],
                         docs_display or "(see month tabs)", s["value"], avg_gap,
                         ", ".join(sheet_name("EWB Detail", m) for m in sorted(s["months"], key=_month_sort_key))))
    rows_out.sort(key=lambda x: -x[6])

    for row in rows_out:
        for ci, v in enumerate(row, 1):
            cell = ws.cell(r, ci, v)
            cell.border = BORDER
            cell.font = Font(size=10)
            cell.alignment = Alignment(wrap_text=(ci in (2, 6, 9)), vertical="top")
            if ci == 7 and isinstance(v, (int, float)):
                cell.number_format = "#,##0.00"
            if ci == 8 and isinstance(v, (int, float)):
                cell.number_format = "0.0"
            if ci == 4 and "CLOSER LOOK" in str(v):
                cell.fill = AMBER
        r += 1
    if not rows_out:
        ws.cell(r, 1, "No recurring EWB pattern-check occurrences found.").font = Font(italic=True)
        r += 1
    r += 1
    n_closer_look = sum(1 for x in rows_out if "CLOSER LOOK" in x[3])
    total_value = sum(x[6] for x in rows_out)
    ws.cell(r, 1, f"Summary: {len(rows_out)} distinct check-subject(s), {n_closer_look} flagged "
                  f"'worth a closer look' (rare-occurrence), total value moved across all "
                  f"subjects Rs {total_value:,.2f}.").font = Font(bold=True, size=10, color="1F3864")
    r += 2
    ws.cell(r, 1, "SUGGESTED FURTHER ANALYSIS from raw data already in this tool's inputs, not yet "
                  "built:").font = Font(bold=True, size=10, color="1F3864")
    r += 1
    for s in ["Vehicle-number reuse ACROSS different consignor/consignee pairs (not just repeated "
              "trips for the same pair, which #27 already covers) -- the same vehicle serving many "
              "unrelated parties in a short window is a different signal (possible EWB-generation-"
              "for-hire without actual movement) from one vehicle doing repeat runs for one route.",
              "A day-of-week / time-of-day distribution for #27's repeated-vehicle trips -- "
              "genuine regular routes cluster on business days at similar hours; movement spread "
              "randomly across all 7 days and all hours is a different pattern worth a second look.",
              "Distance vs assessable-value outliers (very high value on a very short EWB-declared "
              "route, or vice versa) -- not currently computed anywhere in this tool; the EWB "
              "export's own From/To place fields would need geocoding to estimate distance, which "
              "this tool doesn't currently do."]:
        ws.cell(r, 1, "- " + s).font = Font(size=9, italic=True)
        ws.cell(r, 1).alignment = Alignment(wrap_text=True)
        r += 1
    for col, w in zip("ABCDEFGHI", [30, 34, 14, 30, 14, 40, 18, 14, 40]):
        ws.column_dimensions[col].width = w
    ws.auto_filter.ref = f"A4:I{max(r-1,4)}"


def write_irn_late_annual(ws, hsn_findings, einv_month_map=None, months_covered=None, ewb_out_rows=None):
    """NEW sheet: 'IRN Late-Generation -- Annual Detail'. Every #23 occurrence (35 in this
    dataset) as its own complete row -- never aggregated away."""
    ws.cell(1, 1, "IRN LATE-GENERATION -- ANNUAL DETAIL (check #23)").font = TITLEF
    occs = [f for f in hsn_findings if f.ref == "#23" and f.severity != "PASS"]
    total_taxable = sum((f.numbers or {}).get("taxable") or 0 for f in occs)
    months_affected = sorted({(f.numbers or {}).get("month") for f in occs if (f.numbers or {}).get("month")}, key=_month_sort_key)
    worst = max(occs, key=lambda f: (f.numbers or {}).get("gap_days", 0)) if occs else None
    ws.cell(2, 1, f"{len(occs)} late-IRN invoice(s) for the FY across {len(months_affected)} month(s) "
                  f"({', '.join(months_affected)}); total taxable value affected Rs {total_taxable:,.2f}. "
                  + (f"Worst delay: {(worst.numbers or {}).get('gap_days')} days, invoice "
                     f"{(worst.numbers or {}).get('invno')}." if worst else "")).font = Font(size=9, italic=True)
    hdr = ["Month", "Invoice No.", "Invoice Date", "IRN Generation Date", "Delay (days)",
           "Taxable Value (Rs)", "Tax (Rs)", "Recipient GSTIN", "Severity"]
    r = 4
    for i, h in enumerate(hdr, 1):
        ws.cell(r, i, h)
    _style_header(ws, r, len(hdr))
    r += 1
    for f in sorted(occs, key=lambda f: -((f.numbers or {}).get("gap_days") or 0)):
        n = f.numbers or {}
        vals = [n.get("month"), n.get("invno"), n.get("invdate"), n.get("irndate"), n.get("gap_days"),
               n.get("taxable"), n.get("tax"), n.get("recipient_gstin"), f.severity]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(r, ci, v)
            cell.border = BORDER; cell.font = Font(size=10)
            if ci in (6, 7):
                cell.number_format = "#,##0.00"
        if f.severity == "FLAG":
            ws.cell(r, 9).fill = RED
        r += 1
    if not occs:
        ws.cell(r, 1, "No late-IRN invoices found.").font = Font(italic=True)
    for col, w in zip("ABCDEFGHI", [10, 20, 14, 18, 12, 16, 14, 18, 10]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:I{max(r-1,4)}"

    # ================================================================================
    # NEW (per explicit request): "E-Way Bill vs IRN Timing Check + Missing Data Report" --
    # two exception-only tables added below the existing #23 detail, same sheet. Scope is the
    # FULL FY's e-invoiced/EWB'd population (not limited to the 35 already-"late" #23 rows
    # above -- this is a different, independent check: whether the E-Way Bill for an invoice
    # was generated BEFORE its IRN, not whether the IRN itself was late against the invoice
    # date).
    # ================================================================================
    if einv_month_map is not None and months_covered is not None:
        r += 2
        ws.cell(r, 1, "TASK: E-WAY BILL vs IRN TIMING CHECK + MISSING DATA REPORT").font = TITLEF
        r += 1
        ws.cell(r, 1, "Per Rule 138A/48(4), the IRN should exist before an e-way bill is generated "
                      "against the same invoice. Every invoice with BOTH a resolvable IRN date and "
                      "a matching outward EWB is checked below; only the two exception cases are "
                      "shown (rows where the EWB is safely after the IRN are not a finding and are "
                      "excluded, per instruction).").font = Font(size=9, italic=True)
        ws.cell(r, 1).alignment = Alignment(wrap_text=True); r += 2

        # Build FY-wide invoice -> IRN date lookup (deduped -- IRN date is invoice-level, not
        # rate-line-level; a multi-rate invoice's several rows all share the same IRN/IRN date).
        irn_by_invno = {}
        for m in months_covered:
            p = einv_month_map.get(m)
            if not p:
                continue
            for x in ana.read_einv_lines(p, m):
                invno_key = str(x.get("invno", "")).strip().upper()
                if invno_key and invno_key not in irn_by_invno:
                    irn_by_invno[invno_key] = dict(invno=x.get("invno"), invdate=x.get("invdate"),
                                                    irndate=x.get("irndate"))

        # Build EWB doc-number -> ewbdate+ewbtime lookup (first occurrence per doc number wins;
        # dedupe_ewb() has already dropped exact-duplicate rows upstream of ewb_out_rows).
        ewb_by_docno = {}
        for e in (ewb_out_rows or []):
            docno_key = str(e.get("docno", "")).strip().upper()
            if docno_key and docno_key not in ewb_by_docno:
                ewb_by_docno[docno_key] = dict(ewbno=e.get("ewbno"), ewbdate=e.get("ewbdate"), ewbtime=e.get("ewbtime"))

        all_invnos = set(irn_by_invno) | set(ewb_by_docno)
        table1, table2 = [], []
        for key in all_invnos:
            irn_info = irn_by_invno.get(key)
            ewb_info = ewb_by_docno.get(key)
            irndate = irn_info.get("irndate") if irn_info else None
            invdate = irn_info.get("invdate") if irn_info else None
            invno_disp = irn_info.get("invno") if irn_info else (ewb_info.get("ewbno") and key)
            ewbdate = ewb_info.get("ewbdate") if ewb_info else None
            ewbtime = ewb_info.get("ewbtime") if ewb_info else None

            if irndate is None or ewbdate is None:
                missing = []
                if irndate is None:
                    missing.append("IRN Date Missing")
                if ewbdate is None:
                    missing.append("EWB Date Missing")
                reason = "Both Missing" if len(missing) == 2 else missing[0]
                ewb_dt_disp = "MISSING" if ewbdate is None else (
                    f"{ewbdate} {ewbtime}" if ewbtime else str(ewbdate))
                table2.append(dict(invno=invno_disp or key, invdate=invdate,
                                    irndate=("MISSING" if irndate is None else str(irndate)),
                                    ewbdate=ewb_dt_disp, reason=reason))
                continue

            # irndate is DATE-ONLY at the source (no time captured anywhere in the E-Invoice
            # export -- confirmed on the raw file); ewbdate+ewbtime is a real date+time. Per
            # instruction: since IRN's time-of-day is never known, an EWB on the SAME calendar
            # date as the IRN is treated as ambiguous/cannot-rule-out and included here too,
            # not assumed safe. Gap is shown in whole days (IRN date minus EWB date) -- 0 for
            # the same-date/ambiguous case (never a fabricated sub-day figure built on an
            # assumed IRN time this tool does not actually have), negative for a genuinely
            # earlier EWB date.
            gap_days = (irndate - ewbdate).days
            if gap_days > 0:
                # EWB strictly BEFORE the IRN's calendar date -- unambiguous risk.
                table1.append(dict(invno=invno_disp or key, invdate=invdate, irndate=str(irndate),
                                    ewbdate=f"{ewbdate} {ewbtime}" if ewbtime else str(ewbdate),
                                    gap=-gap_days, same_date=False))
            elif gap_days == 0:
                # Same calendar date -- IRN's own time-of-day is unknown, so this cannot be
                # confirmed as safe; included per instruction rather than assumed OK.
                table1.append(dict(invno=invno_disp or key, invdate=invdate, irndate=str(irndate),
                                    ewbdate=f"{ewbdate} {ewbtime}" if ewbtime else str(ewbdate),
                                    gap=0, same_date=True))
            # gap_days < 0 (EWB clearly on a later calendar date than the IRN) -- no risk,
            # excluded from both tables per instruction ("only show exception cases").

        # ---- Table 1: EWB Before IRN -- Risk Flag ----
        ws.cell(r, 1, f"Table 1: EWB Before IRN — Risk Flag ({len(table1)})").font = Font(bold=True, size=11, color="1F3864")
        r += 1
        hdr1 = ["Invoice No.", "Invoice Date", "IRN Generation Date-Time", "E-Way Bill Generation Date-Time",
                "Gap (IRN minus EWB, days)", "Risk Flag"]
        for i, h in enumerate(hdr1, 1):
            ws.cell(r, i, h)
        _style_header(ws, r, len(hdr1))
        r += 1
        t1_hdr_row = r - 1
        table1.sort(key=lambda x: x["gap"])   # most negative/severe first; same-date (0) rows last
        if not table1:
            ws.cell(r, 1, "None -- no invoice has an EWB generated before (or ambiguously on the "
                          "same date as) its IRN.").font = Font(italic=True, color="808080")
            r += 1
        for x in table1:
            risk_label = "EWB Before IRN" if not x["same_date"] else "EWB Before IRN (same calendar date -- IRN time-of-day unknown, cannot confirm sequence)"
            vals = [x["invno"], x["invdate"], x["irndate"], x["ewbdate"], x["gap"], risk_label]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(r, ci, v)
                cell.border = BORDER; cell.font = Font(size=10)
                cell.alignment = Alignment(wrap_text=(ci == 6))
            ws.cell(r, 5).fill = RED if x["gap"] < 0 else AMBER
            r += 1
        ws.column_dimensions["F"].width = 55
        r += 2

        # ---- Table 2: Missing Data -- Cannot Verify ----
        ws.cell(r, 1, f"Table 2: Missing Data — Cannot Verify ({len(table2)})").font = Font(bold=True, size=11, color="1F3864")
        r += 1
        hdr2 = ["Invoice No.", "Invoice Date", "IRN Generation Date-Time", "E-Way Bill Generation Date-Time", "Reason"]
        for i, h in enumerate(hdr2, 1):
            ws.cell(r, i, h)
        _style_header(ws, r, len(hdr2))
        r += 1
        table2.sort(key=lambda x: x["reason"])
        if not table2:
            ws.cell(r, 1, "None -- every invoice checked has both an IRN date and an EWB date to "
                          "compare.").font = Font(italic=True, color="808080")
            r += 1
        for x in table2:
            vals = [x["invno"], x["invdate"], x["irndate"], x["ewbdate"], x["reason"]]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(r, ci, v)
                cell.border = BORDER; cell.font = Font(size=10)
            ws.cell(r, 5).fill = AMBER
            r += 1


def write_itc_blocked_annual(ws, hsn_findings):
    """NEW sheet: 'ITC-Blocked Invoices -- Annual Detail'. Every B3 occurrence (11 in this
    dataset), with an explicit occurrence# per supplier so concentration (e.g. one supplier
    accounting for several of the total occurrences) is visible without manually counting rows."""
    ws.cell(1, 1, "ITC-BLOCKED INVOICES -- ANNUAL DETAIL (check B3: GSTR-2B 'ITC Availability = No')").font = TITLEF
    occs = [f for f in hsn_findings if f.ref == "B3" and f.severity != "PASS"]
    total_tax = sum((f.numbers or {}).get("tax") or 0 for f in occs)
    by_supplier = {}
    for f in occs:
        g = (f.numbers or {}).get("gstin", "")
        by_supplier.setdefault(g, []).append(f)
    top_supplier = max(by_supplier.items(), key=lambda kv: len(kv[1])) if by_supplier else (None, [])
    ws.cell(2, 1, f"{len(occs)} blocked-ITC invoice(s) for the FY; total tax blocked Rs {total_tax:,.2f}. "
                  + (f"Top supplier by occurrence: {top_supplier[0]} ({(occs[0].numbers or {}).get('supplier','') if occs else ''}) "
                     f"-- {len(top_supplier[1])} occurrence(s)." if top_supplier[0] else "")
                  ).font = Font(size=9, italic=True)
    hdr = ["Month", "Invoice No.", "Invoice Date", "Supplier GSTIN", "Supplier Name",
           "Taxable Value (Rs)", "Tax (Rs)", "2B's Ineligibility Reason", "Occurrence # (this supplier, FY)"]
    r = 4
    for i, h in enumerate(hdr, 1):
        ws.cell(r, i, h)
    _style_header(ws, r, len(hdr))
    r += 1
    occ_counter = {}
    for f in sorted(occs, key=lambda f: ((f.numbers or {}).get("gstin", ""), _month_sort_key((f.numbers or {}).get("month", "")))):
        n = f.numbers or {}
        g = n.get("gstin", "")
        occ_counter[g] = occ_counter.get(g, 0) + 1
        vals = [n.get("month"), n.get("invno"), n.get("invdate"), g, n.get("supplier"),
               n.get("taxable"), n.get("tax"), n.get("reason"), occ_counter[g]]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(r, ci, v)
            cell.border = BORDER; cell.font = Font(size=10)
            if ci in (6, 7):
                cell.number_format = "#,##0.00"
        if occ_counter[g] >= 2:
            ws.cell(r, 9).fill = AMBER
        r += 1
    if not occs:
        ws.cell(r, 1, "No ITC-blocked invoices found.").font = Font(italic=True)
    for col, w in zip("ABCDEFGHI", [10, 18, 14, 18, 26, 16, 14, 40, 26]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:I{max(r-1,4)}"


def write_round_number_annual(ws, hsn_findings):
    """NEW sheet: 'Round-Number Invoices -- Annual Detail'. Only 2 occurrences in this dataset
    but a dedicated sheet regardless -- round-number invoices have unusually high fraud-relevance
    (classic accommodation-entry signal) and deserve their own tracking even at low volume."""
    ws.cell(1, 1, "ROUND-NUMBER INVOICES -- ANNUAL DETAIL (check #1)").font = TITLEF
    occs = [f for f in hsn_findings if f.ref == "#1" and f.severity != "PASS"]
    total_val = sum((f.numbers or {}).get("taxable") or 0 for f in occs)
    ws.cell(2, 1, f"{len(occs)} occurrence(s) for the FY; total taxable value Rs {total_val:,.2f}. "
                  "Heuristic only (exact-round taxable value AND exact-round tax on the same line) "
                  "-- genuine bulk/contract pricing can also land on round numbers.").font = Font(size=9, italic=True)
    hdr = ["Month", "Invoice No.", "Invoice Date", "Taxable Value (Rs)", "Tax (Rs)", "Rate (%)", "Buyer GSTIN"]
    r = 4
    for i, h in enumerate(hdr, 1):
        ws.cell(r, i, h)
    _style_header(ws, r, len(hdr))
    r += 1
    for f in occs:
        n = f.numbers or {}
        vals = [n.get("month"), n.get("invno"), n.get("invdate"), n.get("taxable"), n.get("tax"),
               n.get("rate"), n.get("buyer_gstin")]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(r, ci, v)
            cell.border = BORDER; cell.font = Font(size=10)
            if ci in (4, 5):
                cell.number_format = "#,##0.00"
        r += 1
    if not occs:
        ws.cell(r, 1, "No round-number invoices found.").font = Font(italic=True)
    for col, w in zip("ABCDEFG", [10, 20, 14, 16, 14, 10, 18]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:G{max(r-1,4)}"


def write_hsn_timeline_annual(ws, hsn_findings):
    """NEW sheet: 'HSN Timeline -- Annual Detail'. Every #27(HSN-new-code)/#6(mix-shift)
    occurrence -- distinct from EWB-pipeline's OWN #27 (same-vehicle); this is the HSN/Fraud
    pipeline's ref #27, tracked separately by construction (this function never reads
    gst_checks_monthly.py's #27 at all)."""
    ws.cell(1, 1, "HSN TIMELINE -- ANNUAL DETAIL (checks #27 new-code, #6 mix-shift)").font = TITLEF
    occs27 = [f for f in hsn_findings if f.ref == "#27" and f.title.startswith("New HSN") and f.severity != "PASS"]
    occs6 = [f for f in hsn_findings if f.ref == "#6" and f.severity != "PASS"]
    ws.cell(2, 1, f"{len(occs27)} new-HSN-code event(s), {len(occs6)} sharp mix-shift event(s) for the FY -- "
                  "useful for tracking product-line evolution or possible misclassification across the "
                  "year.").font = Font(size=9, italic=True)
    hdr = ["Check Type", "HSN Code", "Month First/Shift Occurred", "Taxable Value That Month (Rs)",
           "Share % That Month", "Share % Prior Month", "Result"]
    r = 4
    for i, h in enumerate(hdr, 1):
        ws.cell(r, i, h)
    _style_header(ws, r, len(hdr))
    r += 1
    for f in occs27:
        n = f.numbers or {}
        vals = ["New HSN code", n.get("hsn"), n.get("month"), n.get("taxable"), None, None, f.severity]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(r, ci, v); cell.border = BORDER; cell.font = Font(size=10)
            if ci == 4 and isinstance(v, (int, float)):
                cell.number_format = "#,##0.00"
            if ci == 7:
                cell.fill = SEV_FILL.get(v, GREY); cell.font = SEV_FONT.get(v, Font(size=10))
        r += 1
    for f in occs6:
        n = f.numbers or {}
        vals = ["Mix share shift", n.get("hsn"), n.get("month"), None, n.get("share_now"), n.get("share_prev"),
                f.severity]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(r, ci, v); cell.border = BORDER; cell.font = Font(size=10)
            if ci in (5, 6) and isinstance(v, (int, float)):
                cell.number_format = "0.0"
            if ci == 7:
                cell.fill = SEV_FILL.get(v, GREY); cell.font = SEV_FONT.get(v, Font(size=10))
        r += 1
    if not occs27 and not occs6:
        ws.cell(r, 1, "No HSN timeline events found.").font = Font(italic=True)
        r += 1
    r += 1
    ws.cell(r, 1, f"Summary: {len(occs27)} new-HSN event(s), {len(occs6)} mix-shift event(s), "
                  f"{sum(1 for f in occs27+occs6 if f.severity=='FLAG')} at FLAG severity."
                  ).font = Font(bold=True, size=10, color="1F3864")
    r += 2
    ws.cell(r, 1, "SUGGESTED FURTHER ANALYSIS from raw data already in this tool's inputs, not yet "
                  "built:").font = Font(bold=True, size=10, color="1F3864")
    r += 1
    for s in ["A full HSN-by-month MATRIX (every HSN down the rows, every month across the "
              "columns, taxable value in each cell) -- this sheet currently only lists the "
              "EXCEPTION events (new codes, sharp shifts); the underlying month-by-month HSN "
              "summary this tool already parses per month would let a reviewer see the whole "
              "product mix evolve, not just the flagged jumps.",
              "HSN-wise effective tax rate trend -- flag an HSN whose average rate changes "
              "mid-year with no corresponding rate-notification change, a possible "
              "misclassification signal distinct from the mix-shift check already here.",
              "Cross-reference against the Machinery HSN Scan and Potential Blocked Credits "
              "sheets' own HSN lists, so a new-code event on a machinery or blocked-credit "
              "heading gets flagged with extra weight instead of being just another row here."]:
        ws.cell(r, 1, "- " + s).font = Font(size=9, italic=True)
        ws.cell(r, 1).alignment = Alignment(wrap_text=True)
        r += 1
    for col, w in zip("ABCDEFG", [16, 14, 22, 22, 16, 18, 10]):
        ws.column_dimensions[col].width = w
    ws.auto_filter.ref = f"A4:G{max(r-1,4)}"


def write_master_dashboard(ws, month_results, months_covered, months_gap, rect_pairs,
                            annual_review_count, hsn_findings=None, forensic_findings=None,
                            cancel_findings=None, flow_findings=None, flow_ref_to_sheet=None):
    flow_ref_to_sheet = flow_ref_to_sheet or {}
    fys = sorted(set(_fy_label_for_month(m) for m in months_covered))
    ws.cell(1, 1, "MASTER DASHBOARD -- ALL MONTHS + ANNUAL SOURCES, RANKED TOGETHER").font = TITLEF
    ws.cell(2, 1, f"GSTIN {raw.SELF_GSTIN}  |  {raw.COMPANY_NAME}  |  "
                  f"generated {_dt.datetime.now():%Y-%m-%d %H:%M:%S}").font = Font(size=9, italic=True)
    ws.cell(3, 1, f"FY(s) covered: {', '.join(fys) or 'none'}  |  "
                  f"Months covered ({len(months_covered)}): {', '.join(months_covered) or 'none'}").font = Font(size=10, bold=True, color="006100")
    ws.cell(4, 1, f"Gaps WITHIN the covered span ({len(months_gap)}): {', '.join(months_gap) or 'none'} -- "
                  "a month between the earliest and latest month supplied that has NO GSTR-1+GSTR-3B pair; "
                  "this is NOT the same as 'nothing was supplied for that FY' if the FY itself wasn't in "
                  "scope for this run at all.").font = Font(size=10, bold=True, color="9C0006")

    # items: (rank, month, pipeline, ref, title, result, detail, amount, source_ref)
    items = []
    RANK = {"FLAG": 0, "MISMATCH": 0, "REVIEW": 1, "INFO": 2, "PASS": 3, "MATCH": 3, "EXPLAINED": 3, "SKIPPED": 4}
    for res in month_results:
        mo = res["month"]
        comp_sheet = sheet_name("Comparison", mo)
        for row in uni._comp_rows_iter(res["comparisons"]):
            sect, check, ll, lv, rl, rv, diff, result, tag = row
            if result == "MISMATCH":
                items.append((RANK[result], mo, "Comparison", sect.split(".")[0], check, result,
                              f"{ll}={lv:,.2f} vs {rl}={rv:,.2f} (diff {diff:,.2f}). {tag}".strip(),
                              diff, f"See '{comp_sheet}', section '{sect}'"))
        for f in res["findings14"]:
            if f.severity in ("FLAG", "REVIEW"):
                items.append((RANK[f.severity], mo, "Analysis(14)", f.ref, f.title, f.severity, f.detail,
                              _extract_amount(f.numbers), f"See '{sheet_name('Analysis', mo)}', ref {f.ref}"))
        for f in res["findings27"]:
            if f.sev in ("FLAG", "REVIEW"):
                items.append((RANK[f.sev], mo, "E-Way-Bill", f.ref, f.title, f.sev, f.detail,
                              _extract_ewb_amount(f),
                              f"See '{sheet_name('EWB', mo)}' or '{sheet_name('EWB Detail', mo)}', ref {f.ref}"))
        for dg in res["doc_gap"]:
            still_unexplained = dg.get("still_unexplained", dg.get("missing", []))
            explained_by_einv = dg.get("explained_by_cancelled_einvoice", [])
            ds_ref = f"See 'Doc-Series Integrity', range {dg['range']} ({mo})"
            if still_unexplained:
                items.append((0, mo, "Doc-Series", "Table13", "Missing invoice serials (unexplained)",
                               "FLAG", f"Range {dg['range']}: missing {still_unexplained}", None, ds_ref))
            elif dg.get("missing") and dg.get("explained_by_declared_cancellation"):
                items.append((1, mo, "Doc-Series", "Table13",
                               "Missing invoice serials (explained by declared cancellation)",
                               "REVIEW", f"Range {dg['range']}: missing {dg['missing']} -- matches "
                               f"Table 13's own declared cancelled count exactly.", None, ds_ref))
            elif explained_by_einv:
                items.append((1, mo, "Doc-Series", "Table13",
                               "Missing invoice serials (explained by cancelled e-invoice)",
                               "REVIEW", f"Range {dg['range']}: {explained_by_einv} -- each of these serials "
                               f"has a CANCELLED e-invoice against it, so its absence from GSTR-1 is expected, "
                               f"not a real gap.", None, ds_ref))
            elif dg.get("missing") and dg.get("explained_by_b2cs"):
                items.append((1, mo, "Doc-Series", "Table13",
                               "Missing invoice serials (explained by B2CS aggregate)",
                               "REVIEW", f"Range {dg['range']}: missing {dg['missing']} -- this "
                               f"taxpayer's B2C sales are legally reported ONLY as a Table 7 (B2CS) "
                               f"state+rate aggregate, which structurally carries no invoice number "
                               f"anywhere in a GSTR-1 export. This is the ONLY unexplained range this "
                               f"month, and B2CS shows non-zero taxable value {dg.get('b2cs_taxable', 0):,.2f} "
                               f"(tax {dg.get('b2cs_tax', 0):,.2f}) this month -- see this month's "
                               f"Comparison sheet, section 'A. Outward Liability'.",
                               dg.get("b2cs_taxable"), f"See '{comp_sheet}', section 'A. Outward Liability'"))

    for f in (hsn_findings or []):
        if f.severity in ("FLAG", "REVIEW"):
            mo = f.numbers.get("month", "FY-wide")
            items.append((RANK[f.severity], mo, "HSN/Fraud", f.ref, f.title, f.severity, f.detail,
                          _extract_amount(f.numbers), f"See 'HSN & Fraud Pattern Checks', ref {f.ref}"))

    # NEW: forensic (R13/R14) and cancelled-e-invoice cross-check findings now also feed the
    # ranked dashboard -- previously these only appeared on their own dedicated sheets, so a
    # FLAG here was invisible from the top-level view unless you opened that specific sheet.
    for f in (forensic_findings or []):
        if f.severity in ("FLAG", "REVIEW"):
            items.append((RANK[f.severity], "FY-wide", "Forensic (R13/R14)", f.ref, f.title, f.severity,
                          f.detail, _extract_amount(f.numbers),
                          f"See 'Forensic Checks (R13-R14)', ref {f.ref}"))
    for f in (cancel_findings or []):
        if f.severity in ("FLAG", "REVIEW"):
            items.append((RANK[f.severity], "FY-wide", "Cancelled E-Inv", f.ref, f.title, f.severity,
                          f.detail, _extract_amount(f.numbers), f"See 'Cancelled E-Invoices', ref {f.ref}"))
    # Flow / stock / ITC roll-forward / payment / counterparty findings (gst_checks_flow.py).
    # Same treatment as every other FY-wide pipeline: only FLAG and REVIEW are promoted to the
    # dashboard; the full set including INFO/PASS/SKIPPED stays on each sheet's own findings block.
    for f in (flow_findings or []):
        if f.severity in ("FLAG", "REVIEW", "MISMATCH"):
            sheet = flow_ref_to_sheet.get(f.ref)
            items.append((RANK[f.severity], "FY-wide", "Flow/Counterparty", f.ref, f.title,
                          f.severity, f.detail, _extract_amount(f.numbers),
                          f"See '{sheet}', ref {f.ref}" if sheet else f"See the Flow/Counterparty sheet for ref {f.ref}"))

    items.sort(key=lambda x: (x[0], _month_sort_key(x[1]) if x[1] != "FY-wide" else (9998, 0), x[2]))

    # ---- Dedup/rollup system (root_id + level) ----
    # Extends every item with (root_id, level). See _root_id_and_level()'s own docstring for the
    # exact, conservative merge rules. This never removes a row -- VIEW/ROLLUP rows stay fully
    # visible on this same sheet, only their counting status changes.
    items = [it + _root_id_and_level(it[2], it[3], it[4], it[1]) for it in items]

    # Sanity check: for every root_id family with both PARENT rows and a ROLLUP row, compare the
    # PARENT sum against the ROLLUP. NOTE this is NOT always expected to match exactly even for a
    # genuine, correctly-scoped rollup: a ROLLUP like F4 sums ALL 12 months unconditionally,
    # while the PARENT rows shown on this dashboard are only the months that individually
    # crossed that check's own materiality threshold -- months with a small offsetting gap in
    # the OTHER direction never become their own PARENT row, so the two totals can legitimately
    # differ. This is reported as CONTEXT, not an alarm -- a genuine mismatch worth a second look
    # is a PARENT-sum that EXCEEDS the rollup by more than the rollup itself (the sign/scale
    # would make no arithmetic sense), not any nonzero difference.
    rollup_checks = []
    by_family = {}
    for it in items:
        root_id, level = it[9], it[10]
        family = root_id.rsplit(":", 1)[0]   # e.g. "outward_gap" from "outward_gap:Aug-25"
        by_family.setdefault(family, {"PARENT": [], "ROLLUP": []})
        if level in ("PARENT", "ROLLUP"):
            by_family[family][level].append(it)
    for family, buckets in by_family.items():
        if buckets["PARENT"] and buckets["ROLLUP"]:
            parent_sum = sum(abs(it[7]) for it in buckets["PARENT"] if it[7] is not None)
            rollup_amt = next((abs(it[7]) for it in buckets["ROLLUP"] if it[7] is not None), None)
            if rollup_amt is not None:
                # Flag only the arithmetically-implausible direction (parent sum more than
                # double the rollup, or the rollup zero while parents are large) -- an ordinary
                # "rollup includes offsetting months the flagged-only parent sum doesn't" gap is
                # expected and shown as CONTEXT, not CHECK.
                implausible = rollup_amt == 0 and parent_sum > 0 or parent_sum > rollup_amt * 2
                rollup_checks.append((family, parent_sum, rollup_amt,
                                      "VERIFY -- LOOKS OFF" if implausible else "context (rollup spans all months; parents are flagged-months only)"))

    n_distinct_root_causes = len({it[9] for it in items})
    parent_items = [it for it in items if it[10] == "PARENT"]
    parent_action_sum = sum(abs(it[7]) for it in parent_items
                            if it[7] is not None and it[5] in ("FLAG", "MISMATCH"))

    FYWIDE_PIPELINES = ("HSN/Fraud", "Forensic (R13/R14)", "Cancelled E-Inv", "Flow/Counterparty")
    monthly_items = [it for it in items if it[2] not in FYWIDE_PIPELINES]
    fywide_items = [it for it in items if it[2] in FYWIDE_PIPELINES]
    nflag_m = sum(1 for it in monthly_items if it[5] in ("FLAG", "MISMATCH"))
    nrev_m = sum(1 for it in monthly_items if it[5] == "REVIEW")
    nflag_f = sum(1 for it in fywide_items if it[5] in ("FLAG", "MISMATCH"))
    nrev_f = sum(1 for it in fywide_items if it[5] == "REVIEW")
    ws.cell(6, 1, f"Monthly actionable items: {len(monthly_items)}  (FLAG/MISMATCH: {nflag_m}  REVIEW: {nrev_m})   |   "
                  f"FY-wide items (HSN/Fraud + Forensic + Cancelled E-Inv + Flow/Counterparty): {len(fywide_items)}  "
                  f"(FLAG: {nflag_f}  REVIEW: {nrev_f})   |   "
                  f"ANNUAL-source REVIEW items: {annual_review_count}   |   "
                  f"Rectification pairs found: {len(rect_pairs)}").font = Font(bold=True, size=11, color="C00000")
    ws.cell(7, 1, f"{n_distinct_root_causes} DISTINCT ROOT-CAUSE ISSUES (dedup applied: rows sharing a "
                  f"root_id are mathematically derived from the same source computation -- see 'Level' "
                  f"column; VIEW/ROLLUP rows are shown but never double-counted). PARENT-level FLAG/"
                  f"MISMATCH total exposure: Rs {parent_action_sum:,.2f}."
                  + ("  Rollup context: " + "; ".join(
                        f"{fam} parent-sum={p:,.0f} vs rollup={r:,.0f} ({s})"
                        for fam, p, r, s in rollup_checks) if rollup_checks else "")
                  ).font = Font(bold=True, size=10, color="1F3864")

    hdr = ["Month", "Pipeline", "Ref/Section", "Check", "Result", "Severity", "Source Reference", "Detail",
           "Root ID", "Level"]
    for i, h in enumerate(hdr, 1):
        ws.cell(8, i, h)
    _style_header(ws, 8, len(hdr))
    r = 9
    LEVEL_FILL = {"VIEW": PatternFill("solid", fgColor="F2F2F2"), "ROLLUP": PatternFill("solid", fgColor="DDEBF7")}
    for (_, mo, pipeline, ref, title, result, detail, amount, source_ref, root_id, level) in items:
        ws.cell(r, 1, mo); ws.cell(r, 2, pipeline); ws.cell(r, 3, ref); ws.cell(r, 4, title)
        cv = ws.cell(r, 5, result); cv.fill = SEV_FILL[result]; cv.font = SEV_FONT[result]
        cv.alignment = Alignment(horizontal="center")
        band = uni.severity_band(amount, resolved=(result == "EXPLAINED"))
        sv = ws.cell(r, 6, band or "")
        if band:
            sv.fill = uni.SEV_BAND_FILL[band]
            sv.font = uni.SEV_BAND_FONT.get(band, Font(bold=True))
        sv.alignment = Alignment(horizontal="center")
        ws.cell(r, 7, source_ref or "")
        ws.cell(r, 8, detail)
        ws.cell(r, 9, root_id)
        lv_cell = ws.cell(r, 10, level)
        lv_cell.font = Font(bold=(level == "PARENT"), size=10,
                            color="1F3864" if level == "PARENT" else "808080")
        for c in range(1, 11):
            cell = ws.cell(r, c); cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(c in (4, 7, 8)))
            if c not in (5, 6, 10):
                cell.font = Font(size=10, italic=(level != "PARENT"))
            if level in LEVEL_FILL and c not in (5, 6):
                cell.fill = LEVEL_FILL[level]
        # Nice-to-have: VIEW/ROLLUP rows outlined and collapsed by default under their PARENT --
        # "one finding, expandable detail" instead of a flat list of near-duplicates. Nothing is
        # hidden in a way that loses data: Excel shows a '+' outline control to expand the group,
        # and AutoFilter (below) still finds these rows regardless of collapse state.
        if level in ("VIEW", "ROLLUP"):
            ws.row_dimensions[r].outlineLevel = 1
            ws.row_dimensions[r].hidden = True
        r += 1
    ws.sheet_properties.outlinePr.summaryBelow = False
    if not items:
        ws.cell(r, 1, "No FLAG / MISMATCH / REVIEW across any supplied month.").font = Font(italic=True, color="006100")
    for col, w in zip("ABCDEFGHIJ", [9, 14, 12, 38, 11, 22, 45, 65, 22, 10]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A9"
    ws.auto_filter.ref = f"A8:J{max(r-1, 8)}"
    return items   # tagged (root_id, level, amount, band, ...) -- reused directly by the
                    # QA-review sheets below, so their amounts/severity never drift from what
                    # this sheet itself shows (this is also the fix for the cross-category
                    # false-duplicate bug in Issue 3 -- see _root_id_and_level()'s docstring).


# ======================================================================
# QA REVIEW LAYER (integrated into the tool -- was previously a set of standalone, manual
# post-processing scripts run after the fact; now runs automatically as part of every build,
# reusing the SAME root_id/level/amount data the Master Dashboard itself already computed, so
# nothing here can drift out of sync with the numbers a reader sees on the main sheets).
# Produces 4 sheets, appended at the very END of the workbook: QA Summary, Action Required,
# EWB Full-Year Reconciliation, Reviewed Master Dashboard -- a reader who wants raw numbers
# works from the front of the workbook; a reader who wants a colour-coded triage view goes to
# the back. Nothing in between is touched.
# ======================================================================
QA_STATUS_FILL = {
    "VERIFIED - NO ACTION": PatternFill("solid", fgColor="D9EAD3"),
    "INFORMATIONAL": PatternFill("solid", fgColor="F3F3F3"),
    "GENUINE - CORRECTLY FLAGGED": PatternFill("solid", fgColor="FFF2CC"),
    "SPOT-CHECK RECOMMENDED": PatternFill("solid", fgColor="CFE2F3"),
    "GENUINE - ACTION NEEDED": PatternFill("solid", fgColor="F4CCCC"),
}
QA_ICON = {
    "VERIFIED - NO ACTION": "\U0001F7E2 VERIFIED - NO ACTION",
    "INFORMATIONAL": "\u26AA INFORMATIONAL",
    "GENUINE - CORRECTLY FLAGGED": "\U0001F7E1 GENUINE - CORRECTLY FLAGGED",
    "SPOT-CHECK RECOMMENDED": "\U0001F535 SPOT-CHECK RECOMMENDED",
    "GENUINE - ACTION NEEDED": "\U0001F534 GENUINE - ACTION NEEDED",
}
QA_SEV_ICON = {"Critical": "\U0001F6A8 CRITICAL", "High": "\U0001F534 HIGH",
               "Medium": "\U0001F7E0 MEDIUM", "Low": "\U0001F7E1 LOW"}
QA_SEV_ORDER = {"\U0001F6A8 CRITICAL": 0, "\U0001F534 HIGH": 1, "\U0001F7E0 MEDIUM": 2,
                "\U0001F7E1 LOW": 3, "": 4}

# (pipeline, ref) -> (status_key, reasoning). Generic per CHECK TYPE, not this run's own numbers
# -- transfers to any taxpayer this tool runs against, since it encodes what each check family
# fundamentally tests. Anything not covered here falls back to a conservative SPOT-CHECK
# RECOMMENDED rather than a guessed category -- see _classify_qa().
QA_REF_RULES = {
    ("Analysis(14)", "#14"): ("INFORMATIONAL",
        "Ratio/threshold heuristic on genuine, correctly-matched numbers (3B liability vs 3B ITC "
        "claimed) -- not a cross-source mismatch. A standing watch-item, not a reconciliation task."),
    ("Analysis(14)", "#2"): ("GENUINE - ACTION NEEDED",
        "Same root cause as the FY-wide rollup and the Comparison 'A' net-of-CN rows for this "
        "month, sharing the same root_id -- one underlying issue, several views."),
    ("Analysis(14)", "#8"): ("GENUINE - CORRECTLY FLAGGED",
        "Real e-invoicing timing fact from IRN-generation timestamps vs invoice/filing dates -- "
        "not a cross-source match. Rule 48(5) treats a late/backdated IRN as a compliance defect "
        "regardless of amount. Correctly flagged; keep as a standing compliance item."),
    ("Comparison", "A"): ("GENUINE - ACTION NEEDED",
        "GSTR-1 vs GSTR-3B outward-value gap -- see the linked root_id family for the other views "
        "of this same underlying fact (CGST/SGST split, tax-figure wording, FY rollup)."),
    ("Comparison", "D"): ("GENUINE - ACTION NEEDED",
        "IGST does not follow the CN-netting identity that resolves CGST/SGST for this filer in "
        "every month -- see the linked root_id family and the FY-wide rollup for full-year context."),
    ("Comparison", "D2"): ("GENUINE - ACTION NEEDED",
        "IGST-specific reversal residual -- kept as its own independent finding (not merged with "
        "the 'D' ITC-claimed gap): 4(B)(2) and 4(A)(5) are different return fields, not a "
        "mathematical derivation of one another, even though both trace to the same broader "
        "'IGST doesn't net cleanly for this filer' pattern."),
    ("Doc-Series", "Table13"): ("VERIFIED - NO ACTION",
        "Already carries an exact tie-out or structural explanation directly in its own Detail "
        "text -- resolved by the tool itself against evidence already in this workbook."),
    ("E-Way-Bill", "#5"): ("SPOT-CHECK RECOMMENDED",
        "Remaining gap after zero-tax exclusion. B2C/sub-threshold EWBs can legitimately have no "
        "e-invoice; cross-reference against #3/#17 for the same month before treating as unresolved."),
    ("E-Way-Bill", "#11"): ("SPOT-CHECK RECOMMENDED",
        ">1% value gap on an invoice already matched by number -- can legitimately reflect "
        "freight/discount, but worth a quick look since it's on an otherwise-confirmed invoice."),
    ("E-Way-Bill", "#12"): ("GENUINE - ACTION NEEDED",
        "Confirmed via the full-year re-match (see 'EWB Full-Year Reconciliation'): genuinely "
        "absent from GSTR-2B for the entire FY, not a same-month timing artifact. Chase the supplier."),
    ("E-Way-Bill", "#13"): ("SPOT-CHECK RECOMMENDED",
        "By design this check cannot distinguish a real Rule 138 lapse from a legitimate "
        "non-goods-movement invoice (services, job-work, delivery challan) -- its own detail text "
        "says so. Needs the taxpayer's own supplier-relationship knowledge, not a tool change."),
    ("E-Way-Bill", "#15"): ("SPOT-CHECK RECOMMENDED",
        "EWB-Out vs EWB-In value on a shared document number -- a real pattern worth a look, not "
        "independently proven as an issue from this check alone."),
    ("E-Way-Bill", "#16"): ("GENUINE - CORRECTLY FLAGGED",
        "Timing fact (EWB generation date vs document date) sourced directly from the EWB "
        "export's own timestamps -- not a cross-source match. Genuine compliance-timing signal."),
    ("E-Way-Bill", "#17"): ("SPOT-CHECK RECOMMENDED",
        "Triangulation gap after zero-tax exclusion. Remaining gaps mix legitimate cross-source "
        "coverage variance with cases already resolved elsewhere (e.g. via a CDNR match on #3) "
        "that this check's own logic doesn't yet cross-reference -- treat as a starting point."),
    ("E-Way-Bill", "#23"): ("SPOT-CHECK RECOMMENDED",
        "Multiple EWBs on one invoice is allowed for genuine partial dispatch, but the same "
        "pattern can be used to stay under sub-threshold limits -- not provable either way from "
        "EWB data alone."),
    ("E-Way-Bill", "#25"): ("SPOT-CHECK RECOMMENDED",
        "Where GSTR-1 B2B taxable value is Rs 0 for the month, this ratio is mathematically "
        "degenerate (EWB value / 0) rather than a meaningful 'goods moved without invoice' signal."),
    ("E-Way-Bill", "#27"): ("SPOT-CHECK RECOMMENDED",
        "Same-vehicle-repeated-trips pattern heuristic -- consistent with an ordinary regular-"
        "route supplier relationship, but not provable as benign from EWB data alone. See the "
        "'EWB Pattern - Annual' sheet for the full-year frequency/value view."),
    ("Flow/Counterparty", "F1"): ("INFORMATIONAL",
        "FY-wide monitoring/context item -- tests purchase-vs-sales value plausibility, a "
        "genuinely different question from GSTR-1-vs-3B filing consistency even though it cites "
        "one shared input figure. The tool's own detail text covers the caveats."),
    ("Flow/Counterparty", "F4"): ("GENUINE - ACTION NEEDED",
        "FY-wide rollup of the confirmed GSTR-1-vs-GSTR-3B outward-value gap -- see the linked "
        "root_id family (level=ROLLUP) for how this relates to the monthly PARENT rows."),
    ("Flow/Counterparty", "F7"): ("GENUINE - ACTION NEEDED",
        "FY-wide rollup of the confirmed IGST ITC residual -- see the linked root_id family "
        "(level=ROLLUP)."),
    ("Flow/Counterparty", "F9"): ("INFORMATIONAL", "FY-wide monitoring/context item -- the tool's own text covers the caveats."),
    ("Flow/Counterparty", "F11"): ("INFORMATIONAL", "FY-wide monitoring/context item -- the tool's own text covers the caveats."),
    ("Flow/Counterparty", "F11a"): ("INFORMATIONAL", "FY-wide monitoring/context item -- the tool's own text covers the caveats."),
    ("Flow/Counterparty", "F12"): ("INFORMATIONAL", "FY-wide monitoring/context item (supplier concentration) -- common for a project-based business."),
    ("Flow/Counterparty", "F12a"): ("INFORMATIONAL", "FY-wide monitoring/context item (buyer concentration) -- common for a project-based business."),
    ("Flow/Counterparty", "G1"): ("INFORMATIONAL", "GSTR-2A cross-check, FY-wide monitoring item -- see the extended 'GSTR-2A Data Quality' sheet for the occurrence# detail."),
    ("Flow/Counterparty", "G2"): ("INFORMATIONAL", "GSTR-2A cross-check, FY-wide monitoring item -- the tool's own detail text explains the caveats."),
    ("Flow/Counterparty", "G2a"): ("INFORMATIONAL", "GSTR-2A cross-check, FY-wide monitoring item -- the tool's own detail text explains the caveats."),
    ("Flow/Counterparty", "G3"): ("INFORMATIONAL", "GSTR-2A cross-check, FY-wide monitoring item -- the tool's own detail text explains the caveats."),
    ("Flow/Counterparty", "G5"): ("INFORMATIONAL", "GSTR-2A cross-check, FY-wide monitoring item -- the tool's own detail text explains the caveats."),
    ("Flow/Counterparty", "G6"): ("INFORMATIONAL", "GSTR-2A cross-check, FY-wide monitoring item -- 2A's RCM flag is supplier-declared and informational only per the tool's own text."),
    ("Flow/Counterparty", "G7"): ("SPOT-CHECK RECOMMENDED",
        "State-code vs tax-head validation -- a wrong tax head on a specific invoice IS "
        "independently checkable and could mean inadmissible ITC even where the total is correct."),
    ("Flow/Counterparty", "G10"): ("INFORMATIONAL", "GSTR-2A cross-check, FY-wide monitoring item -- the tool's own detail text explains the caveats."),
    ("HSN/Fraud", "#1"): ("SPOT-CHECK RECOMMENDED", "Round-number-invoice heuristic -- see the 'Round-Number Inv - Annual' sheet."),
    ("HSN/Fraud", "#23"): ("GENUINE - CORRECTLY FLAGGED",
        "IRN-vs-invoice-date timing fact sourced directly from IRN timestamps -- not a cross-source "
        "match. See the 'IRN Late-Gen - Annual' sheet for the full list."),
    ("HSN/Fraud", "#24"): ("SPOT-CHECK RECOMMENDED", "EWB destination-state mix shift -- see the 'EWB Pattern - Annual' sheet."),
    ("HSN/Fraud", "#27"): ("SPOT-CHECK RECOMMENDED", "New HSN code appearing mid/late year -- see the 'HSN Timeline - Annual' sheet."),
    ("HSN/Fraud", "#47/#55"): ("GENUINE - CORRECTLY FLAGGED",
        "Genuine cash-ledger deposit-timing pattern -- worth knowing for section 50 interest-exposure awareness."),
    ("HSN/Fraud", "#6"): ("INFORMATIONAL", "HSN mix share shift month-over-month -- see the 'HSN Timeline - Annual' sheet; a monitoring heuristic, not a mismatch."),
    ("HSN/Fraud", "B3"): ("GENUINE - CORRECTLY FLAGGED",
        "Sourced directly from GSTR-2B's own 'ITC Availability' flag -- see the 'ITC-Blocked "
        "Invoices - Annual Detail' sheet for the full list with per-supplier occurrence#."),
}


def _classify_qa(pipeline, ref, title):
    key = (pipeline, ref)
    if key in QA_REF_RULES:
        return QA_REF_RULES[key]
    return ("SPOT-CHECK RECOMMENDED", "Not covered by a specific QA rule for this check type -- "
            "classified conservatively pending review rather than guessed.")


def build_qa_layer(dash_items):
    """Classify every Master-Dashboard item once; shared by all 4 QA sheets below so their
    counts can never disagree with each other. Returns a list of dicts, one per dash_items row,
    each carrying: the original fields, plus status_key/status_icon/reasoning/sev_icon."""
    out = []
    for it in dash_items:
        (_rank, mo, pipeline, ref, title, result, detail, amount, source_ref, root_id, level) = it
        status_key, reasoning = _classify_qa(pipeline, ref, title)
        band = uni.severity_band(amount, resolved=(result == "EXPLAINED"))
        sev_icon = QA_SEV_ICON.get(band, "")
        out.append(dict(mo=mo, pipeline=pipeline, ref=ref, title=title, result=result, detail=detail,
                        amount=amount, source_ref=source_ref, root_id=root_id, level=level,
                        status_key=status_key, status_icon=QA_ICON[status_key], reasoning=reasoning,
                        band=band, sev_icon=sev_icon))
    return out


def write_qa_reviewed_dashboard(ws, qa_rows):
    ws.cell(1, 1, "REVIEWED MASTER DASHBOARD -- QA CATEGORISED, COLOUR-CODED").font = TITLEF
    ws.cell(2, 1, "Every row currently on the Master Dashboard (nothing added or removed here), "
                  "classified into a 5-tier QA status and cross-referenced to its root_id/level "
                  "from the dedup system -- VIEW/ROLLUP rows are shown but excluded from any "
                  "total-exposure figure.").font = Font(size=9, italic=True)
    hdr = ["Month", "Pipeline", "Ref", "Check", "Result", "Detail (original)", "Status", "Severity",
           "Rs Impact", "Root ID", "Level", "QA Category", "QA Reasoning"]
    r = 4
    for i, h in enumerate(hdr, 1):
        ws.cell(r, i, h)
    style_header_local(ws, r, len(hdr))
    r += 1
    for row in qa_rows:
        vals = [row["mo"], row["pipeline"], row["ref"], row["title"], row["result"], row["detail"],
               row["status_icon"], row["sev_icon"], row["amount"], row["root_id"], row["level"],
               row["status_icon"], row["reasoning"]]
        fill = QA_STATUS_FILL[row["status_key"]]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(r, ci, v)
            cell.fill = fill
            cell.font = Font(color="1F2937", size=10, italic=(row["level"] != "PARENT"))
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(ci in (4, 6, 13)))
            if ci == 9 and isinstance(v, (int, float)):
                cell.number_format = "#,##0.00"
        r += 1
    for col, w in zip("ABCDEFGHIJKLM", [9, 14, 9, 32, 10, 42, 24, 14, 13, 22, 10, 24, 55]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:M{max(r-1, 4)}"
    if r > 5:
        ws.conditional_formatting.add(f"I5:I{r-1}", DataBarRule(start_type="min", end_type="max", color="638EC6"))


def write_qa_action_required(ws, qa_rows):
    action_rows = [row for row in qa_rows if row["status_key"] == "GENUINE - ACTION NEEDED"
                   and row["level"] == "PARENT"]   # PARENT-only -- see Issue-1/2 dedup fix
    action_rows.sort(key=lambda row: (QA_SEV_ORDER.get(row["sev_icon"], 4), -(row["amount"] or 0)))
    total_exposure = sum(abs(row["amount"]) for row in action_rows if row["amount"] is not None)

    ws.cell(1, 1, f"The genuine, unresolved PARENT-level items only -- worst first. Rows that are "
                  f"mathematically derived from the same root cause (VIEW) or an FY rollup of these "
                  f"same rows (ROLLUP) are NOT repeated here to avoid double-counting -- see the "
                  f"'Reviewed Master Dashboard' tab for those, cross-referenced by Root ID. "
                  f"Total PARENT-level exposure: Rs {total_exposure:,.2f}.").font = Font(italic=True, size=9)
    ws.row_dimensions[1].height = 30
    hdr = ["Severity", "Month", "Check", "Result", "Rs Impact", "Detail (original tool text)",
           "Root ID", "QA Reasoning"]
    r = 3
    for i, h in enumerate(hdr, 1):
        ws.cell(r, i, h)
    style_header_local(ws, r, len(hdr))
    r += 1
    for row in action_rows:
        vals = [row["sev_icon"], row["mo"], row["title"], row["result"], row["amount"], row["detail"],
               row["root_id"], row["reasoning"]]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(r, ci, v)
            cell.border = BORDER
            cell.font = Font(size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=(ci in (3, 6, 8)))
            if ci == 5 and isinstance(v, (int, float)):
                cell.number_format = "#,##0.00"
        r += 1
    if not action_rows:
        ws.cell(r, 1, "No genuine PARENT-level action-needed items.").font = Font(italic=True)
    for col, w in zip("ABCDEFGH", [16, 10, 40, 11, 15, 50, 20, 55]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"
    if action_rows:
        ws.conditional_formatting.add(f"E4:E{r-1}", DataBarRule(start_type="min", end_type="max", color="E06666"))
    ws.auto_filter.ref = f"A3:H{max(r-1,3)}"
    return action_rows, total_exposure


def write_qa_ewb_reconciliation(ws, month_results):
    """Data-driven (not hardcoded to any one taxpayer) -- pulled fresh from this run's own #12
    (inward) and #3 (outward) raw records, so it stays correct for whichever taxpayer's data the
    tool is pointed at.

    UPDATED (per instruction): #3/#12's `.rows` now carry the FULL enriched invoice-detail
    schema (see gst_checks_monthly.OUT_DETAIL_HDR / IN_DETAIL_HDR), not the old bare
    (doc-no, gstin, assess, tax[, classification]) tuples this function used to slice by fixed
    position. Column positions below are resolved BY HEADER NAME against the real header row
    each Finding carries (row[0]), so this stays correct even if the enriched schema's column
    ORDER changes later -- never a silent fixed-offset read of a differently-shaped row again
    (the same class of bug this project's own history (Table 8A, R2A) has already been bitten by
    twice)."""
    ws.cell(1, 1, "E-WAY BILL -- FULL-YEAR RECONCILIATION").font = TITLEF
    ws.cell(2, 1, "Computed by the tool's own corrected EWB matching engine (invoice-number "
                  "date-suffix stripped, FY-wide 2B late-filing index applied) -- not a manual "
                  "one-off re-match.").font = Font(size=9, italic=True)
    r = 4
    only_ewb_out, only_ewb_in = [], []
    for res in month_results:
        for f in res.get("findings27", []):
            if f.ref == "#3" and f.rows:
                hdr3 = f.rows[0]
                i_doc = hdr3.index("EWB Doc No") if "EWB Doc No" in hdr3 else 0
                i_to = hdr3.index("To GSTIN") if "To GSTIN" in hdr3 else 4
                i_assess = hdr3.index("EWB Assess (Rs)") if "EWB Assess (Rs)" in hdr3 else 7
                i_tax = hdr3.index("EWB Tax (Rs)") if "EWB Tax (Rs)" in hdr3 else 8
                i_note = hdr3.index("Note / Classification") if "Note / Classification" in hdr3 else len(hdr3) - 1
                for row in f.rows[1:]:
                    if len(row) > i_note and "NO 2B credit-note match" in str(row[i_note]):
                        only_ewb_out.append((res["month"], row[i_doc], row[i_to], row[i_assess], row[i_tax]))
            if f.ref == "#12" and f.rows:
                hdr12 = f.rows[0]
                i_doc = hdr12.index("Doc/Invoice No") if "Doc/Invoice No" in hdr12 else 0
                i_gstin = hdr12.index("Supplier GSTIN") if "Supplier GSTIN" in hdr12 else 4
                i_assess = hdr12.index("EWB Assess (Rs)") if "EWB Assess (Rs)" in hdr12 else 8
                i_tax = hdr12.index("EWB Tax (Rs)") if "EWB Tax (Rs)" in hdr12 else 9
                for row in f.rows[1:]:
                    only_ewb_in.append((res["month"], row[i_doc], row[i_gstin], row[i_assess], row[i_tax]))
    ws.cell(r, 1, "SUMMARY").font = Font(bold=True, size=12, color="1F3864"); r += 2
    ws.cell(r, 1, f"Genuine OUTWARD gaps (tax-bearing, no GSTR-1 invoice, no 2B credit-note match): {len(only_ewb_out)}"); r += 1
    ws.cell(r, 1, f"Genuine INWARD gaps (tax-bearing, no 2B match anywhere in the FY): {len(only_ewb_in)}"); r += 2

    ws.cell(r, 1, "GENUINE OUTWARD GAPS").font = Font(bold=True, size=12, color="1F3864"); r += 1
    hdr = ["Month", "Doc No.", "Supplier/Buyer GSTIN", "Assessable (Rs)", "Tax (Rs)"]
    for i, h in enumerate(hdr, 1):
        ws.cell(r, i, h)
    style_header_local(ws, r, len(hdr)); r += 1
    for row_data in only_ewb_out:
        for ci, v in enumerate(row_data, 1):
            cell = ws.cell(r, ci, v); cell.border = BORDER; cell.font = Font(size=10)
            if ci in (4, 5) and isinstance(v, (int, float)):
                cell.number_format = "#,##0.00"
        r += 1
    if not only_ewb_out:
        ws.cell(r, 1, "None this FY.").font = Font(italic=True); r += 1
    r += 1

    ws.cell(r, 1, "GENUINE INWARD GAPS").font = Font(bold=True, size=12, color="1F3864"); r += 1
    for i, h in enumerate(hdr, 1):
        ws.cell(r, i, h)
    style_header_local(ws, r, len(hdr)); r += 1
    for row_data in only_ewb_in:
        for ci, v in enumerate(row_data, 1):
            cell = ws.cell(r, ci, v); cell.border = BORDER; cell.font = Font(size=10)
            if ci in (4, 5) and isinstance(v, (int, float)):
                cell.number_format = "#,##0.00"
        r += 1
    if not only_ewb_in:
        ws.cell(r, 1, "None this FY.").font = Font(italic=True); r += 1
    for col, w in zip("ABCDE", [10, 24, 22, 16, 14]):
        ws.column_dimensions[col].width = w
    return only_ewb_out, only_ewb_in


def write_qa_summary(ws, qa_rows, action_rows, total_exposure, ewb_out_gaps, ewb_in_gaps, self_gstin, company_name, fy_label):
    cat_counts = {}
    for row in qa_rows:
        cat_counts[row["status_key"]] = cat_counts.get(row["status_key"], 0) + 1
    sev_counts = {}
    for row in action_rows:
        sev_counts[row["sev_icon"]] = sev_counts.get(row["sev_icon"], 0) + 1
    n_distinct_root_causes = len({row["root_id"] for row in qa_rows})

    ws.cell(1, 1, "GST MASTER DASHBOARD -- QA REVIEW").font = Font(bold=True, size=14, color="1F3864")
    ws.cell(2, 1, f"{company_name}  |  GSTIN {self_gstin}  |  FY {fy_label}  |  "
                  f"Generated automatically as part of the standard build (dedup/root_id system "
                  f"applied).").font = Font(size=9, italic=True)
    r = 4
    ws.cell(r, 1, "5-TIER SYSTEM").font = Font(bold=True, size=11, color="1F3864"); r += 1
    tier_meaning = {
        "VERIFIED - NO ACTION": "Already carries an exact tie-out or structural explanation directly in the tool's own output.",
        "SPOT-CHECK RECOMMENDED": "Real, consistent pattern-detection signal -- not individually proven either way, worth a quick look.",
        "INFORMATIONAL": "Tool's own REVIEW label is already correct -- a monitoring signal, not a mismatch.",
        "GENUINE - CORRECTLY FLAGGED": "Real, factual finding that's already correctly labeled -- nothing to fix in the tool, keep on the radar.",
        "GENUINE - ACTION NEEDED": "Real, unresolved gap or reconciliation item -- see 'Action Required' tab (PARENT-level rows only).",
    }
    for key in ["VERIFIED - NO ACTION", "SPOT-CHECK RECOMMENDED", "INFORMATIONAL",
               "GENUINE - CORRECTLY FLAGGED", "GENUINE - ACTION NEEDED"]:
        ws.cell(r, 1, QA_ICON[key]); ws.cell(r, 2, cat_counts.get(key, 0)); ws.cell(r, 3, tier_meaning[key])
        fill = QA_STATUS_FILL[key]
        for c in range(1, 4):
            ws.cell(r, c).fill = fill; ws.cell(r, c).border = BORDER
        r += 1
    r += 1
    ws.cell(r, 1, f"{n_distinct_root_causes} DISTINCT ROOT-CAUSE ISSUES across all {len(qa_rows)} rows "
                  f"(dedup applied -- this is the number that should actually be quoted as 'how many "
                  f"problems does this taxpayer have').").font = Font(bold=True, size=11, color="C00000")
    r += 2

    ws.cell(r, 1, "SEVERITY SCALE (GENUINE - ACTION NEEDED, PARENT-level rows only)").font = Font(bold=True, size=11, color="1F3864"); r += 1
    for sev, rng in [("\U0001F6A8 CRITICAL", "\u2265 Rs 10,00,000"), ("\U0001F534 HIGH", "Rs 1,00,000 - 9,99,999"),
                     ("\U0001F7E0 MEDIUM", "Rs 10,000 - 99,999"), ("\U0001F7E1 LOW", "< Rs 10,000")]:
        ws.cell(r, 1, sev); ws.cell(r, 2, rng); ws.cell(r, 3, sev_counts.get(sev, 0))
        for c in range(1, 4):
            ws.cell(r, c).border = BORDER
        r += 1
    ws.cell(r, 1, "TOTAL PARENT-LEVEL EXPOSURE").font = Font(bold=True)
    ws.cell(r, 3, total_exposure).number_format = "#,##0.00"; r += 2

    ws.cell(r, 1, "TOP GENUINE FINDINGS").font = Font(bold=True, size=11, color="1F3864"); r += 1
    for row in [a for a in action_rows if a["sev_icon"] == "\U0001F6A8 CRITICAL"][:4]:
        ws.cell(r, 1, f"{row['sev_icon']} -- {row['title']} ({row['mo']})").font = Font(bold=True); r += 1
        ws.cell(r, 1, row["reasoning"]).font = Font(size=9); ws.row_dimensions[r].height = 40; r += 1
    r += 1

    ws.cell(r, 1, "E-WAY BILL FULL-YEAR RECONCILIATION").font = Font(bold=True, size=11, color="1F3864"); r += 1
    ws.cell(r, 1, f"{len(ewb_out_gaps)} genuine outward gap(s), {len(ewb_in_gaps)} genuine inward gap(s) "
                  f"for the FY -- see 'EWB Full-Year Reconciliation' tab for full detail."); r += 2

    ws.cell(r, 1, "WHERE TO GO NEXT").font = Font(bold=True, size=11, color="1F3864"); r += 1
    for line in [
        "1. 'Action Required' tab -- PARENT-level genuine rows only, sorted worst-first by severity then Rs.",
        "2. 'EWB Full-Year Reconciliation' tab -- the complete, invoice-level proof behind every e-way-bill finding.",
        "3. 'Reviewed Master Dashboard' tab -- every row, colour-coded, with Root ID/Level so duplicates are visible but never double-counted.",
        "4. 'EWB Pattern - Annual' and the other Annual Detail sheets -- full-year rollups of checks that recur monthly.",
        "5. All sheets BEFORE this point in the workbook are the tool's own original output, unmodified -- go there for the underlying numbers, formulas, and full detail.",
    ]:
        ws.cell(r, 1, line); r += 1
    for col, w in zip("ABC", [75, 16, 60]):
        ws.column_dimensions[col].width = w


def style_header_local(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.fill = PatternFill("solid", fgColor="1F2937")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.border = BORDER



def write_cn_dn_impact_sheet(ws, data):
    """NEW SHEET (per explicit instruction) -- 'CN-DN ITC Impact - Annual'. Complete
    invoice-level detail of every credit/debit note this taxpayer both received (inward,
    GSTR-2B B2B-CDNR) and issued (outward, GSTR-1 CDNR), each with a plain-language statement
    of the resulting ITC/liability impact and exactly who it falls on -- see
    gst_checks_flow.build_cn_dn_impact_data()'s own docstring for the direction-of-impact
    reasoning this sheet is built from."""
    company = data["company_name"]; gstin = data["self_gstin"]
    ws.cell(1, 1, "CREDIT NOTE / DEBIT NOTE — ITC IMPACT (ANNUAL)").font = TITLEF
    ws.cell(2, 1, f"GSTIN {gstin}  |  {company}").font = Font(size=9, italic=True)
    ws.cell(3, 1, "Every credit/debit note this taxpayer received (inward, from its own "
                  "suppliers) and issued (outward, to its own customers) this FY, with the "
                  "resulting ITC/liability impact stated plainly against each one. Only the "
                  "INWARD side is this taxpayer's own actionable ITC reversal/addition (ties "
                  "to GSTR-3B 4(B)(2) on each month's Comparison sheet); the OUTWARD side's ITC "
                  "consequence belongs to the counterparty -- documented here with exact "
                  "GSTIN/name/amount for reference, since this tool has no visibility into "
                  "whether the counterparty actually complied."
                  ).font = Font(size=9, italic=True)
    ws.cell(3, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[3].height = 45

    r = 5
    ws.cell(r, 1, "QUICK SUMMARY").font = Font(bold=True, size=11, color="1F3864"); r += 1
    _net = data["net_itc_reversal_required"]
    # A negative net figure means inward Debit Notes this FY exceed inward Credit Notes -- i.e.
    # the NET effect is additional ITC available, not a reversal. Both directions are stated
    # plainly rather than always saying "must reverse" regardless of sign.
    if _net >= 0:
        _net_text = (f"Net ITC {company} must REVERSE this FY due to inward credit/debit notes "
                    f"(inward CN tax Rs {data['inward_cn_tax']:,.2f} minus inward DN tax "
                    f"Rs {data['inward_dn_tax']:,.2f}): Rs {_net:,.2f}")
        _net_color = "9C0006" if _net > 0 else "006100"
    else:
        _net_text = (f"Inward Debit Notes this FY (Rs {data['inward_dn_tax']:,.2f} tax) EXCEED "
                    f"inward Credit Notes (Rs {data['inward_cn_tax']:,.2f} tax) -- net effect is "
                    f"ADDITIONAL ITC available to {company} of Rs {abs(_net):,.2f}, not a "
                    f"reversal. Ensure this is reflected in the ITC claimed.")
        _net_color = "006100"
    ws.cell(r, 1, _net_text).font = Font(bold=True, size=10, color=_net_color)
    r += 1
    ws.cell(r, 1, f"Total outward Credit Notes issued this FY: Rs {data['outward_cn_tax']:,.2f} tax "
                  f"(informational -- recipients' ITC reversal obligation; {company}'s own "
                  f"liability is already reduced in its own GSTR-1/3B).").font = Font(size=10)
    r += 1
    ws.cell(r, 1, f"Total outward Debit Notes issued this FY: Rs {data['outward_dn_tax']:,.2f} tax "
                  f"(informational -- recipients' additional ITC eligibility; {company}'s own "
                  f"liability is already increased in its own GSTR-1/3B).").font = Font(size=10)
    r += 2

    hdr = ["Month", "Counterparty GSTIN", "Counterparty Trade Name", "Note No.", "Note Date",
           "Taxable Value", "IGST", "CGST", "SGST", "CESS", "Total Tax",
           "ITC Availability (2B)", "Action / Impact"]
    widths = [10, 18, 30, 18, 12, 16, 13, 13, 13, 11, 14, 16, 90]

    def _write_table(title_text, rows, r, include_itc_col):
        ws.cell(r, 1, title_text).font = Font(bold=True, size=11, color="1F3864"); r += 1
        cols = hdr if include_itc_col else [h for h in hdr if h != "ITC Availability (2B)"]
        for i, h in enumerate(cols, 1):
            ws.cell(r, i, h)
        _style_header(ws, r, len(cols))
        r += 1
        hdr_row = r - 1
        rows_sorted = sorted(rows, key=lambda x: -abs(x["tax_total"]))
        if not rows_sorted:
            ws.cell(r, 1, "None this FY.").font = Font(italic=True, color="808080")
            r += 1
        for idx, x in enumerate(rows_sorted):
            vals = [x["month"], x["gstin"], x["name"], x["note"], x["date"],
                    x["taxable"], x["igst"], x["cgst"], x["sgst"], x["cess"], x["tax_total"]]
            if include_itc_col:
                vals.append(x["itc_avail"])
            vals.append(x["action"])
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(r, ci, v)
                cell.border = BORDER
                cell.font = Font(size=10)
                cell.alignment = Alignment(wrap_text=(ci == len(vals)), vertical="top")
                if isinstance(v, float):
                    cell.number_format = "#,##0.00"
                if idx % 2 == 1:
                    cell.fill = PatternFill("solid", fgColor="F2F2F2")
            r += 1
        total_tax = sum(x["tax_total"] for x in rows_sorted)
        if rows_sorted:
            ws.cell(r, 4, "TOTAL").font = Font(bold=True)
            tc = ws.cell(r, 11, round(total_tax, 2)); tc.font = Font(bold=True); tc.number_format = "#,##0.00"
            for c in range(1, len(cols) + 2):
                ws.cell(r, c).border = BORDER
            r += 1
        ws.auto_filter.ref = None  # only one autofilter allowed per sheet; skip on multi-table sheets
        return r + 1, hdr_row

    inward_cn = [x for x in data["inward"] if x["is_credit"]]
    inward_dn = [x for x in data["inward"] if not x["is_credit"]]
    outward_cn = [x for x in data["outward"] if x["is_credit"]]
    outward_dn = [x for x in data["outward"] if not x["is_credit"]]

    r, _ = _write_table(f"Inward Credit Notes Received ({len(inward_cn)}) -- {company} must reverse ITC", inward_cn, r, True)
    r, _ = _write_table(f"Inward Debit Notes Received ({len(inward_dn)}) -- {company} gains additional ITC", inward_dn, r, True)
    r, _ = _write_table(f"Outward Credit Notes Issued ({len(outward_cn)}) -- recipient must reverse ITC", outward_cn, r, False)
    r, _ = _write_table(f"Outward Debit Notes Issued ({len(outward_dn)}) -- recipient gains additional ITC", outward_dn, r, False)

    for col, w in zip("ABCDEFGHIJKLM", widths):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"


def write_rectification_sheet(ws, pairs, drc_payments):
    # REWORKED FOR READABILITY (bug report #5: "sheet feels very complex, make it easier to
    # read, keep all the data intact"). Nothing below removes a single data point that was
    # there before -- every pair, every column value, the FY summary, the further-analysis
    # notes and the DRC payments block are all still present. What changed:
    #   1. A "Quick Summary" block up top -- counts by kind, and the 8 biggest corrections by
    #      tax impact -- so the reader sees what matters most before wading into the full table.
    #   2. The single 16-column table (which mixed two different kinds of row, most of them
    #      showing blank cells for columns that only apply to the OTHER kind) is now split into
    #      two separately-headed tables: "B2B Invoice Amendments" (which has a matched original
    #      to diff against) and "Credit/Debit Note Amendments" (which doesn't -- so it drops the
    #      five always-blank Original/Delta columns instead of showing them empty on every row).
    #      The CDNR section's static "(note amendments not month-matched...)" explanation, which
    #      used to repeat on every single row, is now stated once in that section's own heading.
    #   3. Each table is sorted by tax-impact (largest |delta| first) within its kind, so the
    #      corrections most worth a look surface at the top rather than being buried among many
    #      small ones in filing-chronology order -- the "Amended In" column still shows exactly
    #      when each one happened, so no chronological information is lost, only the row order.
    #   4. Banded row shading and a frozen header row, matching this tool's existing visual style
    #      elsewhere, so a long table stays readable while scrolling.
    ws.cell(1, 1, "CROSS-MONTH RECTIFICATION PAIRS -- GSTR-1 B2B AMENDMENTS (YEARLY)").font = TITLEF
    ws.cell(2, 1, "Every GSTR-1 amendment-sheet row (b2ba/cdnra) found in any supplied month, "
                  "linked back to the month that first reported the original document -- so an "
                  "error and its later correction both show up, together. Split into two tables "
                  "below by kind, each sorted by tax-impact (largest first); a Quick Summary up "
                  "top surfaces the biggest corrections without needing to scan every row."
                  ).font = Font(size=9, italic=True)
    c2 = ws.cell(2, 1); c2.alignment = Alignment(wrap_text=True, vertical="top")

    inv_pairs = [p for p in pairs if p["kind"] == "B2B Invoice Amendment"]
    note_pairs = [p for p in pairs if p["kind"] == "Credit/Debit Note Amendment"]
    n_not_found = sum(1 for p in inv_pairs if isinstance(p.get("original_month"), str) and "NOT FOUND" in p["original_month"])
    n_matched = len(inv_pairs) - n_not_found
    sum_delta_taxable = sum(p["delta_taxable"] for p in inv_pairs if p.get("delta_taxable") is not None)
    sum_delta_tax = sum((p.get("delta_tax") or 0.0) for p in inv_pairs if p.get("delta_taxable") is not None)

    r = 4
    ws.cell(r, 1, "QUICK SUMMARY").font = Font(bold=True, size=11, color="1F3864")
    r += 1
    ws.cell(r, 1, f"B2B Invoice Amendments: {len(inv_pairs)} total -- {n_matched} matched to an "
                  f"earlier month (net taxable-value change Rs {sum_delta_taxable:,.2f}, net tax "
                  f"change Rs {sum_delta_tax:,.2f}), {n_not_found} where the original could not be "
                  f"located (see amber rows below).").font = Font(size=10)
    r += 1
    ws.cell(r, 1, f"Credit/Debit Note Amendments: {len(note_pairs)} total (not month-matched to an "
                  f"original -- see that table's own heading below for why).").font = Font(size=10)
    r += 2

    top8 = sorted((p for p in inv_pairs if p.get("delta_tax") is not None),
                  key=lambda p: abs(p["delta_tax"]), reverse=True)[:8]
    if top8:
        ws.cell(r, 1, "Biggest corrections by tax impact (top 8, full detail in the table below)").font = Font(bold=True, size=10, color="1F3864")
        r += 1
        mini_hdr = ["GSTIN", "Original Ref", "Reported In", "Amended In", "Delta Taxable", "Delta Tax"]
        for i, h in enumerate(mini_hdr, 1):
            ws.cell(r, i, h)
        _style_header(ws, r, len(mini_hdr))
        r += 1
        for p in top8:
            vals = [p["gstin"], p["original_ref"], p["original_month"], p["amended_in_month"],
                    p.get("delta_taxable"), p.get("delta_tax")]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(r, ci, v)
                cell.border = BORDER
                cell.font = Font(size=10)
                if isinstance(v, float):
                    cell.number_format = "#,##0.00"
            if abs(p["delta_tax"]) >= 100000:
                ws.cell(r, 6).fill = RED
            elif abs(p["delta_tax"]) > 0.005:
                ws.cell(r, 6).fill = AMBER
            r += 1
        r += 2

    # ---- Table 1: B2B Invoice Amendments (has a matched-original diff) ----
    ws.cell(r, 1, f"B2B Invoice Amendments -- Complete Detail ({len(inv_pairs)} row(s)), sorted by "
                  f"tax impact").font = Font(bold=True, size=11, color="1F3864")
    r += 1
    hdr1 = ["GSTIN", "Recipient", "Original Ref", "Reported In (original month)",
            "Revised Ref", "Amended In (month)", "Original Taxable", "Original Tax (Total)",
            "Revised Taxable", "Revised Tax (Total)", "Delta Taxable (Revised - Original)",
            "Delta Tax (Revised - Original)", "IGST (revised)", "CGST (revised)", "SGST (revised)"]
    for i, h in enumerate(hdr1, 1):
        ws.cell(r, i, h)
    _style_header(ws, r, len(hdr1))
    r += 1
    table1_hdr_row = r - 1
    inv_sorted = sorted(inv_pairs, key=lambda p: abs(p.get("delta_tax") or -1), reverse=True)
    for idx, p in enumerate(inv_sorted):
        vals = [p["gstin"], p["recipient"], p["original_ref"], p["original_month"],
                p["revised_ref"], p["amended_in_month"], p.get("original_taxable"), p.get("original_tax"),
                p.get("revised_taxable"), p.get("revised_tax"), p.get("delta_taxable"), p.get("delta_tax"),
                p["igst"], p["cgst"], p["sgst"]]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(r, ci, v)
            cell.border = BORDER
            cell.font = Font(size=10)
            if isinstance(v, float):
                cell.number_format = "#,##0.00"
            if idx % 2 == 1:
                cell.fill = PatternFill("solid", fgColor="F2F2F2")
            if ci == 4 and isinstance(v, str) and "NOT FOUND" in v:
                cell.fill = AMBER
            if ci in (11, 12) and isinstance(v, (int, float)) and abs(v) > 0.005:
                cell.fill = RED if abs(v) >= 100000 else AMBER
        r += 1
    if inv_pairs:
        ws.cell(r, 3, "TOTAL (amendments with a matched original only)").font = Font(bold=True)
        v11 = ws.cell(r, 11, round(sum_delta_taxable, 2)); v11.font = Font(bold=True); v11.number_format = "#,##0.00"
        v12 = ws.cell(r, 12, round(sum_delta_tax, 2)); v12.font = Font(bold=True); v12.number_format = "#,##0.00"
        for c in range(1, len(hdr1) + 1):
            ws.cell(r, c).border = BORDER
        r += 1
    else:
        ws.cell(r, 1, "No invoice-amendment rows found in the month(s) supplied so far -- this "
                      "table fills in automatically as more months are added.").font = Font(italic=True, color="808080")
        r += 1
    ws.auto_filter.ref = f"A{table1_hdr_row}:{get_column_letter(len(hdr1))}{max(r-1, table1_hdr_row)}"
    r += 2

    # ---- Table 2: Credit/Debit Note Amendments (no original-month match attempted -- see note) ----
    ws.cell(r, 1, f"Credit/Debit Note Amendments -- Complete Detail ({len(note_pairs)} row(s)), "
                  f"sorted by tax impact. These are not matched back to an original month: this "
                  f"tool's own GSTR-1 credit-note sheet carries no original-invoice-number field, "
                  f"so any CN-to-original link is an approximation by GSTIN+value at best -- left "
                  f"unmatched here rather than guessed (see 'Original Ref' for the note number "
                  f"itself).").font = Font(bold=True, size=11, color="1F3864")
    ws.cell(r, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 30
    r += 1
    hdr2b = ["GSTIN", "Original Ref (note no.)", "Revised Ref", "Amended In (month)",
             "Revised Taxable", "Revised Tax (Total)", "IGST (revised)", "CGST (revised)", "SGST (revised)"]
    for i, h in enumerate(hdr2b, 1):
        ws.cell(r, i, h)
    _style_header(ws, r, len(hdr2b))
    r += 1
    table2_hdr_row = r - 1
    note_sorted = sorted(note_pairs, key=lambda p: abs(p.get("revised_tax") or 0), reverse=True)
    for idx, p in enumerate(note_sorted):
        vals = [p["gstin"], p["original_ref"], p["revised_ref"], p["amended_in_month"],
                p.get("revised_taxable"), p.get("revised_tax"), p["igst"], p["cgst"], p["sgst"]]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(r, ci, v)
            cell.border = BORDER
            cell.font = Font(size=10)
            if isinstance(v, float):
                cell.number_format = "#,##0.00"
            if idx % 2 == 1:
                cell.fill = PatternFill("solid", fgColor="F2F2F2")
        r += 1
    if not note_pairs:
        ws.cell(r, 1, "No note-amendment rows found in the month(s) supplied so far -- this table "
                      "fills in automatically as more months are added.").font = Font(italic=True, color="808080")
        r += 1
    r += 1
    ws.cell(r, 1, f"Summary: {len(pairs)} amendment pair(s) linked this FY ({len(inv_pairs)} invoice, "
                  f"{len(note_pairs)} note); net taxable-value change across matched invoice "
                  f"amendments: Rs {sum_delta_taxable:,.2f}, net tax change: Rs {sum_delta_tax:,.2f}"
                  + (f". {n_not_found} invoice pair(s) where the ORIGINAL document's month could not "
                     f"be located (amber above) -- the original may be in a prior FY not supplied to "
                     f"this run, or the amendment references a document this tool never saw." if n_not_found else ".")
                  ).font = Font(bold=True, size=10, color="1F3864")
    r += 2
    ws.cell(r, 1, "SUGGESTED FURTHER ANALYSIS from raw data already in this tool's inputs, not yet "
                  "built:").font = Font(bold=True, size=10, color="1F3864")
    r += 1
    for s in ["Time-gap distribution between original filing and amendment (currently shown as "
              "two month labels only) -- a same-quarter correction reads very differently from a "
              "correction filed just before the annual return deadline; the actual date fields "
              "this tool already parses per amendment row would support this without new input.",
              "Cross-check against the 'GSTR-2A Amendments' sheet's own B2BA/CDNRA linkage -- that "
              "sheet does the equivalent match on the INWARD (purchase) side; a taxpayer amending "
              "heavily on the OUTWARD side while their suppliers rarely amend (or vice versa) is a "
              "pattern comparison this tool computes both halves of already but doesn't yet join."]:
        ws.cell(r, 1, "- " + s).font = Font(size=9, italic=True)
        ws.cell(r, 1).alignment = Alignment(wrap_text=True)
        r += 1
    r += 1
    ws.cell(r, 1, "DRC PAYMENTS -- informational reference only (no tax-period field in source; "
                  "match by nearby date, not proven)").font = Font(bold=True, size=11, color="1F3864")
    r += 1
    hdr2 = ["Source ID", "Description", "Date", "Method", "Total (Lakh)"]
    for i, h in enumerate(hdr2, 1):
        ws.cell(r, i, h)
    _style_header(ws, r, 5)
    r += 1
    for d in sorted(drc_payments, key=lambda x: x["date"]):
        vals = [d["source_id"], d["description"], d["date"], d["method"], d["total"]]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(r, ci, v)
            cell.border = BORDER
            cell.font = Font(size=10)
        r += 1
    for col, w in zip("ABCDEFGHIJKLMNO", [18, 22, 18, 20, 18, 20, 15, 15, 15, 15, 15, 15, 13, 13, 13]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"


def write_doc_series(ws, month_results):
    ws.cell(1, 1, "DOCUMENT SERIES INTEGRITY (Table 13 vs actual B2B invoice numbers)").font = TITLEF
    ws.cell(2, 1, "'Found via fuzzy match' = same series, different prefix punctuation/segment than "
                  "Table 13's own printed range (confirmed real in this file's own JWI series, not a "
                  "code defect) -- shown separately so the difference stays visible. A missing count "
                  "that exactly equals Table 13's own declared 'Cancelled' figure is marked explained, "
                  "not left as an unexplained red flag. A range explained by B2CS means: this "
                  "taxpayer's B2C sales are reported only as a Table 7 state+rate aggregate, which "
                  "structurally carries no invoice number anywhere in a GSTR-1 export -- so a range "
                  "that is the ONLY unexplained one that month, with non-zero B2CS value that month, "
                  "is tied to that B2CS figure rather than left as an open question. 'Extra Serials' / "
                  "excess status: Table 13's own sr_from-sr_to range width and its separately-declared "
                  "'Total' field are two independently-reported numbers in the source export and can "
                  "legitimately disagree (confirmed real, Jun-25) -- this is never auto-explained the "
                  "way missing serials can be, since Table 13 carries no equivalent 'expected excess' "
                  "field to tie it out against.").font = Font(size=9, italic=True)
    hdr = ["Month", "Series Range", "Table-13 Total", "Table-13 Cancelled", "Actually Found",
           "Missing Serials", "Extra Serials", "Found via fuzzy match (diff. prefix format)", "Status"]
    r = 4
    for i, h in enumerate(hdr, 1):
        ws.cell(r, i, h)
    _style_header(ws, r, len(hdr))
    r += 1
    for res in month_results:
        for dg in res["doc_gap"]:
            missing = dg.get("missing", [])
            still_unexplained = dg.get("still_unexplained", missing)
            explained_decl = dg.get("explained_by_declared_cancellation", False)
            explained_einv = dg.get("explained_by_cancelled_einvoice", [])
            explained_b2cs = dg.get("explained_by_b2cs", False)
            extra_serials = dg.get("extra_serials", [])
            count_excess = dg.get("count_excess", 0) or 0
            excess_signal = dg.get("excess_signal", False)
            excess_severity = dg.get("excess_severity")
            if dg.get("note"):
                status = "CHECK MANUALLY"
            elif not missing:
                status = "OK"
            elif not still_unexplained and explained_b2cs:
                status = (f"EXPLAINED BY B2CS AGGREGATE (taxable {dg.get('b2cs_taxable', 0):,.2f}, "
                          f"tax {dg.get('b2cs_tax', 0):,.2f})")
            elif not still_unexplained and explained_einv and not explained_decl:
                status = "EXPLAINED BY CANCELLED E-INVOICE"
            elif not still_unexplained and explained_decl:
                status = "EXPLAINED BY DECLARED CANCELLATION"
            elif not still_unexplained:
                status = "EXPLAINED (CANCELLATION + CANCELLED E-INVOICE)"
            else:
                status = "UNEXPLAINED -- REVIEW"
            # NEW (Issue 4 fix): excess-serials never silently downgrades an already-open missing-
            # serials status, and never resolves to plain "OK" on its own -- appended to whatever
            # the missing-serials verdict above produced.
            if excess_signal:
                excess_text = (f"{excess_severity} -- EXTRA INVOICE(S) BEYOND DECLARED RANGE "
                               f"(found {dg.get('actual_count', '?')}, Table-13 declared "
                               f"{dg.get('table13_total', '?')}"
                               + (f", {len(extra_serials)} outside the numeric range boundary" if extra_serials else "")
                               + ")")
                status = excess_text if status == "OK" else f"{status}; {excess_text}"
            missing_display = ", ".join(still_unexplained) if still_unexplained else (
                ", ".join(missing) if missing else "-- none --")
            extra_display = ", ".join(extra_serials) if extra_serials else (
                f"(range/Total mismatch, +{int(count_excess)}, no specific serial attributable)"
                if count_excess > 0 else "-- none --")
            fuzzy_note = ", ".join(dg.get("found_via_fuzzy_match", [])) or "--"
            if explained_einv:
                fuzzy_note += f"  |  cancelled e-invoice: {', '.join(explained_einv)}"
            vals = [res["month"], dg.get("range"), dg.get("table13_total"), dg.get("table13_cancelled"),
                    dg.get("actual_count"), missing_display or dg.get("note", "-- none --"),
                    extra_display, fuzzy_note, status]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(r, ci, v)
                cell.border = BORDER
                cell.font = Font(size=10)
                cell.alignment = Alignment(wrap_text=(ci in (6, 7, 9)))
            if "UNEXPLAINED -- REVIEW" in status or status == "CHECK MANUALLY" or "FLAG --" in status:
                ws.cell(r, 6).fill = RED
                ws.cell(r, 9).fill = RED
            elif excess_signal:
                ws.cell(r, 7).fill = AMBER
                ws.cell(r, 9).fill = AMBER
            elif any(tag in status for tag in ("EXPLAINED BY DECLARED CANCELLATION", "EXPLAINED BY CANCELLED E-INVOICE",
                            "EXPLAINED (CANCELLATION + CANCELLED E-INVOICE)")):
                ws.cell(r, 9).fill = AMBER
            else:
                ws.cell(r, 9).fill = GREEN
            r += 1
    for col, w in zip("ABCDEFGHI", [10, 24, 12, 14, 12, 34, 30, 32, 30]):
        ws.column_dimensions[col].width = w
    ws.auto_filter.ref = f"A4:I{r-1}"


def _excel_safe(v):
    """openpyxl only accepts scalar types (str/int/float/bool/date/None) in a cell -- some
    checks' f.numbers dicts legitimately carry a LIST value (e.g. check A4's 'rates' -- every
    distinct tax rate an HSN code appeared at in one month, e.g. [0.0, 5.0]). Writing that raw
    into ws.cell() raises ValueError: Cannot convert [...] to Excel. This coerces any
    non-scalar value to a readable string instead of crashing the whole run; scalars pass
    through unchanged so their normal number formatting still applies."""
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    if isinstance(v, (list, tuple, set, frozenset)):
        return ", ".join(str(x) for x in v)
    return str(v)


def write_hsn_fraud_checks(ws, findings):
    ws.cell(1, 1, "HSN-CODE-WISE + FRAUD-PATTERN CHECKS (FY-WIDE)").font = TITLEF
    ws.cell(2, 1, "Categories A (HSN-only) / B (POS-state-code) / C (combined) plus the numbered "
                  "fraud-pattern list. Every check is grounded against this taxpayer's real files -- "
                  "see each row's detail for the exact arithmetic, and the module docstring in "
                  "hsn_fraud_checks.py for what could NOT be checked and why.").font = Font(size=9, italic=True)
    nflag = sum(1 for f in findings if f.severity == "FLAG")
    nrev = sum(1 for f in findings if f.severity == "REVIEW")
    ws.cell(3, 1, f"FLAG: {nflag}   REVIEW: {nrev}   (INFO/PASS rows included below for "
                  f"completeness/audit-trail).").font = Font(bold=True, size=11, color="C00000")
    SEV_RANK = {"FLAG": 0, "REVIEW": 1, "INFO": 2, "PASS": 3}
    ordered = sorted(findings, key=lambda f: (SEV_RANK.get(f.severity, 9), f.ref))
    hdr = ["Ref", "Check", "Result", "Key numbers", "Detail"]
    r = 5
    for i, h in enumerate(hdr, 1):
        ws.cell(r, i, h)
    _style_header(ws, r, len(hdr))
    r += 1
    for f in ordered:
        ws.cell(r, 1, f.ref)
        ws.cell(r, 2, f.title)
        cv = ws.cell(r, 3, f.severity)
        cv.fill = SEV_FILL.get(f.severity, GREY)
        cv.font = SEV_FONT.get(f.severity, Font(size=10))
        cv.alignment = Alignment(horizontal="center")
        # NEW: this 'Key numbers' column previously existed on every OTHER findings sheet
        # (Analysis14, Forensic Checks) but was silently never rendered here even though most
        # checks in this module already populate f.numbers -- fixed, not a behaviour change to
        # the checks themselves. List-valued entries (e.g. A4's 'rates') are joined into a
        # readable string here via _excel_safe rather than crashing the f-string format spec.
        ws.cell(r, 4, "  ".join(f"{k}={v:,.2f}" if isinstance(v, (int, float)) else f"{k}={_excel_safe(v)}"
                                for k, v in (f.numbers or {}).items()))
        ws.cell(r, 5, f.detail)
        for c in range(1, 6):
            cell = ws.cell(r, c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(c in (2, 4, 5)))
            if c != 3:
                cell.font = Font(size=10)
        r += 1
    for col, w in zip("ABCDE", [10, 40, 11, 34, 100]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:E{max(r-1, 5)}"

    # ---- COMPLETE DETAILS -- every FLAG/REVIEW check, grouped by check ref, as its own table ----
    # Per instruction: wherever a point in this sheet is FLAG or REVIEW, the complete underlying
    # detail (invoice no., dates, GSTIN, amounts, and -- for the IRN-timing checks -- the matching
    # e-way bill's own detail) should be visible here as a table, not just in the compact 'Key
    # numbers' cell above. Built GENERICALLY off f.numbers (every check that populates it gets a
    # table here automatically) rather than hand-listing individual check refs, so this does not
    # go stale as checks are added/changed -- one table per check-ref, columns = that check's own
    # numbers-dict keys (first-seen order), one row per flagged/reviewed occurrence.
    flagged = [f for f in findings if f.severity in ("FLAG", "REVIEW") and (f.numbers or {})]
    if flagged:
        groups = {}
        for f in flagged:
            groups.setdefault((f.ref, f.title), []).append(f)
        r += 2
        ws.cell(r, 1, "COMPLETE DETAILS -- FLAGGED / REVIEW ITEMS (underlying source-file rows, "
                      "grouped by check)").font = Font(bold=True, size=12, color="1F3864")
        r += 1
        ws.cell(r, 1, "One table per check below -- every invoice/period behind a FLAG or REVIEW "
                      "above, with the full detail from GSTR-1/E-Invoice/EWB (where matched) so you "
                      "can verify without reopening the source files.").font = Font(size=9, italic=True)
        r += 2
        max_cols = 5
        group_list = sorted(groups.items(), key=lambda kv: (SEV_RANK.get(kv[1][0].severity, 9), kv[0][0]))
        for (ref, title), items in group_list:
            sev_here = items[0].severity
            ws.cell(r, 1, f"{ref}  {title}  [{sev_here}]  ({len(items)} item(s))").font = \
                Font(bold=True, color="1F3864")
            r += 1
            # union of keys across this group's findings, first-seen order (some months/rows may
            # be missing a key that others have -- e.g. EWB fields when no EWB matched)
            cols = []
            for f in items:
                for k in f.numbers.keys():
                    if k not in cols:
                        cols.append(k)
            max_cols = max(max_cols, len(cols))
            for j, k in enumerate(cols, 1):
                c = ws.cell(r, j, k); c.font = Font(bold=True, size=9, color="FFFFFF"); c.fill = HEAD; c.border = BORDER
            r += 1
            for f in items:
                for j, k in enumerate(cols, 1):
                    v = _excel_safe(f.numbers.get(k, ""))
                    if isinstance(v, str) and v[:1] in ("=", "+", "-", "@"):
                        v = " " + v
                    c = ws.cell(r, j, v); c.border = BORDER; c.font = Font(size=10)
                    if isinstance(v, (int, float)):
                        c.number_format = "#,##0.00"
                r += 1
            r += 1
        for i in range(1, max_cols + 1):
            col = get_column_letter(i)
            cur = ws.column_dimensions[col].width or 0
            want = 20 if i == 1 else 16
            if want > cur:
                ws.column_dimensions[col].width = want


def _safe_parse_ledger(path, kind):
    """Wraps annual_sources.parse_cash_or_liability_ledger -- previously an
    absent ledger CSV crashed the ENTIRE run (unguarded open(None)). Now
    returns the same empty shape the real parser would return for a file
    with a header but zero transactions, so every downstream consumer
    (build_monthly_rows, write_cover, etc.) keeps working unchanged."""
    if not path or not os.path.exists(path):
        print(f"[info] {kind.title()} ledger not supplied -> annual ledger checks show N/A for this head.")
        return dict(opening=None, transactions=[], monthly_by_tax_period={}, monthly_by_txn_date={})
    return parse_cash_or_liability_ledger(path, kind)


def _safe_parse_credit(path):
    if not path or not os.path.exists(path):
        print("[info] Credit ledger not supplied -> annual credit-ledger checks show N/A.")
        return dict(opening=None, transactions=[], monthly_by_tax_period={})
    return parse_credit_ledger(path)


def _safe_parse_portal(path):
    if not path or not os.path.exists(path):
        print("[info] Portal Tax-Liability-&-ITC-Comparison report not supplied -> that comparison shows N/A.")
        return {}
    return parse_portal_comparison(path)


def _safe_parse_liab_demand(path):
    if not path or not os.path.exists(path):
        print("[info] Electronic Liability Ledger (Part II, demand/DRC) not supplied -> "
              "the ID-matched DRC cross-check (F9c) shows N/A.")
        return dict(opening=None, transactions=[], monthly_by_tax_period={}, monthly_by_txn_date={})
    return parse_cash_or_liability_ledger(path, "liability_demand")


def _safe_parse_r2a(path):
    """GSTR-2A is a genuinely optional whole-FY source, same treatment as portal-comparison/
    BO-Profile above -- degrades to the SAME empty shape dept.parse_r2a_excel() itself returns
    for a missing path, so every G-series check downstream sees available=False and SKIPs
    explicitly rather than crashing or reading a stale/wrong dict shape."""
    if not path or not os.path.exists(path):
        print("[info] GSTR-2A not supplied -> GSTR-2A cross-checks (G1-G10) show SKIPPED.")
    return arp.parse_r2a_excel(path)


_EMPTY_BO_PROFILE = dict(
    self_gstin=None, legal_name=None, trade_name=None, demographic={},
    financial_by_fy={}, bifa_by_fy={}, itc_passed_by_fy={}, itc_received_by_fy={},
    ewb_by_fy={}, einv_by_fy={}, refund_by_fy={},
    top_beneficiaries=[], top_suppliers=[], related_itc_received=[], related_itc_passed=[],
    drc_payments=[], appeals=[], cases=[], transfers=[],
)


def _safe_parse_bo(path):
    """BO/360-degree Profile is a rich but genuinely optional source (not
    every taxpayer/consultant has portal access to pull it). Previously
    parse_bo_profile(None) crashed the entire run; now degrades to an empty
    profile with every key the rest of the codebase expects, matching real
    output shape exactly (see bo_profile_parser.parse_bo_profile's own
    return statement)."""
    if not path or not os.path.exists(path):
        print("[info] BO/360-degree Profile not supplied -> DRC/related-party/BIFA cross-checks show N/A.")
        return dict(_EMPTY_BO_PROFILE)
    return parse_bo_profile(path)


# ======================================================================
# NEW SHEETS: Filing Compliance & Late Fee, Forensic Checks (R13/R14/BS-PL),
# Cancelled E-Invoices
# ======================================================================
def write_filing_compliance(ws, compliance_by_month):
    ws.cell(1, 1, "FILING COMPLIANCE -- ARN DATES, DUE DATES, LATE FEE & INTEREST").font = TITLEF
    ws.cell(2, 1, "Late fee per Section 47 (Rs 50/day normal, Rs 20/day nil, capped per Notification "
                  "07/2023-CT). Interest per Section 50(1), 18% p.a. on cash-paid tax, approximated using "
                  "the filing date as the payment date -- verify against the Liability Register's own "
                  "Interest-head entry (Forensic Framework Part 1, A4).").font = Font(size=9, italic=True)
    hdr = ["Month", "GSTR-1 ARN", "GSTR-1 Filed", "GSTR-1 Due", "GSTR-1 Late Fee (Rs)",
           "GSTR-3B ARN", "GSTR-3B Filed", "GSTR-3B Due", "GSTR-3B Late Fee (Rs)",
           "Sec 50 Interest (Rs)", "GSTR1-vs-3B gap (days)"]
    r = 4
    for i, h in enumerate(hdr, 1):
        ws.cell(r, i, h)
    _style_header(ws, r, len(hdr))
    r += 1
    if not compliance_by_month:
        ws.cell(r, 1, "No ARN date could be extracted -- see notes on the GSTR-1/GSTR-3B parsers; "
                      "check #10 in each month's Analysis(14) sheet also explains what's missing.").font = Font(italic=True)
    for c in (compliance_by_month or []):
        g1fee = (c.get("gstr1_late_fee") or {}).get("fee_payable")
        g3fee = (c.get("gstr3b_late_fee") or {}).get("fee_payable")
        interest = (c.get("gstr3b_interest") or {}).get("interest")
        vals = [c["month"], c.get("gstr1_arn"), str(c.get("gstr1_filing_date") or ""),
                str(c.get("gstr1_due_date") or ""), g1fee,
                c.get("gstr3b_arn"), str(c.get("gstr3b_filing_date") or ""),
                str(c.get("gstr3b_due_date") or ""), g3fee,
                interest, c.get("gstr1_vs_gstr3b_gap_days")]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(r, ci, v)
            cell.border = BORDER
            cell.font = Font(size=10)
            if isinstance(v, float):
                cell.number_format = "#,##0.00"
                if v > 0 and hdr[ci-1].endswith("(Rs)"):
                    cell.fill = AMBER
        r += 1
    for col, w in zip("ABCDEFGHIJK", [9, 18, 12, 12, 16, 18, 12, 12, 16, 14, 16]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:K{max(r-1, 4)}"


def write_forensic_checks(ws, findings):
    ws.cell(1, 1, "FORENSIC CHECKS -- Part 2 of GST_Forensic_Comparison_Framework_v1.md").font = TITLEF
    ws.cell(2, 1, "R13 (turnover-gap rule) and R14 (four-way ITC reconciliation), plus the generic "
                  "Balance-Sheet/P&L rule engine (R0-R12) when structured BS/PL data was supplied. "
                  "See OCR_LIMITATION.md for why a scanned BS/PL PDF is not auto-parsed here.").font = Font(size=9, italic=True)
    hdr = ["Ref", "Check", "Result", "Key numbers", "Detail"]
    r = 4
    for i, h in enumerate(hdr, 1):
        ws.cell(r, i, h)
    _style_header(ws, r, len(hdr))
    r += 1
    SEV_RANK = {"FLAG": 0, "REVIEW": 1, "INFO": 2, "PASS": 3, "SKIPPED": 4}
    for f in sorted(findings, key=lambda x: (SEV_RANK.get(x.severity, 9), x.ref)):
        ws.cell(r, 1, f.ref)
        ws.cell(r, 2, f.title)
        cv = ws.cell(r, 3, f.severity)
        cv.fill = SEV_FILL.get(f.severity, GREY); cv.font = SEV_FONT.get(f.severity, Font(size=10))
        cv.alignment = Alignment(horizontal="center")
        ws.cell(r, 4, "  ".join(f"{k}={v:,.2f}" if isinstance(v, (int, float)) else f"{k}={v}"
                                for k, v in (f.numbers or {}).items()))
        ws.cell(r, 5, f.detail)
        for c in range(1, 6):
            cell = ws.cell(r, c); cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(c in (2, 4, 5)))
            if c != 3:
                cell.font = Font(size=10)
        r += 1
    if not findings:
        ws.cell(r, 1, "No forensic findings produced this run.").font = Font(italic=True)
        r += 1
    r += 1
    nflag = sum(1 for f in findings if f.severity == "FLAG")
    nrev = sum(1 for f in findings if f.severity == "REVIEW")
    ninfo = sum(1 for f in findings if f.severity == "INFO")
    ws.cell(r, 1, f"Summary: {len(findings)} check(s) run -- {nflag} FLAG, {nrev} REVIEW, {ninfo} INFO "
                  f"(INFO here usually means a prerequisite source -- GSTR-9/9C or structured BS/PL "
                  f"data -- wasn't supplied for this taxpayer, not that the check ran clean; read the "
                  f"Detail column for each INFO row to see which).").font = Font(bold=True, size=10, color="1F3864")
    r += 2
    ws.cell(r, 1, "SUGGESTED FURTHER ANALYSIS from raw data already in this tool's inputs, not yet "
                  "built:").font = Font(bold=True, size=10, color="1F3864")
    r += 1
    for s in ["R13/R14 currently need GSTR-9/9C supplied as Excel (this tool does not parse a PDF "
              "GSTR-9/9C -- see the classify step) -- if a taxpayer's GSTR-9/9C is only available "
              "as PDF, these two checks will show INFO/not-computable regardless of how good the "
              "rest of the data is; worth flagging to whoever gathers inputs that Excel export "
              "specifically is needed for this pair of checks.",
              "The R0-R12 Balance-Sheet/P&L rule engine needs a hand-typed structured BS/PL input "
              "(bs_pl_input.py) tagged to the taxpayer's own GSTIN -- currently this file is "
              "tagged to a single reference taxpayer, so it silently no-ops for every other GSTIN. "
              "For a new taxpayer, someone needs to transcribe their real BS/PL into a fresh copy "
              "of that file (see its own docstring/OCR_LIMITATION.md for why this isn't "
              "auto-OCR'd) before R0-R12 will produce anything beyond an explicit skip."]:
        ws.cell(r, 1, "- " + s).font = Font(size=9, italic=True)
        ws.cell(r, 1).alignment = Alignment(wrap_text=True)
        r += 1
    for col, w in zip("ABCDE", [8, 40, 11, 34, 100]):
        ws.column_dimensions[col].width = w
    ws.auto_filter.ref = f"A4:E{max(r-1, 4)}"


def write_cancelled_einvoices(ws, all_cancelled, cross_check_findings, col_found):
    ws.cell(1, 1, "CANCELLED E-INVOICES").font = TITLEF
    note = ("The E-Invoice export's IRN-status column WAS found and read." if col_found else
            "The E-Invoice export does NOT appear to carry a recognisable IRN-status/cancellation "
            "column under any of the header names this tool knows (IRN Status, Status, Cancel Date, "
            "etc.) -- 'zero cancelled e-invoices' below may mean either 'genuinely none' or 'this "
            "export doesn't expose that field'; check the raw file if that distinction matters.")
    ws.cell(2, 1, note).font = Font(size=9, italic=True, color=("006100" if col_found else "9C6500"))
    r = 4
    for f in cross_check_findings:
        cv = ws.cell(r, 1, f"[{f.severity}] {f.ref} {f.title}")
        cv.fill = SEV_FILL.get(f.severity, GREY); cv.font = Font(bold=True, size=10)
        ws.cell(r, 2, f.detail).alignment = Alignment(wrap_text=True)
        r += 1
    r += 1
    hdr = ["Month", "Invoice No.", "Rate", "Taxable", "IGST", "CGST", "SGST", "IRN", "Cancel Date",
           "E-Way Bill Status (for this invoice)"]
    cancel_hdr_row = r
    for i, h in enumerate(hdr, 1):
        ws.cell(r, i, h)
    _style_header(ws, r, len(hdr))
    r += 1
    if not all_cancelled:
        ws.cell(r, 1, "No cancelled e-invoices found (see note above for what this does/doesn't confirm).").font = Font(italic=True)
    for c in all_cancelled:
        vals = [c.get("month"), c.get("invno"), c.get("rate"), c.get("taxable"), c.get("igst"),
                c.get("cgst"), c.get("sgst"), c.get("irn"), c.get("cancel_date"), c.get("ewb_status_note", "")]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(r, ci, v)
            cell.border = BORDER
            cell.font = Font(size=10)
            if isinstance(v, float):
                cell.number_format = "#,##0.00"
            if ci == 10 and isinstance(v, str) and "ACTIVE" in v.upper():
                cell.fill = RED; cell.font = Font(bold=True, size=10, color="9C0006")
        r += 1
    for col, w in zip("ABCDEFGHIJ", [9, 20, 8, 14, 12, 12, 12, 22, 14, 46]):
        ws.column_dimensions[col].width = w
    ws.auto_filter.ref = f"A{cancel_hdr_row}:J{max(r-1, cancel_hdr_row)}"



def write_itc_detailed_recon_table(ws, rows):
    """NEW TABLE (per explicit, detailed instruction) -- appended below the existing 'ITC
    Annual Summary' sheet's own content (same sheet, not a duplicate). See
    gst_checks_flow.build_itc_detailed_recon_data()'s own docstring for the full computation
    and every deliberate design choice (Yes-only filter, strict composite-key mismatch
    matching, inferred-not-parsed carry-forward, etc.)."""
    HEADS = ("IGST", "CGST", "SGST", "CESS", "Total")
    groups = [
        ("ITC as per 2B (Available = Yes only)", "b2b", 5),
        ("ITC as per 2A (all invoices)", "a2a", 5),
        ("Mismatch (invoice-level, strict key)", None, 4),
        ("Credit Note ITC Impact (reversal)", "cn", 5),
        ("Debit Note ITC Impact (additional claim)", "dn", 5),
        ("Net ITC Eligible (2B[Yes] + DN - CN)", "net", 5),
        ("Actual ITC Claimed (GSTR-3B Table 4A)", "claimed", 5),
        ("ITC Carried Fwd from Last FY (Inferred)", "cf", 5),
        ("Excess / Short Claim (Claimed - Net Eligible)", "exs", 5),
    ]
    r0 = ws.max_row + 3
    ws.cell(r0, 1, "ITC RECONCILIATION — TAX-HEAD-WISE, INVOICE-LEVEL (2A/2B/3B/CN-DN)").font = TITLEF
    ws.cell(r0 + 1, 1, "2A and 2B are never netted together: 2B(Yes) is the sole ITC-eligibility "
                       "baseline; 2A feeds ONLY the Mismatch columns (invoice-level, matched on "
                       "invoice no. + supplier GSTIN + invoice value + all four tax heads + invoice "
                       "date -- an invoice differing on any of these counts as unmatched, not a false "
                       "match on number+GSTIN alone). Months with no GSTR-3B supplied show Actual "
                       "Claimed / Excess-Short as blank, marked in Remarks, never a fabricated "
                       "discrepancy.").font = Font(size=9, italic=True)
    ws.cell(r0 + 1, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r0 + 1].height = 40
    r = r0 + 3

    # Two-row header: category (merged across its span) + tax-head sub-row.
    c = 2
    ws.cell(r, 1, "Month")
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=1)
    for title, key, span in groups:
        ws.cell(r, c, title)
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + span - 1)
        ws.cell(r, c).alignment = Alignment(horizontal="center", wrap_text=True)
        if key is None:  # Mismatch group's own 4 sub-columns
            for i, h in enumerate(["In 2A not 2B (Count)", "In 2A not 2B (Value)",
                                    "In 2B not 2A (Count)", "In 2B not 2A (Value)"]):
                ws.cell(r + 1, c + i, h)
        else:
            for i, h in enumerate(HEADS):
                ws.cell(r + 1, c + i, h)
        c += span
    ws.cell(r, c, "Remarks")
    ws.merge_cells(start_row=r, start_column=c, end_row=r + 1, end_column=c)
    last_col = c
    for row_ in ws[r:r + 1]:
        for cell in row_:
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill("solid", fgColor="1F3864")
            cell.border = BORDER
    r += 2
    data_start = r

    for row_data in rows:
        c = 1
        is_fy_total = row_data.get("is_fy_total", False)
        month_cell = ws.cell(r, c, row_data["month"])
        if is_fy_total:
            month_cell.font = Font(bold=True)
        c += 1
        for title, key, span in groups:
            if key is None:
                vals = [row_data["mism_a2a_not_b2b_ct"], row_data["mism_a2a_not_b2b_val"],
                        row_data["mism_b2b_not_a2a_ct"], row_data["mism_b2b_not_a2a_val"]]
            else:
                head_vals = row_data[key]
                total = round(sum(v for v in head_vals if v is not None), 2) if all(v is not None for v in head_vals) else None
                vals = list(head_vals) + [total]
            for v in vals:
                cell = ws.cell(r, c, v)
                cell.border = BORDER
                cell.font = Font(bold=is_fy_total, size=10)
                if isinstance(v, float):
                    cell.number_format = "#,##0.00"
                c += 1
        rem_cell = ws.cell(r, c, row_data["remark"])
        rem_cell.alignment = Alignment(wrap_text=True, vertical="top")
        rem_cell.border = BORDER
        rem_cell.font = Font(size=9, bold=is_fy_total)
        # Conditional RED FILL: any row where Excess/Short TOTAL > 0.
        exs = row_data.get("exs") or [None] * 4
        exs_total = sum(v for v in exs if v is not None) if all(v is not None for v in exs) else None
        if exs_total is not None and exs_total > 0:
            for col in range(1, last_col + 1):
                ws.cell(r, col).fill = RED
        r += 1

    for col_idx in range(1, last_col + 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = 12 if col_idx > 1 else 10
    ws.column_dimensions[get_column_letter(last_col)].width = 55
    ws.freeze_panes = ws.cell(data_start, 2).coordinate


def write_hsn_review_table(ws, hsn_rows, month_label):
    """Appends a 'HSN RATE REVIEW' section into the CURRENT worksheet
    (the same sheet write_comparison() just wrote into), below whatever
    write_comparison() left there. One row per HSN code actually used this
    month (taxable value, rate charged, tax amount -- exactly as reported
    in GSTR-1's own HSN summary table, i.e. already the per-HSN aggregate,
    no further summing needed), plus two reference-rate columns and an
    explicit 'verify' flag on every row -- per the person's own request:
    show the raw HSN-wise breakdown side by side with whatever reference
    data exists, and flag every one for manual verification rather than
    silently trusting either reference (matches this tool's own severity
    discipline throughout -- see HSN_RATE_HISTORY's and
    _load_mcp_india_stack_hsn_table's own docstrings for exactly why
    neither reference is treated as ground truth on its own)."""
    on_date = hfc._month_label_to_date(month_label)
    mcp_table = hfc._load_mcp_india_stack_hsn_table()

    r = ws.max_row + 3
    ws.cell(r, 1, f"HSN RATE REVIEW -- TAXABLE SUPPLY BY HSN CODE ({month_label})").font = TITLEF
    r += 1
    ws.cell(r, 1, "Every HSN code used this month, with the rate/taxable/tax as actually reported "
                  "in GSTR-1's HSN summary, alongside two independent reference rates where "
                  "available. EVERY row is flagged VERIFY -- neither reference column is ground "
                  "truth on its own (see HSN & Fraud Pattern Checks sheet, checks A1/A1-EXT/A7, "
                  "for the full-strength automated comparison); this table is a fast side-by-side "
                  "worksheet for manual review, not a verdict.").font = Font(size=9, italic=True)
    r += 2
    hdr = ["HSN Code", "Description (taxpayer's own)", "Rate Charged (%)", "Taxable Value (Rs)",
           "Tax Amount (Rs)", "Curated Reference Rate (%)", "mcp-india-stack Reference Rate (%)", "Status"]
    for i, h in enumerate(hdr, 1):
        ws.cell(r, i, h)
    _style_header(ws, r, len(hdr))
    r += 1

    if not hsn_rows:
        ws.cell(r, 1, "No HSN summary rows for this month.").font = Font(italic=True)
        return

    for row in sorted(hsn_rows, key=lambda x: x["hsn"]):
        hsn, desc, rate = row["hsn"], row["desc"], row["rate"]
        tax_amt = row["igst"] + row["cgst"] + row["sgst"]

        curated = hfc._hsn_rate_for_date(hsn, on_date) if on_date else None
        curated_rate = curated["rate"] if curated else None
        curated_display = (curated_rate if curated_rate is not None else
                            ("unconfirmed for this period" if curated else "not in curated list"))

        mcp_prefix = hfc._hsn_prefix_lookup(hsn, mcp_table) if mcp_table else None
        mcp_rate = mcp_table[mcp_prefix][0] if mcp_prefix else None
        mcp_display = mcp_rate if mcp_rate is not None else "not found"

        mismatch = ((curated_rate is not None and abs(rate - curated_rate) > 0.01) or
                    (mcp_rate is not None and abs(rate - mcp_rate) > 0.01))
        status = "VERIFY -- reference rate differs" if mismatch else "VERIFY"

        vals = [hsn, desc, rate, row["taxable"], tax_amt, curated_display, mcp_display, status]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(r, ci, v)
            cell.border = BORDER
            cell.font = Font(size=10)
            if ci in (4, 5) and isinstance(v, float):
                cell.number_format = "#,##0.00"
            if ci == 8:
                cell.fill = RED if mismatch else AMBER
                cell.alignment = Alignment(horizontal="center")
        r += 1


def write_hsn_detail(ws, hsn_rows, month_label):
    """<Month> HSN Detail (NEW) -- raw Table-12 HSN rows used by this month's HSN-vs-named-
    invoice comparisons (Comparison 'A'/'A2', Analysis(14) '#13'), so their Source Reference can
    point HERE -- inside this same master output workbook -- instead of back out to the raw
    GSTR1_Merged.xlsx file. Mirrors the existing '<Month> EWB Detail' sheet's layout/pattern."""
    ws.cell(1, 1, f"HSN DETAIL -- Table 12 rows backing this month's HSN-vs-named-invoice checks "
                  f"({month_label})").font = TITLEF
    ws.cell(2, 1, "Backs Comparison 'A' (Outward taxable GROSS before CN), 'A2' (invoices-net vs "
                  "HSN), and Analysis(14) '#13' (HSN IGST vs named-invoice IGST). 'Source Tab' "
                  "shows which GSTR-1 sheet this row came from -- the old single 'hsn' tab, or "
                  "the newer 'hsn(b2b)'/'hsn(b2c)' pair GSTN split it into partway through this "
                  "FY -- both feed the same total, read correctly either way.").font = Font(size=9, italic=True)
    hdr = ["HSN Code", "Description", "UQC", "Qty", "Rate (%)", "Taxable Value", "Integrated Tax",
           "Central Tax", "State/UT Tax", "Cess", "Source Tab"]
    r = 4
    for i, h in enumerate(hdr, 1):
        ws.cell(r, i, h)
    _style_header(ws, r, len(hdr))
    r += 1
    if not hsn_rows:
        ws.cell(r, 1, "No HSN summary rows found for this month (neither 'hsn' nor 'hsn(b2b)'/"
                       "'hsn(b2c)' carried a period marker for it in the source file -- an "
                       "explicit data gap, not a verified zero).").font = Font(italic=True, color="C00000")
        return
    for row in sorted(hsn_rows, key=lambda x: x["hsn"]):
        vals = [row["hsn"], row["desc"], row.get("uqc", ""), row.get("qty", 0.0), row["rate"],
                row["taxable"], row["igst"], row["cgst"], row["sgst"], row["cess"],
                row.get("source_tab", "hsn")]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(r, ci, v)
            cell.border = BORDER
            cell.font = Font(size=10)
            if ci in (4, 6, 7, 8, 9, 10) and isinstance(v, (int, float)):
                cell.number_format = "#,##0.00"
        r += 1
    for col, w in zip("ABCDEFGHIJK", [12, 42, 8, 10, 10, 16, 14, 14, 14, 12, 14]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"


def main(folder="."):
    print("Classifying folder (merged files)...")
    res = classify_folder(folder)

    if res.get("classify_warnings"):
        print("Classification warnings:")
        for w in res["classify_warnings"]:
            print("  -", w)

    if not res["self_gstin"]:
        raise RuntimeError(
            "Could not determine self-GSTIN from the supplied files (no EWB annual "
            "workbooks and no merged GSTR-1 'Read me' sheet found). Stopping -- "
            "every output filename/header depends on this being correct."
        )
    if not res["gstr1_month_map"] or not res["gstr3b_month_map"]:
        raise RuntimeError(
            "No merged GSTR-1 and/or GSTR-3B workbook found in the folder -- these are the only "
            "two genuinely mandatory inputs. Everything else (E-Invoice, GSTR-2B, EWB, ledgers, "
            "portal comparison, BO Profile, GSTR-9/9C/8A, BS/PL) is optional and degrades "
            "gracefully if absent."
        )

    raw.SELF_GSTIN = res["self_gstin"]
    raw.COMPANY_NAME = res["company_name"]

    print("Parsing annual EWB (whole FY(s), filtered per month)...")
    # CHANGED (multi-year + graceful degradation): now pools EWB rows across
    # EVERY outward/inward annual workbook found (any number of FYs), and
    # separately tracks whether each DIRECTION was supplied AT ALL -- see
    # gst_eway_recon.run()'s docstring for why that distinction matters.
    ewb_out_rows = [r for f in res["ewb_out_files"] for r in ewbp.parse_annual_ewb(f)]
    ewb_in_rows = [r for f in res["ewb_in_files"] for r in ewbp.parse_annual_ewb(f)]
    ewb_out_file_supplied = bool(res["ewb_out_files"])
    ewb_in_file_supplied = bool(res["ewb_in_files"])

    # A month is "covered" only if BOTH GSTR-1 and GSTR-3B have a resolved
    # file for it (possibly from DIFFERENT merged workbooks if this is a
    # multi-year run) -- sourced from the month->file maps, no fixed-length
    # calendar assumed, so this naturally spans any number of FYs.
    months_covered = _sort_months_chronologically(
        set(res["gstr1_month_map"]) & set(res["gstr3b_month_map"]))
    # Gaps WITHIN the covered span (not "out of a fixed 12"): any calendar
    # month strictly between the earliest and latest covered month that is
    # NOT itself covered.
    months_gap = []
    if len(months_covered) >= 2:
        y0, m0 = _month_sort_key(months_covered[0])
        y1, m1 = _month_sort_key(months_covered[-1])
        covered_set = set(months_covered)
        y, m = y0, m0
        while (y, m) <= (y1, m1):
            lbl = f"{mpu.CAL_MONTH_ABBR[m]}-{str(y)[2:]}"
            if lbl not in covered_set:
                months_gap.append(lbl)
            m += 1
            if m == 13:
                m = 1; y += 1

    print(f"Months covered: {months_covered}")
    fys_covered = sorted(set(_fy_label_for_month(m) for m in months_covered))
    print(f"FY(s) covered: {fys_covered}")

    # ---- filing compliance: read ARN dates ONCE per unique file (not once per
    #      month -- these functions scan the whole workbook each time) ----
    print("Extracting ARN / filing dates for late-fee & interest computation...")
    gstr1_arn_by_month, gstr3b_arn_by_month = {}, {}
    for f in res["gstr1_files"]:
        m, warns = fc.gstr1_arn_dates_by_month(f)
        gstr1_arn_by_month.update({k: v for k, v in m.items() if k != "_readme_fallback"})
        for w in warns:
            print("  [filing_compliance]", w)
    for f in res["gstr3b_files"]:
        gstr3b_arn_by_month.update(fc.gstr3b_arn_dates_by_month(f))
    # QRMP detection: a GSTR-1 marker that fans one marker into 3 months = quarterly filer.
    # (Approximate signal: if GSTR-1 has fewer distinct ARNs than months, quarterly is likely --
    # exact detection needs the real marker text's own Tax-Period field, already used inside
    # gstr1_arn_dates_by_month(); left as monthly-default here since this taxpayer's GSTR-3B is
    # confirmed one-sheet-per-month = non-QRMP, and QRMP support is otherwise architecturally
    # ready in filing_compliance.py's due_date_gstr1()/due_date_gstr3b() is_qrmp parameter.)
    gstr1_is_qrmp = gstr3b_is_qrmp = False

    # ---- Bug 2 pre-pass: FY-wide 4(B)(2) reversal figures, needed for the month-over-month
    #      outlier check that REPLACES the 4(B)(2)-vs-2B-CN comparison for any month/head where
    #      this filer's CN-into-4A5-netting pattern is confirmed (see build_comparisons()'s D2
    #      section for why that comparison is structurally meaningless for those month/heads).
    #      Read once here, before the month loop, so every month's call can see every OTHER
    #      month's figure -- build_comparisons() only ever sees its own single PERIOD_LABEL.
    print("Reading FY-wide 4(B)(2) reversal figures (for the D2 month-over-month outlier check)...")
    fy_4b2_by_month = {}
    fy_rcm_by_month = {}   # month -> [3.1d IGST,CGST,SGST, 4A3 IGST] -- reused by Bug 5's RCM
                            # trailing-average anomaly escalation, same pre-pass, no extra parse.
    for m in months_covered:
        g3bpath = res["gstr3b_month_map"].get(m)
        if not g3bpath:
            continue
        try:
            g3b_m = raw.parse_gstr3b(g3bpath, m)
            fy_4b2_by_month[m] = g3b_m.get("4B2") or [0.0, 0.0, 0.0, 0.0]
            d31 = g3b_m.get("3.1d") or [0.0, 0.0, 0.0, 0.0]
            a43 = g3b_m.get("4A3") or [0.0, 0.0, 0.0, 0.0]
            fy_rcm_by_month[m] = [d31[1] if len(d31) > 1 else 0.0,
                                   d31[2] if len(d31) > 2 else 0.0,
                                   d31[3] if len(d31) > 3 else 0.0,
                                   a43[0] if len(a43) > 0 else 0.0]
        except Exception:
            continue
    raw.FY_4B2_BY_MONTH = fy_4b2_by_month
    raw.FY_RCM_BY_MONTH = fy_rcm_by_month

    # ---- Bug 3 pre-pass: FY-wide GSTR-2B invoice-level index, keyed by (supplier GSTIN,
    #      invoice/note number), spanning EVERY month -- not just the month an inward EWB
    #      happens to fall in. CONFIRMED against real data: several inward EWB movements have
    #      NO 2B match in their own month purely because the supplier filed LATE (their invoice
    #      shows up in a LATER month's 2B, sometimes an EARLIER one for other reasons) -- e.g. an
    #      EWB dated May with its supplier's invoice only appearing in June's 2B. Matching
    #      against the single month this EWB falls in was treating an ordinary filing-timing gap
    #      as a "supplier hasn't filed" flag. This index lets #12 tell the two apart: genuinely
    #      absent from 2B ALL YEAR (real gap) vs present somewhere else in the FY (timing, not a
    #      gap) -- never silently treated as a clean same-month match, always labelled which one
    #      it is.
    print("Reading FY-wide GSTR-2B invoice index (for the EWB-In late-filing timing check)...")
    fy_2b_index = {}
    for m in months_covered:
        twobpath_m = res["gstr2b_month_map"].get(m)
        if not twobpath_m:
            continue
        try:
            r2b = raw.parse_2b_excel(twobpath_m, m)
            for x in r2b.get("b2b", []):
                key = (str(x["gstin"]).strip().upper(), str(x["invno"]).strip().upper())
                fy_2b_index.setdefault(key, []).append((m, x))
            for c in r2b.get("cdnr", []):
                key = (str(c["gstin"]).strip().upper(), str(c["note"]).strip().upper())
                fy_2b_index.setdefault(key, []).append((m, c))
        except Exception:
            continue
    raw.FY_2B_INVOICE_INDEX = fy_2b_index

    month_results = []
    month_g1_lines = {}
    compliance_records = []
    all_cancelled_einvoices = []
    einv_cancel_col_found_any = False
    run_errors = []
    for m in months_covered:
        print(f"Running month {m}...")
        files = dict(gstr1=res["gstr1_month_map"].get(m), gstr3b=res["gstr3b_month_map"].get(m),
                     einv=res["einv_month_map"].get(m), gstr2b=res["gstr2b_month_map"].get(m))
        try:
            out = run_month(m, files, ewb_out_rows, ewb_in_rows, res["self_gstin"], res["company_name"],
                             ewb_out_file_supplied=ewb_out_file_supplied,
                             ewb_in_file_supplied=ewb_in_file_supplied,
                             gstr1_arn_by_month=gstr1_arn_by_month, gstr3b_arn_by_month=gstr3b_arn_by_month,
                             gstr1_is_qrmp=gstr1_is_qrmp, gstr3b_is_qrmp=gstr3b_is_qrmp)
        except Exception as ex:
            # ROBUSTNESS (new): one month's unexpected parsing issue no longer takes down the
            # entire multi-year run. Logged loudly, and the Dashboard will show this month as
            # having zero findings WITH an explicit note -- never silently dropped.
            print(f"  *** ERROR processing {m}: {ex} -- this month is SKIPPED, all other months continue. ***")
            run_errors.append((m, str(ex)))
            continue
        month_results.append(out)
        month_g1_lines[m] = out["comp_raw"]["g1"].get("lines", {})
        if out.get("compliance"):
            compliance_records.append(out["compliance"])
        if out.get("cancelled_einvoices"):
            all_cancelled_einvoices.extend(dict(c, month=m) for c in out["cancelled_einvoices"])
        einv_cancel_col_found_any = einv_cancel_col_found_any or out.get("einv_cancel_col_found", False)

    if run_errors:
        print(f"\n*** {len(run_errors)} month(s) failed and were skipped: {[m for m, _ in run_errors]} ***\n")

    rect_pairs = build_rectification_pairs(month_results, month_g1_lines, months_covered)

    print("Building Phase-1 annual reconciliation (graceful if any source is missing)...")
    annual_data = dict(
        cash=_safe_parse_ledger(res["cash_ledger"], "cash"),
        credit=_safe_parse_credit(res["credit_ledger"]),
        liab=_safe_parse_ledger(res["liab_ledger"], "liability"),
        liab_demand=_safe_parse_liab_demand(res.get("liab_demand_ledger")),
        comp=_safe_parse_portal(res["portal_comparison"]),
        bo=_safe_parse_bo(res["bo_profile"]),
    )
    annual_rows = annualwb.build_monthly_rows(annual_data, months=months_covered)
    # NOTE: flag_itc/flag_cash_ledger no longer exist as separate keys -- those two
    # comparisons were TPST-only (retired) and have no independent second source left
    # in this run's inputs; build_monthly_rows() now only produces flag_liab (Liability
    # Register vs Portal-Comparison) and flag_credit_ledger (Credit Ledger vs Portal-3B-ITC).
    annual_review_count = sum(1 for r in annual_rows
                               for k in ("flag_liab", "flag_credit_ledger")
                               if r[k] == "REVIEW")

    print("Parsing GSTR-2A (optional; graceful if absent)...")
    r2a_data = _safe_parse_r2a(res.get("r2a_merged"))

    print("Running HSN-code-wise + fraud-pattern checks...")
    files_for_hsn = dict(gstr1=res["gstr1_merged"], gstr3b=res["gstr3b_merged"],
                          einv=res["einv_merged"], gstr2b=res["gstr2b_merged"])
    hsn_findings = hfc.run_all(files_for_hsn, ewb_out_rows, ewb_in_rows, months_covered, annual_data,
                                annual_rows, res["self_gstin"],
                                hsn_sac_master_override=res.get("hsn_sac_master_file"))

    # ---- NEW: annual-return-side sources (GSTR-9, GSTR-9C, Table 8A) + forensic checks ----
    print("Parsing GSTR-9 / GSTR-9C / Table 8A (optional; graceful if absent)...")
    gstr9 = arp.parse_gstr9(res["gstr9_files"][0] if res["gstr9_files"] else None)
    gstr9c = arp.parse_gstr9c(res["gstr9c_files"][0] if res["gstr9c_files"] else None)
    table8a = arp.parse_table_8a(res["table8a_files"][0] if res["table8a_files"] else None)

    print("Running forensic checks (R13 turnover-gap, R14 four-way ITC)...")
    # R13 needs GSTR-1's own Table-8 (exemp) rows per month -- re-derive from what parse_gstr1
    # already read (nil_exempt_taxable is a SUM; R13 needs to know row-count, so re-read the
    # 'exemp' sheet's row presence per month directly here, reusing the same content-based logic).
    exemp_rows_by_month = {}
    for m, res_m in zip(months_covered, month_results):
        g1 = res_m["comp_raw"]["g1"]
        exemp_rows_by_month[m] = [1] * 0 if g1.get("nil_exempt_taxable") in (None,) else (
            [1] if (g1.get("nil_taxable", 0) or g1.get("exempt_taxable", 0) or g1.get("nongst_taxable", 0)) else [])
    r13 = fchk.check_turnover_gap(gstr9c, exemp_rows_by_month)

    gstr2b_fy_total = None  # left None unless a whole-FY GSTR-2B aggregate is separately supplied;
    # architecturally ready (check_four_way_itc accepts it) -- wiring a true FY-sum from the
    # per-month 2B summaries already computed in month_results is the natural next step once a
    # full year of GSTR-2B data is available to sum (each month's summary is already in
    # month_results[i]['comp_raw']['b2b'] when available=True).
    _2b_sums = {"igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0}
    _2b_any = False
    for res_m in month_results:
        b2b = res_m["comp_raw"]["b2b"]
        # Uses '_summary_available' (the narrower 'was the ITC Available/control-total sheet
        # itself readable' signal) -- these ITC_all_other_* fields come SPECIFICALLY from that
        # sheet's own Table 3 breakdown and are all 0.0 (not missing, genuinely zero-filled) when
        # that sheet isn't present, even though b2b/cdnr invoice-level data may be fully there.
        # The broader 'available' flag would wrongly let a taxpayer with no 'ITC Available' sheet
        # sum up a string of real zeros as if they were real data, understating gstr2b_fy_total.
        if b2b.get("_summary_available"):
            _2b_any = True
            _2b_sums["igst"] += b2b.get("ITC_all_other_IGST", 0) or 0
            _2b_sums["cgst"] += b2b.get("ITC_all_other_CGST", 0) or 0
            _2b_sums["sgst"] += b2b.get("ITC_all_other_SGST", 0) or 0
            _2b_sums["cess"] += b2b.get("ITC_all_other_CESS", 0) or 0
    if _2b_any:
        gstr2b_fy_total = _2b_sums

    annual_turnover = gstr9c.get("turnover_declared_gstr9") if gstr9c.get("available") else None
    r14 = fchk.check_four_way_itc(gstr9, gstr2b_fy_total, table8a, gstr9c, annual_turnover=annual_turnover)
    forensic_findings = [r13, r14]

    # ---- NEW: BS/P&L rule engine (R0-R12) -- FIX: this was built and tested standalone
    # (bs_pl_input.py's own __main__ block) but never actually wired into master_build.py's
    # pipeline, so it never appeared in the output workbook even when filled in. Tries to
    # import BS_PL_DATA from bs_pl_input.py in the SAME folder as the running script (not the
    # data folder -- this is a hand-typed structured input, not a content-detected file, per
    # OCR_LIMITATION.md); degrades to a single INFO finding if that file/variable is absent or
    # empty, exactly like every other optional source in this tool. ----
    try:
        import bs_pl_input as bspl_mod
        bs_pl_data = getattr(bspl_mod, "BS_PL_DATA", {})
    except ImportError:
        bs_pl_data = {}
    # SAFETY CHECK (new): refuse to use bs_pl_input.py's data if it's tagged for a DIFFERENT
    # GSTIN than the one actually being processed this run -- guards against a taxpayer's old
    # BS/PL figures silently being reused for a different taxpayer's tool run. A dict with no
    # '_gstin' tag at all is also refused (forces explicit tagging rather than an implicit
    # "assume it matches").
    if bs_pl_data:
        tagged_gstin = bs_pl_data.get("_gstin")
        if tagged_gstin != res["self_gstin"]:
            print(f"[warn] bs_pl_input.py's BS_PL_DATA is tagged for GSTIN {tagged_gstin!r}, "
                  f"but this run is processing {res['self_gstin']!r} -- REFUSING to use it "
                  "(prevents a stale/wrong-taxpayer's Balance Sheet figures being silently "
                  "applied). Update bs_pl_input.py's '_gstin' tag and figures for this taxpayer.")
            bs_pl_data = {}
    if bs_pl_data:
        print("Running BS/P&L rule engine (R0-R12) against bs_pl_input.BS_PL_DATA...")
        bo_for_bspl = annual_data.get("bo") if annual_data.get("bo", {}).get("drc_payments") else None
        forensic_findings += fchk.check_bs_pl_rules(bs_pl_data, gstr9c=gstr9c, bo_profile=bo_for_bspl)
    else:
        forensic_findings.append(fchk.Finding(
            "R0-R12", "Balance Sheet / P&L rule engine", "INFO",
            "bs_pl_input.py not found next to master_build.py, or its BS_PL_DATA dict is empty -- "
            "R0-R12 not run. Fill in bs_pl_input.py (see OCR_LIMITATION.md for why this is a "
            "hand-typed template, not auto-OCR'd) to enable.", {}))

    print("Running cancelled-e-invoice cross-checks...")
    # BUG FIX: a month key must be recorded whenever E-Invoice data was AVAILABLE for that
    # month, even if that month happened to have zero cancelled invoices -- not only when the
    # cancelled-list was non-empty. Otherwise a taxpayer with E-Invoice supplied for every
    # month but genuinely no cancellations anywhere produces a completely EMPTY
    # cancelled_by_month dict, and build_cancelled_einvoice_findings() (which distinguishes
    # "no data supplied" from "data supplied, nothing found" purely by whether this dict has
    # ANY key) then wrongly reports "No E-Invoice data supplied -- cannot test."
    cancelled_by_month = {}
    for m, res_m in zip([r["month"] for r in month_results], month_results):
        if res_m.get("einv_available"):
            cancelled_by_month[m] = res_m.get("cancelled_einvoices") or []
    g1_named_by_month = {r["month"]: r.get("g1_named_invnos", set()) for r in month_results}
    all_cancelled, cancel_findings = fchk.build_cancelled_einvoice_findings(
        cancelled_by_month, g1_named_by_month, ewb_out_rows)

    # Cross-reference Doc-Series 'missing' serials against the cancelled-e-invoice list for
    # that same month -- a Table-13-declared serial that's genuinely absent from GSTR-1 AND
    # turns out to be a cancelled e-invoice is explained, not a real gap. Mutates each month's
    # doc_gap list in place, so both write_master_dashboard (below) and write_doc_series pick
    # up the enriched status.
    for res_m in month_results:
        m = res_m["month"]
        invnos_this_month = {c["invno"] for c in cancelled_by_month.get(m, [])}
        fchk.enrich_doc_gap_with_cancelled_einvoices(res_m["doc_gap"], invnos_this_month)
        # THEN B2CS (Bug 6 fix) -- must run after the above so 'still_unexplained' reflects
        # both prior explanations first; see enrich_doc_gap_with_b2cs()'s own docstring for why
        # this can only ever resolve the ONE remaining ambiguity-free range, never a blanket
        # single-invoice-series rule.
        g1_m = res_m.get("comp_raw", {}).get("g1", {})
        b2cs_taxable_m = g1_m.get("b2cs_taxable", 0.0)
        b2cs_tax_m = (g1_m.get("b2cs_IGST", 0.0) + g1_m.get("b2cs_CGST", 0.0) +
                      g1_m.get("b2cs_SGST", 0.0) + g1_m.get("b2cs_CESS", 0.0))
        fchk.enrich_doc_gap_with_b2cs(res_m["doc_gap"], b2cs_taxable_m, b2cs_tax_m)

    # ---- NEW: flow / stock / ITC roll-forward / payment / counterparty layer ----
    # Built BEFORE the workbook is created so its findings can feed the Master Dashboard;
    # the sheets themselves are written after the annual sheets, keeping sheet order stable.
    print("Running flow / stock / ITC roll-forward / payment / counterparty checks...")
    flow_ctx = None
    try:
        flow_ctx = flow.build_context(
            months_covered, res, month_results, annual_data, ewb_out_rows, ewb_in_rows,
            gstr9, gstr9c, table8a, bs_pl_data, res["self_gstin"],
            fys_covered[0] if len(fys_covered) == 1 else None, r2a_data=r2a_data,
            blocked_credit_master_path=res.get("blocked_itc_master_file"))
        flow_sheets, flow_findings = flow.build_all(flow_ctx)
    except Exception as ex:
        print(f"  *** flow/counterparty layer failed to build: {ex} -- the rest of the run "
              f"continues and the workbook will say so. ***")
        flow_sheets, flow_findings = [], [fchk.Finding(
            "F0", "Flow / counterparty layer", "INFO",
            f"This layer could not be built for this run: {ex!r}. Every other layer ran normally.",
            {})]

    print("Writing workbook...")
    wb = openpyxl.Workbook()
    ws_dash = wb.active; ws_dash.title = "Master Dashboard"
    flow_ref_to_sheet = {f.ref: name for name, _subtitle, built in (flow_sheets or [])
                          for f in built.get("findings", [])}
    dash_items = write_master_dashboard(ws_dash, month_results, months_covered, months_gap,
                            rect_pairs, annual_review_count, hsn_findings,
                            forensic_findings=forensic_findings, cancel_findings=cancel_findings,
                            flow_findings=flow_findings, flow_ref_to_sheet=flow_ref_to_sheet)
    if run_errors:
        r = ws_dash.max_row + 2
        ws_dash.cell(r, 1, f"NOTE: {len(run_errors)} month(s) failed to process and were skipped "
                            "(their data is NOT included above):").font = Font(bold=True, color="C00000")
        for m, err in run_errors:
            r += 1
            ws_dash.cell(r, 1, f"  {m}: {err}")

    _hsn_rows_cache = {}   # keyed by resolved GSTR-1 file path -- avoids re-parsing the same
                           # merged file once per month when several months share one file
    for res_m in month_results:
        m = res_m["month"]
        raw.PERIOD_LABEL = m
        ws_comp = wb.create_sheet(sheet_name("Comparison", m))
        uni.write_comparison(ws_comp, res_m["comparisons"], only_mismatch=False)
        g1_path_this_month = res["gstr1_month_map"].get(m)
        if g1_path_this_month:
            if g1_path_this_month not in _hsn_rows_cache:
                _hsn_rows_cache[g1_path_this_month] = hfc._hsn_rows_by_month(g1_path_this_month)
            hsn_rows_this_month = _hsn_rows_cache[g1_path_this_month].get(m, [])
        else:
            hsn_rows_this_month = []
        write_hsn_review_table(ws_comp, hsn_rows_this_month, m)
        uni.write_analysis14(wb.create_sheet(sheet_name("Analysis", m)), res_m["findings14"])
        # REMOVED (per instruction): the standalone monthly '<Month> HSN Detail' sheet used to be
        # created here from the SAME `hsn_rows_this_month` list that `write_hsn_review_table()`
        # (two lines up) already writes into this month's own Comparison sheet as the "HSN RATE
        # REVIEW" table -- genuinely duplicate data, one raw Table-12 dump, one a rate-review
        # table, both keyed off the identical per-month HSN rows. `write_hsn_detail()` itself is
        # left defined (unused) rather than deleted, in case it's wanted again later.
        uni.write_eway(wb.create_sheet(sheet_name("EWB", m)), wb.create_sheet(sheet_name("EWB Detail", m)),
                        res_m["findings27"])

    write_doc_series(wb.create_sheet("Doc-Series Integrity"), month_results)
    write_rectification_sheet(wb.create_sheet("Rectification Pairs"), rect_pairs, annual_data["bo"]["drc_payments"])
    # NEW SHEET (per explicit instruction): 'CN-DN ITC Impact - Annual' -- see
    # gst_checks_flow.build_cn_dn_impact_data()'s own docstring for what it computes and why.
    # Needs flow_ctx (built above for the flow/counterparty layer); degrades to an explicit
    # skipped sheet, never a crash, if that layer itself failed to build for this run.
    ws_cndn = wb.create_sheet("CN-DN ITC Impact - Annual")
    if flow_ctx is not None:
        try:
            write_cn_dn_impact_sheet(ws_cndn, flow.build_cn_dn_impact_data(flow_ctx))
        except Exception as ex:
            ws_cndn.cell(1, 1, "CREDIT NOTE / DEBIT NOTE — ITC IMPACT (ANNUAL)").font = TITLEF
            ws_cndn.cell(3, 1, f"SKIPPED -- this sheet could not be built for this run: {ex!r}. "
                               f"Every other sheet in this workbook is unaffected.").font = Font(italic=True, color="9C0006")
    else:
        ws_cndn.cell(1, 1, "CREDIT NOTE / DEBIT NOTE — ITC IMPACT (ANNUAL)").font = TITLEF
        ws_cndn.cell(3, 1, "SKIPPED -- the flow/counterparty layer this sheet depends on did not "
                           "build for this run (see the Master Dashboard for the reason). Every "
                           "other sheet in this workbook is unaffected.").font = Font(italic=True, color="9C0006")
    write_hsn_fraud_checks(wb.create_sheet("HSN & Fraud Pattern Checks"), hsn_findings)
    write_filing_compliance(wb.create_sheet("Filing Compliance & Late Fee"), compliance_records)
    write_forensic_checks(wb.create_sheet("Forensic Checks (R13-R14)"), forensic_findings)
    write_cancelled_einvoices(wb.create_sheet("Cancelled E-Invoices"), all_cancelled, cancel_findings,
                               einv_cancel_col_found_any)

    # ---- NEW: annual-detail sheets (this session's dedup/annual-detail prompt) -- additive
    # only, every existing per-month sheet above is completely unchanged. ----
    write_ewb_pattern_annual(wb.create_sheet("EWB Pattern - Annual"), month_results, hsn_findings)
    write_irn_late_annual(wb.create_sheet("IRN Late-Gen - Annual"), hsn_findings,
                          einv_month_map=res.get("einv_month_map"), months_covered=months_covered,
                          ewb_out_rows=ewb_out_rows)
    write_itc_blocked_annual(wb.create_sheet("ITC-Blocked Inv - Annual"), hsn_findings)
    write_round_number_annual(wb.create_sheet("Round-Number Inv - Annual"), hsn_findings)
    write_hsn_timeline_annual(wb.create_sheet("HSN Timeline - Annual"), hsn_findings)

    annualwb.write_monthly(wb.create_sheet("Annual Ledger Walkthrough"), annual_rows,
                            fy_label=(fys_covered[0] if len(fys_covered) == 1 else None))
    annualwb.write_fy_total_vs_bifa(wb.create_sheet("FY Total vs BIFA"), annual_rows, annual_data, ewb_out_rows,
                                     fy=(fys_covered[0] if len(fys_covered) == 1 else None))
    annualwb.write_related_party(wb.create_sheet("Related-Party Alerts"), annual_data)
    annualwb.write_top_counterparties(wb.create_sheet("Top Counterparties"), annual_data)
    flow.write_all(wb, flow_sheets)
    # NEW TABLE (per explicit instruction): appended below the existing 'ITC Annual Summary'
    # sheet's own content, same sheet -- see write_itc_detailed_recon_table()'s own docstring.
    # Guarded on the sheet actually existing: if the flow/counterparty layer failed entirely
    # above, flow_sheets is [] and 'ITC Annual Summary' was never created at all this run.
    if "ITC Annual Summary" in wb.sheetnames:
        try:
            write_itc_detailed_recon_table(wb["ITC Annual Summary"], flow.build_itc_detailed_recon_data(flow_ctx))
        except Exception as ex:
            ws_itcrecon = wb["ITC Annual Summary"]
            r_err = ws_itcrecon.max_row + 3
            ws_itcrecon.cell(r_err, 1, f"SKIPPED -- the tax-head-wise ITC reconciliation table below "
                             f"could not be built for this run: {ex!r}. The sheet's existing content "
                             f"above is unaffected.").font = Font(italic=True, color="9C0006")

    # ---- NEW: Potential Blocked Credits (additive -- new module, new sheet only; does not
    # touch any calculation, merge, or sheet above). Reuses the SAME invoice-level GSTR-2B rows
    # the flow layer already parsed (flow_ctx['twob_lines_fy']) rather than re-parsing 2B a
    # third time. Master keyword list is content-detected by classify_folder() like every other
    # input in this tool -- never by filename. ----
    print("Scanning GSTR-2B for potential blocked credits (keyword/trade-name match)...")
    _b2b_by_month = {}
    for _x in (flow_ctx or {}).get("twob_lines_fy", []):
        _b2b_by_month.setdefault(_x.get("month"), []).append(_x)
    try:
        blocked_credit_status = bcred.build_and_write(
            wb, res.get("blocked_itc_master_file"), _b2b_by_month)
    except Exception as ex:
        blocked_credit_status = f"SKIPPED -- unexpected error building this sheet: {ex!r}"
    print(f"  {blocked_credit_status}")

    # ---- NEW: Machinery HSN Scan (additive -- new module, new sheet only). Reuses the SAME
    # GSTR-1 HSN cache the monthly Comparison sheets already built (_hsn_rows_cache) and the
    # SAME EWB/2B data every other layer above already parsed -- nothing is re-classified or
    # re-read from disk beyond what this run already needed anyway. ----
    print("Scanning GSTR-1 / E-Way Bill / GSTR-2B for machinery HSN purchases and sales...")
    _g1_hsn_by_month = {}
    for _cache in _hsn_rows_cache.values():
        for _m, _rows in _cache.items():
            _g1_hsn_by_month.setdefault(_m, []).extend(_rows)
    try:
        machinery_status = mscan.build_and_write(
            wb, res.get("machinery_hsn_master_file"), _g1_hsn_by_month,
            ewb_out_rows, ewb_in_rows, _b2b_by_month, res["self_gstin"])
    except Exception as ex:
        machinery_status = f"SKIPPED -- unexpected error building this sheet: {ex!r}"
    print(f"  {machinery_status}")

    # ---- QA review layer (integrated, was previously standalone post-processing scripts) --
    # appended at the very END of the workbook, after every original sheet above. Reuses
    # dash_items (already correctly amount-tagged and root_id/level-deduped by
    # write_master_dashboard itself) rather than re-deriving anything from free text.
    fy_tag_for_qa = (fys_covered[0] if len(fys_covered) == 1 else
                     f"{fys_covered[0]}_to_{fys_covered[-1]}" if fys_covered else "UNKNOWN_FY")
    qa_rows = build_qa_layer(dash_items)
    ws_qa_dash = wb.create_sheet("Reviewed Master Dashboard")
    write_qa_reviewed_dashboard(ws_qa_dash, qa_rows)
    ws_qa_action = wb.create_sheet("Action Required")
    action_rows, total_exposure = write_qa_action_required(ws_qa_action, qa_rows)
    ws_qa_ewb = wb.create_sheet("EWB Full-Year Reconciliation")
    ewb_out_gaps, ewb_in_gaps = write_qa_ewb_reconciliation(ws_qa_ewb, month_results)
    ws_qa_summary = wb.create_sheet("QA Summary")
    write_qa_summary(ws_qa_summary, qa_rows, action_rows, total_exposure, ewb_out_gaps, ewb_in_gaps,
                     res["self_gstin"], res["company_name"], fy_tag_for_qa)
    # Reorder the 4 QA sheets so they read QA Summary -> Action Required -> EWB Full-Year
    # Reconciliation -> Reviewed Master Dashboard, exactly at the tail of the workbook.
    desired_tail = ["QA Summary", "Action Required", "EWB Full-Year Reconciliation", "Reviewed Master Dashboard"]
    base_names = [s for s in wb.sheetnames if s not in desired_tail]
    sheets_by_name = {s.title: s for s in wb._sheets}
    wb._sheets = [sheets_by_name[n] for n in base_names + desired_tail]

    fy_tag = (fys_covered[0] if len(fys_covered) == 1 else
              f"{fys_covered[0]}_to_{fys_covered[-1]}" if fys_covered else "UNKNOWN_FY")
    outfile = f"GST_MASTER_{res['self_gstin']}_FY{fy_tag}.xlsx"
    # Unfreeze panes on every sheet, per explicit instruction. Deliberately done here as one
    # guaranteed-complete pass over every worksheet actually in the final workbook, rather than
    # hunting down and editing the ~20 individual `ws.freeze_panes = ...` call sites scattered
    # across gst_report.py/gst_checks_monthly.py/gst_checks_flow.py/master_build.py/
    # gst_machinery_scan.py -- this way nothing is missed (including any sheet added later by a
    # future change) and there's a single place to look if this behaviour ever needs revisiting.
    for _ws in wb.worksheets:
        _ws.freeze_panes = None
    # NUMBERS-STORED-AS-TEXT FIX (per explicit request): one guaranteed-complete pass over the
    # in-memory workbook, converting every text cell that's provably a lossless integer to a
    # real number -- see gst_core.convert_numeric_text_to_numbers's own docstring for exactly
    # what is and isn't touched and why. Must run before save (it changes cell values, unlike
    # fix_ooxml_conformance below, which only touches the saved file's raw ZIP structure).
    _num_converted, _num_skipped = mpu.convert_numeric_text_to_numbers(wb)
    wb.save(outfile)
    # BUG FIX (real, confirmed, not theoretical -- reported against actual Excel by the user,
    # verified by byte-level diff against their own Excel-repaired copy, and independently
    # reproduced on a minimal single-cell workbook): openpyxl 3.1.5 in this environment omits
    # the XML declaration on every part and doesn't order the ZIP per the OOXML spec, both of
    # which are exactly what Excel's own strict parser objects to. See
    # gst_core.fix_ooxml_conformance's own docstring for the full diagnosis. This is a pure
    # post-save fix-up -- no cell content, style, or sheet data is touched, only the raw ZIP
    # structure of the file already written above.
    mpu.fix_ooxml_conformance(outfile)
    print(f"  Numbers-stored-as-text fix: {_num_converted} cell(s) converted to real numbers, "
          f"{_num_skipped} left as text (would have lost a leading zero or similar -- e.g. an "
          f"HSN code like '09' or '035').")
    print(f"\nSaved: {outfile}")
    print(f"Months covered: {months_covered}")
    print(f"Gaps within span: {months_gap}")
    print(f"Rectification pairs: {len(rect_pairs)}")
    print(f"Annual-source REVIEW flags: {annual_review_count}")
    print(f"Cancelled e-invoices found: {len(all_cancelled)}")
    nflag = sum(1 for f in hsn_findings if f.severity == "FLAG")
    nrev = sum(1 for f in hsn_findings if f.severity == "REVIEW")
    print(f"HSN & Fraud Pattern Checks: {len(hsn_findings)} total ({nflag} FLAG, {nrev} REVIEW)")
    print(f"Forensic checks (R13/R14): {[(f.ref, f.severity) for f in forensic_findings]}")
    nf = sum(1 for f in flow_findings if f.severity in ("FLAG", "MISMATCH"))
    nr = sum(1 for f in flow_findings if f.severity == "REVIEW")
    print(f"Flow / counterparty checks: {len(flow_findings)} findings across {len(flow_sheets)} "
          f"sheets ({nf} FLAG, {nr} REVIEW)")
    g_findings = [f for f in flow_findings if f.ref.startswith("G")]
    if g_findings:
        gf = sum(1 for f in g_findings if f.severity in ("FLAG", "MISMATCH"))
        gr = sum(1 for f in g_findings if f.severity == "REVIEW")
        gs = sum(1 for f in g_findings if f.severity == "SKIPPED")
        print(f"  GSTR-2A checks (G1-G10): {len(g_findings)} findings ({gf} FLAG, {gr} REVIEW, "
              f"{gs} SKIPPED)")
    return outfile


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else ".")
