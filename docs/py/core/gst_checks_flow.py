#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GST CHECKS -- FLOW, STOCK, ITC ROLL-FORWARD, PAYMENT & COUNTERPARTY FORENSICS
==============================================================================
This module implements the requirement set added after the first release:

  F1  Purchase vs Sales, monthly and FY, with the running value difference
      ("stock left with the taxpayer" in MONEY terms, per explicit instruction)
  F2  Total ITC availed / reversed, with the reversal split by whether it sits
      in Table 4B(1) (Rules 38/42/43 -- permanent) or 4B(2) (Others -- usually
      temporary/reclaimable), monthly and FY
  F3  Yearly ITC received / utilised / reversed, reconciled to the Electronic
      Credit Ledger's own credits, debits and closing balance
  F4  Three-way: GSTR-1 value vs GSTR-3B value vs E-Way-Bill value
  F5  E-Way-Bill value vs invoice value (document-level, deduplicated)
  F6  B2B -> B2C shift (turnover moved to unregistered buyers, which breaks the
      ITC trail and removes the counterparty's incentive to report)
  F7  ITC claimed in 3B vs ITC available in 2B -- computed INVOICE-LEVEL
      (see the CRITICAL note below)
  F8  RCM ITC vs RCM liability, verified against Electronic Cash Ledger debits
  F9  DRC-03 / voluntary payments, plus every other non-return movement in the
      ledgers (refund debits, Rule 86A blocking)
  F10 Turnover growth vs tax payment across every FY the BO Profile carries
  F11 Counterparty transactions: same-day repeats (NO value floor, per explicit
      instruction) and reciprocal (buy-and-sell-to-the-same-party) pairs
  F12 Top 10 suppliers by ITC received and top 10 buyers by ITC passed on,
      computed from the returns AND compared against the department's own
      BO-Profile lists

CRITICAL NOTE ON GSTR-2B (why F7 does not use the existing summary path)
------------------------------------------------------------------------
gstr2b_parser's 'ITC Available' summary reader takes the FIRST numeric group in
each row of a quarterly block. That group is month-1-of-the-quarter, so months
2 and 3 of every quarter receive month 1's figures. Measured on the reference
taxpayer: Sep-22 +Rs 25.05L, Nov-22 -Rs 19.94L, Mar-23 +Rs 29.77L, FY total
understated by Rs 20.94L. Every 2B figure in this module is therefore computed
from the INVOICE-LEVEL 'B2B' and 'B2B-CDNR' sheets, which carry their own exact
per-row period tag. The summary figure is still shown alongside, as a visible
control total, so the difference is documented rather than silently corrected.

DISCIPLINE (unchanged from the rest of the tool): never fabricate or silently
guess a number. Anything not supplied produces an explicit SKIPPED/INFO row
stating what was missing, never a zero that reads like a verified nil.
"""

import datetime as _dt
import os
import re
from collections import defaultdict

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import gst_core as mpu
import gst_parsers_returns as pr
import gst_parsers_dept as dept
import gst_checks_hsn_fraud as hfc
from gst_checks_forensic import Finding
import gst_config as cfg

# ----------------------------------------------------------------------
# styling (kept local so this module can be dropped into any workbook)
RED = PatternFill("solid", fgColor="FFC7CE")
AMBER = PatternFill("solid", fgColor="FFEB9C")
BLUE = PatternFill("solid", fgColor="DDEBF7")
GREEN = PatternFill("solid", fgColor="C6EFCE")
GREY = PatternFill("solid", fgColor="E7E6E6")
HEADFILL = PatternFill("solid", fgColor="1F3864")
SECTFILL = PatternFill("solid", fgColor="D9E1F2")
TITLEF = Font(bold=True, size=13, color="1F3864")
BORDER = Border(*[Side(style="thin", color="BFBFBF")] * 4)
SEV_FILL = {"FLAG": RED, "MISMATCH": RED, "REVIEW": AMBER, "INFO": BLUE,
            "PASS": GREEN, "MATCH": GREEN, "SKIPPED": GREY}
SEV_FONT = {"FLAG": Font(bold=True, color="9C0006"), "MISMATCH": Font(bold=True, color="9C0006"),
            "REVIEW": Font(bold=True, color="9C6500"), "INFO": Font(bold=True, color="2F5496"),
            "PASS": Font(bold=True, color="006100"), "MATCH": Font(bold=True, color="006100"),
            "SKIPPED": Font(bold=True, color="808080")}

TOL = 1.0          # rupee rounding tolerance
MATERIAL = 100000.0  # Rs 1 lakh -- the "worth an officer's attention" floor


def _n(v):
    if v is None:
        return 0.0
    try:
        return float(str(v).replace(",", "").replace("\u20b9", "").strip())
    except (ValueError, AttributeError):
        return 0.0


def _f(v):
    """Format a rupee amount for a detail string."""
    return f"Rs {v:,.2f}"


def _pct(a, b):
    return (a / b * 100.0) if b else None


# ======================================================================
# SECTION 1 -- FY-WIDE LINE READERS
# ======================================================================
# These read the SAME merged workbooks the rest of the tool reads, but return
# whole-FY invoice-level lists rather than one month's aggregates, because the
# counterparty / stock / roll-forward views are inherently cross-month.

def read_gstr1_b2b_ff(path, month):
    """Invoice-level B2B rows for one month, WITH the invoice header
    forward-filled onto continuation rows.

    Why this exists rather than reusing gst_analysis_checks.read_gstr1_lines():
    a multi-rate invoice occupies several rows in the portal's export, and only
    the FIRST row carries GSTIN / invoice number / date -- the rate lines below
    it have those cells blank. read_gstr1_lines() takes each row as-is, so the
    continuation lines arrive with an empty GSTIN and empty invoice number.
    That is harmless for the rate-level checks it feeds, but it would silently
    drop taxable value out of any per-counterparty or per-invoice aggregate.
    Confirmed present in the reference file (e.g. AJANTA PHARMA MR22-23/0020
    carries a 12% line and an 18% line)."""
    if not path or not os.path.exists(path):
        return []
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if "b2b, sez, de_inv" not in wb.sheetnames:
        return []
    rows = list(wb["b2b, sez, de_inv"].iter_rows(values_only=True))
    hdr = [str(c).strip() if c else "" for c in rows[3]]
    H = {h: i for i, h in enumerate(hdr)}

    def g(r, k):
        i = H.get(k)
        return r[i] if i is not None and i < len(r) else None

    out, carry = [], None
    for r in mpu.rows_for_month(rows, 3, month):
        if not any(r):
            continue
        gstin = str(g(r, "GSTIN/UIN of Recipient") or "").strip()
        if gstin:
            carry = dict(gstin=gstin,
                         name=str(g(r, "Receiver Name") or "").strip(),
                         invno=str(g(r, "Invoice Number") or "").strip(),
                         invdate=g(r, "Invoice date"),
                         invval=_n(g(r, "Invoice Value")),
                         pos=str(g(r, "Place Of Supply") or "").strip(),
                         invtype=str(g(r, "Invoice Type") or "").strip())
        if carry is None:
            continue    # data before any header row -- cannot attribute, skip rather than guess
        out.append(dict(carry, month=month, rate=_n(g(r, "Rate")),
                        taxable=_n(g(r, "Taxable Value")),
                        igst=_n(g(r, "Integrated Tax")), cgst=_n(g(r, "Central Tax")),
                        sgst=_n(g(r, "State/UT Tax")), cess=_n(g(r, "Cess Amount")),
                        header_row=bool(gstin)))
    return out


def read_gstr1_b2c(path, month):
    """B2CS (rate-wise, no invoice) + B2CL (invoice-level) for one month."""
    res = dict(b2cs_taxable=0.0, b2cs_tax=0.0, b2cl_taxable=0.0, b2cl_tax=0.0,
               b2cl_invoices=0, available=True)
    if not path or not os.path.exists(path):
        res["available"] = False
        return res
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    for sheet, pfx in (("b2cs", "b2cs"), ("b2cl", "b2cl")):
        if sheet not in wb.sheetnames:
            continue
        rows = list(wb[sheet].iter_rows(values_only=True))
        hi = None
        for i, r in enumerate(rows[:8]):
            cells = [str(c).strip() if c else "" for c in r]
            if "Taxable Value" in cells:
                hi = i
                break
        if hi is None:
            continue
        hdr = [str(c).strip() if c else "" for c in rows[hi]]
        H = {h: i for i, h in enumerate(hdr)}
        try:
            month_rows = mpu.rows_for_month(rows, hi, month)
        except mpu.PeriodParseError:
            continue

        def g(r, k):
            i = H.get(k)
            return r[i] if i is not None and i < len(r) else None

        for r in month_rows:
            if not any(r):
                continue
            res[pfx + "_taxable"] += _n(g(r, "Taxable Value"))
            res[pfx + "_tax"] += (_n(g(r, "Integrated Tax")) + _n(g(r, "Central Tax"))
                                  + _n(g(r, "State/UT Tax")))
            if pfx == "b2cl" and g(r, "Invoice Number"):
                res["b2cl_invoices"] += 1
    return res


def read_gstr3b_extra(path, month):
    """Table 4 fields the existing parse_gstr3b() does not return: 4A(1)/(2)/(4)
    and the 4D ineligible block. Same content-based approach and the same
    section-anchoring discipline the existing parser uses for 4B (the literal
    '(2) Others' appears in BOTH the 4B and 4D blocks in the pre-Aug-2022
    format, so an unanchored scan silently reads the wrong one)."""
    out = dict(A1=[0.0] * 4, A2=[0.0] * 4, A4=[0.0] * 4, D1=[0.0] * 4, D2=[0.0] * 4,
               available=False)
    if not path or not os.path.exists(path):
        return out
    wb = pr.load_xlsx(path)
    ws = None
    for sn in wb.sheetnames:
        try:
            if pr._gstr3b_sheet_month(wb[sn]) == month:
                ws = wb[sn]
                break
        except mpu.PeriodParseError:
            continue
    if ws is None:
        return out
    rows = [[c.value for c in r] for r in ws.iter_rows()]

    def nums(r):
        return [_n(c) for c in r
                if isinstance(c, (int, float))
                or (isinstance(c, str) and re.match(r"^-?[\d,\.]+$", str(c).strip()))]

    for r in rows:
        j = " ".join(str(c) for c in r if c is not None).strip()
        v = nums(r)
        if not v:
            continue
        if j.startswith("(1) Import of goods"):
            out["A1"] = v[:4]
        elif j.startswith("(2) Import of services"):
            out["A2"] = v[:4]
        elif j.startswith("(4) Inward supplies from ISD"):
            out["A4"] = v[:4]

    d_start = None
    for i, r in enumerate(rows):
        j = " ".join(str(c) for c in r if c is not None).strip()
        if j.startswith("(D)") or j.startswith("D. "):
            d_start = i
            break
    if d_start is not None:
        for r in rows[d_start + 1:d_start + 6]:
            j = " ".join(str(c) for c in r if c is not None).strip()
            v = nums(r)
            if not v:
                continue
            if j.startswith("(1)"):
                out["D1"] = v[:4]
            elif j.startswith("(2)"):
                out["D2"] = v[:4]
    out["available"] = True
    return out


def read_2b_invoice_level(path, month):
    """Invoice-level 2B for one month: B2B invoices + B2B-CDNR notes.
    Returns totals AND the raw rows (needed by the counterparty views).
    Reuses the existing, per-row-period-tagged parser -- it is the SUMMARY
    reader that is unreliable, not the invoice reader.

    UPDATED (per explicit instruction): every row's ITC Availability status is classified into
    exactly one of three buckets -- YES (portal confirms available), NO (portal flags
    ineligible -- e.g. confirmed real: 'POS and supplier state are same but recipient state is
    different'), or UNCONFIRMED (status could not be read -- confirmed real: on some CDNR rows
    with additional misalignment beyond Rate, the 'back half' of the row -- Filing Date, ITC
    Availability, Reason -- has its own separate shift this tool does not yet resolve, so the
    status field itself reads blank/NA rather than a real value; the TAX AMOUNTS on these same
    rows are correctly read via the rate-based shift fix -- only the status is unknown).

    PRIMARY 'available' figure (tot['itc'], tot['cn_itc']) = YES + UNCONFIRMED (excludes only
    a CONFIRMED No -- per instruction, an unconfirmed row is not assumed ineligible, since
    there is no portal statement that it is). Two REFERENCE figures are also computed and
    returned, unrounded, so a reader can pick whichever definition suits: tot['itc_yes_only']
    (excludes unconfirmed too) and tot['itc_grand_total'] (Yes+No+Unconfirmed, no exclusion at
    all). ARITHMETIC IDENTITY that must hold exactly, checked by the caller/tests:
    yes + no + unconfirmed == grand_total, and yes + unconfirmed == grand_total - no."""
    empty = dict(available=False, taxable=0.0, igst=0.0, cgst=0.0, sgst=0.0, cess=0.0,
                 itc=0.0, cn_taxable=0.0, cn_itc=0.0, invoices=0, suppliers=0,
                 rcm_itc=0.0, rows=[], cdnr=[],
                 itc_yes_only=0.0, itc_no_only=0.0, itc_unconfirmed_only=0.0, itc_grand_total=0.0,
                 cn_itc_yes_only=0.0, cn_itc_no_only=0.0, cn_itc_unconfirmed_only=0.0, cn_itc_grand_total=0.0,
                 no_confirmed_rows=[], unconfirmed_rows=[])
    if not path or not os.path.exists(path):
        return empty
    try:
        parsed = pr.parse_2b_excel(path, month)
    except mpu.PeriodParseError:
        return empty
    b2b, cdnr = parsed["b2b"], parsed["cdnr"]
    tot = dict(empty, available=True, rows=b2b, cdnr=cdnr)
    tot["invoices"] = len(b2b)
    tot["suppliers"] = len({x["gstin"] for x in b2b})
    no_confirmed_rows, unconfirmed_rows = [], []
    for x in b2b:
        status = str(x.get("itc_avail", "")).strip().upper()
        line_itc = x["igst"] + x["cgst"] + x["sgst"] + x["cess"]
        if status == "YES":
            tot["itc_yes_only"] += line_itc
            if str(x.get("rcm", "")).strip().upper().startswith("Y"):
                tot["rcm_itc"] += line_itc
        elif status == "NO":
            tot["itc_no_only"] += line_itc
            no_confirmed_rows.append(dict(month=month, source="B2B", **x))
        else:
            tot["itc_unconfirmed_only"] += line_itc
            unconfirmed_rows.append(dict(month=month, source="B2B", **x))
        if status != "NO":
            # PRIMARY aggregate = YES + UNCONFIRMED (every status except a confirmed NO).
            tot["taxable"] += x["taxable"]
            tot["igst"] += x["igst"]; tot["cgst"] += x["cgst"]
            tot["sgst"] += x["sgst"]; tot["cess"] += x["cess"]
    tot["itc"] = tot["igst"] + tot["cgst"] + tot["sgst"] + tot["cess"]
    tot["itc_grand_total"] = tot["itc_yes_only"] + tot["itc_no_only"] + tot["itc_unconfirmed_only"]
    for c in cdnr:
        status = str(c.get("itc_avail", "")).strip().upper()
        sign = -1.0 if str(c.get("ntype", "")).strip().upper().startswith("C") else 1.0
        line_itc = sign * (c["igst"] + c["cgst"] + c["sgst"] + c["cess"])
        if status == "YES":
            tot["cn_itc_yes_only"] += line_itc
        elif status == "NO":
            tot["cn_itc_no_only"] += line_itc
            no_confirmed_rows.append(dict(month=month, source="CDNR", **c))
        else:
            tot["cn_itc_unconfirmed_only"] += line_itc
            unconfirmed_rows.append(dict(month=month, source="CDNR", **c))
        if status != "NO":
            # PRIMARY aggregate = YES + UNCONFIRMED, same rule as B2B above.
            tot["cn_taxable"] += sign * c["taxable"]
            tot["cn_itc"] += line_itc
    tot["cn_itc_grand_total"] = tot["cn_itc_yes_only"] + tot["cn_itc_no_only"] + tot["cn_itc_unconfirmed_only"]
    tot["no_confirmed_rows"] = no_confirmed_rows
    tot["unconfirmed_rows"] = unconfirmed_rows
    return tot


def dedupe_ewb(rows):
    """E-Way-Bill exports can repeat a row verbatim (confirmed: 6 duplicated
    rows out of 1,104 in the reference inward file). Every aggregate here
    dedupes on EWB number first; the count of dropped rows is reported, never
    hidden."""
    seen, out, dropped = set(), [], 0
    for r in rows:
        k = str(r.get("ewbno") or "").strip()
        if k and k in seen:
            dropped += 1
            continue
        if k:
            seen.add(k)
        out.append(r)
    return out, dropped


def bo_financial_rows(bo):
    """Re-derive the BO Profile's Financial Information table.

    BUG FIX (found while investigating a reported always-empty 'Turnover Growth vs Tax' sheet):
    this function was written for the OLD PDF-text-extraction BO Profile parser, which stored
    each FY's row as a list of raw text tokens under rec['raw'] (column count varying 12-13
    tokens depending on whether 'Cash Utilization %' was printed, hence the token-scanning logic
    below). The CURRENT BO Profile parser (gst_parsers_dept.parse_bo_profile, XLSX-based --
    confirmed this is what every real BO Profile supplied so far actually is) returns each FY's
    row as ALREADY-STRUCTURED fields (turnover=, total_tax_liability=, etc.) with NO 'raw' key at
    all. Since `rec.get('raw') or []` is always [] for that format, every single year fell into
    the `len(amounts) < 7` branch and was marked usable=False -- silently, for every FY, on every
    taxpayer, since the XLSX parser was introduced. Confirmed directly against a real BO Profile
    (10 years of genuinely present data, e.g. FY2025-26 turnover Rs 15,450.95 lakh) -- the data
    was always there; this function just never looked in the right place for it.

    Fixed to use the structured fields directly when present (the current/normal case), falling
    back to the old raw-token re-derivation only for a legacy 'raw'-shaped record (kept for
    backward compatibility in case an older cached/PDF-sourced record is ever passed in)."""
    out = {}
    for fy, rec in (bo or {}).get("financial_by_fy", {}).items():
        if rec.get("turnover") is not None and rec.get("total_tax_liability") is not None:
            out[fy] = dict(usable=True, raw=rec.get("raw") or [],
                           turnover=_n(rec.get("turnover")), taxable_turnover=_n(rec.get("taxable_turnover")),
                           rcm=_n(rec.get("rcm")), total_tax_liability=_n(rec.get("total_tax_liability")),
                           tax_paid_by_itc=_n(rec.get("tax_paid_by_itc")),
                           tax_paid_in_cash=_n(rec.get("tax_paid_in_cash")),
                           itc_availed=_n(rec.get("itc_availed")))
            continue
        toks = rec.get("raw") or []
        amounts = []
        for t in toks:
            if "%" in str(t):
                break
            amounts.append(_n(t))
        if len(amounts) < 7:
            out[fy] = dict(usable=False, raw=toks)
            continue
        out[fy] = dict(usable=True, raw=toks,
                       turnover=amounts[0], taxable_turnover=amounts[1], rcm=amounts[2],
                       total_tax_liability=amounts[3], tax_paid_by_itc=amounts[4],
                       tax_paid_in_cash=amounts[5], itc_availed=amounts[6])
    return out


# ======================================================================
# SECTION 2 -- BUILDERS  (each returns dict(header, rows, findings, notes))
# ======================================================================

def ledger_window_note(months, present_periods, what):
    """The electronic ledgers are extracted for a fixed date window (typically
    01-Apr to 31-Mar). A return for the LAST month of that window is filed --
    and its liability discharged -- in the FOLLOWING month, which falls outside
    the extract. Comparing a full year of returns against a ledger that is
    structurally missing the last month's movement produces a shortfall that is
    an artefact of the extract, not a short payment. This returns the months
    that genuinely have no ledger coverage so every ledger comparison can be
    run over the intersection and say plainly what it excluded."""
    missing = [m for m in months if m not in present_periods]
    if not missing:
        return [], ""
    return missing, (f"{len(missing)} month(s) of the period under review have no {what} entry at "
                     f"all ({', '.join(missing)}); for the final month of a financial year this is "
                     f"normally because the return is filed in the following month, outside the "
                     f"ledger extract window. Those months are EXCLUDED from the comparison below "
                     f"rather than being counted as nil.")


def build_purchase_sales_stock(ctx):
    """F1 -- purchase vs sales, monthly and cumulative, in MONEY terms.

    Explicitly NOT a costed stock figure: GSTR-2B carries no quantity column at
    all, and sales values carry the taxpayer's margin while purchase values do
    not. What this table shows is the VALUE FLOW difference (purchases in minus
    sales out) and its running cumulative total. The audited Inventories
    movement is printed underneath as independent context, not folded into the
    arithmetic."""
    rows, findings = [], []
    cum = 0.0
    tot_p = tot_s = 0.0
    any2b = False
    for m in ctx["months"]:
        g1 = ctx["g1_by_month"].get(m, {})
        sales = _n(g1.get("taxable")) - _n(g1.get("cn_taxable"))
        two = ctx["twob_by_month"].get(m, {})
        if two.get("available"):
            any2b = True
            purch = two["taxable"] + two["cn_taxable"]
            diff = purch - sales
            cum += diff
            tot_p += purch; tot_s += sales
            rows.append([m, purch, two["invoices"], sales, ctx["g1_invcount_by_month"].get(m, 0),
                         diff, cum, "" ])
        else:
            rows.append([m, None, None, sales, ctx["g1_invcount_by_month"].get(m, 0),
                         None, cum, "GSTR-2B not available for this month -- purchase side unknown"])
    if any2b:
        rows.append(["FY TOTAL", tot_p, None, tot_s, None, tot_p - tot_s, cum, ""])
        sev = "REVIEW" if abs(tot_p - tot_s) > MATERIAL else "INFO"
        findings.append(Finding(
            "F1", "Purchase vs sales value flow (money terms)", sev,
            f"FY taxable purchases per GSTR-2B (invoice level, net of credit/debit notes) "
            f"{_f(tot_p)}; FY taxable sales per GSTR-1 (net of credit notes) {_f(tot_s)}; "
            f"value flow difference {_f(tot_p - tot_s)}. This is a VALUE difference, not a "
            f"costed closing stock -- sales carry margin, purchases do not, and GSTR-2B has no "
            f"quantity column, so a quantity-based stock figure cannot be derived from GST "
            f"returns at all. Compare against the audited Inventories movement shown on the sheet.",
            dict(purchases=tot_p, sales=tot_s, difference=tot_p - tot_s)))
    else:
        findings.append(Finding("F1", "Purchase vs sales value flow (money terms)", "SKIPPED",
                                "GSTR-2B was not supplied for any month -- the purchase side of this "
                                "comparison has no source. Sales side alone is shown on the sheet.",
                                {}))
    notes = []
    bs = ctx.get("bs_pl_data") or {}
    inv = bs.get("inventories")
    if inv:
        op, cl = _n(inv.get("fy_prior")), _n(inv.get("fy_current"))
        notes.append(f"Audited Inventories (from bs_pl_input.py, hand-typed): opening {_f(op)}, "
                     f"closing {_f(cl)}, movement {_f(cl - op)}.")
        if any2b:
            findings.append(Finding(
                "F1b", "Value flow vs audited inventory movement", "REVIEW",
                f"GST-return value flow (purchases minus sales) is {_f(tot_p - tot_s)} for the year, "
                f"while the audited balance sheet shows inventories moving by {_f(cl - op)}. These two "
                f"figures answer different questions and will not agree by construction (margin, "
                f"non-GST costs, opening/closing timing) -- shown together so the size and direction "
                f"of the gap is on the record and can be explained by the taxpayer, not to assert an "
                f"arithmetic mismatch.",
                dict(flow=tot_p - tot_s, inventory_movement=cl - op)))
    else:
        notes.append("Audited Inventories not available (bs_pl_input.py not filled in for this "
                     "taxpayer) -- no independent inventory context could be shown.")

    # ---- NEW (per instruction, point 1): HSN-wise YEARLY comparison, Inward (E-Way Bill) vs
    # Outward (GSTR-1) supply. Genuine third-party inward movements only (self-to-self branch/
    # stock transfers excluded, same convention as the Machinery HSN Scan and Zero-Tax Scan
    # sheets). Compared at 4-digit HSN heading level, since EWB's own HSN Code and GSTR-1's own
    # HSN summary code are not guaranteed to be reported at the same digit length for the same
    # taxpayer -- 4-digit is the common denominator both can always be reduced to.
    self_gstin = ctx.get("self_gstin") or ""
    # ewb_in_by_hsn: [0]=assess value, [1]=tax value, [2]=EWB count, [3]=a sample desc.
    # BUG FIX/ADDITION (per explicit request): 'Inward Assessable Value (EWB)' had no
    # accompanying tax-value column -- the raw inward EWB row already carries its own 'Tax
    # Val.' figure (same source as Assess Val.), just never summed here. Added as its own
    # column, same per-HSN aggregation as assess value.
    ewb_in_by_hsn = defaultdict(lambda: [0.0, 0.0, 0, ""])
    for x in (ctx.get("ewb_in_rows") or []):
        if x.get("from_gstin") and x.get("from_gstin") != self_gstin and x.get("to_gstin") == self_gstin:
            h4 = str(x.get("hsn") or "").strip()[:4]
            if not h4:
                continue
            e = ewb_in_by_hsn[h4]
            e[0] += _n(x.get("assess")); e[1] += _n(x.get("taxval")); e[2] += 1
            if not e[3] and x.get("hsn_desc"):
                e[3] = x["hsn_desc"]
    g1_out_by_hsn = defaultdict(lambda: [0.0, 0.0, ""])   # taxable, tax, description
    for month_rows in (ctx.get("g1_hsn_by_month") or {}).values():
        for row in month_rows:
            h4 = str(row.get("hsn") or "").strip()[:4]
            if not h4:
                continue
            g = g1_out_by_hsn[h4]
            g[0] += _n(row.get("taxable"))
            g[1] += _n(row.get("igst")) + _n(row.get("cgst")) + _n(row.get("sgst")) + _n(row.get("cess"))
            if not g[2] and row.get("desc"):
                g[2] = row.get("desc")
    all_h4 = sorted(set(ewb_in_by_hsn) | set(g1_out_by_hsn))
    hsn_compare_rows = []
    for h in all_h4:
        ein = ewb_in_by_hsn.get(h, [0.0, 0.0, 0, ""])
        gout = g1_out_by_hsn.get(h, [0.0, 0.0, ""])
        hsn_compare_rows.append([h, gout[2] or ein[3], round(ein[0], 2), round(ein[1], 2), ein[2],
                                 round(gout[0], 2), round(gout[1], 2), round(gout[0] - ein[0], 2)])
    hsn_compare_rows.sort(key=lambda r: -(r[2] + r[5]))

    # ---- NEW (per instruction, point 2): commodity-wise list from inward EWB's own HSN
    # Description -- which commodities, and how many, moved inward across the whole FY.
    # IMPORTANT LIMITATION, stated plainly rather than guessed around: the inward E-Way Bill
    # export has NO quantity/UQC column at all (confirmed directly against the raw file header:
    # EWB No. / From & To GSTIN+Name / From & To Place / EWB No.&Dt. / Doc No.&Dt. / Assess Val.
    # / Tax Val. / HSN Code / HSN Desc. / Latest Vehicle No. -- twelve columns, no unit count
    # anywhere). So "how many" below means number of inward EWB documents for that commodity,
    # NOT physical quantity/units -- the closest genuine proxy this data supports, not a
    # fabricated unit count.
    commodity = defaultdict(lambda: [0, 0.0, set()])   # doc count, total assess value, HSN codes seen
    for x in (ctx.get("ewb_in_rows") or []):
        if x.get("from_gstin") and x.get("from_gstin") != self_gstin and x.get("to_gstin") == self_gstin:
            desc = (x.get("hsn_desc") or "").strip().upper()
            if not desc:
                continue
            c = commodity[desc]
            c[0] += 1; c[1] += _n(x.get("assess")); c[2].add(str(x.get("hsn") or "").strip())
    commodity_rows = [[desc, ", ".join(sorted(v[2])), v[0], round(v[1], 2)]
                       for desc, v in commodity.items()]
    commodity_rows.sort(key=lambda r: -r[2])

    extra_tables = [
        dict(title=f"HSN-wise Yearly Comparison -- Inward (E-Way Bill) vs Outward (GSTR-1) "
                   f"({len(hsn_compare_rows)} HSN heading(s))",
             subtitle="4-digit HSN heading level, whole FY. Inward is genuine third-party e-way "
                      "bill movement only (self-to-self branch/stock transfers excluded). "
                      "'Net' = Outward taxable minus Inward assessable -- a large positive value "
                      "means far more was sold under that heading than physically received "
                      "inward under it that year (worth checking whether that heading is also "
                      "manufactured/processed from a DIFFERENT inward HSN, not necessarily an "
                      "anomaly by itself).",
             header=["HSN Heading (4-digit)", "Description", "Inward Assessable Value (EWB)",
                    "Inward Tax Value (EWB)", "Inward EWB Count", "Outward Taxable Value (GSTR-1)",
                    "Outward Tax (GSTR-1)", "Net (Outward - Inward)"],
             widths=[16, 40, 22, 18, 14, 22, 18, 20], rows=hsn_compare_rows,
             empty_note="No inward EWB or outward GSTR-1 HSN data available."),
        dict(title=f"Commodities Purchased This FY -- from Inward E-Way Bill HSN Description "
                   f"({len(commodity_rows)} distinct commodit(y/ies))",
             subtitle="LIMITATION: the inward e-way bill export has no quantity/UQC column at "
                      "all -- 'Inward EWB Count' below is the number of inward e-way bill "
                      "documents naming that exact commodity description, NOT a physical unit "
                      "count. Treat it as a document-volume proxy, not quantity purchased.",
             header=["Commodity (HSN Description, as printed on the EWB)", "HSN Code(s) seen",
                    "Inward EWB Count (documents, NOT quantity)", "Total Assessable Value (Rs)"],
             widths=[46, 24, 26, 22], rows=commodity_rows,
             empty_note="No inward EWB HSN description data available."),
    ]

    return dict(header=["Month", "Taxable purchases (2B, invoice level)", "2B invoices",
                        "Taxable sales (GSTR-1 net of CN)", "GSTR-1 invoices",
                        "Monthly difference", "Cumulative difference", "Note"],
                widths=[10, 30, 12, 30, 15, 20, 22, 60], rows=rows,
                findings=findings, notes=notes, extra_tables=extra_tables)


def _mom_outlier_check(values_by_month, months, threshold, floor=0.0):
    """Month-over-month outlier detector shared by every numeric column in
    the ITC Roll-Forward sheet (Stage 2, point 5). Compares each month
    against the PRIOR month in `months`' own order (skipping the first month,
    which has no prior). A prior value of (near) zero is reported separately
    as 'new activity this month' rather than run through a ratio (a 0 -> X
    change is an infinite/undefined %, which is meaningless as a flag).
    `floor`: skip a comparison entirely when BOTH values are below this --
    a large RATIO on two trivially small rupee amounts (rounding/interest
    adjustments of a few hundred rupees) isn't a meaningful outlier, just
    noise (confirmed against real data -- see gst_config.ITC_ROLLFORWARD_MOM_FLOOR).
    Returns a list of dicts: month, prior, this, kind ('ratio'|'new'),
    ratio (float or None)."""
    out = []
    prev = None
    for m in months:
        v = _n(values_by_month.get(m))
        if prev is not None:
            pv = _n(prev)
            if max(abs(pv), abs(v)) < floor:
                prev = v
                continue
            if abs(pv) <= TOL:
                if abs(v) > TOL:
                    out.append(dict(month=m, prior=pv, this=v, kind="new", ratio=None))
            else:
                ratio = abs(v) / abs(pv)
                if ratio >= threshold or ratio <= (1.0 / threshold):
                    out.append(dict(month=m, prior=pv, this=v, kind="ratio", ratio=ratio))
        prev = v
    return out


def build_itc_annual_summary(ctx):
    """NEW SHEET -- 'ITC Annual Summary'. FY-level ITC lifecycle (Available ->
    Claimed -> Reversed -> Reclaimed) with a month-wise breakdown above the FY
    row, matching this workbook's existing Month + FY TOTAL convention.

    Explicitly scoped to THIS FY only per instruction -- no multi-year aging,
    no running multi-year pool, no cross-file linking.

    UPDATED (per instruction): now shows ITC Available as per 2A AND 2B in separate columns
    (previously 2B only). The 'carried forward from last FY' concept is now shown TWO ways,
    side by side, since both are genuinely useful and answer slightly different questions:
      - INFERRED (computed): per the exact formula given -- (2B Available - 4B(1) - 4B(2)) is
        what THIS FY's own 2B genuinely supports after reversals; if Current-FY-Claimed exceeds
        that, the excess is inferred to be carry-forward from an earlier period. Computed every
        month AND at FY level, from data already in this sheet -- no new source needed.
      - GSTR-9 Table 13 (authoritative, filed): the taxpayer's own annual return figure for
        'ITC availed for the previous financial year'. FY-level only (no month split in the
        source itself), and only available when GSTR-9 is supplied as Excel (documented
        tool-wide limitation -- PDF isn't parsed). Kept alongside the inferred figure, not
        replaced by it: where they diverge materially, that gap itself is worth a look (either
        the taxpayer's own return has a different reclaim pattern than a same-FY-only formula
        can infer, or there's a genuine discrepancy worth raising).

    ALSO NEW: Table 8A (government-computed, from GSTR-2A, auto-populated into GSTR-9) is now
    cross-referenced as a THIRD independent FY-level ITC-available figure alongside 2A and 2B --
    see F1a. Table 8A's own data additionally carries an ITC=No reason breakdown (e.g. "POS and
    supplier state are same but recipient state is different"), the SAME category of reason this
    session's fix found material amounts under -- a useful independent confirmation source.

    Sources every other figure from data this tool ALREADY computes elsewhere
    (ctx['g3b_by_month']/['g3b_extra_by_month'] for Table 4, ctx['twob_by_month']
    for invoice-level 2B, ctx['r2a_data'] for invoice-level 2A -- the SAME fields
    'ITC 3B vs 2B' (F7) already reads for 2B, so the numbers agree with that sheet
    by construction, not by a second, possibly-diverging computation) plus the
    Cash/Credit Ledger inputs this tool already loads."""
    rows, findings = [], []
    T = defaultdict(float)
    any2a = any2b = False
    r2a = ctx.get("r2a_data") or {}
    r2a_b2b_by_month = r2a.get("b2b", {}) if r2a.get("available") else {}
    for m in ctx["months"]:
        g3b = ctx["g3b_by_month"].get(m, {})
        ex = ctx["g3b_extra_by_month"].get(m, {})
        claimed = (sum(_n(x) for x in (ex.get("A1") or [])[:4]) + sum(_n(x) for x in (ex.get("A2") or [])[:4])
                   + sum(_n(x) for x in (g3b.get("4A3") or [])[:4]) + sum(_n(x) for x in (ex.get("A4") or [])[:4])
                   + sum(_n(x) for x in (g3b.get("4A5") or [])[:4]))
        b1 = sum(_n(x) for x in (g3b.get("4B1") or [])[:4])
        b2 = sum(_n(x) for x in (g3b.get("4B2") or [])[:4])
        two = ctx["twob_by_month"].get(m, {})
        avail_2b = (two["itc"] + two["cn_itc"]) if two.get("available") else None
        if two.get("available"):
            any2b = True
        # NEW (per explicit instruction): two reference figures alongside the PRIMARY (Yes +
        # Unconfirmed) figure above -- computed from the SAME per-row classification
        # read_2b_invoice_level already does (not a second, separately-derived calculation),
        # so these three numbers are guaranteed internally consistent by construction:
        #   avail_2b_yes_only        = excludes unconfirmed too (strictest)
        #   avail_2b_grand_total     = Yes + No + Unconfirmed, no exclusion at all (loosest)
        # Identity that must hold every month (verified for all 12 months of the real taxpayer
        # this was built against before shipping): avail_2b_yes_only + no + unconfirmed ==
        # avail_2b_grand_total, and avail_2b == avail_2b_grand_total - no.
        avail_2b_yes_only = avail_2b_grand_total = None
        if two.get("available"):
            avail_2b_yes_only = two["itc_yes_only"] + two["cn_itc_yes_only"]
            avail_2b_grand_total = two["itc_grand_total"] + two["cn_itc_grand_total"]
            T["avail_2b_yes_only"] += avail_2b_yes_only
            T["avail_2b_grand_total"] += avail_2b_grand_total
        avail_2a = None
        if m in r2a_b2b_by_month:
            any2a = True
            avail_2a = sum(_n(x.get("igst")) + _n(x.get("cgst")) + _n(x.get("sgst")) + _n(x.get("cess"))
                           for x in r2a_b2b_by_month[m])
        # NEW (per instruction, exact formula): (2B Available - 4B(1) - 4B(2)) is what this FY's
        # own 2B genuinely supports after reversals; if Claimed exceeds that, the excess is
        # inferred carry-forward.
        inferred_cf = None
        if avail_2b is not None:
            net_after_reversal = avail_2b - b1 - b2
            inferred_cf = max(0.0, claimed - net_after_reversal)
        rows.append([m, avail_2a, avail_2b, b1, b2, claimed, inferred_cf, None,
                     avail_2b_yes_only, avail_2b_grand_total])
        if avail_2a is not None:
            T["avail_2a"] += avail_2a
        if avail_2b is not None:
            T["avail_2b"] += avail_2b
        T["claimed"] += claimed; T["b1"] += b1; T["b2"] += b2

    # GSTR-9 Table 13 -- FY-level only, no month breakdown in the source itself. Reuses
    # ctx['gstr9'] (already parsed once for R13/R14 upstream) rather than re-reading the file.
    g9 = ctx.get("gstr9") or {}
    carry_fwd_gstr9 = None
    if g9.get("table13_itc_igst") is not None:
        carry_fwd_gstr9 = (_n(g9.get("table13_itc_cgst")) + _n(g9.get("table13_itc_sgst"))
                           + _n(g9.get("table13_itc_igst")) + _n(g9.get("table13_itc_cess")))
    fy_net_after_reversal = (T["avail_2b"] - T["b1"] - T["b2"]) if any2b else None
    fy_inferred_cf = max(0.0, T["claimed"] - fy_net_after_reversal) if fy_net_after_reversal is not None else None
    rows.append(["FY TOTAL", (T["avail_2a"] if any2a else None), (T["avail_2b"] if any2b else None),
                 T["b1"], T["b2"], T["claimed"], fy_inferred_cf, carry_fwd_gstr9,
                 (T["avail_2b_yes_only"] if any2b else None), (T["avail_2b_grand_total"] if any2b else None)])

    findings.append(Finding(
        "F1", "ITC lifecycle -- Available (2A/2B), Claimed, Reversed (FY)", "INFO",
        f"FY ITC available per GSTR-2A (invoice level) {_f(T['avail_2a']) if any2a else 'not available'}; "
        f"per GSTR-2B (invoice level) {_f(T['avail_2b']) if any2b else 'not available'}; "
        f"claimed under Table 4A (FULL -- 4A(1) imports of goods + 4A(2) imports of services + "
        f"4A(3) RCM + 4A(4) ISD + 4A(5) all-other) {_f(T['claimed'])}; "
        f"reversed under 4B(1) {_f(T['b1'])}, under 4B(2) {_f(T['b2'])}. "
        f"CLARIFICATION on a genuine, deliberate difference (not an error, but confirmed confusing "
        f"across sheets so stated explicitly here): the 'ITC 3B vs 2B' sheet's own 'claimed' figure "
        f"will be SMALLER than this one -- that sheet deliberately scopes to 4A(5) 'All Other ITC' "
        f"ONLY, because it compares claimed ITC against GSTR-2B's INVOICE-level data, and imports/ "
        f"RCM/ISD don't originate from any supplier's 2B invoice at all (so there is nothing in 2B "
        f"to check those three sub-heads against). The 2B AVAILABLE figure, 4B(1), and 4B(2) here "
        f"ARE the same as 'ITC Roll-Forward 4A-4B-4C' and 'ITC 3B vs 2B' (sourced identically, not "
        f"independently re-derived) -- only the CLAIMED/4A figure differs between this sheet (full "
        f"Table 4A) and 'ITC 3B vs 2B' (4A(5) only), and now both sheets say so in their own text.",
        dict(available_2a_fy=T["avail_2a"] if any2a else None, available_2b_fy=T["avail_2b"] if any2b else None,
             claimed_fy_full_4a=T["claimed"], reversed_4b1_fy=T["b1"], reversed_4b2_fy=T["b2"])))

    findings.append(Finding(
        "F2", "ITC carried forward from last FY -- INFERRED (computed, this FY's own data only)",
        "INFO" if fy_inferred_cf else "PASS",
        f"Formula (as specified): (2B Available - 4B(1) - 4B(2)) is what this FY's own GSTR-2B "
        f"genuinely supports after reversals = {_f(fy_net_after_reversal) if fy_net_after_reversal is not None else 'n/a'}. "
        f"FY Claimed (full Table 4A) = {_f(T['claimed'])}. Where Claimed exceeds that net-available "
        f"figure, the excess is inferred to be carry-forward from an earlier period: "
        f"{_f(fy_inferred_cf) if fy_inferred_cf is not None else 'n/a'}. Computed every month in the "
        f"table above too, not just at FY level. LIMITATION, stated plainly: this is a same-FY-only "
        f"inference (excess claimed over this year's own net-available), not a proven link to any "
        f"specific prior-year invoice -- it will move with normal month-to-month timing noise (e.g. "
        f"a supplier filing late) even with no real carry-forward at all. Compare against F3 below "
        f"(the taxpayer's own filed GSTR-9 Table 13 figure, where available) -- a large, persistent "
        f"gap between the two is worth a closer look; broad agreement is a good consistency signal.",
        dict(fy_net_after_reversal=fy_net_after_reversal, fy_claimed=T["claimed"], inferred_carry_forward=fy_inferred_cf)))
    findings.append(Finding(
        "F2a", "Reclaim tracking -- this FY's own 4B(2) reversals", "INFO",
        "Left blank rather than estimated: this tool has no transaction-level tag distinguishing "
        "'this is a reclaim of an earlier 4B(2) reversal' from any other 4B(2)/4D entry -- a "
        "month's 4B(2) figure is a single total, not itemised by what it relates to. Populating "
        "this column would mean guessing which part of a later month's ITC is a reclaim, which "
        "this tool does not do. If your working papers track this separately, the correct FY "
        "figure can be entered here manually.", {}))
    if carry_fwd_gstr9 is not None:
        findings.append(Finding(
            "F3", "ITC carried forward from last FY (GSTR-9 Table 13, AUTHORITATIVE/filed)", "INFO",
            f"{_f(carry_fwd_gstr9)} (CGST {_f(_n(g9.get('table13_itc_cgst')))} + SGST "
            f"{_f(_n(g9.get('table13_itc_sgst')))} + IGST {_f(_n(g9.get('table13_itc_igst')))} + "
            f"Cess {_f(_n(g9.get('table13_itc_cess')))}) -- the taxpayer's own filed GSTR-9 Table 13 "
            f"'ITC availed for the previous financial year' figure: ITC pertaining to PRIOR-FY "
            f"invoices/debit notes but claimed within THIS FY's own returns (the Section 16(4) "
            f"carry-forward window). This is an ANNUAL return figure with no month-by-month split "
            f"in the source itself, so it appears on the FY TOTAL row only. Compare against F2's "
            f"INFERRED figure above -- vs the computed estimate, difference "
            f"{_f(carry_fwd_gstr9 - (fy_inferred_cf or 0.0))}. For reference, GSTR-9 "
            f"Table 12 (reversal of ITC availed during the previous FY) shows "
            f"{_f(_n(g9.get('table12_itc_reversed_cgst')) + _n(g9.get('table12_itc_reversed_sgst')) + _n(g9.get('table12_itc_reversed_igst')) + _n(g9.get('table12_itc_reversed_cess')))}.",
            dict(carry_forward_fy_gstr9=carry_fwd_gstr9)))
    else:
        findings.append(Finding(
            "F3", "ITC carried forward from last FY (GSTR-9 Table 13, AUTHORITATIVE/filed)", "SKIPPED",
            "Not available: GSTR-9 was not supplied as Excel for this taxpayer/FY (this tool "
            "does not parse a PDF GSTR-9 -- see the classify step), or its Part V (Items 10-14) "
            "sheet/Table 13 row could not be located. Left blank rather than estimated -- use "
            "F2's INFERRED (computed) figure above instead, with its own stated limitation.", {}))

    # ---- NEW (per instruction): Table 8A cross-check + usage suggestions ----
    t8a = ctx.get("table8a") or {}
    if t8a.get("available"):
        t8a_total = _n((t8a.get("totals") or {}).get("total"))
        diff_2a = t8a_total - T["avail_2a"] if any2a else None
        diff_2b = t8a_total - T["avail_2b"] if any2b else None
        no_reasons = (t8a.get("totals") or {}).get("no_reason_breakdown") or {}
        findings.append(Finding(
            "F1a", "ITC Available cross-check -- Table 8A (government-computed, from 2A, "
                   "auto-populated into GSTR-9)", "INFO",
            f"Table 8A FY total: {_f(t8a_total)}. Vs this sheet's 2A figure: "
            f"{('difference ' + _f(diff_2a)) if diff_2a is not None else 'n/a (2A not available)'}. "
            f"Vs this sheet's 2B figure: {('difference ' + _f(diff_2b)) if diff_2b is not None else 'n/a (2B not available)'}. "
            f"A third, independently-computed FY-level figure -- broad agreement across all three "
            f"(2A, 2B, Table 8A) is a strong consistency signal; a persistent gap against Table 8A "
            f"specifically points at the 2A-vs-8A reconciliation step itself (filing-status/timing "
            f"differences between when 2A shows an invoice and when 8A auto-populates it into the "
            f"annual return) rather than this tool's own computation. Table 8A ALSO carries its own "
            f"ITC-No reason breakdown -- {len(no_reasons)} distinct reason(s) recorded" +
            (f", e.g. {list(no_reasons.items())[0][0]!r} ({list(no_reasons.items())[0][1]} invoice(s))"
             if no_reasons else "") + ". "
            "SUGGESTIONS for further Table 8A use, not yet built: (1) an invoice-level Table 8A vs "
            "2B match, the same way this tool already does 2A-vs-2B on the 'GSTR-2A vs 2B Invoice "
            "Detail' sheet -- would catch invoices GSTN auto-populated into 8A but that never made "
            "it into 2B, or vice versa; (2) a month-wise Table 8A breakdown if the source ever "
            "carries a period column (current export is FY-level only, confirmed against this "
            "taxpayer's real file); (3) cross-tabulating Table 8A's own ITC-No reasons against this "
            "sheet's 2B-side ITC-No exclusion (this session's fix) to confirm both sources agree on "
            "which specific invoices are blocked, not just the aggregate total.",
            dict(table8a_total=t8a_total, diff_vs_2a=diff_2a, diff_vs_2b=diff_2b)))
    else:
        findings.append(Finding(
            "F1a", "ITC Available cross-check -- Table 8A", "SKIPPED",
            t8a.get("reason") or "Table 8A not supplied for this taxpayer/FY.", {}))

    # ---- Credit Ledger tie-out (FY level) ----
    credit = ctx["annual_data"].get("credit") or {}
    credit_txns = credit.get("transactions") or []
    cl_month = credit.get("monthly_by_tax_period", {})
    closing_credit_ledger = None
    if credit_txns:
        opening = _n((credit.get("opening") or {}).get("bal_total"))
        if ctx["months"]:
            first_month = ctx["months"][0]
            for t in credit_txns:
                if dept._tax_period_key(t.get("tax_period")) == first_month:
                    break
                opening = _n(t.get("bal_total"))
        fy_credited = sum(_n(cl_month.get(m, {}).get("credited")) for m in ctx["months"])
        fy_debited = sum(_n(cl_month.get(m, {}).get("debited")) for m in ctx["months"])
        computed_closing = opening + fy_credited - fy_debited
        actual_closing = _n(credit_txns[-1].get("bal_total"))
        closing_credit_ledger = actual_closing
        gap = computed_closing - actual_closing
        findings.append(Finding(
            "F4", "Credit Ledger tie-out (FY) -- Opening + Credited - Debited = Closing",
            "FLAG" if abs(gap) > TOL else "PASS",
            f"Opening (corrected for any pre-FY carry-forward transaction sitting in the same "
            f"ledger file -- see the ITC Roll-Forward sheet's F5 for why that correction matters) "
            f"{_f(opening)} + FY credited {_f(fy_credited)} - FY debited {_f(fy_debited)} = "
            f"computed closing {_f(computed_closing)}, against the ledger's own actual closing "
            f"balance {_f(actual_closing)}; gap {_f(gap)}.",
            dict(opening=opening, credited=fy_credited, debited=fy_debited,
                 computed_closing=computed_closing, actual_closing=actual_closing, gap=gap)))
    else:
        findings.append(Finding("F4", "Credit Ledger tie-out (FY)", "SKIPPED",
                                "Electronic Credit Ledger not supplied.", {}))

    # ---- Cash Ledger sanity check (FY level, ledger's own figures only) ----
    cash = ctx["annual_data"].get("cash") or {}
    cash_txns = cash.get("transactions") or []
    if cash_txns:
        cash_opening = _n((cash.get("opening") or {}).get("balance_total"))
        fy_deposited = sum(_n(t.get("total")) for t in cash_txns if str(t.get("ttype", "")).lower() == "credit")
        fy_utilized = sum(_n(t.get("total")) for t in cash_txns if str(t.get("ttype", "")).lower() == "debit")
        computed_cash_closing = cash_opening + fy_deposited - fy_utilized
        actual_cash_closing = _n(cash_txns[-1].get("balance_total"))
        cash_gap = computed_cash_closing - actual_cash_closing
        findings.append(Finding(
            "F5", "Cash Ledger sanity check (FY) -- Opening + Deposited - Utilized = Closing",
            "FLAG" if abs(cash_gap) > TOL else "PASS",
            f"Using only the Cash Ledger's own figures: opening {_f(cash_opening)} + deposited "
            f"{_f(fy_deposited)} - utilized {_f(fy_utilized)} = computed closing "
            f"{_f(computed_cash_closing)}, against the ledger's own actual closing "
            f"{_f(actual_cash_closing)}; gap {_f(cash_gap)}. This is an internal ledger check "
            f"only (no cross-tool computation) -- cash isn't this tool's main focus.",
            dict(opening=cash_opening, deposited=fy_deposited, utilized=fy_utilized,
                 computed_closing=computed_cash_closing, actual_closing=actual_cash_closing,
                 gap=cash_gap)))
    else:
        findings.append(Finding("F5", "Cash Ledger sanity check (FY)", "SKIPPED",
                                "Electronic Cash Ledger not supplied.", {}))

    closing_pool = T["b2"]  # = FY 4B(2) reversed minus this-FY reclaim-of-own-4B2, which is
                             # unknown (see F2) -- so this figure is the full 4B(2) total, labelled
                             # honestly as such rather than presented as a precise net figure.
    closing_section = dict(
        title="Closing Balances (For Next Year Handoff)",
        subtitle="Manual carry-forward only -- no automatic file-linking or multi-year logic is "
                 "built here, by design (out of scope for this FY's tool). Copy these into next "
                 "year's tool as that year's OPENING figures.",
        header=["Item", "Value", "Note"],
        widths=[40, 20, 70],
        rows=[
            ["Closing Credit Ledger balance (FY end)",
             closing_credit_ledger, "From the actual Credit Ledger file's own last transaction."],
            ["Closing Unreclaimed 4B(2) Pool", closing_pool,
             "This FY's 4B(2) reversed minus this FY's reclaim of its OWN 4B(2) -- but that "
             "reclaim figure isn't trackable in this tool (see F2), so this equals the FULL 4B(2) "
             "total, not a true net figure. Treat as an upper bound, not a precise pool balance, "
             "until reclaim tagging exists."],
        ])

    return dict(header=["Month", "ITC Available (2A)", "ITC Available (2B) -- PRIMARY (Yes+Unconfirmed, excludes confirmed No)",
                        "Reversed 4B(1)",
                        "Reversed 4B(2)", "ITC Availed (Current FY, FULL Table 4A -- Import+RCM+ISD+All-Other)",
                        "ITC Carried Forward -- INFERRED (Claimed minus [2B Available - 4B1 - 4B2], floored at 0)",
                        "ITC Carried Forward -- GSTR-9 Table 13 (AUTHORITATIVE/filed, FY TOTAL row only)",
                        "REFERENCE: 2B if EXCLUDING unconfirmed too (Yes only, strictest)",
                        "REFERENCE: 2B if INCLUDING everything (Yes+No+Unconfirmed, no exclusion at all)"],
                widths=[10, 20, 34, 16, 16, 34, 34, 36, 34, 38], rows=rows, findings=findings,
                notes=["Column order: 2A available, 2B available (PRIMARY), 4B(1) reversed, 4B(2) "
                       "reversed, current-FY availed, then TWO carry-forward columns side by side "
                       "(INFERRED and GSTR-9 Table 13), then the TWO REFERENCE 2B figures requested "
                       "explicitly, at the end of each row.",
                       "PRIMARY 2B definition (per explicit instruction): every GSTR-2B B2B/CDNR row "
                       "is classified YES / NO / UNCONFIRMED by its own 'ITC Availability' field. "
                       "The PRIMARY figure = YES + UNCONFIRMED (excludes only a CONFIRMED No -- an "
                       "unconfirmed row is not assumed ineligible, since the portal never said it "
                       "was). The two REFERENCE columns at the end show the same data two other "
                       "ways: 'excluding unconfirmed too' (Yes only -- the strictest, most "
                       "conservative reading) and 'including everything' (Yes+No+Unconfirmed -- the "
                       "loosest, a plain grand total with no eligibility filter at all). All three "
                       "are computed from the SAME per-row classification, not three separate "
                       "calculations -- verified for every month: (Yes-only column) + (No, see the "
                       "'GSTR-2B ITC No & Unconfirmed -- Invoice Detail' sheet) + (Unconfirmed, same "
                       "sheet) equals the 'including everything' column exactly, and the PRIMARY "
                       "column equals 'including everything' minus No, exactly -- both checked "
                       "independently for all 12 months before this was shipped.",
                       "A REFERENCE column can be NEGATIVE for a given month -- this is correct, not "
                       "an error: GSTR-2B credit notes reduce ITC (they carry a negative sign in "
                       "this tool's own convention, matching how a credit note reduces the buyer's "
                       "ITC), so a month where unconfirmed-status CREDIT notes outweigh unconfirmed-"
                       "status DEBIT notes will show a negative 'unconfirmed' contribution, and that "
                       "can pull the 'including everything' column below the PRIMARY figure for that "
                       "month specifically. The FY TOTAL row is the figure to rely on for the year.",
                       "Full invoice-level detail for every No and Unconfirmed row (GSTIN, invoice/"
                       "note number, date, taxable, tax, and the portal's own stated reason where "
                       "given) is on the 'GSTR-2B ITC No & Unconfirmed -- Invoice Detail' sheet.",
                       "IMPORTANT, explained per explicit question: the 'ITC Availed' column is "
                       "the FULL Table 4A (4A(1) imports of goods + 4A(2) imports of services + "
                       "4A(3) RCM + 4A(4) ISD + 4A(5) all-other) -- it will legitimately be LARGER "
                       "than the 'ITC 3B vs 2B' sheet's own 'claimed' figure, which is 4A(5) ONLY "
                       "(deliberately -- imports/RCM/ISD don't originate from any supplier's 2B "
                       "invoice at all, so that sheet, which exists specifically to compare "
                       "claimed-vs-2B-invoice-available, only covers the ONE Table 4A sub-head "
                       "that 2B invoices can actually support). Both figures are correct; they "
                       "answer different questions -- one is total ITC claimed this FY across every "
                       "source, the other is the subset checkable against supplier invoices."],
                extra_tables=[closing_section])


def build_2b_no_unconfirmed_detail(ctx):
    """NEW SHEET (per explicit instruction): complete invoice-level detail of every GSTR-2B
    B2B/CDNR row whose ITC Availability is confirmed NO (portal-flagged ineligible) or
    UNCONFIRMED (status could not be read for that row -- see read_2b_invoice_level's own
    docstring for why this happens on some CDNR rows), across the whole FY. Two separate
    tables, since they are different categories -- No is a portal statement, Unconfirmed is
    this tool's own inability to read that one field (the tax amounts on those rows ARE
    reliable, only the status is unknown).

    ACCURACY CHECK built into this sheet itself (per explicit 'no calculation mistakes' rule):
    the FY totals shown here for No and Unconfirmed are cross-footed against the exact same
    per-row classification 'ITC Annual Summary' uses for its own reference columns -- both are
    read from the SAME underlying read_2b_invoice_level() output, so they cannot diverge by
    construction, and a total row is provided for both tables so the two can be checked against
    each other directly, cell for cell, without recomputation."""
    no_rows, unconf_rows = [], []
    T_no = T_unconf = 0.0
    for m in ctx["months"]:
        two = ctx["twob_by_month"].get(m, {})
        if not two.get("available"):
            continue
        for x in two.get("no_confirmed_rows", []):
            ref = x.get("invno") if x.get("source") == "B2B" else x.get("note")
            line_tax = _n(x.get("igst")) + _n(x.get("cgst")) + _n(x.get("sgst")) + _n(x.get("cess"))
            if x.get("source") == "CDNR" and str(x.get("ntype", "")).strip().upper().startswith("C"):
                line_tax = -line_tax   # credit note reduces ITC -- same sign convention as the rest of this tool
            T_no += line_tax
            no_rows.append([m, x.get("source"), x.get("gstin"), x.get("supplier"), ref,
                            x.get("date"), round(_n(x.get("taxable")), 2), round(line_tax, 2),
                            x.get("itc_avail_reason") or ""])
        for x in two.get("unconfirmed_rows", []):
            ref = x.get("invno") if x.get("source") == "B2B" else x.get("note")
            line_tax = _n(x.get("igst")) + _n(x.get("cgst")) + _n(x.get("sgst")) + _n(x.get("cess"))
            if x.get("source") == "CDNR" and str(x.get("ntype", "")).strip().upper().startswith("C"):
                line_tax = -line_tax
            T_unconf += line_tax
            unconf_rows.append([m, x.get("source"), x.get("gstin"), x.get("supplier"), ref,
                                x.get("date"), round(_n(x.get("taxable")), 2), round(line_tax, 2),
                                str(x.get("itc_avail") or "") or "(blank)"])
    no_rows.sort(key=lambda r: (str(r[0]), str(r[1]), str(r[2])))
    unconf_rows.sort(key=lambda r: (str(r[0]), str(r[1]), str(r[2])))
    if no_rows:
        no_rows.append(["TOTAL", "", "", "", "", "", "", round(T_no, 2), ""])
    if unconf_rows:
        unconf_rows.append(["TOTAL", "", "", "", "", "", "", round(T_unconf, 2), ""])

    findings = [Finding(
        "Z1", "GSTR-2B ITC No & Unconfirmed -- FY summary", "INFO",
        f"Confirmed NO (portal-flagged ineligible): {len(no_rows) - (1 if no_rows else 0)} row(s), "
        f"net tax Rs {round(T_no, 2):,.2f}. UNCONFIRMED (status not readable): "
        f"{len(unconf_rows) - (1 if unconf_rows else 0)} row(s), net tax Rs {round(T_unconf, 2):,.2f}. "
        f"These two totals are the SAME figures subtracted/added in 'ITC Annual Summary' to move "
        f"between its PRIMARY, 'Yes only', and 'grand total' columns -- cross-check: "
        f"(ITC Annual Summary's 'Yes only' column, FY TOTAL) + {round(T_no, 2):,.2f} + "
        f"{round(T_unconf, 2):,.2f} should equal (ITC Annual Summary's 'grand total' column, FY "
        f"TOTAL) exactly.",
        dict(no_total=round(T_no, 2), unconfirmed_total=round(T_unconf, 2)))]

    return dict(header=["Month", "Source", "GSTIN", "Supplier", "Invoice/Note No", "Date",
                        "Taxable Value (Rs)", "Tax -- IGST+CGST+SGST+Cess (Rs)", "Portal's Stated Reason"],
                widths=[10, 10, 18, 30, 20, 14, 18, 24, 46], rows=no_rows, findings=findings,
                notes=["Table above: CONFIRMED NO (portal-flagged ineligible) -- 'Portal's Stated "
                       "Reason' column carries whatever the source file itself states. Table below: "
                       "UNCONFIRMED (status could not be read for that row -- 'Portal's Stated "
                       "Reason' column instead shows the raw, unreadable status text as-is, e.g. "
                       "'(blank)' or 'NA', for transparency, not a real reason). Tax column sign "
                       "convention: a credit note's tax is shown NEGATIVE (it reduces ITC), matching "
                       "this tool's convention everywhere else -- a debit note is positive."],
                extra_tables=[
                    dict(title=f"UNCONFIRMED Status -- Complete Detail ({len(unconf_rows) - (1 if unconf_rows else 0)} row(s))",
                         subtitle="ITC Availability status could not be read for these rows (see this "
                                  "sheet's own notes above for why) -- the taxable/tax figures ARE "
                                  "reliable (verified: NoteValue = Taxable+Tax reconciles exactly for "
                                  "a sample of these rows), only the eligibility STATUS is unknown. "
                                  "Currently counted as available (Yes+Unconfirmed) in the PRIMARY "
                                  "ITC Available figure, per explicit instruction.",
                         header=["Month", "Source", "GSTIN", "Supplier", "Invoice/Note No", "Date",
                                "Taxable Value (Rs)", "Tax -- IGST+CGST+SGST+Cess (Rs)", "Raw Status Text"],
                         widths=[10, 10, 18, 30, 20, 14, 18, 24, 20], rows=unconf_rows,
                         empty_note="No rows with unconfirmed ITC Availability status this FY."),
                ])


def build_zero_tax_scan(ctx):
    """NEW SHEET (per instruction): every invoice/movement across GSTR-1, GSTR-2B, Inward EWB,
    and Outward EWB that carries a real value but ZERO tax -- grouped heading-wise by source, one
    table per source, so each can be scanned on its own. A non-trivial value with nil tax is
    either a genuinely nil-rated/exempt/zero-rated supply (expected, common) or a signal worth a
    closer look (wrong rate applied, non-supply movement mis-invoiced as a sale, etc.) -- this
    sheet doesn't try to tell those apart (that needs the underlying document), it surfaces every
    candidate so a reviewer can."""
    ZTOL = 1.0
    g1_zero, g2b_zero, ewb_in_zero, ewb_out_zero = [], [], [], []

    for x in ctx["g1_lines_fy"]:
        if x.get("header_row") and _n(x.get("invval")) > ZTOL and \
           _n(x.get("igst")) + _n(x.get("cgst")) + _n(x.get("sgst")) <= ZTOL:
            g1_zero.append([x.get("month"), x.get("gstin"), x.get("name"), x.get("invno"),
                            _dstr(x.get("invdate")), round(_n(x.get("invval")), 2)])

    for x in ctx["twob_lines_fy"]:
        if _n(x.get("invval")) > ZTOL and \
           _n(x.get("igst")) + _n(x.get("cgst")) + _n(x.get("sgst")) + _n(x.get("cess")) <= ZTOL:
            g2b_zero.append([x.get("month"), x.get("gstin"), x.get("supplier"), x.get("invno"),
                             str(x.get("date")), round(_n(x.get("invval")), 2)])

    self_gstin = ctx.get("self_gstin") or ""
    for x in (ctx.get("ewb_in_rows") or []):
        if x.get("from_gstin") and x.get("from_gstin") != self_gstin and x.get("to_gstin") == self_gstin \
           and _n(x.get("assess")) > ZTOL and _n(x.get("taxval")) <= ZTOL:
            ewb_in_zero.append([x.get("month"), x.get("from_gstin"), x.get("from_name"), x.get("docno"),
                                str(x.get("docdate")), round(_n(x.get("assess")), 2), x.get("hsn")])
    for x in (ctx.get("ewb_out_rows") or []):
        if x.get("to_gstin") and x.get("to_gstin") != self_gstin and x.get("from_gstin") == self_gstin \
           and _n(x.get("assess")) > ZTOL and _n(x.get("taxval")) <= ZTOL:
            ewb_out_zero.append([x.get("month"), x.get("to_gstin"), x.get("to_name"), x.get("docno"),
                                 str(x.get("docdate")), round(_n(x.get("assess")), 2), x.get("hsn")])

    for lst in (g1_zero, g2b_zero, ewb_in_zero, ewb_out_zero):
        lst.sort(key=lambda r: (str(r[0] or ""), str(r[1] or "")))

    findings = [Finding(
        "Z1", "Zero-tax invoices/movements across all four sources", "INFO",
        f"GSTR-1 (outward): {len(g1_zero)}. GSTR-2B (inward): {len(g2b_zero)}. Outward EWB "
        f"(third-party only, self-to-self excluded): {len(ewb_out_zero)}. Inward EWB (third-party "
        f"only): {len(ewb_in_zero)}. A non-trivial value with nil tax is commonly a genuine "
        f"nil-rated/exempt/zero-rated/non-supply movement -- this sheet lists every candidate for "
        f"manual review, it does not itself determine which are genuine and which aren't.",
        dict(gstr1=len(g1_zero), gstr2b=len(g2b_zero), ewb_out=len(ewb_out_zero), ewb_in=len(ewb_in_zero)))]

    # NEW: TOTAL row per table (value column is index 5 in every one of these 4 tables).
    def _with_total(lst, label):
        if not lst:
            return lst
        return lst + [[f"TOTAL ({label})", "", "", "", "", round(sum(_n(r[5]) for r in lst), 2)] +
                      ([""] if len(lst[0]) > 6 else [])]

    g1_zero_t = _with_total(g1_zero, "GSTR-1")
    g2b_zero_t = _with_total(g2b_zero, "GSTR-2B")
    ewb_out_zero_t = _with_total(ewb_out_zero, "Outward EWB")
    ewb_in_zero_t = _with_total(ewb_in_zero, "Inward EWB")

    return dict(header=["Month", "GSTIN", "Name", "Invoice No", "Date", "Invoice Value (Rs)"],
                widths=[10, 18, 32, 20, 14, 18], rows=g1_zero_t, findings=findings,
                notes=["GSTR-1 (outward, this table) shown above; GSTR-2B, Outward EWB, and Inward "
                       "EWB are in their own tables below, each headed by source. Self-to-self "
                       "e-way bills (branch/stock transfers) are excluded from both EWB tables -- "
                       "not a purchase or sale, not the subject of this check."],
                extra_tables=[
                    dict(title=f"Zero-Tax -- GSTR-2B (Inward) ({len(g2b_zero)})",
                         subtitle="Every GSTR-2B B2B invoice with a real invoice value but nil "
                                  "IGST+CGST+SGST+Cess.",
                         header=["Month", "GSTIN", "Supplier", "Invoice No", "Date", "Invoice Value (Rs)"],
                         widths=[10, 18, 32, 20, 14, 18], rows=g2b_zero_t, empty_note="None."),
                    dict(title=f"Zero-Tax -- Outward E-Way Bill ({len(ewb_out_zero)})",
                         subtitle="Genuine third-party outward movements (self-to-self excluded) "
                                  "with a real assessable value but nil Tax Val.",
                         header=["Month", "GSTIN", "Name", "Doc No", "Date", "Assessable Value (Rs)", "HSN"],
                         widths=[10, 18, 32, 20, 14, 20, 12], rows=ewb_out_zero_t, empty_note="None."),
                    dict(title=f"Zero-Tax -- Inward E-Way Bill ({len(ewb_in_zero)})",
                         subtitle="Genuine third-party inward movements (self-to-self excluded) "
                                  "with a real assessable value but nil Tax Val.",
                         header=["Month", "GSTIN", "Name", "Doc No", "Date", "Assessable Value (Rs)", "HSN"],
                         widths=[10, 18, 32, 20, 14, 20, 12], rows=ewb_in_zero_t, empty_note="None."),
                ])


def build_itc_rollforward(ctx):
    """F2 + F3 -- ITC availed by head, reversed split 4B(1) vs 4B(2), net,
    ineligible, and the Electronic Credit Ledger's own credits/debits/balance.

    Stage 2 additions (all purely additive -- every column/row/finding above
    this comment is completely unchanged in position, value, or styling):
      F4  cross-link vs the 'Potential Blocked Credits' sheet's flagged ITC,
          per month and FY total, against the taxpayer's own 4D(1) reversal.
      F5  Electronic Credit Ledger running-balance tie-out (opening +
          credited - debited = actual reported closing), month over month.
      F6  note on why an 'as filed vs as computed' comparison isn't built
          (see this finding's own text for why).
      F7  month-over-month outlier flags across every numeric 4A-4D column.
    """
    rows, findings = [], []
    T = defaultdict(float)
    cl_month = (ctx["annual_data"].get("credit") or {}).get("monthly_by_tax_period", {})
    credit = ctx["annual_data"].get("credit") or {}
    credit_txns = credit.get("transactions") or []
    credit_opening = _n((credit.get("opening") or {}).get("bal_total"))

    # index credit-ledger transactions by their OWN tax-period tag, keeping the LAST one seen
    # per period (transactions are in the file's own chronological order) -- gives the actual
    # reported closing balance right after that period's own entries were posted.
    _actual_closing_by_period = {}
    for t in credit_txns:
        pk = dept._tax_period_key(t.get("tax_period"))
        if pk:
            _actual_closing_by_period[pk] = _n(t.get("bal_total"))

    # blocked-credit cross-link source (Stage 1's sheet -- read-only, see build_context())
    blocked_rows = ctx.get("blocked_credit_rows") or []
    blocked_itc_by_month = defaultdict(float)
    for br in blocked_rows:
        blocked_itc_by_month[br.get("month")] += (
            _n(br.get("igst")) + _n(br.get("cgst")) + _n(br.get("sgst")) + _n(br.get("cess")))

    a1_by_m, a2_by_m, a3_by_m, a4_by_m, a5_by_m = {}, {}, {}, {}, {}
    avail_by_m, b1_by_m, b2_by_m, net_by_m, d1_by_m, d2_by_m = {}, {}, {}, {}, {}, {}

    running_opening = credit_opening
    # Absorb any ledger transaction that occurs BEFORE the first transaction tagged to this FY's
    # own first month -- e.g. a prior-FY return (March of the previous FY) actually paid within
    # this FY's ledger-download window. That transaction isn't part of THIS FY's 4A-4D table, but
    # it DID happen, in the SAME ledger, before this FY's own activity -- so the correct starting
    # balance for month 1's tie-out is the balance AFTER it, not the ledger's raw "Opening
    # Balance" header (which predates even that transaction). Confirmed against real data this
    # matters: without this, a genuine prior-FY carryover produced a persistent, misleading
    # multi-crore tie-out gap on every single month of the FY.
    if ctx["months"] and credit_txns:
        first_month = ctx["months"][0]
        for t in credit_txns:
            if dept._tax_period_key(t.get("tax_period")) == first_month:
                break
            running_opening = _n(t.get("bal_total"))
    running_opening_start = running_opening
    ledger_row_extra = {}  # month -> (computed_closing, actual_closing_or_None, gap_or_None)
    for m in ctx["months"]:
        g3b = ctx["g3b_by_month"].get(m, {})
        ex = ctx["g3b_extra_by_month"].get(m, {})

        def s(key, src=None):
            v = (src or g3b).get(key)
            return sum(_n(x) for x in v[:4]) if isinstance(v, (list, tuple)) else 0.0

        a1, a2, a3 = s("A1", ex), s("A2", ex), s("4A3")
        a4, a5 = s("A4", ex), s("4A5")
        b1, b2 = s("4B1"), s("4B2")
        d1, d2 = s("D1", ex), s("D2", ex)
        avail = a1 + a2 + a3 + a4 + a5
        net = avail - b1 - b2
        led = cl_month.get(m, {})

        # ---- Stage 2, point 2: blocked-credit scan vs 4D(1) ----
        blocked_flagged = blocked_itc_by_month.get(m, 0.0)
        blocked_gap = blocked_flagged - d1

        # ---- Stage 2, point 3: credit-ledger running-balance tie-out ----
        credited_m, debited_m = _n(led.get("credited")), _n(led.get("debited"))
        computed_closing = running_opening + credited_m - debited_m
        actual_closing = _actual_closing_by_period.get(m)
        tie_gap = (computed_closing - actual_closing) if actual_closing is not None else None
        ledger_row_extra[m] = (computed_closing, actual_closing, tie_gap)
        running_opening = computed_closing  # feed forward for next month regardless

        rows.append([m, a1, a2, a3, a4, a5, avail, b1, b2, b1 + b2, net, d1, d2,
                     led.get("credited"), led.get("debited"), None,
                     (blocked_flagged if blocked_rows else None), (blocked_gap if blocked_rows else None),
                     round(computed_closing, 2), (round(actual_closing, 2) if actual_closing is not None else None),
                     (round(tie_gap, 2) if tie_gap is not None else None)])
        for k, v in dict(a1=a1, a2=a2, a3=a3, a4=a4, a5=a5, avail=avail, b1=b1, b2=b2,
                         net=net, d1=d1, d2=d2).items():
            T[k] += v
        T["cl_cr"] += credited_m
        T["cl_dr"] += debited_m
        T["blocked"] += blocked_flagged
        a1_by_m[m], a2_by_m[m], a3_by_m[m], a4_by_m[m], a5_by_m[m] = a1, a2, a3, a4, a5
        avail_by_m[m], b1_by_m[m], b2_by_m[m] = avail, b1, b2
        net_by_m[m], d1_by_m[m], d2_by_m[m] = net, d1, d2

    fy_row_idx = len(rows)  # 0-based index the FY TOTAL row will land on, for highlight_cells
    rows.append(["FY TOTAL", T["a1"], T["a2"], T["a3"], T["a4"], T["a5"], T["avail"],
                 T["b1"], T["b2"], T["b1"] + T["b2"], T["net"], T["d1"], T["d2"],
                 T["cl_cr"], T["cl_dr"], None,
                 (T["blocked"] if blocked_rows else None),
                 ((T["blocked"] - T["d1"]) if blocked_rows else None),
                 None, None, None])

    findings.append(Finding(
        "F2", "Total ITC availed and reversed, split 4B(1) vs 4B(2)", "INFO",
        f"FY ITC availed (Table 4A total) {_f(T['avail'])} -- of which RCM 4A(3) {_f(T['a3'])}, "
        f"all-other 4A(5) {_f(T['a5'])}, imports 4A(1)+4A(2) {_f(T['a1'] + T['a2'])}, "
        f"ISD 4A(4) {_f(T['a4'])}. FY ITC reversed {_f(T['b1'] + T['b2'])} -- "
        f"{_f(T['b1'])} under 4B(1) (Rules 38/42/43 -- permanent reversal) and "
        f"{_f(T['b2'])} under 4B(2) (Others -- ordinarily temporary and reclaimable). "
        f"Net ITC per Table 4C {_f(T['net'])}. Ineligible 4D(1) {_f(T['d1'])}, 4D(2) {_f(T['d2'])}.",
        dict(availed=T["avail"], reversed_4b1=T["b1"], reversed_4b2=T["b2"], net=T["net"])))

    # 4B classification: the whole reversal sitting in 'Others' is worth a look
    tot_rev = T["b1"] + T["b2"]
    if tot_rev > MATERIAL and T["b1"] <= TOL:
        findings.append(Finding(
            "F2a", "Entire ITC reversal booked under 4B(2) 'Others'", "REVIEW",
            f"{_f(tot_rev)} of ITC was reversed during the year and all of it sits in Table 4B(2) "
            f"'Others', with nothing in 4B(1). 4B(2) reversals are ordinarily temporary (Rule 37 "
            f"non-payment, later reclaimed); 4B(1) covers permanent reversals under Rules 38/42/43. "
            f"Where a taxpayer has exempt/non-GST outward supply, Rule 42 reversal belongs in 4B(1) "
            f"-- see F2b. Basis of the classification should be obtained.",
            dict(total_reversal=tot_rev, in_4b1=T["b1"], in_4b2=T["b2"])))

    # Rule 42 gate: exempt turnover with no 4B(1) reversal
    exempt_9c = None
    g9c = ctx.get("gstr9c") or {}
    if g9c.get("available"):
        exempt_9c = _n(g9c.get("exempt_nil_nongst_adjustment"))
    exempt_1 = sum(_n(ctx["g1_by_month"].get(m, {}).get("nil_exempt_taxable")) +
                   _n(ctx["g1_by_month"].get(m, {}).get("nongst_taxable"))
                   for m in ctx["months"])
    if (exempt_9c or 0) > MATERIAL and T["b1"] <= TOL:
        findings.append(Finding(
            "F2b", "Exempt turnover declared but zero Rule 42/43 reversal in 4B(1)", "FLAG",
            f"GSTR-9C Table 7B declares {_f(exempt_9c)} of exempted / nil-rated / non-GST supply "
            f"for the year, but Table 4B(1) of GSTR-3B shows NIL reversal under Rules 42/43 in "
            f"every month. Where common inputs or input services feed both taxable and exempt "
            f"supply, Rule 42 requires proportionate reversal. Either the reversal is short, or "
            f"the 7B figure is not genuinely exempt supply. Note that GSTR-1's own Table 8 "
            f"reports {_f(exempt_1)} of nil/exempt/non-GST supply, which is a separate "
            f"inconsistency in its own right (see the R13 forensic check).",
            dict(exempt_per_9c=exempt_9c, exempt_per_gstr1=exempt_1, reversal_4b1=T["b1"])))

    # credit ledger reconciliation
    if T["cl_cr"] > 0:
        missing, wnote = ledger_window_note(ctx["months"], set(cl_month), "Electronic Credit Ledger")
        cmp_months = [m for m in ctx["months"] if m not in missing]
        net_cmp = 0.0
        for m in cmp_months:
            g3b = ctx["g3b_by_month"].get(m, {})
            ex = ctx["g3b_extra_by_month"].get(m, {})
            av = (sum(_n(x) for x in (ex.get("A1") or [])[:4]) + sum(_n(x) for x in (ex.get("A2") or [])[:4])
                  + sum(_n(x) for x in (g3b.get("4A3") or [])[:4]) + sum(_n(x) for x in (ex.get("A4") or [])[:4])
                  + sum(_n(x) for x in (g3b.get("4A5") or [])[:4]))
            net_cmp += av - sum(_n(x) for x in (g3b.get("4B1") or [])[:4]) \
                          - sum(_n(x) for x in (g3b.get("4B2") or [])[:4])
        cr_cmp = sum(_n(cl_month.get(m, {}).get("credited")) for m in cmp_months)
        d = net_cmp - cr_cmp
        findings.append(Finding(
            "F3", "Net ITC per GSTR-3B vs Electronic Credit Ledger credits",
            "REVIEW" if abs(d) > MATERIAL else "PASS",
            f"Compared over the {len(cmp_months)} month(s) the ledger actually covers: net ITC "
            f"availed per Table 4C {_f(net_cmp)} against {_f(cr_cmp)} credited to the Electronic "
            f"Credit Ledger; difference {_f(d)}. The ledger is credited with the 4C net figure on "
            f"filing, so these should agree closely; a residual gap points to a ledger adjustment "
            f"that is not a return entry (refund debit, Rule 86A block) -- those are itemised on "
            f"the DRC-03 & Ledger Movements sheet. "
            + wnote +
            f" For reference, the unadjusted full-year figures are 4C {_f(T['net'])} against "
            f"{_f(T['cl_cr'])} credited.",
            dict(net_itc_3b_in_window=net_cmp, ledger_credited_in_window=cr_cmp, difference=d,
                 months_excluded=", ".join(missing))))
    else:
        findings.append(Finding("F3", "Net ITC per GSTR-3B vs Electronic Credit Ledger credits",
                                "SKIPPED", "Electronic Credit Ledger not supplied.", {}))

    # ---- Stage 2, point 2 (F4): blocked-credit scan vs 4D(1) ----
    highlight_cells = {}
    BLOCKED_FLAG_COL = 18   # 'Blocked Scan minus 4D(1)' column (1-based)
    TIEOUT_COL = 21         # 'Credit Ledger Tie-Out Gap' column (1-based)
    if blocked_rows:
        gap_fy = T["blocked"] - T["d1"]
        flagged_months = []
        for idx, m in enumerate(ctx["months"]):
            g = blocked_itc_by_month.get(m, 0.0) - d1_by_m.get(m, 0.0)
            if abs(g) > MATERIAL:
                flagged_months.append((m, g))
                highlight_cells[(idx, BLOCKED_FLAG_COL)] = RED
        if abs(gap_fy) > MATERIAL:
            highlight_cells[(fy_row_idx, BLOCKED_FLAG_COL)] = RED
        findings.append(Finding(
            "F4", "Blocked Credit Scan (Trade-name match) vs 4D(1) reversed", "FLAG" if flagged_months else "INFO",
            f"FY ITC flagged by the 'Potential Blocked Credits' sheet (Trade-name keyword match) "
            f"{_f(T['blocked'])}, against {_f(T['d1'])} actually reversed under Table 4D(1) "
            f"(sec 17(5)) -- gap {_f(gap_fy)}. "
            + (f"{len(flagged_months)} month(s) individually exceed the materiality floor: "
               + "; ".join(f"{m} (gap {_f(g)})" for m, g in flagged_months) + ". "
               if flagged_months else "No individual month exceeds the materiality floor. ")
            + "A positive gap means more was flagged as potentially blocked than the taxpayer "
              "self-reversed -- worth a sample review; a negative gap can be normal (the taxpayer's "
              "own 4D(1) may cover items this keyword scan can't see, e.g. blocked ITC from a "
              "supplier whose trade name doesn't match any keyword). The keyword scan itself is "
              "Trade-name-only (see that sheet's own notes for why) -- treat this as a screening "
              "cross-check, not a computed liability.",
            dict(blocked_flagged_fy=T["blocked"], reversed_4d1_fy=T["d1"], gap_fy=gap_fy)))
    else:
        findings.append(Finding("F4", "Blocked Credit Scan (Trade-name match) vs 4D(1) reversed",
                                "SKIPPED", "No blocked-credit master file was supplied/detected this "
                                "run, or it produced zero matches -- see the 'Potential Blocked "
                                "Credits' sheet.", {}))

    # ---- Stage 2, point 3 (F5): credit-ledger running-balance tie-out ----
    if credit_txns:
        tie_failures = []
        for idx, m in enumerate(ctx["months"]):
            _, actual_c, gap = ledger_row_extra.get(m, (None, None, None))
            if gap is not None and abs(gap) > TOL:
                tie_failures.append((m, gap))
                highlight_cells[(idx, TIEOUT_COL)] = RED
        n_checked = sum(1 for m in ctx["months"] if ledger_row_extra.get(m, (None, None, None))[1] is not None)
        findings.append(Finding(
            "F5", "Electronic Credit Ledger running-balance tie-out (Opening + Credited - Debited = Closing)",
            "FLAG" if tie_failures else ("INFO" if n_checked == 0 else "PASS"),
            (f"Checked {n_checked} of {len(ctx['months'])} month(s) where the ledger export has a "
             f"transaction tagged to that exact tax period (a month with no such transaction can't "
             f"be checked and isn't counted as a failure). "
             + (f"{len(tie_failures)} month(s) don't tie out: "
                + "; ".join(f"{m} (gap {_f(g)})" for m, g in tie_failures) + ". "
                if tie_failures else "All checked months tie out within the rounding tolerance. ")
             + f"Running balance starts from {_f(running_opening_start)} -- the ledger's own "
               f"reported balance immediately BEFORE this FY's first month's own transactions "
               f"(not the ledger file's raw 'Opening Balance' header of {_f(credit_opening)}, "
               f"which can predate a prior-FY return that was actually paid within this ledger "
               f"download's date window; that difference is absorbed into the starting point "
               f"here rather than showing up as a false gap on every month).")
            if credit_txns else "Electronic Credit Ledger not supplied.",
            dict(months_checked=n_checked, months_failed=len(tie_failures))))
    else:
        findings.append(Finding("F5", "Electronic Credit Ledger running-balance tie-out",
                                "SKIPPED", "Electronic Credit Ledger not supplied.", {}))

    # ---- Stage 2, point 4 (F6): as-filed vs as-computed -- confirmed not buildable ----
    findings.append(Finding(
        "F6", "GSTR-3B 'as filed' vs 'as computed' comparison", "INFO",
        "Not built: every 4A-4D figure in this sheet is ALREADY read directly from the filed "
        "GSTR-3B merged workbook -- this tool does not separately re-derive Table 4 from any other "
        "source, so there is no independently-computed second figure to compare it against. (This "
        "mirrors the GSTR-2A Data Quality sheet's own note on why blocked-credit HSN screening "
        "can't be built from GSTR-2A -- same principle: don't fabricate a comparison the underlying "
        "data doesn't support.)", {}))

    # ---- Stage 2, point 5 (F7): month-over-month outlier flags ----
    _cols = [("4A(1) import goods", a1_by_m), ("4A(2) import svc", a2_by_m), ("4A(3) RCM", a3_by_m),
             ("4A(4) ISD", a4_by_m), ("4A(5) all other ITC", a5_by_m), ("4A TOTAL availed", avail_by_m),
             ("4B(1) Rules 38/42/43", b1_by_m), ("4B(2) Others", b2_by_m), ("4C net ITC", net_by_m),
             ("4D(1) sec 17(5)", d1_by_m), ("4D(2) other ineligible", d2_by_m)]
    outlier_lines = []
    for col_name, by_m in _cols:
        hits = _mom_outlier_check(by_m, ctx["months"], cfg.ITC_ROLLFORWARD_MOM_THRESHOLD,
                                   floor=cfg.ITC_ROLLFORWARD_MOM_FLOOR)
        for h in hits:
            if h["kind"] == "new":
                outlier_lines.append(f"{col_name}/{h['month']}: new activity ({_f(h['this'])}, prior month nil)")
            else:
                direction = "up" if h["this"] > h["prior"] else "down"
                outlier_lines.append(f"{col_name}/{h['month']}: {direction} {h['ratio']:.1f}x vs prior month "
                                      f"({_f(h['prior'])} -> {_f(h['this'])})")
    findings.append(Finding(
        "F7", f"Month-over-month outliers across 4A-4D columns (>={cfg.ITC_ROLLFORWARD_MOM_THRESHOLD:.0f}x change)",
        "REVIEW" if outlier_lines else "PASS",
        (f"{len(outlier_lines)} column/month combination(s) changed by {cfg.ITC_ROLLFORWARD_MOM_THRESHOLD:.0f}x "
         f"or more vs the prior month (or are new activity where the prior month was nil): "
         + "; ".join(outlier_lines) + ". A large swing can be entirely legitimate (business growth, "
         "a one-off import, a quarter-end ISD distribution) -- shown for review, not asserted as an error."
         if outlier_lines else
         f"No column changed by {cfg.ITC_ROLLFORWARD_MOM_THRESHOLD:.0f}x or more vs its prior month."),
        dict(outlier_count=len(outlier_lines))))

    credit_notes = []
    if credit_txns:
        credit_notes.append(f"Credit ledger opening balance {_f(credit_opening)}; closing balance "
                     f"{_f(_n(credit_txns[-1].get('bal_total')))}; "
                     f"{len(credit_txns)} ledger transactions in the period.")
    return dict(header=["Month", "4A(1) import goods", "4A(2) import svc", "4A(3) RCM",
                        "4A(4) ISD", "4A(5) all other ITC", "4A TOTAL availed",
                        "4B(1) Rules 38/42/43", "4B(2) Others", "4B TOTAL reversed",
                        "4C net ITC", "4D(1) sec 17(5)", "4D(2) other ineligible",
                        "Credit ledger credited", "Credit ledger debited", "Note",
                        "Blocked Scan Flagged ITC", "Blocked Scan minus 4D(1)",
                        "Credit Ledger Computed Closing", "Credit Ledger Actual Closing",
                        "Credit Ledger Tie-Out Gap"],
                widths=[10] + [18] * 14 + [40] + [20, 20, 22, 20, 18],
                rows=rows, findings=findings, notes=credit_notes, highlight_cells=highlight_cells)


def build_three_way(ctx):
    """F4 + F5 -- GSTR-1 vs GSTR-3B vs E-Way-Bill, monthly and FY."""
    rows, findings = [], []
    out_by_month = defaultdict(lambda: [0.0, 0.0, 0])
    ewb_out, dropped = dedupe_ewb(ctx["ewb_out_rows"])
    for r in ewb_out:
        if r.get("month"):
            b = out_by_month[r["month"]]
            b[0] += _n(r.get("assess")); b[1] += _n(r.get("taxval")); b[2] += 1
    T = defaultdict(float)
    for m in ctx["months"]:
        g1 = ctx["g1_by_month"].get(m, {})
        g3b = ctx["g3b_by_month"].get(m, {})
        g1_tax = _n(g1.get("taxable")) - _n(g1.get("cn_taxable"))
        g1_tot = (_n(g1.get("IGST")) + _n(g1.get("CGST")) + _n(g1.get("SGST"))
                  - _n(g1.get("cn_IGST")) - _n(g1.get("cn_CGST")) - _n(g1.get("cn_SGST")))
        a = g3b.get("3.1a") or []
        b3_tax = _n(a[0]) if len(a) > 0 else 0.0
        b3_tot = sum(_n(x) for x in a[1:4]) if len(a) > 3 else 0.0
        e_ass, e_tax, e_n = out_by_month.get(m, [0.0, 0.0, 0])
        rows.append([m, g1_tax, b3_tax, e_ass if e_n else None, g1_tax - b3_tax,
                     (e_ass - g1_tax) if e_n else None, g1_tot, b3_tot, e_tax if e_n else None,
                     e_n or None])
        T["g1"] += g1_tax; T["3b"] += b3_tax; T["ewb"] += e_ass
        T["g1t"] += g1_tot; T["3bt"] += b3_tot; T["ewbt"] += e_tax; T["n"] += e_n
    rows.append(["FY TOTAL", T["g1"], T["3b"], T["ewb"] or None, T["g1"] - T["3b"],
                 (T["ewb"] - T["g1"]) if T["n"] else None, T["g1t"], T["3bt"],
                 T["ewbt"] or None, T["n"] or None])

    d13 = T["g1"] - T["3b"]
    findings.append(Finding(
        "F4", "GSTR-1 vs GSTR-3B outward value (FY)", 
        "FLAG" if abs(d13) > MATERIAL else "PASS",
        f"FY taxable outward value per GSTR-1 (net of credit notes) {_f(T['g1'])} vs GSTR-3B "
        f"Table 3.1(a) {_f(T['3b'])}; difference {_f(d13)}.",
        dict(gstr1=T["g1"], gstr3b=T["3b"], difference=d13)))

    if T["n"]:
        de = T["ewb"] - T["g1"]
        sev = "FLAG" if abs(de) > MATERIAL else "PASS"
        # BUG FIX (Bug 4): same reasoning as check #25 -- a services-dominant business (GSTN SAC
        # prefix '99') has little/nothing for an e-way bill to move, so EWB value falling well
        # short of invoiced value is expected for that HSN profile across the WHOLE FY, not a
        # red flag. Computed from the FY-wide GSTR-1 HSN summary (all months combined), gated on
        # the dominant-HSN fact so a genuinely goods-dominant business (or a services business
        # whose gap is unusual even accounting for its own profile) still flags.
        is_services_dom, dom_hsn, dom_share = False, None, 0.0
        g1_files = {ctx["res"]["gstr1_month_map"].get(m) for m in ctx["months"]} - {None}
        fy_hsn_rows = []
        for g1f in g1_files:
            hsn_all = pr.read_gstr1_hsn_all_months(g1f)
            for m in ctx["months"]:
                fy_hsn_rows.extend(hsn_all.get(m, []))
        if fy_hsn_rows:
            is_services_dom, dom_hsn, dom_share = pr.dominant_hsn_is_services(fy_hsn_rows)
        # BUG FIX (bug report §9): the narrative previously attached ONE fixed "under-billing"
        # story regardless of which way the gap pointed -- computing difference = EWB - GSTR1
        # but never branching on its sign before choosing what to say. EWB > GSTR-1 (positive) is
        # the genuine under-billing signature (goods physically moved, not correspondingly
        # invoiced). EWB < GSTR-1 (negative) is the OPPOSITE direction and has mostly benign,
        # common explanations that have nothing to do with under-billing -- printing the
        # under-billing story there was a false-positive risk flag. This branches on sign(de)
        # (and on materiality) before picking the narrative; severity thresholds themselves are
        # unchanged.
        if abs(de) <= MATERIAL:
            narrative = ("EWB assessable value and GSTR-1 taxable value agree within the "
                         "materiality threshold.")
        elif de > 0:
            narrative = (
                "Goods moved under an e-way bill without a correspondingly large taxable invoice "
                "is the classic signature of unbilled or under-billed outward supply; legitimate "
                "explanations exist (job-work challans, branch transfers, sales returns, "
                "non-supply movements) and should be obtained in writing.")
        else:
            coverage_pct = (T["ewb"] / T["g1"] * 100.0) if T["g1"] else None
            narrative = (
                "EWB-covered movement is LOWER than GSTR-1 invoiced value -- the opposite "
                "direction from under-billing. Commonly explained by services (no EWB required), "
                "sub-threshold invoices (below the Rs 50,000 EWB limit), intra-city/short-distance "
                "movements exempt from EWB, or document-only adjustments (debit notes/price "
                "revisions) that inflate GSTR-1 turnover without any corresponding movement. Not "
                "a compliance flag by default"
                + (f"; EWB covers only about {coverage_pct:.0f}% of invoiced value here, which "
                   f"may still warrant a sampling check of invoices lacking an EWB -- just not "
                   f"framed as under-billing." if coverage_pct is not None else "."))
        if abs(de) > MATERIAL and is_services_dom and de < 0:
            sev = "EXPLAINED"
            services_note = (f" FY-wide dominant HSN/SAC is {dom_hsn} ({dom_share:.0%} of taxable "
                             f"value), a SERVICES code -- e-way bills apply only to goods movement "
                             f"(Rule 138), so EWB value falling well short of invoiced value is "
                             f"expected for this HSN profile, not evidence of unbilled supply. See "
                             f"the 'HSN RATE REVIEW' table on any month's Comparison sheet.")
        else:
            services_note = ""
        findings.append(Finding(
            "F5", "E-Way-Bill outward value vs invoice value (FY)", sev,
            f"{int(T['n'])} distinct outward e-way bills carry {_f(T['ewb'])} of assessable value "
            f"({_f(T['ewbt'])} tax), against {_f(T['g1'])} of taxable value invoiced in GSTR-1 "
            f"-- a difference of {_f(de)}. " + narrative + services_note
            + (f" {dropped} duplicate e-way-bill row(s) in the source export were excluded before "
               f"totalling." if dropped else ""),
            dict(ewb_assessable=T["ewb"], gstr1_taxable=T["g1"], difference=de,
                 ewb_count=int(T["n"]))))
    else:
        findings.append(Finding("F5", "E-Way-Bill outward value vs invoice value (FY)", "SKIPPED",
                                "No outward e-way-bill workbook was supplied.", {}))
    notes = []
    if dropped:
        notes.append(f"{dropped} duplicate row(s) were present in the outward e-way-bill export "
                     f"(same EWB number appearing more than once) and were excluded.")
    ewb_in, dropped_in = dedupe_ewb(ctx["ewb_in_rows"])
    if dropped_in:
        notes.append(f"{dropped_in} duplicate row(s) were present in the INWARD e-way-bill export "
                     f"and were excluded from every inward aggregate in this tool.")
    return dict(header=["Month", "GSTR-1 taxable (net CN)", "GSTR-3B 3.1(a) taxable",
                        "EWB-out assessable", "GSTR-1 minus 3B", "EWB minus GSTR-1",
                        "GSTR-1 tax", "GSTR-3B 3.1(a) tax", "EWB tax", "EWB count"],
                widths=[10, 22, 22, 20, 18, 18, 16, 18, 14, 11], rows=rows,
                findings=findings, notes=notes)


def build_itc_3b_vs_2b(ctx):
    """F7 -- ITC claimed in 3B vs ITC available in 2B, computed invoice-level,
    with the quarterly-summary figure shown alongside as a control total."""
    rows, findings = [], []
    T = defaultdict(float)
    any2b = False
    any_summ_available = False
    for m in ctx["months"]:
        g3b = ctx["g3b_by_month"].get(m, {})
        claimed = sum(_n(x) for x in (g3b.get("4A5") or [])[:4])
        two = ctx["twob_by_month"].get(m, {})
        summ = ctx["twob_summary_by_month"].get(m, {})
        # Uses '_summary_available' (the narrower 'was the ITC Available/control-total sheet
        # itself readable' signal), NOT the broader 'available' flag -- see summary_for_month's
        # own docstring for why these are now two different questions. This check specifically
        # needs the narrow one: it's comparing against the SUMMARY sheet's own control total,
        # which can be genuinely absent even when the invoice-level B2B/CDNR data (used by
        # 'avail' below, from ctx['twob_by_month'], a completely separate read) is fully present.
        summ_val = (sum(_n(summ.get(k)) for k in ("ITC_all_other_IGST", "ITC_all_other_CGST",
                                                  "ITC_all_other_SGST", "ITC_all_other_CESS"))
                    if summ.get("_summary_available") else None)
        if summ.get("_summary_available"):
            any_summ_available = True
        if two.get("available"):
            any2b = True
            avail = two["itc"] + two["cn_itc"]
            rows.append([m, claimed, avail, claimed - avail, two["invoices"], two["suppliers"],
                         summ_val, (avail - summ_val) if summ_val is not None else None])
            T["claimed"] += claimed; T["avail"] += avail
            T["summ"] += summ_val or 0.0
        else:
            rows.append([m, claimed, None, None, None, None, summ_val, None])
            T["claimed"] += claimed
    if any2b:
        # BUG FIX (same class already fixed elsewhere -- Turnover Growth vs Tax, FY-Total vs
        # BIFA -- found again here while cross-checking this exact sheet against a taxpayer's
        # own manual verification): T['summ'] silently collapses to 0 when NO month's 'ITC
        # Available' summary sheet was available at all (confirmed real: this taxpayer's GSTR-2B
        # export has no 'ITC Available' sheet whatsoever), and that fake 0 was then shown in F7a
        # as if it were a genuine computed summary-sheet total, against the real invoice-level
        # figure -- producing a huge, meaningless 'difference' that's actually just 'the source
        # sheet doesn't exist', not a reconciliation gap. Fixed: FY TOTAL row and F7a both now
        # distinguish 'summary genuinely summed to zero' from 'summary sheet never existed'.
        rows.append(["FY TOTAL", T["claimed"], T["avail"], T["claimed"] - T["avail"],
                     None, None, (T["summ"] if any_summ_available else None),
                     ((T["avail"] - T["summ"]) if any_summ_available else None)])
        d = T["claimed"] - T["avail"]
        findings.append(Finding(
            "F7", "ITC claimed in GSTR-3B 4A(5) [All-Other ITC only, NOT full Table 4A] vs ITC available in GSTR-2B",
            "FLAG" if d > MATERIAL else ("REVIEW" if abs(d) > MATERIAL else "PASS"),
            f"FY ITC claimed under Table 4A(5) 'All other ITC' SPECIFICALLY {_f(T['claimed'])} -- "
            f"NOT the full Table 4A total (which also includes 4A(1) imports of goods, 4A(2) "
            f"imports of services, 4A(3) RCM, 4A(4) ISD -- see the 'ITC Annual Summary' sheet "
            f"for that full figure, which will legitimately be larger than this one). Scoped to "
            f"4A(5) here deliberately: this sheet compares claimed ITC against GSTR-2B's "
            f"INVOICE-level data, and imports/RCM/ISD don't originate from any supplier's 2B "
            f"invoice at all, so only 4A(5) is the sub-head 2B invoices can actually be checked "
            f"against. FY ITC available per GSTR-2B "
            f"computed from the invoice-level B2B and B2B-CDNR rows (net of B2BA/B2B-CDNRA "
            f"amendments -- see the '_read_b2ba_amendments' fix) {_f(T['avail'])}; "
            f"difference {_f(d)}"
            + (" -- claimed EXCEEDS available, which is an excess-credit exposure under "
               "section 16(2)(aa) read with Rule 36(4)." if d > MATERIAL else "."),
            dict(claimed_3b=T["claimed"], available_2b=T["avail"], difference=d)))
        if not any_summ_available:
            findings.append(Finding(
                "F7a", "GSTR-2B summary sheet vs invoice-level total -- residual difference", "SKIPPED",
                "Not computable: this taxpayer's merged GSTR-2B workbook has no 'ITC Available' "
                "summary sheet at all (confirmed -- not every real export includes it; some carry "
                "only 'B2B'+'B2B-CDNR'). There is no summary-sheet figure to compare against the "
                "invoice-level total in F7 above, which is unaffected -- it never depended on this "
                "summary sheet.", {}))
        else:
            ds = T["avail"] - T["summ"]
            if abs(ds) > TOL:
                findings.append(Finding(
                    "F7a", "GSTR-2B summary sheet vs invoice-level total -- residual difference", "INFO",
                    f"The 'ITC Available' summary sheet in the merged GSTR-2B workbook yields "
                    f"{_f(T['summ'])} for the year, against {_f(T['avail'])} computed from the "
                    f"invoice-level rows (B2B + B2B-CDNR, net of B2BA/B2B-CDNRA amendments) -- a "
                    f"difference of {_f(ds)}. BUG FIX applied: this summary sheet lays four "
                    f"column-groups side by side per row (month 1, month 2, month 3, quarter total) "
                    f"and used to be read by always taking the first group, so months 2 and 3 of "
                    f"every quarter showed month-1's figures -- that misread is now corrected (the "
                    f"right column-group is selected per month, with a built-in month1+2+3-vs-total "
                    f"sanity check). The residual difference shown here is NOT that bug -- likely "
                    f"causes are 2B summary components this module doesn't feed into the "
                    f"invoice-level figure at all (ISD credit, import of goods/services), or a "
                    f"quarter-total column in the source file that doesn't itself foot to its own "
                    f"three months (see '_qtr_total_mismatch' if flagged). Every 2B figure elsewhere "
                    f"in this module is taken from the invoice rows; this summary figure is shown "
                    f"only as a control total, never used in a calculation.",
                    dict(summary_path=T["summ"], invoice_level=T["avail"], difference=ds)))
    else:
        findings.append(Finding("F7", "ITC claimed in GSTR-3B vs available in GSTR-2B", "SKIPPED",
                                "GSTR-2B not supplied for any month.", {}))
    return dict(header=["Month", "ITC claimed 3B 4A(5)", "ITC available 2B (invoice level)",
                        "Claimed minus available", "2B invoices", "2B suppliers",
                        "2B summary-sheet figure", "Invoice level minus summary"],
                widths=[10, 22, 28, 22, 13, 13, 24, 24], rows=rows,
                findings=findings, notes=[])


def build_rcm(ctx):
    """F8 -- RCM liability declared vs RCM ITC taken vs cash actually debited.

    RCM liability cannot be discharged out of the credit ledger (section
    49(4)); it must be paid in cash. So a declared 3.1(d) liability with no
    corresponding cash-ledger debit is a genuine short payment, and RCM ITC
    taken under 4A(3) without the liability being paid is an inadmissible
    credit under section 16(2)(c)."""
    rows, findings = [], []
    cash = (ctx["annual_data"].get("cash") or {}).get("transactions", [])
    liab = (ctx["annual_data"].get("liab") or {}).get("transactions", [])
    cash_rcm = defaultdict(float)
    for t in cash:
        if "reverse charge" in str(t.get("description", "")).lower() and \
           str(t.get("ttype", "")).lower() == "debit":
            k = pr_period_key(t.get("tax_period"))
            cash_rcm[k] += _n(t.get("total"))
    liab_rcm = defaultdict(float)
    for t in liab:
        if "reverse charge" in str(t.get("description", "")).lower() and \
           str(t.get("ttype", "")).lower() == "debit":
            liab_rcm[_ledger_month(t.get("date"))] += _n(t.get("total"))
    T = defaultdict(float)
    highlight_cells = {}
    STATUS_COL = 9   # new 'Status' column, 1-based
    CASH_GAP_COL = 7
    ITC_GAP_COL = 8
    for idx, m in enumerate(ctx["months"]):
        g3b = ctx["g3b_by_month"].get(m, {})
        d = g3b.get("3.1d") or []
        rcm_taxable = _n(d[0]) if d else 0.0
        rcm_liab = sum(_n(x) for x in d[1:4]) if len(d) > 3 else 0.0
        rcm_itc = sum(_n(x) for x in (g3b.get("4A3") or [])[:4])
        two = ctx["twob_by_month"].get(m, {})
        rcm_2b = two.get("rcm_itc") if two.get("available") else None
        paid_cash = cash_rcm.get(m)
        cash_gap = None if paid_cash is None else paid_cash - rcm_liab
        itc_gap = rcm_itc - rcm_liab
        # NEW: per-month Status -- makes the problem months visible at a glance in the table
        # itself, not just buried in the FY-level F8/F8a finding text below.
        _floor_m = max(100.0, 0.01 * rcm_liab)
        status_bits = []
        if cash_gap is not None and cash_gap < -_floor_m:
            status_bits.append("CASH SHORT-PAID")
            highlight_cells[(idx, CASH_GAP_COL)] = RED
        # BUG FIX (caught on a real run, not assumed): the original version only flagged a
        # SHORTFALL (cash < liability). Confirmed against real data this misses a real anomaly
        # the other direction -- one month showed cash-ledger RCM debits of Rs 2.63 crore against
        # a declared liability of just Rs 1.36 lakh (a ~190x overpayment) and the Status column
        # said "OK" because it was never short. Overpaying isn't itself a violation, but a gap
        # this large against the SAME month's own declared liability is very likely a different
        # 'reverse charge' cash movement getting swept into this month's Tax-Period bucket (the
        # ledger's own Tax Period tagging, not this tool's own match) rather than a routine RCM
        # payment -- worth a look either way, so it's now surfaced (amber, not red, since
        # overpayment is not itself a compliance violation the way underpayment is).
        elif cash_gap is not None and rcm_liab > TOL and cash_gap > max(MATERIAL, 2 * rcm_liab):
            status_bits.append("CASH FAR EXCEEDS LIABILITY -- VERIFY (see note)")
            highlight_cells[(idx, CASH_GAP_COL)] = AMBER
        if abs(itc_gap) > MATERIAL:
            status_bits.append("ITC vs LIABILITY GAP")
            highlight_cells[(idx, ITC_GAP_COL)] = RED
        status = "; ".join(status_bits) or "OK"
        rows.append([m, rcm_taxable, rcm_liab, rcm_itc, rcm_2b, paid_cash, cash_gap, itc_gap, status])
        T["taxable"] += rcm_taxable; T["liab"] += rcm_liab; T["itc"] += rcm_itc
        T["cash"] += _n(paid_cash)
    fy_row_idx = len(rows)
    fy_cash_gap = T["cash"] - T["liab"]
    fy_itc_gap = T["itc"] - T["liab"]
    fy_status = []
    if fy_cash_gap < -max(100.0, 0.01 * T["liab"]):
        fy_status.append("CASH SHORT-PAID"); highlight_cells[(fy_row_idx, CASH_GAP_COL)] = RED
    elif T["liab"] > TOL and fy_cash_gap > max(MATERIAL, 2 * T["liab"]):
        fy_status.append("CASH FAR EXCEEDS LIABILITY -- VERIFY"); highlight_cells[(fy_row_idx, CASH_GAP_COL)] = AMBER
    if abs(fy_itc_gap) > MATERIAL:
        fy_status.append("ITC vs LIABILITY GAP"); highlight_cells[(fy_row_idx, ITC_GAP_COL)] = RED
    rows.append(["FY TOTAL", T["taxable"], T["liab"], T["itc"], None, T["cash"],
                 fy_cash_gap, fy_itc_gap, "; ".join(fy_status) or "OK"])

    if cash_rcm:
        missing, wnote = ledger_window_note(ctx["months"], set(cash_rcm), "reverse-charge cash-ledger")
        cmp_months = [m for m in ctx["months"] if m not in missing]
        liab_cmp = 0.0
        for m in cmp_months:
            dd = ctx["g3b_by_month"].get(m, {}).get("3.1d") or []
            liab_cmp += sum(_n(x) for x in dd[1:4]) if len(dd) > 3 else 0.0
        cash_cmp = sum(_n(cash_rcm.get(m)) for m in cmp_months)
        d = cash_cmp - liab_cmp
        # Severity floor: GSTR-3B rounds its payment table to whole rupees while the ledger
        # carries paise, so a few rupees of difference is arithmetic noise, not a short payment.
        # A shortfall only becomes a finding once it exceeds the larger of Rs 100 or 1% of the
        # declared liability.
        _floor = max(100.0, 0.01 * liab_cmp)
        findings.append(Finding(
            "F8", "RCM liability declared vs cash actually debited",
            "FLAG" if d < -_floor else ("REVIEW" if d < -TOL else "PASS"),
            f"Compared over the {len(cmp_months)} month(s) the cash ledger actually covers: "
            f"reverse-charge liability declared in Table 3.1(d) {_f(liab_cmp)} against RCM debits "
            f"in the Electronic Cash Ledger {_f(cash_cmp)}; difference {_f(d)}. Reverse-charge "
            f"liability must be discharged in cash (section 49(4)), so these should agree exactly; "
            f"the cash ledger is the authoritative side. " + wnote +
            f" Unadjusted full-year figures: declared {_f(T['liab'])} against {_f(T['cash'])} paid.",
            dict(declared_in_window=liab_cmp, paid_cash_in_window=cash_cmp, difference=d,
                 months_excluded=", ".join(missing))))
    else:
        findings.append(Finding("F8", "RCM liability declared vs cash actually debited", "SKIPPED",
                                "Electronic Cash Ledger not supplied -- the cash leg of the RCM "
                                "verification has no source.", {}))
    di = T["itc"] - T["liab"]
    findings.append(Finding(
        "F8a", "RCM ITC taken vs RCM liability declared", 
        "FLAG" if di > MATERIAL else ("REVIEW" if abs(di) > TOL else "PASS"),
        f"ITC taken on inward reverse-charge supplies under Table 4A(3) {_f(T['itc'])} against "
        f"reverse-charge liability declared in Table 3.1(d) {_f(T['liab'])}; difference {_f(di)}. "
        f"Credit under 4A(3) is admissible only to the extent the corresponding tax was actually "
        f"paid (section 16(2)(c)).",
        dict(itc_4a3=T["itc"], liability_31d=T["liab"], difference=di)))
    notes = ["RCM rows in the Electronic Cash Ledger are matched by their own 'Tax Period' "
             "column, not by transaction date, so a liability for March paid in April is counted "
             "against March.",
             "'CASH FAR EXCEEDS LIABILITY' months: the cash-ledger RCM debit tagged to that "
             "month's Tax Period is more than double (and materially larger than) that same "
             "month's own declared 3.1(d) RCM liability. Not itself a violation (overpaying tax "
             "isn't illegal), but a gap this large usually means a DIFFERENT reverse-charge cash "
             "movement got tagged to this month's Tax Period rather than this being routine "
             "monthly RCM -- worth tracing the individual ledger entries for that month.",
             "SUGGESTED FURTHER ANALYSIS from raw data already in this tool's inputs, not yet "
             "built: (1) RCM liability by HSN/SAC of the inward supply, to see which expense "
             "categories drive RCM exposure -- needs a Description/HSN field GSTR-2B doesn't "
             "carry, so would need a purchase register; (2) supplier-wise RCM concentration "
             "(which vendors' supplies are routinely RCM-liable -- GTA, legal fees, etc.) from "
             "GSTR-2A's own 'Supply Attract Reverse Charge' flag, cross-checked against this "
             "sheet's declared liability per month; (3) a rolling 12-month RCM cash-shortfall "
             "trend, once more than one FY's output from this tool exists to compare across."]
    return dict(header=["Month", "3.1(d) taxable", "3.1(d) RCM liability", "4A(3) RCM ITC",
                        "2B RCM-flagged ITC", "Cash ledger RCM debit", "Cash minus liability",
                        "ITC minus liability", "Status"],
                widths=[10, 16, 20, 16, 20, 22, 20, 18, 22], rows=rows, findings=findings, notes=notes,
                highlight_cells=highlight_cells)


def pr_period_key(tp):
    """'Mar-22' / 'Mar-2022' -> 'Mar-22'; anything unparseable -> None."""
    if not tp or str(tp).strip() in ("-", ""):
        return None
    m = re.match(r"([A-Za-z]{3})-(\d{2,4})", str(tp).strip())
    if not m:
        return None
    return f"{m.group(1)[:3].title()}-{m.group(2)[-2:]}"


_MON = {1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
        7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"}


def _ledger_month(datestr):
    """Both '-' and '/' day-month-year separators confirmed in real ledger exports
    for different taxpayers -- see gst_parsers_dept._period_key's docstring for the
    same fix applied to the annual-ledger walkthrough (this is the sibling function
    for the F8 RCM triangulation check, same root cause, same fix)."""
    m = re.match(r"(\d{2})[-/](\d{2})[-/](\d{4})", str(datestr or "").strip())
    if not m:
        return None
    return f"{_MON[int(m.group(2))]}-{m.group(3)[2:]}"


def build_drc_and_ledger_movements(ctx):
    """F9 -- DRC-03 / voluntary payments, and every other ledger movement that
    is NOT an ordinary return filing (refund debits, Rule 86A blocking,
    departmental demands). Three independent sources are cross-referenced."""
    rows, findings = [], []
    cash = (ctx["annual_data"].get("cash") or {}).get("transactions", [])
    credit = (ctx["annual_data"].get("credit") or {}).get("transactions", [])
    VOL = ("voluntary", "drc", "demand", "scrutiny", "annual return")
    n_drc = 0
    for src, txns, amt_key in (("Cash ledger", cash, "total"), ("Credit ledger", credit, "total")):
        for t in txns:
            desc = str(t.get("description", ""))
            ref = str(t.get("ref", ""))
            low = desc.lower()
            is_vol = any(k in low for k in VOL)
            is_other = ("refund" in low or "block" in low)
            if not (is_vol or is_other):
                continue
            kind = "DRC-03 / voluntary" if is_vol else (
                "Refund debit" if "refund" in low else "Rule 86A block")
            n_drc += 1 if is_vol else 0
            rows.append([src, t.get("date"), ref, t.get("tax_period") or "-", desc,
                         t.get("ttype"), _n(t.get(amt_key)), kind])
    portal = ctx["annual_data"].get("comp") or {}
    bo = ctx["annual_data"].get("bo") or {}
    fy_months = set(ctx["months"])
    for d in (bo.get("drc_payments") or []):
        dt = str(d.get("date") or d.get("transaction_date") or "")
        in_fy = _bo_date_in_fy(dt, ctx["months"])
        rows.append(["BO Profile", dt, d.get("source_id") or d.get("ref") or "-", "-",
                     str(d.get("description") or d.get("type") or ""),
                     d.get("method") or "-",
                     _n(d.get("total")) if d.get("total") is not None else None,
                     "BO Profile DRC entry -- IN this FY" if in_fy else
                     "BO Profile DRC entry -- OUTSIDE this FY (context only)"])

    if n_drc:
        findings.append(Finding(
            "F9", "DRC-03 / voluntary payments during the year", "REVIEW",
            f"{n_drc} voluntary/DRC-type movement(s) found in the electronic ledgers for the "
            f"period under review. Each is itemised on the sheet with its reference number; "
            f"the underlying DRC-03 and the reason for the payment should be obtained.",
            dict(count=n_drc)))
    elif cash or credit:
        findings.append(Finding(
            "F9", "DRC-03 / voluntary payments during the year", "PASS",
            "No voluntary-payment or DRC-type entry appears in the Electronic Cash or Credit "
            "Ledger for the period under review. Any DRC entries visible in the BO Profile "
            "relate to other periods and are listed separately on the sheet as context.", {}))
    else:
        findings.append(Finding("F9", "DRC-03 / voluntary payments during the year", "SKIPPED",
                                "Neither ledger was supplied.", {}))

    # refund debits out of the credit ledger vs the BO Profile's refund table
    ref_debit = sum(_n(t.get("total")) for t in credit
                    if "refund" in str(t.get("description", "")).lower()
                    and str(t.get("ttype", "")).lower() == "debit")
    ref_credit = sum(_n(t.get("total")) for t in credit
                     if "refund" in str(t.get("description", "")).lower()
                     and str(t.get("ttype", "")).lower() == "credit")
    if ref_debit:
        fy = ctx.get("fy_label")
        bo_ref = (bo.get("refund_by_fy") or {}).get(fy or "", {})
        bo_claimed = _n(bo_ref.get("claimed")) * 1e5 if bo_ref else None   # BO Profile is in lakhs
        detail = (f"{_f(ref_debit)} was debited from the Electronic Credit Ledger during the year "
                  f"as refund claimed from the ITC ledger"
                  + (f", of which {_f(ref_credit)} was re-credited (rejected or withdrawn)"
                     if ref_credit else "") + ". ")
        sev = "REVIEW"
        if bo_ref:
            detail += (f"The BO Profile's own Refund Details table shows {_f(bo_claimed)} claimed "
                       f"for {fy}. ")
            if bo_claimed is not None and abs(bo_claimed - ref_debit) > MATERIAL:
                sev = "FLAG"
                detail += ("These do not reconcile -- a refund flow of this size that the "
                           "department's own profile records as nil should be explained. ")
        detail += ("Note also that no zero-rated or export supply is declared anywhere in GSTR-1 "
                   "or GSTR-9 for this year, so the statutory basis for the refund (inverted duty "
                   "structure under section 54(3)(ii), or some other head) should be established.")
        findings.append(Finding("F9a", "Refund claimed out of the Electronic Credit Ledger", sev,
                                detail, dict(refund_debited=ref_debit, refund_recredited=ref_credit,
                                             bo_profile_claimed=bo_claimed)))
    blocked = [t for t in credit if "block" in str(t.get("description", "")).lower()]
    if blocked:
        amt = sum(_n(t.get("total")) for t in blocked)
        findings.append(Finding(
            "F9b", "Credit blocked in the Electronic Credit Ledger (Rule 86A)", "FLAG",
            f"{len(blocked)} blocking entr{'y' if len(blocked) == 1 else 'ies'} totalling "
            f"{_f(amt)} appear(s) in the Electronic Credit Ledger during the year "
            f"({', '.join(str(t.get('date')) for t in blocked)}). A Rule 86A block is a "
            f"departmental action taken where the credit is believed to be fraudulently availed "
            f"or ineligible; the order and its present status should be on file.",
            dict(count=len(blocked), amount=amt)))

    # F9c -- Electronic Liability LEDGER (Part II, demand/DRC) cross-referenced against BO
    # Profile's DRC Payment Information sheet by EXACT ID match (Reference No./Relevant Demand
    # ID vs BO Profile's Source ID -- confirmed same ID namespace/format on real data: both use
    # DI.../DC.../IP... prefixed IDs). This is a genuine exact tie-out (Hard Safety Rule (a)),
    # a materially stronger link than F9's own keyword-scan of Cash/Credit Ledger description
    # text against unlinked BO Profile context rows above.
    liab_demand = (ctx["annual_data"].get("liab_demand") or {})
    demand_txns = liab_demand.get("transactions", [])
    bo_drc_all = bo.get("drc_payments") or []
    if demand_txns:
        bo_by_source = {}
        for d in bo_drc_all:
            sid = str(d.get("source_id") or "").strip().upper()
            if sid:
                bo_by_source.setdefault(sid, []).append(d)
        matched_bo_sources = set()
        unmatched_ledger = []
        for t in demand_txns:
            key = str(t.get("demand_id") or t.get("ref") or "").strip().upper()
            hit = bo_by_source.get(key)
            if hit:
                matched_bo_sources.add(key)
                classification = f"Matched to BO Profile DRC (Source ID {key})"
            else:
                unmatched_ledger.append(t)
                classification = "NOT found in BO Profile DRC Payment Information -- verify"
            rows.append(["Liability Ledger (Part II)", t.get("date"), t.get("ref"),
                         t.get("tax_period") or "-",
                         (t.get("description") or "") + (f" [Stay: {t['stay_status']}]" if t.get("stay_status") else ""),
                         t.get("ttype"), _n(t.get("total")), classification])
        unmatched_bo_in_fy = [d for d in bo_drc_all
                               if str(d.get("source_id") or "").strip().upper() not in matched_bo_sources
                               and _bo_date_in_fy(str(d.get("date") or ""), ctx["months"])]
        if unmatched_ledger or unmatched_bo_in_fy:
            detail = (f"{len(demand_txns)} entr{'y' if len(demand_txns)==1 else 'ies'} in the Electronic "
                      f"Liability Ledger (Part II -- DRC/demand/voluntary payments) for the year; "
                      f"{len(demand_txns) - len(unmatched_ledger)} matched a BO Profile DRC Payment "
                      f"Information row by exact Source ID/Demand ID. ")
            if unmatched_ledger:
                amt = sum(_n(t.get("total")) for t in unmatched_ledger)
                detail += (f"{len(unmatched_ledger)} Liability-Ledger entr{'y' if len(unmatched_ledger)==1 else 'ies'} "
                          f"(₹{_f(amt)}) have NO corresponding BO Profile DRC record -- the underlying order/"
                          f"voluntary-payment reference should be obtained and the BO Profile gap explained. ")
            if unmatched_bo_in_fy:
                detail += (f"{len(unmatched_bo_in_fy)} BO Profile DRC record(s) dated within this FY have no "
                          f"matching Liability-Ledger entry -- verify whether these were paid through a "
                          f"different ledger/period or the ID format genuinely differs.")
            findings.append(Finding("F9c", "Liability Ledger (Part II) vs BO Profile DRC -- exact ID match",
                                    "REVIEW", detail,
                                    dict(ledger_entries=len(demand_txns), unmatched_ledger=len(unmatched_ledger),
                                         unmatched_bo=len(unmatched_bo_in_fy))))
        else:
            findings.append(Finding("F9c", "Liability Ledger (Part II) vs BO Profile DRC -- exact ID match",
                                    "PASS",
                                    f"All {len(demand_txns)} Electronic Liability Ledger (Part II) entries for "
                                    f"the year matched a BO Profile DRC Payment Information row by exact "
                                    f"Source ID/Demand ID -- no unexplained demand/DRC movement."))
    elif bo_drc_all:
        findings.append(Finding("F9c", "Liability Ledger (Part II) vs BO Profile DRC -- exact ID match", "SKIPPED",
                                "Electronic Liability Ledger (Part II, demand/DRC) not supplied -- BO Profile's "
                                "own DRC records are listed above as unlinked context only; supply the Ledger "
                                "for an exact ID-matched cross-check."))

    notes = ["SUGGESTED FURTHER ANALYSIS from raw data already in this tool's inputs, not yet "
             "built: (1) a running Rule 86A block-vs-release timeline (is a block still open at "
             "FY end, or was it released -- both dates would need to be in the ledger's own "
             "description text to detect automatically); (2) DRC-03 payments cross-tabulated by "
             "the SECTION/reason cited (voluntary vs SCN-driven vs annual-return-driven) once "
             "that's captured as a structured field rather than free text; (3) a total-value "
             "concentration view -- what % of all non-return ledger movement this year sits in "
             "the single largest entry, a quick tell for whether a handful of items decide the "
             "whole picture or it's evenly spread."]
    if portal:
        notes.append("The portal's 'Tax liability and ITC comparison' report also carries a "
                     "'Payment made, if any through DRC-03' column per month; it is read on the "
                     "Annual Ledger Walkthrough sheet.")
    # NEW: per-row highlight -- red for Rule 86A blocks and for ledger entries with no BO Profile
    # match, the two classifications on this sheet that actually need someone to chase a document
    # (everything else is either a clean match or a routine voluntary payment already covered by
    # its own finding above). Total row appended so the aggregate is visible without summing by
    # hand.
    highlight_cells = {}
    KIND_COL = 8
    for idx, row in enumerate(rows):
        cls = str(row[KIND_COL - 1] or "")
        if "Rule 86A" in cls or "NOT found in BO Profile" in cls:
            highlight_cells[(idx, KIND_COL)] = RED
    total_amt = sum(_n(r[6]) for r in rows if isinstance(r[6], (int, float)))
    rows.append(["TOTAL (all sources, all types -- signs not netted)", "", "", "", "", "", total_amt, ""])
    return dict(header=["Source", "Date", "Reference", "Tax period", "Description",
                        "Type", "Amount", "Classification"],
                widths=[14, 12, 22, 12, 42, 10, 16, 46], rows=rows, findings=findings, notes=notes,
                highlight_cells=highlight_cells)


def _bo_date_in_fy(datestr, months):
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", str(datestr or "").strip())
    if not m:
        return False
    lbl = f"{_MON[int(m.group(2))]}-{m.group(1)[2:]}"
    return lbl in set(months)


def build_turnover_vs_tax(ctx):
    """F10 -- turnover growth against tax actually paid, every FY the BO
    Profile carries. A turnover line that rises while the cash contribution
    stays flat is the shape that justifies a closer look at the credit side."""
    rows, findings = [], []
    bo = ctx["annual_data"].get("bo") or {}
    fin = bo_financial_rows(bo)
    if not fin:
        return dict(header=["Financial year"], widths=[16], rows=[],
                    findings=[Finding("F10", "Turnover growth vs tax payment", "SKIPPED",
                                      "BO / 360-degree Profile not supplied, or its Financial "
                                      "Information section could not be read -- this comparison "
                                      "has no source.", {})],
                    notes=[])
    prev = None
    for fy in sorted(fin):
        r = fin[fy]
        if not r.get("usable"):
            rows.append([fy, None, None, None, None, None, None, None, None,
                         "row could not be parsed -- shown for completeness"])
            continue
        to, liab = r["turnover"], r["total_tax_liability"]
        cash, itc = r["tax_paid_in_cash"], r["tax_paid_by_itc"]
        eff = _pct(liab, to)
        cash_share = _pct(cash, liab)
        g_to = _pct(to - prev[0], prev[0]) if prev and prev[0] else None
        g_tx = _pct(liab - prev[1], prev[1]) if prev and prev[1] else None
        rows.append([fy, to, r["taxable_turnover"], liab, itc, cash, r["itc_availed"],
                     eff, cash_share, _growth_note(g_to, g_tx)])
        prev = (to, liab)
    rows.append(["", None, None, None, None, None, None, None, None,
                 "All amounts in LAKHS, exactly as printed in the BO Profile."])

    usable = [(fy, fin[fy]) for fy in sorted(fin) if fin[fy].get("usable")]
    if usable:
        cash_shares = [(_pct(r["tax_paid_in_cash"], r["total_tax_liability"]) or 0.0)
                       for _, r in usable]
        avg_cash = sum(cash_shares) / len(cash_shares)
        first, last = usable[0][1], usable[-1][1]
        gro = _pct(last["turnover"] - first["turnover"], first["turnover"])
        findings.append(Finding(
            "F10", "Turnover growth vs tax paid in cash", 
            "REVIEW" if avg_cash < 5.0 else "INFO",
            f"Across {len(usable)} financial years in the BO Profile, turnover moved from "
            f"Rs {first['turnover']:,.2f} lakh ({usable[0][0]}) to Rs {last['turnover']:,.2f} lakh "
            f"({usable[-1][0]})"
            + (f", a change of {gro:,.1f}%" if gro is not None else "") +
            f", while the share of tax liability discharged in CASH averaged {avg_cash:,.2f}% "
            f"-- the remainder being set off against input tax credit. A rising turnover line "
            f"with a persistently negligible cash contribution is the profile in which the "
            f"genuineness of the input side carries the whole weight of the assessment. This is "
            f"an observation about structure, not an allegation: a manufacturer with genuine "
            f"high-value inputs can legitimately show exactly this shape.",
            dict(avg_cash_share_pct=avg_cash, years=len(usable))))
    return dict(header=["Financial year", "Turnover (lakh)", "Taxable turnover (lakh)",
                        "Total tax liability (lakh)", "Paid by ITC (lakh)", "Paid in cash (lakh)",
                        "ITC availed (lakh)", "Effective tax rate %", "Cash share of liability %",
                        "Year-on-year"],
                widths=[14, 16, 20, 22, 17, 18, 17, 18, 22, 52], rows=rows,
                findings=findings, notes=[])


def _growth_note(g_to, g_tx):
    if g_to is None:
        return ""
    s = f"turnover {g_to:+.1f}%"
    if g_tx is not None:
        s += f", tax liability {g_tx:+.1f}%"
        if g_to > 10 and g_tx < g_to - 10:
            s += "  <-- liability growing materially slower than turnover"
    return s


def build_counterparty(ctx):
    """F11 -- same-day repeat transactions with any counterparty (no value
    floor, per instruction) and reciprocal buy/sell pairs.

    UPDATED (per instruction): reciprocal counterparties now get their OWN table with
    purchases-from-them and sales-to-them broken out into separate columns (instead of being
    squeezed into the same generic 7-column layout as the same-day-repeat rows, with the
    breakdown buried in a text 'Detail' cell). Same-day repeats are now split into two clearly
    labelled tables (Outward / Inward) instead of one mixed list, sorted by repeat-count (the
    actual signal this check is looking for) instead of value, for easier scanning.

    FURTHER UPDATE (per instruction): names now resolved via ctx['gstin_name_lookup'] -- a
    cross-source lookup (GSTR-1, GSTR-2B, GSTR-2A, both EWB directions, BO Profile top-10 lists)
    built once in build_context(), rather than this function's own narrow 2-source fallback,
    which left some counterparties' names blank even though a name for that exact GSTIN was
    sitting in a different source this tool already had open (confirmed against real data: a
    government TDS-deductor counterparty with a genuinely blank name in GSTR-1's own B2B sheet,
    recoverable from GSTR-2A and the BO Profile instead).

    ALSO ADDED (per instruction): the invoice numbers previously crammed into one comma-joined
    cell (up to several dozen for a single busy counterparty-date pair) are now ALSO given a full
    one-row-per-invoice detail table, grouped heading-wise by counterparty+date (same-day repeats)
    or by counterparty (reciprocal pairs) -- purely additive, the original summary rows/tables
    above are unchanged, this is extra detail appended after them."""
    rows, findings = [], []
    out_lines = ctx["g1_lines_fy"]
    in_rows = ctx["twob_lines_fy"]
    names = ctx.get("gstin_name_lookup") or {}

    def _name(gstin, *fallbacks):
        n = names.get(gstin)
        if n:
            return n
        for f in fallbacks:
            if f:
                return f
        return "(name not on file in any source)"

    # --- same-day repeats, outward ---
    by_day = defaultdict(list)
    for x in out_lines:
        if x.get("header_row") and x.get("gstin") and x.get("invdate"):
            by_day[(x["gstin"], _dstr(x["invdate"]))].append(x)
    out_groups = [(k, v) for k, v in by_day.items() if len(v) >= 2]
    out_groups.sort(key=lambda kv: (-len(kv[1]), -sum(y["invval"] for y in kv[1])))
    out_repeat_rows = []
    for (g, d), v in out_groups:
        out_repeat_rows.append([g, _name(g, v[0].get("name"))[:40], d, len(v),
                                round(sum(y["invval"] for y in v), 2),
                                ", ".join(sorted(y["invno"] for y in v))])

    # --- same-day repeats, inward ---
    by_day_in = defaultdict(list)
    for x in in_rows:
        if x.get("gstin") and x.get("date"):
            by_day_in[(x["gstin"], str(x["date"]))].append(x)
    in_groups = [(k, v) for k, v in by_day_in.items() if len(v) >= 2]
    in_groups.sort(key=lambda kv: (-len(kv[1]), -sum(y["invval"] for y in kv[1])))
    in_repeat_rows = []
    for (g, d), v in in_groups:
        in_repeat_rows.append([g, _name(g, v[0].get("supplier"))[:40], d, len(v),
                               round(sum(y["invval"] for y in v), 2),
                               ", ".join(sorted(y["invno"] for y in v))])

    findings.append(Finding(
        "F11", "Same-day repeat transactions with the same counterparty", 
        "REVIEW" if (out_groups or in_groups) else "PASS",
        f"{len(out_groups)} outward and {len(in_groups)} inward counterparty-date combinations "
        f"carry two or more documents on the same day. No value threshold was applied, per the "
        f"scope of this check: every repeat is listed below (Outward / Inward, each its own "
        f"table, sorted by number of same-day documents). Multiple same-day documents to one "
        f"party are ordinary in a high-volume trade and are NOT by themselves an irregularity -- "
        f"the list exists so invoice-splitting (staying under a threshold) and round-tripping can "
        f"be tested against the underlying documents. Full one-row-per-invoice detail for every "
        f"group (not just the comma-joined invoice-number list) is further below, grouped by "
        f"counterparty and date.",
        dict(outward_groups=len(out_groups), inward_groups=len(in_groups))))

    # --- reciprocal pairs -- own table, purchases/sales broken out, not crammed into one cell ---
    sup = defaultdict(lambda: [0.0, 0.0, 0, ""])   # taxable, tax, invoice count, supplier name
    for x in in_rows:
        s = sup[x["gstin"]]
        s[0] += x["taxable"]; s[1] += x["igst"] + x["cgst"] + x["sgst"] + x["cess"]; s[2] += 1
        if not s[3] and x.get("supplier"):
            s[3] = x["supplier"]
    buy = defaultdict(lambda: [0.0, 0.0, set(), ""])   # taxable, tax, invoice numbers, customer name
    for x in out_lines:
        b = buy[x["gstin"]]
        b[0] += x["taxable"]; b[1] += x["igst"] + x["cgst"] + x["sgst"]
        b[2].add(x["invno"])
        if not b[3] and x.get("name"):
            b[3] = x["name"]
    both = sorted(set(sup) & set(buy) - {""},
                  key=lambda g: -(sup[g][1] + buy[g][1]))
    recip_table_rows = []
    for g in both:
        purchases_n, purchases_taxable, purchases_tax = sup[g][2], round(sup[g][0], 2), round(sup[g][1], 2)
        sales_n, sales_taxable, sales_tax = len(buy[g][2]), round(buy[g][0], 2), round(buy[g][1], 2)
        net = round(sales_taxable - purchases_taxable, 2)
        note = ("Net SELLER to this party" if net > 0 else
                "Net BUYER from this party" if net < 0 else "Balanced")
        recip_table_rows.append([g, _name(g, buy[g][3], sup[g][3])[:40], purchases_n, purchases_taxable,
                                 purchases_tax, sales_n, sales_taxable, sales_tax, net, note])
    if both:
        findings.append(Finding(
            "F11a", "Reciprocal counterparties (both supplier and customer)", "REVIEW",
            f"{len(both)} GSTIN(s) appear on BOTH sides of the ledger -- supplying to the taxpayer "
            f"in GSTR-2B and buying from the taxpayer in GSTR-1 within the same year. Two-way "
            f"trading is legitimate in many industries (job work, material returns, group "
            f"companies), but it is also the mechanism by which circular trading inflates turnover "
            f"and credit on both sides without goods moving. Full purchase/sale breakdown for each "
            f"pair is in the 'RECIPROCAL COUNTERPARTIES' table below so the commercial rationale "
            f"can be tested; complete per-invoice detail for each pair follows further below.",
            dict(count=len(both), gstins=", ".join(both[:10]))))

    extra_tables = [
        dict(title="SAME-DAY REPEATS -- INWARD (GSTR-2B)",
             subtitle="Same structure as the outward table above, inward side. No value "
                      "threshold applied -- every repeat is listed.",
             header=["Counterparty GSTIN", "Supplier Name", "Date", "Same-day Documents",
                    "Total Value (Rs)", "Invoice Numbers"],
             widths=[20, 40, 12, 16, 18, 90], rows=in_repeat_rows,
             empty_note="No inward same-day repeat transactions this FY."),
        dict(title="RECIPROCAL COUNTERPARTIES (both supplier and customer)",
             subtitle="Every GSTIN that both sold to the taxpayer (appears in GSTR-2B) "
                      "AND bought from the taxpayer (appears in GSTR-1) within this FY, "
                      "with the purchase and sale legs broken out separately.",
             header=["Counterparty GSTIN", "Name", "Purchase Invoices (from them)",
                    "Purchases Taxable (Rs)", "Purchases Tax (Rs)", "Sale Invoices (to them)",
                    "Sales Taxable (Rs)", "Sales Tax (Rs)", "Net Position (Sales - Purchases, Rs)",
                    "Note"],
             widths=[20, 40, 14, 18, 16, 14, 18, 16, 22, 24], rows=recip_table_rows,
             empty_note="No GSTIN appears on both sides of the ledger this FY."),
    ]

    # ---- NEW: complete per-invoice detail, one heading (its own small table) per counterparty+
    # date group, for same-day repeats -- additive, appended after the summary tables above.
    if out_groups or in_groups:
        extra_tables.append(dict(
            title="SAME-DAY REPEATS -- COMPLETE INVOICE DETAIL (heading per counterparty + date)",
            subtitle="Every group from the two summary tables above, expanded to one row per "
                     "invoice instead of a single comma-joined cell -- so a 49-invoice day can "
                     "be read straight down a column rather than unpacked from one cell by hand.",
            header=["Group", "Direction", "Invoice No", "Date", "Value (Rs)"],
            widths=[46, 10, 22, 14, 18],
            rows=(
                [row for (g, d), v in out_groups
                 for row in ([[f"{_name(g, v[0].get('name'))[:40]} ({g}) -- {d}  [{len(v)} invoices]",
                              "", "", "", ""]]
                             + [["", "Outward", y["invno"], d, round(y["invval"], 2)]
                                for y in sorted(v, key=lambda y: y["invno"])])]
                + [row for (g, d), v in in_groups
                   for row in ([[f"{_name(g, v[0].get('supplier'))[:40]} ({g}) -- {d}  [{len(v)} invoices]",
                                "", "", "", ""]]
                               + [["", "Inward", y["invno"], str(y["date"]), round(y["invval"], 2)]
                                  for y in sorted(v, key=lambda y: y["invno"])])
                   ]),
            empty_note="No same-day repeats this FY.",
        ))

    # ---- NEW: complete per-invoice detail, one heading per reciprocal counterparty ----
    if both:
        recip_out_by_gstin, recip_in_by_gstin = defaultdict(list), defaultdict(list)
        for x in out_lines:
            if x["gstin"] in set(both):
                recip_out_by_gstin[x["gstin"]].append(x)
        for x in in_rows:
            if x["gstin"] in set(both):
                recip_in_by_gstin[x["gstin"]].append(x)
        detail_rows = []
        for g in both:
            detail_rows.append([f"{_name(g, buy[g][3], sup[g][3])[:40]} ({g})", "", "", "", ""])
            for y in sorted(recip_out_by_gstin.get(g, []), key=lambda y: y["invno"]):
                detail_rows.append(["", "Sale (to them)", y["invno"], _dstr(y.get("invdate")), round(y["invval"], 2)])
            for y in sorted(recip_in_by_gstin.get(g, []), key=lambda y: y["invno"]):
                detail_rows.append(["", "Purchase (from them)", y["invno"], str(y.get("date")), round(y["invval"], 2)])
        extra_tables.append(dict(
            title="RECIPROCAL COUNTERPARTIES -- COMPLETE TRANSACTION DETAIL (heading per counterparty)",
            subtitle="Every purchase and sale invoice for every reciprocal counterparty above, "
                     "one row per invoice, so the commercial pattern (steady two-way trade vs "
                     "suspiciously matched value/timing) can be read directly without cross-"
                     "referencing GSTR-1 and GSTR-2B by hand.",
            header=["Counterparty", "Leg", "Invoice No", "Date", "Value (Rs)"],
            widths=[46, 20, 22, 14, 18], rows=detail_rows,
            empty_note="No reciprocal counterparties this FY.",
        ))

    return dict(header=["Counterparty GSTIN", "Name", "Date", "Same-day Documents",
                        "Total Value (Rs)", "Invoice Numbers"],
                widths=[20, 40, 12, 16, 18, 90], rows=out_repeat_rows, findings=findings,
                notes=["OUTWARD same-day repeats (GSTR-1) shown above; INWARD same-day repeats "
                       "(GSTR-2B), RECIPROCAL counterparties, and the complete per-invoice detail "
                       "tables for both are further below. Outward grouping counts invoice HEADER "
                       "rows only, so a multi-rate invoice counts once, not once per rate line."],
                extra_tables=extra_tables)


def _dstr(v):
    if isinstance(v, (_dt.date, _dt.datetime)):
        return v.strftime("%d-%m-%Y")
    return str(v).strip()[:10]


def build_top_counterparties(ctx):
    """F12 -- top 10 suppliers by ITC received and top 10 buyers by tax passed
    on, computed from the returns, then reconciled against the department's own
    BO-Profile top-10 lists (which cover the LAST 12 MONTHS as at the profile
    generation date, not this financial year -- so the two lists are expected
    to differ, and the sheet says so)."""
    rows, findings = [], []
    sup = defaultdict(lambda: [0.0, 0.0, 0, ""])
    for x in ctx["twob_lines_fy"]:
        s = sup[x["gstin"]]
        s[0] += x["igst"] + x["cgst"] + x["sgst"] + x["cess"]
        s[1] += x["taxable"]; s[2] += 1; s[3] = x["supplier"]
    buy = defaultdict(lambda: [0.0, 0.0, set(), ""])
    for x in ctx["g1_lines_fy"]:
        b = buy[x["gstin"]]
        b[0] += x["igst"] + x["cgst"] + x["sgst"]
        b[1] += x["taxable"]; b[2].add(x["invno"]); b[3] = x["name"]

    bo = ctx["annual_data"].get("bo") or {}
    bo_sup = {s["gstin"]: s for s in (bo.get("top_suppliers") or [])}
    bo_ben = {s["gstin"]: s for s in (bo.get("top_beneficiaries") or [])}

    top_sup = sorted(sup.items(), key=lambda kv: -kv[1][0])[:10]
    tot_itc = sum(v[0] for v in sup.values())
    for i, (g, v) in enumerate(top_sup, 1):
        b = bo_sup.get(g)
        rows.append(["Supplier (ITC received)", i, g, v[3][:40], v[0], v[1], v[2],
                     _pct(v[0], tot_itc),
                     (f"BO Profile last-12-months: Rs {b['amount']:,.2f} lakh, risk {b['risk']}"
                      if b else "not in the BO Profile's own top-10 (different period basis)")])
    top_buy = sorted(buy.items(), key=lambda kv: -kv[1][0])[:10]
    tot_tax = sum(v[0] for v in buy.values())
    for i, (g, v) in enumerate(top_buy, 1):
        b = bo_ben.get(g)
        rows.append(["Buyer (tax passed on)", i, g, v[3][:40], v[0], v[1], len(v[2]),
                     _pct(v[0], tot_tax),
                     (f"BO Profile last-12-months: Rs {b['amount']:,.2f} lakh, risk {b['risk']}"
                      if b else "not in the BO Profile's own top-10 (different period basis)")])

    if top_sup:
        g, v = top_sup[0]
        share = _pct(v[0], tot_itc) or 0.0
        sev = "REVIEW" if share >= 10 else "INFO"
        detail = (f"The single largest ITC source for the year is {g} ({v[3]}), contributing "
                  f"{_f(v[0])} of credit across {v[2]} invoice(s) -- {share:,.1f}% of the "
                  f"{_f(tot_itc)} of ITC reflected in GSTR-2B. ")
        risky = [x for x in (bo.get("related_itc_received") or []) if x.get("gstin") == g]
        if risky:
            r = risky[0]
            sev = "FLAG"
            detail += (f"The BO Profile flags this counterparty: status "
                       f"{r.get('status')}, related parameter {r.get('related_parameter')}"
                       + (f", cancelled {r.get('cancellation_date')} "
                          f"({r.get('cancellation_reason')})" if r.get("cancellation_date") else "")
                       + ". A concentrated credit source that the department's own profile has "
                       "flagged warrants supplier-side verification before the credit is accepted.")
        findings.append(Finding("F12", "Concentration of ITC in a single supplier", sev, detail,
                                dict(gstin=g, itc=v[0], share_pct=share)))
    if top_buy:
        g, v = top_buy[0]
        share = _pct(v[0], tot_tax) or 0.0
        findings.append(Finding(
            "F12a", "Concentration of outward supply in a single buyer",
            "REVIEW" if share >= 25 else "INFO",
            f"The largest customer is {g} ({v[3]}), receiving {_f(v[1])} of taxable supply and "
            f"{_f(v[0])} of tax passed on across {len(v[2])} invoice(s) -- {share:,.1f}% of the "
            f"tax passed on for the year.",
            dict(gstin=g, tax_passed=v[0], share_pct=share)))
    return dict(header=["Side", "Rank", "GSTIN", "Name", "Tax / ITC", "Taxable value",
                        "Invoices", "% of total", "Cross-reference to BO Profile"],
                widths=[24, 6, 20, 40, 18, 18, 10, 11, 70], rows=rows, findings=findings,
                notes=["Computed from this FY's own GSTR-2B and GSTR-1. The BO Profile's top-10 "
                       "lists cover the LAST TWELVE MONTHS as at the date that profile was "
                       "generated, which for a past-year scrutiny is a different period entirely "
                       "-- differences between the two lists are expected and are not findings."])


def build_b2b_b2c_shift(ctx):
    """F6 -- turnover moving from B2B to B2C. A B2C sale leaves no ITC trail
    and no counterparty with a reason to report it, so a shift in that
    direction, especially a sudden one, is a standard suppression indicator."""
    rows, findings = [], []
    T = defaultdict(float)
    shares = []
    for m in ctx["months"]:
        g1 = ctx["g1_by_month"].get(m, {})
        b2c = ctx["b2c_by_month"].get(m, {})
        b2b = _n(g1.get("taxable"))
        b2cs, b2cl = _n(b2c.get("b2cs_taxable")), _n(b2c.get("b2cl_taxable"))
        tot = b2b + b2cs + b2cl
        share = _pct(b2cs + b2cl, tot)
        shares.append((m, share))
        rows.append([m, b2b, b2cl, b2cs, tot, share, b2c.get("b2cl_invoices")])
        T["b2b"] += b2b; T["b2cl"] += b2cl; T["b2cs"] += b2cs
    tot = T["b2b"] + T["b2cl"] + T["b2cs"]
    rows.append(["FY TOTAL", T["b2b"], T["b2cl"], T["b2cs"], tot, _pct(T["b2cl"] + T["b2cs"], tot),
                 None])
    b2c_total = T["b2cl"] + T["b2cs"]
    if b2c_total <= TOL:
        findings.append(Finding(
            "F6", "B2B to B2C shift", "PASS",
            f"No B2C turnover is declared in any month: the whole {_f(T['b2b'])} of outward "
            f"taxable supply is B2B, invoice-level and fully traceable to a registered "
            f"counterparty. There is therefore no B2B-to-B2C shift to examine for this year.", {}))
    else:
        jumps = []
        for i in range(1, len(shares)):
            a, b = shares[i - 1][1], shares[i][1]
            if a is not None and b is not None and (b - a) >= 10.0:
                jumps.append(f"{shares[i][0]} ({a:,.1f}% -> {b:,.1f}%)")
        findings.append(Finding(
            "F6", "B2B to B2C shift", "REVIEW" if jumps else "INFO",
            f"B2C turnover for the year is {_f(b2c_total)} ({_pct(b2c_total, tot):,.1f}% of "
            f"outward taxable supply)."
            + (f" Month-on-month jumps of 10 percentage points or more in the B2C share: "
               f"{'; '.join(jumps)}. A B2C sale leaves no input-credit trail and no counterparty "
               f"with an independent reason to report it, so a sudden migration of turnover in "
               f"that direction should be supported by a change in the actual customer base."
               if jumps else " No abrupt month-on-month migration was detected."),
            dict(b2c_total=b2c_total, b2c_share_pct=_pct(b2c_total, tot))))
    return dict(header=["Month", "B2B taxable", "B2C-Large taxable", "B2C-Small taxable",
                        "Total outward taxable", "B2C share %", "B2C-L invoices"],
                widths=[10, 20, 20, 20, 22, 13, 15], rows=rows, findings=findings, notes=[])


# ======================================================================
# SECTION 3 -- SHEET WRITER + ENTRY POINT
# ======================================================================

def _safe_cell(ws, row, col, value):
    """ws.cell() wrapper that guards against a real, confirmed corruption class: openpyxl (like
    Excel itself) treats ANY string cell value starting with '=' as a formula, not text -- it
    sets data_type='f' automatically on assignment. Found one genuine instance of this in this
    tool's own authored note text (ITC Annual Summary's Closing Balances section literally began
    a note with '= this FY's...', which would have made Excel try to parse an English sentence as
    a formula -- a real 'needs repair' cause, confirmed and fixed). The SAME risk exists for any
    raw taxpayer-sourced text this tool writes as-is (a supplier/trade name, invoice number, or
    HSN description that happens to start with '=' in the source data) -- rare, but not
    impossible, and every occurrence anywhere in a taxpayer's real data would hit every sheet that
    quotes it. This wrapper is the systematic guard: if openpyxl auto-detected a formula from a
    plain string value, force the cell's data_type back to string (verified: this preserves the
    exact original text and reloads correctly as text, not a formula) rather than trying to catch
    every individual write site by hand across this codebase."""
    c = ws.cell(row, col, value)
    if c.data_type == "f" and isinstance(value, str):
        c.data_type = "s"
    return c


def _write_table(ws, title, subtitle, built):
    ws.cell(1, 1, title).font = TITLEF
    r = 2
    if subtitle:
        c = ws.cell(r, 1, subtitle)
        c.font = Font(size=9, italic=True)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    r += 1
    hdr_row = r
    for i, h in enumerate(built["header"], 1):
        c = ws.cell(hdr_row, i, h)
        c.fill = HEADFILL; c.font = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        c.border = BORDER
    ws.row_dimensions[hdr_row].height = 32
    r = hdr_row + 1
    # NEW: optional per-cell highlight fills -- e.g. red for a material variance in a specific
    # month's column, without touching any other cell's styling. Purely additive: a built dict
    # with no 'highlight_cells' key renders exactly as before (same "additive, ignored if absent"
    # convention as 'extra_tables' below). Keyed by (0-based index into built['rows'], 1-based
    # column number) -> a PatternFill.
    highlights = built.get("highlight_cells") or {}
    for row_idx, row in enumerate(built["rows"]):
        is_total = isinstance(row[0], str) and row[0].strip().upper().startswith("FY TOTAL")
        for i, v in enumerate(row, 1):
            c = _safe_cell(ws, r, i, v)
            c.border = BORDER
            if isinstance(v, float):
                c.number_format = "#,##0.00"
            if is_total:
                c.font = Font(bold=True)
                c.fill = SECTFILL
            else:
                c.font = Font(size=10)
            hl = highlights.get((row_idx, i))
            if hl is not None:
                c.fill = hl
                c.font = Font(bold=True, size=10)
            c.alignment = Alignment(vertical="top",
                                    wrap_text=(i == len(row) and isinstance(v, str)))
        r += 1
    r += 1
    for n in built.get("notes", []):
        c = _safe_cell(ws, r, 1, "Note: " + n)
        c.font = Font(size=9, italic=True, color="808080")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    # NEW: optional additional labelled tables on the SAME sheet, after the primary table/notes
    # and before the findings summary -- e.g. Counterparty Transactions uses this to give
    # reciprocal counterparties (and each direction of same-day-repeats) their own clean table
    # with columns suited to that data, instead of everything crammed into the primary table's
    # generic column set. Purely additive: a built dict with no "extra_tables" key renders
    # exactly as before.
    for extra in built.get("extra_tables", []):
        r += 1
        if extra.get("title"):
            ws.cell(r, 1, extra["title"]).font = Font(bold=True, size=12, color="1F3864")
            r += 1
        if extra.get("subtitle"):
            c = ws.cell(r, 1, extra["subtitle"])
            c.font = Font(size=9, italic=True)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            r += 1
        r += 1
        ehdr_row = r
        for i, h in enumerate(extra["header"], 1):
            c = ws.cell(ehdr_row, i, h)
            c.fill = HEADFILL; c.font = Font(bold=True, color="FFFFFF", size=10)
            c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            c.border = BORDER
        ws.row_dimensions[ehdr_row].height = 32
        r = ehdr_row + 1
        if not extra["rows"]:
            ws.cell(r, 1, extra.get("empty_note", "None.")).font = Font(italic=True)
            r += 1
        for row in extra["rows"]:
            for i, v in enumerate(row, 1):
                c = _safe_cell(ws, r, i, v)
                c.border = BORDER
                c.font = Font(size=10)
                if isinstance(v, float):
                    c.number_format = "#,##0.00"
                c.alignment = Alignment(vertical="top",
                                        wrap_text=(i == len(row) and isinstance(v, str)))
            r += 1
        for j, w in enumerate(extra.get("widths", []), 1):
            col = get_column_letter(j)
            cur = ws.column_dimensions[col].width or 0
            if w > cur:
                ws.column_dimensions[col].width = w
        r += 1
    if built.get("findings"):
        r += 1
        ws.cell(r, 1, "FINDINGS FROM THIS SHEET").font = Font(bold=True, size=11, color="1F3864")
        r += 1
        for h, i in (("Ref", 1), ("Check", 2), ("Result", 3), ("Detail", 4)):
            c = ws.cell(r, i, h)
            c.fill = HEADFILL; c.font = Font(bold=True, color="FFFFFF", size=10)
            c.border = BORDER
        r += 1
        for f in built["findings"]:
            ws.cell(r, 1, f.ref).border = BORDER
            ws.cell(r, 2, f.title).border = BORDER
            cv = ws.cell(r, 3, f.severity)
            cv.fill = SEV_FILL.get(f.severity, GREY)
            cv.font = SEV_FONT.get(f.severity, Font(bold=True))
            cv.alignment = Alignment(horizontal="center")
            cv.border = BORDER
            d = _safe_cell(ws, r, 4, f.detail)
            d.alignment = Alignment(wrap_text=True, vertical="top")
            d.font = Font(size=10)
            d.border = BORDER
            ws.row_dimensions[r].height = max(15, min(120, 13 * (len(f.detail) // 110 + 1)))
            r += 1
    for i, w in enumerate(built.get("widths", []), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(hdr_row + 1, 1)


# ======================================================================
# SECTION -- GSTR-2A CROSS-CHECKS (G1-G10, NEW)
# ======================================================================
# Merges GSTR-2A into the existing tool per the originating brief's checklist,
# without altering any pre-existing comparison. See gst_parsers_dept.parse_r2a_excel()'s
# docstring for the two real-data quirks (the '-Total' rollup rows, and B2BA/CDNRA's
# repeated sub-header) that drive the parsing side; this section is the check/output side.
#
# Feasibility notes (verified against real data, not assumed):
#   - GSTR-2A's B2B/CDNR sheets carry NO "ITC Availability"/"Reason" column at all -- that
#     determination is 2B-only. G2/G2a use 2B's own flag when a matched invoice exists there.
#   - GSTR-2A carries NO HSN/SAC column anywhere -- blocked-credit HSN screening (point 7 of
#     the brief) is not buildable from this source; recorded as an explicit SKIPPED finding
#     (G-NF1) rather than silently omitted.
#   - "Effective date of cancellation" in 2A is the COUNTERPARTY'S GSTIN cancellation date, not
#     per-invoice IRN-cancellation status -- G10 is scoped to what the field actually means.
#   - GSTR-3B carries no invoice-level data, so a per-invoice ITC "claim date" cannot be traced;
#     G9 checks the SUPPLIER's own filing date against the section 16(4) deadline instead, and
#     says so plainly in its own finding text.


def _r2a_2b_key(gstin, invno, invdate_iso, invtype_canon):
    return (str(gstin or "").strip().upper(), str(invno or "").strip().upper(),
            invdate_iso, invtype_canon)


def _compute_r2a_vs_2b(ctx):
    """Invoice-level match between GSTR-2A B2B and GSTR-2B B2B, computed ONCE and cached on ctx
    -- the monthly-summary sheet (G2a) and the invoice-mismatch sheet (G2) both need it, and it
    is not cheap to redo (tens of thousands of rows on each side). Match key: GSTIN + invoice
    number + invoice date + invoice type (point 6 of the brief -- SEZ vs Regular must not
    collide on a reused invoice number)."""
    cache_key = "_r2a_vs_2b_cache"
    if cache_key in ctx:
        return ctx[cache_key]
    r2a = ctx.get("r2a_data") or {}
    result = dict(available=r2a.get("available", False), reason=r2a.get("reason"),
                  a_side={}, b_side={}, monthly=[])
    if not r2a.get("available"):
        ctx[cache_key] = result
        return result

    a_side = {}
    for month, recs in r2a.get("b2b", {}).items():
        for x in recs:
            k = _r2a_2b_key(x["gstin"], x["invno"], x["invdate"], x["invtype"])
            e = a_side.setdefault(k, dict(month=month, gstin=x["gstin"], supplier=x["supplier"],
                                          invno=x["invno"], invdate=x["invdate"],
                                          taxable=0.0, tax=0.0))
            e["taxable"] += x["taxable"]
            e["tax"] += x["igst"] + x["cgst"] + x["sgst"] + x["cess"]

    b_side = {}
    for m in ctx["months"]:
        two = ctx["twob_by_month"].get(m, {})
        if not two.get("available"):
            continue
        for x in two.get("rows", []):
            invdate_iso = dept.r2a_clean_date(x.get("date"))
            invtype_c = dept.r2a_invtype_canon(x.get("invtype"))
            k = _r2a_2b_key(x.get("gstin"), x.get("invno"), invdate_iso, invtype_c)
            e = b_side.setdefault(k, dict(month=m, gstin=x.get("gstin"), supplier=x.get("supplier"),
                                          invno=x.get("invno"), invdate=invdate_iso,
                                          taxable=0.0, tax=0.0,
                                          itc_avail=x.get("itc_avail", ""),
                                          reason=x.get("itc_avail_reason", "")))
            e["taxable"] += _n(x.get("taxable"))
            e["tax"] += _n(x.get("igst")) + _n(x.get("cgst")) + _n(x.get("sgst")) + _n(x.get("cess"))

    result["a_side"] = a_side
    result["b_side"] = b_side

    monthly = []
    cum_a_taxable = cum_b_taxable = cum_a_tax = cum_b_tax = 0.0
    for m in ctx["months"]:
        a_taxable = sum(e["taxable"] for e in a_side.values() if e["month"] == m)
        a_tax = sum(e["tax"] for e in a_side.values() if e["month"] == m)
        b_taxable = sum(e["taxable"] for e in b_side.values() if e["month"] == m)
        b_tax = sum(e["tax"] for e in b_side.values() if e["month"] == m)
        cum_a_taxable += a_taxable; cum_a_tax += a_tax
        cum_b_taxable += b_taxable; cum_b_tax += b_tax
        monthly.append(dict(month=m, a_taxable=a_taxable, a_tax=a_tax, b_taxable=b_taxable,
                            b_tax=b_tax, cum_a_taxable=cum_a_taxable, cum_a_tax=cum_a_tax,
                            cum_b_taxable=cum_b_taxable, cum_b_tax=cum_b_tax,
                            cum_diff_taxable=cum_a_taxable - cum_b_taxable,
                            cum_diff_tax=cum_a_tax - cum_b_tax))
    result["monthly"] = monthly
    ctx[cache_key] = result
    return result


def build_r2a_data_quality(ctx):
    """G1 -- duplicate/reused invoice numbers in GSTR-2A B2B, plus GSTIN-hygiene and
    rows-excluded-for-safety transparency, plus the explicit not-feasible note for
    blocked-credit HSN screening (point 7 of the brief)."""
    r2a = ctx.get("r2a_data") or {}
    rows, findings = [], []
    if not r2a.get("available"):
        findings.append(Finding("G1", "Duplicate/reused invoice number in GSTR-2A", "SKIPPED",
                                r2a.get("reason") or "GSTR-2A not supplied.", {}))
        return dict(header=["Status"], widths=[100], rows=[["GSTR-2A not supplied -- see findings."]],
                    findings=findings, notes=[])

    groups = defaultdict(list)
    for month, recs in r2a.get("b2b", {}).items():
        for x in recs:
            groups[(x["gstin"], x["invno"])].append(x)
    dupes = 0
    for (gstin, invno), items in groups.items():
        if len(items) < 2:
            continue
        distinct = {(round(it["taxable"] + it["igst"] + it["cgst"] + it["sgst"] + it["cess"], 2),
                    it["invdate"]) for it in items}
        if len(distinct) > 1:
            dupes += 1
            # NEW: occurrence# (1st/2nd/3rd... reuse of this exact GSTIN+invoice-number combo,
            # in month order) and an exact-vs-differs flag, so supplier-level concentration
            # (e.g. one supplier reusing several different invoice numbers) is visible without
            # manually cross-referencing rows -- per this session's annual-detail prompt.
            items_sorted = sorted(items, key=lambda it: (it["month"], it["invdate"] or ""))
            seen_vt = set()
            for occ, it in enumerate(items_sorted, 1):
                vt = (round(it["taxable"] + it["igst"] + it["cgst"] + it["sgst"] + it["cess"], 2), it["invdate"])
                is_exact_repeat = vt in seen_vt
                seen_vt.add(vt)
                rows.append([it["month"], gstin, it["supplier"], invno, it["invdate"],
                            it["taxable"], it["igst"] + it["cgst"] + it["sgst"] + it["cess"],
                            occ, "Exact repeat" if is_exact_repeat else "Value/date DIFFERS"])
    findings.append(Finding(
        "G1", "Duplicate/reused invoice number in GSTR-2A",
        "REVIEW" if dupes else "PASS",
        (f"{dupes} distinct (supplier GSTIN, invoice number) combination(s) appear more than "
         f"once in GSTR-2A with a DIFFERENT value or date across the occurrences -- possible "
         f"duplicate claim, re-upload, or data manipulation. Genuine repeats (identical value "
         f"and date) are not counted.") if dupes else
        "No supplier GSTIN + invoice number combination repeats with a differing value or date "
        "in GSTR-2A.",
        dict(count=dupes)))

    tot_missing = sum(len(v) for v in r2a.get("total_row_missing", {}).values())
    if tot_missing:
        detail = "; ".join(f"{sheet}: {len(items)}" for sheet, items in r2a["total_row_missing"].items())
        findings.append(Finding(
            "G1-DQ1", "GSTR-2A rate-line rows without a matching '-Total' summary row", "INFO",
            f"{tot_missing} document number(s) across the affected sheet(s) ({detail}) had a "
            f"rate-line row but no matching government '-Total' rollup row, so they could not be "
            f"safely summed and are EXCLUDED from every GSTR-2A check in this workbook -- the "
            f"complete list (not just the count) is in the 'GSTR-2A Rows Excluded' table below.", {}))
    if r2a.get("malformed_gstin"):
        findings.append(Finding(
            "G1-DQ2", "Malformed GSTIN strings in GSTR-2A", "INFO",
            f"{len(r2a['malformed_gstin'])} GSTIN string(s) read from GSTR-2A did not match the "
            f"standard 15-character shape even after upper-casing and whitespace-stripping -- the "
            f"complete list is in the 'Malformed GSTIN Strings' table below. Still used as-is for "
            f"every downstream match -- a failed join caused by this would surface as a false "
            f"ONLY-IN-2A/ONLY-IN-2B row rather than a genuine anomaly.", {}))
    findings.append(Finding(
        "G-NF1", "Blocked-credit (Section 17(5)) HSN screening from GSTR-2A", "SKIPPED",
        "GSTR-2A's B2B/CDNR sheets carry no HSN/SAC column at all (confirmed against both the "
        "live header row and the workbook's own Read-me field list), so there is nothing to "
        "filter on from this source. HSN-based blocked-credit screening already exists elsewhere "
        "in this tool (check A5, HSN & Fraud Pattern Checks sheet) using GSTR-1/2B's own HSN data.",
        {}))
    extra_tables = []
    if tot_missing:
        excl_rows = [[sheet, item] for sheet, items in r2a["total_row_missing"].items() for item in items]
        extra_tables.append(dict(
            title=f"GSTR-2A Rows Excluded -- No Matching '-Total' Row ({tot_missing} document(s))",
            subtitle="Complete list of the document numbers referenced in G1-DQ1 above -- every "
                     "one of these is excluded from every GSTR-2A check in this workbook, not "
                     "just summarised as a count.",
            header=["Sheet", "Month: Document No"], widths=[26, 40], rows=excl_rows,
            empty_note="None."))
    if r2a.get("malformed_gstin"):
        extra_tables.append(dict(
            title=f"Malformed GSTIN Strings in GSTR-2A ({len(r2a['malformed_gstin'])})",
            subtitle="Complete list referenced in G1-DQ2 above.",
            header=["GSTIN string as read"], widths=[30],
            rows=[[g] for g in r2a["malformed_gstin"]], empty_note="None."))
    return dict(header=["Month", "GSTIN", "Supplier", "Invoice No", "Invoice Date", "Taxable Value",
                        "Total Tax", "Occurrence # (this GSTIN+Invoice No)", "Repeat Type"],
                widths=[10, 18, 30, 20, 14, 16, 14, 26, 20], rows=rows, findings=findings,
                notes=["Every GSTIN read from GSTR-2A is upper-cased and whitespace-stripped "
                       "before use in any comparison anywhere in this workbook's GSTR-2A checks."],
                extra_tables=extra_tables)


def build_r2a_vs_2b_monthly(ctx):
    """G2a -- monthly and cumulative running-total comparison, the wash-out view (point 9 of
    the brief) that separates ordinary timing noise from a genuine net excess claim."""
    m = _compute_r2a_vs_2b(ctx)
    findings = []
    if not m["available"]:
        findings.append(Finding("G2a", "GSTR-2A vs GSTR-2B -- monthly cumulative", "SKIPPED",
                                m["reason"] or "GSTR-2A not supplied.", {}))
        return dict(header=["Status"], widths=[100], rows=[["GSTR-2A not supplied -- see findings."]],
                    findings=findings, notes=[])
    rows = []
    for r in m["monthly"]:
        rows.append([r["month"], r["a_taxable"], r["a_tax"], r["b_taxable"], r["b_tax"],
                    r["cum_a_taxable"], r["cum_b_taxable"], r["cum_diff_taxable"], r["cum_diff_tax"]])
    last = m["monthly"][-1] if m["monthly"] else None
    fy_diff_taxable = last["cum_diff_taxable"] if last else 0.0
    fy_diff_tax = last["cum_diff_tax"] if last else 0.0
    sev = "PASS" if (last and abs(fy_diff_taxable) <= MATERIAL) else "REVIEW"
    findings.append(Finding(
        "G2a", "GSTR-2A vs GSTR-2B -- cumulative running total", sev,
        (f"By FY end, cumulative 2A taxable value {_f(last['cum_a_taxable'])} against cumulative "
         f"2B taxable value {_f(last['cum_b_taxable'])}; running difference {_f(fy_diff_taxable)} "
         f"(tax difference {_f(fy_diff_tax)}). A small residual difference here despite larger "
         f"MONTH-BY-MONTH swings (see the invoice-level mismatch sheet) usually means the "
         f"mismatches are TIMING -- a late-filed or amended invoice landing in a different "
         f"month's 2A than its matching 2B period -- and wash out over the year rather than "
         f"representing a genuine excess claim.") if last else "No months to compare.",
        dict(cumulative_diff_taxable=fy_diff_taxable, cumulative_diff_tax=fy_diff_tax)))
    return dict(header=["Month", "2A taxable (this month)", "2A tax", "2B taxable (this month)",
                        "2B tax", "2A cumulative taxable", "2B cumulative taxable",
                        "Cumulative taxable diff", "Cumulative tax diff"],
                widths=[10, 20, 14, 20, 14, 20, 20, 18, 16], rows=rows, findings=findings,
                notes=["'2A (this month)' groups by the marker/tax-period each document sits "
                       "under in this GSTR-2A pull, not by invoice date."])


def build_r2a_vs_2b_detail(ctx):
    """G2 -- invoice-level existence and value comparison, matched on GSTIN + invoice number +
    invoice date + invoice type (point 6 of the brief)."""
    m = _compute_r2a_vs_2b(ctx)
    rows, findings = [], []
    if not m["available"]:
        findings.append(Finding("G2", "GSTR-2A vs GSTR-2B -- invoice existence & value", "SKIPPED",
                                m["reason"] or "GSTR-2A not supplied.", {}))
        return dict(header=["Status"], widths=[100], rows=[["GSTR-2A not supplied -- see findings."]],
                    findings=findings, notes=[])
    a_side, b_side = m["a_side"], m["b_side"]
    all_keys = set(a_side) | set(b_side)
    only_2a = only_2b = both_same = both_diff = 0
    CAP = 3000
    # Build the three exception buckets separately first, so the CAP (when it binds) preferentially
    # keeps the categories worth an officer's attention (VALUE DIFFERS, ONLY IN 2B) rather than
    # letting whichever category sorts first in GSTIN order silently crowd them out.
    diff_rows, only2b_rows, only2a_rows = [], [], []
    for k in sorted(all_keys, key=lambda x: (x[0] or "", x[1] or "")):
        a, b = a_side.get(k), b_side.get(k)
        if a and not b:
            only_2a += 1
            only2a_rows.append([a["month"], a["gstin"], a["supplier"], a["invno"], a["invdate"],
                                a["taxable"], a["tax"], None, None, "ONLY IN 2A",
                                "Reported by supplier (2A) but not in 2B for this period -- likely "
                                "lands in a later 2B (post-freeze late filing), or was never "
                                "auto-drafted."])
        elif b and not a:
            only_2b += 1
            only2b_rows.append([b["month"], b["gstin"], b["supplier"], b["invno"], b["invdate"],
                                None, None, b["taxable"], b["tax"], "ONLY IN 2B",
                                f"In 2B but not (any longer) in this 2A pull -- supplier may have "
                                f"since cancelled/amended the invoice; reversal may be due if ITC "
                                f"was already claimed. 2B ITC availability: "
                                f"{b.get('itc_avail') or 'n/a'}" +
                                (f" ({b['reason']})" if b.get("reason") else "")])
        else:
            dv = round(a["taxable"] - b["taxable"], 2)
            dt = round(a["tax"] - b["tax"], 2)
            if abs(dv) <= TOL and abs(dt) <= TOL:
                both_same += 1
            else:
                both_diff += 1
                diff_rows.append([a["month"], a["gstin"], a["supplier"], a["invno"], a["invdate"],
                                  a["taxable"], a["tax"], b["taxable"], b["tax"], "VALUE DIFFERS",
                                  f"Taxable diff {_f(dv)}, tax diff {_f(dt)} -- likely a post-2B-"
                                  f"freeze supplier amendment; 2B's frozen value is what could "
                                  f"legally be claimed for that period. 2B ITC availability: "
                                  f"{b.get('itc_avail') or 'n/a'}" +
                                  (f" ({b['reason']})" if b.get("reason") else "")])
    # Priority order for the cap: VALUE DIFFERS and ONLY IN 2B first (these are the categories the
    # finding text calls "worth an officer's attention"); ONLY IN 2A (ordinary timing) fills
    # whatever budget remains.
    rows = (diff_rows + only2b_rows + only2a_rows)[:CAP]
    total_docs = len(all_keys)
    capped_note = (f" Row list capped at {CAP} for readability, prioritising VALUE-DIFFERS and "
                   f"ONLY-IN-2B rows over ONLY-IN-2A; {only_2a + only_2b + both_diff - CAP} "
                   f"further exception(s) exist but are not listed individually (all are counted "
                   f"in the totals above).") if (only_2a + only_2b + both_diff) > CAP else ""
    findings.append(Finding(
        "G2", "GSTR-2A vs GSTR-2B -- invoice existence and value",
        "REVIEW" if (only_2b or both_diff) else ("INFO" if only_2a else "PASS"),
        f"Of {total_docs} distinct (GSTIN, invoice no., date, invoice type) documents across both "
        f"sources: {both_same} match exactly, {both_diff} match on key but differ in value, "
        f"{only_2a} appear only in 2A (not yet in 2B for that period -- ordinary timing), "
        f"{only_2b} appear only in 2B (no longer in this 2A pull -- worth checking whether ITC "
        f"already claimed on these needs reversal)." + capped_note,
        dict(only_2a=only_2a, only_2b=only_2b, both_same=both_same, both_diff=both_diff)))
    if r2a_malformed := (ctx.get("r2a_data") or {}).get("malformed_gstin"):
        findings.append(Finding(
            "G2-DQ", "Malformed GSTIN strings affecting this match", "INFO",
            f"{len(r2a_malformed)} GSTIN string(s) from GSTR-2A did not match the standard "
            f"15-character shape; see the GSTR-2A Data Quality sheet for the full list.", {}))
    return dict(header=["Month", "GSTIN", "Supplier", "Invoice No", "Invoice Date", "2A Taxable",
                        "2A Tax", "2B Taxable", "2B Tax", "Status", "Note"],
                widths=[10, 18, 26, 18, 14, 14, 12, 14, 12, 14, 55], rows=rows, findings=findings,
                notes=["Only exceptions (ONLY IN 2A / ONLY IN 2B / VALUE DIFFERS) are listed; "
                       "rows that match exactly are counted in the finding above but not printed "
                       "here to keep this sheet readable."])


def build_r2a_vs_3b_and_ledger(ctx):
    """G3 -- cumulative FY-to-date GSTR-2A (non-RCM) availability vs GSTR-3B Table 4A(5) claimed
    (Rule 36(4)/section 16(2)(aa)). G4 -- the same 2A-derived availability vs what was actually
    credited to the Electronic Credit Ledger."""
    r2a = ctx.get("r2a_data") or {}
    rows, findings = [], []
    if not r2a.get("available"):
        findings.append(Finding("G3", "GSTR-2A vs GSTR-3B cumulative claim", "SKIPPED",
                                r2a.get("reason") or "GSTR-2A not supplied.", {}))
        findings.append(Finding("G4", "GSTR-2A vs Electronic Credit Ledger", "SKIPPED",
                                r2a.get("reason") or "GSTR-2A not supplied.", {}))
        return dict(header=["Status"], widths=[100], rows=[["GSTR-2A not supplied -- see findings."]],
                    findings=findings, notes=[])

    cl_month = (ctx["annual_data"].get("credit") or {}).get("monthly_by_tax_period", {})
    cum_2a = cum_3b = cum_led = 0.0
    for m in ctx["months"]:
        a_taxable = a_tax = 0.0
        for x in r2a.get("b2b", {}).get(m, []):
            if x.get("rcm", "").upper().startswith("Y"):
                continue
            a_taxable += x["taxable"]; a_tax += x["igst"] + x["cgst"] + x["sgst"] + x["cess"]
        g3b = ctx["g3b_by_month"].get(m, {})
        b_tax = sum(_n(v) for v in (g3b.get("4A5") or [])[:4])
        led = cl_month.get(m, {})
        led_cr = led.get("credited")
        cum_2a += a_tax; cum_3b += b_tax
        cum_led += _n(led_cr)
        rows.append([m, a_taxable, a_tax, b_tax, (b_tax - a_tax), led_cr,
                    cum_2a, cum_3b, cum_led, cum_3b - cum_2a])
    rows.append(["FY TOTAL", None, cum_2a, cum_3b, cum_3b - cum_2a, cum_led, None, None, None, None])

    d_fy = cum_3b - cum_2a
    findings.append(Finding(
        "G3", "GSTR-2A (non-RCM) vs GSTR-3B Table 4A(5), cumulative FY-to-date",
        "FLAG" if d_fy > MATERIAL else ("REVIEW" if abs(d_fy) > TOL else "PASS"),
        f"Cumulative non-RCM tax available per GSTR-2A B2B {_f(cum_2a)} against cumulative "
        f"'All other ITC' claimed under Table 4A(5) {_f(cum_3b)}; difference {_f(d_fy)}. Where "
        f"4A(5) exceeds 2A's cumulative availability by more than ordinary timing noise from late "
        f"supplier filing, that excess is a Rule 36(4)/section 16(2)(aa) exposure -- the LEGAL "
        f"benchmark for eligibility is GSTR-2B, not 2A (2A is dynamic and pre-determination); "
        f"this comparison uses 2A specifically because it shows whether a supplier had filed AT "
        f"ALL by the time of claim, a fact 2B alone doesn't reveal.",
        dict(cumulative_2a=cum_2a, cumulative_3b_4a5=cum_3b, difference=d_fy)))

    if cum_led:
        missing, wnote = ledger_window_note(ctx["months"], set(cl_month), "Electronic Credit Ledger")
        cmp_months = [m for m in ctx["months"] if m not in missing]
        a_cmp = sum(_n(x["igst"] + x["cgst"] + x["sgst"] + x["cess"])
                    for m in cmp_months for x in r2a.get("b2b", {}).get(m, [])
                    if not x.get("rcm", "").upper().startswith("Y"))
        led_cmp = sum(_n(cl_month.get(m, {}).get("credited")) for m in cmp_months)
        d2 = led_cmp - a_cmp
        findings.append(Finding(
            "G4", "GSTR-2A (non-RCM) availability vs Electronic Credit Ledger credits",
            "REVIEW" if d2 > MATERIAL else "PASS",
            f"Compared over the {len(cmp_months)} month(s) the ledger actually covers: non-RCM "
            f"tax available per GSTR-2A {_f(a_cmp)} against {_f(led_cmp)} actually credited to "
            f"the Electronic Credit Ledger; difference {_f(d2)}. Credit-ledger totals will "
            f"structurally exceed 2A-derived figures by whatever the ledger carries from ISD, "
            f"TDS/TCS, import credit or refund re-credit -- none of which flow through 2A's B2B "
            f"sheet -- so this looks at the direction and materiality of the gap, not an exact "
            f"match. " + wnote,
            dict(r2a_available_in_window=a_cmp, ledger_credited_in_window=led_cmp, difference=d2)))
    else:
        findings.append(Finding("G4", "GSTR-2A (non-RCM) availability vs Electronic Credit Ledger credits",
                                "SKIPPED", "Electronic Credit Ledger not supplied.", {}))

    return dict(header=["Month", "2A taxable (non-RCM)", "2A tax (non-RCM)", "3B 4A(5) claimed",
                        "4A(5) minus 2A", "Credit ledger credited", "Cum. 2A tax", "Cum. 3B 4A(5)",
                        "Cum. ledger credited", "Cum. 3B minus 2A"],
                widths=[10, 18, 16, 16, 14, 20, 14, 14, 18, 16], rows=rows, findings=findings,
                notes=["2A figures here exclude RCM-flagged rows (RCM ITC is claimed under "
                       "4A(3), triangulated separately on the RCM & State-Code sheet) so they "
                       "compare on the same basis as 4A(5) 'All other ITC'."])


def build_r2a_amendments(ctx):
    """G5 -- amended invoices (B2BA) linked back to their original GSTR-2A B2B entry."""
    r2a = ctx.get("r2a_data") or {}
    rows, findings = [], []
    if not r2a.get("available"):
        findings.append(Finding("G5", "GSTR-2A amendment tracking (B2BA)", "SKIPPED",
                                r2a.get("reason") or "GSTR-2A not supplied.", {}))
        return dict(header=["Status"], widths=[100], rows=[["GSTR-2A not supplied -- see findings."]],
                    findings=findings, notes=[])

    b2b_index = defaultdict(list)
    for m, recs in r2a.get("b2b", {}).items():
        for x in recs:
            b2b_index[(x["gstin"], x["invno"])].append(x)

    total_amend = linked = 0
    delta_taxable_total = 0.0
    type_counts = defaultdict(int)
    for m, recs in r2a.get("b2ba", {}).items():
        for x in recs:
            total_amend += 1
            type_counts[x.get("amend_type") or "(blank)"] += 1
            orig_matches = b2b_index.get((x["gstin"], x["orig_invno"]), [])
            orig_taxable = sum(o["taxable"] for o in orig_matches) if orig_matches else None
            delta = (x["taxable"] - orig_taxable) if orig_taxable is not None else None
            if orig_matches:
                linked += 1
                delta_taxable_total += (delta or 0.0)
            rows.append([m, x["gstin"], x["supplier"], x["orig_invno"], x["invno"],
                        x.get("amend_type"), x["orig_tax_period"], x["invdate"],
                        orig_taxable, x["taxable"], delta,
                        "Linked to B2B" if orig_matches else
                        "Original not found in current B2B (may have amended out of a different "
                        "tax period, or the original itself carries an unmatched invoice-number "
                        "change)"])
    findings.append(Finding(
        "G5", "GSTR-2A amendment tracking (B2BA linked to original B2B)",
        "REVIEW" if total_amend else "PASS",
        (f"{total_amend} amended invoice(s) in GSTR-2A this year, of which {linked} could be "
         f"linked back to their original B2B entry by (supplier GSTIN, original invoice number). "
         f"Amendment types: " + ", ".join(f"{k}={v}" for k, v in sorted(type_counts.items())) +
         f". Net taxable-value delta on linked amendments {_f(delta_taxable_total)}.")
        if total_amend else "No amended invoices (B2BA) in GSTR-2A this year.",
        dict(total=total_amend, linked=linked, net_delta_taxable=delta_taxable_total)))
    return dict(header=["Month", "GSTIN", "Supplier", "Original Inv No", "Revised Inv No",
                        "Amendment Type", "Original Tax Period", "Revised Date",
                        "Original Taxable (B2B)", "Revised Taxable (B2BA)", "Delta", "Link Status"],
                widths=[10, 18, 26, 18, 18, 16, 16, 14, 18, 18, 14, 45], rows=rows,
                findings=findings,
                notes=["GSTR-2A shows only the CURRENT state of an invoice under one B2B entry, "
                       "so where an amendment did not change the invoice number, the B2B sheet's "
                       "own row already reflects the amended figures directly -- B2BA specifically "
                       "surfaces amendments GSTN itself logs as a distinct event (GSTIN changed, "
                       "invoice number changed, or other details changed), a narrower set than "
                       "every value correction a supplier makes."])


def build_r2a_rcm_and_statecode(ctx):
    """G6 -- GSTR-2A RCM-flagged rows vs Table 3.1(d) liability and 4A(3) ITC.
    G7 -- state-code vs tax-head (IGST vs CGST+SGST) validation.

    Complete invoice-level detail added on request (extra_tables, additive --
    the primary G6/G7 table/findings above are unchanged): previously G6 only
    showed month totals (no invoice list at all) and G7 only listed the
    MISMATCHED invoices, not every invoice actually checked."""
    r2a = ctx.get("r2a_data") or {}
    rows, findings = [], []
    if not r2a.get("available"):
        findings.append(Finding("G6", "GSTR-2A RCM cross-check", "SKIPPED",
                                r2a.get("reason") or "GSTR-2A not supplied.", {}))
        findings.append(Finding("G7", "GSTR-2A state-code (IGST vs CGST+SGST) validation", "SKIPPED",
                                r2a.get("reason") or "GSTR-2A not supplied.", {}))
        return dict(header=["Status"], widths=[100], rows=[["GSTR-2A not supplied -- see findings."]],
                    findings=findings, notes=[])

    self_gstin = ctx.get("self_gstin") or ""
    self_state = self_gstin[:2] if len(self_gstin) >= 2 else None

    rcm_2a_taxable, rcm_2a_tax = defaultdict(float), defaultdict(float)
    rcm_detail_rows = []
    for m, recs in r2a.get("b2b", {}).items():
        for x in recs:
            if x.get("rcm", "").upper().startswith("Y"):
                rcm_2a_taxable[m] += x["taxable"]
                rcm_2a_tax[m] += x["igst"] + x["cgst"] + x["sgst"] + x["cess"]
                rcm_detail_rows.append([m, x["gstin"], x["supplier"], x["invno"], x.get("invdate"),
                                        x["taxable"], x["igst"], x["cgst"], x["sgst"], x["cess"]])
    fy_rcm_2a_tax = sum(rcm_2a_tax.values())
    fy_rcm_3b_liab = fy_rcm_3b_itc = 0.0
    for m in ctx["months"]:
        g3b = ctx["g3b_by_month"].get(m, {})
        d = g3b.get("3.1d") or []
        fy_rcm_3b_liab += sum(_n(v) for v in d[1:4]) if len(d) > 3 else 0.0
        fy_rcm_3b_itc += sum(_n(v) for v in (g3b.get("4A3") or [])[:4])
    for m in ctx["months"]:
        rows.append(["RCM", m, None, None, None, None, rcm_2a_taxable.get(m, 0.0),
                    rcm_2a_tax.get(m, 0.0)])
    d_rcm = fy_rcm_3b_liab - fy_rcm_2a_tax
    findings.append(Finding(
        "G6", "GSTR-2A RCM-flagged supplies vs 3.1(d) liability & 4A(3) ITC",
        "REVIEW" if abs(d_rcm) > MATERIAL else "PASS",
        f"GSTR-2A flags {_f(fy_rcm_2a_tax)} of tax on inward supplies marked 'Supply Attract "
        f"Reverse Charge = Y' for the year, against {_f(fy_rcm_3b_liab)} RCM liability declared "
        f"in Table 3.1(d) and {_f(fy_rcm_3b_itc)} RCM ITC claimed under 4A(3); 2A-vs-3.1(d) "
        f"difference {_f(d_rcm)}. 2A's RCM flag reflects the SUPPLIER's own declaration and is "
        f"informational only -- RCM tax must be self-assessed and paid by the recipient "
        f"regardless of what the supplier marks, so this is a completeness cross-check on "
        f"3.1(d), not itself an eligibility test. See the RCM Triangulation sheet (F8/F8a) for "
        f"the authoritative cash-ledger-verified comparison. {len(rcm_detail_rows)} invoice(s) "
        f"carry the RCM flag -- see the 'RCM-Flagged GSTR-2A Invoices' table below.",
        dict(rcm_2a_tax=fy_rcm_2a_tax, rcm_3b_liability=fy_rcm_3b_liab, rcm_3b_itc=fy_rcm_3b_itc)))

    mismatches = checked = 0
    statecode_detail_rows = []
    if self_state:
        for m, recs in r2a.get("b2b", {}).items():
            for x in recs:
                sup_state = (x["gstin"] or "")[:2]
                if not sup_state or len(sup_state) != 2 or not sup_state.isdigit():
                    continue
                checked += 1
                same_state = (sup_state == self_state)
                has_igst = x["igst"] > TOL
                has_cgst_sgst = (x["cgst"] > TOL or x["sgst"] > TOL)
                wrong = (same_state and has_igst) or ((not same_state) and has_cgst_sgst and not has_igst)
                tax_head = ("IGST" if has_igst else "") + ("+" if has_igst and has_cgst_sgst else "") \
                           + ("CGST+SGST" if has_cgst_sgst else "") or "NIL"
                if wrong:
                    # Per explicit correction: this detail table lists MISMATCHES only, not every
                    # checked invoice -- 13,778 rows of mostly "Match" was more than needed; the
                    # RCM detail table above (separately requested) stays as complete detail
                    # since that one IS meant to be everything, not just exceptions.
                    mismatches += 1
                    statecode_detail_rows.append([
                        m, x["gstin"], x["supplier"], x["invno"], x.get("invdate"),
                        sup_state, self_state, ("Same-state" if same_state else "Inter-state"),
                        tax_head, "Mismatch", x["taxable"],
                        x["igst"] + x["cgst"] + x["sgst"] + x["cess"]])
                    rows.append(["STATE-CODE", x["month"], x["gstin"], x["supplier"], x["invno"],
                                "Same-state IGST charged" if (same_state and has_igst) else
                                "Inter-state CGST+SGST charged (no IGST)", None, None])
        findings.append(Finding(
            "G7", "GSTR-2A state-code vs tax-head validation (IGST vs CGST+SGST)",
            "REVIEW" if mismatches else "PASS",
            f"Checked {checked} GSTR-2A B2B invoices against the recipient's own state code "
            f"({self_state}, from GSTIN {self_gstin}); {mismatches} carry a tax head inconsistent "
            f"with the supplier's and recipient's state codes (a same-state invoice charging "
            f"IGST, or an inter-state invoice charging CGST+SGST instead of IGST). ITC on a "
            f"wrong tax head is technically inadmissible even where the total amount is correct. "
            f"The {mismatches} mismatched invoice(s) are listed in the 'State-Code Mismatched "
            f"GSTR-2A Invoices' table below.",
            dict(checked=checked, mismatches=mismatches)))
    else:
        findings.append(Finding("G7", "GSTR-2A state-code vs tax-head validation (IGST vs CGST+SGST)",
                                "SKIPPED", "Self-GSTIN could not be determined.", {}))

    return dict(header=["Check", "Month", "GSTIN / n-a", "Supplier / n-a", "Invoice No / n-a",
                        "Issue / n-a", "RCM taxable (2A)", "RCM tax (2A)"],
                widths=[12, 10, 18, 26, 18, 30, 16, 14], rows=rows, findings=findings, notes=[],
                extra_tables=[
                    dict(title=f"RCM-Flagged GSTR-2A Invoices -- Complete Detail ({len(rcm_detail_rows)} invoice(s))",
                         subtitle="Every GSTR-2A B2B invoice where the supplier marked 'Supply Attract Reverse "
                                  "Charge = Y' -- informational (supplier's own declaration), not itself proof "
                                  "of an RCM liability, but every one of them is listed here rather than just "
                                  "the month totals above.",
                         header=["Month", "GSTIN", "Supplier", "Invoice No", "Invoice Date",
                                 "Taxable", "IGST", "CGST", "SGST", "Cess"],
                         widths=[10, 18, 28, 18, 14, 16, 14, 14, 14, 12],
                         rows=sorted(rcm_detail_rows, key=lambda x: (str(x[0]), str(x[1]))),
                         empty_note="No GSTR-2A invoice carries the RCM flag."),
                    dict(title=f"State-Code Mismatched GSTR-2A Invoices -- Complete Detail ({mismatches} invoice(s))",
                         subtitle=f"Only the {mismatches} MISMATCHED invoice(s) -- out of {checked} checked -- "
                                  "not the full checked population; per explicit correction, this table is "
                                  "exceptions only, unlike the RCM-Flagged table above which is complete detail "
                                  "by design.",
                         header=["Month", "Supplier GSTIN", "Supplier", "Invoice No", "Invoice Date",
                                 "Supplier State", "Recipient State", "Same/Inter-state", "Tax Head Charged",
                                 "Status", "Taxable", "Tax"],
                         widths=[10, 18, 28, 18, 14, 12, 12, 14, 16, 10, 16, 14],
                         rows=sorted(statecode_detail_rows, key=lambda x: (str(x[0]), str(x[1]))),
                         empty_note="No mismatches -- every checked GSTR-2A invoice's tax head is consistent "
                                    "with the supplier/recipient state codes."),
                ])


def build_r2a_isd_164_cancelled(ctx):
    """G8 -- ISD credit per GSTR-2A vs Table 4A(4). G9 -- supplier filing dates vs the section
    16(4) deadline (honest version: 3B has no invoice-level claim date to check against).
    G10 -- invoices from a counterparty whose GSTIN was later cancelled."""
    r2a = ctx.get("r2a_data") or {}
    rows, findings = [], []
    if not r2a.get("available"):
        for ref, title in [("G8", "GSTR-2A ISD credit vs 3B Table 4A(4)"),
                           ("G9", "GSTR-2A late supplier filing vs Section 16(4) deadline"),
                           ("G10", "GSTR-2A counterparty GSTIN cancelled but ITC claimed")]:
            findings.append(Finding(ref, title, "SKIPPED", r2a.get("reason") or "GSTR-2A not supplied.", {}))
        return dict(header=["Status"], widths=[100], rows=[["GSTR-2A not supplied -- see findings."]],
                    findings=findings, notes=[])

    isd_2a_tax, isd_rows_present = 0.0, False
    for m, recs in r2a.get("isd", {}).items():
        for x in recs:
            isd_rows_present = True
            isd_2a_tax += x["igst"] + x["cgst"] + x["sgst"] + x["cess"]
    isd_3b_tax, isd_3b_present = 0.0, False
    for m in ctx["months"]:
        v = ctx["g3b_by_month"].get(m, {}).get("4A4")
        if v is not None:
            isd_3b_present = True
            isd_3b_tax += sum(_n(x) for x in v[:4])
    if isd_rows_present or isd_3b_present:
        d_isd = isd_3b_tax - isd_2a_tax
        findings.append(Finding(
            "G8", "GSTR-2A ISD credit vs 3B Table 4A(4)",
            "REVIEW" if abs(d_isd) > MATERIAL else "PASS",
            f"ISD credit per GSTR-2A's own ISD sheet {_f(isd_2a_tax)} against ISD credit claimed "
            f"under Table 4A(4) {_f(isd_3b_tax)}; difference {_f(d_isd)}.",
            dict(isd_2a=isd_2a_tax, isd_3b_4a4=isd_3b_tax)))
    else:
        findings.append(Finding("G8", "GSTR-2A ISD credit vs 3B Table 4A(4)", "PASS",
                                "No ISD credit recorded in GSTR-2A's ISD sheet and no ISD ITC "
                                "claimed under Table 4A(4) in any month -- nothing to reconcile.", {}))

    def _sixteen4_deadline(fy_lbl):
        m = re.match(r"^(\d{4})-(\d{2,4})$", fy_lbl or "")
        if not m:
            return None
        return _dt.date(int(m.group(1)) + 1, 11, 30)

    fy_lbl = ctx.get("fy_label")
    deadline = _sixteen4_deadline(fy_lbl)
    tight_count = late_count = 0
    if deadline:
        for m, recs in r2a.get("b2b", {}).items():
            for x in recs:
                fdate_iso = dept.r2a_clean_date(x.get("g1_filing_date"))
                if not fdate_iso:
                    continue
                try:
                    fd = _dt.date.fromisoformat(fdate_iso)
                except ValueError:
                    continue
                days_to_deadline = (deadline - fd).days
                if days_to_deadline < 0:
                    late_count += 1
                    rows.append(["16(4)", x["month"], x["gstin"], x["supplier"], x["invno"],
                                x["invdate"], fdate_iso, deadline.isoformat(),
                                f"Supplier filed {abs(days_to_deadline)} day(s) AFTER the 16(4) "
                                f"deadline -- if ITC on this invoice was claimed at all, it is "
                                f"time-barred.", x["taxable"],
                                x["igst"] + x["cgst"] + x["sgst"] + x["cess"]])
                elif days_to_deadline < 30:
                    tight_count += 1
        findings.append(Finding(
            "G9", "GSTR-2A late supplier filing vs Section 16(4) deadline",
            "FLAG" if late_count else ("REVIEW" if tight_count else "PASS"),
            f"{late_count} invoice(s) show the SUPPLIER's own GSTR-1 filing date after the "
            f"section 16(4) deadline ({deadline.isoformat()}, 30-Nov following the FY) for this "
            f"year; {tight_count} more filed within 30 days of it. This checks the SUPPLIER's "
            f"filing date only -- GSTR-3B carries no invoice-level data, so the recipient's "
            f"actual CLAIM date for any specific invoice cannot be traced from these files; a "
            f"late-filed invoice is a genuine 'this ITC may be time-barred' signal only once "
            f"matched against when the taxpayer actually claimed it.",
            dict(late=late_count, tight=tight_count)))
    else:
        findings.append(Finding("G9", "GSTR-2A late supplier filing vs Section 16(4) deadline",
                                "SKIPPED", "This run spans more than one financial year (or the FY "
                                "could not be determined), so a single 16(4) deadline does not "
                                "apply uniformly.", {}))

    cancelled_count = 0
    for m, recs in r2a.get("b2b", {}).items():
        for x in recs:
            if x.get("cancel_date"):
                cancelled_count += 1
                rows.append(["CANCELLED-GSTIN", x["month"], x["gstin"], x["supplier"], x["invno"],
                            x["invdate"], None, None,
                            f"Counterparty GSTIN shows an effective cancellation date of "
                            f"{x['cancel_date']} -- this is the SUPPLIER's registration being "
                            f"cancelled, not this specific invoice's e-invoice/IRN being "
                            f"cancelled (GSTR-2A does not carry per-invoice IRN-cancellation "
                            f"status for inward supplies). Verify the invoice date and ITC claim "
                            f"both predate the cancellation.",
                            x["taxable"], x["igst"] + x["cgst"] + x["sgst"] + x["cess"]])
    findings.append(Finding(
        "G10", "GSTR-2A counterparty GSTIN cancelled but invoice present",
        "REVIEW" if cancelled_count else "PASS",
        f"{cancelled_count} GSTR-2A B2B invoice(s) carry a counterparty 'Effective date of "
        f"cancellation' -- the supplier's GST registration was later cancelled. Not itself proof "
        f"of an issue (a supplier can validly cancel registration after correctly filing earlier "
        f"invoices), but worth checking the invoice date and ITC claim both predate the "
        f"cancellation date.",
        dict(count=cancelled_count)))

    return dict(header=["Check", "Month", "GSTIN", "Supplier", "Invoice No", "Invoice Date",
                        "Filing Date / n-a", "16(4) Deadline / Cancel Date", "Note",
                        "Taxable", "Tax"],
                widths=[16, 10, 18, 26, 18, 14, 16, 18, 50, 14, 12], rows=rows, findings=findings,
                notes=["Point 7 of the originating brief (blocked-credit HSN screening) is not "
                       "implemented here: GSTR-2A's B2B/CDNR sheets carry no HSN/SAC column at "
                       "all -- see the GSTR-2A Data Quality sheet's G-NF1 finding."])


def build_cn_dn_impact_data(ctx):
    """NEW (per explicit instruction): whole-FY, invoice-level detail of every credit/debit
    note this taxpayer both RECEIVED (inward, GSTR-2B B2B-CDNR) and ISSUED (outward, GSTR-1
    CDNR), grouped into the four combinations that matter for who owes what. Feeds the
    'CN-DN ITC Impact - Annual' sheet (a custom multi-table writer in master_build.py, not the
    single-table SHEETS mechanism, since this genuinely needs four separate tables plus a net
    summary rather than one).

    Direction of impact, per GST law (credit/debit notes under Sec 34, ITC reversal under Sec
    16(2) read with CGST Rule 37 and the corresponding provisions):
    - INWARD Credit Note (received from a supplier): the value of what THIS taxpayer purchased
      went DOWN -- this taxpayer must reverse the proportionate ITC it had claimed.
    - INWARD Debit Note (received from a supplier): the value of what THIS taxpayer purchased
      went UP -- this taxpayer becomes eligible for additional ITC.
    - OUTWARD Credit Note (issued to a customer): the value of what THIS taxpayer sold went
      DOWN -- this taxpayer's own output tax liability is already correspondingly reduced in
      its own GSTR-1/GSTR-3B (nothing further for THIS taxpayer to do about its own liability),
      but the RECIPIENT is required to reverse the ITC they had claimed on the original invoice.
    - OUTWARD Debit Note (issued to a customer): the value of what THIS taxpayer sold went UP
      -- this taxpayer's own output tax liability is already correspondingly increased in its
      own GSTR-1/GSTR-3B, and the RECIPIENT becomes eligible for additional ITC.

    Only the INWARD side is this taxpayer's own compliance action (ties directly to GSTR-3B
    4(B)(2) -- see the 'D2. ITC Reversal' section on each month's Comparison sheet); the
    OUTWARD side's ITC consequence belongs to the counterparty, and is included here as
    documented, attributable information (GSTIN, trade name, exact amount) rather than
    something this taxpayer can be shown as non-compliant on -- this tool has no visibility
    into whether the counterparty actually reversed/claimed as required."""
    self_gstin = ctx.get("self_gstin", "")
    company_name = ctx.get("company_name", "") or "(self)"
    lookup = ctx.get("gstin_name_lookup", {}) or {}

    def _n(v):
        try:
            return float(v) if v is not None else 0.0
        except (TypeError, ValueError):
            return 0.0

    inward = []
    for m in ctx.get("months", []):
        two = ctx.get("twob_by_month", {}).get(m, {}) or {}
        for c in two.get("cdnr", []) or []:
            gstin = str(c.get("gstin", "") or "").strip()
            name = str(c.get("supplier", "") or "").strip() or lookup.get(gstin, "")
            ntype = str(c.get("ntype", "") or "").strip()
            is_credit = ntype.upper().startswith("C")
            inward.append(dict(
                month=m, direction="Inward (received)", gstin=gstin, name=name,
                note=str(c.get("note", "") or "").strip(), ntype="Credit Note" if is_credit else "Debit Note",
                date=str(c.get("date", "") or "").strip(),
                taxable=_n(c.get("taxable")), igst=_n(c.get("igst")), cgst=_n(c.get("cgst")),
                sgst=_n(c.get("sgst")), cess=_n(c.get("cess")), itc_avail=str(c.get("itc_avail", "") or "").strip(),
                tax_total=_n(c.get("igst")) + _n(c.get("cgst")) + _n(c.get("sgst")) + _n(c.get("cess")),
                is_credit=is_credit,
                action=(f"{company_name} (GSTIN {self_gstin}) must REVERSE ITC previously claimed "
                        f"on the related purchase from {name or 'this supplier'} (GSTIN {gstin}), "
                        f"due to this credit note."
                        if is_credit else
                        f"{company_name} (GSTIN {self_gstin}) becomes eligible for ADDITIONAL ITC "
                        f"on the related purchase from {name or 'this supplier'} (GSTIN {gstin}), "
                        f"due to this debit note -- ensure it is included in the ITC claimed."),
            ))

    outward = []
    for x in ctx.get("g1_cdnr_fy", []) or []:
        gstin = str(x.get("gstin", "") or "").strip()
        name = lookup.get(gstin, "")
        ntype = str(x.get("notetype", "") or "").strip()
        is_credit = ntype.upper().startswith("C")
        outward.append(dict(
            month=x.get("month", ""), direction="Outward (issued)", gstin=gstin, name=name,
            note=str(x.get("noteno", "") or "").strip(), ntype="Credit Note" if is_credit else "Debit Note",
            date="",
            taxable=_n(x.get("taxable")), igst=_n(x.get("igst")), cgst=_n(x.get("cgst")),
            sgst=_n(x.get("sgst")), cess=0.0,
            tax_total=_n(x.get("igst")) + _n(x.get("cgst")) + _n(x.get("sgst")),
            is_credit=is_credit,
            action=(f"{company_name}'s own output tax liability is already reduced accordingly "
                    f"in its own GSTR-1/GSTR-3B. The RECIPIENT, {name or '(name not resolved)'} "
                    f"(GSTIN {gstin}), is required to REVERSE the ITC it had claimed on the "
                    f"related original invoice, due to this credit note."
                    if is_credit else
                    f"{company_name}'s own output tax liability is already increased accordingly "
                    f"in its own GSTR-1/GSTR-3B. The RECIPIENT, {name or '(name not resolved)'} "
                    f"(GSTIN {gstin}), becomes eligible for ADDITIONAL ITC, due to this debit note."),
        ))

    # Net summary -- only the inward side is THIS taxpayer's own actionable ITC impact; matches
    # the "Yes"/"Unconfirmed" convention used everywhere else in this tool (an unconfirmed row
    # is not assumed ineligible; only a CONFIRMED "No" is excluded from the actionable total).
    inward_actionable = [r for r in inward if r["itc_avail"].strip().upper() != "NO"]
    inward_cn_tax = sum(r["tax_total"] for r in inward_actionable if r["is_credit"])
    inward_dn_tax = sum(r["tax_total"] for r in inward_actionable if not r["is_credit"])
    net_itc_reversal_required = round(inward_cn_tax - inward_dn_tax, 2)
    outward_cn_tax = sum(r["tax_total"] for r in outward if r["is_credit"])
    outward_dn_tax = sum(r["tax_total"] for r in outward if not r["is_credit"])

    return dict(
        self_gstin=self_gstin, company_name=company_name,
        inward=inward, outward=outward,
        net_itc_reversal_required=net_itc_reversal_required,
        inward_cn_tax=round(inward_cn_tax, 2), inward_dn_tax=round(inward_dn_tax, 2),
        outward_cn_tax=round(outward_cn_tax, 2), outward_dn_tax=round(outward_dn_tax, 2),
    )


def build_itc_detailed_recon_data(ctx):
    """NEW (per explicit, very detailed instruction): whole-FY, tax-head-wise, invoice-level
    ITC reconciliation -- one row per month plus an FY TOTAL row -- feeding a new table
    appended to the existing 'ITC Annual Summary' sheet (below its current content, per the
    established pattern this session; NOT a second sheet).

    Every figure below is computed FRESH from invoice-level B2B/B2B-CDNR (2B) and B2B (2A)
    rows already available via ctx -- per explicit instruction, NEITHER the 2B summary
    ('ITC Available'/'ITC not available') sheets NOR any pre-aggregated total is used as a
    source for the '2B (Yes only)' or '2A (all invoices)' lines; both are built by summing the
    real invoice rows, tax-head by tax-head.

    Deliberate design choices, stated here rather than left implicit:
    - '2B (Available=Yes only)' is STRICT Yes-status only (excludes both No AND Unconfirmed)
      -- this is a narrower filter than the Yes+Unconfirmed convention this tool uses
      elsewhere (e.g. Section D's fallback), used here because the instruction explicitly
      names 'Available=Yes only'.
    - 2A and 2B are never netted together (per instruction) -- 2A feeds ONLY the Mismatch
      columns, 2B(Yes) is the sole eligibility baseline everything else builds from.
    - Mismatch matching key is deliberately strict: (invoice number, supplier GSTIN, invoice
      value, IGST, CGST, SGST, CESS, invoice date) -- an invoice that exists in both sources
      but differs on ANY of these fields counts as unmatched on that side, not a false "same
      invoice" match on invoice number + GSTIN alone.
    - Credit/Debit Note impact uses the SAME Yes-only convention as the 2B baseline, for
      internal consistency within this one table.
    - 'ITC carried forward from last FY' has no dedicated GSTR-3B field in this data source
      (confirmed: no such field is parsed anywhere in this tool) -- computed the same way
      the existing 'ITC Annual Summary' sheet's own inferred-carry-forward column already
      does (claimed exceeding what this FY's own 2B+reversals support), stated as INFERRED,
      not read directly, and cross-checked against GSTR-9 Table 13 at FY level when available.
    - If GSTR-3B is not available for a month, 'Actual ITC claimed' and 'Excess/Short claim'
      are left as None (rendered blank) and Remarks says so explicitly -- never a computed
      false discrepancy against a missing source."""
    HEADS = ("IGST", "CGST", "SGST", "CESS")

    def _z():
        return [0.0, 0.0, 0.0, 0.0]

    def _norm_date(v):
        """BUG FIX (found during verification, before this table was ever shipped): 2A's
        invdate arrives as an ISO string ('2023-05-23', from r2a_clean_date) while 2B's date
        arrives as a DD/MM/YYYY string ('23/05/2023', unparsed from the raw cell text) --
        same calendar date, two different string shapes. Comparing those strings directly in
        the composite key made EVERY genuinely-matching invoice look unmatched (confirmed on
        a real invoice, SIEPL-INV-79, present in both sources with identical GSTIN/value/tax
        heads, differing only in this string format) -- a near-100% false mismatch rate every
        month. Normalizes both shapes (plus a defensive date/datetime object case) to one
        canonical ISO string before the key is ever built."""
        if v is None or v == "":
            return ""
        if hasattr(v, "isoformat"):
            return v.isoformat()
        s = str(v).strip()
        m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", s)
        if m:
            return s
        m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{4})$", s)
        if m:
            dd, mm, yyyy = m.groups()
            return f"{yyyy}-{int(mm):02d}-{int(dd):02d}"
        return s

    def _key(invno, gstin, invval, igst, cgst, sgst, cess, invdate):
        return (str(invno or "").strip().upper(), str(gstin or "").strip().upper(),
                round(_n(invval), 2), round(_n(igst), 2), round(_n(cgst), 2),
                round(_n(sgst), 2), round(_n(cess), 2), _norm_date(invdate))

    r2a = ctx.get("r2a_data") or {}
    r2a_b2b_by_month = r2a.get("b2b", {}) if r2a.get("available") else {}
    rows = []
    FT = dict(b2b=_z(), a2a=_z(), cn=_z(), dn=_z(), net=_z(), claimed=_z(), cf=_z(), exs=_z(),
              mism_2a_not_2b_ct=0, mism_2a_not_2b_val=0.0, mism_2b_not_2a_ct=0, mism_2b_not_2a_val=0.0)
    any_g3b = False

    for m in ctx["months"]:
        two = ctx["twob_by_month"].get(m, {})
        b2b_rows = two.get("rows", []) if two.get("available") else []
        cdnr_rows = two.get("cdnr", []) if two.get("available") else []

        # ---- 2B (Yes only), tax-head-wise, summed directly from B2B invoice rows ----
        b2b_yes = [r for r in b2b_rows if str(r.get("itc_avail", "")).strip().upper() == "YES"]
        b2b_sum = [round(sum(_n(r.get(h.lower())) for r in b2b_yes), 2) for h in HEADS]

        # ---- 2A (all invoices), tax-head-wise, summed directly from B2B invoice rows ----
        a2a_rows = r2a_b2b_by_month.get(m, [])
        a2a_sum = [round(sum(_n(r.get(h.lower())) for r in a2a_rows), 2) for h in HEADS]

        # ---- Mismatch: invoice-level, strict composite key, both directions ----
        def _agg_by_invoice(rows_in):
            agg = {}
            for r in rows_in:
                k = (str(r.get("invno", "")).strip().upper(), str(r.get("gstin", "")).strip().upper())
                a = agg.setdefault(k, dict(invval=0.0, igst=0.0, cgst=0.0, sgst=0.0, cess=0.0,
                                            invdate=r.get("invdate") or r.get("date")))
                a["invval"] = _n(r.get("invval"))  # invoice-level, not summed across rate-lines
                a["igst"] += _n(r.get("igst")); a["cgst"] += _n(r.get("cgst"))
                a["sgst"] += _n(r.get("sgst")); a["cess"] += _n(r.get("cess"))
            return agg

        b2b_by_inv = _agg_by_invoice(b2b_yes)
        a2a_by_inv = _agg_by_invoice(a2a_rows)
        b2b_keys = {_key(k[0], k[1], v["invval"], v["igst"], v["cgst"], v["sgst"], v["cess"], v["invdate"])
                    for k, v in b2b_by_inv.items()}
        a2a_keys = {_key(k[0], k[1], v["invval"], v["igst"], v["cgst"], v["sgst"], v["cess"], v["invdate"])
                    for k, v in a2a_by_inv.items()}
        a2a_not_b2b = a2a_keys - b2b_keys
        b2b_not_a2a = b2b_keys - a2a_keys
        # value of the mismatched invoices = sum of their (igst+cgst+sgst+cess) from the key tuple
        val_a2a_not_b2b = round(sum(k[3] + k[4] + k[5] + k[6] for k in a2a_not_b2b), 2)
        val_b2b_not_a2a = round(sum(k[3] + k[4] + k[5] + k[6] for k in b2b_not_a2a), 2)

        # ---- Credit/Debit Note impact (2B CDNR, Yes-only, tax-head-wise) ----
        cdnr_yes = [r for r in cdnr_rows if str(r.get("itc_avail", "")).strip().upper() == "YES"]
        cn_rows = [r for r in cdnr_yes if str(r.get("ntype", "")).strip().upper().startswith("C")]
        dn_rows = [r for r in cdnr_yes if str(r.get("ntype", "")).strip().upper().startswith("D")]
        cn_sum = [round(sum(_n(r.get(h.lower())) for r in cn_rows), 2) for h in HEADS]
        dn_sum = [round(sum(_n(r.get(h.lower())) for r in dn_rows), 2) for h in HEADS]

        # ---- Net ITC eligible = 2B(Yes) + DN - CN, tax-head-wise ----
        net_sum = [round(b2b_sum[i] + dn_sum[i] - cn_sum[i], 2) for i in range(4)]

        # ---- Actual ITC claimed (GSTR-3B Table 4A), tax-head-wise ----
        g3b = ctx["g3b_by_month"].get(m, {})
        ex = ctx["g3b_extra_by_month"].get(m, {})
        g3b_available = bool(g3b) or bool(ex.get("available"))
        claimed_sum = cf_sum = exs_sum = [None, None, None, None]
        b1_sum = b2_sum = _z()
        if g3b_available:
            any_g3b = True
            a1 = (ex.get("A1") or [0, 0, 0, 0]); a2 = (ex.get("A2") or [0, 0, 0, 0])
            a3 = (g3b.get("4A3") or [0, 0, 0, 0]); a4 = (ex.get("A4") or [0, 0, 0, 0])
            a5 = (g3b.get("4A5") or [0, 0, 0, 0])
            claimed_sum = [round(_n(a1[i]) + _n(a2[i]) + _n(a3[i]) + _n(a4[i]) + _n(a5[i]), 2) for i in range(4)]
            b1_sum = [_n(x) for x in (g3b.get("4B1") or [0, 0, 0, 0])]
            b2_sum = [_n(x) for x in (g3b.get("4B2") or [0, 0, 0, 0])]
            # INFERRED carry-forward, per head: excess of claimed over what THIS FY's own
            # 2B(Yes)+reversals support -- same formula as the existing 'ITC Annual Summary'
            # sheet's own inferred-CF column, applied per tax-head here instead of combined.
            cf_sum = [round(max(0.0, claimed_sum[i] - (b2b_sum[i] - b1_sum[i] - b2_sum[i])), 2) for i in range(4)]
            exs_sum = [round(claimed_sum[i] - net_sum[i], 2) for i in range(4)]

        remark = ""
        if not g3b_available:
            remark = "Reconciliation only — no 3B comparison (GSTR-3B not supplied for this month)."
        else:
            exs_total = sum(exs_sum)
            cn_total = sum(cn_sum)
            if exs_total > 0 and cn_total != 0:
                remark = ("FLAG: Excess ITC claimed vs eligible, AND a credit-note reversal exists this "
                          "period — verify the credit-note-related ITC reversal was actually applied "
                          "before this excess is treated as a genuine over-claim.")
            elif exs_total > 0:
                remark = "Excess ITC claimed vs eligible this period — verify."

        rows.append(dict(month=m, b2b=b2b_sum, a2a=a2a_sum, cn=cn_sum, dn=dn_sum, net=net_sum,
                          claimed=claimed_sum, cf=cf_sum, exs=exs_sum, remark=remark,
                          mism_a2a_not_b2b_ct=len(a2a_not_b2b), mism_a2a_not_b2b_val=val_a2a_not_b2b,
                          mism_b2b_not_a2a_ct=len(b2b_not_a2a), mism_b2b_not_a2a_val=val_b2b_not_a2a,
                          g3b_available=g3b_available))

        for i in range(4):
            FT["b2b"][i] += b2b_sum[i]; FT["a2a"][i] += a2a_sum[i]
            FT["cn"][i] += cn_sum[i]; FT["dn"][i] += dn_sum[i]; FT["net"][i] += net_sum[i]
            if g3b_available:
                FT["claimed"][i] += claimed_sum[i]; FT["cf"][i] += cf_sum[i]; FT["exs"][i] += exs_sum[i]
        FT["mism_2a_not_2b_ct"] += len(a2a_not_b2b); FT["mism_2a_not_2b_val"] += val_a2a_not_b2b
        FT["mism_2b_not_2a_ct"] += len(b2b_not_a2a); FT["mism_2b_not_2a_val"] += val_b2b_not_a2a

    ft_remark = ""
    if any_g3b:
        exs_total = sum(FT["exs"])
        cn_total = sum(FT["cn"])
        if exs_total > 0 and cn_total != 0:
            ft_remark = "FLAG: FY-level excess claim alongside non-zero credit-note impact — verify."
        elif exs_total > 0:
            ft_remark = "FY-level excess ITC claimed vs eligible — verify."
    else:
        ft_remark = "Reconciliation only — no 3B comparison (GSTR-3B not supplied for any month)."

    rows.append(dict(month="FY TOTAL", b2b=[round(x, 2) for x in FT["b2b"]],
                      a2a=[round(x, 2) for x in FT["a2a"]], cn=[round(x, 2) for x in FT["cn"]],
                      dn=[round(x, 2) for x in FT["dn"]], net=[round(x, 2) for x in FT["net"]],
                      claimed=([round(x, 2) for x in FT["claimed"]] if any_g3b else [None]*4),
                      cf=([round(x, 2) for x in FT["cf"]] if any_g3b else [None]*4),
                      exs=([round(x, 2) for x in FT["exs"]] if any_g3b else [None]*4),
                      remark=ft_remark,
                      mism_a2a_not_b2b_ct=FT["mism_2a_not_2b_ct"], mism_a2a_not_b2b_val=round(FT["mism_2a_not_2b_val"], 2),
                      mism_b2b_not_a2a_ct=FT["mism_2b_not_2a_ct"], mism_b2b_not_a2a_val=round(FT["mism_2b_not_2a_val"], 2),
                      g3b_available=any_g3b, is_fy_total=True))
    return rows


def build_itc_yearly_slim(ctx):
    """NEW SHEET (per explicit request, feature #3 in the bug-fix/feature list) -- a dedicated,
    single-row-per-FY summary with exactly the six columns asked for: ITC Available (2A), ITC
    Available (2B), Reversed 4B(1), Reversed 4B(2), ITC Availed (Current FY), and ITC Availed via
    Carry-Forward from Last FY. Deliberately calls build_itc_annual_summary(ctx) and reads its FY
    TOTAL row rather than recomputing anything independently -- the two sheets are guaranteed to
    agree by construction, never a second, possibly-diverging calculation of the same figures.
    This sheet is a SLIM, FY-only complement to 'ITC Annual Summary' (which keeps the full
    month-by-month breakdown and reference columns) -- nothing in that sheet was removed."""
    full = build_itc_annual_summary(ctx)
    fy_row = next((r for r in full["rows"] if r[0] == "FY TOTAL"), None)
    findings = []
    if fy_row is None:
        findings.append(Finding("F1-SLIM", "ITC Yearly Summary could not be built", "INFO",
                                 "The 'ITC Annual Summary' sheet's own FY TOTAL row was not found "
                                 "-- see that sheet for the underlying reason.", {}))
        return dict(header=["Status"], widths=[120], rows=[["SKIPPED -- see Finding F1-SLIM."]],
                    findings=findings, notes=[])
    # fy_row = [Month, avail_2a, avail_2b, b1, b2, claimed, inferred_cf, carry_fwd_gstr9, yes_only, grand_total]
    _, avail_2a, avail_2b, b1, b2, claimed, inferred_cf, carry_fwd_gstr9, _, _ = fy_row
    # Per instruction, column 6 is ONE figure, not two -- default to the GSTR-9 Table 13 figure
    # when the taxpayer's own filed annual return supplies it (authoritative, filed data beats a
    # same-FY-only inferred estimate); fall back to the inferred figure only when GSTR-9 wasn't
    # supplied. Both are still shown side by side in 'ITC Annual Summary' itself for anyone who
    # wants the full picture -- this sheet is a slim single answer per the request.
    if carry_fwd_gstr9 is not None:
        cf_value, cf_source = carry_fwd_gstr9, "GSTR-9 Table 13 (authoritative, filed)"
    else:
        cf_value, cf_source = inferred_cf, "INFERRED (GSTR-9 not supplied -- see 'ITC Annual Summary')"
    rows = [[
        (round(avail_2a, 2) if avail_2a is not None else None),
        (round(avail_2b, 2) if avail_2b is not None else None),
        round(b1, 2), round(b2, 2), round(claimed, 2),
        (round(cf_value, 2) if cf_value is not None else None),
    ]]
    notes = [
        f"Carry-forward column source: {cf_source}.",
        "Every figure on this sheet is read directly from 'ITC Annual Summary''s own FY TOTAL "
        "row (not recalculated) -- the two sheets always agree; see that sheet for the full "
        "month-by-month breakdown, both carry-forward methods side by side, and the underlying "
        "methodology notes.",
        ("2A available" if avail_2a is not None else "2A available: GSTR-2A was not supplied for this run.") ,
        ("2B available" if avail_2b is not None else "2B available: GSTR-2B was not supplied for this run."),
    ]
    return dict(
        header=["ITC Available (2A)", "ITC Available (2B)", "Reversed 4B(1)", "Reversed 4B(2)",
                "ITC Availed (Current FY)", "ITC Availed via Carry-Forward from Last FY"],
        widths=[22, 22, 18, 18, 24, 34], rows=rows, findings=findings, notes=notes)


SHEETS = [
    ("Purchase vs Sales & Stock", build_purchase_sales_stock,
     "F1 -- purchase against sales in MONEY terms, monthly and cumulative. GSTR-2B carries no "
     "quantity column, so no quantity-based stock figure can be derived from GST returns; this is "
     "a value flow, shown against the audited inventory movement for context."),
    ("ITC Yearly Summary", build_itc_yearly_slim,
     "Dedicated single-row FY summary: ITC Available (2A), ITC Available (2B), Reversed 4B(1), "
     "Reversed 4B(2), ITC Availed (Current FY), ITC Availed via Carry-Forward from Last FY -- "
     "sourced from 'ITC Annual Summary''s own FY TOTAL row, so the two always agree."),
    ("ITC Annual Summary", build_itc_annual_summary,
     "FY-level ITC lifecycle -- Available (2B) -> Claimed (4A) -> Reversed (4B1/4B2) -> Reclaimed "
     "(this-FY tracking not available; prior-FY N/A, first year) -- plus Credit and Cash Ledger "
     "FY-level tie-outs and a Closing Balances section for manual carry-forward into next year's "
     "tool. Scoped to this FY only; no multi-year logic."),
    ("GSTR-2B ITC No & Unconfirmed", build_2b_no_unconfirmed_detail,
     "Complete invoice-level detail of every GSTR-2B B2B/CDNR row flagged ITC=No or with an "
     "unconfirmed eligibility status, whole FY -- the underlying detail behind ITC Annual "
     "Summary's PRIMARY/Yes-only/grand-total reference columns."),
    ("Zero-Tax Invoice Scan", build_zero_tax_scan,
     "Every invoice/movement with a real value but nil tax, across GSTR-1, GSTR-2B, and both "
     "e-way bill directions -- one table per source, heading-wise."),
    ("ITC Roll-Forward 4A-4B-4C", build_itc_rollforward,
     "F2 and F3 -- ITC availed by head, reversal split between Table 4B(1) (Rules 38/42/43, "
     "permanent) and 4B(2) (Others, ordinarily temporary), net credit, ineligible credit, and the "
     "Electronic Credit Ledger's own credits and debits."),
    ("3-Way GSTR1 3B EWB", build_three_way,
     "F4 and F5 -- outward value as invoiced (GSTR-1), as declared (GSTR-3B 3.1(a)), and as moved "
     "(e-way bills, deduplicated on EWB number)."),
    ("ITC 3B vs 2B", build_itc_3b_vs_2b,
     "F7 -- ITC claimed against ITC available, computed from GSTR-2B's INVOICE-LEVEL rows rather "
     "than its quarterly summary sheet; the summary figure is shown alongside as a control total."),
    ("RCM Triangulation", build_rcm,
     "F8 -- reverse-charge liability declared, credit taken on it, and cash actually debited. "
     "RCM cannot be discharged out of the credit ledger (section 49(4)), so the cash ledger is "
     "the authoritative leg."),
    ("DRC-03 & Ledger Movements", build_drc_and_ledger_movements,
     "F9 -- voluntary and DRC-type payments, plus every other non-return movement in the "
     "electronic ledgers: refund debits and Rule 86A credit blocking."),
    ("Turnover Growth vs Tax", build_turnover_vs_tax,
     "F10 -- turnover, liability, and the split between credit and cash across every financial "
     "year the BO Profile carries. All amounts in LAKHS, as printed in that profile."),
    ("Counterparty Transactions", build_counterparty,
     "F11 -- every same-day repeat transaction with a counterparty (no value threshold applied) "
     "and every reciprocal pair that both supplies to and buys from the taxpayer."),
    ("Top Counterparties (Computed)", build_top_counterparties,
     "F12 -- top 10 by ITC received and by tax passed on, computed from this year's own returns "
     "and cross-referenced to the department's BO-Profile lists."),
    ("B2B to B2C Shift", build_b2b_b2c_shift,
     "F6 -- movement of turnover towards unregistered buyers, which removes the input-credit "
     "trail and the counterparty's incentive to report."),
    ("GSTR-2A Data Quality", build_r2a_data_quality,
     "G1 -- duplicate/reused invoice numbers, GSTIN format hygiene, and rows excluded for "
     "safety. Also records why blocked-credit HSN screening could not be built from GSTR-2A."),
    ("GSTR-2A vs 2B Monthly", build_r2a_vs_2b_monthly,
     "G2a -- monthly and cumulative running-total comparison between GSTR-2A and GSTR-2B, the "
     "wash-out view that separates genuine excess claims from ordinary timing noise."),
    ("GSTR-2A vs 2B Invoice Detail", build_r2a_vs_2b_detail,
     "G2 -- invoice-level existence and value comparison between GSTR-2A and GSTR-2B, matched "
     "on GSTIN + invoice number + invoice date + invoice type."),
    ("GSTR-2A vs 3B & Ledger", build_r2a_vs_3b_and_ledger,
     "G3 and G4 -- cumulative FY-to-date ITC available per GSTR-2A against Table 4A(5) claimed "
     "in GSTR-3B, and against what was actually credited to the Electronic Credit Ledger."),
    ("GSTR-2A Amendments", build_r2a_amendments,
     "G5 -- amended invoices (B2BA) linked back to their original GSTR-2A B2B entry."),
    ("GSTR-2A RCM & State-Code", build_r2a_rcm_and_statecode,
     "G6 and G7 -- reverse-charge-flagged inward supplies cross-checked against Table "
     "3.1(d)/4A(3), and a state-code vs tax-head (IGST vs CGST+SGST) validation."),
    ("GSTR-2A ISD, 16(4) & Cancelled GSTIN", build_r2a_isd_164_cancelled,
     "G8, G9, G10 -- ISD credit against Table 4A(4), supplier filing dates against the section "
     "16(4) deadline, and invoices from counterparties whose GSTIN was later cancelled."),
]


def build_context(months, res, month_results, annual_data, ewb_out_rows, ewb_in_rows,
                  gstr9, gstr9c, table8a, bs_pl_data, self_gstin, fy_label, r2a_data=None,
                  blocked_credit_master_path=None):
    """Assemble everything the builders need, reading each merged workbook once
    per month rather than once per check.

    blocked_credit_master_path: optional -- if given, this function ALSO runs
    the Potential Blocked Credits keyword scan (gst_blocked_credit.py) over
    the SAME invoice-level 2B rows it already assembles below, and stores the
    result in the returned ctx (as 'blocked_credit_rows'/'blocked_credit_
    totals') purely so build_itc_rollforward() can cross-link against it (per
    explicit instruction: read that sheet's output, never touch its module).
    The ACTUAL 'Potential Blocked Credits' sheet is written separately by
    master_build.py (which calls gst_blocked_credit.build_and_write() again,
    after the workbook exists) -- this is a second, independent, side-effect-
    free scan for the cross-link numbers only, not a dependency between the
    two sheets' code."""
    g1_by_month, g3b_by_month = {}, {}
    g1_invcount = {}
    for r in month_results:
        m = r["month"]
        g1_by_month[m] = r["comp_raw"]["g1"]
        g3b_by_month[m] = r["comp_raw"]["g3b"]
        g1_invcount[m] = len({k[0] for k in r["comp_raw"]["g1"].get("lines", {}) if k[0]})

    twob, twob_summary, extra, b2c = {}, {}, {}, {}
    g1_lines_fy, twob_lines_fy = [], []
    for m in months:
        g1p = res["gstr1_month_map"].get(m)
        g3p = res["gstr3b_month_map"].get(m)
        g2p = res["gstr2b_month_map"].get(m)
        twob[m] = read_2b_invoice_level(g2p, m)
        twob_summary[m] = (month_results_lookup(month_results, m) or {})
        extra[m] = read_gstr3b_extra(g3p, m)
        b2c[m] = read_gstr1_b2c(g1p, m)
        g1_lines_fy.extend(read_gstr1_b2b_ff(g1p, m))
        for x in twob[m].get("rows", []):
            twob_lines_fy.append(dict(x, month=m))

    # NEW: outward credit/debit notes (GSTR-1's own cdnr sheet), whole FY, flat list with each
    # row tagged by month -- feeds the new CN/DN ITC Impact sheet (inward side already available
    # per-month via twob[m]['cdnr'], see read_2b_invoice_level() above; this is its outward
    # counterpart). _cdnr_rows_by_month() already reads and forward-fills the WHOLE merged file
    # in one call (it internally splits by month itself), so only needs one representative path.
    g1_cdnr_fy = []
    _g1_cdnr_path = next((res["gstr1_month_map"].get(m) for m in months if res["gstr1_month_map"].get(m)), None)
    if _g1_cdnr_path:
        _g1_cdnr_by_month = hfc._cdnr_rows_by_month(_g1_cdnr_path)
        for m in months:
            for x in _g1_cdnr_by_month.get(m, []):
                g1_cdnr_fy.append(dict(x, month=m))

    # NEW: Potential Blocked Credits cross-link data (see docstring above) -- never raises;
    # degrades to empty (which build_itc_rollforward's cross-link finding reports as SKIPPED)
    # exactly like every other optional source in this tool.
    blocked_credit_rows, blocked_credit_totals = [], {}
    if blocked_credit_master_path:
        try:
            import gst_blocked_credit as _bcred
            _master = _bcred.load_master(blocked_credit_master_path)
            _by_month = {}
            for x in twob_lines_fy:
                _by_month.setdefault(x.get("month"), []).append(x)
            blocked_credit_rows, blocked_credit_totals = _bcred.scan(_by_month, _master)
        except Exception:
            blocked_credit_rows, blocked_credit_totals = [], {}

    # NEW: GSTR-1's own HSN summary, whole-FY, by month -- same source/method
    # master_build.py already uses for Machinery HSN Scan, now also available via ctx so
    # build_purchase_sales_stock can use it for the inward-EWB-vs-outward-GSTR1 HSN comparison.
    g1_hsn_by_month = {}
    for g1f in {res["gstr1_month_map"].get(m) for m in months} - {None}:
        hsn_all = hfc._hsn_rows_by_month(g1f)
        for m in months:
            g1_hsn_by_month.setdefault(m, []).extend(hsn_all.get(m, []))

    # NEW: cross-source GSTIN -> name lookup (bug fix: a counterparty can have a genuinely blank
    # name in ONE source's own invoice rows -- confirmed against real data, e.g. a government TDS
    # deductor with a blank name in GSTR-1's own B2B sheet -- while a name for the SAME GSTIN is
    # sitting in a different source this tool already has open, e.g. GSTR-2A or the BO Profile's
    # Top-10 lists. Built once here, from every name-bearing source already parsed in this
    # function or passed in, in a fixed priority order (first non-blank wins), so every sheet
    # that displays a counterparty name can use ONE resolved name instead of each sheet doing its
    # own narrower 1-2-source lookup.
    gstin_name_lookup = {}
    def _learn(gstin, name):
        if gstin and name and gstin not in gstin_name_lookup:
            gstin_name_lookup[gstin] = str(name).strip()
    for x in g1_lines_fy:
        _learn(x.get("gstin"), x.get("name"))
    for x in twob_lines_fy:
        _learn(x.get("gstin"), x.get("supplier"))
    if r2a_data and r2a_data.get("available"):
        for _m, _recs in (r2a_data.get("b2b") or {}).items():
            for x in _recs:
                _learn(x.get("gstin"), x.get("supplier"))
    for x in (ewb_out_rows or []):
        _learn(x.get("to_gstin"), x.get("to_name"))
        _learn(x.get("from_gstin"), x.get("from_name"))
    for x in (ewb_in_rows or []):
        _learn(x.get("from_gstin"), x.get("from_name"))
        _learn(x.get("to_gstin"), x.get("to_name"))
    _bo = (annual_data or {}).get("bo") or {}
    for _key in ("top_beneficiaries", "top_suppliers", "related_itc_received", "related_itc_passed"):
        for x in (_bo.get(_key) or []):
            _learn(x.get("gstin"), x.get("name"))

    return dict(months=months, res=res, month_results=month_results, annual_data=annual_data,
                ewb_out_rows=ewb_out_rows, ewb_in_rows=ewb_in_rows,
                gstin_name_lookup=gstin_name_lookup, g1_hsn_by_month=g1_hsn_by_month,
                gstr9=gstr9, gstr9c=gstr9c, table8a=table8a, bs_pl_data=bs_pl_data,
                self_gstin=self_gstin, fy_label=fy_label,
                company_name=res.get("company_name", ""),
                g1_by_month=g1_by_month, g3b_by_month=g3b_by_month,
                g1_invcount_by_month=g1_invcount, g3b_extra_by_month=extra,
                twob_by_month=twob, twob_summary_by_month=twob_summary,
                b2c_by_month=b2c, g1_lines_fy=g1_lines_fy, twob_lines_fy=twob_lines_fy,
                g1_cdnr_fy=g1_cdnr_fy,
                blocked_credit_rows=blocked_credit_rows, blocked_credit_totals=blocked_credit_totals,
                # NEW: GSTR-2A (r2a_data is None or dept.parse_r2a_excel()'s own return shape --
                # every G-series builder below treats available=False as an explicit SKIP).
                r2a_data=r2a_data or dict(available=False,
                                           reason="GSTR-2A not supplied for this taxpayer/FY.",
                                           months_present=set(), b2b={}, b2ba={}, cdnr={},
                                           cdnra={}, isd={}, total_row_missing={},
                                           malformed_gstin=[]))


def month_results_lookup(month_results, m):
    for r in month_results:
        if r["month"] == m:
            return r["comp_raw"].get("b2b")
    return None


def build_all(ctx):
    """Build every sheet's content WITHOUT touching the workbook, so the caller
    can write the Master Dashboard (which needs these findings) before the
    sheets themselves. Each builder is isolated: one failing produces one
    explicit INFO finding and an explanatory sheet, and never stops the others."""
    sheets, findings = [], []
    for name, fn, subtitle in SHEETS:
        try:
            built = fn(ctx)
        except Exception as exc:   # noqa -- degrade loudly, never crash the run
            f = Finding(name, f"Sheet '{name}' could not be built", "INFO",
                        f"Internal error while building this sheet: {exc!r}. Every other sheet in "
                        f"this module still built normally.", {})
            built = dict(header=["Status"], widths=[120], rows=[[f.detail]], findings=[f], notes=[])
        sheets.append((name, subtitle, built))
        findings.extend(built.get("findings", []))
    return sheets, findings


def write_all(wb, sheets):
    for name, subtitle, built in sheets:
        _write_table(wb.create_sheet(name[:31]), name.upper(), subtitle, built)


def run_and_write(wb, ctx):
    sheets, findings = build_all(ctx)
    write_all(wb, sheets)
    return findings
