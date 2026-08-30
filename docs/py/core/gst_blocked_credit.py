#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
POTENTIAL BLOCKED CREDITS -- scans the merged GSTR-2B invoice-level data for
suppliers whose Trade/Legal name suggests a Section 17(5) blocked-credit
category (accommodation, motor vehicles, club membership, insurance, etc.),
using a taxpayer-supplied keyword/HSN master list. Screening aid only -- every
flagged invoice needs manual review; this module never adjusts, removes, or
blocks any ITC figure anywhere else in the workbook.

SCOPE LIMITATION (confirmed against a real GSTR-2B export, not assumed):
GSTR-2B's 'B2B' sheet carries NO line-item description and NO HSN/SAC column
at all -- it is invoice-level only (GSTIN, trade/legal name, invoice
number/date/value, tax amounts). So of the master list's two match signals,
only the Trade/Legal-name keyword match is buildable from this tool's
existing 2B data; Description-keyword and HSN/SAC-prefix matching are NOT
possible from GSTR-2B and are not attempted here -- this is the SAME
limitation this tool's own gst_checks_hsn_fraud.check_blocked_itc_by_hsn()
(finding A5, "Blocked ITC (Sec 17(5)) by purchase-side HSN") already
documents on the HSN side ("Not computable... no HSN column on the purchase
side... Would need a purchase register"). This module is the keyword/trade-
name-based sheet for the SAME underlying gap A5 already explains; it doesn't
contradict A5, it's the alternative approach the master list's Category/
Search-keyword columns are actually usable for from this data source.
Match_Confidence is therefore capped at "Medium" -- "High" would require a
keyword+HSN combination that can't occur from GSTR-2B alone. If a purchase
register with item-level HSN/description is ever supplied, THAT would be the
source for the fuller three-signal match this master list is designed for.

Purely additive: never changes any existing calculation, merge, or sheet.
If the master list can't be found or the merged 2B data is missing/empty,
this sheet is skipped with a clear reason logged -- never a crash, never a
silently-empty sheet with no explanation (same discipline as every other
optional source in this tool).
"""

import re
from collections import defaultdict

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import gst_config as cfg

# styling -- kept LOCAL to this module, same convention gst_checks_flow.py's
# own top-of-file comment states ("kept local so this module can be dropped
# into any workbook"); values match the rest of the tool's palette exactly.
RED = PatternFill("solid", fgColor="FFC7CE")
AMBER = PatternFill("solid", fgColor="FFEB9C")
GREEN = PatternFill("solid", fgColor="C6EFCE")
HEADFILL = PatternFill("solid", fgColor="1F3864")
TITLEF = Font(bold=True, size=13, color="1F3864")
SECTF = Font(bold=True, size=11, color="1F3864")
BORDER = Border(*[Side(style="thin", color="BFBFBF")] * 4)
CONF_FILL = {"Medium": AMBER, "Low": GREEN}
PLAIN_FONT = Font(size=10)     # shared, single instance -- the complete register can run into
BOLD_FONT = Font(bold=True, size=10)   # the tens of thousands of rows; reusing one Font/Border
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)   # object per style avoids instantiating
WRAP_TOP = Alignment(vertical="top")                      # a fresh one per cell at that scale.
CATBAND = PatternFill("solid", fgColor="DCE6F1")          # category sub-heading band


def _safe_cell(ws, row, col, value):
    """Same guard as gst_checks_flow._write_table's own copy (kept local per this module's own
    convention, not cross-imported): openpyxl auto-detects a leading '=' in a string value as a
    formula, not text. This module writes raw taxpayer trade/legal names as cell values directly
    -- a name that happens to start with '=' would otherwise corrupt the file the same way a
    stray '=' at the start of an authored note did elsewhere in this workbook (found and fixed)."""
    c = ws.cell(row, col, value)
    if c.data_type == "f" and isinstance(value, str):
        c.data_type = "s"
    return c

REQUIRED_HEADER = ["category", "search keyword", "indicative hsn/sac"]


def is_blocked_credit_master(path):
    """Content signature check (exposed for callers/tests that want to
    verify a specific path independently of gst_core.classify_folder's own
    folder-wide scan): exactly one sheet, header row reads Category /
    Search keyword / Indicative HSN/SAC (case/whitespace-insensitive)."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return False
    try:
        if len(wb.sheetnames) != 1:
            return False
        ws = wb[wb.sheetnames[0]]
        row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        cleaned = [str(c).strip().lower() if c else "" for c in row[:3]]
        return cleaned == REQUIRED_HEADER
    finally:
        wb.close()


def load_master(path):
    """Return list of dicts: category, keyword, keyword_upper, hsn_raw.
    Skips rows missing a category or keyword rather than guessing."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[0] or not r[1]:
            continue
        kw = str(r[1]).strip()
        out.append(dict(category=str(r[0]).strip(), keyword=kw,
                         keyword_upper=kw.upper(),
                         hsn_raw=str(r[2]).strip() if len(r) > 2 and r[2] else ""))
    return out


def _whole_phrase_pattern(keyword_upper):
    """Case-insensitive whole-PHRASE match, not just whole-WORD: roughly 30%
    of this master's real keywords are multi-word phrases ('GUEST HOUSE',
    'ROOM RENT', 'CLUB MEMBERSHIP') -- a single-token \\b...\\b match would
    silently never fire on those. Boundary check is against A-Z0-9 so it
    still respects word edges at both ends of the full phrase."""
    escaped = re.escape(keyword_upper)
    return re.compile(r"(?<![A-Z0-9])" + escaped + r"(?![A-Z0-9])")


def scan(b2b_rows_by_month, master):
    """b2b_rows_by_month: {month_label: [invoice-dict, ...]} -- the SAME
    invoice-level dicts gst_parsers_returns.parse_2b_excel() already
    produces per month (gstin, supplier, invno, invtype, date, invval, pos,
    rcm, rate, taxable, igst, cgst, sgst, cess, itc_avail, itc_avail_reason).

    Returns (matched_rows, category_totals):
      matched_rows -- one dict per matching invoice: the original invoice
        fields (plus 'month') and the 5 new columns the brief specified.
      category_totals -- {category: dict(count=, taxable=, itc=)}.

    A single keyword can legitimately belong to more than one category in
    this master list (e.g. 'TV' -> both Residential Furnishing and
    Electronics/Personal Use) -- every matching category is kept, comma-
    separated, not just the first."""
    patterns = [(m, _whole_phrase_pattern(m["keyword_upper"])) for m in master]
    matched_rows = []
    category_totals = {}
    for month, rows in (b2b_rows_by_month or {}).items():
        for row in rows:
            name = str(row.get("supplier") or "").upper()
            if not name:
                continue
            hits = [(m["category"], m["keyword"]) for m, pat in patterns if pat.search(name)]
            if not hits:
                continue
            cats = sorted({c for c, _k in hits})
            kws = sorted({k for _c, k in hits})
            out_row = dict(row)
            out_row["month"] = month
            out_row["Potential_Blocked_Category"] = ", ".join(cats)
            out_row["Matched_Keyword(s)"] = ", ".join(kws)
            out_row["Matched_In"] = "Trade Name"
            out_row["Match_Confidence"] = "Medium"
            out_row["HSN_Also_Matched"] = "N/A -- GSTR-2B carries no HSN/SAC column (see A5)"
            matched_rows.append(out_row)
            itc = (row.get("igst", 0.0) + row.get("cgst", 0.0)
                   + row.get("sgst", 0.0) + row.get("cess", 0.0))
            for c in cats:
                t = category_totals.setdefault(c, dict(count=0, taxable=0.0, itc=0.0))
                t["count"] += 1
                t["taxable"] += row.get("taxable", 0.0) or 0.0
                t["itc"] += itc
    return matched_rows, category_totals


ROW_HEADER = ["Month", "GSTIN", "Trade/Legal Name", "Invoice No", "Invoice Type", "Invoice Date",
              "Invoice Value", "Place of Supply", "RCM", "Rate (%)", "Taxable Value",
              "IGST", "CGST", "SGST", "Cess", "ITC Availability", "ITC Avail Reason",
              "Potential_Blocked_Category", "Matched_Keyword(s)", "Matched_In",
              "Match_Confidence", "HSN_Also_Matched"]

_COL_WIDTHS = [10, 18, 30, 16, 12, 12, 14, 16, 8, 9, 14, 12, 12, 12, 10, 12, 22, 22, 26, 14, 16, 38]


def _row_to_excel(r):
    return [r.get("month"), r.get("gstin"), r.get("supplier"), r.get("invno"), r.get("invtype"),
            r.get("date"), r.get("invval"), r.get("pos"), r.get("rcm"), r.get("rate"),
            r.get("taxable"), r.get("igst"), r.get("cgst"), r.get("sgst"), r.get("cess"),
            r.get("itc_avail"), r.get("itc_avail_reason"),
            r.get("Potential_Blocked_Category"), r.get("Matched_Keyword(s)"),
            r.get("Matched_In"), r.get("Match_Confidence"), r.get("HSN_Also_Matched")]


ALL_ROW_HEADER = ["Month", "GSTIN", "Trade/Legal Name", "Invoice No", "Invoice Type", "Invoice Date",
                   "Invoice Value", "Place of Supply", "RCM", "Rate (%)", "Taxable Value",
                   "IGST", "CGST", "SGST", "Cess", "ITC Availability", "ITC Avail Reason",
                   "Flagged as Potential Blocked Credit", "Flagged Category(ies)"]

_ALL_COL_WIDTHS = [10, 18, 30, 16, 12, 12, 14, 16, 8, 9, 14, 12, 12, 12, 10, 12, 22, 28, 30]


def _all_row_to_excel(r, flag_lookup):
    key = (r.get("month"), r.get("gstin"), str(r.get("invno") or "").strip().upper())
    cats = flag_lookup.get(key)
    return [r.get("month"), r.get("gstin"), r.get("supplier"), r.get("invno"), r.get("invtype"),
            r.get("date"), r.get("invval"), r.get("pos"), r.get("rcm"), r.get("rate"),
            r.get("taxable"), r.get("igst"), r.get("cgst"), r.get("sgst"), r.get("cess"),
            r.get("itc_avail"), r.get("itc_avail_reason"),
            ("Yes" if cats else "No"), (cats or "")]


def write_sheet(wb, matched_rows, category_totals, master_path, notes, all_rows=None):
    """Writes the 'Potential Blocked Credits' sheet. Purely additive -- does
    not touch any other sheet, and reuses this tool's established look
    (header banding, borders, freeze panes, autofilter) rather than
    inventing a new visual style.

    all_rows (added on request): every GSTR-2B B2B invoice for the FY, not
    just the ones the keyword scan flagged -- rendered as a second, complete
    invoice register BELOW the flagged table, each row tagged Yes/No against
    the SAME flag this sheet already computed (so nothing is re-derived
    twice with a chance of disagreeing). Excel allows only one autofilter
    region per sheet, so the filter is placed on THIS table (the one most
    worth slicing by column, being the larger of the two) -- the flagged
    table above keeps its header banding but not its own independent filter."""
    ws = wb.create_sheet("Potential Blocked Credits")
    ws.cell(1, 1, "POTENTIAL BLOCKED CREDITS -- SCREENING ONLY, MANUAL REVIEW REQUIRED").font = TITLEF
    sub = ("Flags invoices whose supplier Trade/Legal name matches a Section 17(5) blocked-credit "
           "keyword list you supplied. This is a screening aid, not a determination -- every row "
           "needs manual review before any ITC figure is adjusted. See the notes at the bottom for "
           "what this sheet can and cannot detect from GSTR-2B.")
    c = ws.cell(2, 1, sub)
    c.font = Font(size=9, italic=True)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    r = 4
    if category_totals:
        ws.cell(r, 1, "Blocked Credit Summary (by category)").font = SECTF
        r += 1
        for i, h in enumerate(("Category", "Invoices Flagged", "Taxable Value (Rs)",
                                "ITC -- IGST+CGST+SGST+Cess (Rs)"), 1):
            cc = ws.cell(r, i, h)
            cc.fill = HEADFILL; cc.font = HEADER_FONT; cc.border = BORDER
        r += 1
        for cat in sorted(category_totals):
            t = category_totals[cat]
            ws.cell(r, 1, cat).border = BORDER
            ws.cell(r, 2, t["count"]).border = BORDER
            v3 = ws.cell(r, 3, round(t["taxable"], 2)); v3.number_format = "#,##0.00"; v3.border = BORDER
            v4 = ws.cell(r, 4, round(t["itc"], 2)); v4.number_format = "#,##0.00"; v4.border = BORDER
            r += 1
        r += 2

    ws.cell(r, 1, f"Flagged Invoices -- Complete Detail ({len(matched_rows)} invoice(s))").font = SECTF
    r += 1
    flagged_hdr_row = r
    for i, h in enumerate(ROW_HEADER, 1):
        cc = ws.cell(flagged_hdr_row, i, h)
        cc.fill = HEADFILL; cc.font = HEADER_FONT
        cc.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cc.border = BORDER
    ws.row_dimensions[flagged_hdr_row].height = 30
    r = flagged_hdr_row + 1
    conf_col = ROW_HEADER.index("Match_Confidence") + 1
    flag_lookup = {}   # (month, gstin, invno-upper) -> category string, shared with the register below
    # Grouped by category on request: previously sorted only by month/GSTIN, so invoices in the
    # same category (e.g. all 47 Accommodation ones) were scattered through the table rather than
    # sitting together under their own heading. The summary table above is untouched -- only this
    # detail table's ROW ORDER changed, plus a banded sub-heading row per category.
    by_cat = defaultdict(list)
    for row in matched_rows:
        cats = [c.strip() for c in (row.get("Potential_Blocked_Category") or "(uncategorized)").split(",")]
        # An invoice matching MULTIPLE categories (a handful of the master's own keywords span more
        # than one category -- e.g. 'ACCOMMODATION'/'GUEST HOUSE' hit both 'Accommodation' and
        # 'Residential/guest house') must appear under EACH of its categories, not just a combined
        # one -- otherwise "Accommodation" here could show fewer invoices than the summary table's
        # own count for that same category (confirmed this actually happened: 45 shown vs 47 in the
        # summary before this fix). Same per-category counting convention the summary table itself
        # already uses.
        for c in cats:
            by_cat[c].append(row)
    for cat in sorted(by_cat.keys()):
        cat_rows = sorted(by_cat[cat], key=lambda x: (str(x.get("month") or ""), str(x.get("gstin") or "")))
        cc = ws.cell(r, 1, f"{cat}  ({len(cat_rows)} invoice(s))")
        cc.font = Font(bold=True, size=10, color="1F3864")
        for i in range(1, len(ROW_HEADER) + 1):
            ws.cell(r, i).fill = CATBAND
        r += 1
        for row in cat_rows:
            for i, v in enumerate(_row_to_excel(row), 1):
                cc = _safe_cell(ws, r, i, v)
                cc.border = BORDER
                cc.font = PLAIN_FONT
                if isinstance(v, float):
                    cc.number_format = "#,##0.00"
            fill = CONF_FILL.get(ws.cell(r, conf_col).value)
            if fill:
                ws.cell(r, conf_col).fill = fill
                ws.cell(r, conf_col).font = BOLD_FONT
            flag_lookup[(row.get("month"), row.get("gstin"),
                         str(row.get("invno") or "").strip().upper())] = row.get("Potential_Blocked_Category")
            r += 1
    if not matched_rows:
        ws.cell(r, 1, "No invoices matched the blocked-credit keyword list.").font = Font(italic=True)
        r += 1
    r += 2

    reg_hdr_row = None
    if all_rows:
        ws.cell(r, 1, f"Complete GSTR-2B Invoice Register -- All Invoices, Flagged and Unflagged "
                      f"({len(all_rows)} invoice(s))").font = SECTF
        r += 1
        c2 = ws.cell(r, 1, "Every B2B invoice parsed from your merged GSTR-2B workbook for this FY -- "
                          "not just the ones flagged above -- so the full source data sits alongside "
                          "the screening result on one sheet. The last two columns tie each row back "
                          "to the flagged table above (same Yes/No, same category) rather than "
                          "re-running the match a second time.")
        c2.font = Font(size=9, italic=True); c2.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
        reg_hdr_row = r
        for i, h in enumerate(ALL_ROW_HEADER, 1):
            cc = ws.cell(reg_hdr_row, i, h)
            cc.fill = HEADFILL; cc.font = HEADER_FONT
            cc.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            cc.border = BORDER
        ws.row_dimensions[reg_hdr_row].height = 30
        r = reg_hdr_row + 1
        flag_col = ALL_ROW_HEADER.index("Flagged as Potential Blocked Credit") + 1
        for row in sorted(all_rows, key=lambda x: (str(x.get("month") or ""), str(x.get("gstin") or ""))):
            vals = _all_row_to_excel(row, flag_lookup)
            for i, v in enumerate(vals, 1):
                cc = _safe_cell(ws, r, i, v)
                cc.border = BORDER
                cc.font = PLAIN_FONT
                if isinstance(v, float):
                    cc.number_format = "#,##0.00"
            if ws.cell(r, flag_col).value == "Yes":
                ws.cell(r, flag_col).fill = AMBER
                ws.cell(r, flag_col).font = BOLD_FONT
            r += 1

    r += 1
    for n in notes:
        c = ws.cell(r, 1, "Note: " + n)
        c.font = Font(size=9, italic=True, color="808080")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    # BUG FIX (reported against a real run): freeze_panes here used to be set to the register's
    # own header row (e.g. row 2561 on a real file) or the flagged table's header row as a
    # fallback. Both are wrong for the SAME reason -- Excel's freeze_panes pins EVERYTHING above
    # the split point as a non-scrolling block, which only behaves like a normal "frozen header"
    # when that block is a handful of rows. Freezing at row 2561 pinned the entire ~2500-row
    # flagged-invoices table above it too, which is exactly what looked like "scrolling doesn't
    # work" -- the frozen block was too large to fit on screen and couldn't itself be scrolled
    # into. This sheet has three stacked sections (summary / flagged table / complete register),
    # each with its own header at a different depth -- there's no single row where a SMALL,
    # header-only freeze makes sense for the whole sheet the way it does on every other sheet in
    # this workbook (which have one table, one header, near the top). So: no freeze_panes here at
    # all, rather than a freeze that's technically "a header row" but practically broken. The
    # autofilter (unaffected by this bug -- it's a dropdown on the header row, not a scroll pin)
    # is kept exactly as before.
    if reg_hdr_row:
        ws.auto_filter.ref = f"A{reg_hdr_row}:{get_column_letter(len(ALL_ROW_HEADER))}{max(r - 2, reg_hdr_row)}"
        for i, w in enumerate(_ALL_COL_WIDTHS[:len(ALL_ROW_HEADER)], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    else:
        ws.auto_filter.ref = f"A{flagged_hdr_row}:{get_column_letter(len(ROW_HEADER))}{max(r - 2, flagged_hdr_row)}"
        for i, w in enumerate(_COL_WIDTHS[:len(ROW_HEADER)], 1):
            ws.column_dimensions[get_column_letter(i)].width = w


def build_and_write(wb, master_path, b2b_rows_by_month):
    """Entry point called from master_build.py. Returns a short status
    string for the run-summary printout. Never raises -- any failure
    degrades to a skipped sheet with a clear reason, matching this tool's
    existing error-handling convention for every other optional source."""
    if not master_path:
        master_path = cfg.BLOCKED_ITC_MASTER_FALLBACK_PATH
    if not master_path:
        return ("SKIPPED -- no blocked-credit master file found (content-detected: single "
                 "sheet, header 'Category / Search keyword / Indicative HSN/SAC') and no "
                 "fallback path configured in gst_config.BLOCKED_ITC_MASTER_FALLBACK_PATH.")
    try:
        master = load_master(master_path)
    except Exception as e:
        return f"SKIPPED -- could not read the master file ({e!r})."
    if not master:
        return f"SKIPPED -- master file {master_path!r} has no usable keyword rows."
    if not b2b_rows_by_month or not any(b2b_rows_by_month.values()):
        return "SKIPPED -- no GSTR-2B invoice-level data available for any month this run."

    matched_rows, category_totals = scan(b2b_rows_by_month, master)
    all_rows = []
    for month, month_rows in b2b_rows_by_month.items():
        for row in month_rows:
            all_rows.append(dict(row, month=row.get("month", month)))
    n_cats_master = len({m["category"] for m in master})
    notes = [
        "Detection is Trade/Legal-name keyword matching ONLY. GSTR-2B's 'B2B' sheet carries no "
        "line-item description and no HSN/SAC column -- Description-keyword and HSN/SAC-prefix "
        "matching (as the master list's own 'Indicative HSN/SAC' column is designed for) are not "
        "possible from this data source and are not attempted. This is the same limitation the "
        "'HSN & Fraud Pattern Checks' sheet's finding A5 already documents for HSN-based blocked-"
        "ITC screening -- this sheet is the keyword/trade-name-based alternative for the SAME gap, "
        "not a contradiction of A5.",
        "Match_Confidence is capped at 'Medium' for the same reason -- 'High' would need a "
        "keyword+HSN combination that can't occur from GSTR-2B alone.",
        f"Master list: {len(master)} keyword(s) across {n_cats_master} categor"
        f"{'y' if n_cats_master == 1 else 'ies'}, read from {master_path}.",
        f"Complete register below lists all {len(all_rows)} GSTR-2B invoice(s) for the FY -- "
        f"{len(matched_rows)} flagged, {len(all_rows) - len(matched_rows)} not flagged.",
        "The flagged-invoices table above is grouped by category, one heading per category, so "
        "every invoice in a category sits together (matching that category's count in the summary "
        "table). An invoice matching more than one category (a handful of this master's own "
        "keywords span two categories, e.g. 'Accommodation' and 'Residential/guest house') appears "
        "once under EACH matching category -- so it can legitimately show up more than once in "
        "this table, by design, not a duplication bug.",
        "Screening aid only -- every flagged invoice needs manual review; this sheet never "
        "adjusts, removes, or blocks any ITC figure anywhere else in this workbook.",
    ]
    write_sheet(wb, matched_rows, category_totals, master_path, notes, all_rows=all_rows)
    n_cats_hit = len(category_totals)
    return (f"OK -- {len(matched_rows)} invoice(s) flagged across {n_cats_hit} "
            f"categor{'y' if n_cats_hit == 1 else 'ies'} (Trade-name match only, see sheet notes); "
            f"complete register of {len(all_rows)} invoice(s) also written.")
