#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GST CHECKS FORENSIC
===================
CONSOLIDATED FILE -- contains what used to be: forensic_checks.py, filing_compliance.py

The tool was reorganised from 19 .py files into 9 for easier sharing. Nothing
in the analytical logic was rewritten during that move: each section below is
the original module's code verbatim, with only (a) intra-project imports
repointed at the new file names, (b) its standalone __main__ demo block
removed, and (c) the renames listed under MERGE NOTES applied where two merged
modules happened to define the same top-level name with different bodies.

MERGE NOTES for this file:
  (none -- no name collisions)
"""


# ============================================================================
# ==== SECTION: forensic_checks.py  (was a standalone module before consolidation)
# ============================================================================
"""
FORENSIC CHECKS  --  Part 2 of GST_Forensic_Comparison_Framework_v1.md
=========================================================================
Implements, generically (not hardcoded to any one taxpayer):

  R13  Turnover-gap rule (zero extra documents needed -- runs off files
       already in the pipeline: GSTR-9C Table 7 vs GSTR-1 Table 8 'exemp').
  R14  Four-way ITC reconciliation (3B vs 2B vs Table-8A vs GSTR-9C books) --
       also zero extra documents beyond what's already optional in the
       pipeline (GSTR-9, GSTR-9C, Table 8A -- all degrade gracefully if any
       one of the four sources is missing; the check simply narrows to
       however many of the four sources ARE available, minimum two, and
       says explicitly which sources were used).
  R0-R12  Generic Balance-Sheet/P&L rule engine. Operates on a STRUCTURED
       INPUT (a plain dict of line-items -> {FY22: val, FY23: val}), not on
       OCR. See bs_pl_input.py / OCR_LIMITATION.md for why: this taxpayer's
       BS/P&L PDF is a SCANNED CamScanner export (confirmed: 0 extractable
       text, 2 raster images/page). OCR on it (pytesseract, tested) mis-reads
       real digits (e.g. Finance Costs 49,73,007.06 misread as 43,73,007.05
       -- a wrong-by-one-digit error that would itself generate a false
       forensic FLAG). A forensic tool cannot silently trust OCR'd rupee
       figures -- see OCR_LIMITATION.md for the full writeup and the two
       supported paths forward.

  Cancelled E-Invoices -- aggregates gst_scrutiny_tool.parse_einv()'s new
       'cancelled' list across every month, and runs the two defensive
       cross-checks already on the project's own backlog (B9 / D2):
       (a) a cancelled e-invoice's invoice number still appearing in GSTR-1
           B2B outward supply (should have been removed/not filed);
       (b) a cancelled e-invoice with a live outward EWB against the same
           invoice number (goods-movement document survives a cancelled tax
           document -- a real red flag).

HARD RULE (same as the rest of this codebot): every check that cannot run
because a source is missing says so explicitly (INFO/SKIP), never silently
skips, never fabricates a number, never lets a missing file crash the run.
"""

import re
import datetime as _dt


FLAG, REVW, INFO, PASS, SKIP = "FLAG", "REVIEW", "INFO", "PASS", "SKIPPED"
SEV_ORDER = {FLAG: 0, REVW: 1, INFO: 2, PASS: 3, SKIP: 4}
TOL_RS = 200.0            # small-rupee rounding tolerance for exact reconciliations
MATERIALITY_PCT = 0.01    # 1% of turnover, per the framework's own R14 threshold
MATERIALITY_FLOOR = 100000.0  # ...or Rs 1L, whichever is LOWER (i.e. the tighter test)


class Finding:
    __slots__ = ("ref", "title", "severity", "detail", "numbers")
    def __init__(self, ref, title, severity, detail, numbers=None):
        self.ref, self.title, self.severity, self.detail = ref, title, severity, detail
        self.numbers = numbers or {}


def _num(v):
    if v is None:
        return None
    try:
        return float(str(v).replace(",", "").strip())
    except ValueError:
        return None


# ======================================================================
# R13 -- Turnover-gap rule (zero-dependency: already-supplied files only)
# ======================================================================
def check_turnover_gap(gstr9c, gstr1_exemp_rows_by_month):
    """gstr9c: dict from annual_return_parser.parse_gstr9c().
    gstr1_exemp_rows_by_month: {month: [row,...]} of GSTR-1's Table-8
    ('exemp' sheet) data rows, ACROSS EVERY MONTH SUPPLIED -- caller builds
    this from gst_scrutiny_tool.parse_gstr1()'s own nil/exempt reader or by
    re-reading the 'exemp' sheet directly; passed in rather than re-parsed
    here so this module has no direct GSTR-1-file dependency of its own."""
    if not gstr9c.get("available"):
        return Finding("R13", "Turnover-gap rule (GSTR-9C Table 7 vs GSTR-1 Table 8)", INFO,
                        "GSTR-9C not supplied -- cannot run. Needs zero NEW documents once GSTR-9C is "
                        "provided (this check is otherwise fully computable from files already in the "
                        "12-file set).")
    adj = gstr9c.get("exempt_nil_nongst_adjustment")
    if adj is None:
        return Finding("R13", "Turnover-gap rule", INFO,
                        "GSTR-9C Table 7B (exempt/nil/non-GST adjustment) not found in the extract -- "
                        "cannot test. If the supplied GSTR-9C is a 'System Drafted' pre-filing copy, "
                        "this table may be blank there even though it is populated on the as-filed PDF.")
    total_exempt_rows = sum(len(v) for v in (gstr1_exemp_rows_by_month or {}).values())
    months_checked = sorted((gstr1_exemp_rows_by_month or {}).keys())
    if abs(adj) < TOL_RS:
        return Finding("R13", "Turnover-gap rule", PASS,
                        f"GSTR-9C Table 7B exempt/nil/non-GST adjustment is Rs {adj:,.2f} (nil) -- "
                        "no turnover-gap risk to test.", {"adjustment": adj})
    if total_exempt_rows == 0 and months_checked:
        return Finding("R13", "Turnover-gap rule -- UNSUPPORTED EXEMPT TURNOVER", FLAG,
                        f"GSTR-9C Table 7B declares an exempt/nil/non-GST turnover ADJUSTMENT of "
                        f"Rs {adj:,.2f}, backing out taxable turnover from gross turnover. But GSTR-1's "
                        f"own Table 8 ('exemp' sheet) shows ZERO data rows across every month supplied "
                        f"({', '.join(months_checked)}) -- this taxpayer declared no nil-rated, exempt, "
                        f"or non-GST outward supply in ANY GSTR-1 filed for these months. "
                        f"Rs {adj:,.2f} of revenue booked in the audited financials has no traceable "
                        f"source anywhere in the GST returns being reconciled. Plausible legitimate "
                        f"explanations exist (high-seas sales, sale-before-customs-clearance, other "
                        f"'no supply' items under Schedule III) but NONE of those has its own row in "
                        f"GSTR-1 Table 8 either -- if that is the real explanation it must be evidenced "
                        f"separately (Bills of Entry, high-seas sale agreements), not inferred from this "
                        f"plug figure alone.",
                        {"unsupported adjustment": adj, "exemp rows found": total_exempt_rows,
                         "months checked": len(months_checked)})
    return Finding("R13", "Turnover-gap rule", REVW,
                    f"GSTR-9C Table 7B adjustment Rs {adj:,.2f}; GSTR-1 Table 8 shows {total_exempt_rows} "
                    f"data row(s) across {len(months_checked)} month(s) checked -- some support exists, "
                    "reconcile the adjustment figure against the sum of those rows to confirm full coverage.",
                    {"adjustment": adj, "exemp rows found": total_exempt_rows})


# ======================================================================
# R14 -- Four-way ITC reconciliation
# ======================================================================
def check_four_way_itc(gstr9, gstr2b_fy_total, table8a, gstr9c, annual_turnover=None):
    """
    gstr9:  dict from annual_return_parser.parse_gstr9() -- uses table6a_total
    gstr2b_fy_total: dict {'igst':..,'cgst':..,'sgst':..,'cess':..} summed by
        the CALLER across every quarter/month of GSTR-2B actually supplied
        (this module does not re-read GSTR-2B itself), or None if 2B wasn't
        supplied at all this run.
    table8a: dict from annual_return_parser.parse_table_8a()
    gstr9c: dict from annual_return_parser.parse_gstr9c() -- uses itc_per_books
    Runs with however many of the 4 sources are actually available (minimum
    2 needed for any comparison at all); explicitly lists which were used.
    """
    sources = {}
    if gstr9.get("available") and gstr9.get("table6a_total") is not None:
        sources["3B (GSTR-9 Table 6A)"] = gstr9["table6a_total"]
    if gstr2b_fy_total:
        t = sum(v or 0 for k, v in gstr2b_fy_total.items() if k != "_source")
        sources["2B (recomputed FY total)"] = t
    if table8a.get("available") and table8a.get("totals"):
        sources["8A (invoice-level, ITC-available=Yes)"] = table8a["totals"].get("total")
    if gstr9c.get("available") and gstr9c.get("itc_per_books") is not None:
        sources["Books (GSTR-9C Table 12A)"] = gstr9c["itc_per_books"]

    if len(sources) < 2:
        return Finding("R14", "Four-way ITC reconciliation (3B / 2B / 8A / Books)", INFO,
                        f"Only {len(sources)} of 4 ITC sources available "
                        f"({', '.join(sources) or 'none'}) -- need at least 2 to compare. Supply "
                        "GSTR-9, GSTR-2B, Table 8A, and/or GSTR-9C to enable.", {})

    # Books, if present, is the odd one out by construction (auditor-certified, ex-GST-return);
    # the three return-side figures should mutually agree closely.
    return_side = {k: v for k, v in sources.items() if k != "Books (GSTR-9C Table 12A)"}
    books = sources.get("Books (GSTR-9C Table 12A)")

    lines = [f"{k}: Rs {v:,.2f}" for k, v in sources.items()]
    detail = "Sources compared: " + "; ".join(lines) + ". "

    return_side_spread = None
    if len(return_side) >= 2:
        vals = list(return_side.values())
        return_side_spread = (max(vals) - min(vals)) / max(vals) * 100 if max(vals) else 0
        detail += f"Return-side spread (max-min)/max = {return_side_spread:.2f}%. "

    materiality = min(MATERIALITY_FLOOR, (annual_turnover or 0) * MATERIALITY_PCT) if annual_turnover else MATERIALITY_FLOOR

    if books is not None and return_side:
        gaps = {k: (v - books) for k, v in return_side.items()}
        gap_lines = "; ".join(f"{k} - Books = Rs {g:,.2f}" for k, g in gaps.items())
        detail += f"Books-vs-return gaps: {gap_lines}. "
        material_gap = any(abs(g) > materiality for g in gaps.values())
        one_directional = len({1 if g > 0 else -1 for g in gaps.values() if abs(g) > 1}) <= 1
        return_side_clean = (return_side_spread is None) or (return_side_spread < 1.0)

        if material_gap and one_directional and return_side_clean:
            sev = FLAG
            detail += (f"All return-side figures are mutually consistent (within ~1%), but ALL of them "
                       f"exceed the books figure by more than the materiality threshold "
                       f"(Rs {materiality:,.2f} = lower of 1% of turnover or Rs 1L), in the SAME "
                       f"direction. This is a BOOKS-VS-RETURNS gap specifically -- distinct from a "
                       f"return-internal mismatch (which would show up as return-side inconsistency "
                       f"instead) -- and points to a different root cause: ITC claimed in the return but "
                       f"not credited to the books ITC ledger, reversed post-facto in the books without a "
                       f"3B reversal, or a genuine GSTR-9C Table 12B/12C under-population (both shown "
                       f"as 0.00 would need to fully explain this gap for the auditor's 'Un-reconciled "
                       f"ITC = 0.00' certification to be defensible).")
        elif material_gap:
            sev = REVW
            detail += "A material gap exists but sources are not cleanly one-directional/consistent -- review individually."
        else:
            sev = PASS
            detail += "No gap exceeds the materiality threshold."
    else:
        sev = INFO
        detail += "Books figure (GSTR-9C 12A) not available -- can only confirm return-side mutual consistency, not test against books."

    return Finding("R14", "Four-way ITC reconciliation (3B / 2B / 8A / Books)", sev, detail,
                    {k: round(v, 2) for k, v in sources.items() if v is not None})


# ======================================================================
# R0-R12 -- Generic Balance-Sheet / P&L rule engine (structured input)
# ======================================================================
# Label-family regex matchers -- taxpayer-agnostic. Match on FAMILIES of
# common line-item naming, not exact strings (per the framework's own
# instruction: "line-item naming isn't standardized across companies'
# financials"). Structured input format (see bs_pl_input.py):
#   {
#     "total_assets": {"fy_prior": .., "fy_current": ..},
#     "total_equity_liab": {...}, "revenue_from_operations": {...},
#     "other_income": {...}, "trade_payables": {...}, "trade_receivables": {...},
#     "inventories": {...}, "fixed_assets_tangible": {...}, "depreciation": {...},
#     "non_current_investments": {...}, "short_term_provisions": {...},
#     "other_expenses": {...}, "finance_costs": {...}, "share_capital": {...},
#     "reserves_and_surplus": {...}, "net_profit_after_tax": {...},
#   }
# Any key the caller doesn't have simply isn't tested -- "not tested, not
# supplied" is emitted for it, never silently skipped without a trace.
BS_PL_LABEL_FAMILIES = {
    "revenue_from_operations": ["revenue from operations", "turnover", "sales", "gross revenue",
                                 "income from operations"],
    "other_income": ["other income", "miscellaneous income", "non-operating income"],
    "trade_payables": ["trade payables", "sundry creditors", "creditors for goods", "creditors for services"],
    "trade_receivables": ["trade receivables", "sundry debtors", "debtors"],
    "inventories": ["inventories", "stock-in-trade", "finished goods", "raw materials", "work-in-progress", "wip"],
    "fixed_assets_tangible": ["fixed assets", "property, plant and equipment", "tangible assets",
                              "capital work in progress", "capital work-in-progress"],
    "non_current_investments": ["non-current investments", "non current investments", "current investments"],
    "short_term_provisions": ["short term provisions", "long term provisions", "contingent liabilities",
                              "provision for taxation"],
    "other_expenses": ["other expenses", "miscellaneous expenses", "administrative expenses"],
    "finance_costs": ["finance costs", "interest and finance charges"],
    "share_capital": ["share capital", "share application money pending allotment"],
    "reserves_and_surplus": ["reserves and surplus"],
}


def _fy_pair(d):
    """Accept either {'fy_prior':x,'fy_current':y} or a bare 2-tuple/list."""
    if d is None:
        return None, None
    if isinstance(d, dict):
        return _num(d.get("fy_prior")), _num(d.get("fy_current"))
    if isinstance(d, (list, tuple)) and len(d) >= 2:
        return _num(d[0]), _num(d[1])
    return None, None


def check_bs_pl_rules(bs_pl_data, gstr9c=None, bo_profile=None, gstr1_table8_all_zero=None):
    """Run rules R0-R12 on a STRUCTURED input dict (see BS_PL_LABEL_FAMILIES
    above for expected keys). Never fed raw OCR text -- see module docstring.
    Returns a list of Finding, one per rule tested, PLUS one INFO Finding per
    rule that could NOT be tested because its key wasn't supplied (so the
    final report always shows a complete checklist, per the framework's own
    'Implementation note' in section 2.6)."""
    F = []
    d = bs_pl_data or {}

    # ---- R0: data-integrity gate, always first ----
    ta = d.get("total_assets"); tel = d.get("total_equity_liab")
    if ta and tel:
        ap, ac = _fy_pair(ta); lp, lc = _fy_pair(tel)
        ok_prior = (ap is not None and lp is not None and abs(ap - lp) < 1.0)
        ok_curr = (ac is not None and lc is not None and abs(ac - lc) < 1.0)
        if ok_prior and ok_curr:
            F.append(Finding("R0", "Balance Sheet self-balances (Total Assets = Total Equity+Liabilities)",
                              PASS, f"Prior FY: {ap:,.2f} = {lp:,.2f}. Current FY: {ac:,.2f} = {lc:,.2f}. "
                              "Pre-flight gate passed -- downstream BS checks may proceed.", {}))
        else:
            F.append(Finding("R0", "Balance Sheet self-balances", FLAG,
                              f"Prior FY: {ap} vs {lp}. Current FY: {ac} vs {lc}. MISMATCH -- this is a "
                              "data-integrity/extraction error, not a real accounting fact. HALTING all "
                              "downstream BS checks below until this is fixed (do not trust any R1-R12 "
                              "result computed alongside a failed R0).", {}))
            F.append(Finding("R1-R12", "Downstream BS/PL checks", SKIP,
                              "Skipped -- R0 pre-flight gate failed (see above).", {}))
            return F
    else:
        F.append(Finding("R0", "Balance Sheet self-balances", INFO,
                          "total_assets / total_equity_liab not supplied -- pre-flight gate not tested. "
                          "Downstream checks proceed but are UNVERIFIED against a balancing BS.", {}))

    # ---- R1: Revenue vs GSTR-9C Table 5 ----
    rev = d.get("revenue_from_operations")
    if rev and gstr9c and gstr9c.get("available"):
        _, rc = _fy_pair(rev)
        g9c_turnover = gstr9c.get("turnover_declared_gstr9")
        if rc is not None and g9c_turnover is not None:
            diff = rc - g9c_turnover
            sev = PASS if abs(diff) < TOL_RS else FLAG
            F.append(Finding("R1", "Revenue from Operations vs GSTR-9C Table 5A/5P/5Q", sev,
                              f"P&L revenue {rc:,.2f} vs GSTR-9C {g9c_turnover:,.2f}, diff {diff:,.2f}. "
                              + ("Exact match." if sev == PASS else
                                 "Variance -- demand revenue-recognition policy note + sales register."),
                              {"P&L revenue": rc, "GSTR-9C turnover": g9c_turnover}))
        else:
            F.append(Finding("R1", "Revenue vs GSTR-9C Table 5", INFO, "Figures incomplete.", {}))
    else:
        F.append(Finding("R1", "Revenue vs GSTR-9C Table 5", INFO,
                          "not tested -- revenue_from_operations and/or GSTR-9C not supplied.", {}))

    # ---- R2: Other Income -- informational only (needs GSTR-1 HSN cross-match, out of scope here) ----
    oi = d.get("other_income")
    if oi:
        op, oc = _fy_pair(oi)
        F.append(Finding("R2", "Other Income -- taxable-component screen", INFO,
                          f"Other Income: prior {op}, current {oc}. Manually screen for scrap/container/"
                          "rent-income components with a plausible HSN/SAC and confirm each is declared "
                          "in GSTR-1 (no automated per-line breakup available from a BS/PL summary "
                          "figure alone -- needs the itemized Other-Income note).", {}))
    else:
        F.append(Finding("R2", "Other Income", INFO, "not tested -- other_income not supplied.", {}))

    # ---- R3: Trade Payables -- 180-day ITC reversal (Rule 37) needs ageing schedule ----
    tp = d.get("trade_payables")
    if tp:
        pp, pc = _fy_pair(tp)
        note = ""
        rev_p, rev_c = _fy_pair(d.get("revenue_from_operations")) if d.get("revenue_from_operations") else (None, None)
        if pp is not None and pc is not None and rev_p is not None and rev_c is not None and rev_p:
            rev_growth = (rev_c - rev_p) / rev_p * 100
            payable_change = (pc - pp) / pp * 100 if pp else None
            if payable_change is not None and rev_growth > 10 and payable_change < 0:
                note = (f" NOTE: revenue grew {rev_growth:.1f}% while trade payables FELL "
                        f"{-payable_change:.1f}% -- worth checking whether large invoices were paid "
                        "faster than normal (reducing 180-day-reversal exposure) or whether an "
                        "unusually large volume of purchases went through cash/other non-payable routes.")
        F.append(Finding("R3", "Trade Payables -- Rule 37 (180-day ITC reversal)", INFO,
                          f"Trade Payables: prior {pp}, current {pc}.{note} Cannot test without an ageing "
                          "schedule (which invoice, which supplier, how many days outstanding) cross-matched "
                          "against GSTR-2B/8A invoice dates -- demand the Trade Payables ageing schedule.",
                          {}))
    else:
        F.append(Finding("R3", "Trade Payables", INFO, "not tested -- trade_payables not supplied.", {}))

    # ---- R4: Trade Receivables -- bad-debt/GST-liability-reduction cross-check ----
    tr = d.get("trade_receivables")
    if tr:
        rp, rc = _fy_pair(tr)
        F.append(Finding("R4", "Trade Receivables -- bad-debt GST relief screen", INFO,
                          f"Trade Receivables: prior {rp}, current {rc}. GST law does not permit output-tax "
                          "adjustment for bad debts. Demand confirmation of any write-off during the year "
                          "and check for a corresponding (impermissible) GST liability reduction in GSTR-1 CN.",
                          {}))
    else:
        F.append(Finding("R4", "Trade Receivables", INFO, "not tested -- trade_receivables not supplied.", {}))

    # ---- R5: Inventories -- Section 17(5)(h) ----
    inv = d.get("inventories")
    if inv:
        ip, ic = _fy_pair(inv)
        swing_note = ""
        if ip is not None and ic is not None and ip:
            swing_pct = (ic - ip) / ip * 100
            if abs(swing_pct) > 30:
                swing_note = (f" Swing of {swing_pct:+.1f}% year-on-year is large enough to warrant a "
                               "quantitative stock reconciliation (Rule 56 register) -- especially since "
                               "GSTR-1's own HSN-summary quantity column is a known structural zero "
                               "(see project context, section 5), so quantity-based verification isn't "
                               "possible from the GST-return side alone.")
        F.append(Finding("R5", "Inventories -- Section 17(5)(h) (lost/destroyed/written-off stock)", INFO,
                          f"Inventories: prior {ip}, current {ic}.{swing_note} Demand the stock/quantity "
                          "register (Rule 56) for the year.", {}))
    else:
        F.append(Finding("R5", "Inventories", INFO, "not tested -- inventories not supplied.", {}))

    # ---- R6: Fixed Assets -- Sec 18(6) disposal / Sec 16(3) dual-claim ----
    fa = d.get("fixed_assets_tangible"); dep = d.get("depreciation")
    if fa:
        fp, fc = _fy_pair(fa)
        _, depc = _fy_pair(dep) if dep else (None, None)
        note = ""
        if fp is not None and fc is not None and depc is not None:
            expected_closing_if_no_disposal = fp - depc
            residual = fc - expected_closing_if_no_disposal
            if abs(residual) > TOL_RS:
                note = (f" Net movement (prior {fp:,.2f} - depreciation {depc:,.2f} = expected closing "
                        f"{expected_closing_if_no_disposal:,.2f}) vs actual closing {fc:,.2f} leaves a "
                        f"residual of {residual:,.2f} -- consistent with either a disposal or an addition "
                        "during the year (net movement alone cannot distinguish which); demand the full "
                        "fixed-asset schedule (gross block, additions, disposals, accumulated depreciation).")
        F.append(Finding("R6", "Fixed Assets -- Sec 18(6) disposal / Sec 16(3) dual-claim", INFO,
                          f"Tangible Fixed Assets: prior {fp}, current {fc}.{note}", {}))
    else:
        F.append(Finding("R6", "Fixed Assets", INFO, "not tested -- fixed_assets_tangible not supplied.", {}))

    # ---- R7: Investments -- Schedule I deemed supply ----
    ni = d.get("non_current_investments")
    if ni:
        np_, nc = _fy_pair(ni)
        note = ""
        if np_ is not None and nc is not None and np_ and (nc - np_) / abs(np_) > 0.5:
            note = (f" Jump of Rs {nc-np_:,.2f} ({(nc-np_)/np_*100:+.1f}%) year-on-year with no "
                    "cash-flow-statement source identified from this summary figure alone -- demand "
                    "the investment movement schedule + funding source (cash vs asset-in-kind, "
                    "counterparty) to rule out a Schedule-I deemed-supply trigger.")
        F.append(Finding("R7", "Non-Current Investments -- Schedule I deemed supply", INFO,
                          f"Non-Current Investments: prior {np_}, current {nc}.{note}", {}))
    else:
        F.append(Finding("R7", "Investments", INFO, "not tested -- non_current_investments not supplied.", {}))

    # ---- R8: Provisions -- undisclosed indirect-tax exposure vs BO Profile DRC/Case ----
    prov = d.get("short_term_provisions")
    if prov:
        pp, pc = _fy_pair(prov)
        note = ""
        if pp is not None and pc is not None and pp and pc / pp > 1.5:
            drc_total = None
            if bo_profile and bo_profile.get("drc_payments"):
                drc_total = sum(x.get("total", 0) for x in bo_profile["drc_payments"])
            note = (f" Grew {pc/pp:.2f}x year-on-year. If any component is a provision for a GST demand "
                    "or disputed tax, it is the company's own admission of a liability not yet reflected "
                    "in the returns -- demand the provisions breakup and cross-map against the BO "
                    f"Profile's DRC/Case sections"
                    + (f" (which already show Rs {drc_total:,.2f} in total recorded DRC payments for "
                       "cross-reference)." if drc_total is not None else " (BO Profile not supplied -- "
                       "cannot cross-check)."))
        F.append(Finding("R8", "Provisions -- undisclosed indirect-tax exposure", INFO,
                          f"Short Term Provisions: prior {pp}, current {pc}.{note}", {}))
    else:
        F.append(Finding("R8", "Provisions", INFO, "not tested -- short_term_provisions not supplied.", {}))

    # ---- R9: Other Expenses -- blocked-credit categories hide here ----
    oe = d.get("other_expenses")
    if oe:
        _, oec = _fy_pair(oe)
        rev_p2, rev_c2 = _fy_pair(d.get("revenue_from_operations")) if d.get("revenue_from_operations") else (None, None)
        pct_note = f" ({oec/rev_c2*100:.1f}% of revenue)" if (oec and rev_c2) else ""
        F.append(Finding("R9", "Other Expenses -- Section 17(5) blocked-credit screen", INFO,
                          f"Other Expenses: current {oec}{pct_note}, zero itemization from a BS/PL summary "
                          "figure alone. Section 17(5) blocked-credit categories (rent-a-cab, catering, "
                          "club/membership, CSR, works contract on immovable property, gifts) typically "
                          "hide inside this catch-all note. Demand the itemized breakup and cross-match "
                          "each line's supplier SAC against GSTR-2B.", {}))
    else:
        F.append(Finding("R9", "Other Expenses", INFO, "not tested -- other_expenses not supplied.", {}))

    # ---- R10: Finance Costs -- interest exempt, bank charges taxable+ITC-eligible ----
    fc_ = d.get("finance_costs")
    if fc_:
        fp, fcur = _fy_pair(fc_)
        note = ""
        if fp is not None and fcur is not None and fp and fcur / fp > 2:
            note = (f" Jumped {fcur/fp:.2f}x year-on-year. Interest itself is GST-exempt, but processing "
                    "fees/guarantee commission/bank charges attract GST and generate eligible ITC -- "
                    "demand the note breakup and confirm ITC on the GST-bearing portion is present in GSTR-2B.")
        F.append(Finding("R10", "Finance Costs -- interest-exempt vs GST-bearing bank charges", INFO,
                          f"Finance Costs: prior {fp}, current {fcur}.{note}", {}))
    else:
        F.append(Finding("R10", "Finance Costs", INFO, "not tested -- finance_costs not supplied.", {}))

    # ---- R11: Share Capital -- non-cash consideration (rare trigger) ----
    F.append(Finding("R11", "Share Capital -- non-cash consideration screen", INFO,
                      "Only triggers if a non-cash allotment is separately disclosed in the Notes -- "
                      "not testable from BS/PL face figures alone.", {}))

    # ---- R12: Reserves & Surplus roll-forward ----
    rs = d.get("reserves_and_surplus"); npat = d.get("net_profit_after_tax")
    if rs and npat:
        rsp, rsc = _fy_pair(rs)
        _, npat_c = _fy_pair(npat)
        if rsp is not None and rsc is not None and npat_c is not None:
            expected = rsp + npat_c
            residual = rsc - expected
            sev = PASS if abs(residual) < TOL_RS else REVW
            F.append(Finding("R12", "Reserves & Surplus roll-forward", sev,
                              f"Opening {rsp:,.2f} + Net Profit {npat_c:,.2f} = {expected:,.2f} vs closing "
                              f"{rsc:,.2f}, residual {residual:,.2f}. "
                              + ("Rolls forward exactly -- no dividend/other adjustment hiding in the movement."
                                 if sev == PASS else "Residual unexplained -- look for an undisclosed adjustment."),
                              {"residual": residual}))
        else:
            F.append(Finding("R12", "Reserves & Surplus roll-forward", INFO, "Figures incomplete.", {}))
    else:
        F.append(Finding("R12", "Reserves & Surplus roll-forward", INFO,
                          "not tested -- reserves_and_surplus and/or net_profit_after_tax not supplied.", {}))

    return F


# ======================================================================
# Cancelled E-Invoices (new) + defensive cross-checks (backlog B9 / D2)
# ======================================================================
def enrich_doc_gap_with_b2cs(doc_gap_list, b2cs_taxable, b2cs_tax_total):
    """Cross-reference amendments.doc_series_gap_check()'s remaining unexplained ranges (after
    declared-cancellation and cancelled-e-invoice enrichment have already run -- call this
    LAST) against this month's B2CS (Table 7, small unregistered B2C sales) aggregate.

    B2CS invoices structurally carry NO invoice number anywhere in a GSTR-1 export -- Table 7
    is a pure state+rate aggregate, by GST design (confirmed against the real sheet: its only
    columns are Place Of Supply/Rate/Taxable Value/tax heads, no invoice-number field at all).
    So a genuinely B2C-only document series can never be resolved by exact invoice-number
    matching the way the two enrichments above resolve things -- the only rigorous tie-out
    available here is at the WHOLE-MONTH level, and only under an unambiguous condition:
    there is non-zero B2CS value THIS month, AND exactly ONE range is still unexplained (so
    there is no ambiguity about which range that value could belong to -- if two or more ranges
    were simultaneously unexplained, this deliberately does NOT fire, per the rule that a
    partial/ambiguous explanation must never silently absorb a flag it can't uniquely account
    for).

    This is deliberately NOT a blanket "series has only one invoice -> skip gap detection" rule
    -- that would just as easily hide a genuinely missing single invoice in an ordinary numbered
    series. It only fires when there is actual B2CS rupee value to point to for this specific
    month, cited in the output.

    Mutates each dg dict in place: adds 'explained_by_b2cs' (bool) to every range, and -- only
    on the one range this fires for -- records the B2CS figures it ties to and clears
    'still_unexplained' to [] (the range's own 'missing' list is left untouched, so Table 13's
    raw declared arithmetic stays fully visible)."""
    unexplained_ranges = [dg for dg in doc_gap_list if dg.get("still_unexplained")]
    for dg in doc_gap_list:
        dg["explained_by_b2cs"] = False
    if len(unexplained_ranges) != 1 or not b2cs_taxable:
        return doc_gap_list
    dg = unexplained_ranges[0]
    dg["explained_by_b2cs"] = True
    dg["b2cs_taxable"] = b2cs_taxable
    dg["b2cs_tax"] = b2cs_tax_total
    dg["still_unexplained"] = []
    return doc_gap_list


def enrich_doc_gap_with_cancelled_einvoices(doc_gap_list, cancelled_invnos_this_month):
    """Cross-reference amendments.doc_series_gap_check()'s 'missing' invoice
    serials against this month's cancelled-e-invoice list (by exact invoice
    number). A serial that Table 13 declares as part of a range, that is
    genuinely absent from GSTR-1's own B2B sheet, IS explained if it turns
    out to be a cancelled e-invoice -- the invoice was legitimately never
    filed. This is a DIFFERENT explanation source from Table 13's own
    declared 'Cancelled' count (which the existing check already handles);
    this one draws on the E-Invoice file's IRN-status column instead, so it
    can explain a gap even when Table 13's own count doesn't line up
    (e.g. because Table 13 groups multiple different reasons under one
    'Cancelled' figure that doesn't 1:1 match a specific serial).
    Mutates each dg dict in place, adding:
      'explained_by_cancelled_einvoice' : sorted list of serials explained this way
      'still_unexplained' : sorted list of 'missing' serials neither declared-cancellation
                             nor cancelled-e-invoice explains -- this is the TRUE residual."""
    cancelled_set = set(cancelled_invnos_this_month or [])
    for dg in doc_gap_list:
        missing = dg.get("missing", [])
        if not missing:
            dg["explained_by_cancelled_einvoice"] = []
            dg["still_unexplained"] = []
            continue
        explained_here = sorted(m for m in missing if m in cancelled_set)
        dg["explained_by_cancelled_einvoice"] = explained_here
        remaining = [m for m in missing if m not in cancelled_set]
        # if Table 13's own declared-cancellation count already explains the (now smaller)
        # remainder exactly, that's still a legitimate full explanation -- re-check, don't
        # just trust the ORIGINAL count-based flag computed before this cross-reference.
        cancelled_declared = dg.get("table13_cancelled", 0) or 0
        if remaining and len(remaining) == cancelled_declared:
            dg["explained_by_declared_cancellation"] = True
            dg["still_unexplained"] = []
        else:
            dg["still_unexplained"] = sorted(remaining)
    return doc_gap_list


def build_cancelled_einvoice_findings(cancelled_by_month, g1_named_invoice_numbers_by_month, ewb_out_rows):
    """cancelled_by_month: {month: [ {invno,rate,taxable,igst,irn,cancel_date}, ... ]}
    -- built by the caller from gst_scrutiny_tool.parse_einv()'s new 'cancelled' key,
    one call per month, since that function is already scoped per-month.
    g1_named_invoice_numbers_by_month: {month: set(invoice_numbers)} from GSTR-1 B2B.
    ewb_out_rows: the full-FY outward EWB list (ewb_annual_parser output).
    Returns (list_of_all_cancelled_rows_for_the_sheet, list_of_Findings)."""
    all_cancelled = []
    for month, rows in (cancelled_by_month or {}).items():
        for r in rows:
            all_cancelled.append(dict(r, month=month))

    if not any(True for _ in cancelled_by_month or {}):
        return all_cancelled, [Finding("D2/B9", "Cancelled E-Invoices", INFO,
                                        "No E-Invoice data supplied -- cannot test.", {})]

    col_found_any = any(True for m in cancelled_by_month)  # caller only populates the dict if einv was available
    findings = []
    if not all_cancelled:
        findings.append(Finding("D2/B9", "Cancelled E-Invoices", PASS,
                                 "No cancelled e-invoices found across the months supplied (or the source "
                                 "export does not carry an IRN-status column -- see the Cancelled E-Invoices "
                                 "sheet's own 'column found' note for which case this is).", {}))
        return all_cancelled, findings

    # (a) cancelled invoice number still present in GSTR-1 B2B outward supply
    still_in_g1 = []
    for c in all_cancelled:
        g1set = (g1_named_invoice_numbers_by_month or {}).get(c["month"], set())
        if c["invno"] in g1set:
            still_in_g1.append(c)
    if still_in_g1:
        det = "; ".join(f"{c['invno']} ({c['month']}, cancelled {c.get('cancel_date') or 'date n/a'})"
                         for c in still_in_g1[:10])
        findings.append(Finding("D2a", f"Cancelled e-invoice STILL reported in GSTR-1 B2B: {len(still_in_g1)}",
                                 FLAG, f"These invoice number(s) were cancelled at the e-invoice/IRN level "
                                 f"but still appear as a live B2B outward supply in GSTR-1 for the same "
                                 f"month -- either the GSTR-1 entry should have been removed/not filed, or "
                                 f"the 'cancelled' status itself needs verifying. {det}.",
                                 {"count": len(still_in_g1)}))
    else:
        findings.append(Finding("D2a", "Cancelled e-invoice vs GSTR-1 B2B", PASS,
                                 "No cancelled e-invoice number still appears in GSTR-1 B2B outward supply.", {}))

    # (b) cancelled e-invoice with a live outward EWB against the same invoice number
    # EXTENDED (per explicit request: "for a cancelled e-invoice that was EWB-eligible, is its
    # EWB active or cancelled?"): checked the taxpayer's actual outward EWB export directly --
    # confirmed it carries NO status/cancellation column at all (fields present: EWB No.,
    # From/To GSTIN & Name, From/To Place, EWB No. & Dt., Doc No. & Dt., Assess Val., Tax Val.,
    # HSN Code/Desc., Latest Vehicle No. -- this matches a known, already-documented structural
    # limitation of this data source, not something this run's file happens to be missing).
    # So the honest, complete answer this tool CAN give is: whether a matching EWB document
    # exists at all for a cancelled invoice (a real, checkable fact) -- but NOT that EWB's own
    # active/cancelled status, which the source simply does not carry. Every cancelled row is
    # now annotated with one of three explicit states, never a guess.
    ewb_by_docno = {}
    for e in (ewb_out_rows or []):
        if e.get("docno"):
            ewb_by_docno.setdefault(str(e["docno"]).strip(), []).append(e)
    for c in all_cancelled:
        matches = ewb_by_docno.get(str(c.get("invno", "")).strip(), [])
        if matches:
            ewbnos = ", ".join(str(m.get("ewbno", "")) for m in matches)
            c["ewb_status_note"] = (
                f"EWB FOUND (No. {ewbnos}) for this cancelled invoice -- its own active/cancelled "
                f"status CANNOT be determined (this taxpayer's outward EWB export carries no "
                f"status/cancellation column at all -- a structural gap in the source, not "
                f"something this tool can compute). Verify directly on the e-way bill portal.")
        elif ewb_out_rows:
            c["ewb_status_note"] = "No EWB found for this invoice number in the outward EWB export."
        else:
            c["ewb_status_note"] = "Cannot check -- no outward EWB data supplied for this run."

    ewb_docnos = set(ewb_by_docno)
    live_ewb_on_cancelled = [c for c in all_cancelled if c["invno"] in ewb_docnos]
    if live_ewb_on_cancelled:
        det = "; ".join(f"{c['invno']} ({c['month']})" for c in live_ewb_on_cancelled[:10])
        findings.append(Finding("D2b", f"E-Invoice cancelled but a matching outward EWB document exists: {len(live_ewb_on_cancelled)}",
                                 FLAG, f"These invoice(s) have a CANCELLED e-invoice but a matching outward "
                                 f"EWB DOCUMENT still exists against the same invoice number -- worth "
                                 f"verifying (this was the project's own backlog item B9/D2, previously "
                                 f"unbuilt). NOTE: this taxpayer's EWB export carries no status/cancellation "
                                 f"column, so whether that EWB is itself still active or was separately "
                                 f"cancelled cannot be determined from this data -- verify on the e-way bill "
                                 f"portal directly; this finding only confirms the DOCUMENT exists, not its "
                                 f"current status. {det}.",
                                 {"count": len(live_ewb_on_cancelled)}))
    else:
        sev = PASS if ewb_out_rows else INFO
        note = ("No cancelled e-invoice has a matching live outward EWB." if ewb_out_rows
                else "No EWB data supplied -- cannot test this cross-check.")
        findings.append(Finding("D2b", "Cancelled e-invoice vs live EWB", sev, note, {}))

    return all_cancelled, findings


# ============================================================================
# ==== SECTION: filing_compliance.py  (was a standalone module before consolidation)
# ============================================================================
"""
FILING COMPLIANCE  --  ARN date extraction, statutory due dates, late fee & interest
========================================================================================
Fixes a real gap found in the previous version of this tool: `gst_unified_scrutiny.py`
had an `_extract_arn_dates()` function, but it lived inside `gather()`, a function
explicitly marked "LEGACY / UNSUPPORTED for the merged-file model" that
`master_build.py` (the actual pipeline) never calls. So in every real run,
`GSTR1_FILING_DATE`/`GSTR3B_FILING_DATE` stayed None, and Analysis checks #8/#10
always fell through to INFO ("Set ... in CONFIG to enable"). Also, even when
supplied, check #10 only compared GSTR-1-filing-date vs GSTR-3B-filing-date GAP --
never against the statutory DUE DATE, and no late-fee/interest RUPEE AMOUNT was
computed anywhere in the codebase.

This module fixes both, PER MONTH (the merged-file model's whole point), and adds
the ledger-actual cross-check the forensic framework asked for (Part 1, A4).

--------------------------------------------------------------------------------
WHERE THE ARN DATE ACTUALLY LIVES (content-based, verified against the real
GSTR-9C export's own ARN/ARN-Date fields as a format cross-check):
  GSTR-1  (merged workbook): every sub-sheet's PERIOD-MARKER row already
      carries the month's own ARN as free text, e.g.
      "Financial Year: 2022-23 | Tax Period: January | ARN: AA0501230730120 | ..."
      merged_period_utils.MARKER_RE only captures FY + Tax Period today; this
      module extends that same marker text with an ARN-date capture. If the
      marker text does NOT also carry a date (some GSTN exports print the ARN
      but not its date on this row), this module falls back to the 'Read me'
      sheet, which -- per gst_unified_scrutiny.py's original comment -- carries
      a single 'ARN date' row. IMPORTANT: for a MERGED whole-FY file, 'Read me'
      is BY DEFINITION only useful as a SINGLE value, which cannot be correct
      for all 12 months at once -- so this module treats a Read-me-only date as
      a per-file fallback (used only for whichever single month, if any, that
      'Read me' actually documents) and prints an explicit WARNING rather than
      silently applying one date to all 12 months. This needs confirming
      against a real multi-month merged GSTR-1 file (not yet supplied) --
      flagged clearly in the output, not guessed past.
  GSTR-3B (merged workbook, one SHEET per month): each month's own sheet
      already carries a 'Date of ARN' row (confirmed in gst_unified_scrutiny.py
      and reused here) -- this one is genuinely per-month already, no fallback
      needed.

HARD RULE: no invented dates, no invented amounts. Every field is either read
from the file or left None with a note. Late fee / interest are computed only
when both a filing date and a due date are known; otherwise the finding says
so explicitly.
"""

import re
import datetime as _dt
import openpyxl
import gst_core as mpu


# ======================================================================
# Statutory due dates
# ======================================================================
# GSTR-1 (monthly filer): 11th of the month following the tax period.
# GSTR-1 (QRMP / quarterly filer): 13th of the month following the quarter.
# GSTR-3B (monthly, non-QRMP, turnover > Rs 5 Cr or opted monthly): 20th.
# GSTR-3B (QRMP): 22nd (Category-X states) or 24th (Category-Y states).
# This taxpayer's GSTR-3B is confirmed one-sheet-per-MONTH (not per-quarter),
# i.e. non-QRMP -- matches hsn_fraud_checks.py's own GSTR3B_DUE_DOM=20 constant.
# For genericity (any taxpayer), filer type is DETECTED from the data itself,
# never assumed: if a GSTR-1 period marker resolves to 3 fanned-out months
# from one marker (see merged_period_utils.QUARTER_TO_MONTHS), that taxpayer
# is QRMP; if GSTR-3B has one sheet per calendar month, that taxpayer reports
# monthly (QRMP taxpayers still file GSTR-3B quarterly, so 4 sheets/FY, not 12).
CATEGORY_X_STATES = {  # 22nd -- confirmed CBIC state grouping for QRMP due dates
    "01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12",
    "13", "14", "15", "16", "17", "18", "19", "20", "21", "22", "23", "24",
    "25", "26", "27", "37",
}  # Chhattisgarh through Maharashtra + Andhra Pradesh(new)/Ladakh -- Category X (West+South+some NE)
CATEGORY_Y_STATES = {"28", "29", "30", "31", "32", "33", "34", "35", "36", "38"}  # Category Y (East+North)

LATE_FEE_PER_DAY_NORMAL = 25.0    # Rs 25 CGST + Rs 25 SGST = Rs 50/day total, per Sec 47
LATE_FEE_PER_DAY_NIL = 10.0       # Rs 10 CGST + Rs 10 SGST = Rs 20/day total, nil return
INTEREST_RATE_ANNUAL = 0.18       # Sec 50(1), 18% p.a. on the cash-paid portion of tax


def _cap_for_turnover(annual_turnover):
    """Late-fee caps per Notification 07/2023-CT (in force for FY22-23 returns
    filed after the notification date) -- max total late fee (both heads
    combined) PER RETURN, based on prior-year aggregate turnover:
      <=Rs 1.5 Cr turnover -> capped at Rs 2,000
      Rs 1.5-5 Cr turnover -> capped at Rs 5,000
      >Rs 5 Cr turnover    -> capped at Rs 10,000
    Nil returns: capped at Rs 500 regardless of turnover.
    Returns (cap_normal, cap_nil). If annual_turnover is None, returns
    (None, None) -- caller must then report the uncapped figure with an
    explicit 'cap not applied -- turnover unknown' note, never silently pick
    a slab."""
    if annual_turnover is None:
        return None, None
    # Amounts in RUPEES throughout this module (Indian digit-grouping used in
    # the literals below purely for readability: 1_50_00_000 = Rs 1.5 Cr).
    if annual_turnover <= 1_50_00_000:
        return 2000.0, 500.0
    if annual_turnover <= 5_00_00_000:
        return 5000.0, 500.0
    return 10000.0, 500.0


def due_date_gstr1(period_start_month_first_day, is_qrmp=False):
    """period_start_month_first_day: date(YYYY, MM, 1) for the tax period
    (or the LAST month of the quarter if QRMP -- caller passes the quarter's
    last calendar month). Returns the statutory due date."""
    y, m = period_start_month_first_day.year, period_start_month_first_day.month
    m2, y2 = (m % 12) + 1, y + (1 if m == 12 else 0)
    return _dt.date(y2, m2, 13 if is_qrmp else 11)


def due_date_gstr3b(period_start_month_first_day, is_qrmp=False, self_gstin=None):
    y, m = period_start_month_first_day.year, period_start_month_first_day.month
    m2, y2 = (m % 12) + 1, y + (1 if m == 12 else 0)
    if not is_qrmp:
        return _dt.date(y2, m2, 20)
    state_code = (self_gstin or "")[:2]
    dom = 24 if state_code in CATEGORY_Y_STATES else 22
    return _dt.date(y2, m2, dom)


def _month_label_to_first_day(label):
    """'Jan-23' -> date(2023,1,1). Uses merged_period_utils' own month-abbr
    map for consistency with the rest of the codebase."""
    m = re.match(r"^([A-Za-z]{3})-(\d{2})$", label)
    if not m:
        return None
    mon_abbr, yy = m.group(1), m.group(2)
    inv = {v: k for k, v in mpu.CAL_MONTH_ABBR.items()}
    mm = inv.get(mon_abbr.title())
    if not mm:
        return None
    yyyy = 2000 + int(yy)
    return _dt.date(yyyy, mm, 1)


# ======================================================================
# ARN date extraction
# ======================================================================
# Extend the marker regex (without touching merged_period_utils.py's own
# MARKER_RE, so nothing else that imports it changes behaviour) to also
# capture an ARN and, if present, a date right after it.
#
# FIXED (real bug, confirmed against the real merged GSTR-1 file's actual
# marker text): the label before the date can be "ARN Date" (ARN comes
# FIRST), not just "Date of Filing"/"Filing Date"/"Date" (Date comes
# first) -- the original pattern only handled the second family, so a
# real marker like "ARN: AA050422057237G | ARN Date: 10-05-2022" matched
# the ARN number but never the date, silently leaving every month's
# GSTR-1 filing date as None (which is why late-fee/filing-gap always
# showed blank even though the real marker text had the date all along).
_ARN_IN_MARKER_RE = re.compile(
    r"ARN:\s*([A-Z0-9]{15})\s*(?:[|,])?\s*"
    r"(?:(?:ARN\s*Date|Date\s*of\s*Filing|Filing\s*Date|Date)\s*[:\s]\s*)?"
    r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})?",
    re.IGNORECASE
)


def _parse_any_date(s):
    if not s:
        return None
    s = str(s).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d-%b-%Y", "%d-%b-%y", "%d/%m/%y"):
        try:
            return _dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def gstr1_arn_dates_by_month(gstr1_path):
    """Return {month_label: {'arn': str|None, 'date': date|None}} by reading
    EVERY period-marker row across every GSTR-1 sub-sheet (not just one),
    since a merged whole-FY file has one marker per month per sub-sheet.
    If NO marker anywhere carries a date (i.e. the export only prints "ARN:
    <no.>" without a date on the marker row itself), falls back to the
    single 'Read me' sheet value and returns it ONLY under whichever month
    that sheet's own 'Tax Period' field (if present) identifies -- otherwise
    returns it under a special '_readme_fallback' key with a clear note,
    rather than guessing which month it belongs to."""
    out = {}
    warnings = []
    wb = openpyxl.load_workbook(gstr1_path, data_only=True)
    found_any_marker_date = False
    for sn in wb.sheetnames:
        if sn.lower() == "read me":
            continue
        ws = wb[sn]
        for row in ws.iter_rows(values_only=True):
            if not row or not row[0]:
                continue
            cell0 = str(row[0])
            if "Financial Year:" not in cell0 or "Tax Period:" not in cell0:
                continue
            try:
                fy, tp, months = mpu.parse_marker_text(cell0)
            except mpu.PeriodParseError:
                continue
            m = _ARN_IN_MARKER_RE.search(cell0)
            arn = m.group(1) if m else None
            dt = _parse_any_date(m.group(2)) if (m and m.group(2)) else None
            if dt:
                found_any_marker_date = True
            for lbl in months:
                slot = out.setdefault(lbl, {"arn": None, "date": None})
                if arn and not slot["arn"]:
                    slot["arn"] = arn
                if dt and not slot["date"]:
                    slot["date"] = dt

    if not found_any_marker_date:
        # Fallback: 'Read me' sheet's single ARN-date row (per gst_unified_scrutiny.py).
        # This is a WHOLE-FILE value, not proven to be per-month -- surfaced with an
        # explicit warning rather than silently stamped onto every month.
        if "Read me" in wb.sheetnames:
            arn_val = date_val = tax_period_val = None
            for r in wb["Read me"].iter_rows(values_only=True):
                cells = [c for c in r if c not in (None, "")]
                if not cells:
                    continue
                label = str(cells[0]).strip().upper()
                if label in ("ARN", "ARN NO", "ARN NUMBER") and len(cells) >= 2:
                    arn_val = str(cells[-1]).strip()
                elif label in ("ARN DATE", "DATE OF ARN") and len(cells) >= 2:
                    date_val = _parse_any_date(cells[-1])
                elif label in ("TAX PERIOD",) and len(cells) >= 2:
                    tax_period_val = str(cells[-1]).strip()
            if date_val:
                warnings.append(
                    "GSTR-1 per-month ARN date not found on any period-marker row for this file -- "
                    "falling back to the single 'ARN date' value on the 'Read me' sheet "
                    f"({date_val}). This is a WHOLE-FILE value; it is only applied to a specific "
                    "month if 'Read me' also states a Tax Period, and is otherwise NOT applied to "
                    "any month automatically (see '_readme_fallback' in the result) -- confirm "
                    "against the portal before using it for a late-fee calculation on a specific month."
                )
                if tax_period_val:
                    try:
                        _, _, months = mpu.parse_marker_text(
                            f"Financial Year: 0000-00 | Tax Period: {tax_period_val}")
                    except Exception:
                        months = []
                    for lbl in months:
                        out.setdefault(lbl, {"arn": arn_val, "date": date_val})
                else:
                    out["_readme_fallback"] = {"arn": arn_val, "date": date_val}
    return out, warnings


def gstr3b_arn_dates_by_month(gstr3b_path):
    """Return {month_label: {'arn': str|None, 'date': date|None}}. GSTR-3B is
    one SHEET per month, and each sheet already carries its own 'Date of ARN'
    key/value row -- genuinely per-month, no fallback needed."""
    out = {}
    wb = openpyxl.load_workbook(gstr3b_path, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        fy = tp = arn = arn_date = None
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() for c in row if c not in (None, "")]
            if not cells:
                continue
            key = cells[0].upper()
            if key in ("YEAR", "FINANCIAL YEAR") and len(cells) >= 2:
                fy = cells[1]
            elif key == "TAX PERIOD" and len(cells) >= 2:
                tp = cells[1]
            elif key == "ARN" and len(cells) >= 2:
                arn = cells[-1]
            elif key in ("DATE OF ARN", "ARN DATE") and len(cells) >= 2:
                arn_date = _parse_any_date(cells[-1])
            if fy and tp and arn_date:
                break
        if not (fy and tp):
            continue
        try:
            labels = mpu.months_for_tax_period(fy, tp)
        except mpu.PeriodParseError:
            continue
        for lbl in labels:
            out[lbl] = {"arn": arn, "date": arn_date}
    return out


# ======================================================================
# Late fee / interest computation
# ======================================================================
def compute_late_fee(filing_date, due_date, is_nil_return=False, annual_turnover=None):
    """Section 47. Returns dict(days_late, fee_per_day_combined, gross_fee,
    cap, fee_payable, capped). fee_payable is the LOWER of the day-count
    formula and the applicable cap (Notification 07/2023-CT). If
    annual_turnover is None, cap is not applied and that is stated
    explicitly -- never silently uncapped without saying so."""
    if not filing_date or not due_date:
        return dict(days_late=None, fee_payable=None,
                     note="Filing date or due date not available -- cannot compute.")
    days_late = (filing_date - due_date).days
    if days_late <= 0:
        return dict(days_late=days_late, fee_payable=0.0, note="Filed on or before due date -- no late fee.")
    per_day = LATE_FEE_PER_DAY_NIL if is_nil_return else LATE_FEE_PER_DAY_NORMAL
    gross = days_late * per_day * 2  # both CGST+SGST heads combined
    cap_normal, cap_nil = _cap_for_turnover(annual_turnover)
    cap = cap_nil if is_nil_return else cap_normal
    if cap is None:
        return dict(days_late=days_late, gross_fee=gross, cap=None, fee_payable=gross, capped=False,
                     note="Annual turnover not supplied -- Notification 07/2023-CT cap NOT applied; "
                          "figure shown is the uncapped day-count formula only.")
    payable = min(gross, cap)
    return dict(days_late=days_late, gross_fee=gross, cap=cap, fee_payable=payable, capped=(payable < gross),
                note=f"{days_late} day(s) late x Rs {per_day*2:.0f}/day = Rs {gross:,.2f}, "
                     f"capped at Rs {cap:,.2f} per Notification 07/2023-CT" if payable < gross else
                     f"{days_late} day(s) late x Rs {per_day*2:.0f}/day = Rs {gross:,.2f} (within cap).")


def compute_interest(cash_paid_tax, filing_date, due_date):
    """Section 50(1): 18% p.a. simple interest on the CASH-portion of tax
    liability, for every day between due date and actual payment (approximated
    here as the filing date, since the cash ledger debit for a self-assessed
    liability happens at filing -- exact if paid via DRC-03 on the same date
    as filing, otherwise this is a lower bound; noted explicitly)."""
    if not filing_date or not due_date or cash_paid_tax is None:
        return dict(days_late=None, interest=None, note="Missing filing date, due date, or cash-tax figure.")
    days_late = (filing_date - due_date).days
    if days_late <= 0:
        return dict(days_late=days_late, interest=0.0, note="Filed on or before due date -- no interest.")
    interest = cash_paid_tax * INTEREST_RATE_ANNUAL * days_late / 365.0
    return dict(days_late=days_late, interest=round(interest, 2),
                note=f"Rs {cash_paid_tax:,.2f} (cash-paid tax) x 18% p.a. x {days_late} days / 365 "
                     f"= Rs {interest:,.2f}. Approximated using the FILING date as the payment date "
                     "(exact only if cash was actually debited same-day as filing) -- verify against "
                     "the Liability Register's own Interest-head entry for this period, per Forensic "
                     "Framework Part 1 A4.")


def month_filing_compliance(month_label, gstr1_arn_by_month, gstr3b_arn_by_month,
                             gstr1_is_qrmp=False, gstr3b_is_qrmp=False, self_gstin=None,
                             is_nil_return=False, annual_turnover=None, cash_paid_tax=None):
    """Full per-month compliance record: dates, due dates, late fee, interest,
    for BOTH GSTR-1 and GSTR-3B. Every field degrades to None + a note if the
    underlying date isn't available -- never fabricated."""
    first_day = _month_label_to_first_day(month_label)
    out = dict(month=month_label)
    for ret_type, arn_map, is_qrmp, due_fn in [
        ("gstr1", gstr1_arn_by_month, gstr1_is_qrmp, due_date_gstr1),
        ("gstr3b", gstr3b_arn_by_month, gstr3b_is_qrmp, due_date_gstr3b),
    ]:
        rec = arn_map.get(month_label, {})
        filing_date = rec.get("date")
        due = due_fn(first_day, is_qrmp) if first_day else None
        if ret_type == "gstr3b":
            due = due_date_gstr3b(first_day, is_qrmp, self_gstin) if first_day else None
        late = compute_late_fee(filing_date, due, is_nil_return, annual_turnover)
        out[f"{ret_type}_arn"] = rec.get("arn")
        out[f"{ret_type}_filing_date"] = filing_date
        out[f"{ret_type}_due_date"] = due
        out[f"{ret_type}_late_fee"] = late
        if ret_type == "gstr3b":
            out["gstr3b_interest"] = compute_interest(cash_paid_tax, filing_date, due)
    if out.get("gstr1_filing_date") and out.get("gstr3b_filing_date"):
        out["gstr1_vs_gstr3b_gap_days"] = (out["gstr3b_filing_date"] - out["gstr1_filing_date"]).days
    else:
        out["gstr1_vs_gstr3b_gap_days"] = None
    return out


