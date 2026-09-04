#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GST CORE
========
CONSOLIDATED FILE -- contains what used to be: merged_period_utils.py, folder_classifier.py

The tool was reorganised from 19 .py files into 9 for easier sharing. Nothing
in the analytical logic was rewritten during that move: each section below is
the original module's code verbatim, with only (a) intra-project imports
repointed at the new file names, (b) its standalone __main__ demo block
removed, and (c) the renames listed under MERGE NOTES applied where two merged
modules happened to define the same top-level name with different bodies.

MERGE NOTES for this file:
  (none -- no name collisions)
"""


# ============================================================================
# ==== SECTION: merged_period_utils.py  (was a standalone module before consolidation)
# ============================================================================
"""
MERGED PERIOD UTILS
====================
Shared helpers for the new MERGED input files (one workbook per document
type covering many months, instead of one file per month).

Every GSTR-1 / E-Invoice / GSTR-2B sub-sheet in a merged workbook carries a
period-MARKER row before each month's (or quarter's) block of data, e.g.:

    "Financial Year: 2022-23  |  Tax Period: January  |  ARN: ..."          (GSTR-1)
    "Financial Year: 2022-23  |  Tax Period: 042022  |  Date Updated ..."   (E-Invoice, numeric MMYYYY)
    "Financial Year: 2022-23  |  Tax Period: Apr-Jun  |  Date of Gen ..."   (GSTR-2B, quarterly)

This module finds those marker rows by CONTENT (never by position/sheet name)
and slices the sheet into {month_label: [data_rows]} blocks. GSTR-3B is
different -- it merges as one SHEET PER MONTH rather than marker rows inside
one sheet -- so GSTR-3B sheets are identified by their own in-sheet
'Year'/'Tax Period' key-value rows (content-based, per user instruction:
ignore the sheet's NAME entirely, e.g. 'Jan_2022-23' is not to be trusted).

HARD RULE (per explicit instruction): no safety nets. If a marker can't be
parsed, or a requested month isn't present, this raises -- it does not
silently return zero/empty/a guess.
"""

import re

MONTH_NAME_TO_NUM = {
    "JANUARY": 1, "FEBRUARY": 2, "MARCH": 3, "APRIL": 4, "MAY": 5, "JUNE": 6,
    "JULY": 7, "AUGUST": 8, "SEPTEMBER": 9, "OCTOBER": 10, "NOVEMBER": 11, "DECEMBER": 12,
    "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "JUN": 6, "JUL": 7, "AUG": 8,
    "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12,
}
MONTH_NUM_TO_ABBR = {1: "Apr", 2: "May", 3: "Jun", 4: "Jul", 5: "Aug", 6: "Sep",
                     7: "Oct", 8: "Nov", 9: "Dec", 10: "Jan", 11: "Feb", 12: "Mar"}
# ^ deliberately NOT used -- keep a plain calendar map instead (clearer, no FY-offset tricks here)
CAL_MONTH_ABBR = {1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
                   7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"}

QUARTER_TO_MONTHS = {
    "JAN-MAR": [1, 2, 3], "APR-JUN": [4, 5, 6], "JUL-SEP": [7, 8, 9], "OCT-DEC": [10, 11, 12],
}

MARKER_RE = re.compile(
    r"Financial Year:\s*([0-9]{4}\s*-\s*[0-9]{2,4})\s*\|\s*Tax Period:\s*([^\|]+?)\s*(?:\||$)"
)


class PeriodParseError(ValueError):
    pass


def fy_years(fy):
    """'2022-23' -> (2022, 2023). Raises PeriodParseError if malformed."""
    m = re.match(r"^\s*(\d{4})\s*-\s*(\d{2,4})\s*$", fy)
    if not m:
        raise PeriodParseError(f"Unrecognised Financial Year format: {fy!r}")
    y1 = int(m.group(1))
    y2s = m.group(2)
    y2 = int(y2s) if len(y2s) == 4 else int(str(y1)[:2] + y2s)
    return y1, y2


def months_for_tax_period(fy, tax_period):
    """Return list of 'Mon-YY' calendar-month labels this marker covers.
    1 label for a month marker, 3 for a quarter marker. Raises PeriodParseError
    if the Tax Period text isn't in any recognised format (month name, numeric
    MMYYYY, or a quarter like 'Apr-Jun') -- stays flexible across formats
    since real files may use any of the three, but does not guess beyond them."""
    y1, y2 = fy_years(fy)
    tp = tax_period.strip().upper()

    if re.match(r"^\d{6}$", tp):                 # numeric MMYYYY, e.g. '042022'
        month_nums = [int(tp[:2])]
        if not (1 <= month_nums[0] <= 12):
            raise PeriodParseError(f"Unrecognised numeric Tax Period: {tax_period!r}")
    elif tp in MONTH_NAME_TO_NUM:                 # 'January' / 'Jan' style
        month_nums = [MONTH_NAME_TO_NUM[tp]]
    elif tp in QUARTER_TO_MONTHS:                 # 'Apr-Jun' style
        month_nums = QUARTER_TO_MONTHS[tp]
    else:
        raise PeriodParseError(f"Unrecognised Tax Period format: {tax_period!r} (FY {fy})")

    labels = []
    for mm in month_nums:
        cal_year = y2 if mm <= 3 else y1
        labels.append(f"{CAL_MONTH_ABBR[mm]}-{str(cal_year)[2:]}")
    return labels


def parse_marker_text(text):
    """Return (fy, tax_period_raw, [month_labels]) for one marker cell's text.
    Raises PeriodParseError if the text is not a period marker at all."""
    m = MARKER_RE.search(text or "")
    if not m:
        raise PeriodParseError(f"Not a period-marker cell: {text!r}")
    fy, tp = m.group(1).strip(), m.group(2).strip()
    return fy, tp, months_for_tax_period(fy, tp)


def is_marker_row(row):
    """A marker row carries its text in cell 0 only; every other cell is empty."""
    if not row or not row[0]:
        return False
    return bool(MARKER_RE.search(str(row[0])))


def split_rows_by_month(data_rows):
    """data_rows: rows AFTER the header row (may start with a marker).
    Returns {month_label: [row, row, ...]}, excluding marker rows themselves
    and genuinely blank spacer rows. A quarter marker fans its rows out into
    all 3 of that quarter's month buckets (used for GSTR-2B's quarter-level
    summary sheet; invoice-level 2B sheets should instead use each row's own
    per-line period column -- see gstr2b_parser.py).
    A month whose block has ZERO data rows (its marker is immediately
    followed by the next marker) still gets registered with an empty list --
    that is a legitimate 'this month had nothing on this sub-sheet' state,
    not a missing month.
    Raises PeriodParseError if data is found before any marker has been seen."""
    blocks = {}
    current_labels = None
    for row in data_rows:
        if is_marker_row(row):
            _, _, current_labels = parse_marker_text(str(row[0]))
            for lbl in current_labels:
                blocks.setdefault(lbl, [])
            continue
        if not any(c not in (None, "") for c in row):
            continue
        if current_labels is None:
            raise PeriodParseError(
                "Data row encountered before any period marker in this sheet -- "
                "cannot determine which month it belongs to: " + repr(row)
            )
        for lbl in current_labels:
            blocks.setdefault(lbl, []).append(row)
    return blocks


def rows_for_month(all_rows, header_row_idx, month_label):
    """Convenience wrapper: split rows[header_row_idx+1:] by month, return only
    the requested month's rows. Raises PeriodParseError if that month has no
    block in this sheet at all (distinct from the month having zero DATA rows,
    which is legitimate and returns [])."""
    blocks = split_rows_by_month(all_rows[header_row_idx + 1:])
    if month_label not in blocks:
        raise PeriodParseError(
            f"Month {month_label!r} not found as a period marker in this sheet. "
            f"Months present: {sorted(blocks)}"
        )
    return blocks[month_label]


def find_block_for_month(all_rows, month_label):
    """For sheets where period markers sit directly among the data (no single
    fixed header row to split from -- e.g. GSTR-2B's 'ITC Available' summary,
    which has several small tables per quarter block), return (start, end)
    row-index bounds (start is the row right after the marker; end is the
    next marker's row, or len(all_rows)) for the block covering `month_label`.
    Raises PeriodParseError if no marker in the sheet covers that month."""
    marker_positions = []  # (row_idx, [month_labels])
    for i, row in enumerate(all_rows):
        if is_marker_row(row):
            _, _, labels = parse_marker_text(str(row[0]))
            marker_positions.append((i, labels))
    for idx, (row_idx, labels) in enumerate(marker_positions):
        if month_label in labels:
            start = row_idx + 1
            end = marker_positions[idx + 1][0] if idx + 1 < len(marker_positions) else len(all_rows)
            return start, end
    raise PeriodParseError(
        f"Month {month_label!r} not covered by any period marker in this sheet. "
        f"Markers found: {[lbl for _, lbl in marker_positions]}"
    )


def find_block_and_index_for_month(all_rows, month_label):
    """Like find_block_for_month, but ALSO returns WHICH position (0-based)
    month_label occupies within its marker's own label list, and how many
    months that marker covers in total (1 for a month marker, 3 for a
    quarter marker).

    BUG FIX (bug report §1-5): GSTR-2B's quarterly 'ITC Available' summary
    sheet lays three months' figures side by side on the SAME row inside one
    quarter block (month-1 columns, month-2 columns, month-3 columns, then a
    quarter-total group) -- find_block_for_month() alone can only say WHICH
    rows belong to the quarter, not which of the 3 side-by-side column-groups
    within those rows belongs to the specific month being asked for. A reader
    that didn't know this always took the first (month-1) group, so months 2
    and 3 of every quarter silently got month-1's figures. This function
    gives the caller the position needed to pick the RIGHT group instead."""
    start, end = find_block_for_month(all_rows, month_label)
    # Recover the marker's own label list the same way find_block_for_month did,
    # by re-reading the marker row immediately before `start`.
    marker_row = all_rows[start - 1]
    _, _, labels = parse_marker_text(str(marker_row[0]))
    return start, end, labels.index(month_label), len(labels)


def months_present(all_rows, header_row_idx):
    """Return sorted set of every month label found via markers in this sheet
    (including months whose block turned out to have zero data rows)."""
    months = set()
    current_labels = None
    for row in all_rows[header_row_idx + 1:]:
        if is_marker_row(row):
            _, _, current_labels = parse_marker_text(str(row[0]))
            months.update(current_labels)
    return months


# ============================================================================
# ==== SECTION: folder_classifier.py  (was a standalone module before consolidation)
# ============================================================================
"""
FOLDER CLASSIFIER  (merged-file model)
========================================
Scans a folder and identifies ONE merged workbook per document type (whole
FY, however many months it currently covers), plus the whole-FY annual
sources -- all by CONTENT signature, never by filename.

Returns:
  gstr1_merged, gstr3b_merged, einv_merged, gstr2b_merged : paths (or None)
  gstr1_months, gstr3b_months, einv_months, gstr2b_months : sets of 'Mon-YY'
      labels actually found inside each merged file (for coverage reporting --
      discovering a month here does NOT mean every sub-sheet has data for it,
      just that the file's period markers include it)
  ewb_out_annual, ewb_in_annual : whole-FY EWB workbook paths
  cash_ledger, credit_ledger, liab_ledger (Register, Part I) : CSV paths
  liab_demand_ledger (Ledger, Part II -- DRC/demand) : CSV path
  portal_comparison : xlsx path
  bo_profile, gstr9, gstr9c : xlsx paths (all three previously PDF -- now Excel)
  self_gstin, company_name
"""

import os
import re
import glob
import csv
import openpyxl
import sys as _sys_selfalias
mpu = _sys_selfalias.modules[__name__]


def _sheetnames(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sn = list(wb.sheetnames)
    wb.close()
    return sn


def _looks_like_r2a_merged(path):
    """Content signature for the merged GSTR-2A workbook: its own sheets carry
    the literal 'GSTR 2A' / 'GSTR-2A' / 'GSTR2A' banner text (spacing/hyphen
    varies across the sheets in a real export -- confirmed 'GSTR 2A' on B2B,
    'GSTR-2A' on B2BA, 'GSTR2A' on CDNRA). Checked BEFORE the Table 8A test
    below: a real GSTR-2A export's sheet set -- {Read me, B2B, B2BA, CDNR,
    CDNRA, ECO, ECOA, ISD, ISDA, TDS, TDSA, TCS, IMPG, IMPG SEZ} -- is a
    strict SUPERSET of Table 8A's {B2B, B2BA, CDNR, CDNRA} + 'Read me', so
    the pre-existing Table-8A sheet-set check alone would misclassify every
    GSTR-2A file as Table 8A (confirmed: this is not a hypothetical -- a
    real GSTR-2A export was fed through classify_folder() unmodified and
    landed in table8a_files). Banner text, not sheet names, is the safer
    signal because it doesn't depend on which optional 2A sheets happen to
    be present in a given export."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for sn in ("B2B", "B2BA", "CDNR", "CDNRA"):
            if sn not in wb.sheetnames:
                continue
            for row in wb[sn].iter_rows(min_row=1, max_row=4, values_only=True):
                for c in row:
                    if not c:
                        continue
                    cu = str(c).upper()
                    if "GSTR 2A" in cu or "GSTR-2A" in cu or "GSTR2A" in cu:
                        return True
    finally:
        wb.close()
    return False


def _looks_like_gstr3b_merged(path):
    """Content signature for the merged GSTR-3B workbook: at least one sheet
    contains the literal 'Form GSTR-3B' banner text. Sheet NAMES (e.g.
    'Jan_2022-23') are never consulted, per instruction."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for sn in wb.sheetnames:
            for row in wb[sn].iter_rows(min_row=1, max_row=3, values_only=True):
                if any(c and "Form GSTR-3B" in str(c) for c in row):
                    return True
    finally:
        wb.close()
    return False


def _gstr1_months(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["b2b, sez, de_inv"]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    return mpu.months_present(rows, 3)


def _einv_months(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["b2b, sez, de"]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    return mpu.months_present(rows, 3)


def _gstr2b_months(path):
    """Which months a GSTR-2B file covers, from its period markers.

    BUG FIX (same root cause as parse_2b_excel's own fix, found on the same real taxpayer's
    file): this unconditionally read 'ITC Available' for its markers, so a file lacking that
    sheet (confirmed real: some exports carry only 'B2B'+'B2B-CDNR') would raise here too --
    caught by _build_month_file_map's own try/except, but the practical effect was this file
    silently contributing ZERO months to gstr2b_month_map even after classify_folder correctly
    recognised it as a GSTR-2B file. 'B2B' carries the identical period-marker format ('Financial
    Year: ... | Tax Period: ...' rows) -- confirmed directly against this taxpayer's real file --
    so it's a safe, equivalent fallback source for month coverage when 'ITC Available' is absent."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_name = "ITC Available" if "ITC Available" in wb.sheetnames else "B2B"
        if sheet_name not in wb.sheetnames:
            return set()
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    return mpu.months_present(rows, 0)


def _gstr3b_months(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    months = set()
    try:
        for sn in wb.sheetnames:
            ws = wb[sn]
            fy = tp = None
            for row in ws.iter_rows(values_only=True):
                cells = [str(c).strip() for c in row if c not in (None, "")]
                if not cells:
                    continue
                key = cells[0].upper()
                if key in ("YEAR", "FINANCIAL YEAR") and len(cells) >= 2:
                    fy = cells[1]
                elif key == "TAX PERIOD" and len(cells) >= 2:
                    tp = cells[1]
                if fy and tp:
                    break
            if fy and tp:
                try:
                    months.update(mpu.months_for_tax_period(fy, tp))
                except mpu.PeriodParseError:
                    pass  # a stray non-data sheet with unrelated Year/Tax Period-looking text
    finally:
        wb.close()
    return months


def _csv_first_line(path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        for row in csv.reader(f):
            line = " ".join(str(c) for c in row if c).strip()
            if line:
                return line
    return ""


def _read_me_gstin_and_name(gstr1_path):
    """Content-based GSTIN + Legal Name from the merged GSTR-1's 'Read me' sheet."""
    wb = openpyxl.load_workbook(gstr1_path, read_only=True, data_only=True)
    try:
        sn = "Read me" if "Read me" in wb.sheetnames else wb.sheetnames[0]
        gstin = name = None
        for row in wb[sn].iter_rows(values_only=True):
            cells = [str(c).strip() for c in row if c not in (None, "")]
            if not cells:
                continue
            key = cells[0].upper()
            if key == "GSTIN" and len(cells) >= 2:
                gstin = cells[1]
            elif key == "LEGAL NAME" and len(cells) >= 2:
                name = cells[1]
        return gstin, name
    finally:
        wb.close()


def _build_month_file_map(files, months_fn, label):
    """files: list of candidate paths for ONE document type (e.g. every merged
    GSTR-1 workbook found in the folder, however many FYs that spans).
    months_fn: the _gstr1_months/_gstr3b_months/etc function for that type.
    Returns ({month_label: filepath}, warnings). If two files both claim the
    SAME month, that is a genuine ambiguity (e.g. two overlapping exports)
    and is NOT silently resolved by picking one -- it's reported as a
    warning and the LATER file (by mtime) wins, so at least the behaviour is
    deterministic and visible, not silently arbitrary."""
    month_map = {}
    warnings = []
    # newest-mtime-last, so a later duplicate legitimately overrides an
    # earlier one (e.g. a corrected re-export) rather than the reverse
    for f in sorted(files, key=lambda p: os.path.getmtime(p)):
        try:
            months = months_fn(f)
        except Exception as ex:
            warnings.append(f"{label} file {f!r} could not be read for month coverage: {ex}")
            continue
        for m in months:
            if m in month_map and month_map[m] != f:
                warnings.append(f"{label}: month {m!r} found in BOTH {month_map[m]!r} and {f!r} -- "
                                 f"using {f!r} (newer file). Verify these aren't two different FYs "
                                 f"that happen to reuse the same 'Mon-YY' label by mistake.")
            month_map[m] = f
    return month_map, warnings


def _looks_like_gstr9_excel(sn):
    """Content signature for the GSTR-9 (Annual Return) Excel export. Sheet names in
    this export are truncated to Excel's 31-char sheet-name limit, so match on the
    stable PREFIX of two sheets that only ever appear in GSTR-9 (never GSTR-9C):
    'Item 4 - Advances & Outward+Inw...' and 'Items 17-18 - HSN Wise Summary...'.
    Verified against a real GSTR9_<GSTIN>_<period>.xlsx export."""
    return (any(s.startswith("Item 4 - Advances") for s in sn)
            and any(s.startswith("Items 17-18") for s in sn))


def _looks_like_gstr9c_excel(sn):
    """Content signature for the GSTR-9C (Reconciliation Statement) Excel export.
    'Item 5 - Reconciliation of Gros...' and 'Item 12 - Reconciliation of Net...'
    only ever appear in GSTR-9C (GSTR-9's own Item 5 sheet is titled 'Outward
    Supplies...' instead). Verified against a real GSTR-9C_<GSTIN>_<period>.xlsx export."""
    return (any(s.startswith("Item 5 - Reconciliation") for s in sn)
            and any(s.startswith("Item 12 - Reconciliation") for s in sn))


def _looks_like_bo_profile_excel(sn):
    """Content signature for the BO / 360-degree Profile Excel export -- these three
    sheet names together are distinctive and stable (verified against a real
    <GSTIN>_BO_Profile_<date>.xlsx export)."""
    return {"Demographic Details", "Financial Information", "BIFA Specific Information"}.issubset(sn)


def _has_cdnr_pair(sn, suffix=""):
    """Table 8A / GSTR-2A exports have been seen using EITHER 'CDNR'/'CDNRA' or the
    newer 'B2B-CDNR'/'B2B-CDNRA' sheet names for the same content (confirmed: same
    column layout under either name, verified against a real R9_8A export using the
    'B2B-CDNR' variant) -- accept either alias rather than assuming one."""
    name = f"CDNR{suffix}"
    alt = f"B2B-CDNR{suffix}"
    return name in sn or alt in sn


def _first_row_lower(path, sheet_name):
    """First row of one sheet, as lowercased/stripped strings -- used for
    single-sheet content signatures (e.g. the blocked-ITC keyword master)
    where the sheet NAME isn't distinctive enough to key off on its own."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        return [str(c).strip().lower() if c else "" for c in row]
    finally:
        wb.close()


def classify_folder(folder="."):
    xlsx = sorted(glob.glob(os.path.join(folder, "*.xlsx")) + glob.glob(os.path.join(folder, "*.xlsm")))
    csvs = sorted(glob.glob(os.path.join(folder, "*.csv")))

    # CHANGED (multi-year support): every doc type now collects a LIST of
    # matching files, not a single overwritten path -- a taxpayer with 5
    # years of data may supply 5 separate merged GSTR-1 workbooks (one per
    # FY), each internally covering up to 12 months via its own period
    # markers. Previously this loop did `gstr1_merged = f` on every match,
    # so only the LAST file (by glob/sort order) survived and every earlier
    # FY's data was silently discarded without any error or warning.
    gstr1_files, gstr3b_files, einv_files, gstr2b_files = [], [], [], []
    portal_comparison_files, ewb_candidates = [], []
    bo_profile_files = []
    gstr9_files, gstr9c_files, table8a_files, bs_pl_files = [], [], [], []
    hsn_sac_master_files = []
    r2a_files = []
    blocked_itc_master_files = []
    machinery_hsn_master_files = []

    for f in xlsx:
        sn = set(_sheetnames(f))
        if not sn:
            continue
        if "b2b, sez, de_inv" in sn and "hsn" in sn:
            gstr1_files.append(f); continue
        if "ITC Available" in sn and "B2B" in sn:
            gstr2b_files.append(f); continue
        if "b2b, sez, de" in sn and "b2b, sez, de_inv" not in sn:
            einv_files.append(f); continue
        if "Comparison Summary" in sn:
            portal_comparison_files.append(f); continue
        # GSTR-9 / GSTR-9C / BO Profile: now supplied as Excel exports (previously PDF --
        # the PDF-classification path for these three was removed; nothing else in this
        # tool parses a PDF, so there is no PDF fallback for these three doc types).
        if _looks_like_gstr9c_excel(sn):
            gstr9c_files.append(f); continue
        if _looks_like_gstr9_excel(sn):
            gstr9_files.append(f); continue
        if _looks_like_bo_profile_excel(sn):
            bo_profile_files.append(f); continue
        # GSTR-2A: checked BEFORE Table 8A below -- see _looks_like_r2a_merged()'s docstring for
        # why the Table-8A sheet-set test alone would otherwise swallow every GSTR-2A file.
        # CDNR/CDNRA accepted under either the classic or the 'B2B-CDNR(A)' sheet-naming variant.
        if ({"B2B", "B2BA"}.issubset(sn) and _has_cdnr_pair(sn) and _has_cdnr_pair(sn, "A")
                and "Read me" in sn and _looks_like_r2a_merged(f)):
            r2a_files.append(f); continue
        # Table 8A: government-standard export, always has this sheet set (CDNR/CDNRA
        # OR the newer B2B-CDNR/B2B-CDNRA naming -- same column layout under either name)
        if ({"B2B", "B2BA"}.issubset(sn) and _has_cdnr_pair(sn) and _has_cdnr_pair(sn, "A")
                and "Read me" in sn):
            table8a_files.append(f); continue
        # GSTR-2B, minimal shape (confirmed real, not hypothetical: one real taxpayer's export
        # carried ONLY 'B2B' + 'B2B-CDNR' -- no 'ITC Available', no 'Read me', no B2BA/CDNRA at
        # all -- and was silently never classified as GSTR-2B at all, so it never even reached
        # the parser). Deliberately placed AFTER the GSTR-2A/Table-8A checks above (both of
        # which require B2BA to ALSO be present) so a genuine 4+-sheet 2A/Table-8A-shaped file
        # is never mis-caught here: "B2BA not in sn" alone is sufficient to guarantee this isn't
        # one of those, since both require B2BA unconditionally. gst_parsers_returns.
        # parse_2b_excel already handles 'ITC Available' being absent gracefully (summary
        # marked not-available, invoice-level B2B/B2B-CDNR parsing unaffected).
        if "B2B" in sn and "B2BA" not in sn:
            gstr2b_files.append(f); continue
        # HSN/SAC code-and-description master (e.g. the NIC e-Invoice system's own
        # downloadable HSN_SAC.xlsx) -- content signature: exactly these two sheet
        # names. NOTE: this master has CODE + DESCRIPTION columns only, no GST rate
        # column -- used for code-existence/description validation (see
        # hsn_fraud_checks.check_hsn_master_validity), NOT for rate comparison
        # (that stays HSN_RATE_HISTORY's job, a separate curated table).
        if {"HSN_MSTR", "SAC_MSTR"}.issubset(sn):
            hsn_sac_master_files.append(f); continue
        # Blocked-ITC keyword/HSN master (taxpayer-supplied, e.g. 'GSTR_2B_Blocked_ITC_
        # Master_Updated.xlsx') -- content signature: exactly one sheet, header row reads
        # Category / Search keyword / Indicative HSN/SAC. Never matched by filename.
        if len(sn) == 1:
            _hdr = _first_row_lower(f, next(iter(sn)))
            if _hdr[:3] == ["category", "search keyword", "indicative hsn/sac"]:
                blocked_itc_master_files.append(f); continue
        # Machinery HSN master (taxpayer-supplied, e.g. 'Machinery_HSN_Master_v2.xlsx') --
        # content signature: ANY sheet's header row reads S.No / HSN Heading (4-digit) /
        # Chapter / Description / Category / Flag as Machine Purchase?... / Match Rule...
        # (this workbook has multiple sheets -- per-chapter references, a Read Me -- so the
        # check has to scan each sheet's header, not assume sheet count like the one above).
        if any(_first_row_lower(f, s)[:7] == [
                "s.no", "hsn heading (4-digit)", "chapter", "description", "category",
                "flag as machine purchase? (y/n/review)",
                "match rule (use as hsn prefix match on gstr-2b)"] for s in sn):
            machinery_hsn_master_files.append(f); continue
        if _looks_like_gstr3b_merged(f):
            gstr3b_files.append(f); continue
        # Annual EWB: has 'EWB No.' + 'From GSTIN & Name' header on some sheet
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        found_ewb = False
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
                hdr = [str(c).strip() if c else "" for c in row]
                if "EWB No." in hdr and "From GSTIN & Name" in hdr and "To GSTIN & Name" in hdr:
                    ewb_candidates.append(f)
                    found_ewb = True
                    break
            if found_ewb:
                break
        wb.close()

    # NOTE: GSTR-9 / GSTR-9C / BO Profile are now classified above, from the xlsx content
    # signatures -- this tool no longer reads any PDF (BS/PL was never PDF-parsed either;
    # it's always the hand-typed bs_pl_input.py, GSTIN-tagged -- see that file). bs_pl_files
    # is kept as an always-empty list purely so any code that still reads that key
    # (defensively) sees a list, not a missing key.

    # ---- direction for the annual EWB files: self-GSTIN mostly in From -> outward ----
    # (unchanged logic; still works across multiple years' EWB files pooled together)
    self_gstin = None
    ewb_out_files, ewb_in_files = [], []
    if ewb_candidates:
        import gst_parsers_returns as ewbp
        from collections import Counter
        parsed = {f: ewbp.parse_annual_ewb(f) for f in ewb_candidates}
        freq = Counter()
        gstin_re = re.compile(r"^\d{2}[A-Z]{5}\d{4}[A-Z]\dZ[A-Z\d]$")
        for f, rows in parsed.items():
            for r in rows:
                for g in (r["from_gstin"], r["to_gstin"]):
                    if gstin_re.match(g or ""):
                        freq[g] += 1
        if freq:
            self_gstin = freq.most_common(1)[0][0]
        for f, rows in parsed.items():
            fr = sum(1 for r in rows if r["from_gstin"] == self_gstin)
            to = sum(1 for r in rows if r["to_gstin"] == self_gstin)
            (ewb_out_files if fr >= to else ewb_in_files).append(f)

    # Liability: the GST portal exports TWO distinct reports that both contain the word
    # "liability" -- confirmed genuinely different column layouts against real files, not
    # duplicates of each other:
    #   "Electronic Liability Register" (Part I, return-related: liability created by
    #     filing GSTR-1/3B, discharged via cash/credit -- 6-column preamble before the
    #     tax-head groups) -- classified as liab_register_files.
    #   "Electronic Liability Ledger" (Part II, non-return: DRC/demand/voluntary-payment
    #     liability, carries 'Relevant Demand ID / Liability ID' + 'Stay status' columns
    #     the Register doesn't have -- 8-column preamble) -- classified as liab_demand_files.
    # Distinguished by the report's own title cell (real file content, not a filename
    # guess) -- verified against a real pair of these files for the same GSTIN/FY.
    cash_ledgers, credit_ledgers, liab_register_files, liab_demand_files = [], [], [], []
    for c in csvs:
        line = _csv_first_line(c).lower()
        if "cash ledger" in line:
            cash_ledgers.append(c)
        elif "credit ledger" in line:
            credit_ledgers.append(c)
        elif "liability register" in line:
            liab_register_files.append(c)
        elif "liability ledger" in line:
            liab_demand_files.append(c)

    # self_gstin / company_name refinement from the FIRST merged GSTR-1's Read me sheet
    company_name = None
    if gstr1_files:
        g1_gstin, g1_name = _read_me_gstin_and_name(gstr1_files[0])
        self_gstin = self_gstin or g1_gstin
        company_name = g1_name

    gstr1_month_map, w1 = _build_month_file_map(gstr1_files, _gstr1_months, "GSTR-1")
    gstr3b_month_map, w2 = _build_month_file_map(gstr3b_files, _gstr3b_months, "GSTR-3B")
    einv_month_map, w3 = _build_month_file_map(einv_files, _einv_months, "E-Invoice")
    gstr2b_month_map, w4 = _build_month_file_map(gstr2b_files, _gstr2b_months, "GSTR-2B")
    warnings = w1 + w2 + w3 + w4

    return dict(
        # NEW multi-year keys (month-level file resolution -- use these)
        gstr1_month_map=gstr1_month_map, gstr3b_month_map=gstr3b_month_map,
        einv_month_map=einv_month_map, gstr2b_month_map=gstr2b_month_map,
        gstr1_files=gstr1_files, gstr3b_files=gstr3b_files,
        einv_files=einv_files, gstr2b_files=gstr2b_files,
        classify_warnings=warnings,
        # BACKWARD-COMPAT single-path keys (first file found; old single-FY code
        # that doesn't know about *_month_map still works unchanged)
        gstr1_merged=gstr1_files[0] if gstr1_files else None,
        gstr3b_merged=gstr3b_files[0] if gstr3b_files else None,
        einv_merged=einv_files[0] if einv_files else None,
        gstr2b_merged=gstr2b_files[0] if gstr2b_files else None,
        gstr1_months=set(gstr1_month_map), gstr3b_months=set(gstr3b_month_map),
        einv_months=set(einv_month_map), gstr2b_months=set(gstr2b_month_map),
        # EWB -- now lists (one set of annual workbooks per FY supplied)
        ewb_out_files=ewb_out_files, ewb_in_files=ewb_in_files,
        ewb_out_annual=ewb_out_files[0] if ewb_out_files else None,
        ewb_in_annual=ewb_in_files[0] if ewb_in_files else None,
        # Annual-level sources -- now lists (one set per FY)
        cash_ledgers=cash_ledgers, credit_ledgers=credit_ledgers,
        liab_register_files=liab_register_files, liab_demand_files=liab_demand_files,
        cash_ledger=cash_ledgers[0] if cash_ledgers else None,
        credit_ledger=credit_ledgers[0] if credit_ledgers else None,
        # BACKWARD-COMPAT: liab_ledger/liab_ledgers keep pointing at the RETURN register
        # (Part I) -- this is what every existing monthly-liability-vs-3B comparison in the
        # codebase already expects and is column-layout-compatible with.
        liab_ledger=liab_register_files[0] if liab_register_files else None,
        liab_ledgers=liab_register_files,
        liab_demand_ledger=liab_demand_files[0] if liab_demand_files else None,
        portal_comparison_files=portal_comparison_files,
        bo_profile_files=bo_profile_files,
        portal_comparison=portal_comparison_files[0] if portal_comparison_files else None,
        bo_profile=bo_profile_files[0] if bo_profile_files else None,
        # NEW optional annual-return-side documents
        gstr9_files=gstr9_files, gstr9c_files=gstr9c_files,
        table8a_files=table8a_files, bs_pl_files=bs_pl_files,
        # GSTR-2A -- one whole-FY snapshot file, same "annual/dept-side" treatment as
        # portal_comparison/bo_profile (a list plus a first-found singular convenience key).
        r2a_files=r2a_files, r2a_merged=(r2a_files[0] if r2a_files else None),
        hsn_sac_master_files=hsn_sac_master_files,
        hsn_sac_master_file=(sorted(hsn_sac_master_files, key=os.path.getmtime)[-1]
                              if hsn_sac_master_files else None),
        blocked_itc_master_files=blocked_itc_master_files,
        blocked_itc_master_file=(sorted(blocked_itc_master_files, key=os.path.getmtime)[-1]
                                  if blocked_itc_master_files else None),
        machinery_hsn_master_files=machinery_hsn_master_files,
        machinery_hsn_master_file=(sorted(machinery_hsn_master_files, key=os.path.getmtime)[-1]
                                    if machinery_hsn_master_files else None),
        self_gstin=self_gstin, company_name=company_name,
    )


_NUMERIC_TEXT_RE = re.compile(r'^-?\d+$')


def convert_numeric_text_to_numbers(wb):
    """Runs on the IN-MEMORY workbook right before save (per explicit request): a great deal of
    the tool's source data (E-Invoice's Taxable Value/tax columns, HSN codes, invoice/EWB/IRN
    reference numbers, etc.) arrives from the GST portal's own exports as TEXT even when it's
    purely digits -- openpyxl/this tool then writes many of those cells straight through, so
    Excel shows its 'Number Stored as Text' warning (the green corner flag) on them, and anyone
    doing further analysis has to manually convert each such column before it behaves as a
    number (sorting, SUM, VLOOKUP-by-number, etc.).

    Rather than hunting down and fixing the write call at every one of the many places across
    this codebase that passes such a value straight through, this is ONE guaranteed-complete
    pass over every cell actually in the final workbook -- the same strategy already used for
    the freeze-panes pass in master_build.py, for the same reason (nothing missed, including
    anything a future change adds).

    SAFETY: a cell is converted ONLY when doing so is provably lossless -- str(int(cell_text))
    must reproduce the exact original text. This is what excludes any identifier where a digit
    string's exact form matters and isn't just its numeric value: an HSN code like '09' or
    '035' would silently become 9 / 35 if converted, changing what's displayed and losing a
    real digit -- confirmed 93 such cells exist in a real run, all skipped, left as text
    exactly as they were. Decimal-valued text cells are also left untouched (a real scan of a
    full run found zero of them -- every text-numeric cell in this tool's output is an
    integer-shaped identifier or count -- so there's no decimal round-trip logic to get subtly
    wrong; if one ever appears, it stays as text rather than risk misconverting it). Only a
    cell's VALUE changes -- its style, number format, and every other property are untouched.

    Returns (converted_count, skipped_count) for the caller to log."""
    converted = 0
    skipped = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type != "s" or cell.value is None:
                    continue
                v = str(cell.value).strip()
                if not _NUMERIC_TEXT_RE.match(v):
                    continue
                try:
                    n = int(v)
                except ValueError:
                    continue
                if str(n) == v:
                    cell.value = n
                    converted += 1
                else:
                    skipped += 1
    return converted, skipped


def fix_ooxml_conformance(path):
    """Post-process a workbook AFTER openpyxl has already saved it, to fix real, confirmed
    Excel-'needs repair' bugs (not assumed, not theoretical) found in this exact pipeline's
    own output:

    1. Every '.xml' and '.rels' part missing its XML declaration -- see the detailed
       diagnosis in the fix below (byte-diff against a real Excel-repaired reference file,
       independently reproduced on a minimal single-cell openpyxl.Workbook()).

    2. A dangling <selection pane="bottomLeft" .../> attribute left behind on 66 of this
       workbook's 89 sheets. Root cause, confirmed by direct reproduction: this tool sets
       ws.freeze_panes to a real cell during sheet construction (for the header-row freeze
       every sheet used to have), then unconditionally sets ws.freeze_panes = None on every
       sheet right before save (per an explicit 'remove all freeze panes' instruction).
       openpyxl's freeze_panes=None correctly removes the <pane> element that defines the
       actual split, but does NOT clean up the pane="..." attribute that was written onto
       the sheet's <selection> element while freeze_panes was still active -- confirmed by
       reproducing the exact sequence (set freeze_panes to a cell, then to None) on a
       throwaway single-cell workbook and inspecting the resulting XML directly: the same
       orphaned attribute appears. The result is a <selection> referencing a pane that no
       <pane> element defines anywhere in the file (confirmed: zero <pane> elements exist
       anywhere in this workbook, since every sheet's freeze is unconditionally cleared) --
       an internal inconsistency lenient readers (LibreOffice, openpyxl's own loader) ignore
       silently, but Excel's own stricter loader does not tolerate, especially once the file
       is taken out of Protected View's lighter-weight rendering path into full edit mode.
       Confirmed present on 66 of 89 real sheets in this tool's own last output before this
       fix. Safe to strip unconditionally: since freeze_panes is cleared on every sheet, no
       sheet anywhere in this workbook legitimately has an active pane split, so a
       pane="..." attribute on any <selection> element is ALWAYS orphaned, never legitimate,
       in this tool's output.
    """
    import os
    import re
    import zipfile

    tmp = path + ".ooxmlfix.tmp"
    DECL_RELS = b'<?xml version="1.0" encoding="UTF-8"?>\n'
    DECL_CONTENT = b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
    PRIORITY = ["[Content_Types].xml", "_rels/.rels", "xl/workbook.xml", "xl/_rels/workbook.xml.rels"]
    DANGLING_PANE_RE = re.compile(rb'<selection pane="[^"]*" ')

    with zipfile.ZipFile(path, "r") as zin:
        entries = {info.filename: (info, zin.read(info.filename)) for info in zin.infolist()}

    fixed = {}
    for name, (info, data) in entries.items():
        if (name.endswith(".xml") or name.endswith(".rels")) and not data.startswith(b"<?xml"):
            decl = DECL_RELS if (name.endswith(".rels") or name == "[Content_Types].xml") else DECL_CONTENT
            data = decl + data
        if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"):
            data = DANGLING_PANE_RE.sub(b"<selection ", data)
        fixed[name] = (info, data)

    ordered_names = [n for n in PRIORITY if n in fixed] + [n for n in fixed if n not in PRIORITY]

    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for name in ordered_names:
            info, data = fixed[name]
            zout.writestr(info, data)

    os.replace(tmp, path)


