"""
merge_gstr3b.py
Drop this file (+ gst_merge_common.py) into a folder containing any number of
GSTR-3B workbooks and run:

    python merge_gstr3b.py

It will detect every GSTR-3B file in the folder (by content, not filename),
sort them chronologically by Tax Period, and produce GSTR3B_Merged.xlsx with
one full-formatting sheet per source file, ordered left to right by month.
"""
from openpyxl import Workbook, load_workbook
from gst_merge_common import (
    find_xlsx_files, detect_file_type, fy_start_year, month_key, copy_sheet_full,
)


def _first_nonempty(ws, cells):
    """Return the value of the first cell (from a list of cell refs) that isn't blank."""
    for cell in cells:
        val = ws[cell].value
        if val not in (None, ""):
            return val
    return None


def read_meta(ws):
    return {
        "gstin": _first_nonempty(ws, ["E5", "G5"]),
        "fy": _first_nonempty(ws, ["E6", "G6"]),
        "tax_period": _first_nonempty(ws, ["E7", "G7"]),
        "arn": _first_nonempty(ws, ["E8", "G8"]),
        "arn_date": _first_nonempty(ws, ["E9", "G9"]),
        "legal_name": _first_nonempty(ws, ["E10", "G10"]),
        "generated_on": _first_nonempty(ws, ["E12", "G12"]),
    }


def main(folder="."):
    files = find_xlsx_files(folder)
    gstr3b_files = [f for f in files if detect_file_type(f) == "GSTR3B"]
    if not gstr3b_files:
        print("No GSTR-3B files found in this folder.")
        return

    records = []
    for f in gstr3b_files:
        wb = load_workbook(f, data_only=True)
        ws = wb["GSTR-3B"]
        meta = read_meta(ws)
        if not meta["fy"] or not meta["tax_period"]:
            raise ValueError(
                f"'{f}' file mein FY ya Tax Period nahi mila (E6/G6, E7/G7 sab khali hain). "
                f"Is file ka 'GSTR-3B' sheet check karo."
            )
        key = (fy_start_year(meta["fy"]), month_key(meta["tax_period"]))
        records.append({"path": f, "wb": wb, "ws": ws, "meta": meta, "key": key})

    records.sort(key=lambda r: r["key"])

    print("Merge order (GSTR-3B):")
    for r in records:
        print(f"  {r['path']}  ->  FY {r['meta']['fy']}, Tax Period {r['meta']['tax_period']}")

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    used_names = set()
    for rec in records:
        meta = rec["meta"]
        base_name = f"{meta['tax_period']}_{meta['fy']}".replace("/", "-")
        name = base_name
        i = 2
        while name[:31] in used_names:
            name = f"{base_name}_{i}"
            i += 1
        used_names.add(name[:31])
        copy_sheet_full(rec["ws"], wb_out, name)

    out_path = "GSTR3B_Merged.xlsx"
    wb_out.save(out_path)
    print(f"\nSaved: {out_path}")


if __name__ == "__main__":
    main()
