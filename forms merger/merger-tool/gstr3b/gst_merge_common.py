"""
Shared helpers for GSTR1 / GSTR3B / E-Invoice merge scripts.  (v2)
Keep this file in the same folder as the merge_*.py scripts.

v2 adds two small defensive helpers (find_label_cell, dump_region) used by
merge_gstr3b.py to cope with GSTR-3B files whose header layout has shifted
away from the E5:E12 addresses the script originally assumed.
"""
import copy
import glob
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SEP_FILL = PatternFill("solid", fgColor="FFF2CC")
SEP_FONT = Font(bold=True, size=10)
HEADER_FONT = Font(bold=True)

MONTH_ORDER_IN_FY = {
    "apr": 1, "april": 1,
    "may": 2,
    "jun": 3, "june": 3,
    "jul": 4, "july": 4,
    "aug": 5, "august": 5,
    "sep": 6, "sept": 6, "september": 6,
    "oct": 7, "october": 7,
    "nov": 8, "november": 8,
    "dec": 9, "december": 9,
    "jan": 10, "january": 10,
    "feb": 11, "february": 11,
    "mar": 12, "march": 12,
}


def find_xlsx_files(folder="."):
    return sorted(glob.glob(os.path.join(folder, "*.xlsx")))


def detect_file_type(path):
    """Return 'EINV', 'GSTR1', 'GSTR3B', 'GSTR2B', or None based on sheet-name signature."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None
    names = set(wb.sheetnames)
    wb.close()
    if names == {"GSTR-3B"}:
        return "GSTR3B"
    if "Read me" in names and "ITC Available" in names:
        return "GSTR2B"
    if "Read me" in names and "b2cl" in names:
        return "GSTR1"
    if "Read me" in names and "b2b, sez, de" in names:
        return "EINV"
    return None


def fy_start_year(fy_string):
    # "2022-23" -> 2022
    return int(str(fy_string).split("-")[0].strip())


def month_key(month_name):
    # Quarterly tax periods look like "Apr-Jun" / "Jan-Mar" -> use the start month
    raw = str(month_name).strip()
    first_token = raw.split("-")[0].strip().lower()[:9]
    for k, v in MONTH_ORDER_IN_FY.items():
        if first_token.startswith(k) or k.startswith(first_token):
            return v
    raise ValueError(f"Unrecognised month/tax-period value: {month_name!r}")


def einv_period_to_key(mmyyyy):
    # "042022" -> (fy_start_year, order_in_fy)
    mmyyyy = str(mmyyyy).strip()
    month, year = int(mmyyyy[:2]), int(mmyyyy[2:])
    if month >= 4:
        return (year, month - 3)
    return (year - 1, month + 9)


def warn_duplicates(records):
    """Print a warning if two records share the same sort key (same period)."""
    seen = {}
    for r in records:
        if r["key"] in seen:
            print(
                f"WARNING: duplicate Tax Period detected - '{r['meta']['tax_period']}' "
                f"appears in both {seen[r['key']]} and {r['path']}. "
                f"Both will be merged in as separate blocks."
            )
        else:
            seen[r["key"]] = r["path"]


def sheet_max_data_row(ws, min_row):
    """Last row index (1-based) that has at least one non-empty cell, from min_row onward.

    PERFORMANCE: uses iter_rows(values_only=True) rather than repeated
    ws[r]/.cell().value access -- openpyxl builds a full Cell wrapper object
    per access for the latter, which is the dominant cost on a large sheet
    (measured: a real ~2000-row GSTR-2B 'B2B' sheet spent the bulk of a
    28s merge step in exactly this kind of cell-by-cell scanning).
    values_only=True yields plain values directly, skipping that overhead."""
    last = min_row - 1
    for offset, row in enumerate(ws.iter_rows(min_row=min_row, values_only=True)):
        if any(v is not None and str(v).strip() != "" for v in row):
            last = min_row + offset
    return last


def write_separator(ws_out, row_idx, text, n_cols):
    ws_out.cell(row=row_idx, column=1, value=text)
    for col in range(1, n_cols + 1):
        cell = ws_out.cell(row=row_idx, column=col)
        cell.font = SEP_FONT
        cell.fill = SEP_FILL
    if n_cols > 1:
        ws_out.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=n_cols)


def copy_sheet_full(src_ws, dst_wb, title):
    """Copy an entire worksheet (values, styles, merges, dimensions) into a new
    sheet of dst_wb. Used for GSTR-3B where each source file becomes one tab."""
    ws_out = dst_wb.create_sheet(title=title[:31])

    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = ws_out.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy.copy(cell.font)
                new_cell.fill = copy.copy(cell.fill)
                new_cell.border = copy.copy(cell.border)
                new_cell.alignment = copy.copy(cell.alignment)
                new_cell.number_format = cell.number_format

    for merged_range in src_ws.merged_cells.ranges:
        ws_out.merge_cells(str(merged_range))

    for col_letter, dim in src_ws.column_dimensions.items():
        if dim.width:
            ws_out.column_dimensions[col_letter].width = dim.width
    for row_idx, dim in src_ws.row_dimensions.items():
        if dim.height:
            ws_out.row_dimensions[row_idx].height = dim.height

    return ws_out


def find_label_cell(ws, label, max_row=25, max_col=10):
    """Scan the top-left block of a worksheet for the first cell whose text
    matches `label` (case-insensitive, ignoring a trailing colon). Returns
    (row, col) of that cell, or None if not found. Used as a fallback when a
    fixed cell address (e.g. 'E6') no longer holds what we expect, because
    the source template's rows or columns have shifted."""
    target = label.strip().lower()
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip().lower().rstrip(":").strip() == target:
                return r, c
    return None


def dump_region(ws, max_row=15, max_col=8):
    """Printable dump of a worksheet's top-left block, for diagnosing a file
    whose header layout doesn't match what the script expects."""
    lines = []
    for r in range(1, max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        lines.append(f"  row {r}: {row_vals}")
    return "\n".join(lines)


def value_after_label(ws, row, col):
    """Given a label cell's position, return the value that belongs to it -
    accounting for the label itself being merged across several columns
    (e.g. a template revision that widens the label column from B:D to
    B:F). openpyxl only stores a value in the top-left cell of a merged
    range, so the true value cell is the one immediately after wherever the
    label's own merge ends, not just 'one column to the right of the label
    cell' - that assumption breaks the moment the label merge gets wider."""
    end_col = col
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            end_col = rng.max_col
            break
    return ws.cell(row=row, column=end_col + 1).value


def label_lookup(ws, labels):
    """Try each label string in turn (first match wins) and return the
    value sitting next to whichever one is found in the sheet."""
    for label in labels:
        pos = find_label_cell(ws, label)
        if pos:
            r, c = pos
            return value_after_label(ws, r, c)
    return None


def robust_read_meta(ws, fixed_cells, label_fallbacks, path, key_validators=None):
    """Read a set of named fields from a worksheet's header block, in a way
    that survives the source template shifting rows or columns between
    filing periods (which the GST portal has done at least twice: a row
    shift on GSTR-3B, a column shift on GSTR-3B, and GSTR-2B's 'Read me'
    sheet uses the same label+merged-value layout so it's exposed to the
    same risk even though it hasn't broken yet).

    fixed_cells: {key: "A1"-style address} - the fast path, tried first,
        keeps this cheap for the common case where nothing has shifted.
    label_fallbacks: {key: [label strings to search for, in priority
        order]} - used whenever the fixed-cell value fails validation.
    key_validators: optional {key: callable(value) -> bool}. Keys listed
        here get their VALUE checked, not just "is it non-empty" - this
        catches a shifted-but-non-empty cell holding the wrong field
        entirely, not only a genuinely blank one. Keys without a validator
        are only checked for emptiness.

    Returns the meta dict, or None if any validated key still can't be
    found after both the fixed-cell and label-search attempts - in which
    case a dump of the sheet's header block is printed so the exact new
    layout can be identified from the printed output.
    """
    key_validators = key_validators or {}
    meta = {key: ws[cell].value for key, cell in fixed_cells.items()}

    for key in meta:
        validator = key_validators.get(key)
        ok = validator(meta[key]) if validator else meta[key] not in (None, "", "None")
        if not ok:
            meta[key] = label_lookup(ws, label_fallbacks.get(key, [key]))

    failed = [k for k, v in key_validators.items() if not v(meta.get(k))]
    if failed:
        print(f"\n[SKIP] {path}")
        print(f"  Could not find a valid value for: {failed}")
        print("  (checked the expected fixed cells, then scanned the sheet for")
        print("  the label text itself). This file's header layout differs from")
        print("  the others. Dump of its top-left block:")
        print(dump_region(ws))
        return None

    return meta


def looks_like_fy(v):
    return v is not None and re.match(r"^\s*\d{4}\s*-\s*\d{2,4}\s*$", str(v)) is not None


def looks_like_tax_period(v):
    if v is None or str(v).strip() == "":
        return False
    try:
        month_key(v)
        return True
    except ValueError:
        return False
