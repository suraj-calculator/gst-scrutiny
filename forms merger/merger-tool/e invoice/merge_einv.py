"""
merge_einv.py
Drop this file (+ gst_merge_common.py) into a folder containing any number of
E-Invoice GSTR-1 auto-populated workbooks and run:

    python merge_einv.py

It will detect every E-Invoice file in the folder (by content, not filename),
sort them chronologically by Tax Period, and produce EINV_Merged.xlsx with the
same 4 data sheets (Read me sheet dropped), one continuous month-wise block per
sheet, separated by a labelled separator row.
"""
from openpyxl import Workbook, load_workbook
from gst_merge_common import (
    find_xlsx_files, detect_file_type, einv_period_to_key,
    sheet_max_data_row, write_separator, HEADER_FONT,
)

DATA_SHEETS = ["b2b, sez, de", "cdnr", "cdnur", "exp"]
HEADER_ROWS = 4  # title, blank, section title, column headers


def read_meta(wb):
    ws = wb["Read me"]
    return {
        "fy": ws["C4"].value,
        "tax_period": ws["C5"].value,
        "gstin": ws["C6"].value,
        "legal_name": ws["C7"].value,
        "date_updated_till": ws["C9"].value,
    }


def main(folder="."):
    files = find_xlsx_files(folder)
    einv_files = [f for f in files if detect_file_type(f) == "EINV"]
    if not einv_files:
        print("No E-Invoice files found in this folder.")
        return

    records = []
    for f in einv_files:
        wb = load_workbook(f, data_only=True)
        meta = read_meta(wb)
        key = einv_period_to_key(meta["tax_period"])
        records.append({"path": f, "wb": wb, "meta": meta, "key": key})

    records.sort(key=lambda r: r["key"])

    print("Merge order (E-Invoice):")
    for r in records:
        print(f"  {r['path']}  ->  FY {r['meta']['fy']}, Tax Period {r['meta']['tax_period']}")

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    for sheet_name in DATA_SHEETS:
        first_ws = records[0]["wb"][sheet_name]
        n_cols = first_ws.max_column
        ws_out = wb_out.create_sheet(title=sheet_name)

        for r in range(1, HEADER_ROWS + 1):
            for c in range(1, n_cols + 1):
                src_cell = first_ws.cell(row=r, column=c)
                dst_cell = ws_out.cell(row=r, column=c, value=src_cell.value)
                if r == HEADER_ROWS:
                    dst_cell.font = HEADER_FONT

        current_row = HEADER_ROWS + 1
        for idx, rec in enumerate(records):
            meta = rec["meta"]
            sep_text = (
                f"Financial Year: {meta['fy']}  |  Tax Period: {meta['tax_period']}  |  "
                f"Date Updated till: {meta['date_updated_till']}"
            )
            write_separator(ws_out, current_row, sep_text, n_cols)
            current_row += 1

            ws_src = rec["wb"][sheet_name]
            last_row = sheet_max_data_row(ws_src, HEADER_ROWS + 1)
            for r in range(HEADER_ROWS + 1, last_row + 1):
                for c in range(1, n_cols + 1):
                    ws_out.cell(row=current_row, column=c, value=ws_src.cell(row=r, column=c).value)
                current_row += 1

        for col_letter, dim in first_ws.column_dimensions.items():
            if dim.width:
                ws_out.column_dimensions[col_letter].width = dim.width

    out_path = "EINV_Merged.xlsx"
    wb_out.save(out_path)
    print(f"\nSaved: {out_path}")


if __name__ == "__main__":
    main()
