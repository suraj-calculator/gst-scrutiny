"""
Shared helpers for GSTR1 / GSTR3B / E-Invoice merge scripts.
Keep this file in the same folder as the merge_*.py scripts.
"""
import copy
import glob
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SEP_FILL = PatternFill("solid", fgColor="FFF2CC")
SEP_FONT = Font(bold=True, size=10)
HEADER_FONT = Font(bold=True)

MONTH_ORDER_IN_FY = {
    "apr": 1, "april": 1,
    "may": 2,
    "jun": 3, "june": 3,
    "jul": 4, "july": 4,
    "aug": 5, "august": 5,
    "sep": 6, "sept": 6, "september": 6,
    "oct": 7, "october": 7,
    "nov": 8, "november": 8,
    "dec": 9, "december": 9,
    "jan": 10, "january": 10,
    "feb": 11, "february": 11,
    "mar": 12, "march": 12,
}


def find_xlsx_files(folder="."):
    return sorted(glob.glob(os.path.join(folder, "*.xlsx")))


def detect_file_type(path):
    """Return 'EINV', 'GSTR1', 'GSTR3B', 'GSTR2B', or None based on sheet-name signature."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None
    names = set(wb.sheetnames)
    wb.close()
    if names == {"GSTR-3B"}:
        return "GSTR3B"
    if "Read me" in names and "ITC Available" in names:
        return "GSTR2B"
    if "Read me" in names and "b2cl" in names:
        return "GSTR1"
    if "Read me" in names and "b2b, sez, de" in names:
        return "EINV"
    return None


def fy_start_year(fy_string):
    # "2022-23" -> 2022
    return int(str(fy_string).split("-")[0].strip())


def month_key(month_name):
    # Quarterly tax periods look like "Apr-Jun" / "Jan-Mar" -> use the start month
    raw = str(month_name).strip()
    first_token = raw.split("-")[0].strip().lower()[:9]
    for k, v in MONTH_ORDER_IN_FY.items():
        if first_token.startswith(k) or k.startswith(first_token):
            return v
    raise ValueError(f"Unrecognised month/tax-period value: {month_name!r}")


def einv_period_to_key(mmyyyy):
    # "042022" -> (fy_start_year, order_in_fy)
    mmyyyy = str(mmyyyy).strip()
    month, year = int(mmyyyy[:2]), int(mmyyyy[2:])
    if month >= 4:
        return (year, month - 3)
    return (year - 1, month + 9)


def warn_duplicates(records):
    """Print a warning if two records share the same sort key (same period)."""
    seen = {}
    for r in records:
        if r["key"] in seen:
            print(
                f"WARNING: duplicate Tax Period detected - '{r['meta']['tax_period']}' "
                f"appears in both {seen[r['key']]} and {r['path']}. "
                f"Both will be merged in as separate blocks."
            )
        else:
            seen[r["key"]] = r["path"]


def sheet_max_data_row(ws, min_row):
    """Last row index (1-based) that has at least one non-empty cell, from min_row onward."""
    last = min_row - 1
    for r in range(min_row, ws.max_row + 1):
        if any(c.value is not None and str(c.value).strip() != "" for c in ws[r]):
            last = r
    return last


def write_separator(ws_out, row_idx, text, n_cols):
    ws_out.cell(row=row_idx, column=1, value=text)
    for col in range(1, n_cols + 1):
        cell = ws_out.cell(row=row_idx, column=col)
        cell.font = SEP_FONT
        cell.fill = SEP_FILL
    if n_cols > 1:
        ws_out.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=n_cols)


def copy_sheet_full(src_ws, dst_wb, title):
    """Copy an entire worksheet (values, styles, merges, dimensions) into a new
    sheet of dst_wb. Used for GSTR-3B where each source file becomes one tab."""
    ws_out = dst_wb.create_sheet(title=title[:31])

    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = ws_out.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy.copy(cell.font)
                new_cell.fill = copy.copy(cell.fill)
                new_cell.border = copy.copy(cell.border)
                new_cell.alignment = copy.copy(cell.alignment)
                new_cell.number_format = cell.number_format

    for merged_range in src_ws.merged_cells.ranges:
        ws_out.merge_cells(str(merged_range))

    for col_letter, dim in src_ws.column_dimensions.items():
        if dim.width:
            ws_out.column_dimensions[col_letter].width = dim.width
    for row_idx, dim in src_ws.row_dimensions.items():
        if dim.height:
            ws_out.row_dimensions[row_idx].height = dim.height

    return ws_out
