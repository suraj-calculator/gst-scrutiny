"""
merge_r2a.py
Drop this file (+ gst_merge_common.py) into a folder containing any number of
GSTR-2A (R2A) workbooks - monthly, any FY, any count - and run:

    python merge_r2a.py

It detects every R2A file in the folder by content (sheet-name signature,
not filename), sorts them chronologically by Tax Period, and produces
R2A_Merged.xlsx with the same sheets as the source files (Read me sheet kept
as-is from the earliest period), one continuous month-wise block per sheet,
separated by a labelled separator row.

Source layout (confirmed against sample files): every data sheet
(B2B, B2BA, CDNR, CDNRA, ECO, ECOA, ISD, ISDA, TDS, TDSA, TCS, IMPG,
IMPG SEZ) has the same 6-row header (title / blank / blank / section title /
column headers spanning rows 5-6) before data starts at row 7 - unlike
GSTR-2B, R2A doesn't have any sheet whose header text itself changes per
period, so a single shared header per sheet is safe to reuse for every block.
"""
from openpyxl import Workbook
from gst_merge_common import (
    find_xlsx_files, detect_file_type, einv_period_to_key,
    sheet_max_data_row, write_separator, HEADER_FONT, copy_sheet_full,
    warn_duplicates, load_data_workbook,
)

HEADER_ROWS = 6  # title, blank, blank, section title, column headers (2 rows)


def read_meta(wb):
    ws = wb["Read me"]
    return {
        "gstin": ws["C2"].value,
        "tax_period": ws["E2"].value,
        "legal_name": ws["C3"].value,
        "fy": ws["E3"].value,
        "trade_name": ws["C4"].value,
        "generated_on": ws["E4"].value,
    }


def main(folder="."):
    files = find_xlsx_files(folder)
    r2a_files = [f for f in files if detect_file_type(f) == "R2A"]
    if not r2a_files:
        print("No GSTR-2A (R2A) files found in this folder.")
        return

    records = []
    for f in r2a_files:
        wb = load_data_workbook(f)
        meta = read_meta(wb)
        key = einv_period_to_key(meta["tax_period"])
        records.append({"path": f, "wb": wb, "meta": meta, "key": key})

    records.sort(key=lambda r: r["key"])

    print("Merge order (GSTR-2A / R2A):")
    for r in records:
        print(f"  {r['path']}  ->  FY {r['meta']['fy']}, Tax Period {r['meta']['tax_period']}")

    warn_duplicates(records)

    data_sheets = [s for s in records[0]["wb"].sheetnames if s != "Read me"]

    # sanity check: all files must have the same sheet set
    for r in records:
        other = [s for s in r["wb"].sheetnames if s != "Read me"]
        if other != data_sheets:
            print(f"WARNING: {r['path']} has a different sheet layout, sheets may not align correctly.")

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    # Keep the Read me sheet from the earliest period as-is (generic instructions).
    copy_sheet_full(records[0]["wb"]["Read me"], wb_out, "Read me")

    for sheet_name in data_sheets:
        first_ws = records[0]["wb"][sheet_name]
        n_cols = first_ws.max_column
        ws_out = wb_out.create_sheet(title=sheet_name[:31])

        for r in range(1, HEADER_ROWS + 1):
            for c in range(1, n_cols + 1):
                src_cell = first_ws.cell(row=r, column=c)
                dst_cell = ws_out.cell(row=r, column=c, value=src_cell.value)
                if r == HEADER_ROWS:
                    dst_cell.font = HEADER_FONT

        current_row = HEADER_ROWS + 1
        for rec in records:
            meta = rec["meta"]
            sep_text = (
                f"Financial Year: {meta['fy']}  |  Tax Period: {meta['tax_period']}  |  "
                f"GSTIN: {meta['gstin']}  |  Date of Generation: {meta['generated_on']}"
            )
            write_separator(ws_out, current_row, sep_text, n_cols)
            current_row += 1

            ws_src = rec["wb"][sheet_name]
            n_cols_src = ws_src.max_column
            last_row = sheet_max_data_row(ws_src, HEADER_ROWS + 1)
            for r in range(HEADER_ROWS + 1, last_row + 1):
                for c in range(1, n_cols_src + 1):
                    ws_out.cell(row=current_row, column=c, value=ws_src.cell(row=r, column=c).value)
                current_row += 1

        for col_letter, dim in first_ws.column_dimensions.items():
            if dim.width:
                ws_out.column_dimensions[col_letter].width = dim.width

    out_path = "R2A_Merged.xlsx"
    wb_out.save(out_path)
    print(f"\nSaved: {out_path}")


if __name__ == "__main__":
    main()
