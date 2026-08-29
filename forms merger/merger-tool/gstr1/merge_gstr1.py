"""
merge_gstr1.py  (v2)
Drop this file (+ gst_merge_common.py) into a folder containing any number of
GSTR-1 workbooks and run:

    python merge_gstr1.py

It will detect every GSTR-1 file in the folder (by content, not filename),
sort them chronologically by Tax Period, and produce GSTR1_Merged.xlsx with
one continuous month-wise block per sheet, separated by a labelled
separator row.

v2 rebuild - the previous version assumed every input file has the same
sheet set (data_sheets taken from the first file only) and a fixed 4-row
header for every sheet. Both assumptions broke on real filings from a
single GSTIN within one financial year:

  - Sheet set: the April-2025 file has one 'hsn' sheet; every other month
    (May-2025 onward) splits it into 'hsn(b2b)' and 'hsn(b2c)' (a portal
    template change, same family of change as GSTR-2B's IMS rollout). The
    old script used whichever file sorted first as the "reference" sheet
    list and then did wb[sheet_name] against every other file's workbook
    unconditionally - which throws a KeyError the moment a period is
    missing a sheet the reference file has.

  - Header rows: hardcoded to 4 for every sheet. That happens to be
    correct for GSTR-1 as of this filing year, but it's exactly the kind
    of assumption that broke GSTR-2B's merge before (some of its sheets
    need 6 header rows, some need 7) - so this version uses the same
    shared, self-detecting logic instead of a fixed number, rather than
    waiting for GSTR-1 to eventually grow that same kind of variation.

This version now shares its architecture with merge_gstr2b.py:
  - Sheet list is the UNION across all files, not just the first one; each
    sheet only pulls in the periods that actually contain it, with a note
    printed for any period that's missing it.
  - Header-row count is detected per sheet via
    gst_merge_common.detect_header_rows() - it finds the actual boundary
    between header and data (the first row with a cell that looks like
    real GST data: a GSTIN, a date, a number) rather than assuming a fixed
    row count or relying on merged cells, which GSTR-1's flat-row headers
    don't use.
  - read_meta() goes through gst_merge_common.robust_read_meta(), so a
    'Read me' sheet whose layout shifts (as has now happened on both
    GSTR-3B and, in spirit, GSTR-2B/GSTR-1's HSN split) gets a label-based
    fallback and a diagnostic dump instead of crashing the batch.
"""
from openpyxl import Workbook, load_workbook
from gst_merge_common import (
    find_xlsx_files, detect_file_type, fy_start_year, month_key,
    sheet_max_data_row, write_separator, HEADER_FONT, copy_sheet_full, warn_duplicates,
    robust_read_meta, looks_like_fy, looks_like_tax_period, detect_header_rows,
)

FIXED_CELLS = {
    "fy": "C4",
    "tax_period": "C5",
    "gstin": "C6",
    "legal_name": "C7",
    "arn": "C9",
    "arn_date": "C10",
    "generated_on": "C11",
}

LABEL_FALLBACKS = {
    "fy": ["Financial Year", "Year", "FY"],
    "tax_period": ["Tax Period"],
    "gstin": ["GSTIN"],
    "legal_name": ["Legal Name", "Legal name of the registered person"],
    "arn": ["ARN"],
    "arn_date": ["ARN date", "Date of ARN"],
    "generated_on": ["Date and Time of Generation", "Generated on", "Date of generation"],
}

VALIDATORS = {
    "fy": looks_like_fy,
    "tax_period": looks_like_tax_period,
}

# Sheet types seen and verified as of this script version - used only to
# flag genuinely new sheet types the portal introduces later, so you get a
# heads-up rather than the tool silently doing its best guess with no notice.
KNOWN_SHEETS = {
    "b2b, sez, de_inv", "b2cl", "exp", "b2cs", "exemp", "b2ba", "b2cla", "expa",
    "cdnr", "cdnur", "cdnra", "cdnura", "b2csa", "at", "atadj", "ata", "atadja",
    "hsn", "hsn(b2b)", "hsn(b2c)", "docs", "eco", "ecoa", "eco_9(5)", "ecoa_9(5)",
}


def read_meta(wb, path):
    return robust_read_meta(wb["Read me"], FIXED_CELLS, LABEL_FALLBACKS, path, VALIDATORS)


def sep_text(meta):
    return (
        f"Financial Year: {meta['fy']}  |  Tax Period: {meta['tax_period']}  |  "
        f"ARN: {meta['arn']}  |  ARN Date: {meta['arn_date']}  |  "
        f"Date and Time of Generation: {meta['generated_on']}"
    )


def merge_sheet(wb_out, sheet_name, records_with_sheet):
    first_ws = records_with_sheet[0]["wb"][sheet_name]
    header_rows = detect_header_rows(first_ws)
    n_cols = first_ws.max_column
    ws_out = wb_out.create_sheet(title=sheet_name[:31])

    for r in range(1, header_rows + 1):
        for c in range(1, n_cols + 1):
            src_cell = first_ws.cell(row=r, column=c)
            dst_cell = ws_out.cell(row=r, column=c, value=src_cell.value)
            if r == header_rows:
                dst_cell.font = HEADER_FONT

    current_row = header_rows + 1
    for rec in records_with_sheet:
        write_separator(ws_out, current_row, sep_text(rec["meta"]), n_cols)
        current_row += 1

        ws_src = rec["wb"][sheet_name]
        n_cols_src = ws_src.max_column
        last_row = sheet_max_data_row(ws_src, header_rows + 1)
        for r in range(header_rows + 1, last_row + 1):
            for c in range(1, n_cols_src + 1):
                ws_out.cell(row=current_row, column=c, value=ws_src.cell(row=r, column=c).value)
            current_row += 1

    for col_letter, dim in first_ws.column_dimensions.items():
        if dim.width:
            ws_out.column_dimensions[col_letter].width = dim.width


def main(folder="."):
    import os
    out_name = "GSTR1_Merged.xlsx"
    files = [f for f in find_xlsx_files(folder) if os.path.basename(f) != out_name]
    gstr1_files = [f for f in files if detect_file_type(f) == "GSTR1"]
    if not gstr1_files:
        print("No GSTR-1 files found in this folder.")
        return

    records = []
    skipped = []
    for f in gstr1_files:
        wb = load_workbook(f, data_only=True)
        meta = read_meta(wb, f)
        if meta is None:
            skipped.append(f)
            continue
        key = (fy_start_year(meta["fy"]), month_key(meta["tax_period"]))
        records.append({"path": f, "wb": wb, "meta": meta, "key": key})

    if not records:
        print("\nNo GSTR-1 file could be read - see the dump(s) above.")
        return

    records.sort(key=lambda r: r["key"])

    print("Merge order (GSTR-1):")
    for r in records:
        print(f"  {r['path']}  ->  FY {r['meta']['fy']}, Tax Period {r['meta']['tax_period']}")
    if skipped:
        print(f"\n{len(skipped)} file(s) skipped - see dump(s) above:")
        for f in skipped:
            print(f"  {f}")

    warn_duplicates(records)

    # Union of every sheet seen across ALL files - a period can drop or
    # split a sheet (like the hsn / hsn(b2b)+hsn(b2c) split here) without
    # warning, so this can't be based on just one "reference" file.
    all_sheet_names = []
    for r in records:
        for name in r["wb"].sheetnames:
            if name != "Read me" and name not in all_sheet_names:
                all_sheet_names.append(name)

    new_sheets = [s for s in all_sheet_names if s not in KNOWN_SHEETS]
    if new_sheets:
        print(f"\nHeads up: {len(new_sheets)} sheet type(s) not seen before - "
              f"header rows were auto-detected, worth a quick look at the "
              f"merged output to confirm they line up:")
        for s in new_sheets:
            print(f"  {s}")

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    copy_sheet_full(records[0]["wb"]["Read me"], wb_out, "Read me")

    for sheet_name in all_sheet_names:
        records_with_sheet = [r for r in records if sheet_name in r["wb"].sheetnames]
        records_missing = [r for r in records if sheet_name not in r["wb"].sheetnames]

        if records_missing:
            print(f"\nNote: '{sheet_name}' is missing from {len(records_missing)} file(s) "
                  f"- merging without them:")
            for r in records_missing:
                print(f"  {r['path']}  (Tax Period {r['meta']['tax_period']})")

        if not records_with_sheet:
            continue

        merge_sheet(wb_out, sheet_name, records_with_sheet)

    out_path = "GSTR1_Merged.xlsx"
    wb_out.save(out_path)
    print(f"\nSaved: {out_path}")


if __name__ == "__main__":
    main()
