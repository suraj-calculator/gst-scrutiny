"""
complete_workbooks.py
----------------------
Kai saari Excel workbooks lo, jinme kisi mein sab worksheets hain aur
kisi mein kam. Ye tool sabhi files ka "master" sheet-name list nikal
kar, jin files mein koi sheet missing hai unme usi naam se worksheet
add kar deta hai, sahi order mein.

- Zyaadatar missing sheets BLANK add hoti hain (khali rows/columns).
- Kuch chuni hui sheets (default: "Read me") ke liye - agar wo missing
  hai to use blank nahi rakha jaata, balki reference file (jisme sabse
  zyada sheets hain / jo master order ke liye use hui) ki "Read me"
  sheet ka content copy karke paste kar diya jaata hai.
- Agar kisi file mein "Read me" (ya jo bhi copy-sheet ho) pehle se
  MAUJOOD hai (chahe khali ho ya bhari), use bilkul CHHEDA nahi jaata -
  wo as-it-is rehti hai.

IMPORTANT: Ye tool files ko IN-PLACE update karta hai, matlab original
input file hi overwrite ho jaati hai updated version se. Koi alag
"_completed.xlsx" copy nahi banti - folder mein sirf updated files
rahengi. Isliye agar original files ka backup chahiye to run karne se
pehle khud copy bana lena.

USAGE:
    python complete_workbooks.py file1.xlsx file2.xlsx file3.xlsx ...

    Ya poori folder ki saari .xlsx files ke liye:
    python complete_workbooks.py --folder /path/to/folder

    Master sheet order aur reference content us file se liya jaata hai
    jisme sabse zyada sheets hain (agar tie ho to sabse pehle mili file
    se). Khud fix karna ho to --master-order flag se file do.

    Jin sheet naamon ka content (blank ke bajaye) copy hona chahiye,
    unki list --copy-sheets se do (comma-separated). Default: "Read me"
    Example: --copy-sheets "Read me,Instructions"
"""

import argparse
import copy
import glob
import os
import sys

from openpyxl import load_workbook


DEFAULT_COPY_SHEETS = ["Read me"]


def get_master_sheet_order(files, master_order_file=None):
    """
    Sabhi files ke sheet names ka union nikaalta hai, order preserve
    karte hue. Reference file bhi return karta hai - jiska content
    'copy-sheets' (jaise Read me) ke liye source banega.
    """
    if master_order_file:
        wb = load_workbook(master_order_file, read_only=True)
        order = list(wb.sheetnames)
        wb.close()
        extra = []
        for f in files:
            wb = load_workbook(f, read_only=True)
            for name in wb.sheetnames:
                if name not in order and name not in extra:
                    extra.append(name)
            wb.close()
        return order + extra, master_order_file

    # Sabse zyada sheets wali file ko base/reference maano
    best_file, best_names = None, []
    for f in files:
        wb = load_workbook(f, read_only=True)
        names = list(wb.sheetnames)
        wb.close()
        if len(names) > len(best_names):
            best_file, best_names = f, names

    order = list(best_names)
    for f in files:
        wb = load_workbook(f, read_only=True)
        for name in wb.sheetnames:
            if name not in order:
                order.append(name)
        wb.close()
    return order, best_file


def copy_sheet_content(src_ws, dst_ws):
    """Ek worksheet ka data + basic formatting doosri (khali) worksheet mein copy karta hai."""
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy.copy(cell.font)
                new_cell.fill = copy.copy(cell.fill)
                new_cell.border = copy.copy(cell.border)
                new_cell.alignment = copy.copy(cell.alignment)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy.copy(cell.protection)

    # Column widths
    for col_letter, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = dim.width

    # Row heights
    for row_idx, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_idx].height = dim.height

    # Merged cells
    for merged_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged_range))


def complete_workbook(path, master_order, reference_path, copy_sheet_names):
    """
    File ko complete karke USI path par overwrite kar deta hai.
    - Missing normal sheets -> blank add hoti hain.
    - Missing 'copy_sheet_names' (jaise Read me) -> reference file se
      content copy karke add hoti hain.
    - Jo sheet pehle se maujood hai (chahe copy-sheet ho ya normal),
      use bilkul touch nahi kiya jaata.
    """
    wb = load_workbook(path)
    existing = set(wb.sheetnames)
    added_blank = []
    added_copied = []

    reference_wb = None
    if reference_path and os.path.abspath(reference_path) != os.path.abspath(path):
        reference_wb = load_workbook(reference_path, data_only=False)

    for name in master_order:
        if name in existing:
            continue  # pehle se hai -> as-it-is, chheda nahi

        new_ws = wb.create_sheet(title=name)

        if name in copy_sheet_names and reference_wb is not None and name in reference_wb.sheetnames:
            copy_sheet_content(reference_wb[name], new_ws)
            added_copied.append(name)
        else:
            added_blank.append(name)  # blank hi rahegi

    # Master order ke exact sequence mein reorder karo
    ordered_sheets = [wb[name] for name in master_order if name in wb.sheetnames]
    wb._sheets = ordered_sheets

    # Pehle ek temp file mein save karo, fir original ko replace karo.
    tmp_path = path + ".tmp_writing.xlsx"
    wb.save(tmp_path)
    os.replace(tmp_path, path)
    return path, added_blank, added_copied


def main():
    parser = argparse.ArgumentParser(description="Missing worksheets ko complete karo (files ko in-place update karta hai).")
    parser.add_argument("files", nargs="*", help="Excel files (.xlsx)")
    parser.add_argument("--folder", help="Is folder ki saari .xlsx files process karo")
    parser.add_argument("--master-order", dest="master_order_file",
                         help="Sheet order/reference fix karne ke liye file (optional)")
    parser.add_argument("--copy-sheets", default=",".join(DEFAULT_COPY_SHEETS),
                         help='Comma-separated sheet naam jinka content copy hona chahiye (blank ke bajaye). '
                              f'Default: "{", ".join(DEFAULT_COPY_SHEETS)}"')
    args = parser.parse_args()

    files = list(args.files)
    if args.folder:
        files += sorted(glob.glob(os.path.join(args.folder, "*.xlsx")))

    files = [f for f in files if os.path.isfile(f)]
    if not files:
        print("Koi valid .xlsx file nahi mili. File paths ya --folder do.")
        sys.exit(1)

    copy_sheet_names = [s.strip() for s in args.copy_sheets.split(",") if s.strip()]

    master_order, reference_path = get_master_sheet_order(files, args.master_order_file)
    print(f"Master sheet list ({len(master_order)} sheets): {master_order}")
    print(f"Reference file (content-copy source): {os.path.basename(reference_path) if reference_path else 'N/A'}")
    print(f"Copy-content sheets: {copy_sheet_names}")
    print("NOTE: Original files ko hi update kiya ja raha hai (overwrite). "
          "Koi extra '_completed' copy nahi banegi.\n")

    for f in files:
        _, added_blank, added_copied = complete_workbook(f, master_order, reference_path, copy_sheet_names)
        if added_blank or added_copied:
            print(f"[OK] {os.path.basename(f)}")
            if added_copied:
                print(f"     Content copy hui (reference se): {added_copied}")
            if added_blank:
                print(f"     Blank add hui: {added_blank}")
        else:
            print(f"[OK] {os.path.basename(f)} -> pehle se hi complete thi (dobara save ho gayi)")
        print()


if __name__ == "__main__":
    main()
