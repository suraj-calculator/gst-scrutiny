import os
import glob
import pandas as pd
from openpyxl import Workbook
import re

def convert_html_to_excel(html_file, output_file):
    """Convert HTML file to proper Excel format"""
    try:
        # Read HTML content
        with open(html_file, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Try to extract table data
        tables = pd.read_html(html_file)
        if tables:
            df = tables[0]
            df.to_excel(output_file, index=False, engine='openpyxl')
            return True
    except:
        pass
    return False

def convert_with_openpyxl(file_path, output_file):
    """Convert using openpyxl"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, data_only=True)
        sheet = wb.active
        
        # Read data
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        
        if data:
            df = pd.DataFrame(data[1:], columns=data[0])
            df.to_excel(output_file, index=False, engine='openpyxl')
            return True
    except:
        pass
    return False

def convert_with_csv_method(file_path, output_file):
    """Try to read as CSV"""
    try:
        # Try different encodings
        for encoding in ['utf-8', 'latin1', 'cp1252']:
            try:
                df = pd.read_csv(file_path, encoding=encoding, on_bad_lines='skip')
                if len(df) > 0:
                    df.to_excel(output_file, index=False, engine='openpyxl')
                    return True
            except:
                continue
    except:
        pass
    return False

def convert_file(file_path):
    """Convert a single file using multiple methods"""
    file_name = os.path.basename(file_path)
    output_file = file_path.replace('.xls', '.xlsx')
    
    print(f"  Converting: {file_name}")
    
    # Method 1: Try reading as HTML
    if convert_html_to_excel(file_path, output_file):
        print(f"    ✅ Converted (HTML method)")
        return True
    
    # Method 2: Try openpyxl
    if convert_with_openpyxl(file_path, output_file):
        print(f"    ✅ Converted (Openpyxl method)")
        return True
    
    # Method 3: Try CSV method
    if convert_with_csv_method(file_path, output_file):
        print(f"    ✅ Converted (CSV method)")
        return True
    
    print(f"    ❌ Failed to convert")
    return False

def main():
    print("="*60)
    print("EWB FILE CONVERTER - .xls to .xlsx")
    print("="*60)
    
    folders = ["inward_eway bill", "outward_eway bill"]
    converted = 0
    failed = 0
    
    for folder in folders:
        if os.path.exists(folder):
            print(f"\n📁 Processing: {folder}")
            files = glob.glob(os.path.join(folder, "*.xls"))
            
            for file in files:
                if convert_file(file):
                    converted += 1
                else:
                    failed += 1
    
    print("\n" + "="*60)
    print(f"✅ Converted: {converted} files")
    print(f"❌ Failed: {failed} files")
    print("="*60)
    
    if converted > 0:
        print("\n🎉 Conversion complete! Now you can run the auto-merger.")
    else:
        print("\n⚠️  No files were converted. Files may be corrupted.")
        print("   Try opening them manually in Excel and save as .xlsx")
    
    input("\nPress Enter to exit...")

if __name__ == "__main__":
    main()