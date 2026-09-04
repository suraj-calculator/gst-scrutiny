#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MACHINERY HSN SCAN -- flags likely plant & machinery purchases/sales (Chapter
84 + 85 HSN headings) from a taxpayer-supplied master list, so the review
team can judge whether the taxpayer is engaged in manufacturing (inward) or
has disposed of capital goods (outward, which can carry its own GST
consequences under Sec 18(6)).

THREE independent signals, cross-checked against each other -- not one:

  1. GSTR-1 HSN Summary (Outward / Sale) -- REAL HSN-based match. GSTR-1's
     own HSN-wise summary carries actual HSN codes; every row is checked as
     a 4-digit-prefix match against the master's 135 headings.

  2. E-Way Bill, both directions (Outward / Sale AND Inward / Purchase) --
     REAL HSN-based match. CONFIRMED (not assumed): unlike GSTR-2A/2B,
     e-way bill exports DO carry an 'HSN Code' column per movement. This is
     what makes the INWARD side possible at all -- see point 3.

  3. Trade-Name keyword screen (Inward only) -- GSTR-2B's own B2B sheet has
     NO HSN/SAC column (confirmed against a real export -- same finding
     already documented in gst_checks_hsn_fraud.check_blocked_itc_by_hsn,
     finding A5, and in gst_blocked_credit.py), so it CANNOT be matched by
     HSN directly. This is a lower-confidence, name-based cross-check
     instead: a curated (not machine-derived) subset of ~51 of the more
     distinctive machine-type terms from the master's own HSN descriptions,
     hand-reviewed one by one against the actual heading text before being
     used for matching -- generic/ambiguous terms ('parts', 'boards',
     'taps', 'electric', 'agricultural') were deliberately left out because
     they would be common in unrelated trade names and would just add noise
     rather than signal. TWO terms that looked reasonable on paper were
     REMOVED after testing against real data showed them producing near-total
     false positives ('crusher' matched stone-aggregate traders, not
     crushing-machine sellers; 'motor' matched vehicle-parts stores) -- see
     KEYWORD_MAP's own comments for exactly what was cut and the real
     evidence for why. This keyword list is provided for review/editing in
     the sheet's own notes, not hidden in code.

Self-to-self e-way bill movements (own-GSTIN-to-own-GSTIN, e.g. branch
transfers) are EXCLUDED from both EWB directions -- confirmed against real
data these are common (89% of this taxpayer's own outward EWBs) and are not
a purchase or a sale.

Purely additive: new module, new sheet only. Never changes any existing
calculation or sheet. Degrades to a clearly-explained skip if the master
file or any given source isn't available -- never crashes the pipeline.
"""

import re

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# styling -- kept LOCAL to this module (same convention as gst_blocked_credit.py
# and gst_checks_flow.py's own stated reason: any module can be dropped into
# any workbook on its own).
RED = PatternFill("solid", fgColor="FFC7CE")
AMBER = PatternFill("solid", fgColor="FFEB9C")
GREEN = PatternFill("solid", fgColor="C6EFCE")
HEADFILL = PatternFill("solid", fgColor="1F3864")
TITLEF = Font(bold=True, size=13, color="1F3864")
SECTF = Font(bold=True, size=11, color="1F3864")
BORDER = Border(*[Side(style="thin", color="BFBFBF")] * 4)
PLAIN_FONT = Font(size=10)
BOLD_FONT = Font(bold=True, size=10)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
FLAG_FILL = {"Y": AMBER, "Review": GREEN}

REQUIRED_HEADER = ["s.no", "hsn heading (4-digit)", "chapter", "description", "category",
                    "flag as machine purchase? (y/n/review)", "match rule (use as hsn prefix match on gstr-2b)"]


def _safe_cell(ws, row, col, value):
    """Same guard as the other modules' own copy (kept local, not cross-imported, per this
    project's convention). Protects against a real, confirmed corruption class: a raw
    taxpayer-sourced value (a counterparty name here) that happens to start with '=' gets
    auto-detected by openpyxl as a formula, not text."""
    c = ws.cell(row, col, value)
    if c.data_type == "f" and isinstance(value, str):
        c.data_type = "s"
    return c

# ---------------------------------------------------------------------------
# Curated Trade-Name keyword map (see module docstring, point 3) -- hand-
# reviewed against every one of the master's 89 "Y" (Machine/Equipment)
# headings; only kept where the term is distinctive enough to be a genuine
# signal in a company's own trade name. ~36 "Y" headings were deliberately
# left WITHOUT a keyword (too generic/narrow/risky to match on a name) --
# those headings are still fully covered by the HSN-based EWB/GSTR-1 checks
# above, just not by this name-based layer.
# ---------------------------------------------------------------------------
KEYWORD_MAP = {
    "8401": "REACTOR", "8402": "BOILER", "8403": "BOILER", "8405": "GAS GENERATOR",
    "8406": "TURBINE", "8407": "ENGINE", "8408": "ENGINE", "8410": "TURBINE",
    "8411": "TURBINE", "8413": "PUMP", "8414": "COMPRESSOR",
    "8415": "AIR CONDITIONING", "8416": "BURNER", "8417": "FURNACE",
    "8418": "REFRIGERATION", "8419": "DRYER", "8420": "CALENDERING",
    "8421": "CENTRIFUGE", "8423": "WEIGHING MACHINE", "8425": "HOIST",
    "8426": "CRANE", "8427": "FORKLIFT", "8429": "BULLDOZER", "8430": "EXCAVATOR",
    "8433": "HARVESTER", "8435": "PRESS", "8443": "PRINTING MACHINE",
    "8445": "SPINNING MACHINE", "8446": "LOOM", "8447": "KNITTING MACHINE",
    "8452": "SEWING MACHINE", "8455": "ROLLING MILL", "8456": "MACHINE TOOL",
    "8457": "MACHINING CENTRE", "8458": "LATHE", "8459": "MACHINE TOOL",
    "8460": "GRINDING MACHINE", "8462": "PRESS", "8463": "MACHINE TOOL",
    "8464": "MACHINE TOOL", "8465": "WOODWORKING MACHINE", "8467": "POWER TOOL",
    "8468": "WELDING MACHINE", "8477": "PLASTIC MACHINERY",
    "8480": "MOULD", "8485": "3D PRINTING", "8501": "ELECTRIC MOTOR",
    "8502": "GENERATOR", "8504": "TRANSFORMER", "8514": "FURNACE",
    "8515": "WELDING MACHINE", "8535": "SWITCHGEAR", "8537": "CONTROL PANEL",
}
# REMOVED after testing against real data (confirmed, not assumed):
#   8474 CRUSHER -- 844 of 857 total keyword matches on one real taxpayer's data were
#        'XYZ STONE CRUSHER' suppliers, i.e. aggregate/crushed-stone TRADERS whose own
#        business name happens to include 'crusher' -- not sellers of crushing machinery
#        (HSN 8474). Massively overwhelmed the genuine signal from every other keyword
#        combined (13 matches). Still fully covered by the HSN-based EWB/GSTR-1 checks.
#   8412 MOTOR -- every real match was a vehicle-parts store ('XYZ MOTOR STORE',
#        'UTTRAKHAND OLD MOTOR PARTS') -- 'motor' in common Indian trade-name usage means
#        automobile, not the Ch.84/85 electric-motor/engine sense this master intends.


def is_machinery_master(path):
    """Content signature: ANY sheet in the workbook whose header row matches
    the 'Master HSN Flag List' 7-column header exactly (case/whitespace-
    insensitive). Never matched by filename or by assuming a sheet name,
    since this workbook legitimately has 4 sheets (per-chapter references +
    the actual master + a Read Me) and only one carries the match-ready list."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return False
    try:
        for sn in wb.sheetnames:
            ws = wb[sn]
            row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
            cleaned = [str(c).strip().lower() if c else "" for c in row[:7]]
            if cleaned == REQUIRED_HEADER:
                return True
        return False
    finally:
        wb.close()


def _find_master_sheet(wb):
    for sn in wb.sheetnames:
        ws = wb[sn]
        row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        cleaned = [str(c).strip().lower() if c else "" for c in row[:7]]
        if cleaned == REQUIRED_HEADER:
            return ws
    return None


def load_master(path):
    """Return list of dicts: heading (4-digit str), chapter, desc, category, flag."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = _find_master_sheet(wb)
    if ws is None:
        raise ValueError("No sheet in this workbook matches the expected "
                          "'Master HSN Flag List' 7-column header.")
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[1]:
            continue
        heading = str(r[1]).strip()
        if not heading:
            continue
        out.append(dict(heading=heading, chapter=str(r[2]).strip() if r[2] else "",
                         desc=str(r[3]).strip() if r[3] else "",
                         category=str(r[4]).strip() if r[4] else "",
                         flag=str(r[5]).strip() if r[5] else ""))
    return out


def _by_heading(master):
    return {m["heading"]: m for m in master}


def scan_gstr1_outward(g1_hsn_by_month, master):
    """GSTR-1's own HSN summary, real HSN codes, 4-digit-prefix match."""
    by_h = _by_heading(master)
    out = []
    for month, rows in (g1_hsn_by_month or {}).items():
        for row in rows:
            hsn = str(row.get("hsn") or "").strip()
            m = by_h.get(hsn[:4])
            if not m:
                continue
            out.append(dict(month=month, hsn=hsn, master_heading=m["heading"], desc=m["desc"],
                             category=m["category"], flag=m["flag"], source="GSTR-1 HSN Summary",
                             counterparty=None, counterparty_gstin=None, taxable=row.get("taxable", 0.0),
                             tax=(row.get("igst", 0.0) + row.get("cgst", 0.0)
                                  + row.get("sgst", 0.0) + row.get("cess", 0.0)),
                             ref=row.get("desc") or ""))
    return out


def scan_ewb(ewb_rows, master, self_gstin, direction):
    """direction: 'Inward' (from != self, to == self) or 'Outward' (from ==
    self, to != self) -- self-to-self movements are excluded in both cases
    (own-branch transfers, confirmed common in real data, not a purchase or
    a sale)."""
    by_h = _by_heading(master)
    out = []
    for r in ewb_rows or []:
        if direction == "Inward":
            if not (r.get("from_gstin") and r.get("from_gstin") != self_gstin
                    and r.get("to_gstin") == self_gstin):
                continue
            counterparty = r.get("from_name")
            counterparty_gstin = r.get("from_gstin")
        else:
            if not (r.get("to_gstin") and r.get("to_gstin") != self_gstin
                    and r.get("from_gstin") == self_gstin):
                continue
            counterparty = r.get("to_name")
            counterparty_gstin = r.get("to_gstin")
        hsn = str(r.get("hsn") or "").strip()
        m = by_h.get(hsn[:4])
        if not m:
            continue
        out.append(dict(month=r.get("month"), hsn=hsn, master_heading=m["heading"], desc=m["desc"],
                         category=m["category"], flag=m["flag"], source=f"E-Way Bill ({direction})",
                         counterparty=counterparty, counterparty_gstin=counterparty_gstin,
                         taxable=r.get("assess", 0.0), tax=r.get("taxval", 0.0),
                         ref=f"EWB {r.get('ewbno')}, Doc {r.get('docno')}"))
    return out


def _whole_phrase_pattern(keyword_upper):
    escaped = re.escape(keyword_upper)
    return re.compile(r"(?<![A-Z0-9])" + escaped + r"(?![A-Z0-9])")


def scan_names_inward(b2b_rows_by_month):
    """Trade-Name keyword screen against GSTR-2B's own invoice-level data
    (the only field 2B actually carries that's usable here -- see module
    docstring point 3). Uses KEYWORD_MAP, not the full master -- only the
    curated, distinctive subset."""
    patterns = [(h, kw, _whole_phrase_pattern(kw)) for h, kw in KEYWORD_MAP.items()]
    out = []
    for month, rows in (b2b_rows_by_month or {}).items():
        for row in rows:
            name = str(row.get("supplier") or "").upper()
            if not name:
                continue
            hits = [(h, kw) for h, kw, pat in patterns if pat.search(name)]
            if not hits:
                continue
            headings = sorted({h for h, _k in hits})
            kws = sorted({k for _h, k in hits})
            itc = (row.get("igst", 0.0) + row.get("cgst", 0.0)
                   + row.get("sgst", 0.0) + row.get("cess", 0.0))
            out.append(dict(month=month, hsn=None, master_heading=", ".join(headings),
                             desc=", ".join(kws), category="Trade-Name match", flag="Review",
                             source="Trade-Name keyword (GSTR-2B)", counterparty=row.get("supplier"),
                             counterparty_gstin=row.get("gstin"),
                             taxable=row.get("taxable", 0.0), tax=itc,
                             ref=f"Inv {row.get('invno')}, matched keyword(s): {', '.join(kws)}"))
    return out


DETAIL_HEADER = ["Month", "Direction", "Source", "HSN / Heading", "Description", "Category", "Flag",
                  "Counterparty", "Counterparty GSTIN", "Taxable / Assessable Value", "Tax", "Reference"]
_COL_WIDTHS = [10, 10, 22, 14, 42, 20, 9, 30, 18, 20, 14, 40]


def _row_to_excel(r, direction):
    return [r["month"], direction, r["source"], r["hsn"] or r["master_heading"], r["desc"], r["category"],
            r["flag"], r["counterparty"], r.get("counterparty_gstin"), r["taxable"], r["tax"], r["ref"]]


def write_sheet(wb, sale_rows, purchase_rows, master_path, notes):
    ws = wb.create_sheet("Machinery HSN Scan")
    ws.cell(1, 1, "MACHINERY HSN SCAN -- PURCHASE / SALE OF PLANT & MACHINERY (Ch.84 + Ch.85)").font = TITLEF
    sub = ("Screens GSTR-1, E-Way Bill (both directions), and GSTR-2B supplier names against a "
           "machinery HSN master list, to flag likely capital-goods purchases (possible manufacturing "
           "activity) and sales (possible Sec 18(6) consequence) for manual review. See the notes at "
           "the bottom for exactly what each source can and cannot detect.")
    c = ws.cell(2, 1, sub); c.font = Font(size=9, italic=True); c.alignment = Alignment(wrap_text=True, vertical="top")

    r = 4
    ws.cell(r, 1, "Summary").font = SECTF
    r += 1
    for i, h in enumerate(("Direction", "Source", "Count", "Taxable / Assessable Value (Rs)"), 1):
        cc = ws.cell(r, i, h); cc.fill = HEADFILL; cc.font = HEADER_FONT; cc.border = BORDER
    r += 1
    summary_rows = [
        ("Sale (Outward)", "GSTR-1 HSN Summary", [x for x in sale_rows if x["source"] == "GSTR-1 HSN Summary"]),
        ("Sale (Outward)", "E-Way Bill (Outward)", [x for x in sale_rows if x["source"] == "E-Way Bill (Outward)"]),
        ("Purchase (Inward)", "E-Way Bill (Inward)", [x for x in purchase_rows if x["source"] == "E-Way Bill (Inward)"]),
        ("Purchase (Inward)", "Trade-Name keyword (GSTR-2B)",
         [x for x in purchase_rows if x["source"] == "Trade-Name keyword (GSTR-2B)"]),
    ]
    for direction, source, items in summary_rows:
        ws.cell(r, 1, direction).border = BORDER
        ws.cell(r, 2, source).border = BORDER
        ws.cell(r, 3, len(items)).border = BORDER
        v = ws.cell(r, 4, round(sum(x["taxable"] for x in items), 2)); v.number_format = "#,##0.00"; v.border = BORDER
        r += 1
    r += 2

    ws.cell(r, 1, f"Sale (Outward) -- Complete Detail ({len(sale_rows)} row(s))").font = SECTF
    r += 1
    hdr = r
    for i, h in enumerate(DETAIL_HEADER, 1):
        cc = ws.cell(hdr, i, h); cc.fill = HEADFILL; cc.font = HEADER_FONT
        cc.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center"); cc.border = BORDER
    ws.row_dimensions[hdr].height = 28
    r = hdr + 1
    flag_col = DETAIL_HEADER.index("Flag") + 1
    for row in sorted(sale_rows, key=lambda x: (str(x["month"] or ""), x["source"])):
        for i, v in enumerate(_row_to_excel(row, "Outward / Sale"), 1):
            cc = _safe_cell(ws, r, i, v); cc.border = BORDER; cc.font = PLAIN_FONT
            if isinstance(v, float):
                cc.number_format = "#,##0.00"
        fill = FLAG_FILL.get(ws.cell(r, flag_col).value)
        if fill:
            ws.cell(r, flag_col).fill = fill; ws.cell(r, flag_col).font = BOLD_FONT
        r += 1
    if not sale_rows:
        ws.cell(r, 1, "No machinery-HSN sale detected.").font = Font(italic=True); r += 1
    r += 2

    ws.cell(r, 1, f"Purchase (Inward) -- Complete Detail ({len(purchase_rows)} row(s))").font = SECTF
    r += 1
    hdr2 = r
    for i, h in enumerate(DETAIL_HEADER, 1):
        cc = ws.cell(hdr2, i, h); cc.fill = HEADFILL; cc.font = HEADER_FONT
        cc.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center"); cc.border = BORDER
    ws.row_dimensions[hdr2].height = 28
    r = hdr2 + 1
    for row in sorted(purchase_rows, key=lambda x: (str(x["month"] or ""), x["source"])):
        for i, v in enumerate(_row_to_excel(row, "Inward / Purchase"), 1):
            cc = _safe_cell(ws, r, i, v); cc.border = BORDER; cc.font = PLAIN_FONT
            if isinstance(v, float):
                cc.number_format = "#,##0.00"
        fill = FLAG_FILL.get(ws.cell(r, flag_col).value)
        if fill:
            ws.cell(r, flag_col).fill = fill; ws.cell(r, flag_col).font = BOLD_FONT
        r += 1
    if not purchase_rows:
        ws.cell(r, 1, "No machinery-HSN purchase detected.").font = Font(italic=True); r += 1
    r += 1

    for n in notes:
        c = ws.cell(r, 1, "Note: " + n)
        c.font = Font(size=9, italic=True, color="808080")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    ws.auto_filter.ref = f"A{hdr2}:{get_column_letter(len(DETAIL_HEADER))}{max(r - len(notes) - 2, hdr2)}"
    ws.freeze_panes = ws.cell(hdr2 + 1, 1)
    for i, w in enumerate(_COL_WIDTHS[:len(DETAIL_HEADER)], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_skipped_sheet(wb, sheet_name, title_text, reason_text):
    """BUG FIX (bug report #4, 'Blocked Credit / Machinery HSN sheets missing from the final
    workbook') -- see gst_blocked_credit._write_skipped_sheet for the full explanation; same
    fix applied here. build_and_write() used to return its status STRING without ever calling
    write_sheet() when the master list couldn't be found/read, so the sheet never existed in
    the output workbook at all rather than being empty with a reason. Now the sheet always
    exists, explicit about why it has no data this run when it can't be built."""
    ws = wb.create_sheet(sheet_name)
    ws.cell(1, 1, title_text).font = TITLEF
    c = ws.cell(3, 1, "SKIPPED -- " + reason_text)
    c.font = Font(size=11, italic=True, color="9C0006")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 110
    ws.row_dimensions[3].height = 60
    return ws


def build_and_write(wb, master_path, g1_hsn_by_month, ewb_out_rows, ewb_in_rows,
                     b2b_rows_by_month, self_gstin):
    """Entry point called from master_build.py. Never raises -- degrades to
    a skipped sheet with a clear reason."""
    title = "MACHINERY HSN SCAN -- PURCHASE / SALE OF PLANT & MACHINERY (Ch.84 + Ch.85)"
    if not master_path:
        reason = ("no machinery HSN master file found in the input folder (content-detected: a "
                   "sheet with header 'S.No / HSN Heading (4-digit) / Chapter / Description / "
                   "Category / Flag.../ Match Rule...'). Add Machinery_HSN_Master_v2.xlsx (or "
                   "your own copy with the same header row) to the input folder and rerun to "
                   "enable this sheet.")
        _write_skipped_sheet(wb, "Machinery HSN Scan", title, reason)
        return "SKIPPED -- " + reason
    try:
        master = load_master(master_path)
    except Exception as e:
        reason = f"could not read the master file ({e!r})."
        _write_skipped_sheet(wb, "Machinery HSN Scan", title, reason)
        return "SKIPPED -- " + reason
    if not master:
        reason = f"master file {master_path!r} has no usable heading rows."
        _write_skipped_sheet(wb, "Machinery HSN Scan", title, reason)
        return "SKIPPED -- " + reason

    sale_rows = scan_gstr1_outward(g1_hsn_by_month, master)
    sale_rows += scan_ewb(ewb_out_rows, master, self_gstin, "Outward")
    purchase_rows = scan_ewb(ewb_in_rows, master, self_gstin, "Inward")
    purchase_rows += scan_names_inward(b2b_rows_by_month)

    n_y = sum(1 for m in master if m["flag"] == "Y")
    notes = [
        f"Master list: {len(master)} HSN headings ({n_y} flagged 'Y' -- core Machine/Equipment) "
        f"read from {master_path}.",
        "GSTR-1 HSN Summary and both E-Way Bill directions are matched on the ACTUAL HSN code "
        "(4-digit prefix) -- these are the reliable signals.",
        "GSTR-2A/2B carry NO HSN/SAC column at all (confirmed against real exports) -- an inward "
        "purchase genuinely cannot be HSN-matched from 2A/2B directly. This is the same gap "
        "documented in the 'HSN & Fraud Pattern Checks' sheet's finding A5 and in the 'Potential "
        "Blocked Credits' sheet. The 'E-Way Bill (Inward)' row above is the real substitute -- "
        "e-way bills DO carry HSN, confirmed against this taxpayer's own data.",
        "The 'Trade-Name keyword (GSTR-2B)' row is a SEPARATE, lower-confidence cross-check on top "
        "of the E-Way Bill (Inward) signal, not a replacement for it -- matches a supplier's "
        "Trade/Legal name against a curated ~51-term subset of the master's own HSN descriptions "
        "(see KEYWORD_MAP in gst_machinery_scan.py for the full list, including 2 terms removed "
        "after real-data testing showed them producing near-total false positives).",
        "Self-to-self e-way bills (same GSTIN both ends -- branch/stock transfers) are excluded "
        "from both E-Way Bill rows above; confirmed common in real data (most of this taxpayer's "
        "own outward e-way bills were exactly this) and not a genuine purchase or sale.",
        "Screening aid only -- a machinery-HSN purchase or sale is not itself a compliance issue; "
        "it's a prompt to check registration particulars (nature of business), ITC eligibility on "
        "capital goods, and Sec 18(6) on any disposal. Manual review required before any conclusion.",
    ]
    write_sheet(wb, sale_rows, purchase_rows, master_path, notes)
    return (f"OK -- {len(sale_rows)} sale row(s), {len(purchase_rows)} purchase row(s) "
            f"(includes both HSN-based and Trade-Name-based signals, see sheet).")
