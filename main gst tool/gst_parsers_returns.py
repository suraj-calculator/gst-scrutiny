#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GST PARSERS RETURNS
===================
CONSOLIDATED FILE -- contains what used to be: gst_scrutiny_tool.py, gstr2b_parser.py, ewb_annual_parser.py, amendments.py

The tool was reorganised from 19 .py files into 9 for easier sharing. Nothing
in the analytical logic was rewritten during that move: each section below is
the original module's code verbatim, with only (a) intra-project imports
repointed at the new file names, (b) its standalone __main__ demo block
removed, and (c) the renames listed under MERGE NOTES applied where two merged
modules happened to define the same top-level name with different bodies.

MERGE NOTES for this file:
  - gstr2b_parser._num -> _num_2b
  - ewb_annual_parser._num -> _num_ewb
  - amendments._num -> _num_amd
"""


# ============================================================================
# ==== SECTION: gst_scrutiny_tool.py  (was a standalone module before consolidation)
# ============================================================================
"""
GST SCRUTINY COMPARISON TOOL
============================
Plain, raw, side-by-side comparison of a single tax period across:
  GSTR-1  |  GSTR-2B  |  GSTR-3B  |  E-Invoice  |  Cash / Credit / Liability Ledgers

NO analysis. NO interpretation. NO safety nets.
Just: LEFT value | RIGHT value | DIFFERENCE | MATCH? (highlighted if mismatch)

Output: one Excel workbook. Every mismatch row shaded RED.
An "Exceptions" sheet on top lists ONLY the mismatched rows.

USAGE:
    Edit the CONFIG block below (file paths + period), then:
        python gst_scrutiny_tool.py
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import csv, re, sys
import gst_core as mpu

# ======================================================================
# CONFIG  --  MERGED FILES (whole FY, one workbook per document type)
# ======================================================================
# master_build.py is the single entry point. It classifies the run folder
# once (folder_classifier.py), then before each month's run it sets the
# globals below: the merged-workbook paths stay the SAME across every month
# (there is only one file per type); only PERIOD_LABEL changes, since that's
# what tells the parsers below which month's block to read out of each
# merged file.
#
# No auto-detection and no fallback values live in this module any more --
# if these are wrong, that must be visible immediately, not masked.

GSTR1_FILE   = None   # merged GSTR-1 workbook path (marker rows inside each sub-sheet)
GSTR3B_FILE  = None   # merged GSTR-3B workbook path (one sheet per month)
EINV_FILE    = None   # merged E-Invoice workbook path, or None if not supplied at all
GSTR2B_FILE  = None   # merged GSTR-2B workbook path (monthly or quarterly marker blocks)

SELF_GSTIN   = ""
COMPANY_NAME = ""

PERIOD_LABEL = None   # e.g. 'Jan-23' -- which month to read out of every merged file above
FY_4B2_BY_MONTH = {}   # {month_label: [IGST,CGST,SGST,CESS]} -- set once by master_build.py
                        # before its month loop starts; used by build_comparisons()'s D2 section
                        # for the month-over-month reversal-outlier check (see that section).
FY_RCM_BY_MONTH = {}    # {month_label: [3.1d IGST,CGST,SGST, 4A3 IGST]} -- same pre-pass timing
                         # as FY_4B2_BY_MONTH; used by the RCM section's trailing-average
                         # anomaly escalation (Bug 5 fix).
FY_2B_INVOICE_INDEX = {}   # {(supplier_gstin, invoice_or_note_no): [(month, row), ...]} across
                            # the WHOLE FY -- set once by master_build.py before its month loop;
                            # used by gst_checks_monthly.run()'s EWB-In check #12 to distinguish
                            # a genuine "supplier never filed" gap from an ordinary cross-month
                            # filing-timing difference (Bug 3 fix).

# Filing dates for analysis checks #8 (IRN-lag) and #10 (GSTR-1 vs 3B filing gap).
# Auto-extracted from the ARN date inside the files by the unified tool;
# leave None here (do NOT hardcode).
GSTR1_FILING_DATE  = None
GSTR3B_FILING_DATE = None

OUTPUT_FILE  = "GST_Scrutiny_Comparison.xlsx"

TOLERANCE    = 1.0   # rupee tolerance; abs(diff) <= this  => treated as MATCH

def get_gstr2b_values():
    """Return the GSTR-2B summary dict for PERIOD_LABEL, read straight out of
    the merged GSTR-2B workbook. Requires GSTR2B_FILE + PERIOD_LABEL to be
    set. No hardcoded/zero fallback: if 2B isn't available for this month,
    that must stop the run for this month, not silently zero out the ITC
    comparison rows."""
    import gst_parsers_returns as _g2b
    return _g2b.summary_for_month(GSTR2B_FILE, PERIOD_LABEL)

# ======================================================================
# HELPERS
# ======================================================================
def num(v):
    """Convert any cell/string to float. '-', '', None -> 0.0 . Strips commas."""
    if v is None: return 0.0
    s = str(v).strip()
    if s in ("", "-", "–"): return 0.0
    s = s.replace(",", "").replace("₹", "").strip()
    try: return float(s)
    except: return 0.0

def load_xlsx(path):
    return openpyxl.load_workbook(path, data_only=True)

def read_csv_rows(path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        return list(csv.reader(f))

# ======================================================================
# PARSERS  -- each returns the numbers this tool needs from one source
# ======================================================================

_HSN_ALL_MONTHS_CACHE = {}   # path -> {month: [row-dict,...]} -- avoids re-scanning the small
                             # hsn/hsn(b2b)/hsn(b2c) tabs once per month per caller (parse_gstr1
                             # runs once per month; the HSN Detail sheet and the review table
                             # both also need this same data).


def read_gstr1_hsn_all_months(path):
    """Table 12 (HSN-wise summary of outward supplies), ALL months in ONE pass.

    BUG FIX (confirmed against a real taxpayer's export, GSTIN 05AACFT2702L1ZD): two tab formats
    exist across a single FY's merged workbook --
      OLD: a single 'hsn' tab (per-product table). Confirmed: carries a marker for April ONLY.
      NEW: 'hsn(b2b)' + 'hsn(b2c)' tabs (GSTN split Table 12 by customer type at some point).
           Confirmed: carry markers for May-25 onward, IDENTICAL column layout to the old tab
           (HSN, Description, UQC, Total Quantity, Rate, Taxable Value, Integrated Tax, Central
           Tax, State/UT Tax, Cess Amount) -- so the same reader works for both, only the sheet
           name(s) consulted differ.
    Every caller that previously read ONLY the 'hsn' tab (parse_gstr1()'s own hsn_* accumulator,
    and gst_checks_hsn_fraud._hsn_rows_by_month()) saw HSN=0.00 for every month from May-25
    onward -- not because there was no HSN data, but because it was sitting in a tab this
    function never looked at. That silently produced a 100% gap on every HSN-vs-named-invoice
    comparison (Comparison 'A'/'A2', Analysis(14) '#13') every month from May onward.

    Resolution is PER MONTH, not a single hardcoded tab choice for the whole file: 'hsn' is tried
    first for a given month; only if it has no marker for that month does this fall back to
    summing 'hsn(b2b)' + 'hsn(b2c)' for that month. Nothing rules out a merged workbook mixing
    formats within its own FY the way this one does, so no format is assumed in advance.

    Returns {month_label: [dict(hsn, desc, uqc, qty, rate, taxable, igst, cgst, sgst, cess,
                                  source_tab), ...]}. A month genuinely absent from the returned
    dict means NEITHER format had a marker for it anywhere in the file -- callers must treat
    that as an explicit gap, never silently as a zero total."""
    if path in _HSN_ALL_MONTHS_CACHE:
        return _HSN_ALL_MONTHS_CACHE[path]
    wb = openpyxl.load_workbook(path, data_only=True)

    def _read_sheet_month(sheet_name, month, source_tag):
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        hdr = [str(c).strip() if c else "" for c in rows[3]]
        H = {h: i for i, h in enumerate(hdr)}
        recs = []
        for r in mpu.rows_for_month(rows, 3, month):   # raises PeriodParseError if no marker
            if not any(r):
                continue
            recs.append(dict(
                hsn=str(r[H["HSN"]] or "").strip() if "HSN" in H and H["HSN"] < len(r) else "",
                desc=str(r[H["Description"]] or "").strip() if "Description" in H and H["Description"] < len(r) else "",
                uqc=str(r[H["UQC"]] or "").strip() if "UQC" in H and H["UQC"] < len(r) else "",
                qty=num(r[H["Total Quantity"]]) if "Total Quantity" in H and H["Total Quantity"] < len(r) else 0.0,
                rate=num(r[H["Rate"]]) if "Rate" in H and H["Rate"] < len(r) else 0.0,
                taxable=num(r[H["Taxable Value"]]) if "Taxable Value" in H and H["Taxable Value"] < len(r) else 0.0,
                igst=num(r[H["Integrated Tax"]]) if "Integrated Tax" in H and H["Integrated Tax"] < len(r) else 0.0,
                cgst=num(r[H["Central Tax"]]) if "Central Tax" in H and H["Central Tax"] < len(r) else 0.0,
                sgst=num(r[H["State/UT Tax"]]) if "State/UT Tax" in H and H["State/UT Tax"] < len(r) else 0.0,
                cess=num(r[H["Cess Amount"]]) if "Cess Amount" in H and H["Cess Amount"] < len(r) else 0.0,
                source_tab=source_tag,
            ))
        return recs

    # Union of every month either format declares a marker for, across the whole file.
    all_months = set()
    for sn in ("hsn", "hsn(b2b)", "hsn(b2c)"):
        if sn not in wb.sheetnames:
            continue
        for row in wb[sn].iter_rows(values_only=True):
            if row and row[0] and isinstance(row[0], str) and row[0].startswith("Financial Year:"):
                try:
                    _, _, labels = mpu.parse_marker_text(row[0])
                    all_months.update(labels)
                except mpu.PeriodParseError:
                    continue

    out = {}
    for month in all_months:
        recs, used_old = [], False
        if "hsn" in wb.sheetnames:
            try:
                recs = _read_sheet_month("hsn", month, "hsn")
                used_old = True
            except mpu.PeriodParseError:
                used_old = False
        if not used_old:
            for sn, tag in (("hsn(b2b)", "hsn(b2b)"), ("hsn(b2c)", "hsn(b2c)")):
                if sn not in wb.sheetnames:
                    continue
                try:
                    recs.extend(_read_sheet_month(sn, month, tag))
                except mpu.PeriodParseError:
                    continue
        out[month] = recs
    _HSN_ALL_MONTHS_CACHE[path] = out
    return out


def parse_gstr1(path, month):
    """Sum outward tax & taxable value from GSTR-1 sub-sheets, for ONE month
    out of the merged (whole-FY) workbook. `month` e.g. 'Jan-23'. Every
    sub-sheet in the merged file carries its own period-marker rows (see
    merged_period_utils.py); this reads ONLY the block matching `month` --
    it raises if that month has no marker at all in a given sub-sheet (that
    sub-sheet is simply skipped for scoring only when the SHEET itself is
    entirely absent from the workbook, not when the month is missing from it)."""
    wb = load_xlsx(path)
    out = {"taxable":0.0,"IGST":0.0,"CGST":0.0,"SGST":0.0,"CESS":0.0,
           "b2b_count":0,"b2b_no_irn":0,"lines":{},"blank_invno_lines":0,"blank_invno_taxable":0.0,"named_taxable":0.0,"named_IGST":0.0,"named_CGST":0.0,"named_SGST":0.0,
           "cn_taxable":0.0,"cn_IGST":0.0,"cn_CGST":0.0,"cn_SGST":0.0,"cn_CESS":0.0,
           "hsn_IGST":0.0,"hsn_CGST":0.0,"hsn_SGST":0.0,"hsn_CESS":0.0,"hsn_taxable":0.0,
           "b2cs_taxable":0.0,"b2cs_IGST":0.0,"b2cs_CGST":0.0,"b2cs_SGST":0.0,"b2cs_CESS":0.0,
           "nil_taxable":0.0,"exempt_taxable":0.0,"nongst_taxable":0.0,"nil_exempt_taxable":None,
           # BUG FIX (confirmed on a real taxpayer's export): a sub-sheet's marker set can
           # legitimately be INCOMPLETE across the FY (confirmed case: the 'hsn' -- Table 12 --
           # sub-sheet carried a marker for April only, out of 12 months, on a real GSTR-1 export;
           # every other sub-sheet had all 12). mpu.rows_for_month() correctly raises
           # PeriodParseError when a month has no marker at all in a given sub-sheet (that is NOT
           # the same as the month having zero data rows, which is legitimate). Previously that
           # exception was uncaught here, so ONE sub-sheet missing ONE month's marker killed this
           # entire month's GSTR-1 parse -- and by extension every other check for that month too,
           # since parse_gstr1() feeds the whole per-month pipeline. Each sub-sheet block below is
           # now individually guarded; a genuinely-missing month-marker on a sub-sheet degrades
           # that sub-sheet's contribution to 0 for this month (matching its existing zero-default)
           # AND is recorded here by name, so it reads as an explicit gap, never a silent verified
           # zero.
           "month_marker_gaps": []}

    # --- b2b ---
    if "b2b, sez, de_inv" in wb.sheetnames:
      try:
        ws=wb["b2b, sez, de_inv"]
        rows=list(ws.iter_rows(values_only=True))
        hdr=[str(c).strip() if c else "" for c in rows[3]]
        H={h:i for i,h in enumerate(hdr)}
        c_gstin, c_invno = H.get("GSTIN/UIN of Recipient"), H.get("Invoice Number")
        # BUG FIX -- found on real data while chasing a reported false "gap" (invoice
        # CNN24A000001 appeared to be missing its 18% rate-line from GSTR-1, and the same
        # pattern independently reported again for ACC24A0007's 12%/18% lines). ROOT CAUSE:
        # the GST portal's own Excel export for a multi-rate-line invoice leaves GSTIN,
        # Receiver Name, Invoice Number and Invoice date BLANK on every rate-line after the
        # first (a merged cell in the source spreadsheet, visually spanning all of that
        # invoice's rows) -- only Rate/Taxable/tax are populated on those continuation rows.
        # This loop used to read each row's OWN (blank) Invoice Number/GSTIN cell in
        # isolation, so every continuation row silently fell into the "blank invoice number"
        # bucket -- completely disconnected from the invoice it actually belongs to, and
        # therefore invisible everywhere invoice number mattered (the B2 E-Invoice
        # comparison above all). Confirmed on this taxpayer's real April data: 5,836 of
        # 21,493 b2b rows across the FY (27%) are continuation rows of this shape. Fixed by
        # forward-filling the last SEEN gstin/invno whenever a row's own cells are blank but
        # its Rate/Taxable cells are populated -- reset to None at the top of every call
        # (i.e. every month), so this can never carry an invoice identity across a month
        # boundary.
        last_gstin, last_invno = None, None
        for r in mpu.rows_for_month(rows, 3, month):
            if not any(r): continue
            out["taxable"]+=num(r[H.get("Taxable Value")])
            out["IGST"]+=num(r[H.get("Integrated Tax")])
            out["CGST"]+=num(r[H.get("Central Tax")])
            out["SGST"]+=num(r[H.get("State/UT Tax")])
            out["CESS"]+=num(r[H.get("Cess Amount")])
            irn_i=H.get("IRN")
            if irn_i is None or not str(r[irn_i] if irn_i<len(r) else "").strip():
                out["b2b_no_irn"]+=1
            raw_gstin = r[c_gstin] if c_gstin is not None and c_gstin < len(r) else None
            raw_invno = r[c_invno] if c_invno is not None and c_invno < len(r) else None
            if raw_gstin not in (None, "") and raw_invno not in (None, ""):
                # A primary row -- this is where a "new invoice" is actually counted (a
                # multi-rate invoice's continuation rows are the SAME invoice, not additional
                # ones; b2b_count is meant to be an invoice count, not a rate-line count).
                last_gstin = str(raw_gstin).strip()
                last_invno = str(raw_invno).strip()
                gstin_key, invno = last_gstin, last_invno
                out["b2b_count"]+=1
            elif last_gstin is not None:
                # A continuation row of the immediately preceding primary row.
                gstin_key, invno = last_gstin, last_invno
            else:
                # No primary row has been seen yet this month -- genuinely can't attribute
                # this row to an invoice (shouldn't happen in a real export; never crash).
                gstin_key, invno = "", "None"
            # line-level key for reconciliation
            # BUG FIX (found during verification of the invoice-level aggregation fix above):
            # confirmed on real data that the SAME invoice number can legitimately be reused by
            # this taxpayer's counterparty across TWO DIFFERENT GST registrations of that same
            # company (e.g. invoice 'CNN24A002674' filed once against the counterparty's
            # Uttarakhand GSTIN in GSTR-1 and, separately, a different transaction under the same
            # invoice number against the counterparty's Maharashtra GSTIN in the e-invoice data --
            # two unrelated transactions that happen to share a number). Keying by invoice number
            # alone would silently cross-match those two unrelated transactions against each
            # other. GSTIN is now part of the key; invno stays element [0] so every existing
            # consumer of these keys elsewhere in this codebase (which reads only k[0]) is
            # unaffected.
            rate=num(r[H.get("Rate")]) if H.get("Rate") is not None else 0.0
            k=(invno,gstin_key,rate)
            L=out["lines"].setdefault(k,[0.0,0.0])
            L[0]+=num(r[H.get("Taxable Value")]); L[1]+=num(r[H.get("Integrated Tax")])
            if not invno or invno.lower()=="none":
                out["blank_invno_lines"]+=1
                out["blank_invno_taxable"]+=num(r[H.get("Taxable Value")])
            else:
                out["named_taxable"]+=num(r[H.get("Taxable Value")])
                out["named_IGST"]+=num(r[H.get("Integrated Tax")])
                out["named_CGST"]+=num(r[H.get("Central Tax")])
                out["named_SGST"]+=num(r[H.get("State/UT Tax")])
      except mpu.PeriodParseError:
        out["month_marker_gaps"].append("b2b, sez, de_inv")

    # --- b2cl (inter-state large, IGST only) ---
    for sn,cols in [("b2cl",("Taxable Value","Integrated Tax",None,None,"Cess Amount")),
                    ("exp",("Taxable Value","Integrated Tax",None,None,"Cess Amount"))]:
        if sn in wb.sheetnames:
          try:
            rows=list(wb[sn].iter_rows(values_only=True))
            hdr=[str(c).strip() if c else "" for c in rows[3]]
            H={h:i for i,h in enumerate(hdr)}
            for r in mpu.rows_for_month(rows, 3, month):
                if not any(r): continue
                out["taxable"]+=num(r[H.get("Taxable Value")]) if "Taxable Value" in H else 0
                out["IGST"]+=num(r[H.get("Integrated Tax")]) if "Integrated Tax" in H else 0
                out["CESS"]+=num(r[H.get("Cess Amount")]) if "Cess Amount" in H else 0
          except mpu.PeriodParseError:
            out["month_marker_gaps"].append(sn)

    # --- b2cs (intra/inter small) ---
    # ALSO tracked separately (b2cs_*, distinct from the combined out["taxable"]/out["IGST"]
    # etc. above) because B2CS invoices structurally carry no invoice number anywhere in a
    # GSTR-1 export (Table 7 is a pure state+rate aggregate, by design -- confirmed against the
    # real sheet, whose only columns are Place Of Supply/Rate/Taxable Value/tax heads, no
    # invoice-number field at all). That means B2CS value can never appear in named_taxable/
    # named_IGST (b2b-only), which produces a real, EXPLAINABLE (not fixable-by-matching)
    # residual on any check that compares "named invoices" against a broader total that DOES
    # include B2CS, such as Analysis(14) #13's HSN-vs-named-invoice gap. Tracked here so that
    # check can net it out with a visible, sourced explanation instead of leaving an
    # unexplained residual.
    if "b2cs" in wb.sheetnames:
      try:
        rows=list(wb["b2cs"].iter_rows(values_only=True))
        hdr=[str(c).strip() if c else "" for c in rows[3]]
        H={h:i for i,h in enumerate(hdr)}
        for r in mpu.rows_for_month(rows, 3, month):
            if not any(r): continue
            t=num(r[H.get("Taxable Value")]) if "Taxable Value" in H else 0
            i_=num(r[H.get("Integrated Tax")]) if "Integrated Tax" in H else 0
            c_=num(r[H.get("Central Tax")]) if "Central Tax" in H else 0
            s_=num(r[H.get("State/UT Tax")]) if "State/UT Tax" in H else 0
            ce=num(r[H.get("Cess Amount")]) if "Cess Amount" in H else 0
            out["taxable"]+=t; out["IGST"]+=i_; out["CGST"]+=c_; out["SGST"]+=s_; out["CESS"]+=ce
            out["b2cs_taxable"]+=t; out["b2cs_IGST"]+=i_; out["b2cs_CGST"]+=c_
            out["b2cs_SGST"]+=s_; out["b2cs_CESS"]+=ce
      except mpu.PeriodParseError:
        out["month_marker_gaps"].append("b2cs")

    # --- credit notes (cdnr + cdnur) ---
    for sn in ("cdnr","cdnur"):
        if sn in wb.sheetnames:
          try:
            rows=list(wb[sn].iter_rows(values_only=True))
            hdr=[str(c).strip() if c else "" for c in rows[3]]
            H={h:i for i,h in enumerate(hdr)}
            for r in mpu.rows_for_month(rows, 3, month):
                if not any(r): continue
                out["cn_taxable"]+=num(r[H.get("Taxable Value")]) if "Taxable Value" in H else 0
                out["cn_IGST"]+=num(r[H.get("Integrated Tax")]) if "Integrated Tax" in H else 0
                out["cn_CGST"]+=num(r[H.get("Central Tax")]) if "Central Tax" in H else 0
                out["cn_SGST"]+=num(r[H.get("State/UT Tax")]) if "State/UT Tax" in H else 0
                out["cn_CESS"]+=num(r[H.get("Cess Amount")]) if "Cess Amount" in H else 0
          except mpu.PeriodParseError:
            out["month_marker_gaps"].append(sn)

    # --- HSN summary (internal cross-check of GSTR-1) ---
    # Reads via read_gstr1_hsn_all_months() -- see that function's docstring for why this can no
    # longer just read the 'hsn' tab directly (two tab formats exist across a real FY; this
    # picks the right one PER MONTH). A month genuinely absent from BOTH formats is recorded in
    # month_marker_gaps, same convention as every other sub-sheet block in this function.
    hsn_all = read_gstr1_hsn_all_months(path)
    if month in hsn_all:
        for r in hsn_all[month]:
            out["hsn_taxable"] += r["taxable"]
            out["hsn_IGST"] += r["igst"]
            out["hsn_CGST"] += r["cgst"]
            out["hsn_SGST"] += r["sgst"]
            out["hsn_CESS"] += r["cess"]
    else:
        out["month_marker_gaps"].append("hsn")

    # --- Table 8: Nil rated / Exempted / Non-GST outward supplies (sheet 'exemp') ---
    # Header row is found dynamically (its index varies), then scoped to `month`
    # the same way as every other sub-sheet above.
    if "exemp" in wb.sheetnames:
        rows=list(wb["exemp"].iter_rows(values_only=True))
        hdr=None; hi=None
        for i,r in enumerate(rows):
            j=" ".join(str(c) for c in r if c is not None)
            if "Nil Rated" in j and "Non-GST" in j:
                hdr=[str(c).strip() if c else "" for c in r]; hi=i; break
        if hdr is not None:
          try:
            H={h:i for i,h in enumerate(hdr)}
            def col(*names):
                for n in names:
                    for h,idx in H.items():
                        if n.lower() in h.lower():
                            return idx
                return None
            i_nil=col("Nil Rated"); i_exm=col("Exempted"); i_non=col("Non-GST")
            for r in mpu.rows_for_month(rows, hi, month):
                if not any(c not in (None,"") for c in r): continue
                if i_nil is not None and i_nil<len(r): out["nil_taxable"]+=num(r[i_nil])
                if i_exm is not None and i_exm<len(r): out["exempt_taxable"]+=num(r[i_exm])
                if i_non is not None and i_non<len(r): out["nongst_taxable"]+=num(r[i_non])
            # header present => figures are known (0 if no rows). Set the combined nil+exempt
            # (matches 3B 3.1(b zero-rated is reported separately; 3.1(c) = nil+exempt).
            out["nil_exempt_taxable"]=out["nil_taxable"]+out["exempt_taxable"]
          except mpu.PeriodParseError:
            out["month_marker_gaps"].append("exemp")
    return out


def _gstr3b_sheet_month(ws):
    """Read a GSTR-3B sheet's own 'Year' + 'Tax Period' key/value rows and
    return the 'Mon-YY' label they represent. Content-based only -- the
    sheet's NAME (e.g. 'Jan_2022-23') is never trusted, per instruction."""
    fy = tp = None
    for row in ws.iter_rows(values_only=True):
        cells = [str(c).strip() for c in row if c not in (None, "")]
        if not cells:
            continue
        key = cells[0].upper()
        if key in ("YEAR", "FINANCIAL YEAR") and len(cells) >= 2:
            fy = cells[1]
        elif key == "TAX PERIOD" and len(cells) >= 2:
            tp = cells[1]
        if fy and tp:
            break
    if not (fy and tp):
        raise mpu.PeriodParseError(
            f"Sheet {ws.title!r} has no readable 'Year'/'Tax Period' fields.")
    labels = mpu.months_for_tax_period(fy, tp)
    return labels[0]  # GSTR-3B's Tax Period is always a single month, never a quarter


def parse_gstr3b(path, month):
    """Pull Table 3.1 and Table 4 values from GSTR-3B for ONE month, out of
    the merged workbook (one sheet per month). The sheet is located by its
    OWN in-sheet 'Year'/'Tax Period' content, not by its sheet name."""
    wb=load_xlsx(path)
    ws = None
    months_found = []
    for sn in wb.sheetnames:
        candidate = wb[sn]
        try:
            m = _gstr3b_sheet_month(candidate)
        except mpu.PeriodParseError:
            continue  # not a GSTR-3B data sheet (e.g. a stray blank/help sheet)
        months_found.append(m)
        if m == month:
            ws = candidate
            break
    if ws is None:
        raise mpu.PeriodParseError(
            f"Month {month!r} not found as a GSTR-3B sheet in {path!r}. "
            f"Months present: {sorted(months_found)}")
    rows=[[c.value for c in r] for r in ws.iter_rows()]
    g={}
    def find(label):
        for r in rows:
            joined=" ".join(str(c) for c in r if c is not None)
            if label in joined:
                vals=[num(c) for c in r if isinstance(c,(int,float)) or (isinstance(c,str) and re.match(r'^-?[\d,\.]+$',str(c).strip()))]
                return r
        return None
    def vals_after(rowlist, n=5):
        nums=[num(c) for c in rowlist if (isinstance(c,(int,float)) or (isinstance(c,str) and re.match(r'^-?[\d,\.]+$',str(c).replace(',','').strip())))]
        return nums

    # ---- Table 4(B) ITC Reversed: anchor to the section boundary ----
    # BUG FIX (confirmed against the real file): the literal label "(2) Others"
    # appears TWICE in every Apr/May/Jun/Jul-22 sheet -- once under "B. ITC
    # Reversed" (the real 4B(2) figure) and once under "(D) Ineligibe ITC"
    # (always 0, a completely different field). A loose "if '(2) Others' in j"
    # scan over every row (no anchoring) picks whichever occurs LAST, which is
    # always the D-section zero -- silently zeroing out 4B(2) for those 4
    # months. From Aug-22 onward the GSTR-3B Table 4 format itself changed
    # (Circular 170/02/2022-GST): the old "(D) Ineligibe ITC -> (1) As per
    # section 17(5)" row is gone, replaced by "(D) Other Details" with
    # unrelated sub-items -- so there's only one "(2) Others" match there and
    # the old loose scan happened to still work for those months by luck, not
    # by design. Fixed properly here: find the "B. ITC Reversed" header row
    # and the "C. Net ITC available" header row, then match "(1)"/"(2) Others"
    # ONLY within that bounded slice -- safe regardless of which Table-4
    # format the sheet uses, and regardless of row order in the file.
    b_start = b_end = None
    for i, r in enumerate(rows):
        j = " ".join(str(c) for c in r if c is not None).strip()
        if j.startswith("B. ITC Reversed"):
            b_start = i
        elif j.startswith("C. Net ITC available") and b_start is not None:
            b_end = i
            break
    if b_start is not None and b_end is not None:
        for r in rows[b_start + 1:b_end]:
            j = " ".join(str(c) for c in r if c is not None).strip()
            nums = vals_after(r)
            if not j or not nums:
                continue
            if j.startswith("(2) Others"):
                g["4B2"] = nums
            elif j.startswith("(1)"):
                g["4B1"] = nums   # Rules 42/43 (Apr-Jul) or Rules 38/42/43 + Sec 17(5) (Aug onward)
    g.setdefault("4B1", [0.0, 0.0, 0.0, 0.0])
    g.setdefault("4B2", [0.0, 0.0, 0.0, 0.0])

    for r in rows:
        j=" ".join(str(c) for c in r if c is not None)
        nums=vals_after(r)
        # 3.1(a) outward taxable
        if "Outward Taxable" in j and "other than zero" in j:
            g["3.1a"]=nums  # [taxable, IGST, CGST, SGST, CESS]
        # 3.1(b) zero-rated, 3.1(c) nil/exempt, 3.1(e) non-GST.
        # Guard against the 3.1(a) line (which also contains 'zero rated, nil rated and exempted'):
        # match the bracketed sub-label, not the words inside 3.1(a)'s 'other than ...' clause.
        if ("(b)" in j or "(zero rated )" in j) and "other than zero" not in j and "zero rated" in j.lower():
            g["3.1b"]=nums  # [taxable, IGST, ?, ?, CESS]
        if "(c)" in j and "Nil rated" in j and "Other Outward" in j:
            g["3.1c"]=nums  # [taxable]
        if "(e)" in j and "Non-GST" in j:
            g["3.1e"]=nums  # [taxable]
        if "Inward supplies (liable to reverse charge)" in j:
            g["3.1d"]=nums
        if "(5) All other ITC" in j:
            g["4A5"]=nums   # [IGST,CGST,SGST,CESS]
        if "(3) Inward supplies liable to reverse charge" in j:
            g["4A3"]=nums
        # NEW (added for the GSTR-2A ISD cross-check, G8): 4A(4) was not previously extracted
        # at all. Exact label confirmed against a real GSTR-3B sheet: "(4) Inward supplies from
        # ISD". Same anchor-free match style as 4A(3)/4A(5) above -- this label is unique in the
        # sheet (unlike the "(2) Others" ambiguity fixed elsewhere in this codebase), so no
        # section-boundary anchoring is needed here.
        if "(4) Inward supplies from ISD" in j:
            g["4A4"]=nums   # [IGST,CGST,SGST,CESS]
        if "Net ITC available" in j:
            g["4C"]=nums
    return g


def parse_einv(path, month):
    """E-Invoice file totals (B2B), for ONE month out of the merged workbook.
    E-Invoice is legitimately OPTIONAL, at both the whole-file level (some
    taxpayers/periods genuinely have none) AND the per-month level within an
    existing file (e-invoicing coverage can start partway through a file).
    Both cases produce the SAME clearly-surfaced 'not available' state
    (available=False) -- callers already branch on this explicitly, so it is
    not hidden, just not a hard stop for what is a documented PARTIAL source."""
    import os
    out={"taxable":0.0,"IGST":0.0,"CGST":0.0,"SGST":0.0,"CESS":0.0,"count":0,"errors":0,"available":True,"lines":{},
         "cancel_col_found":False,"cancel_date_col_found":False,"cancelled":[]}
    if not path or not os.path.exists(path):
        print(f"[info] E-Invoice file not supplied -> EINV checks skipped for {month}")
        out["available"]=False; return out
    wb=load_xlsx(path)
    if "b2b, sez, de" in wb.sheetnames:
        rows=list(wb["b2b, sez, de"].iter_rows(values_only=True))
        if month not in mpu.months_present(rows, 3):
            print(f"[info] E-Invoice file does not cover {month} -> EINV checks skipped for {month}")
            out["available"]=False; return out
        hdr=[str(c).strip() if c else "" for c in rows[3]]
        H={h:i for i,h in enumerate(hdr)}
        # Cancelled-e-invoice detection: try every known real-world header variant for the
        # IRN status / cancel-date columns (GSTN's own e-invoice/GSTR-1-auto-populate export
        # has used different header text across portal versions). CASE-INSENSITIVE match (fixed:
        # a real export's exact header was 'E-invoice status' -- lowercase 'i'/'s' -- which the
        # original case-sensitive exact-dict-key lookup never matched, even though "E-Invoice
        # Status" was already in the candidate list). Content-based, never a fixed column index.
        # If NONE of these are found, cancel_col_found stays False and the 'Cancelled E-Invoices'
        # sheet says so explicitly rather than reporting zero cancellations as if verified.
        H_LOWER = {h.strip().lower(): i for h, i in H.items()}
        STATUS_HDRS = ["irn status", "status", "e-invoice status", "einvoice status",
                       "cancel status", "invoice status"]
        CANCELDATE_HDRS = ["cancel date", "irn cancel date", "cancelled date",
                           "date of cancellation", "cancellation date"]
        status_col = next((H_LOWER[h] for h in STATUS_HDRS if h in H_LOWER), None)
        canceldate_col = next((H_LOWER[h] for h in CANCELDATE_HDRS if h in H_LOWER), None)
        out["cancel_col_found"] = status_col is not None
        out["cancel_date_col_found"] = canceldate_col is not None
        for r in mpu.rows_for_month(rows, 3, month):
            if not any(r): continue
            # BUG FIX (found via a real taxpayer's monthly Comparison sheets showing FALSE
            # 'line-level gap' mismatches for invoices that genuinely exist in both GSTR-1 and
            # E-Invoice -- confirmed real, not assumed): this looked up H.get("Invoice number")
            # (lowercase 'n'), but this file's actual header is "Invoice Number" (capital N) --
            # confirmed directly against the raw file. Since the lookup never matched, invno fell
            # back to the LITERAL STRING "None" for EVERY row in EVERY month, not just one
            # invoice -- meaning the entire 'lines' dict this function builds was keyed under
            # ("None", rate) instead of real invoice numbers, and every single line-level
            # comparison in section B2 was comparing against nothing. read_einv_lines (a separate
            # function, used elsewhere) already tries both capitalisations as fallback candidates
            # and was unaffected -- this function's own single-candidate lookup was the one place
            # this broke. Fixed the same way: try both capitalisations, not just one.
            invno_col = H.get("Invoice Number", H.get("Invoice number"))
            invno=str(r[invno_col]).strip() if invno_col is not None else "None"
            rate=num(r[H.get("Rate")]) if H.get("Rate") is not None else 0.0
            gstin_col = H.get("GSTIN/UIN of Recipient", H.get("GSTIN of Recipient"))
            gstin_key = str(r[gstin_col]).strip() if gstin_col is not None else ""

            # FIX (was Bug 1's second half): check cancellation status FIRST. A cancelled
            # e-invoice is correctly ABSENT from GSTR-1 (GSTR-1's own auto-population/deletion
            # status marks it Deleted), so including it in the totals/line-map used for every
            # E-Invoice-vs-GSTR-1 comparison manufactures a false gap on every such invoice --
            # confirmed on the real file: 6 cancelled invoices totalling Rs 44,12,291 taxable
            # were previously producing 6 false "LINE-LEVEL GAP" mismatches, one per month.
            # Cancelled rows are recorded in out['cancelled'] for the Cancelled-E-Invoices sheet
            # and cross-checks, but do NOT contribute to taxable/IGST/CGST/SGST/CESS/count/lines.
            is_cancelled = False
            if status_col is not None and status_col < len(r):
                status_val = str(r[status_col] or "").strip().upper()
                if status_val in ("CANCELLED", "CANCEL", "CANCELED"):
                    is_cancelled = True
                    out["cancelled"].append(dict(
                        invno=invno, rate=rate,
                        gstin=gstin_key,
                        invdate=str(r[H.get("Invoice date")] or "").strip() if H.get("Invoice date") is not None else "",
                        taxable=num(r[H.get("Taxable Value")]) if "Taxable Value" in H else 0.0,
                        igst=num(r[H.get("Integrated Tax")]) if "Integrated Tax" in H else 0.0,
                        # BUG FIX (reported): only IGST was ever captured here -- CGST/SGST were
                        # silently dropped from this sheet even though they're read into the
                        # main totals two lines below. Added, same source columns, same 'or 0'
                        # graceful-missing-column handling as every other field here.
                        cgst=num(r[H.get("Central Tax")]) if "Central Tax" in H else 0.0,
                        sgst=num(r[H.get("State/UT Tax")]) if "State/UT Tax" in H else 0.0,
                        cess=num(r[H.get("Cess Amount")]) if "Cess Amount" in H else 0.0,
                        irn=str(r[H.get("IRN")] or "").strip() if H.get("IRN") is not None else "",
                        cancel_date=(str(r[canceldate_col]).strip()
                                     if canceldate_col is not None and canceldate_col < len(r) else None),
                        month=month,
                    ))
            if is_cancelled:
                continue

            out["taxable"]+=num(r[H.get("Taxable Value")]) if "Taxable Value" in H else 0
            out["IGST"]+=num(r[H.get("Integrated Tax")]) if "Integrated Tax" in H else 0
            out["CGST"]+=num(r[H.get("Central Tax")]) if "Central Tax" in H else 0
            out["SGST"]+=num(r[H.get("State/UT Tax")]) if "State/UT Tax" in H else 0
            out["CESS"]+=num(r[H.get("Cess Amount")]) if "Cess Amount" in H else 0
            out["count"]+=1
            # BUG FIX -- see the matching comment in parse_gstr1's b2b loop above: GSTIN is now
            # part of this key too (invno stays element [0] for backward compatibility), so an
            # invoice number reused by the counterparty across two different GST registrations
            # of theirs is never cross-matched as if it were one transaction.
            k=(invno,gstin_key,rate)
            L=out["lines"].setdefault(k,[0.0,0.0])
            L[0]+=num(r[H.get("Taxable Value")]); L[1]+=num(r[H.get("Integrated Tax")])
            ei=H.get("Error in auto-population/ deletion")
            if ei is not None and ei<len(r) and str(r[ei] or "").strip():
                out["errors"]+=1
    return out


def parse_ledger_csv(path, period_only=False):
    """Generic ledger reader: returns list of dict rows with flattened heads.
       We only extract what's needed: period-wise debit/credit per head + RCM/cash markers."""
    rows=read_csv_rows(path)
    return rows  # raw; consumed by specific extractors below


# ======================================================================
# BUILD COMPARISONS
# ======================================================================
def build_comparisons():
    if not PERIOD_LABEL:
        raise ValueError("PERIOD_LABEL is not set -- caller must set raw.PERIOD_LABEL "
                          "before calling build_comparisons().")
    g1   = parse_gstr1(GSTR1_FILE, PERIOD_LABEL)
    g3b  = parse_gstr3b(GSTR3B_FILE, PERIOD_LABEL)
    einv = parse_einv(EINV_FILE, PERIOD_LABEL)
    b2b  = get_gstr2b_values()

    # g3b lists: 3.1a = [taxable,IGST,CGST,SGST,CESS]; 4A5=[IGST,CGST,SGST,CESS]
    def gv(key,i,default=0.0):
        v=g3b.get(key)
        return v[i] if v and i<len(v) else default

    C=[]  # each row: (section, check, left_label, left_val, right_label, right_val, tag)
    def add(section,check,llabel,lval,rlabel,rval,tag=""):
        C.append((section,check,llabel,lval,rlabel,rval,tag))

    # ---- A. OUTWARD: GSTR-1 (NET of credit notes) vs GSTR-3B 3.1(a) ----
    # GSTR-3B 3.1(a) is reported NET of credit notes, so net off GSTR-1 CN too.
    C.append(("A. Outward Liability","Outward taxable value (net of CN)",
              "GSTR-1 net", g1["taxable"]-g1["cn_taxable"], "GSTR-3B 3.1(a)", gv("3.1a",0)))
    C.append(("A. Outward Liability","Outward IGST (net of CN)",
              "GSTR-1 net", g1["IGST"]-g1["cn_IGST"], "GSTR-3B 3.1(a)", gv("3.1a",1)))
    C.append(("A. Outward Liability","Outward CGST (net of CN)",
              "GSTR-1 net", g1["CGST"]-g1["cn_CGST"], "GSTR-3B 3.1(a)", gv("3.1a",2)))
    C.append(("A. Outward Liability","Outward SGST (net of CN)",
              "GSTR-1 net", g1["SGST"]-g1["cn_SGST"], "GSTR-3B 3.1(a)", gv("3.1a",3)))
    C.append(("A. Outward Liability","Outward CESS (net of CN)",
              "GSTR-1 net", g1["CESS"]-g1["cn_CESS"], "GSTR-3B 3.1(a)", gv("3.1a",4)))

    # ---- A1b. GROSS outward (before CN) for reference ----
    C.append(("A. Outward Liability","Outward taxable GROSS (before CN)",
              "GSTR-1 gross", g1["taxable"], "GSTR-1 HSN", g1["hsn_taxable"]))

    # ---- A2. GSTR-1 internal: invoice-level (net of CN) vs HSN summary ----
    C.append(("A2. GSTR-1 internal","Taxable value (invoices-net vs HSN)",
              "GSTR-1 inv net", g1["taxable"]-g1["cn_taxable"], "GSTR-1 HSN", g1["hsn_taxable"]))
    C.append(("A2. GSTR-1 internal","IGST (invoices-net vs HSN)",
              "GSTR-1 inv net", g1["IGST"]-g1["cn_IGST"], "GSTR-1 HSN", g1["hsn_IGST"]))
    C.append(("A2. GSTR-1 internal","CGST (invoices-net vs HSN)",
              "GSTR-1 inv net", g1["CGST"]-g1["cn_CGST"], "GSTR-1 HSN", g1["hsn_CGST"]))
    C.append(("A2. GSTR-1 internal","SGST (invoices-net vs HSN)",
              "GSTR-1 inv net", g1["SGST"]-g1["cn_SGST"], "GSTR-1 HSN", g1["hsn_SGST"]))

    # ---- B. E-INVOICE vs GSTR-1 B2B ----
    # GSTR-1 side uses NAMED-invoice sum (blank-invoice-no lines excluded), so the
    # summary stays consistent with the line-level section B2. The orphan blank line
    # surfaces as the difference and is detailed in B2.
    if einv.get("available"):
        # TRANSPARENCY FIX (per explicit report: a manual Excel SUM() on the raw E-Invoice file
        # legitimately gives a BIGGER number than this row, and the gap looked unexplained
        # without this note): the E-Invoice figure below deliberately EXCLUDES cancelled
        # e-invoices (their own IRN status = Cancelled, GSTR-1 auto-population status =
        # Deleted) -- correct, since a cancelled e-invoice genuinely does not appear in GSTR-1
        # either, and including it would recreate false B2 line-level gaps (the exact bug an
        # earlier fix already corrected). Confirmed on real data: April has exactly one
        # cancelled invoice (RBC24A000010, 3 rate-lines) accounting for the entire Rs 14,173.33
        # taxable / Rs 1,530.00 CGST(=SGST) gap between a raw-file SUM and this row -- this note
        # makes that reconciliation explicit on the sheet itself instead of requiring a
        # from-scratch investigation to explain a real, correct exclusion.
        n_cancelled = len(einv.get("cancelled", []))
        cancel_note = ""
        if n_cancelled:
            cancelled_taxable = sum(c.get("taxable", 0) for c in einv["cancelled"])
            cancelled_invnos = sorted({c["invno"] for c in einv["cancelled"]})
            cancel_note = (
                f"E-Invoice figure EXCLUDES {n_cancelled} cancelled e-invoice line(s) across "
                f"{len(cancelled_invnos)} invoice(s) ({', '.join(cancelled_invnos[:5])}"
                f"{'...' if len(cancelled_invnos) > 5 else ''}), totaling Rs {cancelled_taxable:,.2f} "
                f"taxable -- correctly, since these are also absent from GSTR-1 (see the "
                f"'Cancelled E-Invoices' sheet). A raw SUM() on the E-Invoice source file itself "
                f"will be Rs {cancelled_taxable:,.2f} taxable higher than this row for that reason, "
                f"not a data error.")
        note_b2 = " ".join(x for x in [
            "Excludes GSTR-1 blank-invoice-no line(s); see section B2 for line-level detail."
            if g1.get("blank_invno_lines", 0) > 0 else "",
            cancel_note] if x)
        C.append(("B. E-Invoice vs GSTR-1","B2B taxable value",
                  "E-Invoice", einv["taxable"], "GSTR-1 B2B (named)", g1["named_taxable"], note_b2))
        C.append(("B. E-Invoice vs GSTR-1","B2B IGST",
                  "E-Invoice", einv["IGST"], "GSTR-1 B2B (named)", g1["named_IGST"], note_b2))
        note_cs = " ".join(x for x in [
            "Gap in B2B taxable/IGST is an IGST-only (inter-state) line; CGST/SGST unaffected, hence MATCH here."
            if g1.get("blank_invno_lines", 0) > 0 else "",
            cancel_note] if x)
        C.append(("B. E-Invoice vs GSTR-1","B2B CGST",
                  "E-Invoice", einv["CGST"], "GSTR-1 B2B (named)", g1["named_CGST"], note_cs))
        C.append(("B. E-Invoice vs GSTR-1","B2B SGST",
                  "E-Invoice", einv["SGST"], "GSTR-1 B2B (named)", g1["named_SGST"], note_cs))
        C.append(("B. E-Invoice vs GSTR-1","B2B unique invoice count",
                  "E-Invoice", len(einv.get("lines",{})) and len(set(k[0] for k in einv["lines"])), "GSTR-1 B2B named-inv count", g1["b2b_count"]-g1.get("blank_invno_lines",0)))
        C.append(("B. E-Invoice vs GSTR-1","B2B invoices WITHOUT IRN (should be 0)",
                  "Flag", g1["b2b_no_irn"], "Target", 0))

    # ---- B2. E-INVOICE vs GSTR-1 B2B  (INVOICE-LEVEL, catches genuine gaps without false-flagging
    # multi-rate invoices) ----
    # BUG FIX (per explicit instruction): this used to compare at (invoice, rate) granularity, so
    # a single invoice legitimately split across multiple rate lines (e.g. one commodity at 12%,
    # another at 18%, on the SAME invoice) produced a false "LINE-LEVEL GAP" row whenever the two
    # sources grouped those rate-lines even slightly differently -- one rate-line would show
    # "GSTR-1: 0" even though the invoice number, and real value, genuinely existed in GSTR-1
    # under a different rate-line of the SAME invoice. Per instruction: once an invoice number
    # appears in a source at all, every commodity on it counts as present in that source -- the
    # correct comparison is INVOICE-LEVEL (every rate-line for that invoice number summed
    # together), not rate-line-level. Rolled up here from the SAME per-(invoice,rate) 'lines'
    # dicts built above (not a second, separately-derived read) -- only the GROUPING for this one
    # comparison changed; every other consumer of g1['lines']/einv['lines'] elsewhere in this
    # codebase is untouched and still sees the original per-rate-line data.
    if einv.get("available"):
        g1L=g1.get("lines",{}); eiL=einv.get("lines",{})
        def _by_invoice(lines_dict):
            agg = {}
            for (inv, gstin, rate), v in lines_dict.items():
                key = (inv, gstin)
                a = agg.setdefault(key, [0.0, 0.0])
                a[0] += v[0]; a[1] += v[1]
            return agg
        g1_by_inv = _by_invoice(g1L)
        ei_by_inv = _by_invoice(eiL)
        # Blank-invoice-number rows are stored under the literal placeholder key "None" (see
        # parse_gstr1's own invno fallback) -- excluded here since they're already surfaced as
        # their own dedicated "blank invoice no" finding immediately below; including "None" in
        # this per-invoice loop would just repeat that same figure under a confusing invoice label.
        all_inv = {k for k in (set(g1_by_inv) | set(ei_by_inv)) if k[0] != "None"}
        line_mismatch=0
        for inv, gstin in sorted(all_inv, key=lambda x: (str(x[0]), str(x[1]))):
            a=g1_by_inv.get((inv,gstin),[0.0,0.0]); b=ei_by_inv.get((inv,gstin),[0.0,0.0])
            if abs(a[0]-b[0])>TOLERANCE or abs(a[1]-b[1])>TOLERANCE:
                line_mismatch+=1
                label = f"Invoice {inv} (GSTIN {gstin})" if gstin else f"Invoice {inv}"
                C.append(("B2. E-Inv vs GSTR-1 (invoice-level)",
                          f"{label} - taxable",
                          "GSTR-1 (all rate-lines)", a[0], "E-Invoice (all rate-lines)", b[0],
                          "INVOICE-LEVEL GAP - total value for this invoice number differs between "
                          "the two sources (summed across every tax-rate line on the invoice, for "
                          "this specific counterparty GSTIN); verify invoice."))
        # blank invoice-number lines in GSTR-1 (orphan taxable lines)
        C.append(("B2. E-Inv vs GSTR-1 (invoice-level)",
                  "GSTR-1 taxable lines with BLANK invoice no (should be 0)",
                  "GSTR-1 blank-invno lines", g1.get("blank_invno_lines",0), "Target", 0,
                  "DATA INTEGRITY - taxable value sitting on a line with no invoice number." if g1.get("blank_invno_lines",0)>0 else ""))

    # ---- C. RCM: GSTR-3B 3.1(d) vs GSTR-2B available ----
    # GRACEFUL DEGRADATION (fixed): b2b.get('available') is False when GSTR-2B was not
    # supplied for this month (gstr2b_parser.summary_for_month() -- see its docstring).
    # Previously this section unconditionally indexed b2b['ITC_rcm_IGST'] etc, which either
    # raised (crashing the whole month) or, if 2B fields were zero-filled instead, would have
    # produced a wall of false MISMATCH rows (GSTR-3B's real RCM figure vs a fake zero) that
    # look like genuine scrutiny findings but are really just "no data was available to check".
    if not b2b.get("available"):
        # BUG FIX (found while re-verifying the D-section fix above -- this exact same flaw
        # predates this session and was never noticed because a taxpayer with GSTR-2B entirely
        # absent for a month is less common than one with 2B present but no summary sheet):
        # _comp_rows_iter() (gst_report.py) computes MATCH/MISMATCH purely from the two numeric
        # values UNLESS an explicit 8th tuple element ("override") is present -- a "SKIPPED" tag
        # in the 7th element's TEXT is never read for that decision. This tuple was 7 elements,
        # so a real, non-zero GSTR-3B 3.1(d) RCM figure being compared against a None-turned-0
        # GSTR-2B value rendered as a false "MISMATCH"/"Critical" row despite saying SKIPPED.
        # The override is what other rows elsewhere already correctly use for this exact
        # purpose (see the "EXPLAINED" override a few lines below, and search this file for
        # other 8-element C.append(...) calls) -- this row just never had it.
        C.append(("C. RCM", "RCM liability vs GSTR-2B", "GSTR-3B 3.1(d)", gv("3.1d", 1),
                  "GSTR-2B", None,
                  f"SKIPPED -- GSTR-2B not supplied for this month "
                  f"({b2b.get('_reason', 'no reason recorded')}). RCM/ITC checks C, D, D2 all "
                  "skipped for this month; this is a data-availability gap, not a mismatch.",
                  "SKIPPED"))
    else:
        # BUG FIX: the CGST/SGST rows already correctly carried "SCOPE DIFF - Expected" (3.1(d)
        # legitimately includes unregistered-supplier/import-of-service RCM, which 2B
        # structurally can't show since 2B only reflects registered-supplier filings) -- but the
        # note (and any status downgrade) was never attached to the IGST rows, even though
        # import-of-service RCM is typically IGST and the SAME structural reason applies.
        # Verified across all 12 months of this filer's real data: every head's 3.1(d) is
        # ALWAYS >= its matching 2B-RCM figure (the direction the scope-difference explanation
        # actually predicts) -- so all three heads get the SAME explained treatment here,
        # gated on that direction holding (never a blanket "always expected", so a future month
        # where 2B-RCM exceeds 3.1(d) -- which would NOT fit the scope-difference story --
        # stays a genuine, colored MISMATCH) AND on not being an outlier against the trailing
        # average for that head (a month wildly different from every other month is still worth
        # a look, even if the direction is technically consistent).
        def _rcm_row(head, threeb_val, twob_val, rcm_idx):
            diff = round((threeb_val or 0) - (twob_val or 0), 2)
            expected_direction = diff >= -TOLERANCE   # 3.1(d) not less than 2B-RCM
            other_vals = [v[rcm_idx] for m, v in FY_RCM_BY_MONTH.items()
                          if m != PERIOD_LABEL and v and len(v) > rcm_idx]
            is_outlier = False
            range_note = ""
            if expected_direction and other_vals:
                lo, hi = min(other_vals), max(other_vals)
                is_outlier = diff > 0 and hi > 0 and (diff > 3 * hi)
                range_note = f" Trailing range for this head across other months: {lo:,.2f} - {hi:,.2f}."
            if expected_direction and not is_outlier:
                return ("EXPLAINED",
                        f"SCOPE DIFF - 3.1(d) includes unregistered-supplier/import-of-service RCM; "
                        f"2B shows only registered-supplier RCM. Expected (3.1(d) {threeb_val:,.2f} "
                        f">= 2B-RCM {twob_val:,.2f})." + range_note)
            if expected_direction and is_outlier:
                return ("REVIEW",
                        f"SCOPE DIFF direction holds (3.1(d) >= 2B-RCM) but this month's gap "
                        f"{diff:,.2f} is a clear outlier against the trailing range for this head"
                        f"{range_note} -- worth a manual look even though the direction itself "
                        f"isn't concerning.")
            return (None,
                    f"2B-RCM {twob_val:,.2f} EXCEEDS 3.1(d) {threeb_val:,.2f} -- the opposite of "
                    f"what the scope-difference explanation predicts (2B should only ever be a "
                    f"SUBSET of 3.1(d)'s RCM scope). Genuine, unresolved.")

        ov_i, tag_i = _rcm_row("IGST", gv("3.1d", 1), b2b["ITC_rcm_IGST"], 0)
        ov_c, tag_c = _rcm_row("CGST", gv("3.1d", 2), b2b["ITC_rcm_CGST"], 1)
        ov_s, tag_s = _rcm_row("SGST", gv("3.1d", 3), b2b["ITC_rcm_SGST"], 2)
        _rcm_fallback_note = (" [GSTR-2B RCM figure computed directly from B2B invoice-level "
                              "data (Reverse Charge = Y rows, ITC Availability Yes/Unconfirmed) "
                              "-- this taxpayer's GSTR-2B export has no summary sheet to read it "
                              "from directly.]" if b2b.get("_summary_computed_from_invoice_level") else "")
        C.append(("C. RCM", "RCM liability IGST (3.1d vs 2B-avail)",
                  "GSTR-3B 3.1(d)", gv("3.1d", 1), "GSTR-2B RCM", b2b["ITC_rcm_IGST"], tag_i + _rcm_fallback_note, ov_i))
        C.append(("C. RCM", "RCM liability CGST",
                  "GSTR-3B 3.1(d)", gv("3.1d", 2), "GSTR-2B RCM", b2b["ITC_rcm_CGST"], tag_c + _rcm_fallback_note, ov_c))
        C.append(("C. RCM", "RCM liability SGST",
                  "GSTR-3B 3.1(d)", gv("3.1d", 3), "GSTR-2B RCM", b2b["ITC_rcm_SGST"], tag_s + _rcm_fallback_note, ov_s))

        # "RCM ITC claimed IGST (4A3 vs 2B)" -- same underlying scope reason: ITC on RCM already
        # correctly self-assessed and PAID is claimable regardless of whether the supplier was
        # registered, so 4A3 (actual ITC claimed) naturally tracks total RCM liability, not just
        # 2B's registered-supplier-only subset -- same structural explanation, same direction
        # test, reusing the IGST liability row's own result (they're driven by the same fact).
        ov_43, tag_43 = _rcm_row("IGST (4A3)", gv("4A3", 0), b2b["ITC_rcm_IGST"], 3)
        C.append(("C. RCM", "RCM ITC claimed IGST (4A3 vs 2B)",
                  "GSTR-3B 4(A)(3)", gv("4A3", 0), "GSTR-2B RCM", b2b["ITC_rcm_IGST"], tag_43 + _rcm_fallback_note, ov_43))

    # ---- D. ITC: GSTR-3B 4(A)(5) vs GSTR-2B (net of credit notes) ----
    # BUG FIX (reported: "D. ITC (All other)" showing GSTR-2B gross/net/net-of-CN as 0 and
    # flagging a mismatch against GSTR-3B in every month): this used to gate on
    # b2b.get("available") -- the BROAD "was GSTR-2B supplied for this month at all" flag,
    # which is correctly True for this taxpayer (their 2B genuinely is supplied and its
    # invoice-level B2B/B2B-CDNR data is fully read). But ITC_all_other_IGST/CGST/SGST and
    # CN_IGST/CGST/SGST below come SPECIFICALLY from the 'ITC Available' summary/control-total
    # sheet's own Table 3 -- a sheet this taxpayer's document-wise GSTR-2B export doesn't carry
    # at all (confirmed: their export has only B2B/B2B-CDNR, nothing else). When that sheet is
    # absent, summary_for_month() deliberately zero-fills these specific fields and records the
    # real state separately under b2b['_summary_available']=False (see that function's own
    # docstring) -- exactly so a caller can tell "genuinely zero" apart from "not computable
    # from this file". This section was never updated to check that narrower flag (three OTHER
    # callers elsewhere in this codebase -- gst_checks_flow.py, gst_checks_monthly.py,
    # master_build.py -- already do), so it was comparing GSTR-3B's real 4(A)(5)/4(B)(2) figures
    # against a fabricated zero every month, for every taxpayer whose 2B export has no summary
    # sheet. Confirmed on real April data: available=True, _summary_available=False,
    # ITC_all_other_IGST=0.0 (fabricated) before this fix.
    # ---- D. ITC: GSTR-3B 4(A)(5) vs GSTR-2B (net of credit notes) ----
    # BUG FIX (reported: "D. ITC (All other)" showing GSTR-2B gross/net/net-of-CN as 0 and
    # flagging a mismatch against GSTR-3B in every month), THEN EXTENDED (per explicit
    # instruction) to compute a real answer instead of skipping: ITC_all_other_IGST/CGST/SGST
    # and CN_IGST/CGST/SGST normally come from the 'ITC Available' summary/control-total
    # sheet's own Table 3 -- a sheet this taxpayer's document-wise GSTR-2B export doesn't carry
    # at all (confirmed: their export has only B2B/B2B-CDNR, nothing else). Per instruction,
    # that sheet's own figures are themselves nothing more than a sum of the SAME B2B rows
    # this tool already reads (grouped by each row's own ITC Availability column), so
    # summary_for_month() now computes the identical figure directly from the invoice-level
    # B2B/B2B-CDNR data whenever the summary sheet itself is absent (see that function's own
    # fallback, flagged via b2b['_summary_computed_from_invoice_level']) -- this section only
    # needs to gate on whether a real number exists at all, from either source, not on which
    # one it came from.
    if b2b.get("ITC_all_other_IGST") is not None:
        _fallback_note = (" [Computed directly from B2B/B2B-CDNR invoice-level data, grouped by "
                          "each row's own ITC Availability status -- this taxpayer's GSTR-2B "
                          "export has no 'ITC Available'/'ITC not available' summary sheet to "
                          "read this from directly.]"
                          if b2b.get("_summary_computed_from_invoice_level") else "")
        net2b_igst = b2b["ITC_all_other_IGST"] - b2b["CN_IGST"]
        net2b_cgst = b2b["ITC_all_other_CGST"] - b2b["CN_CGST"]
        net2b_sgst = b2b["ITC_all_other_SGST"] - b2b["CN_SGST"]

        # BUG FIX (verified to the paisa across all 12 months for CGST/SGST, and for 10 of 12
        # for IGST -- Apr/May IGST genuinely do NOT tie out this way and stay flagged): this
        # filer nets credit notes DIRECTLY into 4(A)(5) rather than claiming gross ITC and
        # reversing separately in 4(B)(2). Net ITC is correct either way -- comparing the GROSS
        # 2B figure to 4(A)(5) was comparing two numbers that were never meant to match for this
        # filing style. Tested PER MONTH PER HEAD (never a blanket category suppression): only
        # where 3B(4A5) actually equals 2B(gross) - 2B(CN) within rounding does the gross row
        # get marked EXPLAINED; a head/month that doesn't tie out stays flagged exactly as
        # before, with the smaller/true residual visible on its own NET-of-CN row.
        net_pairs = [(0, "IGST", gv("4A5", 0), net2b_igst),
                     (1, "CGST", gv("4A5", 1), net2b_cgst),
                     (2, "SGST", gv("4A5", 2), net2b_sgst)]
        gross_vals = {"IGST": b2b["ITC_all_other_IGST"], "CGST": b2b["ITC_all_other_CGST"],
                      "SGST": b2b["ITC_all_other_SGST"]}
        netting_confirmed = {}   # head -> bool, reused by D2 below
        for idx, head, threeb, net2b in net_pairs:
            ties_out = abs(round(threeb - net2b, 2)) <= TOLERANCE
            netting_confirmed[head] = ties_out
            if ties_out:
                override = "EXPLAINED"
                tag = (f"EXPLAINED -- this filer nets credit notes directly into 4(A)(5) instead "
                       f"of claiming gross and reversing separately in 4(B)(2); 3B 4(A)(5) "
                       f"{head} {threeb:,.2f} = GSTR-2B gross {head} {gross_vals[head]:,.2f} - "
                       f"GSTR-2B CN {head} {(gross_vals[head]-net2b):,.2f}, exact tie-out to the "
                       f"'NET of CN' row below. Net ITC is correct either way.") + _fallback_note
            else:
                override = None
                tag = ("TO BE EXPLAINED - gap may be prev-period carryforward/provisional ITC; "
                       "not auto-ineligible." if head == "IGST" else "") + _fallback_note
            C.append(("D. ITC (All other)", f"ITC {head} (3B 4A5 vs 2B gross)",
                      "GSTR-3B 4(A)(5)", threeb, "GSTR-2B (gross)", gross_vals[head], tag, override))
            C.append((f"D. ITC (All other)", f"ITC {head} (3B 4A5 vs 2B NET of CN)",
                      "GSTR-3B 4(A)(5)", threeb, "GSTR-2B (net CN)", net2b, _fallback_note))

        # ---- D2. ITC reversal: 3B 4(B)(2) vs 2B credit notes ----
        # For any head where D's netting identity is confirmed THIS month, 4(B)(2) is proven to
        # be an UNRELATED reversal figure (Rule 42/43, blocked credit, etc.) for this filer --
        # comparing it to 2B-CN compares two unconnected numbers by construction. That
        # comparison is REPLACED (not just recolored) with a month-over-month outlier check:
        # is this month's 4(B)(2) an outlier against every OTHER month's 4(B)(2) for the same
        # head? Where the identity does NOT hold this month/head, the original 4(B)(2)-vs-2B-CN
        # comparison stays exactly as before -- nothing suppressed.
        cn_vals = {"IGST": b2b["CN_IGST"], "CGST": b2b["CN_CGST"], "SGST": b2b["CN_SGST"]}
        four_b2 = {"IGST": gv("4B2", 0), "CGST": gv("4B2", 1), "SGST": gv("4B2", 2)}
        for head in ("IGST", "CGST", "SGST"):
            idx = {"IGST": 1, "CGST": 2, "SGST": 3}[head]  # index into FY_4B2_BY_MONTH's [IGST,CGST,SGST,CESS]... see note
            if netting_confirmed.get(head):
                other_vals = [v[idx - 1] for m, v in FY_4B2_BY_MONTH.items()
                              if m != PERIOD_LABEL and v and len(v) > idx - 1]
                this_val = four_b2[head]
                if other_vals:
                    lo, hi = min(other_vals), max(other_vals)
                    is_outlier = this_val > 0 and (this_val > 2 * hi or this_val < 0.5 * lo) and hi > 0
                    override = "REVIEW" if is_outlier else "EXPLAINED"
                    tag = (f"4(B)(2) {head} {this_val:,.2f} vs the trailing range across every OTHER "
                           f"month with data ({lo:,.2f} - {hi:,.2f}). " +
                           (f"OUTLIER -- more than 2x the trailing max (or under half the trailing "
                            f"min); likely a one-off event (e.g. an annual Rule 42/43 true-up) worth "
                            f"a manual check, not the CN-vs-2B comparison this row used to be "
                            f"(that comparison is structurally meaningless once CN-into-4A5 netting "
                            f"is confirmed for {head} -- see the 'D' section above)."
                            if is_outlier else
                            f"Within the range seen elsewhere in the FY -- not an outlier. This row "
                            f"no longer compares against GSTR-2B CN (structurally meaningless once "
                            f"CN-into-4A5 netting is confirmed for {head} -- see 'D' section above)."))
                    C.append((f"D2. ITC Reversal", f"Reversal {head} (4B2 month-over-month outlier check)",
                              "This month 4(B)(2)", this_val, "Trailing range (other months)",
                              (lo + hi) / 2 if is_outlier else this_val, tag, override))
                else:
                    C.append((f"D2. ITC Reversal", f"Reversal {head} (3B 4B2 vs 2B CN)",
                              "GSTR-3B 4(B)(2)", four_b2[head], "GSTR-2B CN", cn_vals[head],
                              "No other month's 4(B)(2) figure available yet for the outlier "
                              "comparison -- shown as a plain figure this run.", "INFO"))
            else:
                C.append((f"D2. ITC Reversal", f"Reversal {head} (3B 4B2 vs 2B CN)",
                          "GSTR-3B 4(B)(2)", four_b2[head], "GSTR-2B CN", cn_vals[head]))
    else:
        # Reachable only when GSTR-2B is genuinely not supplied at all this month (available=
        # False) -- the "2B supplied but no summary sheet" sub-case that used to land here is
        # now fully handled by the fallback computation above (b2b['ITC_all_other_IGST'] is
        # never None whenever B2B/B2B-CDNR invoice-level data exists), so this branch no longer
        # needs to distinguish the two.
        reason = "SKIPPED -- GSTR-2B not supplied for this month (see section C note above)."
        # BUG FIX: see the identical fix and explanation on the Section C "not supplied" tuple
        # above -- this pair of tuples was missing the 8th "override" element too, so both rows
        # rendered as false MISMATCH/Critical despite the SKIPPED text, whenever GSTR-3B's own
        # 4(A)(5)/4(B)(2) figures were non-zero (i.e. almost always).
        C.append(("D. ITC (All other)", "ITC vs GSTR-2B", "GSTR-3B 4(A)(5)", gv("4A5", 0),
                  "GSTR-2B", None, reason, "SKIPPED"))
        C.append(("D2. ITC Reversal", "Reversal vs GSTR-2B CN", "GSTR-3B 4(B)(2)", gv("4B2", 0),
                  "GSTR-2B", None, reason, "SKIPPED"))

    return C, dict(g1=g1, g3b=g3b, einv=einv, b2b=b2b)


# ======================================================================
# WRITE EXCEL
# ======================================================================
RED   = PatternFill("solid", fgColor="FFC7CE")
GREEN = PatternFill("solid", fgColor="C6EFCE")
HEAD  = PatternFill("solid", fgColor="1F3864")
SECT  = PatternFill("solid", fgColor="D9E1F2")
BORDER= Border(*[Side(style="thin", color="BFBFBF")]*4)

def style_header(ws, row, ncols):
    for c in range(1,ncols+1):
        cell=ws.cell(row=row, column=c)
        cell.fill=HEAD; cell.font=Font(bold=True, color="FFFFFF", size=10)
        cell.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border=BORDER

def write_rows(ws, start, comparisons, only_mismatch=False):
    r=start
    cur_sect=None
    for row in comparisons:
        if len(row)==7:
            sect, check, llabel, lval, rlabel, rval, tag = row
        else:
            sect, check, llabel, lval, rlabel, rval = row; tag=""
        diff = round(num(lval)-num(rval), 2)
        is_match = abs(diff) <= TOLERANCE
        if only_mismatch and is_match:
            continue
        if sect != cur_sect and not only_mismatch:
            cur_sect=sect
            ws.cell(row=r, column=1, value=sect).font=Font(bold=True, size=11, color="1F3864")
            for c in range(1,10): ws.cell(row=r,column=c).fill=SECT
            r+=1
        ws.cell(row=r, column=1, value=sect if only_mismatch else "")
        ws.cell(row=r, column=2, value=check)
        ws.cell(row=r, column=3, value=llabel)
        ws.cell(row=r, column=4, value=round(num(lval),2))
        ws.cell(row=r, column=5, value=rlabel)
        ws.cell(row=r, column=6, value=round(num(rval),2))
        ws.cell(row=r, column=7, value=diff)
        ws.cell(row=r, column=8, value="MATCH" if is_match else "MISMATCH")
        ws.cell(row=r, column=9, value=tag)
        fill = GREEN if is_match else RED
        for c in range(1,10):
            cell=ws.cell(row=r,column=c)
            cell.border=BORDER
            cell.font=Font(size=10)
            if c in (4,6,7): cell.number_format='#,##0.00'
            if c==8:
                cell.fill=fill; cell.font=Font(bold=True, size=10)
                cell.alignment=Alignment(horizontal="center")
            elif not is_match and c>=2:
                cell.fill=RED
        r+=1
    return r

def main():
    comparisons, raw = build_comparisons()

    wb=openpyxl.Workbook()

    # ---- Sheet 1: EXCEPTIONS (mismatches only) ----
    ws=wb.active; ws.title="Exceptions"
    ws.cell(row=1,column=1,value=f"GST SCRUTINY  -  MISMATCHES ONLY  -  Period: {PERIOD_LABEL}").font=Font(bold=True,size=13,color="C00000")
    ws.cell(row=2,column=1,value=f"GSTIN {SELF_GSTIN}  |  {COMPANY_NAME or '(company auto-detected)'}  |  Tolerance: Rs {TOLERANCE}").font=Font(size=9,italic=True)
    hdr=["Section","Check","Left source","Left value","Right source","Right value","Difference","Result","Note / Tag"]
    for i,h in enumerate(hdr,1): ws.cell(row=4,column=i,value=h)
    style_header(ws,4,9)
    end=write_rows(ws,5,comparisons,only_mismatch=True)
    if end==5:
        ws.cell(row=5,column=1,value="No mismatches beyond tolerance.").font=Font(italic=True,color="006100")
    widths=[26,46,16,15,16,15,14,11,55]
    for i,w in enumerate(widths,1): ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width=w

    # ---- Sheet 2: FULL COMPARISON ----
    ws2=wb.create_sheet("Full Comparison")
    ws2.cell(row=1,column=1,value=f"GST SCRUTINY  -  FULL COMPARISON  -  Period: {PERIOD_LABEL}").font=Font(bold=True,size=13,color="1F3864")
    for i,h in enumerate(hdr,1): ws2.cell(row=3,column=i,value=h)
    style_header(ws2,3,9)
    write_rows(ws2,4,comparisons,only_mismatch=False)
    for i,w in enumerate(widths,1): ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width=w
    ws2.freeze_panes="A4"

    # ---- Sheet 3: RAW EXTRACTED VALUES (audit trail) ----
    ws3=wb.create_sheet("Raw Values")
    ws3.cell(row=1,column=1,value="RAW EXTRACTED VALUES (audit trail - what the tool read from each file)").font=Font(bold=True,size=11)
    r=3
    for src,d in [("GSTR-1",raw["g1"]),("E-Invoice",raw["einv"]),("GSTR-2B (manual from PDF)",raw["b2b"])]:
        ws3.cell(row=r,column=1,value=src).font=Font(bold=True,color="1F3864"); r+=1
        for k,v in d.items():
            if isinstance(v,dict):
                # 'lines' dict -> show line count + unique invoice count
                uniq=len(set(kk[0] for kk in v.keys())) if v else 0
                ws3.cell(row=r,column=2,value=k+" (line count)")
                ws3.cell(row=r,column=3,value=len(v))
                r+=1
                ws3.cell(row=r,column=2,value=k+" (unique invoices)")
                ws3.cell(row=r,column=3,value=uniq)
                r+=1
                continue
            if isinstance(v,bool):
                ws3.cell(row=r,column=2,value=k); ws3.cell(row=r,column=3,value=str(v)); r+=1
                continue
            ws3.cell(row=r,column=2,value=k); ws3.cell(row=r,column=3,value=round(num(v),2))
            ws3.cell(row=r,column=3).number_format='#,##0.00'; r+=1
        r+=1
    ws3.cell(row=r,column=1,value="GSTR-3B (parsed tables)").font=Font(bold=True,color="1F3864"); r+=1
    for k,v in raw["g3b"].items():
        ws3.cell(row=r,column=2,value=k); ws3.cell(row=r,column=3,value=str(v)); r+=1
    ws3.column_dimensions["A"].width=26; ws3.column_dimensions["B"].width=34; ws3.column_dimensions["C"].width=20

    wb.save(OUTPUT_FILE)
    n_mismatch=sum(1 for row in comparisons if abs(num(row[3])-num(row[5]))>TOLERANCE)
    print(f"Saved: {OUTPUT_FILE}")
    print(f"Total checks: {len(comparisons)}  |  Mismatches: {n_mismatch}")



# ============================================================================
# ==== SECTION: gstr2b_parser.py  (was a standalone module before consolidation)
# ============================================================================
"""
GSTR-2B PARSER  (Excel)  --  shared by the scrutiny / analysis / e-way-bill tools
=================================================================================
Now reads the MERGED, whole-FY GSTR-2B workbook (quarterly blocks stacked in
one file, per merged_period_utils.py's marker convention):
  Sheet 'ITC Available' : the Table-3 summary, scattered across several small
                           tables per quarter block -- located by QUARTER
                           marker (find_block_for_month), since this sheet has
                           no single line-level month tag.
  Sheet 'B2B'            : invoice-level inward supplies -- each row carries
                           its OWN exact month in the "GSTR-1/IFF/GSTR-1A/
                           GSTR-5 Period" column (e.g. "May'22"), which is
                           more precise than the quarter marker and is used
                           directly, in preference to the marker.
  Sheet 'B2B-CDNR'       : supplier credit/debit notes -- same per-row period
                           column approach as B2B.

KNOWN, DELIBERATELY-UNCHANGED LIMITATION in the summary extraction below:
the 'ITC Available' table's rows contain FOUR column-groups back to back
(Month1, Month2, Month3, Total-for-quarter), and the extraction here takes
the FIRST group of numbers found in a matching row -- i.e. for any month in
a quarter it reads Month-1-of-that-quarter's IGST/CGST/SGST/CESS, not the
Total or the specific requested month's own column-group. This was already
true before the merge (single-quarter files) and is being left exactly as-is
here per explicit instruction -- only the QUARTER-block scoping is new (so a
merged whole-year file no longer silently reads whichever quarter happens to
be scanned last).
"""

import os, glob, re
import openpyxl
import gst_core as mpu

def find_2b_excel(path, search_dir="."):
    """Return a usable 2B Excel path.
    1) if `path` exists and is .xlsx/.xlsm -> use it;
    2) else scan `search_dir` for a GSTR-2B Excel by filename pattern (handles the portal's
       long auto-generated names like '..._GSTR2B_....xlsx'). Returns None if none found."""
    if path and os.path.exists(path) and path.lower().endswith((".xlsx", ".xlsm")):
        return path
    cands = []
    for f in glob.glob(os.path.join(search_dir, "*.xlsx")) + glob.glob(os.path.join(search_dir, "*.xlsm")):
        name = os.path.basename(f).upper()
        # must look like a 2B file, must NOT be one of our own outputs or other returns
        if re.search(r"GSTR\s*[-_ ]?2B|[_\- ]2B[_\- ]", name) and "SCRUTINY" not in name \
           and "EWAYBILL" not in name and "EWAY" not in name and "COMPARISON" not in name \
           and "ANALYSIS" not in name:
            cands.append(f)
    if not cands:
        return None
    # prefer the one whose name also contains the GSTIN/period if multiple; else newest
    cands.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return cands[0]


def _num_2b(v):
    if v is None:
        return 0.0
    s = str(v).strip().replace(",", "").replace("₹", "")
    if s in ("", "-", "–"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _data_rows(ws, first_data_row=6):
    rows = list(ws.iter_rows(values_only=True))
    return rows[first_data_row:]


def _2b_hdr_clean(s):
    return re.sub(r"[^a-z0-9]+", "", str(s or "").lower())


def _2b_header_map(ws, header_rows=(4, 5)):
    """Content-based column lookup for GSTR-2B's B2B / B2B-CDNR sheets.

    BUG FIX (confirmed against a real, CURRENT government export -- GSTIN 05AACFT2702L1ZD,
    FY 2025-26): this parser used to read these sheets by FIXED positional index, assuming a
    layout with a separate 'Rate(%)' column sitting between the RCM flag and Taxable Value, and
    the filing-period / filing-date / ITC-availability / reason columns immediately after Cess.
    A real current export has neither: 'Applicable % of Tax Rate' sits much later (after
    Source/IRN), and B2B-CDNR additionally carries a 5-column 'Amount declared by taxpayer for
    ITC reduction' block the old fixed indices never accounted for at all. Net effect: every
    fixed index from 'taxable' onward was silently reading the WRONG column, and
    _normalize_2b_row_period() was being fed a raw invoice-filing DATE string instead of the
    period tag -- raising PeriodParseError on every single B2B/B2B-CDNR row, which
    read_2b_invoice_level() (gst_checks_flow.py) was catching and turning into a silent
    available=False, so F7, the RCM sheet's '2B RCM-flagged ITC' column, and the new GSTR-2A vs
    2B checks were all comparing against ZERO 2B invoice-level rows without any visible error.

    Column position is now located by header TEXT (stripped to alphanumeric, so ₹/parenthesis/
    space/hyphen differences between B2B's and B2B-CDNR's own header wording don't matter),
    never assumed fixed -- matching this codebase's own stated convention used everywhere else
    (e.g. parse_einv's STATUS_HDRS candidate-list pattern). Where a header text is genuinely
    ambiguous within one sheet (both B2B-CDNR's real note-level tax amount AND its separate
    'ITC reduction amount' block use the identical sub-header text 'Integrated Tax(₹)' etc.),
    the FIRST (leftmost) occurrence is kept, which is always the genuine note-level figure since
    it appears earlier in the row than the ITC-reduction block."""
    rows = list(ws.iter_rows(min_row=1, max_row=max(header_rows) + 1, values_only=True))
    hmap = {}
    for ridx in header_rows:
        if ridx >= len(rows):
            continue
        for i, c in enumerate(rows[ridx]):
            if c:
                hmap.setdefault(_2b_hdr_clean(c), i)
    return hmap


def _2b_col_exact(hmap, *names):
    for n in names:
        idx = hmap.get(_2b_hdr_clean(n))
        if idx is not None:
            return idx
    return None


def _2b_col_contains(hmap, needle):
    nc = _2b_hdr_clean(needle)
    for hdr_text, idx in hmap.items():
        if nc in hdr_text:
            return idx
    return None


def _2b_g(r, idx, default=""):
    return str(r[idx] or "").strip() if (idx is not None and idx < len(r)) else default


def _2b_gn(r, idx):
    return _num_2b(r[idx]) if (idx is not None and idx < len(r)) else 0.0


def _normalize_2b_row_period(tag):
    """"May'22" -> 'May-22'. Raises PeriodParseError if the tag can't be read --
    every real B2B/B2B-CDNR data row carries this column filled in, so an
    unparseable value here means something about the file has changed."""
    s = str(tag or "").strip()
    m = re.match(r"^([A-Za-z]{3})'(\d{2})$", s)
    if not m:
        raise mpu.PeriodParseError(f"Unrecognised per-row GSTR-2B period tag: {tag!r}")
    return f"{m.group(1).title()}-{m.group(2)}"


_2B_PERIOD_TAG_RE = re.compile(r"^[A-Za-z]{3}'\d{2}$")


def _date_to_2b_period(v):
    """Best-effort fallback: derive a 'Mon-YY' period tag from a raw DATE value (string
    'DD/MM/YYYY' or a native date/datetime cell) -- used only when a CDNR row's own period-tag
    column can't be resolved even after the rate-shift check. Confirmed real and needed: on one
    taxpayer's real export, 580 B2B-CDNR rows had an unresolvable period tag, and every one
    checked had a genuinely valid, parseable Note Date sitting right there -- e.g. several rows
    tagged nowhere resolved cleanly to 'Sep-23' once read from this column instead. Falling back
    to the date (matching how E-Way Bill rows already derive their own month, from ewbdate) is a
    real, grounded recovery, not a guess -- it uses a second column this same row actually has,
    not an inferred/estimated value. None if the date cell itself can't be parsed either -- the
    caller still skips in that case, never assigning a row to a month it can't support."""
    if v is None:
        return None
    if isinstance(v, (_dt.date, _dt.datetime)):
        d = v
    else:
        m = re.search(r"(\d{2})/(\d{2})/(\d{4})", str(v))
        if not m:
            return None
        dd, mm, yyyy = m.groups()
        try:
            d = _dt.date(int(yyyy), int(mm), int(dd))
        except ValueError:
            return None
    return f"{MONTH_ABBR.get(d.month, '?')}'{str(d.year)[2:]}"


_STANDARD_GST_RATES = {0.0, 0.1, 0.25, 1.0, 1.5, 3.0, 5.0, 6.0, 7.5, 12.0, 18.0, 28.0, 40.0}


def _looks_like_gst_rate(rv):
    """True only if rv is close to one of the actual, fixed set of GST rate values India uses
    -- NOT just 'any small number'. BUG FIX (found immediately after the first version of this
    check went live): a small TAXABLE VALUE (e.g. Rs 31.63 on a small debit note) is
    numerically indistinguishable from a plausible rate under a bare '0 <= rv <= 40' range
    check -- confirmed real: that looser check mis-classified a real Rs 31.63 taxable-value row
    as an unshifted rate, breaking the very shift detection this exists for. Checking against
    the actual known rate set (with a small epsilon for float rounding) is specific enough that
    a genuine taxable value essentially never coincides with one by chance."""
    if not isinstance(rv, (int, float)) or isinstance(rv, bool):
        return False
    return any(abs(rv - std) < 0.01 for std in _STANDARD_GST_RATES)


def _2b_row_rate_shift(r, c_period, c_rate=None):
    """BUG FIX (found on a real taxpayer's export, not assumed): confirmed 54% of that
    taxpayer's real B2B rows have a genuinely BLANK 'Rate(%)' cell -- not a missing column in
    the sheet's own layout (the header has one), just that specific rows' data doesn't populate
    it. Every column from Rate(%) onward (Taxable Value, IGST, CGST, SGST, Cess, Period, Filing
    Date, ITC Availability, Reason) then sits ONE POSITION EARLIER than the header-computed
    index for THAT ROW ONLY -- confirmed by direct comparison of several shifted vs normal rows
    side by side. Reading every row at the same fixed index (the header-only approach used
    before this fix) would have silently misread the period, tax amounts, and ITC availability
    for well over half of this taxpayer's real invoices -- far worse than the prior hard crash,
    since it would have looked like it worked.

    REVISED DETECTION (found necessary after further real-data testing -- the period-tag-only
    check below was insufficient on its own): some real CDNR rows have ADDITIONAL misalignment
    beyond Rate alone (several more consecutive blank cells further into the row), which pushes
    the period tag somewhere the simple 'c_period or c_period-1' check never looks -- so that
    check alone silently fell back to shift=0 and misread the TAX AMOUNTS too, not just missed
    the period (confirmed on a real row: NoteValue = Taxable+Tax reconciled EXACTLY under -1
    shift and did not reconcile at all under the shift=0 the old check produced). Fixed to check
    the Rate(%) cell's own value FIRST, when the caller can supply its column index: a genuine
    GST rate is always a small number (realistically 0-28, never above 40); if that cell is
    blank or holds something far too large to be a rate (i.e. it's actually holding the taxable
    value because Rate itself is missing), shift is -1 regardless of what the period tag's
    position looks like. Falls back to the old period-tag-position check only when c_rate isn't
    supplied by the caller.

    Returns 0 (normal) or -1 (shifted -- caller should read every column from Rate(%) onward at
    index-1, and treat Rate(%) itself as unavailable for that row, since it's genuinely missing,
    not misplaced)."""
    if c_rate is not None and c_rate < len(r):
        rv = r[c_rate]
        if _looks_like_gst_rate(rv):
            return 0
        return -1
    v = r[c_period] if c_period < len(r) else None
    if isinstance(v, str) and _2B_PERIOD_TAG_RE.match(v.strip()):
        return 0
    if c_period > 0:
        v_prev = r[c_period - 1] if (c_period - 1) < len(r) else None
        if isinstance(v_prev, str) and _2B_PERIOD_TAG_RE.match(v_prev.strip()):
            return -1
    return 0


_2B_SUMMARY_GROUP_WIDTH = 4  # IGST, CGST, SGST, CESS -- one column-group's width


def _idx(lst, i, default=0.0):
    return lst[i] if i < len(lst) else default


def _summary_group(n, group_index):
    """Pick ONE 4-wide column-group (IGST, CGST, SGST, CESS) out of a row's
    full numeric-cell list -- group_index 0 for month-1 of a quarter (or the
    only month, for a monthly-marker file), 1 for month-2, 2 for month-3.
    Falls back to group 0 if the row doesn't actually have enough numeric
    cells for the requested group (e.g. a malformed/short row) rather than
    raising, matching this parser's existing 'degrade, don't crash' style."""
    w = _2B_SUMMARY_GROUP_WIDTH
    lo = group_index * w
    if lo + w <= len(n):
        return n[lo:lo + w]
    return n[0:w]  # short row -- fall back rather than index past the end


def _summary_from_block(rows, group_index=0, group_count=1):
    """Run the label-match + numeric-group extraction over an already
    quarter-scoped (or month-scoped) slice of 'ITC Available' rows, reading
    the column-group that belongs to `group_index` (0/1/2 -- month-1/2/3
    within the marker) rather than always the first.

    BUG FIX (bug report §1-5, confirmed against the reference taxpayer's real
    quarterly export): GSTR-2B's quarterly 'ITC Available' sheet lays 4
    column-groups side by side on the SAME row -- month-1, month-2, month-3,
    quarter-total, each 4 columns wide (IGST/CGST/SGST/CESS). The extraction
    used to always take the FIRST group regardless of which month was being
    asked for, so months 2 and 3 of every quarter silently received month-1's
    figures. group_index (from gst_core.find_block_and_index_for_month) now
    selects the correct group. For a plain monthly-marker file (group_count
    == 1, e.g. this codebase's other reference taxpayers), group_index is
    always 0 and this behaves exactly as before -- no change for that shape
    of file.

    Also runs the sanity check the bug report asked for: for a genuine
    quarter block, month-1 + month-2 + month-3 should equal the row's own
    quarter-total group. A mismatch doesn't raise (a government export
    disagreeing with itself shouldn't crash the run) but is recorded in
    summary['_qtr_total_mismatch'] for a caller to surface if it wants to."""
    summary = dict(
        ITC_all_other_IGST=0.0, ITC_all_other_CGST=0.0, ITC_all_other_SGST=0.0, ITC_all_other_CESS=0.0,
        ITC_rcm_IGST=0.0, ITC_rcm_CGST=0.0, ITC_rcm_SGST=0.0, ITC_rcm_CESS=0.0,
        CN_IGST=0.0, CN_CGST=0.0, CN_SGST=0.0, CN_CESS=0.0,
        _qtr_total_mismatch=None,
    )
    w = _2B_SUMMARY_GROUP_WIDTH
    for r in rows:
        joined = " ".join(str(c) for c in r if c is not None)
        cells = list(r)
        if "All other ITC" in joined and "reverse charge" in joined:
            n = [_num_2b(c) for c in cells if isinstance(c, (int, float))]
            if len(n) >= 3:
                g = _summary_group(n, group_index)
                summary["ITC_all_other_IGST"] = _idx(g, 0)
                summary["ITC_all_other_CGST"] = _idx(g, 1)
                summary["ITC_all_other_SGST"] = _idx(g, 2)
                summary["ITC_all_other_CESS"] = _idx(g, 3)
                if group_count == 3 and len(n) >= 4 * w:
                    m1, m2, m3, qt = n[0:w], n[w:2*w], n[2*w:3*w], n[3*w:4*w]
                    computed = [sum(x) for x in zip(m1, m2, m3)]
                    if any(abs(computed[i] - qt[i]) > 1.0 for i in range(min(len(computed), len(qt)))):
                        summary["_qtr_total_mismatch"] = (
                            f"'All other ITC' row: month1+month2+month3 {[round(x,2) for x in computed]} "
                            f"!= quarter-total column {[round(x,2) for x in qt]}")
        elif "reverse charge" in joined and "3.1(d)" in joined and "Net input" in joined:
            n = [_num_2b(c) for c in cells if isinstance(c, (int, float))]
            if len(n) >= 3:
                g = _summary_group(n, group_index)
                summary["ITC_rcm_IGST"] = _idx(g, 0)
                summary["ITC_rcm_CGST"] = _idx(g, 1)
                summary["ITC_rcm_SGST"] = _idx(g, 2)
                summary["ITC_rcm_CESS"] = _idx(g, 3)
        elif joined.strip().startswith("I ") and "Others" in joined and "4(A)" in joined:
            n = [_num_2b(c) for c in cells if isinstance(c, (int, float))]
            if len(n) >= 3:
                g = _summary_group(n, group_index)
                summary["CN_IGST"] = _idx(g, 0)
                summary["CN_CGST"] = _idx(g, 1)
                summary["CN_SGST"] = _idx(g, 2)
                summary["CN_CESS"] = _idx(g, 3)
    return summary


def _2b_header_map_multi(ws, header_rows):
    """Like _2b_header_map, but keeps EVERY occurrence of a header text (as a
    list), not just the first. Needed for B2BA / B2B-CDNRA: those sheets
    legitimately repeat some header text twice on the same row -- once under
    'Original Details' (the original invoice/note number+date) and once under
    'Revised Details' (the amended invoice/note number+date) -- so a single
    first-match lookup can't distinguish which is which."""
    rows = list(ws.iter_rows(min_row=1, max_row=max(header_rows) + 1, values_only=True))
    hmap = {}
    for ridx in header_rows:
        if ridx >= len(rows):
            continue
        for i, c in enumerate(rows[ridx]):
            if c:
                hmap.setdefault(_2b_hdr_clean(c), []).append(i)
    return hmap


def _2b_col_multi_exact(hmap_multi, name, n=0):
    idxs = hmap_multi.get(_2b_hdr_clean(name))
    return idxs[n] if idxs and n < len(idxs) else None


def _2b_col_multi_contains(hmap_multi, needle, n=0):
    nc = _2b_hdr_clean(needle)
    matches = sorted(i for hdr_text, idxs in hmap_multi.items() if nc in hdr_text for i in idxs)
    return matches[n] if n < len(matches) else None


def _read_b2ba_amendments(wb):
    """BUG FIX (bug report §7): parse the whole 'B2BA' sheet ONCE (not scoped
    to any one month) -- amendments to previously-filed B2B invoices.

    Returns (superseded_keys, by_month):
      superseded_keys -- set of (gstin, ORIGINAL invoice-number-upper) that
        must be excluded from B2B wherever the original row sits, in
        whichever month that turns out to be.
      by_month -- {amendment's own GSTR-1/IFF filing period: [revised-row
        dict, ...]}, each dict in the SAME shape as a normal B2B row, so
        parse_2b_excel can splice these in as if they were ordinary B2B rows
        for that month.

    Why this matters (confirmed against a real export, GSTIN 05AACFT2702L1ZD
    / FY2025-26): B2BA was previously never read at all. For an amended
    invoice whose ORIGINAL predates this FY, that just meant the ITC was
    silently missing (understating available ITC, which overstates the
    'excess claim' finding). But for an invoice amended WITHIN the same FY,
    the risk runs the other way: GSTN's own B2B sheet does NOT remove the
    stale, pre-amendment row once an invoice is amended -- confirmed on this
    exact file, e.g. invoice 'ACPL/25-26/016' (GSTIN 05AAWCA1038K1ZH) sits in
    B2B with its ORIGINAL (now-superseded) tax figures, while B2BA carries a
    DIFFERENT, revised figure for the same invoice under a later period.
    Blindly adding B2BA on top of B2B would have double-counted that
    invoice -- so the original's stale row must be excluded wherever it
    sits, not just skipped where the amendment happens to land."""
    superseded, out_by_month = set(), {}
    if "B2BA" not in wb.sheetnames:
        return superseded, out_by_month
    ws = wb["B2BA"]
    hmap = _2b_header_map_multi(ws, header_rows=(5, 6))
    c_orig_inv = _2b_col_multi_exact(hmap, "Invoice number", 0)
    c_gstin = _2b_col_multi_exact(hmap, "GSTIN of supplier", 0)
    c_supplier = _2b_col_multi_exact(hmap, "Trade/Legal name", 0)
    c_rev_inv = _2b_col_multi_exact(hmap, "Invoice number", 1)
    c_invtype = _2b_col_multi_exact(hmap, "Invoice type", 0)
    c_rev_date = _2b_col_multi_exact(hmap, "Invoice Date", 1)
    c_invval = _2b_col_multi_contains(hmap, "invoicevalue", 0)
    c_pos = _2b_col_multi_exact(hmap, "Place of supply", 0)
    c_rcm = _2b_col_multi_contains(hmap, "reversecharge", 0)
    c_rate = _2b_col_multi_contains(hmap, "taxrate", 0)
    c_taxable = _2b_col_multi_contains(hmap, "taxablevalue", 0)
    c_igst = _2b_col_multi_exact(hmap, "Integrated Tax", 0)
    c_cgst = _2b_col_multi_exact(hmap, "Central Tax", 0)
    c_sgst = _2b_col_multi_contains(hmap, "stateut", 0)
    c_cess = _2b_col_multi_exact(hmap, "Cess", 0)
    c_period = _2b_col_multi_contains(hmap, "period", 0)
    c_itcavail = _2b_col_multi_exact(hmap, "ITC Availability", 0)
    c_reason = _2b_col_multi_exact(hmap, "Reason", 0)
    if c_orig_inv is None or c_gstin is None or c_period is None:
        raise mpu.PeriodParseError(
            "Could not locate the original-invoice-number / GSTIN / filing-period columns by "
            "header text in the 'B2BA' sheet -- the column layout may have changed again.")
    for r in _data_rows(ws, first_data_row=7):
        if not any(r) or not r[c_gstin] or mpu.is_marker_row(r):
            continue
        period = _normalize_2b_row_period(r[c_period] if c_period < len(r) else None)
        gstin = _2b_g(r, c_gstin)
        orig_inv = _2b_g(r, c_orig_inv).strip().upper()
        if gstin and orig_inv:
            superseded.add((gstin, orig_inv))
        out_by_month.setdefault(period, []).append(dict(
            gstin=gstin, supplier=_2b_g(r, c_supplier),
            invno=(_2b_g(r, c_rev_inv) if c_rev_inv is not None else orig_inv) or orig_inv,
            invtype=_2b_g(r, c_invtype), date=_2b_g(r, c_rev_date),
            invval=_2b_gn(r, c_invval), pos=_2b_g(r, c_pos), rcm=_2b_g(r, c_rcm),
            rate=_2b_gn(r, c_rate), taxable=_2b_gn(r, c_taxable),
            igst=_2b_gn(r, c_igst), cgst=_2b_gn(r, c_cgst), sgst=_2b_gn(r, c_sgst),
            cess=_2b_gn(r, c_cess),
            itc_avail=_2b_g(r, c_itcavail), itc_avail_reason=_2b_g(r, c_reason),
            via_amendment=True, original_invno=orig_inv,
        ))
    return superseded, out_by_month


def _read_cdnra_amendments(wb):
    """Same as _read_b2ba_amendments, for 'B2B-CDNRA' (amendments to
    previously-filed credit/debit notes) against 'B2B-CDNR'. Netted the same
    way, keyed on (gstin, ORIGINAL note-number-upper)."""
    superseded, out_by_month = set(), {}
    if "B2B-CDNRA" not in wb.sheetnames:
        return superseded, out_by_month
    ws = wb["B2B-CDNRA"]
    hmap = _2b_header_map_multi(ws, header_rows=(5, 6))
    c_ntype0 = _2b_col_multi_exact(hmap, "Note type", 0)
    c_orig_note = _2b_col_multi_exact(hmap, "Note number", 0)
    c_orig_date = _2b_col_multi_exact(hmap, "Note date", 0)
    c_gstin = _2b_col_multi_exact(hmap, "GSTIN of supplier", 0)
    c_supplier = _2b_col_multi_exact(hmap, "Trade/Legal name", 0)
    c_rev_note = _2b_col_multi_exact(hmap, "Note number", 1)
    c_rev_ntype = _2b_col_multi_exact(hmap, "Note type", 1)
    c_supplytype = _2b_col_multi_exact(hmap, "Note Supply type", 0)
    c_rev_date = _2b_col_multi_exact(hmap, "Note date", 1)
    c_noteval = _2b_col_multi_contains(hmap, "notevalue", 0)
    c_pos = _2b_col_multi_exact(hmap, "Place of supply", 0)
    c_rate = _2b_col_multi_contains(hmap, "taxrate", 0)
    c_taxable = _2b_col_multi_contains(hmap, "taxablevalue", 0)
    c_igst = _2b_col_multi_exact(hmap, "Integrated Tax", 0)
    c_cgst = _2b_col_multi_exact(hmap, "Central Tax", 0)
    c_sgst = _2b_col_multi_contains(hmap, "stateut", 0)
    c_cess = _2b_col_multi_exact(hmap, "Cess", 0)
    c_period = _2b_col_multi_contains(hmap, "period", 0)
    if c_orig_note is None or c_gstin is None or c_period is None:
        raise mpu.PeriodParseError(
            "Could not locate the original-note-number / GSTIN / filing-period columns by "
            "header text in the 'B2B-CDNRA' sheet -- the column layout may have changed again.")
    for r in _data_rows(ws, first_data_row=7):
        if not any(r) or not r[c_gstin] or mpu.is_marker_row(r):
            continue
        period = _normalize_2b_row_period(r[c_period] if c_period < len(r) else None)
        gstin = _2b_g(r, c_gstin)
        orig_note = _2b_g(r, c_orig_note).strip().upper()
        if gstin and orig_note:
            superseded.add((gstin, orig_note))
        out_by_month.setdefault(period, []).append(dict(
            gstin=gstin, supplier=_2b_g(r, c_supplier),
            note=(_2b_g(r, c_rev_note) if c_rev_note is not None else orig_note) or orig_note,
            ntype=_2b_g(r, c_rev_ntype) if c_rev_ntype is not None else _2b_g(r, c_ntype0),
            supplytype=_2b_g(r, c_supplytype), date=_2b_g(r, c_rev_date),
            noteval=_2b_gn(r, c_noteval), pos=_2b_g(r, c_pos),
            rate=_2b_gn(r, c_rate), taxable=_2b_gn(r, c_taxable),
            igst=_2b_gn(r, c_igst), cgst=_2b_gn(r, c_cgst), sgst=_2b_gn(r, c_sgst),
            cess=_2b_gn(r, c_cess),
            via_amendment=True, original_note=orig_note,
        ))
    return superseded, out_by_month


_2B_FILE_CACHE = {}


def _2b_b2b_columns(hmap, path):
    c = dict(
        gstin=_2b_col_exact(hmap, "GSTIN of supplier"),
        supplier=_2b_col_exact(hmap, "Trade/Legal name"),
        invno=_2b_col_exact(hmap, "Invoice number"),
        invtype=_2b_col_exact(hmap, "Invoice type"),
        date=_2b_col_exact(hmap, "Invoice Date"),
        invval=_2b_col_contains(hmap, "invoicevalue"),
        pos=_2b_col_exact(hmap, "Place of supply"),
        rcm=_2b_col_contains(hmap, "reversecharge"),
        rate=_2b_col_exact(hmap, "Rate(%)"),
        taxable=_2b_col_contains(hmap, "taxablevalue"),
        igst=_2b_col_exact(hmap, "Integrated Tax"),
        cgst=_2b_col_exact(hmap, "Central Tax"),
        sgst=_2b_col_contains(hmap, "stateut"),
        cess=_2b_col_exact(hmap, "Cess"),
        period=_2b_col_contains(hmap, "period"),
        itcavail=_2b_col_exact(hmap, "ITC Availability"),
        reason=_2b_col_exact(hmap, "Reason"),
        source=_2b_col_exact(hmap, "Source"),
        irn=_2b_col_exact(hmap, "IRN"),
    )
    if c["period"] is None:
        raise mpu.PeriodParseError(
            f"Could not locate the GSTR-1/IFF/GSTR-5 filing-period column by header text in "
            f"the 'B2B' sheet of {path!r} -- the column layout may have changed again.")
    return c


def _2b_cdnr_columns(hmap, path):
    c = dict(
        gstin=_2b_col_exact(hmap, "GSTIN of supplier"),
        supplier=_2b_col_exact(hmap, "Trade/Legal name"),
        note=_2b_col_exact(hmap, "Note number"),
        ntype=_2b_col_exact(hmap, "Note type"),
        supplytype=_2b_col_exact(hmap, "Note Supply type"),
        date=_2b_col_exact(hmap, "Note date"),
        noteval=_2b_col_contains(hmap, "notevalue"),
        pos=_2b_col_exact(hmap, "Place of supply"),
        rate=_2b_col_exact(hmap, "Rate(%)"),
        taxable=_2b_col_contains(hmap, "taxablevalue"),
        igst=_2b_col_exact(hmap, "Integrated Tax"),
        cgst=_2b_col_exact(hmap, "Central Tax"),
        sgst=_2b_col_contains(hmap, "stateut"),
        cess=_2b_col_exact(hmap, "Cess"),
        period=_2b_col_contains(hmap, "period"),
        itcavail=_2b_col_exact(hmap, "ITC Availability"),
        reason=_2b_col_exact(hmap, "Reason"),
        source=_2b_col_exact(hmap, "Source"),
        irn=_2b_col_exact(hmap, "IRN"),
    )
    if c["period"] is None:
        raise mpu.PeriodParseError(
            f"Could not locate the GSTR-1/IFF/GSTR-5 filing-period column by header text in "
            f"the 'B2B-CDNR' sheet of {path!r} -- the column layout may have changed again.")
    return c


def _load_2b_file_data(path):
    """The expensive, MONTH-INDEPENDENT half of parsing a merged (whole-FY) GSTR-2B
    workbook: one load_workbook() call, one full read of 'ITC Available' (if present),
    'B2B' and 'B2B-CDNR' (rows + header-column map), and the whole-file amendment
    indices -- everything parse_2b_excel(path, month) needs that does NOT depend on
    `month`.

    PERFORMANCE (restores the fix from 'Fix the 10-minute full-scrutiny timeout':
    parse_2b_excel() was 24x redundant -- that cache was lost when this file was
    replaced wholesale on 2026-09-04 while adding the marker-block-scoping fix below,
    reintroducing the exact same cost this existed to remove): parse_2b_excel() is
    called once per month from several independent call sites (master_build.py's
    per-month loop AND its own FY-wide 2B invoice index, gst_checks_hsn_fraud.py,
    gst_checks_flow.py) -- all against the exact same file. Without this cache, every
    one of those calls independently repeats a full load_workbook() plus a full scan
    of every 2B sheet. Cached by path, computed once; a failed load is never cached
    (the exception propagates before the cache is written), so a genuinely broken
    file still fails on every call, matching parse_2b_excel()'s original behaviour."""
    if path in _2B_FILE_CACHE:
        return _2B_FILE_CACHE[path]
    if not path or not os.path.exists(path) or not path.lower().endswith((".xlsx", ".xlsm")):
        raise mpu.PeriodParseError(f"Not a GSTR-2B Excel file: {path!r}")
    wb = openpyxl.load_workbook(path, data_only=True)

    # BUG FIX (found on a real taxpayer's export): 'ITC Available' missing used to be
    # fatal for the whole file, even though B2B/B2B-CDNR parsing is structurally
    # independent of it (each row carries its own columns; nothing about parsing them
    # needs the summary sheet's marker blocks). Some real GSTR-2B exports genuinely
    # only carry 'B2B'/'B2B-CDNR'. `itc_rows=None` here means "summary not available",
    # handled per-month below.
    itc_rows = list(wb["ITC Available"].iter_rows(values_only=True)) if "ITC Available" in wb.sheetnames else None

    # ---------- amendment indices (whole-file, not month-scoped -- bug report §7) ----------
    superseded_inv, b2ba_by_month = _read_b2ba_amendments(wb)
    superseded_note, cdnra_by_month = _read_cdnra_amendments(wb)

    b2b_all_rows, b2b_cols = None, None
    if "B2B" in wb.sheetnames:
        ws_b2b = wb["B2B"]
        b2b_cols = _2b_b2b_columns(_2b_header_map(ws_b2b), path)
        b2b_all_rows = list(ws_b2b.iter_rows(values_only=True))

    cdnr_all_rows, cdnr_cols = None, None
    if "B2B-CDNR" in wb.sheetnames:
        ws_cdnr = wb["B2B-CDNR"]
        cdnr_cols = _2b_cdnr_columns(_2b_header_map(ws_cdnr), path)
        cdnr_all_rows = list(ws_cdnr.iter_rows(values_only=True))

    data = dict(
        itc_rows=itc_rows,
        superseded_inv=superseded_inv, b2ba_by_month=b2ba_by_month,
        superseded_note=superseded_note, cdnra_by_month=cdnra_by_month,
        b2b_all_rows=b2b_all_rows, b2b_cols=b2b_cols,
        cdnr_all_rows=cdnr_all_rows, cdnr_cols=cdnr_cols,
    )
    _2B_FILE_CACHE[path] = data
    return data


def parse_2b_excel(path, month):
    """Return dict(summary=..., b2b=[...], cdnr=[...], available=True) for ONE
    month out of the merged (whole-FY) GSTR-2B workbook.

    B2B and B2B-CDNR include amendment-adjusted figures (bug report §7): a
    row superseded by a later B2BA/B2B-CDNRA entry is excluded wherever its
    stale original sits, and the amendment's own revised row is spliced in
    under the amendment's own filing period instead -- see
    _read_b2ba_amendments()/_read_cdnra_amendments() for why both halves of
    that (exclude AND re-add, not just add) are necessary.

    The expensive, month-independent parsing (one load_workbook(), one scan
    each of every 2B sheet) happens once per file in _load_2b_file_data();
    this function just does the cheap per-month lookup against that cache."""
    d = _load_2b_file_data(path)

    # ---------- Summary (Table 3), quarter-block-scoped ----------
    if d["itc_rows"] is None:
        summary = dict(
            ITC_all_other_IGST=0.0, ITC_all_other_CGST=0.0, ITC_all_other_SGST=0.0, ITC_all_other_CESS=0.0,
            ITC_rcm_IGST=0.0, ITC_rcm_CGST=0.0, ITC_rcm_SGST=0.0, ITC_rcm_CESS=0.0,
            CN_IGST=0.0, CN_CGST=0.0, CN_SGST=0.0, CN_CESS=0.0,
            _qtr_total_mismatch=None, available=False,
            _reason="'ITC Available' sheet not present in this GSTR-2B export -- summary "
                    "control-total not computable from this file; invoice-level B2B/B2B-CDNR "
                    "figures below are unaffected and are what every real check in this tool uses.",
        )
    else:
        start, end, group_index, group_count = mpu.find_block_and_index_for_month(d["itc_rows"], month)
        summary = _summary_from_block(d["itc_rows"][start:end], group_index=group_index, group_count=group_count)

    # ---------- B2B invoice list, scoped by the MARKER BLOCK for `month` ----------
    # BUG FIX (bug report #1, ITC Annual Summary wrong every month -- confirmed on real data):
    # this used to filter the WHOLE sheet (every month's rows at once) by matching each row's
    # own 'GSTR-1/IFF/GSTR-5 Period' column against `month` -- i.e. it trusted the INVOICE's
    # own reporting period, not which 2B statement (merged-file marker block) the row actually
    # sits under. That is wrong for "ITC available in 2B this month": Rule 36(4)/Sec 16(2)(aa)
    # eligibility is governed by which of the RECIPIENT's own monthly 2B statements first shows
    # the invoice, not by the invoice's own original filing period -- a March invoice the
    # supplier files late, that only appears in the recipient's April 2B, is available to the
    # recipient in April, not March. Confirmed on this taxpayer's real April 2B: 52 rows
    # genuinely present in April's own downloaded 2B (physically under the April marker block)
    # carry an internal period tag of "Mar'23" and were being silently DROPPED from every month
    # in this FY (Mar-23 isn't even a valid month in an Apr-23..Mar-24 FY, so a row tagged that
    # way could never match ANY month query) -- while 43 rows physically sitting in a LATER
    # month's block (tagged "Apr'23" internally, e.g. a correction reported the following month)
    # were being pulled INTO April's figures instead of counting toward the month they actually
    # appeared in. Net effect on April alone: understated 2B-available ITC by Rs 21,366.02 --
    # confirmed against a manual invoice-by-invoice count of the raw April file (1222 rows, 1
    # "No"/1221 "Yes", matching exactly). Fixed: scope to the marker block for `month` (i.e.
    # exactly the rows physically present in that month's own downloaded 2B file, matching a
    # manual count of that file), never re-filtered by the row's own internal period tag. The
    # per-row period is still read into `filed_period` below (best-effort) as reference data --
    # useful for spotting late-reported invoices -- but no longer gates inclusion.
    #
    # NOTE: for a QUARTERLY GSTR-2B filer, one marker block covers all 3 months of the quarter,
    # so this month-scoped `b2b`/`cdnr` list is identical across those 3 months (the whole
    # quarter, every time) -- correct for ITC-availability timing (the question this scoping was
    # built for), but NOT the right list to match a specific calendar month's EWB movements
    # against (see gst_checks_monthly.run()'s own per-row date filter for that -- checks #10-#13
    # further narrow this list to each row's own invoice/note date before matching).
    b2b = []
    if d["b2b_cols"] is not None:
        c = d["b2b_cols"]
        blk_start, blk_end = mpu.find_block_for_month(d["b2b_all_rows"], month)
        for r in d["b2b_all_rows"][blk_start:blk_end]:
            if not any(r) or not r[0] or mpu.is_marker_row(r):
                continue
            shift = _2b_row_rate_shift(r, c["period"], c["rate"])
            eff_period_col = c["period"] + shift
            try:
                filed_period = _normalize_2b_row_period(
                    r[eff_period_col] if eff_period_col < len(r) else None)
            except mpu.PeriodParseError:
                filed_period = None
            gstin_val = _2b_g(r, c["gstin"])
            invno_val = _2b_g(r, c["invno"])
            if (gstin_val, invno_val.strip().upper()) in d["superseded_inv"]:
                # This exact original row has since been amended (B2BA carries a later, revised
                # figure for it, spliced in below under the AMENDMENT's own period) -- counting
                # this stale row too would double-count the invoice. See _read_b2ba_amendments().
                continue
            b2b.append(dict(
                gstin=gstin_val, supplier=_2b_g(r, c["supplier"]),
                invno=invno_val, invtype=_2b_g(r, c["invtype"]),
                date=_2b_g(r, c["date"]), invval=_2b_gn(r, c["invval"]),
                pos=_2b_g(r, c["pos"]), rcm=_2b_g(r, c["rcm"]),
                rate=(None if shift else _2b_gn(r, c["rate"])),
                taxable=_2b_gn(r, c["taxable"] + shift),
                igst=_2b_gn(r, c["igst"] + shift), cgst=_2b_gn(r, c["cgst"] + shift), sgst=_2b_gn(r, c["sgst"] + shift),
                cess=_2b_gn(r, c["cess"] + shift),
                itc_avail=_2b_g(r, c["itcavail"] + shift), itc_avail_reason=_2b_g(r, c["reason"] + shift),
                # 2B's own 'Source' column (== "E-Invoice" with a real IRN when the SUPPLIER
                # e-invoiced this row) -- used by gst_checks_monthly's #28. Named 'einv_source',
                # not 'source': gst_checks_flow.py does `dict(month=month, source="B2B", **x)` on
                # these same dicts elsewhere, and a plain 'source' key here collided with that
                # literal kwarg (TypeError: multiple values for 'source' -- confirmed on real data).
                einv_source=(_2b_g(r, c["source"] + shift) if c["source"] is not None else ""),
                irn=(_2b_g(r, c["irn"] + shift) if c["irn"] is not None else ""),
                filed_period=filed_period, via_amendment=False,
            ))
    # splice in this month's amended (revised) invoice rows -- see _read_b2ba_amendments()
    b2b.extend(d["b2ba_by_month"].get(month, []))

    # ---------- B2B-CDNR (credit/debit notes), scoped by the MARKER BLOCK for `month` ----------
    # BUG FIX (bug report #1, same root cause as B2B above): rows are now scoped to the marker
    # block for `month` (i.e. exactly what's physically in that month's own downloaded 2B file),
    # never re-filtered by the row's own internal period tag. The period-tag/date-fallback logic
    # below is KEPT, but repurposed: it now only fills `filed_period` for reference (useful for
    # spotting a note reported under a different period than the 2B statement it appears in) and
    # feeds `cdnr_skipped` as an informational data-quality count -- it no longer excludes the
    # row from this month's totals, since the row's inclusion is now decided by the marker block
    # it's physically in, not by whether its own period tag could be read.
    cdnr = []
    cdnr_skipped = 0
    if d["cdnr_cols"] is not None:
        c = d["cdnr_cols"]
        blk_start, blk_end = mpu.find_block_for_month(d["cdnr_all_rows"], month)
        for r in d["cdnr_all_rows"][blk_start:blk_end]:
            if not any(r) or not r[0] or mpu.is_marker_row(r):
                continue
            shift = _2b_row_rate_shift(r, c["period"], c["rate"])
            eff_period_col = c["period"] + shift
            v = r[eff_period_col] if eff_period_col < len(r) else None
            if not (isinstance(v, str) and _2B_PERIOD_TAG_RE.match(v.strip())):
                # Some B2B-CDNR rows have a MORE complex misalignment than the B2B sheet's clean
                # 'Rate(%) missing, shift everything else by one' pattern -- the period tag isn't
                # sitting where shift alone would put it. Try recovering it from this row's own
                # Note Date column instead (same fallback as before this fix); this is now only
                # used to populate the reference field below, not to decide inclusion.
                v_from_date = _date_to_2b_period(r[c["date"]] if c["date"] is not None and c["date"] < len(r) else None)
                v = v_from_date if v_from_date else None
            if v is None:
                cdnr_skipped += 1
                filed_period = None
            else:
                try:
                    filed_period = _normalize_2b_row_period(v)
                except mpu.PeriodParseError:
                    filed_period = None
            gstin_val = _2b_g(r, c["gstin"])
            note_val = _2b_g(r, c["note"])
            if (gstin_val, note_val.strip().upper()) in d["superseded_note"]:
                # Superseded by a later B2B-CDNRA entry -- see _read_cdnra_amendments().
                continue
            cdnr.append(dict(
                gstin=gstin_val, supplier=_2b_g(r, c["supplier"]),
                note=note_val, ntype=_2b_g(r, c["ntype"]),
                supplytype=_2b_g(r, c["supplytype"]), date=_2b_g(r, c["date"]),
                noteval=_2b_gn(r, c["noteval"]), pos=_2b_g(r, c["pos"]),
                rate=(None if shift else _2b_gn(r, c["rate"])),
                taxable=_2b_gn(r, c["taxable"] + shift),
                igst=_2b_gn(r, c["igst"] + shift), cgst=_2b_gn(r, c["cgst"] + shift), sgst=_2b_gn(r, c["sgst"] + shift),
                cess=_2b_gn(r, c["cess"] + shift),
                itc_avail=(_2b_g(r, c["itcavail"] + shift) if c["itcavail"] is not None else ""),
                itc_avail_reason=(_2b_g(r, c["reason"] + shift) if c["reason"] is not None else ""),
                einv_source=(_2b_g(r, c["source"] + shift) if c["source"] is not None else ""),
                irn=(_2b_g(r, c["irn"] + shift) if c["irn"] is not None else ""),
                filed_period=filed_period, via_amendment=False,
            ))
    # splice in this month's amended (revised) note rows -- see _read_cdnra_amendments()
    cdnr.extend(d["cdnra_by_month"].get(month, []))

    # Only defaults to True if not already explicitly set (the 'ITC Available' sheet missing
    # branch above sets it to False and must not be clobbered back to True here).
    summary.setdefault("available", True)
    summary["cdnr_skipped_unparseable_this_month"] = cdnr_skipped
    return dict(summary=summary, b2b=b2b, cdnr=cdnr, available=True)


_ZERO_SUMMARY_KEYS = (
    "ITC_all_other_IGST", "ITC_all_other_CGST", "ITC_all_other_SGST", "ITC_all_other_CESS",
    "ITC_rcm_IGST", "ITC_rcm_CGST", "ITC_rcm_SGST", "ITC_rcm_CESS",
    "CN_IGST", "CN_CGST", "CN_SGST", "CN_CESS",
)


def summary_for_month(path, month):
    """Return the 2B summary dict for `month`, read directly from the merged
    GSTR-2B workbook.

    GRACEFUL DEGRADATION (fixed -- previously this raised PeriodParseError
    with NOTHING catching it anywhere in run_monthly_pipeline.py or
    master_build.py's per-month loop, so a taxpayer/month with no GSTR-2B
    supplied crashed the ENTIRE run, not just that month's 2B-dependent
    checks. GSTR-2B is auto-generated by GSTN and usually present, but a
    small/new taxpayer's early months, or a partial upload, can genuinely
    lack it -- exactly the 'limited data must not error out' requirement.

    Now: if `path` is missing, or doesn't cover `month`, returns a summary
    with available=False and every numeric field as None (NOT zero -- zero
    would look like a real, verified nil balance and get diffed against
    GSTR-3B as if it were data, producing a wall of false MISMATCH rows;
    None makes 'not available' visually and programmatically distinct from
    'available and nil'). Callers (gst_scrutiny_tool.build_comparisons(),
    gst_eway_recon.run()) must check summary.get('available') before using
    any numeric field -- both have been updated to do so; see their own
    docstrings for the resulting INFO/SKIP behaviour instead of a false
    numeric compare."""
    try:
        parsed = parse_2b_excel(path, month)
    except mpu.PeriodParseError as ex:
        s = {k: None for k in _ZERO_SUMMARY_KEYS}
        s["available"] = False
        s["_reason"] = str(ex)
        s["_source"] = "unavailable"
        s["_file"] = os.path.basename(path) if path else None
        s["_lines"] = None
        return s
    s = dict(parsed["summary"])
    # BUG FIX (found via a real taxpayer's monthly sheets all wrongly saying "GSTR-2B not
    # supplied" despite a full, correctly-parsed GSTR-2B Excel being supplied for the whole
    # year): parsed['summary']['available'] is a NARROWER flag added later -- it means
    # specifically 'the ITC Available summary/control-total sheet could not be read' (some real
    # GSTR-2B exports genuinely lack that sheet while still carrying full B2B/B2B-CDNR
    # invoice-level data -- see parse_2b_excel's own docstring). Blindly copying that narrower
    # flag into s['available'] made every caller that reads THIS function's own 'available' key
    # for the ORIGINAL, broader question -- 'was GSTR-2B supplied for this month at all' --
    # wrongly conclude it wasn't, even though b2b/cdnr rows were sitting right there, correctly
    # parsed. Fixed: s['available'] now reflects whether parse_2b_excel succeeded at all (True
    # here, since reaching this line means it did) -- the narrower "was the summary/control-total
    # itself readable" question is preserved separately under s['_summary_available'] for the one
    # caller (ITC 3B vs 2B's F7a) that actually needs that specific distinction.
    s["_summary_available"] = s.get("available", True)
    s["available"] = True
    s["_source"] = "excel"
    s["_file"] = os.path.basename(path)
    s["_lines"] = parsed
    if not s["_summary_available"]:
        # FALLBACK (per explicit instruction): when a taxpayer's GSTR-2B export has no 'ITC
        # Available'/'ITC not available' summary sheet (document-wise export -- only B2B and
        # B2B-CDNR), those two summary sheets' own figures are themselves nothing more than a
        # sum of the SAME invoice-level B2B rows this tool already reads, grouped by the row's
        # own 'ITC Availability' column ('ITC Available' = the Yes/Unconfirmed rows; 'ITC not
        # available' = the No rows) -- so rather than leaving 'All other ITC' uncomputable,
        # compute the identical figure directly from parsed['b2b']/['cdnr'] (both already
        # correctly scoped to this exact month -- see parse_2b_excel's own marker-block fix).
        # 'Yes'+'Unconfirmed' matches this tool's existing convention everywhere else it makes
        # this exact distinction (an unconfirmed row is not assumed ineligible, since nothing
        # states it is) -- confirmed rows only are excluded, matching the portal's own 'ITC not
        # available' sheet.
        def _yes_unconf(rows):
            return [r for r in rows if str(r.get("itc_avail", "")).strip().upper() in ("YES", "UNCONFIRMED")]
        b2b_ok = _yes_unconf(parsed.get("b2b", []))
        s["ITC_all_other_IGST"] = round(sum(r["igst"] for r in b2b_ok), 2)
        s["ITC_all_other_CGST"] = round(sum(r["cgst"] for r in b2b_ok), 2)
        s["ITC_all_other_SGST"] = round(sum(r["sgst"] for r in b2b_ok), 2)
        s["ITC_all_other_CESS"] = round(sum(r["cess"] for r in b2b_ok), 2)
        cdnr_ok = _yes_unconf(parsed.get("cdnr", []))
        cn_rows = [r for r in cdnr_ok if str(r.get("ntype", "")).strip().upper().startswith("C")]
        dn_rows = [r for r in cdnr_ok if str(r.get("ntype", "")).strip().upper().startswith("D")]
        # CN_IGST etc. are the NET amount Section D SUBTRACTS from ITC_all_other (i.e. positive
        # = net reduction) -- Credit Notes reduce available ITC, Debit Notes restore/add to it,
        # so CN total minus DN total gives exactly the net subtraction Section D's own formula
        # (ITC_all_other - CN_IGST) expects.
        s["CN_IGST"] = round(sum(r["igst"] for r in cn_rows) - sum(r["igst"] for r in dn_rows), 2)
        s["CN_CGST"] = round(sum(r["cgst"] for r in cn_rows) - sum(r["cgst"] for r in dn_rows), 2)
        s["CN_SGST"] = round(sum(r["sgst"] for r in cn_rows) - sum(r["sgst"] for r in dn_rows), 2)
        s["CN_CESS"] = round(sum(r["cess"] for r in cn_rows) - sum(r["cess"] for r in dn_rows), 2)
        # Section C (RCM, GSTR-3B 3.1(d)/4(A)(3) vs GSTR-2B) has the identical exposure --
        # ITC_rcm_IGST/CGST/SGST also only ever came from the summary sheet, defaulting to a
        # fabricated 0.0 the same way. The B2B sheet carries its own per-row Reverse Charge
        # flag (captured as 'rcm' by parse_2b_excel already), so RCM-liable rows within the
        # SAME Yes/Unconfirmed B2B rows already selected above give the identical fallback.
        rcm_rows = [r for r in b2b_ok if str(r.get("rcm", "")).strip().upper().startswith("Y")]
        s["ITC_rcm_IGST"] = round(sum(r["igst"] for r in rcm_rows), 2)
        s["ITC_rcm_CGST"] = round(sum(r["cgst"] for r in rcm_rows), 2)
        s["ITC_rcm_SGST"] = round(sum(r["sgst"] for r in rcm_rows), 2)
        s["ITC_rcm_CESS"] = round(sum(r["cess"] for r in rcm_rows), 2)
        s["_summary_computed_from_invoice_level"] = True
    return s




# ============================================================================
# ==== SECTION: ewb_annual_parser.py  (was a standalone module before consolidation)
# ============================================================================
"""
ANNUAL E-WAY-BILL PARSER
=========================
The whole-FY Inward/Outward EWB workbooks (one file each, not one-per-month)
don't follow one consistent sheet-name or header-text convention:
  - Outward file: real data on sheet 'OUT EWB', doc-no/doc-date columns are
    literally named 'INVOICE'/'DATE'. A second sheet 'R1' is a bonus annual
    B2B invoice register (not EWB data) -- kept separately if useful later.
  - Inward file: sheet 'Sheet1' is empty; real data is on 'merged_sheet_1',
    where the doc-no/doc-date columns are mislabelled 'Doc'/'&' (a broken
    merged-cell header from the source export).

Despite the text differences, both files share the SAME column order for the
first 13 columns. So this parser finds the right sheet by header CONTENT
(must contain 'EWB No.' + 'From GSTIN & Name' + 'To GSTIN & Name'), then
reads the doc-number/doc-date pair by POSITION (the two columns immediately
after 'EWB No. & Dt.'), not by their (unreliable) header text.

Every row also gets a `month` tag ('Mon-YY') derived from the EWB date, so a
per-month engine can filter this annual list down to one period.
"""

import re
import datetime as _dt
import openpyxl

GSTIN_RE = re.compile(r"(\d{2}[A-Z]{5}\d{4}[A-Z]\dZ[A-Z\d])")
MONTH_ABBR = {1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
              7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"}


def _num_ewb(v):
    if v is None:
        return 0.0
    s = str(v).replace(",", "").strip()
    if s in ("", "-", "#N/A", "NA"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _gstin_of(cell):
    if not cell:
        return ""
    m = GSTIN_RE.search(str(cell))
    return m.group(1) if m else str(cell).split("/")[0].strip()


def _as_date(v):
    if isinstance(v, _dt.datetime):
        return v.date()
    if isinstance(v, _dt.date):
        return v
    if isinstance(v, str):
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
            try:
                return _dt.datetime.strptime(v.strip()[:10], fmt).date()
            except ValueError:
                continue
    return None


def _split_ewb_no_dt(v):
    """'351569103816 - 04/03/2023 11:43:00' -> (ewb_no, date, time).
    BUG FIX (found while implementing the EWB-vs-IRN timing check, per explicit request): the
    source text for 'EWB No. & Dt.' genuinely carries a time-of-day (confirmed on this
    taxpayer's real export -- e.g. '341599265378 - 12/05/2023 13:13:00'), but this function
    used to discard it, returning only the date. That silently starved an EXISTING check
    (gst_checks_hsn_fraud.py's same-window EWB clustering check, which reads e.get('ewbtime'))
    of any real data -- e.get('ewbtime') was always None, so that check could never fire.
    Now returns the time too (a THIRD tuple element, so every existing caller that only ever
    unpacked (ewbno, date) two-at-a-time still works unchanged) -- callers that want it store
    it under a separate 'ewbtime' key rather than folding it into 'ewbdate', so every existing
    piece of `.days`-based date arithmetic on ewbdate/docdate (several real checks depend on
    date-minus-date subtraction, which breaks if either side becomes a datetime) is completely
    unaffected. 'Doc No. & Dt.' (the other caller) genuinely carries no time in this data
    source -- confirmed on the raw file -- so its own third return value is simply always None,
    correctly reflecting that it was never available, not a parsing failure."""
    if not v:
        return "", None, None
    s = str(v)
    parts = s.split(" - ", 1)
    ewbno = parts[0].strip()
    date = None
    time = None
    if len(parts) > 1:
        m = re.search(r"(\d{2})/(\d{2})/(\d{4})", parts[1])
        if m:
            dd, mm, yyyy = m.groups()
            date = _dt.date(int(yyyy), int(mm), int(dd))
        mt = re.search(r"(\d{1,2}):(\d{2})(?::(\d{2}))?", parts[1])
        if mt:
            hh, mi, ss = mt.groups()
            try:
                time = _dt.time(int(hh), int(mi), int(ss or 0))
            except ValueError:
                time = None
    return ewbno, date, time


def _find_data_sheet(wb):
    """Return worksheet whose header row matches the EWB column signature."""
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            hdr = [str(c).strip() if c else "" for c in row]
            if "EWB No." in hdr and "From GSTIN & Name" in hdr and "To GSTIN & Name" in hdr:
                return ws, hdr, row
    return None, None, None


def dominant_hsn_is_services(hsn_rows):
    """True if the single largest-taxable-value HSN/SAC code across `hsn_rows` (a list of the
    dict shape read_gstr1_hsn_all_months() returns, or any list of dicts with 'hsn'/'taxable'
    keys) is a SERVICES code (SAC, which GSTN prefixes '99') rather than a goods HSN. Used by
    Bug-4's fix: e-way bills apply to goods movement only (Rule 138) -- a services-dominant
    business genuinely has little/nothing to move, so a low EWB-value/invoice-value ratio is
    expected for that HSN profile, not a red flag on its own. Returns (is_services, dominant_hsn,
    dominant_share) -- dominant_share is the fraction of total taxable value the dominant code
    represents, so the caller can cite it."""
    if not hsn_rows:
        return False, None, 0.0
    by_hsn = {}
    for r in hsn_rows:
        by_hsn[r["hsn"]] = by_hsn.get(r["hsn"], 0.0) + (r.get("taxable") or 0.0)
    total = sum(by_hsn.values())
    if total <= 0:
        return False, None, 0.0
    dominant_hsn, dominant_val = max(by_hsn.items(), key=lambda kv: kv[1])
    return (str(dominant_hsn).strip().startswith("99"), dominant_hsn, dominant_val / total)


def parse_annual_ewb(path):
    """Return list of dicts: ewbno, ewbdate, month, docno, docdate, from_gstin,
    to_gstin, assess, taxval, hsn, hsn_desc, vehicle, rate.

    BUG FIX (confirmed against real outward/inward EWB exports -- this was THE largest source of
    false positives across checks #1/#3/#5/#12/#13/#17): docno/docdate used to be computed by
    POSITIONAL OFFSET from the 'EWB No. & Dt.' column (idx+1 for doc number, idx+2 for doc
    date), assuming a 3-column layout (EWB-combined, Doc-No-only, Doc-Date-only). The real file
    has only a 2-column layout: 'EWB No. & Dt.' (combined) and 'Doc No. & Dt.' (ALSO combined,
    e.g. 'TCP/DC/25-26/029 - 20/06/2025') -- there is no separate Doc-Date column at all. The
    offset happened to land doc_no_col on the right COLUMN by coincidence (idx+1), but never
    split the combined string, so docno silently retained its ' - DD/MM/YYYY' suffix forever
    (exact-string matching against GSTR-1/2B/E-Invoice's clean invoice numbers then failed
    almost every time -- the ~0% match rate every single month, all year, was the tell). The
    doc_dt_col offset (idx+2) was worse: it landed on 'Assess Val.' (a number), not a date, so
    docdate was always None.

    Fixed the same way 'EWB No. & Dt.' already was: looked up BY NAME ('Doc No. & Dt.'), split
    with the same _split_ewb_no_dt() helper -- never a positional offset from a DIFFERENT
    column."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws, hdr, _ = _find_data_sheet(wb)
    if ws is None:
        return []
    H = {h: i for i, h in enumerate(hdr) if h}

    def g(r, name):
        i = H.get(name)
        return r[i] if i is not None and i < len(r) else None

    out = []
    rows = ws.iter_rows(min_row=1, values_only=True)
    next(rows)  # header
    for r in rows:
        if not any(r):
            continue
        if str(r[0] or "").strip() in ("EWB No.", ""):
            continue
        ewbno, ewbdate, ewbtime = _split_ewb_no_dt(g(r, "EWB No. & Dt."))
        docno, docdate, _docdate_time = _split_ewb_no_dt(g(r, "Doc No. & Dt."))
        docno = docno.strip().upper()   # per the brief: cleaned doc_number, uppercased, for matching
        month = f"{MONTH_ABBR.get(ewbdate.month,'?')}-{str(ewbdate.year)[2:]}" if ewbdate else None
        out.append(dict(
            ewbno=str(g(r, "EWB No.") or ewbno).strip(), ewbdate=ewbdate, ewbtime=ewbtime, month=month,
            docno=docno, docdate=docdate,
            from_gstin=_gstin_of(g(r, "From GSTIN & Name")),
            from_name=str(g(r, "From GSTIN & Name") or "").split("/", 1)[-1].strip(),
            to_gstin=_gstin_of(g(r, "To GSTIN & Name")),
            to_name=str(g(r, "To GSTIN & Name") or "").split("/", 1)[-1].strip(),
            from_place=str(g(r, "From Place & Pin") or "").strip(),
            to_place=str(g(r, "To Place & Pin") or "").strip(),
            assess=_num_ewb(g(r, "Assess Val.")), taxval=_num_ewb(g(r, "Tax Val.")),
            hsn=str(g(r, "HSN Code") or "").strip(),
            hsn_desc=str(g(r, "HSN Desc.") or "").strip(),
            vehicle=str(g(r, "Latest Vehicle No.") or "").strip(),
            rate=_num_ewb(g(r, "TAX RATE")),
        ))
    return out


def filter_by_month(ewb_rows, month_key):
    """month_key e.g. 'Jan-23' -- matches on EWB date's month (not doc date)."""
    return [r for r in ewb_rows if r["month"] == month_key]




# ============================================================================
# ==== SECTION: amendments.py  (was a standalone module before consolidation)
# ============================================================================
"""
AMENDMENTS + DOC-SERIES  (new checks enabled by fields the earlier single-
month tool never read: GSTR-1's b2ba/cdnra/b2csa/expa amendment sheets, and
the 'docs' sheet = Table 13, Summary of Documents Issued).

A GSTR-1 filed in month N can contain amendment rows that correct an invoice
originally reported in an EARLIER month M (via 'Original Invoice Number' /
'Original Invoice date'). This is the direct evidence for the "error in an
earlier month, corrected later" requirement -- read these sheets across every
month you have and match Original -> that original month's B2B line.
"""

import openpyxl
import gst_core as mpu


def _num_amd(v):
    if v is None:
        return 0.0
    try:
        return float(str(v).replace(",", ""))
    except ValueError:
        return 0.0


def _hdr_row_idx(rows, must_contain):
    for i, r in enumerate(rows):
        cells = [str(c).strip() if c else "" for c in r]
        if all(any(m.lower() in c.lower() for c in cells) for m in must_contain):
            return i
    return None


def parse_b2ba(path, month):
    """9A amendment sheet: corrections to B2B invoices reported in an earlier
    period, scoped to ONE month's block out of the merged workbook.
    BUG FIX -- same root cause as the b2b/cdnr/b2cl continuation-row fixes: a multi-rate
    amendment leaves every identity column (GSTIN, Receiver Name, Original/Revised Invoice
    Number+date) blank on its continuation rows. Confirmed on real data (4 such rows in this
    taxpayer's FY, e.g. invoice CNN24A000302 amended with both an 18% and a 28% line, only the
    first of which used to carry the invoice identity). This feeds both the B2 comparison's
    amendment-aware splicing AND the Rectification Pairs sheet, so an unfixed continuation row
    here would understate a revised invoice's value in both places. Forward-filled the same way."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if "b2ba" not in wb.sheetnames:
        return []
    rows = list(wb["b2ba"].iter_rows(values_only=True))
    hi = _hdr_row_idx(rows, ["Original Invoice Number", "Revised Invoice Number"])
    if hi is None:
        return []
    hdr = [str(c).strip() if c else "" for c in rows[hi]]
    H = {h: i for i, h in enumerate(hdr)}
    out = []
    last = None  # (gstin, recipient, orig_invno, orig_date, revised_invno, revised_date, invval, pos)
    for r in mpu.rows_for_month(rows, hi, month):
        if not r or not any(r):
            continue
        g = lambda k: r[H[k]] if k in H and H[k] < len(r) else None
        raw_orig = g("Original Invoice Number")
        raw_gstin = g("GSTIN/UIN of Recipient")
        if raw_orig not in (None, "") and raw_gstin not in (None, ""):
            last = (str(raw_gstin).strip(), str(g("Receiver Name") or "").strip(),
                    str(raw_orig).strip(), g("Original Invoice date"),
                    str(g("Revised Invoice Number") or "").strip(), g("Revised Invoice date"),
                    _num_amd(g("Invoice Value")), str(g("Place Of Supply") or "").strip())
        elif last is None:
            continue  # no primary row seen yet this month -- can't attribute, skip rather than guess
        gstin_v, recipient_v, orig_invno_v, orig_date_v, revised_invno_v, revised_date_v, invval_v, pos_v = last
        out.append(dict(
            gstin=gstin_v,
            recipient=recipient_v,
            orig_invno=orig_invno_v,
            orig_date=orig_date_v,
            revised_invno=revised_invno_v,
            revised_date=revised_date_v,
            invval=invval_v, pos=pos_v,
            rate=_num_amd(g("Rate")), taxable=_num_amd(g("Taxable Value")),
            igst=_num_amd(g("Integrated Tax")), cgst=_num_amd(g("Central Tax")),
            sgst=_num_amd(g("State/UT Tax")), cess=_num_amd(g("Cess Amount")),
        ))
    return out


def parse_cdnra(path, month):
    """9C amendment sheet: corrections to credit/debit notes reported earlier,
    scoped to ONE month's block out of the merged workbook.
    BUG FIX -- same root cause as parse_b2ba() above; confirmed on real data (e.g. note
    VOU24A000545 amended with both an 18% and a 28% line). Forward-filled the same way."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if "cdnra" not in wb.sheetnames:
        return []
    rows = list(wb["cdnra"].iter_rows(values_only=True))
    hi = _hdr_row_idx(rows, ["Original Note Number", "Revised Note Number"])
    if hi is None:
        return []
    hdr = [str(c).strip() if c else "" for c in rows[hi]]
    H = {h: i for i, h in enumerate(hdr)}
    out = []
    last = None  # (gstin, orig_noteno, orig_date, revised_noteno, revised_date, note_type)
    for r in mpu.rows_for_month(rows, hi, month):
        if not r or not any(r):
            continue
        g = lambda k: r[H[k]] if k in H and H[k] < len(r) else None
        raw_orig = g("Original Note Number")
        raw_gstin = g("GSTIN/UIN of Recipient")
        if raw_orig not in (None, "") and raw_gstin not in (None, ""):
            last = (str(raw_gstin).strip(), str(raw_orig).strip(), g("Original Note Date"),
                    str(g("Revised Note Number") or "").strip(), g("Revised Note Date"),
                    str(g("Note Type") or "").strip())
        elif last is None:
            continue
        gstin_v, orig_noteno_v, orig_date_v, revised_noteno_v, revised_date_v, note_type_v = last
        out.append(dict(
            gstin=gstin_v,
            orig_noteno=orig_noteno_v,
            orig_date=orig_date_v,
            revised_noteno=revised_noteno_v,
            revised_date=revised_date_v,
            note_type=note_type_v,
            taxable=_num_amd(g("Taxable Value")), igst=_num_amd(g("Integrated Tax")),
            cgst=_num_amd(g("Central Tax")), sgst=_num_amd(g("State/UT Tax")),
        ))
    return out


def parse_docs(path, month):
    """Table 13: Summary of Documents Issued, scoped to ONE month's block out
    of the merged workbook. Returns list of dicts for gap analysis against B2B."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if "docs" not in wb.sheetnames:
        return []
    rows = list(wb["docs"].iter_rows(values_only=True))
    hi = _hdr_row_idx(rows, ["Sr. No. From", "Sr. No. To"])
    if hi is None:
        return []
    hdr = [str(c).strip() if c else "" for c in rows[hi]]
    H = {h: i for i, h in enumerate(hdr)}
    out = []
    for r in mpu.rows_for_month(rows, hi, month):
        if not r or not r[H.get("Nature of Document", 0)]:
            continue
        g = lambda k: r[H[k]] if k in H and H[k] < len(r) else None
        out.append(dict(
            nature=str(g("Nature of Document") or "").strip(),
            sr_from=str(g("Sr. No. From") or "").strip(),
            sr_to=str(g("Sr. No. To") or "").strip(),
            total=_num_amd(g("Total Number")), cancelled=_num_amd(g("Cancelled")),
        ))
    return out


def _split_series(invno):
    """'MR22-23/509' -> ('MR22-23/', 509). Returns (None, None) if not numeric-suffixed."""
    import re
    m = re.match(r"^(.*?)(\d+)$", invno.strip())
    if not m:
        return None, None
    return m.group(1), int(m.group(2))


def _normalize_prefix(p):
    """Strip slashes/hyphens/spaces and uppercase, for punctuation-tolerant prefix matching.
    'JWI/22-23/' and 'JWI22-23/' both normalize to 'JWI2223' -- confirmed necessary against
    the real file, where Table 13's own printed range header and the actual invoice numbers
    for the SAME series don't always agree on punctuation (this is a real inconsistency in
    the source export, not a code bug to silently paper over -- see doc_series_gap_check)."""
    import re
    return re.sub(r"[^A-Za-z0-9]", "", p or "").upper()


def doc_series_gap_check(docs, actual_invoice_numbers):
    """For each 'Invoices for outward supply' Sr-No range in Table 13, find
    which serials in that range are MISSING from the actual B2B invoice list
    (and not accounted for by the 'Cancelled' count).

    Matching is two-tier, because the real source file is not internally
    consistent about series-prefix punctuation:
      Tier 1 -- exact prefix match (fast path, works for most series).
      Tier 2 -- punctuation-normalized prefix match (handles 'JWI/22-23/'
                declared vs 'JWI22-23/' actually used -- same series).
      Tier 3 -- fuzzy fallback: does any actual invoice's normalized text
                CONTAIN the declared series' normalized prefix, with the
                right trailing number? (handles 'MR/JWI/22-23/001' actually
                used for a series Table 13 declares as 'JWI/22-23/001' --
                confirmed real, in the April data.) Numbers resolved only at
                this tier are marked so the difference stays visible rather
                than silently treated as an exact match.

    A number still unresolved after all three tiers is genuinely reported as
    missing UNLESS the count of such numbers exactly equals Table 13's own
    declared 'Cancelled' count for that range, in which case it's reported as
    explained (not a real gap) -- still shown, not hidden, just not painted
    as an open question when the source document already accounts for it.
    """
    findings = []
    actual_by_prefix = {}      # exact prefix -> {numbers}
    actual_by_norm_prefix = {} # normalized prefix -> {numbers}
    actual_norm_full = []      # [(normalized full invno, number)] for the tier-3 fallback
    for inv in actual_invoice_numbers:
        prefix, num = _split_series(inv)
        if prefix is None:
            continue
        actual_by_prefix.setdefault(prefix, set()).add(num)
        actual_by_norm_prefix.setdefault(_normalize_prefix(prefix), set()).add(num)
        actual_norm_full.append((_normalize_prefix(inv), num))

    for d in docs:
        if d["nature"] != "Invoices for outward supply":
            continue
        p_from, n_from = _split_series(d["sr_from"])
        p_to, n_to = _split_series(d["sr_to"])
        if p_from is None or p_to is None or p_from != p_to:
            findings.append(dict(range=f"{d['sr_from']} - {d['sr_to']}", missing=[],
                                  note="Could not parse series prefix/number -- check manually"))
            continue
        expected = set(range(n_from, n_to + 1))
        norm_prefix = _normalize_prefix(p_from)

        exact_have = actual_by_prefix.get(p_from, set())
        norm_have = actual_by_norm_prefix.get(norm_prefix, set())
        still_missing = expected - exact_have - norm_have
        fuzzy_found = set()
        for norm_inv, num in actual_norm_full:
            if num in still_missing and norm_prefix in norm_inv:
                fuzzy_found.add(num)
        missing_nums = sorted(still_missing - fuzzy_found)
        found_via_fuzzy = sorted(fuzzy_found)
        actual_count = len(expected) - len(missing_nums)

        # NEW (Issue 4 fix): the reverse case -- MORE invoices genuinely found and matched than
        # Table 13 itself claims. CONFIRMED real and re-diagnosed against the actual data (not
        # assumed): Jun-25's range 'TCP/UK/25-26/010 - TCP/UK/25-26/018' is 9 serials wide
        # (010..018 inclusive) and all 9 are genuinely, individually matched invoices -- but
        # Table 13's OWN separately-declared 'Total' field says 8. These are two independently
        # -reported numbers in the source government export (the sr_from/sr_to range and the
        # Total count are not derived from each other) and they can legitimately disagree. The
        # previous check only ever compared actual_count against the RANGE's own width, so a
        # range-vs-declared-Total mismatch like this fell through with no test at all and
        # printed a bare, unexplained "OK".
        #
        # A second, distinct excess signal is also captured: any actual invoice sharing this
        # exact series prefix whose number falls OUTSIDE [n_from, n_to] entirely (a genuinely
        # different scenario -- a wrong-series or duplicate invoice, not a range/Total mismatch).
        # Both are reported; neither is auto-explained -- unlike missing serials (where "missing
        # count == declared cancelled count" is an exact tie-out straight from Table 13's own
        # data), Table 13 carries no equivalent "expected excess" field to tie either of these
        # out against, so inventing an explanation here would not meet this tool's evidence bar.
        declared_total = d["total"]
        count_excess = actual_count - declared_total   # >0: range legitimately has more than Total says
        extra_nums = sorted((exact_have | norm_have) - expected)   # invoices outside the range boundary
        excess_signal = count_excess > 0 or bool(extra_nums)
        excess_n = max(count_excess, 0) + len(extra_nums)
        if excess_n >= 5:
            excess_severity = "FLAG"   # systematic gap at this scale is a stronger manipulation signal
        elif excess_signal:
            excess_severity = "REVIEW"
        else:
            excess_severity = None

        cancelled = d["cancelled"]
        explained_by_cancellation = bool(missing_nums) and len(missing_nums) == cancelled
        findings.append(dict(
            range=f"{d['sr_from']} - {d['sr_to']}", prefix=p_from,
            table13_total=declared_total, table13_cancelled=cancelled,
            actual_count=actual_count,
            missing=[f"{p_from}{n}" for n in missing_nums],
            found_via_fuzzy_match=[f"{p_from}{n}" for n in found_via_fuzzy],
            explained_by_declared_cancellation=explained_by_cancellation,
            extra_serials=[f"{p_from}{n}" for n in extra_nums],
            count_excess=count_excess,
            excess_signal=excess_signal,
            excess_severity=excess_severity,
        ))
    return findings


