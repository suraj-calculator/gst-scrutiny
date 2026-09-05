#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GST CHECKS MONTHLY
==================
CONSOLIDATED FILE -- contains what used to be: gst_analysis_checks.py, gst_eway_recon.py

The tool was reorganised from 19 .py files into 9 for easier sharing. Nothing
in the analytical logic was rewritten during that move: each section below is
the original module's code verbatim, with only (a) intra-project imports
repointed at the new file names, (b) its standalone __main__ demo block
removed, and (c) the renames listed under MERGE NOTES applied where two merged
modules happened to define the same top-level name with different bodies.

MERGE NOTES for this file:
  - gst_analysis_checks.main -> main_analysis14
  - gst_eway_recon.main -> main_eway
"""


# ============================================================================
# ==== SECTION: gst_analysis_checks.py  (was a standalone module before consolidation)
# ============================================================================
"""
GST SCRUTINY  --  ANALYSIS LAYER  (Sooraj's 14 checks)
======================================================
This sits ON TOP of gst_scrutiny_tool.py (the raw comparison tool).

The raw tool does: LEFT | RIGHT | DIFF | MATCH.
This layer does the *interpretive* scrutiny a GST officer / CA actually runs:
arithmetic consistency, effective-rate suppression, POS vs GSTIN tax-head,
RCM routing, duplicate invoices, blank-invoice detective work, timing/late-fee,
rate-wise e-invoice vs HSN, and ratio-based red flags.

Design rules (same spirit as the raw tool):
  - No hand-waving. Every flag shows the numbers it was computed from.
  - A check is either PASS / FLAG / INFO, with the exact arithmetic in the note.
  - Line-level checks list the offending rows by invoice number.

USAGE:
    Put this file next to gst_scrutiny_tool.py, configured for the same period, then:
        python gst_analysis_checks.py
    It reuses the CONFIG + parsers from gst_scrutiny_tool and writes:
        GST_Scrutiny_Analysis.xlsx
    (Findings sheet + Line Detail sheet + the existing Raw Values.)

Some checks need fields the original parser did not capture (POS, recipient
GSTIN, invoice date, RCM flag, IRN date, per-rate split). Those are pulled here
by re-reading the GSTR-1 / e-invoice sub-sheets directly, so the original tool
file does NOT have to change. If a needed column is absent in your export, the
check degrades to INFO ("column not found") instead of crashing.
"""

import os, re, datetime as _dt
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

import gst_parsers_returns as raw
import gst_core as mpu

num = raw.num
TOL = raw.TOLERANCE

# ----------------------------------------------------------------------
# Severity model
# ----------------------------------------------------------------------
PASS = "PASS"      # reconciles / nothing to do
INFO = "INFO"      # informational, manual eyeball
FLAG = "FLAG"      # genuine exception, needs explanation
REVW = "REVIEW"    # needs manual verification (allowed-but-watch)

SEV_ORDER = {FLAG: 0, REVW: 1, INFO: 2, PASS: 3, "EXPLAINED": 3}


class Finding:
    __slots__ = ("ref", "title", "severity", "detail", "numbers", "rows")
    def __init__(self, ref, title, severity, detail, numbers=None, rows=None):
        self.ref = ref            # e.g. "#3"
        self.title = title
        self.severity = severity
        self.detail = detail      # plain-English explanation w/ arithmetic
        self.numbers = numbers or {}   # {label: value} shown in a compact column
        # NEW: optional list of tuples, first row = header -- full invoice-level evidence for
        # this finding (e.g. every invoice behind an IRN-lag or duplicate-invoice flag), so the
        # reviewer doesn't need to reopen GSTR-1/E-Invoice to see which invoices/rupee-amounts
        # are actually involved. write_analysis14() renders these as a table under the finding
        # list on the SAME Analysis(14) sheet. Purely additive -- a Finding with no rows behaves
        # exactly as before.
        self.rows = rows or []


# ----------------------------------------------------------------------
# Extra raw-row readers  (fields the base parser doesn't keep)
# ----------------------------------------------------------------------
def _open(path):
    return openpyxl.load_workbook(path, data_only=True)

def _sheet_rows(wb, name):
    if name not in wb.sheetnames:
        return None, None
    rows = list(wb[name].iter_rows(values_only=True))
    if len(rows) < 5:
        return rows, {}
    hdr = [str(c).strip() if c else "" for c in rows[3]]
    H = {h: i for i, h in enumerate(hdr)}
    return rows, H

def _g(r, H, *names):
    """First matching column value from a row, by any of the given header names."""
    for n in names:
        if n in H and H[n] < len(r):
            return r[H[n]]
    return None

def _parse_date(v):
    if v is None or str(v).strip() in ("", "-"):
        return None
    if isinstance(v, _dt.datetime):
        return v.date()
    if isinstance(v, _dt.date):
        return v
    s = str(v).strip()
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%b-%Y", "%d-%b-%y", "%m/%d/%Y", "%d.%m.%Y"):
        try:
            return _dt.datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def read_gstr1_lines(path, month):
    """Detailed B2B + CDNR rows for line-level scrutiny, scoped to ONE month
    out of the merged GSTR-1 workbook.
    Returns list of dicts with the fields the analysis checks need.
    BUG FIX -- same root cause as parse_gstr1()'s b2b continuation-row fix in
    gst_parsers_returns.py, found while tracing a reported false gap through to every
    consumer of GSTR-1 invoice-level data: the portal's own export leaves GSTIN/Invoice
    (or Note) Number/Date blank -- a merged cell -- on every rate-line after a multi-rate
    document's first row. This function used to read each row's own (blank) cells in
    isolation, so a continuation row's real invoice/note identity was lost for every check
    that consumes this function's output (~15 checks across gst_checks_hsn_fraud.py,
    gst_report.py and master_build.py). Fixed the same way: forward-fill the last-seen
    gstin/invno/invdate/pos/rcm within each month's block, reset per sheet (a cdnr
    continuation row must never inherit identity left over from the b2b sheet)."""
    wb = _open(path)
    out = []
    for sn, kind in (("b2b, sez, de_inv", "INV"), ("cdnr", "CN"), ("cdnur", "CN")):
        rows, H = _sheet_rows(wb, sn)
        if not rows:
            continue
        last = None  # (gstin, invno, invdate_raw, pos, rcm) from the last primary row
        for r in mpu.rows_for_month(rows, 3, month):
            if not any(r):
                continue
            raw_gstin = _g(r, H, "GSTIN/UIN of Recipient", "GSTIN/UIN", "Recipient GSTIN")
            raw_invno = _g(r, H, "Invoice Number", "Note Number", "Invoice/Advance Receipt Number")
            if raw_gstin not in (None, "") and raw_invno not in (None, ""):
                last = (raw_gstin, raw_invno,
                        _g(r, H, "Invoice date", "Invoice Date", "Note date", "Note Date"),
                        _g(r, H, "Place Of Supply", "Place of Supply"),
                        _g(r, H, "Reverse Charge", "Reverse charge"))
                gstin_v, invno_v, invdate_v, pos_v, rcm_v = last
            elif last is not None:
                gstin_v, invno_v, invdate_v, pos_v, rcm_v = last
            else:
                gstin_v, invno_v, invdate_v, pos_v, rcm_v = "", "", None, "", ""
            out.append(dict(
                sheet=sn, kind=kind,
                gstin=str(gstin_v or "").strip(),
                invno=str(invno_v or "").strip(),
                invdate=_parse_date(invdate_v),
                pos=str(pos_v or "").strip(),
                rate=num(_g(r, H, "Rate", "Rate (%)")),
                taxable=num(_g(r, H, "Taxable Value")),
                igst=num(_g(r, H, "Integrated Tax")),
                cgst=num(_g(r, H, "Central Tax")),
                sgst=num(_g(r, H, "State/UT Tax")),
                cess=num(_g(r, H, "Cess Amount")),
                rcm=str(rcm_v or "").strip(),
                irn=str(_g(r, H, "IRN") or "").strip(),
                irndate=_parse_date(_g(r, H, "IRN date", "IRN Date", "Ack Date", "Acknowledgement Date")),
            ))
    return out


def read_einv_lines(path, month):
    if not path or not os.path.exists(path):
        return []   # E-Invoice legitimately not supplied at all -- graceful
    wb = _open(path)
    out = []
    rows, H = _sheet_rows(wb, "b2b, sez, de")
    if not rows:
        return out
    if month not in mpu.months_present(rows, 3):
        return out  # E-Invoice doesn't cover this month -- same graceful state
    for r in mpu.rows_for_month(rows, 3, month):
        if not any(r):
            continue
        out.append(dict(
            gstin=str(_g(r, H, "GSTIN/UIN of Recipient", "GSTIN/UIN", "Recipient GSTIN") or "").strip(),
            invno=str(_g(r, H, "Invoice number", "Invoice Number", "Document number") or "").strip(),
            invdate=_parse_date(_g(r, H, "Invoice date", "Document date")),
            pos=str(_g(r, H, "Place Of Supply", "Place of Supply") or "").strip(),
            rate=num(_g(r, H, "Rate", "Rate (%)")),
            taxable=num(_g(r, H, "Taxable Value")),
            igst=num(_g(r, H, "Integrated Tax")),
            cgst=num(_g(r, H, "Central Tax")),
            sgst=num(_g(r, H, "State/UT Tax")),
            rcm=str(_g(r, H, "Reverse Charge", "Reverse charge") or "").strip(),
            irn=str(_g(r, H, "IRN", "Irn") or "").strip(),
            irndate=_parse_date(_g(r, H, "IRN date", "IRN Date", "Ack Date", "Acknowledgement Date")),
            err=str(_g(r, H, "Error in auto-population/ deletion", "Error", "Errors") or "").strip(),
        ))
    return out


def _state_code(s):
    """First 2 chars of a GSTIN or POS string, if numeric state code."""
    s = (s or "").strip()
    m = re.match(r"\s*0?(\d{1,2})", s)
    if m:
        return m.group(1).zfill(2)
    return None


# ----------------------------------------------------------------------
# THE 14 CHECKS
# ----------------------------------------------------------------------
def run_checks(g1, g3b, einv, b2b, g1_lines, einv_lines):
    F = []
    def gv(key, i, d=0.0):
        v = g3b.get(key); return v[i] if v and i < len(v) else d

    # ---- #0  Totals reconciliation (disambiguate gross vs named vs orphan) ----
    gross = g1.get("taxable", 0.0)          # sum of ALL B2B lines incl any orphan
    named = g1.get("named_taxable", 0.0)    # sum of properly-numbered invoices only
    orphan_tax = round(gross - named, 2)
    F.append(Finding("#0", "GSTR-1 B2B totals (gross / named / orphan)",
                     INFO,
                     f"GSTR-1 B2B all-lines taxable = {gross:,.2f}; properly-numbered-invoices taxable = "
                     f"{named:,.2f}; difference = {orphan_tax:,.2f} sits on line(s) whose invoice number was "
                     "dropped (see #5). NOTE: the all-lines total can coincide with the e-invoice total when "
                     "the orphan line is the same rate-line that IS numbered in the e-invoice — it is still a "
                     "GSTR-1 figure, not an e-invoice figure. The GROSS-vs-HSN gap is explained ONCE by credit "
                     "notes (CDNR taxable), not separately in two places.",
                     {"all-lines": gross, "named": named, "orphan": orphan_tax}))

    # ---- #1  Nil / Exempt / Non-GST: GSTR-1 'exemp' vs 3B 3.1(c)/(e) ----
    # GSTR-1 Table 8 'exemp' sheet carries Nil-rated, Exempted and Non-GST outward supplies.
    # Mapping to 3B:
    #   GSTR-1 (Nil + Exempted)  ->  3B 3.1(c) 'Other Outward (Nil rated, exempted)'
    #   GSTR-1 (Non-GST)         ->  3B 3.1(e) 'Non-GST Outward supplies'
    # Zero-rated (3B 3.1(b)) is reported in GSTR-1 via the 'exp'/SEZ heads, a different sheet,
    # so it is shown for context but not differenced here (would create a false mismatch).
    nil_exempt_g1 = g1.get("nil_exempt_taxable", None)
    nongst_g1 = g1.get("nongst_taxable", 0.0)
    if nil_exempt_g1 is None:
        # 'exemp' sheet/header not found in this export -> cannot auto-compare, say so honestly.
        F.append(Finding("#1", "Nil / exempt / non-GST (GSTR-1 vs 3B 3.1c/3.1e)",
                         INFO,
                         "GSTR-1 'exemp' (Table 8) sheet/header not found in this export, so nil/exempt/"
                         "non-GST could not be auto-read. Manually compare GSTR-1 Table 8 against 3B "
                         "3.1(c)+3.1(e).",
                         {}))
    else:
        b3c = gv("3.1c", 0)            # 3B nil+exempt taxable
        b3e = gv("3.1e", 0)            # 3B non-GST taxable
        diff_ne = nil_exempt_g1 - b3c
        diff_ng = nongst_g1 - b3e
        worst = PASS if (abs(diff_ne) <= TOL and abs(diff_ng) <= TOL) else FLAG
        F.append(Finding("#1", "Nil / exempt / non-GST (GSTR-1 Table 8 vs 3B 3.1c/3.1e)",
                         worst,
                         f"GSTR-1 nil+exempt {nil_exempt_g1:,.2f} vs 3B 3.1(c) {b3c:,.2f} (diff {diff_ne:,.2f}); "
                         f"GSTR-1 non-GST {nongst_g1:,.2f} vs 3B 3.1(e) {b3e:,.2f} (diff {diff_ng:,.2f}). "
                         f"(For context, 3B 3.1(b) zero-rated taxable = {gv('3.1b',0):,.2f}; zero-rated is "
                         "reported via GSTR-1 export/SEZ heads, not Table 8, so not differenced here.) "
                         + ("All nil/exempt/non-GST figures reconcile." if worst == PASS
                            else "Mismatch — GSTR-1 Table 8 does not tie to 3B 3.1(c)/(e); reconcile."),
                         {"G1 nil+exempt": nil_exempt_g1, "3B 3.1c": b3c,
                          "G1 non-GST": nongst_g1, "3B 3.1e": b3e}))

    # ---- #2  Credit-note effect: GSTR-1 net vs gross on 3B liability ----
    g1_gross_tax = g1["IGST"] + g1["CGST"] + g1["SGST"]
    g1_net_tax = (g1["IGST"]-g1["cn_IGST"]) + (g1["CGST"]-g1["cn_CGST"]) + (g1["SGST"]-g1["cn_SGST"])
    b3b_tax = gv("3.1a",1)+gv("3.1a",2)+gv("3.1a",3)
    d_net = g1_net_tax - b3b_tax
    F.append(Finding("#2", "Credit-note effect on outward liability (GSTR-1 net vs 3B 3.1a)",
                     PASS if abs(d_net) <= TOL else FLAG,
                     f"CN total tax = {g1['cn_IGST']+g1['cn_CGST']+g1['cn_SGST']:,.2f}. "
                     f"GSTR-1 GROSS tax {g1_gross_tax:,.2f} -> NET {g1_net_tax:,.2f}. "
                     f"3B 3.1(a) tax {b3b_tax:,.2f}. Net-vs-3B diff {d_net:,.2f}. "
                     "3B 3.1(a) must be reported NET of credit notes; if GROSS matched but NET didn't, "
                     "CN was not given effect in 3B.",
                     {"G1 gross": g1_gross_tax, "G1 net": g1_net_tax, "3B 3.1a": b3b_tax}))

    # ---- #3  Arithmetic consistency: 4C = 4A5 + 4A3 - 4B1 - 4B2 (per head) ----
    # NOTE: Net ITC includes RCM ITC (4A3). Omitting 4A3 produces a false mismatch.
    # FIX: 4B1 (Rule 42/43/38 + Sec 17(5) reversal) is now included -- it was
    # silently dropped from this formula before (always 0 for this taxpayer,
    # so it was invisible, but would misfire for any period with a real
    # Rule-42/43 reversal). Also depends on gst_scrutiny_tool.parse_gstr3b's
    # corrected, section-anchored 4B(1)/4B(2) extraction -- see that function's
    # docstring for the duplicate-label bug this fixes.
    heads = ["IGST", "CGST", "SGST"]
    rows3 = []
    worst = PASS
    for i, h in enumerate(heads):
        should = gv("4A5", i) + gv("4A3", i) - gv("4B1", i) - gv("4B2", i)
        actual = gv("4C", i)
        d = round(should - actual, 2)
        rows3.append(f"{h}: 4A5 {gv('4A5',i):,.2f} + 4A3 {gv('4A3',i):,.2f} - 4B1 {gv('4B1',i):,.2f} "
                     f"- 4B2 {gv('4B2',i):,.2f} = {should:,.2f}  | 3B 4C {actual:,.2f}  | diff {d:,.2f}")
        if abs(d) > TOL:
            worst = FLAG
    F.append(Finding("#3", "ITC arithmetic: Net ITC (4C) = 4A5 + 4A3 (RCM) - 4B1 - 4B2",
                     worst,
                     "Net ITC must include RCM ITC (Table 4A3); a check using only 4A5-4B2 wrongly "
                     "flags a gap equal to the RCM ITC. Per head:\n   " + "\n   ".join(rows3),
                     {}))

    # ---- #4  Effective tax-rate comparison (GSTR-1 vs 3B 3.1a) ----
    r1 = (g1["IGST"]+g1["CGST"]+g1["SGST"]) / g1["taxable"] * 100 if g1["taxable"] else 0
    den3b = gv("3.1a",0)
    r3b = (gv("3.1a",1)+gv("3.1a",2)+gv("3.1a",3)) / den3b * 100 if den3b else 0
    drate = r1 - r3b
    if abs(drate) <= 0.10:
        sev4, msg4 = PASS, "Rates align (within 0.10pp) -> no suppression signal."
    elif drate > 0.10:
        sev4, msg4 = FLAG, "GSTR-1 rate HIGHER than 3B -> possible suppression of liability in 3B."
    else:
        sev4, msg4 = REVW, "3B rate higher than GSTR-1 -> over-reported in 3B / under-reported in GSTR-1; verify."
    F.append(Finding("#4", "Effective tax-rate (GSTR-1 vs 3B 3.1a)",
                     sev4,
                     f"GSTR-1 eff rate {r1:.3f}% vs 3B 3.1(a) {r3b:.3f}% (diff {drate:+.3f}pp). {msg4} "
                     "Blended ~10.7% indicates a 5%/12% supply mix.",
                     {"G1 %": round(r1,3), "3B %": round(r3b,3)}))

    # ---- #5  Orphan invoice-number lines -> RE-LINK to e-invoice, don't just call it "blank" ----
    # A GSTR-1 B2B line whose invoice-number cell is empty is a real export defect, but the
    # actionable finding is WHICH invoice it belongs to. Re-link by (rate, taxable, tax) against
    # e-invoice lines that are absent from GSTR-1's named set -> names the true invoice + rate-line.
    blanks = [L for L in g1_lines if L["kind"] == "INV" and not L["invno"]]
    if blanks:
        g1_named_keys = {(L["invno"], round(L["rate"], 2)) for L in g1_lines
                         if L["kind"] == "INV" and L["invno"]}
        ei_index = {}
        for E in einv_lines:
            ei_index.setdefault((round(E["rate"], 2), round(E["taxable"], 2), round(E["igst"], 2)),
                                []).append(E)
        relinked, unresolved = [], []
        for b in blanks:
            key = (round(b["rate"], 2), round(b["taxable"], 2), round(b["igst"], 2))
            cands = [E for E in ei_index.get(key, [])
                     if (E["invno"], round(E["rate"], 2)) not in g1_named_keys]
            if len(cands) == 1:
                relinked.append((b, cands[0]["invno"]))
            else:
                unresolved.append(b)
        parts = []
        for b, inv in relinked:
            parts.append(f"GSTR-1 B2B line lost its invoice number; re-linked by (rate, value, tax) to "
                         f"invoice {inv} @ {b['rate']:g}% (taxable {b['taxable']:,.2f}, IGST {b['igst']:,.2f}). "
                         f"This rate-line is correctly numbered in the e-invoice and appears in the HSN summary, "
                         f"but is mis-recorded (number dropped) in the GSTR-1 B2B detail. Tax impact {b['igst']:,.2f}. "
                         f"Amend GSTR-1 / adjust next period.")
        for b in unresolved:
            parts.append(f"Unnumbered GSTR-1 B2B line @ {b['rate']:g}% (taxable {b['taxable']:,.2f}, "
                         f"IGST {b['igst']:,.2f}) — no unique e-invoice match; locate the source invoice manually.")
        F.append(Finding("#5", f"GSTR-1 B2B line(s) with dropped invoice number: {len(blanks)} "
                                f"({len(relinked)} re-linked)",
                         FLAG,
                         " ".join(parts),
                         {"orphan lines": len(blanks), "re-linked": len(relinked),
                          "taxable": sum(b["taxable"] for b in blanks),
                          "tax impact": sum(b["igst"]+b["cgst"]+b["sgst"] for b in blanks)},
                         rows=[("Re-linked invoice no.", "GSTIN", "POS", "Invoice Date", "Rate %",
                                "Taxable", "IGST", "CGST", "SGST", "Status")] +
                              [(inv, b.get("gstin", ""), b.get("pos", ""), b.get("invdate"), b["rate"],
                                round(b["taxable"], 2), round(b["igst"], 2), round(b["cgst"], 2),
                                round(b["sgst"], 2), "re-linked from e-invoice by rate/value/tax")
                               for b, inv in relinked] +
                              [("(no invoice no. found)", b.get("gstin", ""), b.get("pos", ""),
                                b.get("invdate"), b["rate"], round(b["taxable"], 2), round(b["igst"], 2),
                                round(b["cgst"], 2), round(b["sgst"], 2),
                                "unresolved -- no unique e-invoice match, locate manually")
                               for b in unresolved]))
    else:
        F.append(Finding("#5", "GSTR-1 B2B lines with dropped invoice number", PASS,
                         "Every GSTR-1 B2B line carries an invoice number.", {}))

    # ---- #6  Duplicate invoice numbers (allowed if multi-rate; flag for manual) ----
    seen = {}
    for L in g1_lines:
        if L["kind"] != "INV" or not L["invno"]:
            continue
        seen.setdefault(L["invno"], []).append(L)
    dups = {k: v for k, v in seen.items() if len(v) > 1}
    if dups:
        lines = []
        dup_detail_rows = [("Invoice No.", "GSTIN", "POS", "Invoice Date", "Rate %", "Taxable",
                             "IGST", "CGST", "SGST", "Note")]
        for inv, dup_rows in list(dups.items())[:15]:
            rates = ",".join(f"{r['rate']:g}%" for r in dup_rows)
            tot = sum(r["taxable"] for r in dup_rows)
            same_rate = len({r["rate"] for r in dup_rows}) < len(dup_rows)
            tag = "  <-- SAME RATE REPEATED, verify not a true duplicate" if same_rate else ""
            lines.append(f"{inv}: {len(dup_rows)} lines @ rates {rates}, taxable {tot:,.2f}{tag}")
            for r in dup_rows:
                dup_detail_rows.append((inv, r.get("gstin", ""), r.get("pos", ""), r.get("invdate"),
                                        r["rate"], round(r["taxable"], 2), round(r["igst"], 2),
                                        round(r["cgst"], 2), round(r["sgst"], 2),
                                        "SAME RATE REPEATED -- verify not a true duplicate" if same_rate else ""))
        F.append(Finding("#6", f"Duplicate invoice numbers: {len(dups)} invoice(s) on >1 line",
                         REVW,
                         "Multiple lines per invoice are allowed for multi-rate invoices (taxable adds up). "
                         "Same rate repeated on one invoice = likely true duplicate, verify manually.\n   "
                         + "\n   ".join(lines),
                         {"dup invoices": len(dups)},
                         rows=dup_detail_rows))
    else:
        F.append(Finding("#6", "Duplicate invoice numbers", PASS,
                         "Every B2B invoice number appears on a single line.", {}))

    # ---- #7  E-invoice errors field ----
    err_lines = [L for L in einv_lines if L.get("err")]
    if einv.get("available"):
        if err_lines:
            det = "; ".join(f"{L['invno']}: {L['err']}" for L in err_lines[:10])
            F.append(Finding("#7", f"E-invoice error/auto-population flags: {len(err_lines)}",
                             FLAG, f"Errors present: {det}. Investigate (e.g. IRN already used, deletion).",
                             {"errors": len(err_lines)},
                             rows=[("Invoice No.", "GSTIN", "Invoice Date", "IRN", "IRN Date",
                                    "Rate %", "Taxable", "IGST", "CGST", "SGST", "Error")] +
                                  [(L.get("invno", ""), L.get("gstin", ""), L.get("invdate"),
                                    L.get("irn", ""), L.get("irndate"), L.get("rate"),
                                    round(L.get("taxable", 0.0), 2), round(L.get("igst", 0.0), 2),
                                    round(L.get("cgst", 0.0), 2), round(L.get("sgst", 0.0), 2), L.get("err", ""))
                                   for L in err_lines]))
        else:
            F.append(Finding("#7", "E-invoice error field", PASS,
                             "No errors flagged in the e-invoice auto-population column.", {}))
    else:
        F.append(Finding("#7", "E-invoice error field", INFO, "E-invoice file not supplied.", {}))

    # ---- #8  Time-lag: IRN date vs GSTR-1 filing (>30 days) ----
    filing = _filing_date("GSTR1")
    lagged = []
    lag_src_by_inv = {}
    for L in einv_lines or g1_lines:
        idt = L.get("irndate")
        if idt and filing:
            lag = (filing - idt).days
            if lag > 30:
                lagged.append((L.get("invno", "?"), idt, lag))
                lag_src_by_inv[L.get("invno", "?")] = L
    if filing is None:
        F.append(Finding("#8", "IRN-date vs GSTR-1 filing lag (>30d)", INFO,
                         "Set GSTR1_FILING_DATE in CONFIG to enable. IRN-date column also required in the export.",
                         {}))
    elif lagged:
        det = "; ".join(f"{inv} IRN {d} ({lag}d)" for inv, d, lag in lagged[:10])
        # Full invoice detail per instruction: not just invoice-no + IRN date, but the complete
        # underlying invoice record (GSTIN, POS, taxable/tax amounts, IRN, GSTR-1 filing date,
        # and the computed lag) for every invoice behind this flag.
        lag_rows = [("Invoice No.", "GSTIN", "POS", "Invoice Date", "IRN", "IRN Date",
                     "GSTR-1 Filing Date", "Lag (days)", "Rate %", "Taxable", "IGST", "CGST", "SGST")]
        for inv, idt, lag in lagged:
            L = lag_src_by_inv.get(inv, {})
            lag_rows.append((inv, L.get("gstin", ""), L.get("pos", ""), L.get("invdate"),
                             L.get("irn", ""), idt, filing, lag, L.get("rate"),
                             round(L.get("taxable", 0.0), 2), round(L.get("igst", 0.0), 2),
                             round(L.get("cgst", 0.0), 2), round(L.get("sgst", 0.0), 2)))
        F.append(Finding("#8", f"IRN-to-filing lag >30 days: {len(lagged)} invoice(s)",
                         FLAG, f"Delayed reporting: {det}.", {"lagged": len(lagged)},
                         rows=lag_rows))
    else:
        F.append(Finding("#8", "IRN-date vs GSTR-1 filing lag", PASS,
                         "No invoice with IRN-to-filing gap beyond 30 days.", {}))


    # ---- #9  Rate-wise: e-invoice vs GSTR-1 HSN summary ----
    rate_buckets = {}
    src_lines = einv_lines if einv.get("available") and einv_lines else g1_lines
    for L in src_lines:
        if L.get("kind") == "CN":
            continue
        rb = rate_buckets.setdefault(round(L["rate"], 2), 0.0)
        rate_buckets[round(L["rate"], 2)] = rb + L["taxable"]
    if rate_buckets:
        buckets = "; ".join(f"{r:g}%: {v:,.2f}" for r, v in sorted(rate_buckets.items()))
        F.append(Finding("#9", "Rate-wise taxable split (e-invoice / GSTR-1 lines)",
                         INFO,
                         f"Rate-wise taxable from source lines -> {buckets}. "
                         "Compare against the GSTR-1 HSN summary rate-wise rows; any rate present here but "
                         "absent/short in HSN = misclassification. (HSN rate split must come from the HSN sheet.)",
                         {}))
    else:
        F.append(Finding("#9", "Rate-wise e-invoice vs HSN", INFO, "No rate-bearing source lines parsed.", {}))

    # ---- #10  Late-fee timing: GSTR-1 vs 3B filing gap (>20 days... uses statutory due dates) ----
    f1 = _filing_date("GSTR1"); f3 = _filing_date("GSTR3B")
    if f1 and f3:
        gap = abs((f3 - f1).days)
        F.append(Finding("#10", "GSTR-1 vs GSTR-3B filing-gap",
                         REVW if gap > 20 else PASS,
                         f"GSTR-1 filed {f1}, GSTR-3B filed {f3}, gap {gap} days. "
                         + ("Gap >20d -> check late-fee/interest exposure." if gap > 20
                            else "Within 20 days."),
                         {"gap_days": gap}))
    else:
        F.append(Finding("#10", "Filing-gap / late fee", INFO,
                         "Set GSTR1_FILING_DATE and GSTR3B_FILING_DATE in CONFIG to enable.", {}))

    # ---- #11  POS vs recipient-GSTIN state-code -> correct tax head ----
    pos_errors = []
    for L in g1_lines:
        if L["kind"] == "CN":
            continue
        pos_sc = _state_code(L["pos"]); g_sc = _state_code(L["gstin"])
        if not pos_sc:
            continue
        has_igst = L["igst"] > TOL
        has_local = (L["cgst"] > TOL or L["sgst"] > TOL)
        if g_sc:  # registered recipient
            same = (pos_sc == g_sc)
            if same and has_igst and not has_local:
                pos_errors.append((L["invno"], L["pos"], L["gstin"], "intra-state but charged IGST"))
            if (not same) and has_local and not has_igst:
                pos_errors.append((L["invno"], L["pos"], L["gstin"], "inter-state but charged CGST+SGST"))
    if pos_errors:
        det = "; ".join(f"{inv} POS={p} GSTIN={g}: {why}" for inv, p, g, why in pos_errors[:12])
        F.append(Finding("#11", f"POS vs GSTIN tax-head mismatch: {len(pos_errors)} invoice(s)",
                         FLAG, f"Wrong tax head charged -> direct exposure. {det}.",
                         {"mismatches": len(pos_errors)}))
    else:
        F.append(Finding("#11", "POS vs recipient-GSTIN tax head", PASS,
                         "Every registered-recipient line uses the correct head for its POS/GSTIN state pair "
                         "(or POS/GSTIN not available to test).", {}))

    # ---- #12  RCM flag routing: e-invoice/GSTR-1 RCM=Y -> 3B 3.1(d) & ITC 4A3 ----
    rcm_lines = [L for L in (einv_lines or g1_lines)
                 if str(L.get("rcm", "")).strip().upper() in ("Y", "YES", "TRUE")]
    rcm_tax = sum(L["igst"]+L.get("cgst",0)+L.get("sgst",0) for L in rcm_lines) if rcm_lines else 0
    d31 = gv("3.1d",1)+gv("3.1d",2)+gv("3.1d",3)
    if rcm_lines:
        F.append(Finding("#12", f"RCM-flagged invoices: {len(rcm_lines)}",
                         REVW,
                         f"RCM-marked lines total tax {rcm_tax:,.2f}. 3B 3.1(d) liability tax {d31:,.2f}, "
                         f"3B 4A3 RCM-ITC {gv('4A3',0)+gv('4A3',1)+gv('4A3',2):,.2f}. "
                         "RCM liability must appear in 3.1(d) and the matching ITC in 4A3; reconcile.",
                         {"RCM lines": len(rcm_lines), "3.1d tax": d31}))
    else:
        F.append(Finding("#12", "RCM routing (3.1d & 4A3)",
                         INFO if not (einv_lines or g1_lines) else PASS,
                         f"No RCM=Y line found in source. (3B shows 3.1d tax {d31:,.2f}, "
                         f"4A3 ITC {gv('4A3',0)+gv('4A3',1)+gv('4A3',2):,.2f} — RCM on unregistered/import "
                         "of services won't carry a line-level flag.)",
                         {}))

    # ---- #13  HSN IGST vs named-invoice IGST gap == credit-note effect + B2CS? ----
    # B2CS (Table 7, small unregistered B2C sales) structurally carries NO invoice number
    # anywhere in a GSTR-1 export -- it is legally reported only as a state+rate aggregate, so
    # it can never appear in named_taxable/named_IGST (b2b-only) even though it IS included in
    # the HSN summary (Table 12 covers every outward supply, B2B and B2C alike). Netting it out
    # here, with its own figure shown, is an EXACT tie-out against data already computed
    # elsewhere in this same run (g1["b2cs_IGST"]) -- not a suppression.
    gap13 = g1["named_IGST"] - g1["hsn_IGST"]
    resid_cn = gap13 - g1["cn_IGST"]
    resid = resid_cn + g1["b2cs_IGST"]
    if abs(resid) <= 200:
        sev13 = PASS
        why = ("Gap is explained by credit notes (HSN summary is net of CN, invoice lines are gross)"
               if abs(resid_cn) <= 200 else
               f"Gap is explained by B2CS (Table 7 unregistered-B2C aggregate IGST "
               f"{g1['b2cs_IGST']:,.2f} -- see this month's Comparison sheet, section 'A. Outward "
               f"Liability' -- which is included in the HSN summary but structurally can never "
               f"carry an invoice number, so it's never in the named-invoice figure).")
    else:
        sev13 = FLAG
        why = "Residual NOT explained by credit notes or B2CS — reconcile line-by-line."
    F.append(Finding("#13", "HSN-summary IGST vs named-invoice IGST",
                     sev13,
                     f"named-invoice IGST {g1['named_IGST']:,.2f} - HSN IGST {g1['hsn_IGST']:,.2f} "
                     f"= gap {gap13:,.2f}. Credit-note IGST = {g1['cn_IGST']:,.2f}. B2CS IGST = "
                     f"{g1['b2cs_IGST']:,.2f}. Residual after both = {resid:,.2f}. " + why,
                     {"gap": gap13, "cn_IGST": g1["cn_IGST"], "b2cs_IGST": g1["b2cs_IGST"],
                      "residual": resid}))

    # ---- #14  ITC / Liability ratio ----
    liab = gv("3.1a",1)+gv("3.1a",2)+gv("3.1a",3)
    itc = gv("4A5",0)+gv("4A5",1)+gv("4A5",2)
    ratio = (itc/liab*100) if liab else 0
    if ratio > 95:
        sev14 = FLAG
        m = ("ITC >95% of output liability -> taxpayer pays almost no cash. "
             "Watch for circular trading / fake billing. Here ITC even EXCEEDS liability "
             "(ratio >100%): net ITC accumulation this period — confirm it's genuine input build-up "
             "(capex/inventory) and not inflated.") if ratio > 100 else \
            "ITC >95% of liability -> minimal cash payout; verify input genuineness."
    elif ratio < 20:
        sev14, m = REVW, "ITC <20% of liability -> low-margin/high-value-add; still verify."
    else:
        sev14, m = PASS, "ITC/Liability ratio in a normal band."
    F.append(Finding("#14", "ITC / Output-liability ratio",
                     sev14,
                     f"Liability (3.1a tax) {liab:,.2f}; ITC (4A5) {itc:,.2f}; ratio {ratio:.1f}%. {m}",
                     {"liability": liab, "ITC": itc, "ratio %": round(ratio,1)}))

    F.sort(key=lambda x: (SEV_ORDER[x.severity], x.ref))
    return F


# ----------------------------------------------------------------------
# Optional CONFIG additions (filing dates). Read from gst_scrutiny_tool if set.
# ----------------------------------------------------------------------
def _filing_date(which):
    attr = "GSTR1_FILING_DATE" if which == "GSTR1" else "GSTR3B_FILING_DATE"
    v = getattr(raw, attr, None)
    return _parse_date(v) if v else None


# ----------------------------------------------------------------------
# WRITE EXCEL
# ----------------------------------------------------------------------
FILL = {FLAG: PatternFill("solid", fgColor="FFC7CE"),
        REVW: PatternFill("solid", fgColor="FFEB9C"),
        INFO: PatternFill("solid", fgColor="DDEBF7"),
        PASS: PatternFill("solid", fgColor="C6EFCE")}
FONT_SEV = {FLAG: Font(bold=True, color="9C0006"),
            REVW: Font(bold=True, color="9C6500"),
            INFO: Font(bold=True, color="2F5496"),
            PASS: Font(bold=True, color="006100")}
HEAD = PatternFill("solid", fgColor="1F3864")
BORDER = Border(*[Side(style="thin", color="BFBFBF")]*4)


def write_analysis(findings, raw_bundle, outpath):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Findings"
    ws.cell(1, 1, f"GST SCRUTINY — ANALYSIS (14 checks) — Period {raw.PERIOD_LABEL}").font = Font(bold=True, size=13, color="1F3864")
    ws.cell(2, 1, f"GSTIN {getattr(raw,'SELF_GSTIN','')}  |  {getattr(raw,'COMPANY_NAME','') or '(company auto-detected)'}").font = Font(size=9, italic=True)

    nflag = sum(1 for f in findings if f.severity == FLAG)
    nrev = sum(1 for f in findings if f.severity == REVW)
    ws.cell(3, 1, f"FLAGS: {nflag}   REVIEW: {nrev}   "
                  f"INFO: {sum(1 for f in findings if f.severity==INFO)}   "
                  f"PASS: {sum(1 for f in findings if f.severity==PASS)}").font = Font(size=10, bold=True)

    hdr = ["Ref", "Check", "Result", "Key numbers", "Detail / arithmetic"]
    for i, h in enumerate(hdr, 1):
        c = ws.cell(5, i, h); c.fill = HEAD; c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = BORDER
    r = 6
    for f in findings:
        ws.cell(r, 1, f.ref)
        ws.cell(r, 2, f.title)
        cv = ws.cell(r, 3, f.severity); cv.fill = FILL[f.severity]; cv.font = FONT_SEV[f.severity]
        cv.alignment = Alignment(horizontal="center")
        ws.cell(r, 4, "  ".join(f"{k}={v:,.2f}" if isinstance(v, (int, float)) else f"{k}={v}"
                                for k, v in f.numbers.items()))
        ws.cell(r, 5, f.detail)
        for c in range(1, 6):
            cell = ws.cell(r, c); cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(c in (2, 4, 5)))
            if c != 3:
                cell.font = Font(size=10)
        r += 1
    for col, w in zip("ABCDE", [6, 44, 10, 30, 95]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A6"

    wb.save(outpath)
    return nflag, nrev


def main_analysis14():
    """Legacy/manual-testing entry point -- master_build.py does not call this;
    it drives run_checks() per month via run_monthly_pipeline.py instead.
        python gst_analysis_checks.py <month, e.g. Jan-23>
    """
    import sys as _sys
    if len(_sys.argv) < 2:
        raise SystemExit("Usage: python gst_analysis_checks.py <month, e.g. Jan-23>")
    month = _sys.argv[1]
    raw.PERIOD_LABEL = month
    g1 = raw.parse_gstr1(raw.GSTR1_FILE, month)
    g3b = raw.parse_gstr3b(raw.GSTR3B_FILE, month)
    einv = raw.parse_einv(raw.EINV_FILE, month)
    b2b = raw.get_gstr2b_values()
    g1_lines = read_gstr1_lines(raw.GSTR1_FILE, month)
    einv_lines = read_einv_lines(raw.EINV_FILE, month) if raw.EINV_FILE else []

    # remap 3b keys 3.1a etc. to match parser output keys
    g3b_norm = {}
    for k, v in g3b.items():
        g3b_norm[k] = v
    # parser stores '3.1a','3.1d','4A5','4A3','4B2','4C' already

    findings = run_checks(g1, g3b_norm, einv, b2b, g1_lines, einv_lines)
    out = "GST_Scrutiny_Analysis.xlsx"
    nflag, nrev = write_analysis(findings, dict(g1=g1, g3b=g3b_norm, einv=einv, b2b=b2b), out)
    print(f"Saved: {out}")
    print(f"Findings: {len(findings)}  |  FLAG: {nflag}  REVIEW: {nrev}")
    for f in findings:
        print(f"  [{f.severity:6}] {f.ref}  {f.title}")




# ============================================================================
# ==== SECTION: gst_eway_recon.py  (was a standalone module before consolidation)
# ============================================================================
"""
GST SCRUTINY  --  E-WAY BILL RECONCILIATION LAYER  (27-check matrix)
====================================================================
Sixth-source layer: brings E-Way Bill OUTWARD and INWARD into the
GSTR-1 / 2B / 3B / E-Invoice reconciliation.

Reuses CONFIG + parsers from gst_scrutiny_tool.py.

Honest scope (decided with the user):
  BUILD   : #1-9, #15-18, #23, #25-27   (full, file-driven)
  PARTIAL : #10-13 (2B is a PDF summary -> aggregate only, no line list)
            #22 (no validity/expiry col), #24 (no cancel-status col)
  SKIP    : #14 (no books), #19/#20 (no purchase e-invoice / books),
            #21 (no filing date)  -- #7/#16 still run on dates inside the files

NO safety nets, NO invented data. Every PARTIAL/SKIP is labelled, not faked.

USAGE:
    Put next to gst_scrutiny_tool.py + gst_analysis_checks.py, set the two
    EWB filenames in CONFIG below, then:
        python gst_eway_recon.py
    -> GST_Scrutiny_EWayBill.xlsx
"""

import os, re, datetime as _dt
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

import gst_parsers_returns as raw
import gst_parsers_returns as g2b
import gst_core as mpu

num = raw.num
TOL = 1.0                 # rupee tolerance for value matches
VALUE_TOL_PCT = 0.01      # 1% tolerance for EWB-vs-return value compares

# ----------------------------------------------------------------------
# CONFIG  --  set by run_monthly_pipeline.py per month (merged-file model)
# ----------------------------------------------------------------------
GSTR2B_FILE  = None                                   # merged GSTR-2B workbook path
SELF_GSTIN   = ""
COMPANY_NAME = ""
EWB_THRESHOLD = 50000.0   # Rule 138 inter-state EWB threshold (consignment value)

OUTPUT_FILE = "GST_Scrutiny_EWayBill.xlsx"
_LAST_2B_SRC = "pdf"    # set in main_eway(): 'excel' enables line-level checks
_LAST_2B_FILE = None    # actual 2B Excel filename used (for the header stamp)
_LAST_EINV_FILE = None  # actual e-invoice filename used


def find_file(configured, patterns, search_dir=".", exclude=("SCRUTINY", "EWAYBILL", "EWAY", "COMPARISON", "ANALYSIS")):
    """Return a usable file path. 1) use `configured` if it exists; 2) else scan search_dir for
    a file whose name matches any regex in `patterns` (case-insensitive) and isn't one of our outputs."""
    import glob as _glob, re as _re
    if configured and os.path.exists(configured):
        return configured
    cands = []
    for f in _glob.glob(os.path.join(search_dir, "*.xlsx")) + _glob.glob(os.path.join(search_dir, "*.xlsm")):
        name = os.path.basename(f).upper()
        if any(x in name for x in exclude):
            continue
        if any(_re.search(p, name, _re.I) for p in patterns):
            cands.append(f)
    if not cands:
        return None
    cands.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return cands[0]


# ----------------------------------------------------------------------
# State code <-> name (needed: GSTR-1 POS is a name, EWB carries GSTIN codes)
# ----------------------------------------------------------------------
STATE = {
 "01":"Jammu and Kashmir","02":"Himachal Pradesh","03":"Punjab","04":"Chandigarh",
 "05":"Uttarakhand","06":"Haryana","07":"Delhi","08":"Rajasthan","09":"Uttar Pradesh",
 "10":"Bihar","11":"Sikkim","12":"Arunachal Pradesh","13":"Nagaland","14":"Manipur",
 "15":"Mizoram","16":"Tripura","17":"Meghalaya","18":"Assam","19":"West Bengal",
 "20":"Jharkhand","21":"Odisha","22":"Chhattisgarh","23":"Madhya Pradesh","24":"Gujarat",
 "25":"Daman and Diu","26":"Dadra and Nagar Haveli","27":"Maharashtra","28":"Andhra Pradesh",
 "29":"Karnataka","30":"Goa","31":"Lakshadweep","32":"Kerala","33":"Tamil Nadu",
 "34":"Puducherry","35":"Andaman and Nicobar Islands","36":"Telangana","37":"Andhra Pradesh",
 "38":"Ladakh","97":"Other Territory",
}
NAME2CODE = {v.lower(): k for k, v in STATE.items()}
NAME2CODE["andhra pradesh"] = "37"

def state_code(s):
    """Pull a 2-digit state code from a GSTIN or a leading-coded string."""
    m = re.match(r"\s*(\d{2})", str(s or ""))
    return m.group(1) if m else None

def name_to_code(name):
    return NAME2CODE.get(str(name or "").strip().lower())


# ----------------------------------------------------------------------
# Severity model (shared style with analysis layer)
# ----------------------------------------------------------------------
FLAG, REVW, INFO, PASS, SKIP = "FLAG", "REVIEW", "INFO", "PASS", "SKIPPED"
SEV_ORDER = {FLAG: 0, REVW: 1, INFO: 2, PASS: 3, "EXPLAINED": 3, SKIP: 4}

class F:
    __slots__ = ("ref", "title", "sev", "detail", "rows", "raw")
    def __init__(self, ref, title, sev, detail, rows=None, raw=None):
        self.ref, self.title, self.sev, self.detail, self.rows = ref, title, sev, detail, rows or []
        # NEW: `raw` -- optional list of per-occurrence dicts, richer than what `.rows` shows on
        # the per-month EWB sheet (which stays exactly as it was -- nothing here changes that
        # display). This is what the new FY-wide annual-detail sheets consume, so a check's
        # underlying records survive past the point where the per-month Finding used to discard
        # them into a single summary line.
        self.raw = raw or []


# ----------------------------------------------------------------------
# EWB parser
# ----------------------------------------------------------------------
def _split_doc(v):
    """'MR22-23/0519 - 10/01/2023' -> ('MR22-23/0519', date(2023,1,10))."""
    s = str(v or "").strip()
    if not s:
        return "", None
    parts = re.split(r"\s+-\s+", s, maxsplit=1)
    docno = parts[0].strip()
    dt = None
    if len(parts) > 1:
        dt = _parse_dt(parts[1])
    return docno, dt

def _split_ewb(v):
    """'301546430758 - 10/01/2023 17:01:00' -> ('301546430758', datetime)."""
    s = str(v or "").strip()
    if not s:
        return "", None
    parts = re.split(r"\s+-\s+", s, maxsplit=1)
    ewbno = parts[0].strip()
    dt = _parse_dt(parts[1]) if len(parts) > 1 else None
    return ewbno, dt

def _parse_dt(s):
    s = str(s or "").strip()
    if not s:
        return None
    s = s.split()[0]  # drop time part
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return _dt.datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None

def _gstin_of(combined):
    """'05AAECM6380J1ZA / M.R...' -> '05AAECM6380J1ZA'."""
    return str(combined or "").split("/")[0].strip()

def parse_ewb(path):
    """Return list of dicts for one EWB file (outward or inward)."""
    out = []
    if not path or not os.path.exists(path):
        return out
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c).strip() if c else "" for c in rows[0]]
    H = {h: i for i, h in enumerate(hdr)}
    def g(r, *names):
        for n in names:
            if n in H and H[n] < len(r):
                return r[H[n]]
        return None
    for r in rows[1:]:
        if not any(r):
            continue
        docno, docdt = _split_doc(g(r, "Doc No. & Dt."))
        ewbno, ewbdt = _split_ewb(g(r, "EWB No. & Dt."))
        out.append(dict(
            ewbno=ewbno or str(g(r, "EWB No.") or "").strip(),
            ewbdate=ewbdt,
            docno=docno, docdate=docdt,
            from_gstin=_gstin_of(g(r, "From GSTIN & Name")),
            to_gstin=_gstin_of(g(r, "To GSTIN & Name")),
            assess=num(g(r, "Assess Val.")),
            taxval=num(g(r, "Tax Val.")),
            hsn=str(g(r, "HSN Code") or "").strip(),
            vehicle=str(g(r, "Latest Vehicle No.") or "").strip(),
        ))
    return out


# ----------------------------------------------------------------------
# Pull GSTR-1 & e-invoice line detail (reuse analysis layer's reader if present)
# ----------------------------------------------------------------------
def read_gstr1_invoices(path, month):
    """invno -> dict(taxable, igst, cgst, sgst, pos, gstin, rate_lines, consignment).
    Scoped to ONE month's block out of the merged GSTR-1 workbook.
    BUG FIX -- same root cause as read_gstr1_lines() above and parse_gstr1() in
    gst_parsers_returns.py: continuation rows of a multi-rate invoice carry a blank
    Invoice Number cell (merged in the source export), so they used to fall into their own
    d[""] bucket instead of accumulating into their real parent invoice's totals -- silently
    losing that portion of the invoice's taxable/tax value from this function's output.
    Fixed the same way: forward-fill the last-seen invoice identity within this month."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["b2b, sez, de_inv"]; rows = list(ws.iter_rows(values_only=True))
    H = {h: i for i, h in enumerate([str(c).strip() if c else "" for c in rows[3]])}
    def g(r, k): return r[H[k]] if k in H and H[k] < len(r) else None
    inv = {}
    last_no, last_pos, last_gstin = None, "", ""
    for r in mpu.rows_for_month(rows, 3, month):
        if not any(r):
            continue
        raw_no = g(r, "Invoice Number")
        raw_gstin = g(r, "GSTIN/UIN of Recipient")
        if raw_no not in (None, "") and raw_gstin not in (None, ""):
            last_no = str(raw_no).strip()
            last_pos = str(g(r, "Place Of Supply") or "").strip()
            last_gstin = str(raw_gstin).strip()
        no = last_no if last_no is not None else str(raw_no or "").strip()
        pos_v = last_pos if last_no is not None else str(g(r, "Place Of Supply") or "").strip()
        gstin_v = last_gstin if last_no is not None else str(raw_gstin or "").strip()
        d = inv.setdefault(no, dict(taxable=0.0, igst=0.0, cgst=0.0, sgst=0.0,
                                    invval=0.0, pos=pos_v,
                                    gstin=gstin_v,
                                    rates=set()))
        d["taxable"] += num(g(r, "Taxable Value"))
        d["igst"] += num(g(r, "Integrated Tax"))
        d["cgst"] += num(g(r, "Central Tax"))
        d["sgst"] += num(g(r, "State/UT Tax"))
        d["invval"] += num(g(r, "Invoice Value"))
        d["rates"].add(num(g(r, "Rate")))
    return inv

def read_einv_invoices(path, month):
    out = {}
    if not path or not os.path.exists(path):
        return out   # E-Invoice legitimately not supplied at all -- graceful
    wb = openpyxl.load_workbook(path, data_only=True)
    if "b2b, sez, de" not in wb.sheetnames:
        return out
    ws = wb["b2b, sez, de"]; rows = list(ws.iter_rows(values_only=True))
    if month not in mpu.months_present(rows, 3):
        return out   # E-Invoice doesn't cover this month -- same graceful state
    H = {h: i for i, h in enumerate([str(c).strip() if c else "" for c in rows[3]])}
    invcol = "Invoice number" if "Invoice number" in H else "Invoice Number"
    def g(r, k): return r[H[k]] if k in H and H[k] < len(r) else None
    for r in mpu.rows_for_month(rows, 3, month):
        if not any(r):
            continue
        no = str(g(r, invcol) or "").strip()
        d = out.setdefault(no, dict(taxable=0.0, igst=0.0, cgst=0.0, sgst=0.0, invval=0.0))
        d["taxable"] += num(g(r, "Taxable Value"))
        d["igst"] += num(g(r, "Integrated Tax"))
        d["cgst"] += num(g(r, "Central Tax"))
        d["sgst"] += num(g(r, "State/UT Tax"))
        d["invval"] += num(g(r, "Invoice Value"))
    return out


def _val_mismatch(a, b):
    if max(abs(a), abs(b)) == 0:
        return False
    return abs(a - b) > max(TOL, VALUE_TOL_PCT * max(abs(a), abs(b)))


# ----------------------------------------------------------------------
# NEW: complete invoice-detail row builders for the "PRE CHECK DETAIL ROWS
# (E-Way Bill)" sheet -- per explicit instruction, checks that used to show
# only a bare doc-number (or doc-number + consignment value) now show the
# FULL invoice/EWB record plus a GSTR-1 / E-Invoice / EWB triangulation
# flag, so the reviewer never has to re-open the source workbooks to see
# what a flagged doc-number actually was. Only touches the .rows CONTENT
# each check already builds -- no check's severity/detail-text logic
# changes.
# ----------------------------------------------------------------------
OUT_DETAIL_HDR = ("EWB Doc No", "EWB No.", "EWB Date", "Doc Date", "To GSTIN", "HSN",
                   "Vehicle No.", "EWB Assess (Rs)", "EWB Tax (Rs)",
                   "GSTR-1 Taxable (Rs)", "GSTR-1 IGST", "GSTR-1 CGST", "GSTR-1 SGST",
                   "GSTR-1 Invoice Value", "GSTR-1 POS",
                   "In GSTR-1?", "In E-Invoice?", "In EWB?", "Note / Classification")

def _out_detail_row(d, out_by_doc, g1inv, einv, note=""):
    """Full outward-side (EWB-Out) invoice record for one doc-no, with GSTR-1/
    E-Invoice/EWB triangulation flags -- used by #1, #3, #5, #15."""
    items = out_by_doc.get(d, [])
    i0 = items[0] if items else {}
    assess = sum(x.get("assess", 0.0) for x in items)
    taxval = sum(x.get("taxval", 0.0) for x in items)
    g1 = (g1inv or {}).get(d)
    ei = (einv or {}).get(d) if einv else None
    return (d, i0.get("ewbno", ""), i0.get("ewbdate"), i0.get("docdate"), i0.get("to_gstin", ""),
            i0.get("hsn", ""), i0.get("vehicle", ""), round(assess, 2), round(taxval, 2),
            round(g1["taxable"], 2) if g1 else None,
            round(g1["igst"], 2) if g1 else None,
            round(g1["cgst"], 2) if g1 else None,
            round(g1["sgst"], 2) if g1 else None,
            round(g1["invval"], 2) if g1 else None,
            g1["pos"] if g1 else None,
            "Y" if g1 else "-", "Y" if ei else "-", "Y" if items else "-", note)

IN_DETAIL_HDR = ("Doc/Invoice No", "EWB No.", "EWB Date", "Doc Date", "Supplier GSTIN",
                  "Supplier Name", "HSN", "Vehicle No.", "EWB Assess (Rs)", "EWB Tax (Rs)",
                  "2B Taxable (Rs)", "2B IGST", "2B CGST", "2B SGST", "2B Invoice Value",
                  "2B POS", "2B Rate (%)", "In EWB?", "In GSTR-2B?", "Note")

def _in_detail_row_from_ewb(k, ewbin_map, b2b_by_key, note=""):
    """Full inward-side (EWB-In) record keyed off (supplier_gstin, doc-no), with a 2B
    lookup by the SAME key -- used by #12 (EWB-In exists, no 2B match at all)."""
    gstin, docno = k
    items = ewbin_map.get(k, [])
    i0 = items[0] if items else {}
    assess = sum(x.get("assess", 0.0) for x in items)
    taxval = sum(x.get("taxval", 0.0) for x in items)
    hit = (b2b_by_key or {}).get(k)
    if hit:
        b0 = hit[0]
        taxable = round(sum(x["taxable"] for x in hit), 2)
        igst = round(sum(x["igst"] for x in hit), 2)
        cgst = round(sum(x["cgst"] for x in hit), 2)
        sgst = round(sum(x["sgst"] for x in hit), 2)
        invval = round(sum(x["invval"] for x in hit), 2)
        supplier, pos, rate, in2b = b0.get("supplier", ""), b0.get("pos", ""), b0.get("rate"), "Y"
    else:
        supplier = pos = rate = taxable = igst = cgst = sgst = invval = None
        in2b = "-"
    return (docno, i0.get("ewbno", ""), i0.get("ewbdate"), i0.get("docdate"), gstin, supplier,
            i0.get("hsn", ""), i0.get("vehicle", ""), round(assess, 2), round(taxval, 2),
            taxable, igst, cgst, sgst, invval, pos, rate,
            "Y" if items else "-", in2b, note)

def _in_detail_row_from_2b(xs, consign, note=""):
    """Full 2B-invoice record for one (supplier, invoice-no) with NO inward EWB at all --
    used by #13. `xs` is the list of matching B2B-sheet line dicts for this invoice."""
    x0 = xs[0]
    taxable = round(sum(x["taxable"] for x in xs), 2)
    igst = round(sum(x["igst"] for x in xs), 2)
    cgst = round(sum(x["cgst"] for x in xs), 2)
    sgst = round(sum(x["sgst"] for x in xs), 2)
    invval = round(sum(x["invval"] for x in xs), 2)
    return (x0.get("invno", ""), "", None, x0.get("date"), x0.get("gstin", ""), x0.get("supplier", ""),
            "", "", None, None,
            taxable, igst, cgst, sgst, invval, x0.get("pos", ""), x0.get("rate"),
            "-", "Y", note + f"  (consignment Rs {round(consign, 2):,.2f})")


def _own_month(date_str):
    """'DD/MM/YYYY' -> 'Mon-YY' (same label format as raw.PERIOD_LABEL / EWB rows' own
    'month' field), or None if unparseable.

    BUG FIX (reported: EWB-In checks #10-#13 flagging real, matched invoices as having
    'no matching inward EWB' -- confirmed against real data, GSTIN 05ASQPB9012R1ZA
    FY2023-24, a QUARTERLY GSTR-2B filer): twob_lines['b2b']/['cdnr'] are scoped to the
    marker BLOCK for the current month (see parse_2b_excel()'s own docstring) -- correct
    for ITC-availability timing, but for a quarterly filer that block is the WHOLE
    quarter, identical across all 3 of its months. ewb_in, by contrast, is scoped to
    ONLY the exact calendar month (filter_by_month() on the EWB's own date). Matching
    a quarter-wide invoice list against a single-month EWB list meant every quarterly
    invoice got checked 3 times (once per month of its quarter): it matched correctly
    in its own month, but was FALSELY flagged 'no matching inward EWB' in the other two
    -- confirmed exactly this pattern on real data (e.g. invoice '313-450-23IVBR', dated
    23/06/2023 with a real 23/06/2023 inward EWB, correctly absent from Jun-23's own
    sheet but wrongly flagged under both Apr-23 and May-23).

    Fixed at the point these checks build their invoice/note lookup, not in
    parse_2b_excel() itself (which other callers -- e.g. the FY-wide 2B invoice index,
    ITC-availability checks -- correctly rely on staying quarter-wide): narrow the
    quarter-wide list down to rows whose own invoice/note date falls in the CURRENT
    calendar month before it's used for EWB matching, restoring one-invoice-checked-
    once. The invoice's own date (not its GSTR-1 filing period) is used because EWB
    movement timing tracks the invoice date, not whenever the supplier happened to
    file."""
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", str(date_str or ""))
    if not m:
        return None
    dd, mm, yyyy = m.groups()
    try:
        mm = int(mm)
    except ValueError:
        return None
    return f"{mpu.CAL_MONTH_ABBR.get(mm, '?')}-{yyyy[2:]}"


# ----------------------------------------------------------------------
# THE 27 CHECKS
# ----------------------------------------------------------------------
def run(ewb_out, ewb_in, g1inv, einv, g3b, b2b, ewb_out_file_supplied=True, ewb_in_file_supplied=True,
        is_services_dominant=False, dominant_hsn=None, dominant_hsn_share=0.0):
    """ewb_out_file_supplied / ewb_in_file_supplied: whether the ANNUAL EWB
    workbook for that direction was supplied AT ALL this run (independent of
    whether this particular month happens to have zero rows in it, which is
    a legitimate business state, not a data gap). Defaults to True so any
    existing caller that doesn't pass these keeps its old behaviour.

    GRACEFUL DEGRADATION (fixed): previously, when an entire EWB direction
    was never supplied for this taxpayer (common for smaller taxpayers below
    the Rule-138 threshold, or service businesses with no goods movement),
    `ewb_out`/`ewb_in` arrived here as an empty list indistinguishable from
    "file present, zero rows this month" -- and several checks silently
    produced a MISLEADING result instead of an honest 'no data' state:
      #1  showed PASS ("0/0 EWB-Out doc-numbers found in GSTR-1") -- looked
          like a clean reconciliation; it was actually zero data to check.
      #4  flagged EVERY inter-state >Rs50k invoice as REVIEW ("no EWB found"),
          since the empty ewbset made every such invoice look EWB-less --
          flooding the dashboard with noise that isn't a real finding.
      #10 showed REVIEW ("0 of 0 inward-EWB documents matched") -- a false
          alarm in the opposite direction from #1's false PASS.
      #12 showed PASS ("every inward EWB matches a 2B invoice") -- true only
          because there were zero inward EWBs to fail to match.
    All four (plus #3, #13) now check the file-supplied flag FIRST and emit
    an explicit SKIP with the reason, before running their normal logic."""
    R = []
    def gv(k, i, d=0.0):
        v = g3b.get(k); return v[i] if v and i < len(v) else d

    b2b_available = b2b.get("available", True)

    # Computed unconditionally (safe: sums to 0 on an empty list) so later
    # sections that reference these can do so regardless of which branch
    # ran above -- only the FINDINGS that report them are gated behind
    # ewb_out_file_supplied, not the sums themselves.
    ewb_out_total = sum(e["assess"] for e in ewb_out)
    ewb_out_tax = sum(e["taxval"] for e in ewb_out)

    # EWB outward keyed by doc number (invoice number)
    out_by_doc = {}
    for e in ewb_out:
        out_by_doc.setdefault(e["docno"], []).append(e)
    in_by_doc = {}
    for e in ewb_in:
        in_by_doc.setdefault(e["docno"], []).append(e)

    if not ewb_out_file_supplied:
        for ref, title in [("#1", "EWB-Out invoice present in GSTR-1"), ("#3", "EWB-Out with NO matching GSTR-1 invoice"),
                            ("#4", "GSTR-1 inter-state >\u20b950k with NO EWB-Out")]:
            R.append(F(ref, title, SKIP,
                       "No outward EWB workbook was supplied for this taxpayer/FY at all -- this is a "
                       "data-availability gap (common for taxpayers under the Rule-138 threshold or with "
                       "no goods movement), not evidence of a missing EWB on any specific invoice. "
                       "Skipped rather than shown as PASS or REVIEW to avoid a false-clean or "
                       "false-flagged result."))
    if not ewb_in_file_supplied:
        for ref, title in [("#10", "EWB-In invoice matched to GSTR-2B"), ("#12", "EWB-In with NO matching GSTR-2B invoice"),
                            ("#13", "GSTR-2B inter-state >\u20b950k with NO inward EWB")]:
            R.append(F(ref, title, SKIP,
                       "No inward EWB workbook was supplied for this taxpayer/FY at all -- data-availability "
                       "gap, not a finding. Skipped rather than shown as PASS or REVIEW."))

    # ===== A. EWB-Out vs GSTR-1 =====
    if ewb_out_file_supplied:
        g1set = set(g1inv) - {"None", ""}
        ewbset = set(out_by_doc) - {""}
        matched = g1set & ewbset
        no_g1 = sorted(ewbset - g1set)

        # Tax-bearing/zero-tax/purchase-return classification -- computed ONCE, shared by #1 and
        # #3 (used to be #3-only; #1's OWN severity used to fire on ANY unmatched doc, including
        # ordinary zero-tax stock-transfer movements that #3 already correctly treats as
        # expected -- same classification now drives both checks' severity consistently).
        cdnr_idx = {}
        twob_lines = b2b.get("_lines") if b2b_available else None
        if twob_lines:
            # Narrowed to this month's own notes (by their own date) -- same quarter-vs-month
            # scoping mismatch as _own_month() documents for #10-#13 below applies here too
            # (out_by_doc is month-scoped; twob_lines['cdnr'] is quarter-wide for a quarterly
            # filer), and a VALUE-based match (state+taxable+tax, not invoice/note number) only
            # gets MORE collision-prone against a 3x-larger, quarter-wide candidate pool.
            for c in [c for c in twob_lines["cdnr"] if _own_month(c.get("date")) == raw.PERIOD_LABEL]:
                ct = c["igst"] + c["cgst"] + c["sgst"]
                cdnr_idx.setdefault((state_code(c["gstin"]) or c["gstin"][:2],
                                     round(c["taxable"], 2), round(ct, 2)), []).append(c)
        pur_return, taxed, zero = [], [], []
        for d in no_g1:
            t = sum(x["taxval"] for x in out_by_doc[d])
            a = sum(x["assess"] for x in out_by_doc[d])
            to = out_by_doc[d][0]["to_gstin"]
            if t <= TOL:
                zero.append((d, to, round(a, 2), round(t, 2), "zero-tax movement (approval/stock-transfer)"))
                continue
            key = (state_code(to) or to[:2], round(a, 2), round(t, 2))
            match = cdnr_idx.get(key)
            if match:
                note = match[0]["note"]
                pur_return.append((d, to, round(a, 2), round(t, 2), f"matches 2B credit note {note} (purchase return)"))
            else:
                taxed.append((d, to, round(a, 2), round(t, 2), "tax-bearing, NO 2B credit-note match — investigate"))

        # BUG FIX (reported against real output): #1's detail table used to list the SAME
        # (taxed + pur_return + zero) rows as #3 -- i.e. the UNMATCHED docs -- even though #1 is
        # titled "invoice present in GSTR-1". That made #1 and #3's detail tables look identical
        # (confusing on its own), and on a taxpayer whose outward EWB export has 'Tax Val.' as 0
        # for nearly every row (confirmed against real data: 123/132 for this taxpayer -- a
        # source-file characteristic, not a parsing bug; inward EWB on the SAME taxpayer is 98%
        # populated) almost everything fell into the 'zero' bucket, making the two tables not just
        # similarly-shaped but near-IDENTICAL in content. Fixed at the source of the confusion:
        # #1's own detail table now shows the docs it's actually named for -- the MATCHED ones --
        # so #1 and #3 are complementary (matched vs unmatched), not duplicates.
        sev1 = FLAG if taxed else (INFO if (pur_return or zero) else PASS)
        R.append(F("#1", "EWB-Out invoice present in GSTR-1", sev1,
                   f"{len(matched)}/{len(ewbset)} EWB-Out doc-numbers found in GSTR-1 "
                   f"(GSTR-1 invoices: {len(g1set)}). {len(no_g1)} EWB doc(s) not in GSTR-1: "
                   f"{len(taxed)} genuinely unexplained (tax-bearing, no GSTR-1 invoice, no 2B "
                   f"credit-note match), {len(pur_return)} reconciled as purchase returns, "
                   f"{len(zero)} zero-tax movements (expected). Detail below is the "
                   f"{len(matched)} MATCHED doc(s) this check is named for; see #3 for the "
                   f"unmatched ones. Note on 'EWB Tax (Rs)' showing 0 for many/most rows: that is "
                   f"the outward EWB export's own 'Tax Val.' column, not derived by this tool -- "
                   f"where it's genuinely 0 in the source file, the real tax (if any) is still "
                   f"shown separately in the 'GSTR-1 IGST/CGST/SGST' columns of the same row, "
                   f"sourced from the matching GSTR-1 invoice instead.",
                   [OUT_DETAIL_HDR] +
                   [_out_detail_row(d, out_by_doc, g1inv, einv, "Matched to GSTR-1 invoice")
                    for d in sorted(matched)]))


        # #2 value mismatch (EWB consignment 'assess' is post-discount taxable; compare to GSTR-1 taxable)
        vm = []
        for doc in sorted(matched):
            ewb_assess = sum(x["assess"] for x in out_by_doc[doc])
            g1_tax = g1inv[doc]["taxable"]
            if _val_mismatch(ewb_assess, g1_tax):
                vm.append((doc, ewb_assess, g1_tax, round(ewb_assess - g1_tax, 2)))
        R.append(F("#2", "EWB-Out value vs GSTR-1 taxable", PASS if not vm else REVW,
                   f"{len(vm)} invoice(s) with >1% value gap (EWB assessable can differ from GSTR-1 "
                   "taxable due to freight/discount; review, not auto-flag).",
                   [("Invoice", "EWB assess", "GSTR-1 taxable", "diff")] + vm))

        # #3 EWB-Out exists but NO GSTR-1 -- same three buckets as before, now reusing the
        # classification computed above instead of recomputing it.
        sev3 = FLAG if taxed else (INFO if (pur_return or zero) else PASS)
        detail3 = []
        if taxed:
            detail3.append(f"{len(taxed)} TAX-BEARING outward EWB(s) with NO GSTR-1 invoice AND no matching 2B "
                           "credit note -> goods moved with tax but not in GSTR-1; investigate. ")
        if pur_return:
            detail3.append(f"{len(pur_return)} outward EWB(s) RECONCILED as PURCHASE RETURNS — each matches a "
                           "supplier credit note in GSTR-2B (GSTIN + taxable + tax), so the goods went back to the "
                           "supplier and correctly do NOT appear as a GSTR-1 outward sale. No action. ")
        if zero:
            detail3.append(f"{len(zero)} ZERO-TAX movement(s) (approval / stock-transfer / job-work challan) "
                           "correctly absent from GSTR-1; verify invoiced on sale/return. ")
        if not no_g1:
            detail3 = ["Every outward EWB has a GSTR-1 invoice."]
        R.append(F("#3", "EWB-Out with NO matching GSTR-1 invoice", sev3, "".join(detail3),
                   [OUT_DETAIL_HDR] +
                   [_out_detail_row(d, out_by_doc, g1inv, einv, note)
                    for (d, to, a, t, note) in (taxed + pur_return + zero)]))

        # #4 GSTR-1 inter-state >50k with NO EWB-Out.
        # NOT an automatic Rule 138 violation: EWB is required only if goods physically move.
        # Job-work (JWI) / pure-service / delivery-challan cases may legitimately have no EWB.
        # Classify by document-number prefix and route to REVIEW with the right question, not FLAG.
        miss_ewb = []
        own = SELF_GSTIN[:2]
        for no, d in g1inv.items():
            if no in ("None", ""):
                continue
            pos_code = state_code(d["pos"]) or name_to_code(d["pos"])
            inter = pos_code and pos_code != own
            consign = d["taxable"] + d["igst"] + d["cgst"] + d["sgst"]
            if inter and consign > EWB_THRESHOLD and no not in ewbset:
                pfx = re.match(r"([A-Za-z]+)", no)
                pfx = pfx.group(1) if pfx else ""
                jobwork = pfx.upper() in ("JWI", "JW")
                miss_ewb.append((no, d["pos"], round(consign, 2),
                                 "job-work invoice — verify goods moved" if jobwork
                                 else "verify goods movement"))
        # severity: REVIEW (a missing EWB on an inter-state >50k supply is a question, not a proven breach)
        sev4 = REVW if miss_ewb else PASS
        R.append(F("#4", "GSTR-1 inter-state >₹50k with NO EWB-Out", sev4,
                   ("EWB is mandatory only where goods physically move. These inter-state >₹50k supplies have "
                    "no outward EWB — Rule 138 applies ONLY if goods actually moved. Job-work (JWI) / pure-service "
                    "/ delivery-challan movements may legitimately have no EWB under this GSTIN. VERIFY physical "
                    "movement per invoice before treating as a violation; this is a review item, not a proven breach.")
                   if miss_ewb else "All inter-state >₹50k GSTR-1 invoices have an EWB.",
                   [("Invoice", "POS", "consignment ₹", "action")] + miss_ewb))
    else:
        # SKIP findings for #1/#3/#4 already appended above; ewbset must still exist
        # (as an empty set) for later sections (#5-9, #15-18) that reference it.
        ewbset = set()
        matched = set()



    # ===== B. EWB-Out vs E-Invoice =====
    if einv:
        einvset = set(einv) - {"None", ""}
        # BUG FIX: ewbset includes zero-tax movements (delivery challan / stock transfer), which
        # legitimately have NO e-invoice (e-invoicing applies to taxable B2B supplies only) --
        # comparing the FULL ewbset against einvset fired every month regardless of whether the
        # genuinely taxable EWBs matched. Scoped to tax-bearing docs only, mirroring #1/#3's own
        # zero-tax exclusion.
        ewbset_taxbearing = {d for d in ewbset if sum(x["taxval"] for x in out_by_doc.get(d, [])) > TOL}
        # #5 match. RENAMED (same reasoning as #1's own title fix above): the detail table
        # here has always been the GAP list -- ewbset_taxbearing - einvset, i.e. the docs with
        # NO e-invoice match -- but the title read as if it listed the matched ones, and
        # gst_report.py's EWB Detail sheet silently excluded this ref entirely on the (here
        # wrong) assumption that every ref in that exclusion list only ever holds confirmatory/
        # matched rows. Title now says what the rows actually are; no longer excluded from the
        # per-month EWB Detail sheet (see write_eway() in gst_report.py).
        em = ewbset_taxbearing & einvset
        R.append(F("#5", "EWB-Out doc(s) NOT present in E-Invoice", PASS if ewbset_taxbearing <= einvset else REVW,
                   f"{len(em)}/{len(ewbset_taxbearing)} TAX-BEARING EWB-Out doc-numbers found in e-invoice "
                   f"({len(ewbset) - len(ewbset_taxbearing)} zero-tax movements excluded -- e-invoicing "
                   f"applies to taxable B2B supplies only). (B2C / sub-threshold EWBs may legitimately "
                   f"have no e-invoice.) Detail below is the docs with NO e-invoice match.",
                   [OUT_DETAIL_HDR] +
                   [_out_detail_row(d, out_by_doc, g1inv, einv, "no matching e-invoice")
                    for d in sorted(ewbset_taxbearing - einvset)]))
        # #6 value mismatch
        vm6 = []
        for doc in sorted(em):
            ea = sum(x["assess"] for x in out_by_doc[doc]); et = einv[doc]["taxable"]
            if _val_mismatch(ea, et):
                vm6.append((doc, ea, et, round(ea - et, 2)))
        R.append(F("#6", "EWB-Out value vs E-Invoice taxable", PASS if not vm6 else REVW,
                   f"{len(vm6)} invoice(s) with >1% gap.",
                   [("Invoice", "EWB assess", "E-inv taxable", "diff")] + vm6))
        # #7 EWB date vs e-invoice... e-invoice file has no per-invoice date kept here -> use doc date vs ewb date
        gap7 = []
        for doc in sorted(em):
            for x in out_by_doc[doc]:
                if x["ewbdate"] and x["docdate"]:
                    g = (x["ewbdate"] - x["docdate"]).days
                    if g > 1:
                        gap7.append((doc, x["docdate"], x["ewbdate"], g))
        R.append(F("#7", "EWB-date vs invoice(doc)-date gap (>1 day)",
                   PASS if not gap7 else REVW,
                   "EWB generated well after invoice date -> delayed generation; verify movement timing. "
                   "(Return-filing-date checks #10/#21 skipped: no filing date supplied.)",
                   [("Invoice", "Doc date", "EWB date", "gap days")] + gap7))
    else:
        for ref, t in [("#5", "EWB-Out vs E-Invoice match"), ("#6", "EWB-Out vs E-Invoice value"),
                       ("#7", "EWB-date vs invoice-date gap")]:
            R.append(F(ref, t, INFO, "E-invoice file not supplied."))

    # ===== C. EWB-Out vs GSTR-3B =====
    # #8 aggregate outward EWB vs 3B 3.1(a): BOTH assessable and tax.
    if ewb_out_file_supplied:
        b3b_tax = gv("3.1a", 0)                                  # taxable value
        b3b_outtax = gv("3.1a", 1) + gv("3.1a", 2) + gv("3.1a", 3)   # IGST+CGST+SGST
        ratio8 = ewb_out_total / b3b_tax if b3b_tax else 0
        tax_gap = ewb_out_tax - b3b_outtax
        sev8 = REVW if tax_gap > TOL else INFO
        R.append(F("#8", "EWB-Out aggregate vs GSTR-3B 3.1(a)", sev8,
                   f"Assessable: EWB-Out {ewb_out_total:,.2f} vs 3B 3.1(a) taxable {b3b_tax:,.2f} "
                   f"(ratio {ratio8:.2f}). TAX: EWB-Out tax {ewb_out_tax:,.2f} vs 3B 3.1(a) output tax "
                   f"{b3b_outtax:,.2f} -> EWB tax higher by {abs(tax_gap):,.2f}. "
                   + ("NOT A NEW FINDING: the outward EWB carries more tax than 3B output tax because some outward "
                      "EWBs are PURCHASE RETURNS (goods sent back to suppliers, matching supplier credit notes in "
                      "2B — see #3), not outward sales. Those returns inflate outward EWB tax but correctly never "
                      "hit 3B output liability. The gap is explained by #3's reclassification, not a separate "
                      "suppression item. " if tax_gap > TOL else "")
                   + "EWB covers only goods-movement supplies (not services / B2C sub-threshold), so EWB < 3B "
                     "is normal; a true EWB-tax > 3B-output-tax gap that ISN'T purchase-returns is the signal.",
                   []))

        # #9 tax type consistency: EWB inter/intra (from->to state) vs GSTR-1 head
        tt = []
        for doc in sorted(matched):
            e = out_by_doc[doc][0]
            fr, to = state_code(e["from_gstin"]), state_code(e["to_gstin"])
            if not (fr and to):
                continue
            inter = fr != to
            d = g1inv[doc]
            has_igst = d["igst"] > TOL
            has_local = d["cgst"] > TOL or d["sgst"] > TOL
            if inter and has_local and not has_igst:
                tt.append((doc, f"{fr}->{to} inter", "GSTR-1 has CGST/SGST"))
            if (not inter) and has_igst and not has_local:
                tt.append((doc, f"{fr}->{to} intra", "GSTR-1 has IGST"))
        R.append(F("#9", "Tax-type (inter/intra) EWB vs GSTR-1 head", PASS if not tt else FLAG,
                   "EWB movement direction contradicts the tax head charged in GSTR-1."
                   if tt else "EWB direction matches GSTR-1 tax head on all matched invoices.",
                   [("Invoice", "EWB direction", "GSTR-1")] + tt))
    else:
        R.append(F("#8", "EWB-Out aggregate vs GSTR-3B 3.1(a)", SKIP,
                   "No outward EWB workbook supplied -- cannot aggregate."))
        R.append(F("#9", "Tax-type (inter/intra) EWB vs GSTR-1 head", SKIP,
                   "No outward EWB workbook supplied -- cannot test."))

    # ===== D. EWB-In vs GSTR-2B =====
    ewb_in_assess = sum(e["assess"] for e in ewb_in)
    ewb_in_tax = sum(e["taxval"] for e in ewb_in)
    # b2b_itc specifically needs '_summary_available' (the narrower 'ITC Available sheet itself
    # was readable' signal) -- ITC_all_other_* are SUMMARY-SHEET fields, genuinely 0.0 (not
    # missing) when that sheet isn't present, even on a taxpayer whose full B2B/CDNR
    # invoice-level data parsed perfectly. twob_lines below correctly keeps using the broader
    # 'available' flag, since _lines holds that same invoice-level data, unaffected either way.
    b2b_itc = ((b2b["ITC_all_other_IGST"] + b2b["ITC_all_other_CGST"] + b2b["ITC_all_other_SGST"])
               if b2b.get("_summary_available") else None)
    twob_lines = b2b.get("_lines") if b2b_available else None   # set when 2B came from Excel

    if not ewb_in_file_supplied:
        pass  # SKIP findings for #10/#12/#13 already appended at top; #11 handled just below.
    elif twob_lines:
        # ---- LINE-LEVEL (2B Excel invoice list available) ----
        # NARROWED to this calendar month's own invoices/notes (by their own date) --
        # twob_lines['b2b']/['cdnr'] are quarter-wide for a quarterly GSTR-2B filer; see
        # _own_month()'s docstring for why EWB matching specifically needs month, not
        # quarter, granularity (checks #10-#13 below). Other B2B/CDNR consumers (e.g. the
        # ITC-availability summary) are untouched -- they read `b2b`/`twob_lines` directly,
        # not this narrowed `b2b_inv`.
        this_month = raw.PERIOD_LABEL
        b2b_inv = [x for x in twob_lines["b2b"] if _own_month(x.get("date")) == this_month]
        cdnr_inv = [c for c in twob_lines.get("cdnr", []) if _own_month(c.get("date")) == this_month]
        def nkey(g, n): return (str(g).strip().upper(), str(n).strip().upper())
        # primary index: (supplier, invoice-no).  value index: (supplier_state, taxable, tax)
        b2b_map, b2b_val = {}, {}
        for x in b2b_inv:
            b2b_map.setdefault(nkey(x["gstin"], x["invno"]), []).append(x)
            vk = (str(x["gstin"]).strip().upper(), round(x["taxable"], 2),
                  round(x["igst"]+x["cgst"]+x["sgst"], 2))
            b2b_val.setdefault(vk, []).append(x)
        # NEW (Bug 3, point 4): a supplier CREDIT/DEBIT NOTE against the original invoice is
        # also a legitimate match for an inward movement -- e.g. goods physically moving back
        # under a note reference rather than the original invoice number. Indexed the same way
        # as b2b_map, by (supplier_gstin, note_number).
        cdnr_map = {}
        for c in cdnr_inv:
            cdnr_map.setdefault(nkey(c["gstin"], c["note"]), []).append(c)
        ewbin_map = {}
        for e in ewb_in:
            ewbin_map.setdefault(nkey(e["from_gstin"], e["docno"]), []).append(e)

        def matches_2b(k):
            """('invoice-no'|'cdnr'|'value'|None, matched_rows). Invoice-number and CDNR-note
            matches are both EXACT-key matches (either is a genuine, no-ambiguity resolution);
            the value fallback is NOT -- see the point-4/point-2 note below on why it's kept
            structurally separate rather than folded into the same bucket."""
            if k in b2b_map:
                return ("invoice-no", b2b_map[k])
            if k in cdnr_map:
                return ("cdnr", cdnr_map[k])
            es = ewbin_map[k]
            ea = round(sum(x["assess"] for x in es), 2); et = round(sum(x["taxval"] for x in es), 2)
            vk = (k[0], ea, et)
            if vk in b2b_val:
                return ("value", b2b_val[vk])
            return (None, None)

        # BUG FIX (Bug 3, points 2-3): the supplier+value fallback used to be silently counted
        # as an equal-weight "matched" result alongside exact invoice-number matches, AND
        # zero-tax inward movements (delivery challans / stock transfers -- the SAME pattern #3
        # already excludes on the outward side) were never excluded from the "chase supplier"
        # count. Now: exact (invoice-no or CDNR-note) matches are the only ones counted as
        # genuinely resolved; value-only matches are their OWN "possible match, verify manually"
        # bucket, never silently folded in; zero-tax movements are excluded from #12 entirely,
        # mirroring #3.
        matched_exact, matched_value, by_value = {}, {}, 0
        for k in ewbin_map:
            how, hit = matches_2b(k)
            if how in ("invoice-no", "cdnr"):
                matched_exact[k] = (how, hit)
            elif how == "value":
                matched_value[k] = (how, hit)
                by_value += 1
        matched_in = {**matched_exact, **matched_value}   # kept for #11's value-match reuse below

        # #10 invoice matching
        R.append(F("#10", "EWB-In invoice matched to GSTR-2B (no. or value)",
                   PASS if matched_exact else (REVW if matched_value else REVW),
                   f"{len(matched_exact)} of {len(ewbin_map)} inward-EWB documents matched a 2B "
                   f"invoice or credit/debit note EXACTLY (by invoice number or note number). "
                   f"{len(matched_value)} more matched ONLY by supplier+value (invoice-number "
                   f"format differed) -- these are a SEPARATE 'possible match, verify manually' "
                   f"bucket, not counted as a clean match; see #12 for exactly which. 2B B2B "
                   f"invoices: {len(b2b_map)}.",
                   []))

        # #11 value match on matched-by-number set (value-matched are equal by construction)
        vmism = []
        for k, (how, hit) in matched_exact.items():
            if how != "invoice-no":
                continue
            ea = sum(x["assess"] for x in ewbin_map[k]); ba = sum(x["taxable"] for x in hit)
            if _val_mismatch(ea, ba):
                vmism.append((k[1], k[0], round(ea, 2), round(ba, 2), round(ea-ba, 2)))
        R.append(F("#11", "EWB-In vs GSTR-2B value (matched invoices)",
                   PASS if not vmism else REVW,
                   f"{len(vmism)} matched invoice(s) with >1% taxable gap (under/over-invoicing signal). "
                   "EWB assessable can differ from 2B taxable for freight/discount; review." if vmism
                   else "All number-matched inward-EWB/2B invoices agree on value within tolerance.",
                   [("Invoice", "Supplier", "EWB assess", "2B taxable", "diff")] + vmism))

        # #12 EWB-In with NO 2B match by exact key (neither invoice-no nor CDNR-note) --
        # value-only matches are reported as their OWN "possible match" bucket (point 2), and
        # zero-tax movements are excluded entirely (point 3, mirrors #3's outward treatment).
        # A fourth check, BEFORE calling anything a genuine gap: is this supplier's invoice
        # sitting in a DIFFERENT month's 2B (late filing) rather than genuinely absent all year?
        # Confirmed real against this taxpayer's data -- treating same-month-only absence as a
        # gap was flagging ordinary filing-timing differences as "supplier hasn't filed".
        no_exact = sorted(k for k in ewbin_map if k not in matched_exact)
        zero12, possible12, timing12, only_ewb = [], [], [], []
        fy_index = getattr(raw, "FY_2B_INVOICE_INDEX", None) or {}
        for k in no_exact:
            t = sum(x["taxval"] for x in ewbin_map[k])
            a = sum(x["assess"] for x in ewbin_map[k])
            if t <= TOL:
                zero12.append((k[1], k[0], round(a, 2), round(t, 2)))
                continue
            if k in matched_value:
                possible12.append((k[1], k[0], round(a, 2), round(t, 2)))
                continue
            fy_hits = fy_index.get(k)
            if fy_hits:
                other_months = sorted({m for m, _ in fy_hits})
                timing12.append((k[1], k[0], round(a, 2), round(t, 2),
                                 f"filed in 2B for {', '.join(other_months)} (not this EWB's own month)"))
                continue
            only_ewb.append(k)
        R.append(F("#12", "EWB-In with NO matching GSTR-2B invoice",
                   FLAG if only_ewb else (INFO if (zero12 or possible12 or timing12) else PASS),
                   f"{len(only_ewb)} TAX-BEARING inward EWB(s) with NO 2B match ANYWHERE in the FY "
                   f"(not this month, not any other month, not by number, not by supplier+value) -> "
                   f"supplier hasn't filed at all; ITC not yet available -- chase these suppliers. "
                   f"{len(timing12)} more matched a 2B invoice in a DIFFERENT month (ordinary filing-"
                   f"timing difference, not a gap). {len(possible12)} matched ONLY by supplier+value "
                   f"in this same month -- verify manually. {len(zero12)} zero-tax movement(s) "
                   f"excluded (mirrors #3's outward treatment). Full EWB record for each below "
                   f"(no matching 2B invoice, so all 2B-side columns are blank).",
                   [IN_DETAIL_HDR] +
                   [_in_detail_row_from_ewb(k, ewbin_map, b2b_map, "no 2B match anywhere in FY")
                    for k in only_ewb]))

        # #13 2B invoice >50k inter-state with NO inward EWB
        own = SELF_GSTIN[:2]
        matched_2b_ids = set()
        for k, (how, hit) in matched_in.items():
            for x in hit:
                matched_2b_ids.add((x["gstin"].upper(), x["invno"].upper()))
        only_2b = []
        seen2b = {}
        for x in b2b_inv:
            seen2b.setdefault((x["gstin"].upper(), x["invno"].upper()), []).append(x)
        for key, xs in seen2b.items():
            if key in matched_2b_ids:
                continue
            sup_state = state_code(xs[0]["gstin"])
            taxable = sum(x["taxable"] for x in xs)
            consign = taxable + sum(x["igst"]+x["cgst"]+x["sgst"] for x in xs)
            inter = sup_state and sup_state != own
            if inter and consign > EWB_THRESHOLD:
                only_2b.append((key, round(consign, 2)))
        raw13 = [dict(invno=k[1], gstin=k[0], consignment=c,
                      supplier=seen2b[k][0].get("supplier", ""))
                 for k, c in only_2b]
        R.append(F("#13", "GSTR-2B inter-state >₹50k with NO inward EWB",
                   PASS if not only_2b else REVW,
                   f"{len(only_2b)} inter-state 2B invoice(s) >₹50k with no matching inward EWB -> supplier "
                   "may not have generated an EWB (their Rule 138 issue) or goods moved on a challan. Verify; "
                   "affects your defensibility if questioned." if only_2b
                   else "All inter-state >₹50k 2B invoices have a matching inward EWB. Full invoice detail "
                   "for each 2B line below.",
                   [IN_DETAIL_HDR] +
                   [_in_detail_row_from_2b(seen2b[k], c, "no matching inward EWB") for k, c in only_2b],
                   raw=raw13))

        # #28 NEW CHECK (per explicit request): the inward mirror of #5's "EWB doc NOT in
        # E-Invoice". There's no separate inward e-invoice FILE to compare against -- an
        # e-invoice portal export for a GSTIN only ever contains invoices where THAT GSTIN is
        # the supplier (outward). What actually tells you whether the SUPPLIER e-invoiced an
        # inward supply is GSTR-2B's own per-invoice 'Source'/'IRN' columns (now read into
        # parse_2b_excel()'s b2b/cdnr dicts): 'Source' reads "E-Invoice" with a real IRN when
        # the supplier e-invoiced it, blank otherwise. Scoped to inward EWB docs that DO have a
        # matched 2B invoice/note (#10/#12) -- an unmatched one is already #12's problem, not
        # this one's. INFO, not FLAG: this tool cannot verify the SUPPLIER's own e-invoicing
        # turnover threshold, so a missing IRN is a "verify", not a proven breach.
        not_einvoiced = []
        for k, (how, hit) in matched_exact.items():
            src = (hit[0].get("einv_source") or "").strip()
            irn = (hit[0].get("irn") or "").strip()
            if src.lower() != "e-invoice" and not irn:
                not_einvoiced.append((k[1], k[0], hit[0].get("supplier", ""),
                                       round(sum(x["taxable"] for x in hit), 2),
                                       "supplier invoice not e-invoiced (no IRN in 2B) -- verify if "
                                       "e-invoicing was mandatory for this supplier"))
        R.append(F("#28", "Inward EWB doc(s) -- supplier invoice NOT e-invoiced",
                   INFO if not_einvoiced else PASS,
                   f"{len(not_einvoiced)} matched inward EWB document(s) whose 2B invoice/note carries no "
                   f"e-invoice IRN (GSTR-2B 'Source' != 'E-Invoice'). Informational: this tool cannot verify "
                   f"whether e-invoicing was mandatory for that specific supplier -- verify against their "
                   f"turnover." if not_einvoiced
                   else "Every matched inward EWB document's 2B invoice/note carries a real e-invoice IRN.",
                   [("Doc/Invoice No", "Supplier GSTIN", "Supplier Name", "2B Taxable (Rs)", "Note")]
                   + not_einvoiced))
    else:
        # ---- AGGREGATE ONLY: either 2B wasn't supplied at all, or it was supplied but only
        # as a PDF summary (no line-level invoice list) -- these are different situations and
        # get different messages rather than one generic "insufficient" note. ----
        if not b2b_available:
            reason = (f"GSTR-2B not supplied for this month ({b2b.get('_reason', 'no reason recorded')})"
                      " -- line-level matching not possible.")
        else:
            reason = ("GSTR-2B supplied as PDF summary (no invoice list) -> line-level matching not possible. "
                      "Supply the GSTR-2B Excel download to enable #10-#13.")
        R.append(F("#10", "EWB-In invoice match GSTR-2B (line-level)", INFO if b2b_available else SKIP,
                   reason + " Aggregate compare in #11/#26.", []))
        if b2b_itc is not None:
            R.append(F("#11", "EWB-In aggregate vs GSTR-2B ITC", INFO,
                       f"Inward EWB assessable {ewb_in_assess:,.2f}, EWB tax {ewb_in_tax:,.2f}. "
                       f"2B 'all other ITC' tax {b2b_itc:,.2f}. CAUTION — DIFFERENT BASES, NOT A LIKE-FOR-LIKE "
                       "RATIO: EWB-In is goods-only, 2B ITC includes services+goods+RCM. Scale context only.", []))
        else:
            R.append(F("#11", "EWB-In aggregate vs GSTR-2B ITC", SKIP,
                       f"GSTR-2B not supplied -- cannot compare. Inward EWB assessable {ewb_in_assess:,.2f}, "
                       f"EWB tax {ewb_in_tax:,.2f} shown for reference only.", []))
        R.append(F("#12", "EWB-In exists but no 2B entry", INFO if b2b_available else SKIP,
                   "Needs 2B invoice list (PDF summary insufficient)." if b2b_available
                   else "GSTR-2B not supplied.", []))
        R.append(F("#13", "2B entry but no EWB-In (>50k inter)", INFO if b2b_available else SKIP,
                   "Needs 2B invoice list (PDF summary insufficient)." if b2b_available
                   else "GSTR-2B not supplied.", []))
        R.append(F("#28", "Inward EWB doc(s) -- supplier invoice NOT e-invoiced", INFO if b2b_available else SKIP,
                   "Needs 2B invoice list (PDF summary insufficient)." if b2b_available
                   else "GSTR-2B not supplied.", []))

    # ===== E. skipped (no books) =====
    R.append(F("#14", "Unaccounted purchases (EWB-In vs books)", SKIP, "No purchase register supplied."))

    # ===== F. EWB-Out vs EWB-In (same transaction) =====
    # For a single GSTIN, outward and inward EWBs are different counterparties; overlap only if
    # the same doc-no appears on both sides (rare). Report overlaps, else N/A.
    overlap = set(out_by_doc) & set(in_by_doc) - {""}
    raw15 = []
    for d in sorted(overlap):
        o_assess = sum(x["assess"] for x in out_by_doc[d]); o_tax = sum(x["taxval"] for x in out_by_doc[d])
        i_assess = sum(x["assess"] for x in in_by_doc[d]); i_tax = sum(x["taxval"] for x in in_by_doc[d])
        raw15.append(dict(docno=d, out_gstin=out_by_doc[d][0]["to_gstin"], in_gstin=in_by_doc[d][0]["from_gstin"],
                          out_assess=o_assess, out_tax=o_tax, in_assess=i_assess, in_tax=i_tax,
                          ewbdate=out_by_doc[d][0]["ewbdate"]))
    SAME_DOC_HDR = ("Doc No", "EWB Date", "Out To-GSTIN", "In From-GSTIN",
                     "Out EWB Assess (Rs)", "Out EWB Tax (Rs)", "In EWB Assess (Rs)", "In EWB Tax (Rs)",
                     "In GSTR-1?", "In E-Invoice?")
    rows15 = [SAME_DOC_HDR]
    for rec in raw15:
        d = rec["docno"]
        rows15.append((d, rec["ewbdate"], rec["out_gstin"], rec["in_gstin"],
                       round(rec["out_assess"], 2), round(rec["out_tax"], 2),
                       round(rec["in_assess"], 2), round(rec["in_tax"], 2),
                       "Y" if (g1inv or {}).get(d) else "-",
                       "Y" if (einv or {}).get(d) else "-"))
    R.append(F("#15", "EWB-Out vs EWB-In value (same doc-no)",
               PASS if not overlap else REVW,
               "No document number appears on both outward and inward EWB (expected for a single GSTIN: "
               "your outward = others' inward, not in your own download)."
               if not overlap else f"{len(overlap)} doc-no on both sides; verify.",
               rows15,
               raw=raw15))

    # #16 EWB time gap generation vs document date (both sides)
    gap16 = []
    for tag, lst in (("OUT", ewb_out), ("IN", ewb_in)):
        for e in lst:
            if e["ewbdate"] and e["docdate"]:
                g = (e["ewbdate"] - e["docdate"]).days
                if abs(g) > 2:
                    gap16.append((tag, e["docno"], e["docdate"], e["ewbdate"], g))
    R.append(F("#16", "EWB generation vs document date gap (>2 days)",
               PASS if not gap16 else REVW,
               "Large gap between document date and EWB generation -> verify genuine movement timing."
               if gap16 else "All EWBs generated within 2 days of document date.",
               [("Side", "doc-no", "doc date", "EWB date", "gap")] + gap16[:50]))

    # ===== G. Triangulation GSTR-1 + E-Inv + EWB-Out =====
    if einv:
        # BUG FIX: same zero-tax exclusion as #5 -- a delivery-challan/stock-transfer EWB
        # legitimately has no GSTR-1 invoice and no e-invoice; including it in the triangulation
        # universe flagged it every month regardless of whether genuinely taxable invoices
        # triangulated correctly.
        tri = []
        raw17 = []
        allinv = (set(g1inv) | set(einv) | ewbset_taxbearing) - {"None", ""}
        for no in sorted(allinv):
            in_g1 = no in g1inv; in_ei = no in einv; in_ew = no in ewbset_taxbearing
            if not (in_g1 and in_ei and in_ew):
                tri.append((no, "Y" if in_g1 else "-", "Y" if in_ei else "-", "Y" if in_ew else "-"))
                val = g1inv.get(no, {}).get("taxable") if no in g1inv else (
                      sum(x["assess"] for x in out_by_doc.get(no, [])) or None)
                raw17.append(dict(invno=no, in_g1=in_g1, in_einv=in_ei, in_ewb=in_ew, taxable=val))
        R.append(F("#17", "Triangulation: GSTR-1 / E-Invoice / EWB-Out", PASS if not tri else REVW,
                   "Invoices not present in all three sources (zero-tax EWB movements excluded -- see #5). "
                   "EWB-absent (goods) can be legitimate for services/B2C; GSTR-1-absent but EWB-present "
                   "= suppression signal (see #3)."
                   if tri else "Every tax-bearing invoice appears in all three sources.",
                   [("Invoice", "GSTR-1", "E-Inv", "EWB-Out")] + tri[:80],
                   raw=raw17))
    else:
        R.append(F("#17", "Triangulation GSTR-1/E-Inv/EWB-Out", INFO, "E-invoice file not supplied."))

    # #18 HSN rate-wise across sources -- EWB has HSN code, GSTR-1 HSN sheet has rate-wise.
    hsn_out = {}
    for e in ewb_out:
        h = e["hsn"][:4] if e["hsn"] else "?"
        hsn_out[h] = hsn_out.get(h, 0.0) + e["assess"]
    R.append(F("#18", "HSN rate-wise across sources", INFO,
               "EWB-Out HSN-wise assessable (4-digit): "
               + "; ".join(f"{k}:{v:,.0f}" for k, v in sorted(hsn_out.items()))
               + ". Cross-check against GSTR-1 HSN summary rate rows for misclassification "
               "(HSN-summary rate split must be read from the GSTR-1 'hsn' sheet).",
               []))

    # ===== H. skipped =====
    R.append(F("#19", "3B ITC (4A5) vs EWB-In aggregate", SKIP,
               "Skipped per scope (purchase/books side). Aggregate context shown in #11/#26."))
    R.append(F("#20", "E-Invoice (purchase) vs EWB-In", SKIP, "No purchase-side e-invoice data."))

    # ===== I. timing =====
    R.append(F("#21", "EWB gen date vs GSTR-1 filing date", SKIP, "No GSTR-1 filing date supplied."))
    R.append(F("#22", "EWB validity expiry vs supply date", INFO,
               "EWB export has no validity/expiry column -> cannot test. Download detailed EWB with "
               "'Valid Upto' to enable."))

    # #23 multiple EWBs per invoice (partial dispatch / threshold-splitting)
    multi = [(doc, len(lst), round(sum(x["assess"] for x in lst), 2))
             for doc, lst in out_by_doc.items() if doc and len(lst) > 1]
    R.append(F("#23", "Multiple EWBs per invoice (partial dispatch)",
               PASS if not multi else REVW,
               "Allowed for partial dispatch, but multiple sub-threshold EWBs on one invoice can be "
               "used to dodge e-invoice/EWB limits -> verify." if multi else
               "No invoice has more than one outward EWB.",
               [("Invoice", "#EWBs", "total assess")] + multi))

    R.append(F("#24", "EWB cancelled after return filing", INFO,
               "EWB export has no cancellation-status column -> cannot test. Download with status to enable."))

    # ===== J. risk ratios =====
    # #25 outward EWB value / GSTR-1 taxable
    g1_total_tax_val = sum(d["taxable"] for no, d in g1inv.items())
    if ewb_out_file_supplied:
        ratio25 = ewb_out_total / g1_total_tax_val if g1_total_tax_val else 0
        # BUG FIX (Bug 4): e-way bills apply to goods movement only (Rule 138) -- a business
        # whose dominant HSN/SAC is a SERVICES code (GSTN prefixes these '99', e.g. 995424
        # "General construction services...") genuinely has little/nothing to move, so a low
        # EWB-value/invoice-value ratio is EXPECTED for that HSN profile, not evidence of
        # accommodation billing. Gated on the dominant-HSN fact (computed from this month's own
        # GSTR-1 HSN summary, cited by name and share below -- see the 'HSN RATE REVIEW' table on this month's Comparison sheet), not a
        # blanket "always explained" -- a services-dominant business whose ratio is unusually
        # low even for THAT profile, or a goods-dominant business with a low ratio, still flags.
        if ratio25 < 0.9 and is_services_dominant:
            sev25 = "EXPLAINED"
            why25 = (f"Ratio <0.9, but this month's dominant HSN/SAC is {dominant_hsn} "
                     f"({dominant_hsn_share:.0%} of taxable value), a SERVICES code -- e-way "
                     f"bills apply only to goods movement (Rule 138), so a low ratio is expected "
                     f"for this HSN profile, not a red flag. See the 'HSN RATE REVIEW' table on this month's Comparison sheet.")
        elif ratio25 < 0.9:
            sev25 = FLAG
            why25 = "Ratio <0.9 -> sizeable B2B supply with no goods movement; check accommodation bills."
        else:
            sev25 = PASS
            why25 = "Ratio ≥0.9 -> most B2B supply backed by goods movement. (B2C/services not in EWB-Out.)"
        R.append(F("#25", "Ratio: EWB-Out assessable / GSTR-1 B2B taxable",
                   sev25,
                   f"EWB-Out {ewb_out_total:,.2f} / GSTR-1 B2B {g1_total_tax_val:,.2f} = {ratio25:.3f}. "
                   + why25,
                   []))
    else:
        R.append(F("#25", "Ratio: EWB-Out assessable / GSTR-1 B2B taxable", SKIP,
                   "No outward EWB workbook supplied -- cannot compute.", []))

    # #26 inward EWB value / 2B taxable value (Excel) or ITC tax (PDF fallback)
    if twob_lines:
        b2b_taxable = sum(x["taxable"] for x in twob_lines["b2b"])
        ratio26 = ewb_in_assess / b2b_taxable if b2b_taxable else 0
        R.append(F("#26", "Ratio: EWB-In assessable / 2B B2B taxable", INFO,
                   f"EWB-In assessable {ewb_in_assess:,.2f} / 2B B2B taxable {b2b_taxable:,.2f} = {ratio26:.2f}. "
                   "Now a like-for-like VALUE ratio (both taxable value) from the 2B Excel. EWB-In covers only "
                   "goods movement while 2B B2B also includes goods received without an e-way bill (sub-"
                   "threshold / local), so EWB-In < 2B is normal; EWB-In >> 2B would suggest goods received "
                   "without matching ITC documents.",
                   []))
    elif b2b_itc is not None:
        R.append(F("#26", "Ratio: EWB-In assessable / 2B ITC", INFO,
                   f"EWB-In assessable {ewb_in_assess:,.2f}; 2B ITC tax {b2b_itc:,.2f}; EWB-In tax "
                   f"{ewb_in_tax:,.2f}. DO NOT OVER-WEIGHT: numerator is goods-movement value, denominator is "
                   "ITC tax (goods+services+RCM) — different bases/units. A true value ratio needs the 2B "
                   "Excel (taxable value), which isn't in the PDF summary. Scale context only.",
                   []))
    else:
        R.append(F("#26", "Ratio: EWB-In assessable / 2B", SKIP,
                   f"GSTR-2B not supplied -- cannot compute. EWB-In assessable {ewb_in_assess:,.2f} shown "
                   "for reference only.", []))

    # #27 same vehicle multiple trips between same GSTIN pair (circular-trading indicator)
    # BUG FIX (per this session's dedup/annual-detail prompt): the underlying per-trip records
    # (vehicle, doc-no, date) were already being computed into `veh` below, then discarded --
    # only the (vehicle, from, to, #trips) SUMMARY made it into `.rows`. Now retained in full via
    # `.raw` so the new "EWB Pattern Checks -- Annual Summary" sheet can show actual EWB/vehicle
    # numbers instead of a narrative "eyeball the high-frequency pairs" line. `.rows` (the
    # per-month EWB sheet's own display) is unchanged.
    veh = {}
    for tag, lst in (("OUT", ewb_out), ("IN", ewb_in)):
        for e in lst:
            if not e["vehicle"]:
                continue
            key = (e["vehicle"], e["from_gstin"], e["to_gstin"])
            veh.setdefault(key, []).append((tag, e["docno"], e["ewbdate"], e["assess"], e["taxval"]))
    rep = [(v, fr, to, len(trips)) for (v, fr, to), trips in veh.items() if len(trips) > 1]
    rep.sort(key=lambda x: -x[3])
    raw27 = []
    for (v, fr, to), trips in veh.items():
        if len(trips) <= 1:
            continue
        for tag, docno, ewbdate, assess, taxval in trips:
            raw27.append(dict(vehicle=v, from_gstin=fr, to_gstin=to, side=tag, docno=docno,
                               ewbdate=ewbdate, assess=assess, taxval=taxval))
    R.append(F("#27", "Same vehicle, repeated trips on same GSTIN pair",
               PASS if not rep else REVW,
               "Repeated vehicle between the same GSTIN pair can be normal (regular supplier) but is also "
               "a circular-trading indicator -> eyeball the high-frequency pairs." if rep else
               "No vehicle repeats on the same GSTIN pair.",
               [("Vehicle", "From", "To", "#trips")] + rep[:30],
               raw=raw27))

    R.sort(key=lambda x: (SEV_ORDER[x.sev], x.ref))
    return R


# ----------------------------------------------------------------------
# WRITE
# ----------------------------------------------------------------------
FILL = {FLAG: PatternFill("solid", fgColor="FFC7CE"), REVW: PatternFill("solid", fgColor="FFEB9C"),
        INFO: PatternFill("solid", fgColor="DDEBF7"), PASS: PatternFill("solid", fgColor="C6EFCE"),
        SKIP: PatternFill("solid", fgColor="E7E6E6")}
FONTC = {FLAG: Font(bold=True, color="9C0006"), REVW: Font(bold=True, color="9C6500"),
         INFO: Font(bold=True, color="2F5496"), PASS: Font(bold=True, color="006100"),
         SKIP: Font(bold=True, color="808080")}
HEAD = PatternFill("solid", fgColor="1F3864")
BORDER = Border(*[Side(style="thin", color="BFBFBF")] * 4)

def write(R, outpath):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "EWB Findings"
    ws.cell(1, 1, f"E-WAY BILL RECONCILIATION (27-check matrix) — {raw.PERIOD_LABEL}").font = Font(bold=True, size=13, color="1F3864")
    ws.cell(2, 1, f"GSTIN {SELF_GSTIN}  |  {COMPANY_NAME or '(company auto-detected)'}").font = Font(size=9, italic=True)
    import datetime as _d
    _src = f"2B Excel: {_LAST_2B_FILE}" if (_LAST_2B_SRC == "excel") else "2B PDF summary (no Excel found)"
    ws.cell(4, 1, f"BUILD v3 (line-level 2B + e-invoice + #3 purchase-return reclass)  |  "
                  f"generated {_d.datetime.now():%Y-%m-%d %H:%M:%S}  |  source: {_src}").font = \
        Font(size=9, italic=True, color="C00000")
    counts = {s: sum(1 for x in R if x.sev == s) for s in (FLAG, REVW, INFO, PASS, SKIP)}
    ws.cell(3, 1, "  ".join(f"{s}: {c}" for s, c in counts.items())).font = Font(bold=True, size=10)
    hdr = ["Ref", "Check", "Result", "Detail"]
    for i, h in enumerate(hdr, 1):
        c = ws.cell(5, i, h); c.fill = HEAD; c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = BORDER
    r = 6
    detail_blocks = []
    for f in R:
        ws.cell(r, 1, f.ref)
        ws.cell(r, 2, f.title)
        cv = ws.cell(r, 3, f.sev); cv.fill = FILL[f.sev]; cv.font = FONTC[f.sev]
        cv.alignment = Alignment(horizontal="center")
        ws.cell(r, 4, f.detail)
        for c in range(1, 5):
            cell = ws.cell(r, c); cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(c in (2, 4)))
            if c != 3:
                cell.font = Font(size=10)
        r += 1
        if len(f.rows) > 1:
            detail_blocks.append(f)
    for col, w in zip("ABCD", [6, 42, 10, 110]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A6"

    # Detail sheet: the actual offending rows for each check that has them
    ws2 = wb.create_sheet("EWB Detail")
    rr = 1
    ws2.cell(rr, 1, "PER-CHECK DETAIL ROWS").font = Font(bold=True, size=12, color="1F3864"); rr += 2
    for f in detail_blocks:
        ws2.cell(rr, 1, f"{f.ref}  {f.title}  [{f.sev}]").font = Font(bold=True, color="1F3864"); rr += 1
        head = f.rows[0]
        for j, h in enumerate(head, 1):
            c = ws2.cell(rr, j, h); c.font = Font(bold=True, size=9, color="FFFFFF"); c.fill = HEAD; c.border = BORDER
        rr += 1
        for row in f.rows[1:]:
            for j, v in enumerate(row, 1):
                # guard: a string starting with = + - @ is treated by Excel as a formula; prefix a space
                if isinstance(v, str) and v[:1] in ("=", "+", "-", "@"):
                    v = " " + v
                c = ws2.cell(rr, j, v)
                c.border = BORDER; c.font = Font(size=10)
                if isinstance(v, (int, float)):
                    c.number_format = "#,##0.00"
            rr += 1
        rr += 1
    for col, w in zip("ABCDE", [22, 18, 18, 14, 10]):
        ws2.column_dimensions[col].width = w

    wb.save(outpath)
    return counts


def main_eway():
    """Legacy/manual-testing entry point -- master_build.py does NOT call this;
    it drives read_gstr1_invoices/read_einv_invoices/run() directly per month
    via run_monthly_pipeline.py. Kept working for standalone spot-checks:
        python gst_eway_recon.py <month, e.g. Jan-23>
    (still requires raw.GSTR1_FILE / GSTR3B_FILE / EINV_FILE / GSTR2B_FILE and
    ewb_annual_parser-based ewb_out/ewb_in lists to be set up by the caller.)
    """
    import sys as _sys, os as _os
    global _LAST_2B_SRC, _LAST_2B_FILE, _LAST_EINV_FILE
    if len(_sys.argv) < 2:
        raise SystemExit("Usage: python gst_eway_recon.py <month, e.g. Jan-23>")
    month = _sys.argv[1]

    print("="*70)
    print(f"GST E-WAY BILL RECONCILIATION — input check ({month})")
    print("="*70)
    for label, fn in [("GSTR-1", raw.GSTR1_FILE), ("E-Invoice", raw.EINV_FILE),
                      ("GSTR-3B", raw.GSTR3B_FILE), ("GSTR-2B xlsx", GSTR2B_FILE)]:
        ok = fn and _os.path.exists(fn)
        print(f"  {'OK ' if ok else 'MISS'} {label:13} {fn if fn else '(not found)'}")
    print("="*70)

    g1inv = read_gstr1_invoices(raw.GSTR1_FILE, month)
    einv = read_einv_invoices(raw.EINV_FILE, month) if raw.EINV_FILE else {}
    _LAST_EINV_FILE = _os.path.basename(raw.EINV_FILE) if raw.EINV_FILE else None
    g3b = raw.parse_gstr3b(raw.GSTR3B_FILE, month)
    b2b = g2b.summary_for_month(GSTR2B_FILE, month)

    R = run([], [], g1inv, einv, g3b, b2b)
    _LAST_2B_SRC = b2b.get("_source", "pdf-hardcoded").split("-")[0]
    _LAST_2B_FILE = b2b.get("_file")
    counts = write(R, OUTPUT_FILE)
    src = b2b.get("_source", "pdf-hardcoded")
    print(f"Saved: {OUTPUT_FILE}")
    print(f"EWB-Out/In: not loaded in this standalone stub (use master_build.py for the full run)  |  2B source: {src}"
          + (f" ({len(b2b['_lines']['b2b'])} B2B invoices, {len(b2b['_lines']['cdnr'])} CDNR notes)"
             if b2b.get("_lines") else ""))
    print(f"E-invoice: {'loaded ('+str(len(einv))+' invoices)' if einv else 'NOT FOUND -> #5/#6/#7/#17 limited'}")
    if src != "excel":
        print("  >>> NOTE: 2B Excel was NOT used. #10-#13 and #3 reclass are LIMITED.")
    print("  " + "  ".join(f"{s}:{c}" for s, c in counts.items()))
    for f in R:
        print(f"  [{f.sev:8}] {f.ref:4} {f.title}")


