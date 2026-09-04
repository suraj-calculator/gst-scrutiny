#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GST REPORT
==========
CONSOLIDATED FILE -- contains what used to be: gst_unified_scrutiny.py, build_annual_workbook.py

The tool was reorganised from 19 .py files into 9 for easier sharing. Nothing
in the analytical logic was rewritten during that move: each section below is
the original module's code verbatim, with only (a) intra-project imports
repointed at the new file names, (b) its standalone __main__ demo block
removed, and (c) the renames listed under MERGE NOTES applied where two merged
modules happened to define the same top-level name with different bodies.

MERGE NOTES for this file:
  - gst_unified_scrutiny.gather -> gather_legacy_single_month
  - gst_unified_scrutiny.main -> main_unified_legacy
  - build_annual_workbook.main -> main_annual_standalone
  - build_annual_workbook.SEV_FILL -> SEV_FILL_ANNUAL
  - build_annual_workbook.SEV_FONT -> SEV_FONT_ANNUAL
  - build_annual_workbook.gather -> gather_annual
"""


# ============================================================================
# ==== SECTION: gst_unified_scrutiny.py  (was a standalone module before consolidation)
# ============================================================================
"""
GST UNIFIED SCRUTINY TOOL  (single workbook, both pipelines)
============================================================
Combines, in ONE Excel workbook, the two reconciliations that were previously
produced as two separate files:

  PIPELINE 1  (was GST_Scrutiny_Comparison_Jan2023.xlsx)
      - Raw side-by-side comparison : GSTR-1 | 2B | 3B | E-Invoice
        (sections A, A2, B, B2, C, D, D2 — exactly as the original tool)
      - PLUS Sooraj's 14 interpretive checks (#0-#14) that were already coded
        in gst_analysis_checks.py but had never been written into the
        Comparison workbook. They are now included as the "Analysis (14 checks)"
        sheet, ON THE SAME LINES that the comparison already reconciles.

  PIPELINE 2  (was GST_Scrutiny_EWayBill.xlsx)
      - The 27-check E-Way-Bill matrix (inward + outward), unchanged.

  CROSS-FILE  (new)
      - A "Dashboard" sheet that puts every FLAG / MISMATCH / REVIEW from BOTH
        pipelines in one ranked list, so the two reconciliations are read
        together rather than in two files.

HARD RULE honoured:
  - Comparison runs on EXACTLY the same lines/sections as before.
  - E-Way-Bill runs on EXACTLY the same 27 checks as before.
  - Sooraj's 14 points map 1:1 onto the existing analysis checks (#0-#14);
    nothing new is invented and NO safety net / fabricated data is added.
  - All compute logic is REUSED from the three existing modules
    (gst_scrutiny_tool, gst_analysis_checks, gst_eway_recon). This file only
    orchestrates them and writes the combined book. The engines are untouched,
    so results are bit-for-bit identical to the two original files.

A NOTE ON SOORAJ POINT #3 (the "4C mismatch"):
  Sooraj computed 4C = 4A5 - 4B2 and got a gap vs the filed 4C. That manual
  formula OMITS the RCM-ITC row (Table 4A3). The correct identity is
  4C = 4A5 + 4A3 - 4B2, which reconciles to the rupee (diff 0.00). The tool
  uses the correct identity (analysis check #3) and therefore PASSES it. The
  "mismatch" was in the manual arithmetic, not in the return. This is stated
  explicitly in the #3 finding row.

USAGE
    Put this next to the four modules + the input files (same folder the two
    original tools already use), set nothing else, then:
        python gst_unified_scrutiny.py
    -> GST_Scrutiny_Unified.xlsx
"""

import os
import datetime as _dt
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- reuse the three existing engines (no logic change) -------------------
import gst_parsers_returns as raw
import gst_checks_monthly as ana
import gst_checks_monthly as eway
try:
    import gst_parsers_returns as g2b
except ImportError:
    g2b = None

num = raw.num
TOL = raw.TOLERANCE
OUTPUT_FILE = "GST_Scrutiny_Unified.xlsx"


# ----------------------------------------------------------------------
# ARN / filing-date auto-extract (enables analysis checks #8 and #10)
# ----------------------------------------------------------------------
# Sooraj asked for #8 (IRN-date vs GSTR-1 filing lag) and #10 (GSTR-1 vs
# GSTR-3B filing gap). Both need the filing dates. The GST-portal exports
# ALREADY carry these as the ARN date:
#   GSTR-1  : 'Read me' sheet, row 'ARN date'   (value in col C)
#   GSTR-3B : 'GSTR-3B'  sheet, row 'Date of ARN'(value in col E)
# The base analysis module only read them from a hardcoded CONFIG attr that
# was left blank, so #8/#10 silently fell to INFO. We read the ARN date
# straight from the file and inject it, so the existing checks fire. No new
# check is added — we only supply the date the existing checks already want.
def _extract_arn_dates():
    out = {"GSTR1_FILING_DATE": None, "GSTR3B_FILING_DATE": None}
    # ---- GSTR-1 'Read me' -> 'ARN date' ----
    try:
        wb = openpyxl.load_workbook(raw.GSTR1_FILE, data_only=True)
        sn = "Read me" if "Read me" in wb.sheetnames else wb.sheetnames[0]
        for r in wb[sn].iter_rows(values_only=True):
            label = next((str(c).strip() for c in r if c not in (None, "")), "")
            if label.upper() in ("ARN DATE", "DATE OF ARN"):
                vals = [c for c in r if c not in (None, "")]
                if len(vals) >= 2:
                    out["GSTR1_FILING_DATE"] = str(vals[-1]).strip()
                break
    except Exception:
        pass
    # ---- GSTR-3B 'GSTR-3B' -> 'Date of ARN' ----
    try:
        wb = openpyxl.load_workbook(raw.GSTR3B_FILE, data_only=True)
        sn = "GSTR-3B" if "GSTR-3B" in wb.sheetnames else wb.sheetnames[0]
        for r in wb[sn].iter_rows(values_only=True):
            cells = [str(c).strip() for c in r if c not in (None, "")]
            if cells and cells[0].upper() in ("DATE OF ARN", "ARN DATE") and len(cells) >= 2:
                out["GSTR3B_FILING_DATE"] = cells[-1]
                break
    except Exception:
        pass
    return out

# ----------------------------------------------------------------------
# Shared styling (matches the two original workbooks)
# ----------------------------------------------------------------------
RED    = PatternFill("solid", fgColor="FFC7CE")
GREEN  = PatternFill("solid", fgColor="C6EFCE")
AMBER  = PatternFill("solid", fgColor="FFEB9C")
BLUE   = PatternFill("solid", fgColor="DDEBF7")
GREY   = PatternFill("solid", fgColor="E7E6E6")
HEAD   = PatternFill("solid", fgColor="1F3864")
SECT   = PatternFill("solid", fgColor="D9E1F2")
TITLEF = Font(bold=True, size=13, color="1F3864")
BORDER = Border(*[Side(style="thin", color="BFBFBF")] * 4)


def _safe_cell(ws, row, col, value):
    """Guard against a real, confirmed corruption class: openpyxl auto-detects a leading '='
    in a string cell value as a formula, not text (found and fixed one real instance of this
    in this workbook's own authored note text; applied here as a systematic guard against the
    same thing happening from raw taxpayer-sourced text -- a name, description, or reference
    string that happens to start with '=' in the source data)."""
    c = ws.cell(row, col, value)
    if c.data_type == "f" and isinstance(value, str):
        c.data_type = "s"
    return c

# ---- NEW: rupee-magnitude severity bands (distinct from the Result/status colors above --
# this is a SECOND, additive axis: "how big is the unresolved gap", not "what kind of finding
# is this"). Applied via a new 'Severity' column on the Master Dashboard and, where a rupee
# amount exists, on monthly Comparison/EWB Detail rows -- never replaces the existing Result
# column or its own coloring.
CRIT_FILL = PatternFill("solid", fgColor="C00000")     # dark red
CRIT_FONT = Font(bold=True, color="FFFFFF")
HIGH_FILL = PatternFill("solid", fgColor="F4B183")     # red-orange
MED_FILL  = AMBER                                       # amber (reused -- same "needs attention" tone)
LOW_FILL  = PatternFill("solid", fgColor="E2EFDA")     # soft green (paler than GREEN, which means "verified correct")
RESOLVED_FILL = PatternFill("solid", fgColor="D6DCE5") # light grey-blue -- "explained, visible, not actionable"

SEV_BAND_FILL = {"Critical": CRIT_FILL, "High": HIGH_FILL, "Medium": MED_FILL, "Low": LOW_FILL,
                  "Resolved / Informational": RESOLVED_FILL}
SEV_BAND_FONT = {"Critical": CRIT_FONT}


def severity_band(amount, resolved=False):
    """Map a rupee amount to one of the 5 severity bands from the guiding principle. `resolved`
    (a flag genuinely explained per the guiding-principle's two conditions) always wins over the
    magnitude -- a huge but fully-explained gap is 'Resolved / Informational', not 'Critical'.
    amount=None (nothing to band, e.g. a purely structural/count-based finding) returns None --
    callers should leave the Severity cell blank rather than force a band onto a non-rupee row."""
    if resolved:
        return "Resolved / Informational"
    if amount is None:
        return None
    a = abs(amount)
    if a >= 1000000:
        return "Critical"
    if a >= 100000:
        return "High"
    if a >= 10000:
        return "Medium"
    return "Low"


SEV_FILL = {"FLAG": RED, "MISMATCH": RED, "REVIEW": AMBER, "INFO": BLUE,
            "PASS": GREEN, "MATCH": GREEN, "SKIPPED": GREY, "EXPLAINED": RESOLVED_FILL}
SEV_FONT = {"FLAG": Font(bold=True, color="9C0006"),
            "MISMATCH": Font(bold=True, color="9C0006"),
            "REVIEW": Font(bold=True, color="9C6500"),
            "INFO": Font(bold=True, color="2F5496"),
            "PASS": Font(bold=True, color="006100"),
            "MATCH": Font(bold=True, color="006100"),
            "SKIPPED": Font(bold=True, color="808080"),
            "EXPLAINED": Font(bold=True, color="31708F")}


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


# ======================================================================
# GATHER  --  parse every source ONCE, then drive each engine
# ======================================================================
def gather_legacy_single_month():
    """** LEGACY / UNSUPPORTED for the merged-file model **
    This function pre-dates the merged-workbook migration: it still assumes
    one-file-per-month and calls raw.parse_gstr1()/parse_gstr3b()/parse_einv()
    etc. WITHOUT a month argument, plus references EWB globals that no longer
    exist on gst_eway_recon (EWB_OUT_FILE/EWB_IN_FILE/find_file). It will
    raise if called. master_build.py is the supported entry point -- it
    drives the same underlying engines correctly per month via
    run_monthly_pipeline.run_month(). This function is kept only because
    master_build.py reuses this module's write_* sheet writers, not gather_legacy_single_month()
    itself. Left as a clear, visible failure rather than quietly patched to
    "sort of" work, since fully re-plumbing a standalone single-merged-month
    CLI path was out of scope for this pass.
    """
    # ---- auto-fill filing dates from ARN (enables analysis #8 and #10) ----
    arn = _extract_arn_dates()
    if arn["GSTR1_FILING_DATE"] and not getattr(raw, "GSTR1_FILING_DATE", None):
        raw.GSTR1_FILING_DATE = arn["GSTR1_FILING_DATE"]
    if arn["GSTR3B_FILING_DATE"] and not getattr(raw, "GSTR3B_FILING_DATE", None):
        raw.GSTR3B_FILING_DATE = arn["GSTR3B_FILING_DATE"]

    # ---- shared parses ----
    g1   = raw.parse_gstr1(raw.GSTR1_FILE)
    g3b  = raw.parse_gstr3b(raw.GSTR3B_FILE)
    einv = raw.parse_einv(raw.EINV_FILE)
    b2b  = raw.get_gstr2b_values() if hasattr(raw, "get_gstr2b_values") else dict(raw.GSTR2B_VALUES)

    # ---- pipeline 1a: raw comparison rows (unchanged) ----
    comparisons, comp_raw = raw.build_comparisons()

    # ---- pipeline 1b: Sooraj's 14 checks (unchanged) ----
    g1_lines   = ana.read_gstr1_lines(raw.GSTR1_FILE)
    einv_lines = ana.read_einv_lines(raw.EINV_FILE)
    findings14 = ana.run_checks(g1, g3b, einv, b2b, g1_lines, einv_lines)

    # ---- pipeline 2: the 27-check EWB matrix (unchanged logic) ----
    # paths come from the content-based finder (see CONFIG), not filename patterns.
    ewb_out_path = eway.EWB_OUT_FILE or eway.find_file(None, [r"OUT.*EWAY|EWAY.*OUT|OUTWARD"])
    ewb_in_path  = eway.EWB_IN_FILE  or eway.find_file(None, [r"IN.*EWAY|EWAY.*IN|INWARD"])
    einv_path    = raw.EINV_FILE     or eway.find_file(None, [r"EINV", r"E[-_ ]?INVOICE"])
    ewb_out = eway.parse_ewb(ewb_out_path) if ewb_out_path else []
    ewb_in  = eway.parse_ewb(ewb_in_path) if ewb_in_path else []
    g1inv   = eway.read_gstr1_invoices(raw.GSTR1_FILE)
    einv_ew = eway.read_einv_invoices(einv_path) if einv_path else {}
    if g2b and eway.GSTR2B_FILE:
        b2b_ew = g2b.summary_or_fallback(eway.GSTR2B_FILE, raw.GSTR2B_VALUES, ".")
    else:
        b2b_ew = dict(raw.GSTR2B_VALUES); b2b_ew["_lines"] = None; b2b_ew["_source"] = "pdf-hardcoded"; b2b_ew["_file"] = None
    findings27 = eway.run(ewb_out, ewb_in, g1inv, einv_ew, g3b, b2b_ew)

    return dict(
        comparisons=comparisons, comp_raw=comp_raw,
        findings14=findings14, findings27=findings27,
        meta=dict(
            ewb_out_n=len(ewb_out), ewb_in_n=len(ewb_in),
            twob_src=b2b_ew.get("_source", "pdf-hardcoded"),
            twob_file=b2b_ew.get("_file"),
            einv_file=os.path.basename(einv_path) if einv_path else None,
            g2b_ok=bool(g2b),
        ),
    )


# ======================================================================
# WRITERS  (each produces one sheet; styling matches the originals)
# ======================================================================
HDR_COMP = ["Section", "Check", "Left source", "Left value", "Right source",
            "Right value", "Difference", "Result", "Severity", "Source Reference", "Note / Tag"]
WID_COMP = [26, 46, 16, 15, 16, 15, 14, 11, 20, 30, 55]


def _comp_rows_iter(comparisons):
    for row in comparisons:
        override = None
        if len(row) == 8:
            sect, check, llabel, lval, rlabel, rval, tag, override = row
        elif len(row) == 7:
            sect, check, llabel, lval, rlabel, rval, tag = row
        else:
            sect, check, llabel, lval, rlabel, rval = row; tag = ""
        diff = round(num(lval) - num(rval), 2)
        is_match = abs(diff) <= TOL
        # BUG FIX: status used to be computed from the raw numeric diff ONLY, completely
        # ignoring `tag` even when the tag text itself already states the gap is explained/
        # expected/a scope difference -- narrative and status were set independently. `override`
        # (8th tuple element, optional) is how a row now explicitly drives its own status when
        # the narrative has PROVEN it via one of the two guiding-principle conditions (exact
        # tie-out to a named source, or structural/legal certainty) -- never inferred from the
        # tag text by keyword-sniffing, always an explicit decision made where the row is built.
        result = override if override else ("MATCH" if is_match else "MISMATCH")
        yield (sect, check, llabel, round(num(lval), 2), rlabel,
               round(num(rval), 2), diff, result, tag)


def write_comparison(ws, comparisons, only_mismatch):
    """Replicates the original Comparison/Exceptions sheets exactly."""
    title = ("GST SCRUTINY  -  MISMATCHES ONLY  -  Period: " + raw.PERIOD_LABEL
             if only_mismatch else
             "GST SCRUTINY  -  FULL COMPARISON  -  Period: " + raw.PERIOD_LABEL)
    ws.cell(1, 1, title).font = Font(bold=True, size=13,
                                     color="C00000" if only_mismatch else "1F3864")
    if only_mismatch:
        ws.cell(2, 1, f"GSTIN {raw.SELF_GSTIN}  |  {raw.COMPANY_NAME or '(company auto-detected)'}  |  "
                      f"Tolerance: Rs {TOL}").font = Font(size=9, italic=True)
        hdr_row = 4
    else:
        hdr_row = 3
    for i, h in enumerate(HDR_COMP, 1):
        ws.cell(hdr_row, i, h)
    _style_header(ws, hdr_row, len(HDR_COMP))

    r = hdr_row + 1
    cur_sect = None
    wrote = 0
    for (sect, check, ll, lv, rl, rv, diff, result, tag) in _comp_rows_iter(comparisons):
        is_match = (result == "MATCH")
        if only_mismatch and is_match:
            continue
        if (sect != cur_sect) and not only_mismatch:
            cur_sect = sect
            ws.cell(r, 1, sect).font = Font(bold=True, size=11, color="1F3864")
            for c in range(1, len(HDR_COMP) + 1):
                ws.cell(r, c).fill = SECT
            r += 1
        ws.cell(r, 1, sect if only_mismatch else "")
        ws.cell(r, 2, check); ws.cell(r, 3, ll); ws.cell(r, 4, lv)
        ws.cell(r, 5, rl); ws.cell(r, 6, rv); ws.cell(r, 7, diff)
        ws.cell(r, 8, result)
        band = severity_band(diff, resolved=(result == "EXPLAINED"))
        sv = ws.cell(r, 9, band or "")
        if band:
            sv.fill = SEV_BAND_FILL[band]
            sv.font = SEV_BAND_FONT.get(band, Font(bold=True))
        sv.alignment = Alignment(horizontal="center")
        if "HSN" in check or sect.startswith("A2") or "HSN" in sect:
            src_ref = f"See the 'HSN RATE REVIEW' table further down this same sheet for the underlying HSN-wise rows."
        elif sect.startswith("C. RCM"):
            src_ref = f"See 'RCM Triangulation' (F8/F8a) sheet for the cash-ledger-verified version of this comparison."
        else:
            src_ref = f"Underlying figures: {ll} vs {rl} (this row's own Left/Right source values above)."
        ws.cell(r, 10, src_ref)
        ws.cell(r, 11, tag)
        for c in range(1, len(HDR_COMP) + 1):
            cell = ws.cell(r, c); cell.border = BORDER; cell.font = Font(size=10)
            if c in (4, 6, 7):
                cell.number_format = '#,##0.00'
            if c == 8:
                cell.fill = SEV_FILL[result]; cell.font = SEV_FONT[result]
                cell.alignment = Alignment(horizontal="center")
            elif not is_match and c >= 2 and c != 9:
                cell.fill = RED
        r += 1
        wrote += 1
    if only_mismatch and wrote == 0:
        ws.cell(hdr_row + 1, 1, "No mismatches beyond tolerance.").font = Font(italic=True, color="006100")
    for i, w in enumerate(WID_COMP, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    if not only_mismatch:
        ws.freeze_panes = f"A{hdr_row + 1}"
        ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(len(HDR_COMP))}{max(r-1, hdr_row)}"


def _excel_safe(v):
    """openpyxl only accepts scalar cell values (str/int/float/bool/date/None). Defensive
    guard for any table-row cell built from a check's own data -- if a list/tuple/set ever
    ends up in a .rows tuple or a .numbers dict (as happened with check A4's 'rates' list,
    which crashed the HSN & Fraud sheet's new detail table with 'Cannot convert [...] to
    Excel'), this coerces it to a readable comma-joined string instead of crashing the whole
    run. Scalars pass through unchanged so existing number formatting is untouched."""
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    if isinstance(v, (list, tuple, set, frozenset)):
        return ", ".join(str(x) for x in v)
    return str(v)


def write_analysis14(ws, findings):
    """Sooraj's 14 checks (#0-#14) — same content gst_analysis_checks writes."""
    ws.cell(1, 1, f"GST SCRUTINY — ANALYSIS (Sooraj's 14 checks) — Period {raw.PERIOD_LABEL}").font = TITLEF
    ws.cell(2, 1, f"GSTIN {raw.SELF_GSTIN}  |  {raw.COMPANY_NAME or '(company auto-detected)'}").font = Font(size=9, italic=True)
    counts = {s: sum(1 for f in findings if f.severity == s) for s in ("FLAG", "REVIEW", "INFO", "PASS")}
    ws.cell(3, 1, "   ".join(f"{k}: {v}" for k, v in counts.items())).font = Font(size=10, bold=True)

    hdr = ["Ref", "Check", "Result", "Key numbers", "Detail / arithmetic"]
    for i, h in enumerate(hdr, 1):
        ws.cell(5, i, h)
    _style_header(ws, 5, 5)
    r = 6
    detail_blocks = []
    for f in findings:
        ws.cell(r, 1, f.ref); ws.cell(r, 2, f.title)
        cv = ws.cell(r, 3, f.severity); cv.fill = SEV_FILL[f.severity]; cv.font = SEV_FONT[f.severity]
        cv.alignment = Alignment(horizontal="center")
        ws.cell(r, 4, "  ".join(f"{k}={v:,.2f}" if isinstance(v, (int, float)) else f"{k}={v}"
                                for k, v in f.numbers.items()))
        ws.cell(r, 5, f.detail)
        for c in range(1, 6):
            cell = ws.cell(r, c); cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(c in (2, 4, 5)))
            if c != 3:
                cell.font = Font(size=10)
        r += 1
        if getattr(f, "rows", None) and len(f.rows) > 1:
            detail_blocks.append(f)
    for col, w in zip("ABCDE", [6, 44, 10, 30, 95]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:E{max(r-1, 5)}"

    # ---- COMPLETE DETAILS table(s) at the bottom of this same sheet ----
    # Per instruction: wherever a check flags/reviews specific invoices (IRN-to-filing lag,
    # duplicate invoices, dropped invoice numbers, e-invoice errors), the FULL underlying
    # invoice-level evidence goes here as its own table, so the reviewer never has to re-open
    # GSTR-1/E-Invoice to see which invoices and rupee amounts are actually involved.
    if detail_blocks:
        r += 2
        ws.cell(r, 1, "COMPLETE DETAILS -- FLAGGED / REVIEW INVOICES (underlying source-file rows)").font = \
            Font(bold=True, size=12, color="1F3864")
        r += 1
        ws.cell(r, 1, "Full invoice-level detail behind every check above that names specific invoices, "
                      "pulled directly from GSTR-1/E-Invoice -- so you can verify without reopening the "
                      "source workbooks.").font = Font(size=9, italic=True)
        r += 2
        max_cols = 5
        for f in detail_blocks:
            ws.cell(r, 1, f"{f.ref}  {f.title}  [{f.severity}]").font = Font(bold=True, color="1F3864")
            r += 1
            head = f.rows[0]
            max_cols = max(max_cols, len(head))
            for j, h in enumerate(head, 1):
                c = ws.cell(r, j, h); c.font = Font(bold=True, size=9, color="FFFFFF"); c.fill = HEAD; c.border = BORDER
            r += 1
            for row in f.rows[1:]:
                for j, v in enumerate(row, 1):
                    v = _excel_safe(v)
                    if isinstance(v, str) and v[:1] in ("=", "+", "-", "@"):
                        v = " " + v
                    c = ws.cell(r, j, v); c.border = BORDER; c.font = Font(size=10)
                    if isinstance(v, (int, float)):
                        c.number_format = "#,##0.00"
                r += 1
            r += 1
        for i in range(1, max_cols + 1):
            col = get_column_letter(i)
            cur = ws.column_dimensions[col].width or 0
            want = 24 if i == 1 else (16 if i <= 9 else 20)
            if want > cur:
                ws.column_dimensions[col].width = want


def write_eway(ws_find, ws_det, findings):
    """The 27-check EWB matrix + per-check detail — same content gst_eway_recon writes."""
    ws_find.cell(1, 1, f"E-WAY BILL RECONCILIATION (27-check matrix) — {raw.PERIOD_LABEL}").font = TITLEF
    ws_find.cell(2, 1, f"GSTIN {eway.SELF_GSTIN}  |  {raw.COMPANY_NAME or '(company auto-detected)'}").font = Font(size=9, italic=True)
    counts = {s: sum(1 for x in findings if x.sev == s) for s in ("FLAG", "REVIEW", "INFO", "PASS", "SKIPPED")}
    ws_find.cell(3, 1, "  ".join(f"{s}: {c}" for s, c in counts.items())).font = Font(bold=True, size=10)
    hdr = ["Ref", "Check", "Result", "Detail"]
    for i, h in enumerate(hdr, 1):
        ws_find.cell(5, i, h)
    _style_header(ws_find, 5, 4)
    r = 6
    detail_blocks = []
    for f in findings:
        ws_find.cell(r, 1, f.ref); ws_find.cell(r, 2, f.title)
        cv = ws_find.cell(r, 3, f.sev); cv.fill = SEV_FILL[f.sev]; cv.font = SEV_FONT[f.sev]
        cv.alignment = Alignment(horizontal="center")
        ws_find.cell(r, 4, f.detail)
        for c in range(1, 5):
            cell = ws_find.cell(r, c); cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(c in (2, 4)))
            if c != 3:
                cell.font = Font(size=10)
        r += 1
        # BUG FIX (reported: headings inconsistent month to month, e.g. December's EWB Detail
        # missing #3 entirely, May showing 7 headings while other months show fewer): the old
        # filter `len(f.rows) > 1` dropped a check's WHOLE block whenever it had zero data rows
        # that month (a check that always attaches [HEADER] + [...] ends up length 1 on a clean
        # month, since the header alone doesn't count as "having rows"). That made the set of
        # headings shown vary month to month based on which checks happened to have something to
        # report, rather than showing every check consistently with an explicit "nothing this
        # month" where genuinely clean. Fixed: any check that attaches a header at all (length
        # >= 1) now always gets its block, whether or not it has data rows underneath.
        # PER EXPLICIT REQUEST: the EWB Detail sheet should only carry rows worth investigating,
        # not a full listing of everything that's confirmed present/matched -- #1 ("EWB-Out
        # invoice present in GSTR-1"), #5 ("EWB-Out invoice present in E-Invoice"), and #10
        # ("EWB-In invoice matched to GSTR-2B") are confirmatory checks whose detail table is a
        # listing of the MATCHED documents, not a mismatch -- removed from this sheet only.
        # Nothing about the check itself changes: it still runs, and its ref/title/severity/
        # summary-detail line still appears on the 'EWB' matrix sheet above (ws_find) exactly as
        # before, so the pass/fail record for these three checks is not lost, only the bulky
        # "here are all the matched documents" table that added no mismatch information here.
        if f.rows and f.ref not in ("#1", "#5", "#10"):
            detail_blocks.append(f)
    for col, w in zip("ABCD", [6, 42, 10, 110]):
        ws_find.column_dimensions[col].width = w
    ws_find.freeze_panes = "A6"
    ws_find.auto_filter.ref = f"A5:D{max(r-1, 5)}"

    # detail sheet
    rr = 1
    ws_det.cell(rr, 1, "PER-CHECK DETAIL ROWS (E-Way Bill)").font = Font(bold=True, size=12, color="1F3864"); rr += 1
    ws_det.cell(rr, 1, "Per instruction: every check below now shows the COMPLETE invoice/EWB record "
                       "(not just a bare doc-number or consignment value), including a GSTR-1 / E-Invoice / "
                       "GSTR-2B / EWB triangulation flag where applicable, so a flagged/reviewed row can be "
                       "verified without reopening the source workbooks.").font = Font(size=9, italic=True)
    rr += 2
    max_cols = 5
    for f in detail_blocks:
        ws_det.cell(rr, 1, f"{f.ref}  {f.title}  [{f.sev}]").font = Font(bold=True, color="1F3864"); rr += 1
        head = f.rows[0]
        max_cols = max(max_cols, len(head))
        for j, h in enumerate(head, 1):
            c = ws_det.cell(rr, j, h); c.font = Font(bold=True, size=9, color="FFFFFF"); c.fill = HEAD; c.border = BORDER
        rr += 1
        for row in f.rows[1:]:
            for j, v in enumerate(row, 1):
                v = _excel_safe(v)
                if isinstance(v, str) and v[:1] in ("=", "+", "-", "@"):
                    v = " " + v
                c = ws_det.cell(rr, j, v); c.border = BORDER; c.font = Font(size=10)
                if isinstance(v, (int, float)):
                    c.number_format = "#,##0.00"
            rr += 1
        rr += 1
    # column widths: wide enough for the enriched multi-column tables (doc-no/GSTIN/name/note
    # columns need more room than the old 4-5 column layout did) -- sized for however many
    # columns the widest block on THIS sheet actually used, not a fixed A-E guess.
    for i in range(1, max_cols + 1):
        col = get_column_letter(i)
        ws_det.column_dimensions[col].width = 24 if i == 1 else (16 if i <= 9 else 20)


def write_dashboard(ws, data):
    """Cross-file: every actionable item from BOTH pipelines, ranked together."""
    meta = data["meta"]
    ws.cell(1, 1, f"UNIFIED GST SCRUTINY — CROSS-FILE DASHBOARD — Period {raw.PERIOD_LABEL}").font = TITLEF
    ws.cell(2, 1, f"GSTIN {raw.SELF_GSTIN}  |  {raw.COMPANY_NAME or '(company auto-detected)'}  |  "
                  "one ranked view of Comparison + Analysis + E-Way-Bill").font = Font(size=9, italic=True)
    stamp = (f"generated {_dt.datetime.now():%Y-%m-%d %H:%M:%S}  |  "
             f"2B source: {meta['twob_src']}"
             + (f" ({meta['twob_file']})" if meta['twob_file'] else "")
             + f"  |  E-Invoice: {meta['einv_file'] or 'not found'}"
             + f"  |  EWB-Out lines: {meta['ewb_out_n']}  EWB-In lines: {meta['ewb_in_n']}")
    ws.cell(3, 1, stamp).font = Font(size=9, italic=True, color="C00000")

    # collect actionable items
    items = []  # (rank_sev, pipeline, ref, title, result, detail)
    RANK = {"FLAG": 0, "MISMATCH": 0, "REVIEW": 1, "INFO": 2, "PASS": 3, "MATCH": 3, "EXPLAINED": 3, "SKIPPED": 4}

    # P1a: comparison mismatches
    for (sect, check, ll, lv, rl, rv, diff, result, tag) in _comp_rows_iter(data["comparisons"]):
        if result == "MISMATCH":
            det = f"{ll}={lv:,.2f} vs {rl}={rv:,.2f} (diff {diff:,.2f}). {tag}".strip()
            items.append((RANK[result], "Comparison", sect.split(".")[0], check, result, det))
    # P1b: Sooraj 14 — only FLAG/REVIEW shown in the dashboard (PASS/INFO live on their sheet)
    for f in data["findings14"]:
        if f.severity in ("FLAG", "REVIEW"):
            items.append((RANK[f.severity], "Analysis(14)", f.ref, f.title, f.severity, f.detail))
    # P2: EWB 27 — FLAG/REVIEW
    for f in data["findings27"]:
        if f.sev in ("FLAG", "REVIEW"):
            items.append((RANK[f.sev], "E-Way-Bill", f.ref, f.title, f.sev, f.detail))

    items.sort(key=lambda x: (x[0], x[1], x[2]))

    nflag = sum(1 for it in items if it[4] in ("FLAG", "MISMATCH"))
    nrev = sum(1 for it in items if it[4] == "REVIEW")
    ws.cell(4, 1, f"ACTIONABLE ITEMS: {len(items)}   (FLAG/MISMATCH: {nflag}   REVIEW: {nrev})").font = Font(bold=True, size=11, color="C00000")

    hdr = ["Pipeline", "Ref / Section", "Check", "Result", "Detail"]
    for i, h in enumerate(hdr, 1):
        ws.cell(6, i, h)
    _style_header(ws, 6, 5)
    r = 7
    for (_, pipeline, ref, title, result, detail) in items:
        ws.cell(r, 1, pipeline); ws.cell(r, 2, ref); ws.cell(r, 3, title)
        cv = ws.cell(r, 4, result); cv.fill = SEV_FILL[result]; cv.font = SEV_FONT[result]
        cv.alignment = Alignment(horizontal="center")
        ws.cell(r, 5, detail)
        for c in range(1, 6):
            cell = ws.cell(r, c); cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(c in (3, 5)))
            if c != 4:
                cell.font = Font(size=10)
        r += 1
    if not items:
        ws.cell(7, 1, "No FLAG / MISMATCH / REVIEW across either pipeline.").font = Font(italic=True, color="006100")
    for col, w in zip("ABCDE", [14, 16, 42, 11, 110]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A7"


def write_raw(ws, comp_raw):
    """Audit trail — identical to the original Comparison 'Raw Values' sheet."""
    ws.cell(1, 1, "RAW EXTRACTED VALUES (audit trail - what the tool read from each file)").font = Font(bold=True, size=11)
    r = 3
    for src, d in [("GSTR-1", comp_raw["g1"]), ("E-Invoice", comp_raw["einv"]),
                   ("GSTR-2B", comp_raw["b2b"])]:
        ws.cell(r, 1, src).font = Font(bold=True, color="1F3864"); r += 1
        for k, v in d.items():
            if k.startswith("_"):
                continue
            if isinstance(v, dict):
                uniq = len(set(kk[0] for kk in v.keys())) if v else 0
                ws.cell(r, 2, k + " (line count)"); ws.cell(r, 3, len(v)); r += 1
                ws.cell(r, 2, k + " (unique invoices)"); ws.cell(r, 3, uniq); r += 1
                continue
            if isinstance(v, bool):
                ws.cell(r, 2, k); ws.cell(r, 3, str(v)); r += 1
                continue
            ws.cell(r, 2, k); c = ws.cell(r, 3, round(num(v), 2)); c.number_format = '#,##0.00'; r += 1
        r += 1
    ws.cell(r, 1, "GSTR-3B (parsed tables)").font = Font(bold=True, color="1F3864"); r += 1
    for k, v in comp_raw["g3b"].items():
        ws.cell(r, 2, k); ws.cell(r, 3, str(v)); r += 1
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 20


# ======================================================================
# MAIN
# ======================================================================
def main_unified_legacy():
    data = gather_legacy_single_month()
    wb = openpyxl.Workbook()

    # Sheet order: Dashboard first (read both pipelines together), then each pipeline.
    ws_dash = wb.active; ws_dash.title = "Dashboard"
    write_dashboard(ws_dash, data)

    write_comparison(wb.create_sheet("Exceptions"), data["comparisons"], only_mismatch=True)
    write_comparison(wb.create_sheet("Full Comparison"), data["comparisons"], only_mismatch=False)
    write_analysis14(wb.create_sheet("Analysis (14 checks)"), data["findings14"])
    write_eway(wb.create_sheet("EWB Findings"), wb.create_sheet("EWB Detail"), data["findings27"])
    write_raw(wb.create_sheet("Raw Values"), data["comp_raw"])

    wb.save(OUTPUT_FILE)

    # console summary
    n_comp_mis = sum(1 for row in _comp_rows_iter(data["comparisons"]) if row[7] == "MISMATCH")
    n14_flag = sum(1 for f in data["findings14"] if f.severity == "FLAG")
    n27_flag = sum(1 for f in data["findings27"] if f.sev == "FLAG")
    print(f"Saved: {OUTPUT_FILE}")
    print(f"  Comparison mismatches : {n_comp_mis}")
    print(f"  Analysis-14 FLAGs     : {n14_flag}")
    print(f"  E-Way-Bill FLAGs      : {n27_flag}")
    print(f"  2B source             : {data['meta']['twob_src']}  "
          f"({'line-level ON' if data['meta']['twob_src']=='excel' else 'PDF summary — line-level OFF'})")




# ============================================================================
# ==== SECTION: build_annual_workbook.py  (was a standalone module before consolidation)
# ============================================================================
"""
PHASE 1 -- ANNUAL / BIFA RECONCILIATION WORKBOOK
==================================================
Builds a standalone workbook from the annual-level sources only (does NOT
need monthly GSTR-1/2B/3B/EWB/E-Inv -- that is Phase 2, once those are
uploaded). Sources used here:
  - Electronic Cash Ledger (CSV)
  - Electronic Credit Ledger (CSV)
  - Electronic Liability Register (CSV)
  - Portal's own "Tax liability and ITC comparison" report (Excel)
  - BO / 360-degree Profile (Excel)

Output: GST_Annual_Reconciliation_FY2022-23.xlsx
"""

import datetime as _dt
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from gst_parsers_dept import (parse_cash_or_liability_ledger, parse_credit_ledger,
                             parse_portal_comparison, MONTH_ABBR)
from gst_parsers_dept import parse_bo_profile

RED = PatternFill("solid", fgColor="FFC7CE")
GREEN = PatternFill("solid", fgColor="C6EFCE")
AMBER = PatternFill("solid", fgColor="FFEB9C")
BLUE = PatternFill("solid", fgColor="DDEBF7")
GREY = PatternFill("solid", fgColor="E7E6E6")
HEAD = PatternFill("solid", fgColor="1F3864")
SECT = PatternFill("solid", fgColor="D9E1F2")
TITLEF = Font(bold=True, size=13, color="1F3864")
BORDER = Border(*[Side(style="thin", color="BFBFBF")] * 4)
SEV_FILL_ANNUAL = {"FLAG": RED, "REVIEW": AMBER, "INFO": BLUE, "OK": GREEN, "N/A": GREY}
SEV_FONT_ANNUAL = {"FLAG": Font(bold=True, color="9C0006"), "REVIEW": Font(bold=True, color="9C6500"),
            "INFO": Font(bold=True, color="2F5496"), "OK": Font(bold=True, color="006100"),
            "N/A": Font(bold=True, color="808080")}

FY_MONTHS = ["Apr-22", "May-22", "Jun-22", "Jul-22", "Aug-22", "Sep-22",
             "Oct-22", "Nov-22", "Dec-22", "Jan-23", "Feb-23", "Mar-23"]
TOL_LAKH = 0.5  # Rs 50,000 -- ledger/portal-comparison reconciliation tolerance (rounding + timing noise)


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def gather_annual(folder="."):
    import os
    j = lambda f: os.path.join(folder, f)
    cash = parse_cash_or_liability_ledger(j("Electronic_Cash_Ledger__2___1_.csv"), "cash")
    credit = parse_credit_ledger(j("ElectronicCreditLedger__2___1_.csv"))
    liab = parse_cash_or_liability_ledger(
        j("Electronic_Liability_Register_Electronic_Liability_Register_2__1_.csv"), "liability")
    comp = parse_portal_comparison(
        j("2022-23_05AAECM6380J1ZA_Tax_liability_and_ITC_comparison__1_.xlsx"))
    bo = parse_bo_profile(j("05AAECM6380J1ZA_BO_Profile_15_06_2026.pdf"))
    return dict(cash=cash, credit=credit, liab=liab, comp=comp, bo=bo)


def build_monthly_rows(data, months=None):
    """Monthly annual-ledger walkthrough. Previously cross-checked the ledgers against the
    GST-Prime TPST 12-month self-filing summary (retired -- no longer supplied); every
    comparison below is now sourced from sources that ARE still supplied: the portal's own
    "Tax liability and ITC comparison" report, the Electronic Liability REGISTER (Part I,
    return-related -- this is what 'liab' has always been fed, see gst_core.classify_folder),
    and the Electronic Credit Ledger. A comparison that genuinely has no second independent
    source left (cash paid, return-filing date) is shown as informational/'N/A', never
    silently dropped or backed by a guessed figure.

    months: the actual 'Mon-YY' labels for the FY being processed (e.g. from
    master_build.py's months_covered) -- REQUIRED for any taxpayer/FY other than the original
    FY2022-23 reference taxpayer. Falls back to the hardcoded FY_MONTHS (2022-23) only for the
    standalone gather_annual()/main_annual_standalone() demo path, which is itself hardcoded
    to that one taxpayer's filenames -- confirmed a real bug for every OTHER taxpayer/FY if
    this fallback is silently relied on: every row comes back empty (no crash, no warning)
    because none of the hardcoded 'Apr-22'..'Mar-23' keys exist in that taxpayer's data."""
    months = months or FY_MONTHS
    cash, credit, liab, comp = data["cash"], data["credit"], data["liab"], data["comp"]
    rows = []
    for mo in months:
        c = comp.get(mo, {})
        cr = credit["monthly_by_tax_period"].get(mo, {})
        ca_period = cash["monthly_by_tax_period"].get(mo, {})   # cash DEBITED against this period
        ca_txndate = cash["monthly_by_txn_date"].get(mo, {})     # cash DEPOSITED in this calendar month
        li = liab["monthly_by_txn_date"].get(mo, {})              # liability booked in this calendar month

        portal_g3b_liab = c.get("gstr3b_liability")
        portal_itc_3b = c.get("itc_3b_adj") if c.get("itc_3b_adj") is not None else c.get("itc_3b_unadj")
        credit_accrued = cr.get("credited")
        cash_debited_period = ca_period.get("debited")
        liab_register_booked = li.get("debited")

        # Liability Register (Part I, return-related booking) vs Portal-comparison's own GSTR-3B
        # liability figure -- two independent government-sourced records of the same fact.
        # NOTE: the portal's "Tax liability and ITC comparison" Comparison-Summary sheet is in
        # absolute Rs (verified: its FY 'Total' row for GSTR-1 liability, 33,479,969.41, equals
        # BIFA's 334.80 Lakh) -- NOT lakhs, so no unit conversion here.
        flag_liab = "N/A"
        if liab_register_booked is not None and portal_g3b_liab is not None:
            d = abs(liab_register_booked - portal_g3b_liab)
            flag_liab = "OK" if d <= TOL_LAKH * 100000 else "REVIEW"

        # Credit Ledger's own accrued ITC vs Portal-comparison's own 3B ITC figure.
        flag_credit_ledger = "N/A"
        if portal_itc_3b is not None and credit_accrued is not None:
            d = abs(portal_itc_3b - credit_accrued)
            flag_credit_ledger = "OK" if d <= TOL_LAKH * 100000 else "REVIEW"

        # Cash paid has no remaining independent second source (TPST was the only one) --
        # shown as informational figures only, never a fabricated OK/REVIEW comparison.
        cash_ratio_pct = (round(100 * cash_debited_period / portal_g3b_liab, 3)
                           if (cash_debited_period is not None and portal_g3b_liab) else None)

        rows.append(dict(
            month=mo,
            liab_register_booked=liab_register_booked, portal_g3b_liability=portal_g3b_liab, flag_liab=flag_liab,
            portal_g1_liability=c.get("gstr1_liability"),
            portal_itc_3b=portal_itc_3b, portal_itc_2b=c.get("itc_2b"),
            credit_ledger_accrued=credit_accrued, flag_credit_ledger=flag_credit_ledger,
            cash_ledger_debited_period=cash_debited_period,
            cash_ledger_deposited_calmonth=ca_txndate.get("credited"),
            liability_register_calmonth=li.get("debited"),
            cash_ratio_pct=cash_ratio_pct,
        ))
    return rows


def write_cover(ws, data, fy_label=None):
    bo = data["bo"]
    ws.cell(1, 1, "GST ANNUAL RECONCILIATION -- PHASE 1 (Ledgers + Portal Comparison + BO Profile)").font = TITLEF
    ws.cell(2, 1, f"GSTIN {bo['self_gstin']}  |  {bo['legal_name']}  |  FY {fy_label or '(unspecified)'}").font = Font(size=10, bold=True)
    ws.cell(3, 1, f"Generated {_dt.datetime.now():%Y-%m-%d %H:%M:%S}").font = Font(size=9, italic=True)
    r = 5
    ws.cell(r, 1, "Sources used").font = Font(bold=True, size=11, color="1F3864"); r += 1
    for s in ["Electronic Cash Ledger (CSV)", "Electronic Credit Ledger (CSV)",
              "Electronic Liability Register (CSV)",
              "Portal Tax-liability & ITC Comparison report (Excel)",
              "BO / 360-degree Profile (Excel)"]:
        ws.cell(r, 2, "- " + s); r += 1
    r += 1
    ws.cell(r, 1, "Known limitations (structural gaps in the source data -- see README for full list)").font = Font(bold=True, size=11, color="9C0006"); r += 1
    for s in ["HSN code is not linked to individual invoices anywhere except the EWB files -- "
              "GSTR-1's own HSN sheet and GSTR-2B's purchase-side data are both monthly aggregates",
              "GSTR-1's B2C-Small (Table 7) sheet is a state+rate summary with no invoice numbers at "
              "all -- invoice-level B2C splitting cannot be detected from this data by design",
              "Credit notes (GSTR-1 'cdnr' sheet) carry no original-invoice-number reference -- any "
              "check linking a CN back to its original sale is approximate (by recipient + value), "
              "never a proven document link",
              "DRC Payment Information is linked to months ONLY by nearby transaction date "
              "(no tax-period field exists in that Excel sheet) -- treat as informational reference, "
              "not a proven link"]:
        ws.cell(r, 2, "- " + s); r += 1
    r += 1
    ws.cell(r, 1, "Known data-quality caveats").font = Font(bold=True, size=11, color="9C6500"); r += 1
    for s in ["Credit-ledger monthly figures are grouped by the ledger's own 'Tax Period' column; "
              "entries with a blank tax-period ('-') are excluded from the monthly view (see raw "
              "transactions for those)"]:
        ws.cell(r, 2, "- " + s); r += 1
    r += 1
    ws.cell(r, 1, "Departmental proceedings (from BO Profile Appeal/Case/Transfer sections)").font = Font(bold=True, size=11, color="9C0006"); r += 1
    appeals = bo.get("appeals", [])
    cases = bo.get("cases", [])
    transfers = bo.get("transfers", [])
    if not (appeals or cases or transfers):
        ws.cell(r, 2, "- None found in the BO Profile's Appeal/Case/Transfer Information sections."); r += 1
    for a in appeals:
        ws.cell(r, 2, f"- APPEAL: ARN {a.get('arn')} filed {a.get('filing_date')} (FY {a.get('fy')}) -- "
                       f"{a.get('case_type')}, status {a.get('status')}"); r += 1
    for c in cases:
        ws.cell(r, 2, f"- CASE: {c.get('case_id')} / ref {c.get('reference_id')}, "
                       f"action date {c.get('action_date')} -- {c.get('case_type')}, status {c.get('status')}"); r += 1
    for t in transfers:
        ws.cell(r, 2, f"- TRANSFER: {t.get('source_case_id')} on {t.get('date')} -- "
                       f"{t.get('source_module')} -> {t.get('target_module')}, status {t.get('status')}"); r += 1
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 110


def write_monthly(ws, rows, fy_label=None):
    ws.cell(1, 1, f"MONTHLY WALKTHROUGH -- Liability Register vs Portal-Comparison vs Ledgers "
                  f"(FY {fy_label or '(unspecified)'})").font = TITLEF
    ws.cell(2, 1, "Amounts in Rs unless stated. Tolerance: Rs 50,000 (liability/ITC). Cash paid has no "
                  "independent second source in this run's inputs -- shown informationally, not cross-checked.").font = Font(size=9, italic=True)
    hdr = ["Month",
           "Liability Register: booked", "Liability: Portal(3B)", "Liab Check", "Liability: Portal(GSTR-1)",
           "ITC: Portal(3B)", "ITC: Portal(2B)",
           "Credit Ledger: ITC accrued (this period)", "Credit-Ledger Check",
           "Cash Ledger: debited (this period)",
           "Cash Ledger: deposited (this cal.month)", "Liability Register: booked (this cal.month)",
           "Cash / Liability(3B) %"]
    r = 4
    for i, h in enumerate(hdr, 1):
        ws.cell(r, i, h)
    _style_header(ws, r, len(hdr))
    r += 1
    for row in rows:
        vals = [row["month"],
                row["liab_register_booked"], row["portal_g3b_liability"], row["flag_liab"],
                row["portal_g1_liability"],
                row["portal_itc_3b"], row["portal_itc_2b"],
                row["credit_ledger_accrued"], row["flag_credit_ledger"],
                row["cash_ledger_debited_period"],
                row["cash_ledger_deposited_calmonth"], row["liability_register_calmonth"],
                row["cash_ratio_pct"]]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(r, ci, v)
            cell.border = BORDER
            cell.font = Font(size=10)
            if isinstance(v, (int, float)):
                cell.number_format = "#,##0.00"
            if hdr[ci - 1].endswith("Check"):
                cell.fill = SEV_FILL_ANNUAL.get(v, GREY)
                cell.font = SEV_FONT_ANNUAL.get(v, Font(size=10))
                cell.alignment = Alignment(horizontal="center")
            if hdr[ci - 1] == "Cash / Liability(3B) %" and isinstance(v, (int, float)) and v < 2:
                cell.fill = AMBER
        r += 1
    for i, w in enumerate([9, 16, 14, 11, 15, 13, 13, 16, 13, 14, 16, 16, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"


def write_fy_total_vs_bifa(ws, rows, data, ewb_out_annual_rows=None, fy=None):
    bo = data["bo"]
    bifa_by_fy = bo.get("bifa_by_fy", {})
    # FIX (genericity -- was hardcoded to "2022-23", so a different taxpayer/FY would
    # silently read an empty {} and show every BIFA figure as 0, indistinguishable from a
    # real mismatch). Now: use the FY explicitly passed in; if not given, and there's
    # exactly one FY in bifa_by_fy, use that; otherwise leave bifa empty AND say so on the
    # sheet, rather than guessing a key that might not exist.
    fy_used = fy
    if not fy_used:
        if len(bifa_by_fy) == 1:
            fy_used = next(iter(bifa_by_fy))
        elif bifa_by_fy:
            fy_used = sorted(bifa_by_fy)[-1]  # most recent, if multiple and none specified
    bifa = bifa_by_fy.get(fy_used, {}) if fy_used else {}
    ws.cell(1, 1, f"FY-TOTAL vs BIFA (DEPARTMENT'S OWN CROSS-CHECK) -- FY {fy_used or '(unknown)'}").font = TITLEF
    ws.cell(2, 1, "BIFA = the department's own pre-computed FY-level figures from the BO Profile. "
                  "'Recomputed' = summed from the Liability Register/Credit Ledger/EWB by this tool, independently."
                  + ("" if bifa else f"  **BIFA figures below are all blank/zero because FY "
                     f"{fy_used!r} was not found in the BO Profile's BIFA table -- FYs actually "
                     f"present there: {sorted(bifa_by_fy) or 'none'}. This is a data-availability "
                     "note, not a real mismatch; do not read the REVIEW flags below at face value "
                     "until this is resolved.")).font = Font(size=9, italic=True,
                     color=("C00000" if not bifa else "000000"))

    # sum_liab previously came from the TPST source (retired); now sourced from the Liability
    # Register (Part I, return-related) -- the same figure Portal-Comparison is cross-checked
    # against per month on the Annual Ledger Walkthrough sheet.
    # BUG FIX (caught while verifying this session's Diff% column): sum_liab collapsed a
    # genuinely-not-supplied Liability Register into 0 (via `or 0`), which then compared against
    # BIFA's real figure as if 0 were an actual recomputed result -- producing a nonsensical
    # '-100% diff, REVIEW' flag instead of an honest 'not supplied' entry. Confirmed on a real
    # run: this taxpayer's Liability Ledger genuinely wasn't supplied (the pipeline's own log
    # already says so for the per-month ledger checks), yet this FY-total row showed a fake
    # mismatch. Same guard pattern the 'Cash paid' row below already uses, now applied here too.
    liab_data_available = any(r["liab_register_booked"] is not None for r in rows)
    sum_liab = sum(r["liab_register_booked"] or 0 for r in rows) if liab_data_available else None
    sum_credit_accrued = sum(r["credit_ledger_accrued"] or 0 for r in rows)
    # Cash paid has no remaining independent second source in this run's inputs (TPST was the
    # only one) -- shown as an informational recomputed figure only, never cross-checked.
    cash_data_available = any(r["cash_ledger_debited_period"] is not None for r in rows)
    sum_cash = sum(r["cash_ledger_debited_period"] or 0 for r in rows) if cash_data_available else None
    # B1: recompute annual Outward-EWB "tax value" total directly from the raw EWB file,
    # independent of BIFA's own pre-computed figure -- new addition (was previously unused).
    sum_ewb_tax = sum(e["taxval"] for e in ewb_out_annual_rows) if ewb_out_annual_rows else None

    def _bifa_val(key):
        """None (not 0) when bifa is empty -- a real 0 in the BO Profile is a legitimate
        comparator; an EMPTY bifa dict (FY not found at all) must not silently look like one."""
        if not bifa:
            return None
        return (bifa.get(key) or 0) * 100000

    lines = [
        ("Liability as per GSTR-3B (Rs)", sum_liab, _bifa_val("liability_gstr3b")),
        ("Credit Ledger accrued vs BIFA 'ITC Availed in R3B' (Rs)", sum_credit_accrued,
         _bifa_val("itc_r3b")),
        ("Credit Ledger accrued vs BIFA 'ITC Accrued in R2B/R2A' (Rs)", sum_credit_accrued,
         _bifa_val("itc_r2b_r2a")),
        ("Outward EWB Tax Val. (recomputed) vs BIFA 'Liability as per EWB' (Rs)",
         sum_ewb_tax, _bifa_val("liability_ewb")),
        ("Cash paid (Rs) -- informational only, no independent second source in this run's inputs"
         if sum_cash is None else "Cash paid (Rs) -- informational only, no independent second source in this run's inputs",
         sum_cash, None),
    ]
    ws.cell(3, 1, "NOTE (A5): BIFA carries two distinct ITC columns -- 'ITC Availed in R3B' and 'ITC "
                  "Accrued in R2B/R2A' -- which differ by the department's own 'Excess ITC claimed' "
                  "figure. The Credit Ledger accrual is shown against BOTH below rather than picking "
                  "one implicitly; which is the more meaningful comparator is a judgement call for the "
                  "reviewer, not something this tool should decide silently.").font = Font(size=9, italic=True, color="9C6500")
    hdr = ["Metric", "Recomputed (Register/Ledger sum)", "BIFA (dept figure, converted from Lakh)",
           "Diff", "Diff %", "Check"]
    r = 5
    for i, h in enumerate(hdr, 1):
        ws.cell(r, i, h)
    _style_header(ws, r, len(hdr))
    r += 1
    for label, a, b in lines:
        ws.cell(r, 1, label)
        ws.cell(r, 2, round(a, 2) if a is not None else "n/a")
        ws.cell(r, 3, round(b, 2) if b is not None else "n/a")
        if a is not None and b is not None:
            d = round(a - b, 2)
            ws.cell(r, 4, d)
            ws.cell(r, 5, round(d / b * 100.0, 2) if b else "n/a")
            chk = "OK" if abs(d) <= TOL_LAKH * 100000 else "REVIEW"
            c = ws.cell(r, 6, chk); c.fill = SEV_FILL_ANNUAL[chk]; c.font = SEV_FONT_ANNUAL[chk]
            c.alignment = Alignment(horizontal="center")
            if chk == "REVIEW":
                for cc in (2, 3, 4):
                    ws.cell(r, cc).fill = RED
        else:
            ws.cell(r, 4, "n/a"); ws.cell(r, 5, "n/a"); ws.cell(r, 6, "N/A")
        for c in range(1, len(hdr) + 1):
            ws.cell(r, c).border = BORDER
        r += 1
    r += 1
    ws.cell(r, 1, "SUGGESTED FURTHER ANALYSIS from raw data already in this tool's inputs, not yet "
                  "built:").font = Font(bold=True, size=10, color="1F3864")
    r += 1
    for s in ["A month-by-month version of this same comparison (currently FY-total only) -- the "
              "'Annual Ledger Walkthrough' sheet already has the monthly Liability/Credit figures, "
              "and BIFA's own table is FY-level only so a monthly BIFA comparator isn't possible, "
              "but a monthly RECOMPUTED trend against the flat FY BIFA line would still show "
              "whether the gap is concentrated in specific months or spread evenly.",
              "A 'days since BIFA generation' note -- BO Profile is generated at a point in time; "
              "if late-filed returns or amendments landed after that date, a gap here can be "
              "explained by pure timing rather than a genuine discrepancy. This tool doesn't "
              "currently capture the BO Profile's own generation date to test that.",
              "Cross-FY trend of this same Diff %, once more than one year's output from this "
              "tool exists to compare -- a persistent, growing gap is a different signal than a "
              "one-off."]:
        ws.cell(r, 1, "- " + s).font = Font(size=9, italic=True)
        ws.cell(r, 1).alignment = Alignment(wrap_text=True)
        r += 1
    for i, w in enumerate([48, 24, 30, 14, 12, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_drc(ws, data):
    drc = data["bo"]["drc_payments"]
    ws.cell(1, 1, "DRC PAYMENT INFORMATION (informational reference -- BO Profile PDF)").font = TITLEF
    ws.cell(2, 1, "No tax-period field exists in this PDF section -- linkage to a specific GST month is "
                  "by nearby transaction date ONLY. Treat as a lead to investigate, not a proven match.").font = Font(size=9, italic=True, color="C00000")
    hdr = ["Source ID", "Description", "Date", "Method", "CGST", "SGST", "IGST", "CESS", "Other", "Total"]
    r = 4
    for i, h in enumerate(hdr, 1):
        ws.cell(r, i, h)
    _style_header(ws, r, len(hdr))
    r += 1
    for d in sorted(drc, key=lambda x: x["date"]):
        vals = [d["source_id"], d["description"], d["date"], d["method"],
                d["cgst"], d["sgst"], d["igst"], d["cess"], d["other"], d["total"]]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(r, ci, v)
            cell.border = BORDER
            cell.font = Font(size=10)
            cell.alignment = Alignment(wrap_text=(ci == 2))
            if isinstance(v, float):
                cell.number_format = "#,##0.00"
        r += 1
    for i, w in enumerate([18, 40, 12, 14, 10, 10, 10, 10, 10, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"


def write_related_party(ws, data):
    bo = data["bo"]
    ws.cell(1, 1, "RELATED / CANCELLED-PARTY ITC ALERTS (BO Profile)").font = TITLEF
    ws.cell(2, 1, "Fraud-risk indicator: ITC exchanged with a party sharing a related parameter "
                  "(mobile/PAN/etc.) with this taxpayer, or with a since-cancelled GSTIN.").font = Font(size=9, italic=True)
    r = 4
    for label, key in [("ITC RECEIVED from related/cancelled supplier", "related_itc_received"),
                        ("ITC PASSED ON to related/cancelled recipient", "related_itc_passed")]:
        ws.cell(r, 1, label).font = Font(bold=True, size=11, color="1F3864")
        for c in range(1, 9):
            ws.cell(r, c).fill = SECT
        r += 1
        hdr = ["FY", "GSTIN", "Name", "Related Param", "Status", "Cancellation Date", "Reason", "Total ITC (Lakh)"]
        for i, h in enumerate(hdr, 1):
            ws.cell(r, i, h)
        _style_header(ws, r, 8)
        r += 1
        section_start = r
        for x in bo[key]:
            vals = [x["fy"], x["gstin"], x["name"], x["related_parameter"], x["status"],
                    x["cancellation_date"], x["reason"], x["total_itc"]]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(r, ci, v)
                cell.border = BORDER
                cell.font = Font(size=10)
                if isinstance(v, float):
                    cell.number_format = "#,##0.00"
                if x["status"] == "Cancelled":
                    cell.fill = RED
            r += 1
        if not bo[key]:
            ws.cell(r, 1, "No rows -- either genuinely none in the source BO Profile, or this "
                          "section's marker text wasn't found during parsing (see "
                          "bo_profile_parser.py --diagnose for which). Do not read this as a "
                          "confirmed 'no related-party ITC' result without checking.").font = Font(
                          italic=True, color="9C6500")
            r += 1
        else:
            # NEW: total row -- sum of the amount column for this section, so the aggregate
            # exposure is visible without adding the column up by hand.
            total = round(sum(num(x.get("total_itc")) for x in bo[key]), 2)
            ws.cell(r, 7, "TOTAL").font = Font(bold=True)
            tc = ws.cell(r, 8, total); tc.font = Font(bold=True); tc.number_format = "#,##0.00"
            for c in range(1, 9):
                ws.cell(r, c).border = BORDER
            r += 1
        r += 1

    # RELOCATED from the removed 'Annual Cover & Caveats' sheet (per explicit instruction to
    # remove that sheet) -- this is the only genuinely unique DATA it carried (the rest was
    # static methodology/source-list text); everything else from that sheet is not reproduced
    # here since it wasn't taxpayer-specific data. Same BO Profile source as the rest of this
    # sheet, so it belongs here rather than nowhere.
    appeals = bo.get("appeals", [])
    cases = bo.get("cases", [])
    transfers = bo.get("transfers", [])
    ws.cell(r, 1, "Departmental Proceedings (BO Profile Appeal / Case / Transfer sections)").font = \
        Font(bold=True, size=11, color="9C0006")
    for c in range(1, 9):
        ws.cell(r, c).fill = SECT
    r += 1
    if not (appeals or cases or transfers):
        ws.cell(r, 1, "None found in the BO Profile's Appeal/Case/Transfer Information sections."
                      ).font = Font(italic=True)
        r += 1
    for a in appeals:
        ws.cell(r, 1, f"APPEAL: ARN {a.get('arn')} filed {a.get('filing_date')} (FY {a.get('fy')}) -- "
                       f"{a.get('case_type')}, status {a.get('status')}").font = Font(size=10)
        r += 1
    for c_ in cases:
        ws.cell(r, 1, f"CASE: {c_.get('case_id')} / ref {c_.get('reference_id')}, "
                       f"action date {c_.get('action_date')} -- {c_.get('case_type')}, status {c_.get('status')}"
                      ).font = Font(size=10)
        r += 1
    for t in transfers:
        ws.cell(r, 1, f"TRANSFER: {t.get('source_case_id')} on {t.get('date')} -- "
                       f"{t.get('source_module')} -> {t.get('target_module')}, status {t.get('status')}"
                      ).font = Font(size=10)
        r += 1
    for i, w in enumerate([9, 18, 26, 14, 11, 16, 40, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_top_counterparties(ws, data):
    bo = data["bo"]
    ws.cell(1, 1, "TOP-10 COUNTERPARTIES (context -- BO Profile, trailing 12 months)").font = TITLEF
    r = 3
    for label, key, amtlabel in [("Top 10 Beneficiaries (ITC Passed On)", "top_beneficiaries", "ITC Passed (Lakh)"),
                                   ("Top 10 Suppliers (ITC Received)", "top_suppliers", "ITC Received (Lakh)")]:
        ws.cell(r, 1, label).font = Font(bold=True, size=11, color="1F3864")
        for c in range(1, 7):
            ws.cell(r, c).fill = SECT
        r += 1
        hdr = ["GSTIN", "Name", "Reg. Start", "Status", "Risk", amtlabel]
        for i, h in enumerate(hdr, 1):
            ws.cell(r, i, h)
        _style_header(ws, r, 6)
        r += 1
        for x in bo[key]:
            vals = [x["gstin"], x["name"], x["reg_start"], x["status"], x["risk"], x["amount"]]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(r, ci, v)
                cell.border = BORDER
                cell.font = Font(size=10)
                if x["status"] == "Cancelled":
                    cell.fill = RED
            r += 1
        if not bo[key]:
            ws.cell(r, 1, "No rows -- either genuinely none in the source BO Profile, or this "
                          "section's marker text wasn't found during parsing (see "
                          "bo_profile_parser.py --diagnose for which).").font = Font(italic=True, color="9C6500")
            r += 1
        r += 1
    for i, w in enumerate([18, 30, 12, 11, 16, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main_annual_standalone(folder=".", outfile="GST_Annual_Reconciliation_FY2022-23.xlsx"):
    data = gather_annual(folder)
    rows = build_monthly_rows(data)

    wb = openpyxl.Workbook()
    write_cover(wb.active, data); wb.active.title = "Cover & Caveats"
    write_monthly(wb.create_sheet("Monthly Walkthrough"), rows)
    write_fy_total_vs_bifa(wb.create_sheet("FY Total vs BIFA"), rows, data)
    write_drc(wb.create_sheet("DRC Payments (info)"), data)
    write_related_party(wb.create_sheet("Related-Party Alerts"), data)
    write_top_counterparties(wb.create_sheet("Top Counterparties"), data)
    wb.save(outfile)

    n_review = sum(1 for r in rows for k in ("flag_liab", "flag_itc", "flag_credit_ledger", "flag_cash_ledger")
                    if r[k] == "REVIEW")
    avg_cash_ratio = sum(r["cash_ratio_pct"] or 0 for r in rows) / len(rows)
    print(f"Saved: {outfile}")
    print(f"  Monthly REVIEW flags: {n_review}")
    print(f"  Avg cash-utilization ratio across FY: {avg_cash_ratio:.2f}%")


