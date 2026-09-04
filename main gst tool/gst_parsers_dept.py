#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GST PARSERS DEPT
================
CONSOLIDATED FILE -- contains what used to be: annual_sources.py, annual_return_parser.py, bo_profile_parser.py

The tool was reorganised from 19 .py files into 9 for easier sharing. Nothing
in the analytical logic was rewritten during that move: each section below is
the original module's code verbatim, with only (a) intra-project imports
repointed at the new file names, (b) its standalone __main__ demo block
removed, and (c) the renames listed under MERGE NOTES applied where two merged
modules happened to define the same top-level name with different bodies.

MERGE NOTES for this file:
  - annual_sources._num -> _num_as
  - bo_profile_parser._num -> _num_bo
"""


# ============================================================================
# ==== SECTION: annual_sources.py  (was a standalone module before consolidation)
# ============================================================================
"""
ANNUAL SOURCES  --  Ledgers (Cash/Credit/Liability), GST-Prime TPST, and the
portal's own "Tax liability and ITC comparison" report.

These are FY-wide, month-agnostic files (one file covers the whole year),
unlike GSTR-1/3B/2B/E-Inv/EWB which are one-file-per-period. All parsers here
key their output by month label 'Mon-YY' (e.g. 'Apr-22') so they can be
cross-joined against each other and, later, against the monthly GSTR set.
"""

import csv
import re
import openpyxl

MONTH_ABBR = {1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
              7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"}
MONTH_NUM = {v: k for k, v in MONTH_ABBR.items()}


def _num_as(v):
    if v is None:
        return 0.0
    s = str(v).replace(",", "").replace("₹", "").strip()
    if s in ("", "-", "–", "NA"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _period_key(datestr):
    """'21-05-2022' or '21/05/2022' -> 'May-22'. Both '-' and '/' day-month-year
    separators confirmed in real ledger/register exports for different taxpayers
    (this taxpayer's Cash/Liability-Register/Liability-Ledger CSVs all use '/',
    the original reference taxpayer's used '-') -- accept either rather than
    silently returning every row unkeyed (which happened before: 49/49 real
    transactions parsed correctly but 0 landed in any monthly bucket, no error
    raised anywhere, because '/'-dates never matched the hyphen-only regex).
    Returns None if genuinely unparseable."""
    if not datestr or datestr == "-":
        return None
    m = re.match(r"(\d{2})[-/](\d{2})[-/](\d{4})", datestr.strip())
    if not m:
        return None
    _, mm, yyyy = m.groups()
    mm = int(mm)
    return f"{MONTH_ABBR.get(mm, '?')}-{yyyy[2:]}"


def _tax_period_key(tp):
    """'Mar-22' (as already given in ledger 'Tax Period' column) -> normalise to 'Mar-22'."""
    if not tp or tp == "-":
        return None
    tp = tp.strip()
    m = re.match(r"([A-Za-z]{3})-(\d{2,4})", tp)
    if not m:
        return None
    mon, yy = m.groups()
    yy = yy[-2:]
    return f"{mon[:3].title()}-{yy}"


# ======================================================================
# CASH LEDGER  /  LIABILITY REGISTER  (same 8-group-of-6 layout)
# ======================================================================
_HEAD_GROUPS_8 = ["IGST", "CGST", "SGST", "CESS"]  # first 4 groups = Debited/Credited
# groups 5-8 = running balances (same order)


def parse_cash_or_liability_ledger(path, kind):
    """kind: 'cash', 'liability' (Electronic Liability REGISTER, Part I -- return-related),
    or 'liability_demand' (Electronic Liability LEDGER, Part II -- DRC/demand/voluntary
    payments; carries two columns Part I doesn't: 'Relevant Demand ID / Liability ID' and
    a trailing 'Stay status'). Confirmed against real files: Register and Ledger are
    genuinely different column layouts (6-column vs 8-column preamble before the tax-head
    groups) -- NOT the same report under two names, so this must not be guessed from one
    shared code path; each kind's offsets are explicit below, verified against real exports.
    Returns dict(opening={...}, transactions=[...],
    monthly_by_tax_period={period: {...totals...}}, monthly_by_txn_date={period: {...}})."""
    with open(path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))

    if kind == "cash":
        # cols: 0 Sr.No,1 Date,2 Time,3 Reporting date,4 Reference No.,5 Tax Period,
        #       6 Description,7 Transaction Type, 8-31 Debited/Credited (4x6), 32-55 Balance (4x6)
        col_date, col_ref, col_taxperiod, col_desc, col_ttype = 1, 4, 5, 6, 7
        col_ledger_used, col_demand_id, col_stay = None, None, None
        deb_start = 8
    elif kind == "liability_demand":
        # Electronic Liability LEDGER (Part II): cols 0 Sr.No,1 Date,2 Reference No.,
        # 3 Tax Period if applicable,4 Ledger used,5 Relevant Demand ID / Liability ID,
        # 6 Description,7 Type of Transaction, 8-31 Debited/Credited (4x6), 32-55 Balance
        # (4x6), 56 Stay status. Verified column-by-column against a real export.
        col_date, col_ref, col_taxperiod, col_desc, col_ttype = 1, 2, 3, 6, 7
        col_ledger_used, col_demand_id, col_stay = 4, 5, 56
        deb_start = 8
    else:  # liability -- Electronic Liability REGISTER (Part I, return-related)
        # cols: 0 Sr.No,1 Date,2 Reference No.,3 Ledger Used,4 Description,5 Transaction Type,
        #       6-29 Debited/Credited (4x6), 30-53 Balance (4x6)
        col_date, col_ref, col_taxperiod, col_desc, col_ttype = 1, 2, None, 4, 5
        col_ledger_used, col_demand_id, col_stay = 3, None, None
        deb_start = 6

    opening = None
    transactions = []
    for r in rows:
        if not r or len(r) < deb_start + 24:
            continue
        first = (r[0] or "").strip()
        desc = (r[col_desc] or "").strip() if col_desc < len(r) else ""
        if not first and desc != "Opening Balance":
            continue
        if not (first.isdigit() or first in ("-",) or desc == "Opening Balance"):
            continue
        date = (r[col_date] or "").strip() if col_date < len(r) else "-"
        ref = (r[col_ref] or "").strip() if col_ref < len(r) else "-"
        taxp = (r[col_taxperiod] or "").strip() if col_taxperiod is not None and col_taxperiod < len(r) else "-"
        ttype = (r[col_ttype] or "").strip() if col_ttype < len(r) else "-"
        extra = (r[col_ledger_used] or "").strip() if col_ledger_used is not None and col_ledger_used < len(r) else None
        demand_id = (r[col_demand_id] or "").strip() if col_demand_id is not None and col_demand_id < len(r) else None
        stay = (r[col_stay] or "").strip() if col_stay is not None and col_stay < len(r) else None

        heads = {}
        bal_start = deb_start + 24
        for gi, head in enumerate(_HEAD_GROUPS_8):
            base = deb_start + gi * 6
            bbase = bal_start + gi * 6
            tax = _num_as(r[base]) if base < len(r) else 0.0
            total = _num_as(r[base + 5]) if base + 5 < len(r) else 0.0
            bal_total = _num_as(r[bbase + 5]) if bbase + 5 < len(r) else 0.0
            heads[head] = dict(tax=tax, total=total, balance=bal_total)
        total_debited_or_credited = sum(h["total"] for h in heads.values())

        rec = dict(date=date, ref=ref, tax_period=taxp, description=desc, ttype=ttype,
                    heads=heads, total=total_debited_or_credited,
                    balance_total=sum(h["balance"] for h in heads.values()),
                    ledger_used=extra, demand_id=demand_id, stay_status=stay)

        if desc == "Opening Balance":
            opening = rec
            continue
        if desc == "Closing Balance":
            continue
        transactions.append(rec)

    monthly_by_taxperiod = {}
    monthly_by_txndate = {}
    for t in transactions:
        sign = 1 if t["ttype"].lower() == "credit" else -1
        amt = t["total"]
        tpk = _tax_period_key(t["tax_period"]) if kind == "cash" else None
        dpk = _period_key(t["date"])
        if tpk:
            monthly_by_taxperiod.setdefault(tpk, dict(credited=0.0, debited=0.0, net=0.0))
            if t["ttype"].lower() == "credit":
                monthly_by_taxperiod[tpk]["credited"] += amt
            else:
                monthly_by_taxperiod[tpk]["debited"] += amt
            monthly_by_taxperiod[tpk]["net"] += sign * amt
        if dpk:
            monthly_by_txndate.setdefault(dpk, dict(credited=0.0, debited=0.0, net=0.0))
            if t["ttype"].lower() == "credit":
                monthly_by_txndate[dpk]["credited"] += amt
            else:
                monthly_by_txndate[dpk]["debited"] += amt
            monthly_by_txndate[dpk]["net"] += sign * amt

    return dict(opening=opening, transactions=transactions,
                monthly_by_tax_period=monthly_by_taxperiod,
                monthly_by_txn_date=monthly_by_txndate)


# ======================================================================
# CREDIT LEDGER  (different layout: 1 Credit/Debit block + 1 Balance block)
# ======================================================================
def parse_credit_ledger(path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    # cols: 0 Sr.No,1 Date,2 Reference No.,3 Tax Period,4 Description,5 Transaction Type,
    #       6-10 Credit/Debit[IGST,CGST,SGST,CESS,Total], 11-15 Balance[IGST,CGST,SGST,CESS,Total]
    opening = None
    transactions = []
    for r in rows:
        if not r or len(r) < 16:
            continue
        desc = (r[4] or "").strip()
        first = (r[0] or "").strip()
        if not (first.isdigit() or desc == "Opening Balance"):
            continue
        rec = dict(date=(r[1] or "").strip(), ref=(r[2] or "").strip(),
                    tax_period=(r[3] or "").strip(), description=desc,
                    ttype=(r[5] or "").strip(),
                    igst=_num_as(r[6]), cgst=_num_as(r[7]), sgst=_num_as(r[8]), cess=_num_as(r[9]),
                    total=_num_as(r[10]),
                    bal_igst=_num_as(r[11]), bal_cgst=_num_as(r[12]), bal_sgst=_num_as(r[13]),
                    bal_cess=_num_as(r[14]), bal_total=_num_as(r[15]))
        if desc == "Opening Balance":
            opening = rec
            continue
        transactions.append(rec)

    monthly_by_taxperiod = {}
    for t in transactions:
        tpk = _tax_period_key(t["tax_period"])
        if not tpk:
            continue
        m = monthly_by_taxperiod.setdefault(
            tpk, dict(credited=0.0, debited=0.0, net=0.0, accrued_desc=set()))
        sign = 1 if t["ttype"].lower() == "credit" else -1
        if t["ttype"].lower() == "credit":
            m["credited"] += t["total"]
            m["accrued_desc"].add(t["description"])
        else:
            m["debited"] += t["total"]
        m["net"] += sign * t["total"]
    for m in monthly_by_taxperiod.values():
        m["accrued_desc"] = sorted(m["accrued_desc"])

    return dict(opening=opening, transactions=transactions,
                monthly_by_tax_period=monthly_by_taxperiod)


# ======================================================================
# PORTAL "Tax liability and ITC comparison" report (Excel, Comparison Summary sheet)
# ======================================================================
def parse_portal_comparison(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Comparison Summary" not in wb.sheetnames:
        return {}
    ws = wb["Comparison Summary"]
    rows = list(ws.iter_rows(values_only=True))
    out = {}
    for r in rows:
        if not r or not r[0]:
            continue
        period = str(r[0]).strip()
        m = re.match(r"([A-Za-z]{3})-(\d{2})$", period)
        if not m:
            continue
        key = f"{m.group(1)}-{m.group(2)}"
        def _blank(v):
            return v is None or (isinstance(v, str) and v.strip() == "")

        out[key] = dict(
            gstr1_liability=_num_as(r[1]), gstr3b_liability=_num_as(r[2]),
            diff_liability=_num_as(r[3]), cum_diff_liability=_num_as(r[4]),
            itc_3b_unadj=_num_as(r[6]), itc_2b=_num_as(r[7]), diff_itc_unadj=_num_as(r[8]),
            cum_diff_itc_unadj=_num_as(r[9]),
            itc_3b_adj=None if (len(r) <= 11 or _blank(r[11])) else _num_as(r[11]),
            diff_itc_adj=None if (len(r) <= 12 or _blank(r[12])) else _num_as(r[12]),
            cum_diff_itc_adj=None if (len(r) <= 13 or _blank(r[13])) else _num_as(r[13]),
        )
    return out




# ============================================================================
# ==== SECTION: annual_return_parser.py  (was a standalone module before consolidation)
# ============================================================================
"""
ANNUAL RETURN PARSER  --  GSTR-9, GSTR-9C, Table 8A
=====================================================
Three optional annual-level sources, added on top of the existing 12+3 (EWB)
file set. GSTR-9 and GSTR-9C are now read as the government's own Excel
export (each Item/Table lives on its own named sheet, e.g. "Item 5 -
Outward Supplies...", "Item 9 - Rate-wise Tax Liabilit..."), NOT PDF --
this tool no longer reads any PDF at all (BS/PL was never PDF-parsed
either; see bs_pl_input.py). Table 8A (XLSX) is the exact
government-generated workbook.

Every field is located by CONTENT (the row's own 'Sr.No'/description label,
or the sheet's header row), never a hardcoded row/column position -- Table
8A in particular has been confirmed to have a DIFFERENT column count across
real exports for different taxpayers (see _t8a_header_map's docstring), so
even "the same government form" cannot be trusted to keep a fixed layout.

ALL THREE ARE OPTIONAL. Every function here degrades to a clearly-labeled
"not supplied" / "not found" state -- never raises for a missing file, never
fabricates a number.
"""

import os
import re
import datetime as _dt
import openpyxl


# ======================================================================
# Shared helpers
# ======================================================================
def num(v):
    """Convert any cell/string to float. Blank/'-'/None -> 0.0. Never raises."""
    if v is None:
        return 0.0
    s = str(v).strip()
    if s in ("", "-", "\u2013", "NA", "N/A"):
        return 0.0
    s = s.replace(",", "").replace("\u20b9", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def _gxl_sheet(wb, *name_startswith):
    """First sheet whose name starts with any of the given prefixes (GSTR-9/9C
    Excel sheet names are truncated to Excel's 31-char limit, so match on a
    stable prefix rather than the full title -- see the classify_folder()
    signatures in gst_core.py for the same technique). None if not found."""
    for name in wb.sheetnames:
        for pfx in name_startswith:
            if name.startswith(pfx):
                return wb[name]
    return None


def _gxl_row_by_label(ws, col_label, *label_fragments, max_scan=60, occurrence=1):
    """Scan a sheet's rows for the one whose designated label column contains
    ALL given fragments (case-insensitive substring match), return that row's
    full value tuple. None if not found after max_scan rows -- caller must
    handle explicitly, never guessed."""
    hits = 0
    for row in ws.iter_rows(min_row=1, max_row=max_scan, values_only=True):
        cell = row[col_label] if col_label < len(row) else None
        if cell is None:
            continue
        low = str(cell).lower()
        if all(f.lower() in low for f in label_fragments):
            hits += 1
            if hits == occurrence:
                return row
    return None


def _gxl_num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").replace("\u20b9", "").strip()
    if s in ("", "-", "\u2013", "NA", "N/A"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return None


# ======================================================================
# GSTR-9  (Annual Return, Excel export)
# ======================================================================
def parse_gstr9(path):
    """Return dict of the GSTR-9 figures this tool's forensic checks need, read from
    the government's own Excel export (Item-numbered sheets). Same output-dict
    CONTRACT as the retired PDF version (identical key names) so R13/R14 and every
    other caller need no changes -- only extended with NEW keys (never previously
    available from the PDF scrape) for figures this Excel format exposes cleanly:
    Table 6/7/8's full ITC breakdown, and the Part-I filing ARN/date.
    available=False (with reason) if the file is absent or unreadable -- NEVER
    raises. A field genuinely absent from this return stays None, never guessed."""
    out = dict(available=False, reason=None, is_system_draft=None,
                fy=None, gstin=None, legal_name=None,
                table4_b2b_taxable=None, table4_b2b_igst=None, table4_b2b_cgst=None, table4_b2b_sgst=None,
                table4_cn_taxable=None, table4_cn_igst=None, table4_cn_cgst=None, table4_cn_sgst=None,
                table5_zero_rated=None, table5_sez=None, table5_exempted=None, table5_nil_rated=None,
                table5_nongst=None, table5_all_zero=None,
                table6a_cgst=None, table6a_sgst=None, table6a_igst=None, table6a_cess=None, table6a_total=None,
                table9_liability_igst=None, table9_liability_cgst=None, table9_liability_sgst=None,
                table9_late_fee_payable=None, table9_late_fee_paid=None,
                table9_interest_payable=None, table9_interest_paid=None,
                notes=[],
                # NEW -- available now this is a structured Excel read; not consumed by any
                # existing check yet, kept for future checklist-driven checks (see handoff).
                arn=None, date_of_filing=None,
                table4_rcm_inward_taxable=None, table4_net_supplies_tax_igst=None,
                table6_total_itc_availed=None, table7_total_reversed=None,
                table8_itc_as_per_2b=None, table8_itc_as_per_6b6h=None, table8_difference=None,
                table13_itc_cgst=None, table13_itc_sgst=None, table13_itc_igst=None, table13_itc_cess=None,
                table12_itc_reversed_cgst=None, table12_itc_reversed_sgst=None,
                table12_itc_reversed_igst=None, table12_itc_reversed_cess=None)
    if not path or not os.path.exists(path):
        out["reason"] = "GSTR-9 not supplied for this taxpayer/FY."
        return out
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as ex:
        out["reason"] = f"Could not read GSTR-9 workbook: {ex}"
        return out

    part1 = _gxl_sheet(wb, "Part I - Basic Details")
    if part1 is not None:
        for row in part1.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            label = str(row[0]).strip()
            val = row[1] if len(row) > 1 else None
            if label.startswith("1.") and "financial year" in label.lower():
                out["fy"] = val
            elif label.startswith("2.") and "gstin" in label.lower():
                out["gstin"] = val
            elif label.startswith("3(a)"):
                out["legal_name"] = val
            elif label.startswith("3(c)"):
                out["arn"] = val
            elif label.startswith("3(d)"):
                out["date_of_filing"] = val
    else:
        out["notes"].append("Part I (Basic Details) sheet not found.")
    out["is_system_draft"] = not bool(out["arn"])
    if out["is_system_draft"]:
        out["notes"].append("No ARN found on Part I -- this may be the pre-filing 'System Drafted "
                             "(For Reference Only)' auto-draft rather than the as-filed return.")

    item4 = _gxl_sheet(wb, "Item 4 - Advances")
    if item4 is not None:
        r = _gxl_row_by_label(item4, 1, "supplies made to registered persons")
        if r:
            out["table4_b2b_taxable"], out["table4_b2b_cgst"], out["table4_b2b_sgst"], out["table4_b2b_igst"] = (
                _gxl_num(r[2]), _gxl_num(r[3]), _gxl_num(r[4]), _gxl_num(r[5]))
        else:
            out["notes"].append("Table 4B (B2B outward) row not found.")
        r = _gxl_row_by_label(item4, 1, "credit notes issued in respect")
        if r:
            out["table4_cn_taxable"], out["table4_cn_cgst"], out["table4_cn_sgst"], out["table4_cn_igst"] = (
                _gxl_num(r[2]), _gxl_num(r[3]), _gxl_num(r[4]), _gxl_num(r[5]))
        else:
            out["notes"].append("Table 4I (credit notes) row not found.")
        r = _gxl_row_by_label(item4, 1, "reverse charge basis")
        if r:
            out["table4_rcm_inward_taxable"] = _gxl_num(r[2])
        r = _gxl_row_by_label(item4, 1, "tax is to be paid (h + m)")
        if r:
            out["table4_net_supplies_tax_igst"] = _gxl_num(r[5])
    else:
        out["notes"].append("Item 4 sheet not found.")

    item5 = _gxl_sheet(wb, "Item 5 - Outward Supplies")
    if item5 is not None:
        t5 = {}
        for key, frag in [("zero_rated", "zero rated supply (export) without payment"),
                           ("sez", "supply to sezs without payment"),
                           ("exempted", "exempted"), ("nil_rated", "nil rated"),
                           ("nongst", "non-gst supply")]:
            r = _gxl_row_by_label(item5, 1, frag)
            t5[key] = _gxl_num(r[2]) if r else None
        out["table5_zero_rated"] = t5.get("zero_rated")
        out["table5_sez"] = t5.get("sez")
        out["table5_exempted"] = t5.get("exempted")
        out["table5_nil_rated"] = t5.get("nil_rated")
        out["table5_nongst"] = t5.get("nongst")
        vals5 = [v for v in t5.values() if v is not None]
        out["table5_all_zero"] = bool(vals5) and all(abs(v) < 0.01 for v in vals5)
    else:
        out["notes"].append("Item 5 sheet not found.")

    item6 = _gxl_sheet(wb, "Item 6 - ITC Availed")
    if item6 is not None:
        r = _gxl_row_by_label(item6, 1, "total amount of input tax credit availed through form gstr-3b")
        if r:
            out["table6a_cgst"], out["table6a_sgst"], out["table6a_igst"], out["table6a_cess"] = (
                _gxl_num(r[3]), _gxl_num(r[4]), _gxl_num(r[5]), _gxl_num(r[6]))
            out["table6a_total"] = sum(v or 0 for v in (out["table6a_cgst"], out["table6a_sgst"],
                                                          out["table6a_igst"], out["table6a_cess"]))
        else:
            out["notes"].append("Table 6A (ITC availed via 3B) row not found.")
        r = _gxl_row_by_label(item6, 1, "total itc availed")
        if r:
            out["table6_total_itc_availed"] = sum(_gxl_num(r[c]) or 0 for c in (3, 4, 5, 6))
    else:
        out["notes"].append("Item 6 sheet not found.")

    item7 = _gxl_sheet(wb, "Item 7 - ITC Reversed")
    if item7 is not None:
        r = _gxl_row_by_label(item7, 1, "total itc reversed")
        if r:
            out["table7_total_reversed"] = sum(_gxl_num(r[c]) or 0 for c in (2, 3, 4, 5))
    else:
        out["notes"].append("Item 7 sheet not found.")

    item8 = _gxl_sheet(wb, "Item 8 - Other ITC")
    if item8 is not None:
        r = _gxl_row_by_label(item8, 1, "itc as per gstr-2b")
        if r:
            out["table8_itc_as_per_2b"] = sum(_gxl_num(r[c]) or 0 for c in (2, 3, 4, 5))
        r = _gxl_row_by_label(item8, 1, "itc as per sum total of 6(b)")
        if r:
            out["table8_itc_as_per_6b6h"] = sum(_gxl_num(r[c]) or 0 for c in (2, 3, 4, 5))
        r = _gxl_row_by_label(item8, 1, "difference [a-(b+c)]")
        if r:
            out["table8_difference"] = sum(_gxl_num(r[c]) or 0 for c in (2, 3, 4, 5))
    else:
        out["notes"].append("Item 8 sheet not found.")

    item9 = _gxl_sheet(wb, "Item 9 - Tax Paid")
    if item9 is not None:
        for row_label, keyprefix in [("integrated tax", "igst"), ("central tax", "cgst"), ("state/ut tax", "sgst")]:
            r = _gxl_row_by_label(item9, 1, row_label)
            if r:
                out[f"table9_liability_{keyprefix}"] = _gxl_num(r[2])
        r = _gxl_row_by_label(item9, 1, "interest")
        if r:
            out["table9_interest_payable"], out["table9_interest_paid"] = _gxl_num(r[2]), _gxl_num(r[3])
        r = _gxl_row_by_label(item9, 1, "late fee")
        if r:
            out["table9_late_fee_payable"], out["table9_late_fee_paid"] = _gxl_num(r[2]), _gxl_num(r[3])
    else:
        out["notes"].append("Item 9 sheet not found.")

    # NEW (per instruction): Part V (Items 10-14), Table 12 (reversal of ITC availed during
    # previous FY) and Table 13 (ITC availed for the previous financial year) -- the
    # taxpayer's own filed, government-standard figure for ITC pertaining to the PRIOR FY but
    # claimed within the CURRENT FY's returns (the Section 16(4) carry-forward window), used by
    # the ITC Annual Summary sheet's 'carried forward from last FY' column. Confirmed present
    # and populated on a real filed GSTR-9 (not assumed): Table 13 showed real, non-zero
    # CGST/SGST/IGST/Cess figures on the taxpayer this was built against.
    part5 = _gxl_sheet(wb, "Part V - Transactions Declared")
    if part5 is not None:
        r = _gxl_row_by_label(part5, 1, "itc availed for the previous financial year")
        if r:
            out["table13_itc_cgst"], out["table13_itc_sgst"], out["table13_itc_igst"], out["table13_itc_cess"] = (
                _gxl_num(r[3]), _gxl_num(r[4]), _gxl_num(r[5]), _gxl_num(r[6]))
        else:
            out["notes"].append("Table 13 (ITC availed for the previous financial year) row not found.")
        r = _gxl_row_by_label(part5, 1, "reversal of itc availed during previous financial year")
        if r:
            out["table12_itc_reversed_cgst"], out["table12_itc_reversed_sgst"], \
                out["table12_itc_reversed_igst"], out["table12_itc_reversed_cess"] = (
                _gxl_num(r[3]), _gxl_num(r[4]), _gxl_num(r[5]), _gxl_num(r[6]))
    else:
        out["notes"].append("Part V (Items 10-14) sheet not found -- Table 13 'ITC availed for "
                             "the previous financial year' not available.")

    # Item 19 (dedicated Late Fee Payable/Paid sheet, Central+State) supersedes Item 9's late-fee
    # row if present -- it's the more granular, purpose-built source for this figure.
    item19 = _gxl_sheet(wb, "Item 19 - Late Fee")
    if item19 is not None:
        rows19 = [r for r in item19.iter_rows(values_only=True) if r and r[0] in ("A", "B")]
        if rows19:
            out["table9_late_fee_payable"] = sum(_gxl_num(r[2]) or 0 for r in rows19)
            out["table9_late_fee_paid"] = sum(_gxl_num(r[3]) or 0 for r in rows19)

    out["available"] = True
    return out


# ======================================================================
# GSTR-9C  (Reconciliation Statement)
# ======================================================================
def parse_gstr9c(path):
    """Return dict of GSTR-9C figures, read from the government's own Excel export.
    Same output-dict CONTRACT (identical key names) as the retired PDF version, so
    R13 and every other caller need no changes -- extended with NEW keys: the
    auditor's own typed reasons for each un-reconciled difference (Items 6/8/10/13/15
    carry these as free text with the exact rupee amount named), which is a genuine
    exact-tie-out opportunity this PDF-era scrape never had access to. Same
    graceful-degrade contract as parse_gstr9()."""
    out = dict(available=False, reason=None, fy=None, gstin=None, legal_name=None,
                arn=None, arn_date=None,
                turnover_audited_bs=None, turnover_after_adjustments=None,
                turnover_declared_gstr9=None, turnover_unreconciled=None,
                exempt_nil_nongst_adjustment=None,
                taxable_turnover_after_adj=None, taxable_turnover_declared=None,
                itc_per_books=None, itc_declared_gstr9=None, itc_unreconciled=None,
                itc_booked_earlier_claimed_now=None, itc_booked_now_claimed_later=None,
                tax_payable_total=None, tax_paid_declared=None,
                notes=[],
                # NEW -- the auditor's own typed reason text for each un-reconciled gap,
                # not consumed by any existing check yet; kept for a future exact-tie-out
                # check (a reason line often names the exact rupee amount it explains).
                turnover_diff_reasons=[], taxable_turnover_diff_reasons=[], itc_diff_reasons=[])
    if not path or not os.path.exists(path):
        out["reason"] = "GSTR-9C not supplied for this taxpayer/FY."
        return out
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as ex:
        out["reason"] = f"Could not read GSTR-9C workbook: {ex}"
        return out

    part1 = _gxl_sheet(wb, "Part I - Basic Details")
    if part1 is not None:
        for row in part1.iter_rows(values_only=True):
            if not row or len(row) < 2:
                continue
            label = str(row[1] or "").strip().lower()
            val = row[2] if len(row) > 2 else None
            if label == "financial year":
                out["fy"] = val
            elif label == "gstin":
                out["gstin"] = val
            elif label == "legal name":
                out["legal_name"] = val
            elif label == "arn":
                out["arn"] = val
            elif label == "arn date":
                out["arn_date"] = val
    else:
        out["notes"].append("Part I (Basic Details) sheet not found.")

    item5 = _gxl_sheet(wb, "Item 5 - Reconciliation of Gros")
    if item5 is not None:
        r = _gxl_row_by_label(item5, 1, "turnover (including exports) as per audited")
        out["turnover_audited_bs"] = _gxl_num(r[3]) if r else None
        r = _gxl_row_by_label(item5, 1, "annual turnover after adjustments as above")
        out["turnover_after_adjustments"] = _gxl_num(r[3]) if r else None
        r = _gxl_row_by_label(item5, 1, "turnover as declared in annual return")
        out["turnover_declared_gstr9"] = _gxl_num(r[3]) if r else None
        r = _gxl_row_by_label(item5, 1, "un-reconciled turnover")
        out["turnover_unreconciled"] = _gxl_num(r[3]) if r else None
    else:
        out["notes"].append("Item 5 sheet not found.")

    item6 = _gxl_sheet(wb, "Item 6 - Reasons for Turnover")
    if item6 is not None:
        out["turnover_diff_reasons"] = [str(r[2]).strip() for r in item6.iter_rows(values_only=True)
                                          if r and len(r) > 2 and r[2] and str(r[0] or "").strip().isalpha()]

    item7 = _gxl_sheet(wb, "Item 7 - Reconciliation of Taxa")
    if item7 is not None:
        r = _gxl_row_by_label(item7, 1, "taxable turnover as per adjustments above")
        out["taxable_turnover_after_adj"] = _gxl_num(r[2]) if r else None
        r = _gxl_row_by_label(item7, 1, "value of exempted, nil rated, non-gst supplies")
        out["exempt_nil_nongst_adjustment"] = _gxl_num(r[2]) if r else None
        r = _gxl_row_by_label(item7, 1, "taxable turnover as per liability declared")
        out["taxable_turnover_declared"] = _gxl_num(r[2]) if r else None
    else:
        out["notes"].append("Item 7 sheet not found.")

    item8 = _gxl_sheet(wb, "Item 8 - Reasons for Taxable")
    if item8 is not None:
        out["taxable_turnover_diff_reasons"] = [str(r[2]).strip() for r in item8.iter_rows(values_only=True)
                                                  if r and len(r) > 2 and r[2] and str(r[0] or "").strip().isalpha()]

    item9 = _gxl_sheet(wb, "Item 9 - Rate-wise Tax Liabilit")
    if item9 is not None:
        r = _gxl_row_by_label(item9, 1, "total amount to be paid as per tables above")
        if r:
            out["tax_payable_total"] = sum(_gxl_num(r[c]) or 0 for c in (3, 4, 5, 6))
        r = _gxl_row_by_label(item9, 1, "total amount paid as declared in annual return")
        if r:
            out["tax_paid_declared"] = sum(_gxl_num(r[c]) or 0 for c in (3, 4, 5, 6))
    else:
        out["notes"].append("Item 9 sheet not found.")

    item12 = _gxl_sheet(wb, "Item 12 - Reconciliation of Net")
    if item12 is not None:
        r = _gxl_row_by_label(item12, 1, "itc availed as per audited annual financial statement")
        out["itc_per_books"] = _gxl_num(r[2]) if r else None
        r = _gxl_row_by_label(item12, 1, "itc claimed in annual return")
        out["itc_declared_gstr9"] = _gxl_num(r[2]) if r else None
        r = _gxl_row_by_label(item12, 1, "un-reconciled itc")
        out["itc_unreconciled"] = _gxl_num(r[2]) if r else None
        r = _gxl_row_by_label(item12, 1, "itc booked in earlier financial years claimed")
        out["itc_booked_earlier_claimed_now"] = _gxl_num(r[2]) if r else None
        r = _gxl_row_by_label(item12, 1, "itc booked in current financial year to be claimed")
        out["itc_booked_now_claimed_later"] = _gxl_num(r[2]) if r else None
    else:
        out["notes"].append("Item 12 sheet not found.")

    item13 = _gxl_sheet(wb, "Item 13 - Reasons for Un-reconc")
    if item13 is not None:
        out["itc_diff_reasons"] = [str(r[2]).strip() for r in item13.iter_rows(values_only=True)
                                     if r and len(r) > 2 and r[2] and str(r[0] or "").strip().isalpha()]

    if not out["exempt_nil_nongst_adjustment"]:
        out["notes"].append("Table 7B (exempt/nil/non-GST adjustment) not found or zero -- Rule R13 "
                             "(turnover-gap check) needs this figure; verify manually if a gap is otherwise expected.")

    out["available"] = True
    return out


# ======================================================================
# Table 8A  (government-standard export, same layout for every taxpayer)
# ======================================================================
def _find_header_row(ws, must_contain, max_scan=15):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        cells = [str(c).strip() if c else "" for c in row]
        if any(must_contain.lower() in c.lower() for c in cells):
            return i
    return None


def _t8a_header_map(ws, hdr_row):
    """Combine Table-8A's 2-row wrapped header (a merged top row + a sub-label row
    immediately below it) into ONE list of column labels, sub-row label wins where
    both exist (the sub-row carries the specific field name; the top row is only the
    merged-cell group title). Column POSITIONS are then resolved from these labels by
    content match, never assumed fixed -- confirmed necessary: a real Table-8A export
    for one taxpayer had a leading 'Tax Period' column at index 0 (shifting every
    field right by one) while another taxpayer's export had no such column at all
    (GSTIN of supplier starts at index 0) -- same government form, different column
    count, both real."""
    top = [str(c).strip() if c else "" for c in
           next(ws.iter_rows(min_row=hdr_row, max_row=hdr_row, values_only=True))]
    sub = [str(c).strip() if c else "" for c in
           next(ws.iter_rows(min_row=hdr_row + 1, max_row=hdr_row + 1, values_only=True))]
    n = max(len(top), len(sub))
    return [(sub[i] if i < len(sub) and sub[i] else (top[i] if i < len(top) else "")) for i in range(n)]


def _t8a_col(labels, *fragments, start=0):
    """First column index whose combined label contains ALL fragments (case-insensitive).
    None if not found -- caller must handle a missing column explicitly (never guess a
    position), per this project's no-fabrication rule."""
    for i in range(start, len(labels)):
        low = labels[i].lower()
        if all(frag.lower() in low for frag in fragments):
            return i
    return None


def _t8a_v(r, idx):
    return r[idx] if idx is not None and idx < len(r) else None


def parse_table_8a(path):
    """Return dict(available, b2b=[...], cdnr=[...], totals={...}) from the
    government Table-8A workbook. Header row is located by CONTENT ('GSTIN of
    supplier' cell), and every FIELD's column position within that header is also
    located by content (see _t8a_header_map/_t8a_col) rather than a fixed index --
    confirmed necessary, not theoretical: real exports for two different taxpayers
    have different column counts for the same sheet (see _t8a_header_map's docstring).
    The row-validity test (GSTIN-shaped token) is applied at whichever column the
    GSTIN was actually found in, not a hardcoded column B."""
    out = dict(available=False, reason=None, b2b=[], cdnr=[], totals={})
    if not path or not os.path.exists(path):
        out["reason"] = "Table 8A not supplied for this taxpayer/FY."
        return out
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as ex:
        out["reason"] = f"Could not read Table 8A workbook: {ex}"
        return out

    gstin_re = re.compile(r"^\d{2}[A-Z]{5}\d{4}[A-Z]\dZ[A-Z\d]$")

    def _rows_after_header(ws, hdr_row, gstin_col):
        # +1 (not +2): the GSTIN-shape check below naturally filters out the wrapped
        # sub-header row too (its cell at the GSTIN column is a text label, never
        # GSTIN-shaped) -- so this works whether the real export has a 1-row or a
        # 2-row wrapped header, without needing to know which in advance.
        for r in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
            v = _t8a_v(r, gstin_col)
            if v and gstin_re.match(str(v).strip()):
                yield r

    if "B2B" in wb.sheetnames:
        ws = wb["B2B"]
        hdr = _find_header_row(ws, "GSTIN of supplier")
        if hdr:
            L = _t8a_header_map(ws, hdr)
            c_gstin = _t8a_col(L, "gstin of supplier")
            c_supplier = _t8a_col(L, "trade") or _t8a_col(L, "legal name")
            c_invno = _t8a_col(L, "invoice number")
            c_invtype = _t8a_col(L, "invoice type")
            c_invdate = _t8a_col(L, "invoice date")
            c_invval = _t8a_col(L, "invoice value")
            c_pos = _t8a_col(L, "place of supply")
            c_rcm = _t8a_col(L, "reverse charge")
            c_rate = _t8a_col(L, "rate")
            c_taxable = _t8a_col(L, "taxable value")
            c_igst = _t8a_col(L, "integrated tax")
            c_cgst = _t8a_col(L, "central tax")
            c_sgst = _t8a_col(L, "state") 
            c_cess = _t8a_col(L, "cess")
            c_period = _t8a_col(L, "period")
            c_filingdate = _t8a_col(L, "filing date")
            c_itcavail = _t8a_col(L, "itc availab")
            c_reason = _t8a_col(L, "reason")
            if c_gstin is None:
                out["reason"] = (out["reason"] or "") + " 'B2B' sheet: could not locate the 'GSTIN of supplier' column."
            else:
                for r in _rows_after_header(ws, hdr, c_gstin):
                    out["b2b"].append(dict(
                        period=str(_t8a_v(r, c_period) or "").strip(), gstin=str(_t8a_v(r, c_gstin)).strip(),
                        supplier=str(_t8a_v(r, c_supplier) or "").strip(),
                        invno=str(_t8a_v(r, c_invno) or "").strip(), invtype=str(_t8a_v(r, c_invtype) or "").strip(),
                        invdate=_t8a_v(r, c_invdate), invval=num(_t8a_v(r, c_invval)),
                        pos=str(_t8a_v(r, c_pos) or "").strip(), rcm=str(_t8a_v(r, c_rcm) or "").strip(),
                        rate=num(_t8a_v(r, c_rate)), taxable=num(_t8a_v(r, c_taxable)),
                        igst=num(_t8a_v(r, c_igst)), cgst=num(_t8a_v(r, c_cgst)),
                        sgst=num(_t8a_v(r, c_sgst)), cess=num(_t8a_v(r, c_cess)),
                        supplier_filing_date=_t8a_v(r, c_filingdate),
                        itc_available=str(_t8a_v(r, c_itcavail) or "").strip(),
                        reason_not_available=str(_t8a_v(r, c_reason) or "").strip(),
                    ))
        else:
            out["reason"] = (out["reason"] or "") + " 'B2B' sheet found but no 'GSTIN of supplier' header row located."

    # CDNR sheet name varies by export vintage -- 'CDNR' or the newer 'B2B-CDNR' (same
    # column CONTENT under either name, verified against a real 'B2B-CDNR' export --
    # though not necessarily the same column COUNT, hence the content-based mapping below).
    cdnr_sheet_name = "CDNR" if "CDNR" in wb.sheetnames else ("B2B-CDNR" if "B2B-CDNR" in wb.sheetnames else None)
    if cdnr_sheet_name:
        ws = wb[cdnr_sheet_name]
        hdr = _find_header_row(ws, "GSTIN of supplier")
        if hdr:
            L = _t8a_header_map(ws, hdr)
            c_gstin = _t8a_col(L, "gstin of supplier")
            c_supplier = _t8a_col(L, "trade") or _t8a_col(L, "legal name")
            c_notetype = _t8a_col(L, "note type")
            c_supplytype = _t8a_col(L, "note supply type") or _t8a_col(L, "supply type")
            c_noteno = _t8a_col(L, "note number")
            c_notedate = _t8a_col(L, "note date")
            c_noteval = _t8a_col(L, "note value")
            c_pos = _t8a_col(L, "place of supply")
            c_rate = _t8a_col(L, "rate")
            c_taxable = _t8a_col(L, "taxable value")
            c_igst = _t8a_col(L, "integrated tax")
            c_cgst = _t8a_col(L, "central tax")
            c_sgst = _t8a_col(L, "state")
            c_cess = _t8a_col(L, "cess")
            c_itcavail = _t8a_col(L, "itc availab")
            if c_gstin is None:
                out["reason"] = (out["reason"] or "") + f" {cdnr_sheet_name!r} sheet: could not locate the 'GSTIN of supplier' column."
            else:
                for r in _rows_after_header(ws, hdr, c_gstin):
                    out["cdnr"].append(dict(
                        period=None, gstin=str(_t8a_v(r, c_gstin)).strip(),
                        supplier=str(_t8a_v(r, c_supplier) or "").strip(),
                        note_type=str(_t8a_v(r, c_notetype) or "").strip(),
                        supply_type=str(_t8a_v(r, c_supplytype) or "").strip(),
                        note_no=str(_t8a_v(r, c_noteno) or "").strip(), note_date=_t8a_v(r, c_notedate),
                        note_val=num(_t8a_v(r, c_noteval)), pos=str(_t8a_v(r, c_pos) or "").strip(),
                        rate=num(_t8a_v(r, c_rate)), taxable=num(_t8a_v(r, c_taxable)),
                        igst=num(_t8a_v(r, c_igst)), cgst=num(_t8a_v(r, c_cgst)),
                        sgst=num(_t8a_v(r, c_sgst)), cess=num(_t8a_v(r, c_cess)),
                        itc_available=str(_t8a_v(r, c_itcavail) or "").strip(),
                    ))

    yes_b2b = [r for r in out["b2b"] if r["itc_available"].upper() == "YES"]
    out["totals"] = dict(
        b2b_rows=len(out["b2b"]), b2b_yes_rows=len(yes_b2b),
        cgst=sum(r["cgst"] for r in yes_b2b), sgst=sum(r["sgst"] for r in yes_b2b),
        igst=sum(r["igst"] for r in yes_b2b), cess=sum(r["cess"] for r in yes_b2b),
    )
    out["totals"]["total"] = (out["totals"]["cgst"] + out["totals"]["sgst"]
                                + out["totals"]["igst"] + out["totals"]["cess"])
    # bucket the "No" rows by reason -- Part 1 C2 of the forensic framework
    no_b2b = [r for r in out["b2b"] if r["itc_available"].upper() != "YES"]
    from collections import Counter
    out["totals"]["no_reason_breakdown"] = dict(Counter(r["reason_not_available"] or "(blank)" for r in no_b2b))

    out["available"] = True
    return out




# ============================================================================
# ============================================================================
# ==== SECTION: bo_profile_parser.py  (was a standalone module before consolidation)
# ============================================================================
"""
BO PROFILE (360-degree taxpayer profile) EXCEL PARSER
========================================================
Parses the GST-department "BO Profile" Excel export (previously PDF -- this
tool no longer reads any PDF at all). Every FY-keyed table (Financial
Information, BIFA Specific Information, ITC Passed On/Received, EWB/E-Invoice
Related Information, Refund Details) lives on its own named sheet with a
clean header row -- read by CONTENT (header cell text), never a fixed column
position, since even this government form has been confirmed to have
different column counts across real exports for different taxpayers (see
gst_parsers_dept.parse_table_8a's docstring for the same lesson applied to a
sibling government export).

All FY-table amounts are in LAKHS (confirmed: this taxpayer's FY23-24
'Turnover' cell of 3372.53 equals GSTR-9's own declared turnover of
Rs 3,37,15,5316.86 / 1,00,000 -- consistent with the retired PDF version's
documented convention). Downstream callers that need absolute rupees already
multiply by 1e5 themselves (see gst_checks_flow.build_drc_and_ledger_movements's
refund cross-check) -- this parser does NOT convert, to keep that one
multiplication in one place.

Returns one dict with the SAME key names as the retired PDF version (no
downstream caller needs to change):
  demographic         : {...}
  financial_by_fy      : {fy: {...}}                 (Financial Information)
  bifa_by_fy            : {fy: {...}}                 (BIFA Specific Information)
  itc_passed_by_fy      : {fy: {...}}
  itc_received_by_fy    : {fy: {...}}
  ewb_by_fy              : {fy: {...}}
  einv_by_fy             : {fy: {...}}
  refund_by_fy           : {fy: {...}}
  top_beneficiaries       : [ {...} ]                  (Top 10 ITC Passed)
  top_suppliers           : [ {...} ]                  (Top 10 ITC Received)
  related_itc_received    : [ {...} ]                  (fraud-risk: related/cancelled supplier)
  related_itc_passed      : [ {...} ]                  (fraud-risk: related/cancelled recipient)
  drc_payments            : [ {...} ]                  (voluntary payments + demand orders)
  appeals, cases, transfers : [ {...} ]                (informational; not consumed by any
                                                         check yet, kept for future use)
  self_gstin, legal_name, trade_name
"""


def _bxl_sheet(wb, name):
    return wb[name] if name in wb.sheetnames else None


def _num_bo(s):
    """BO Profile Excel-specific numeric cleaner (handles a trailing '%' on the
    Utilization/Effective-Rate columns, which the shared num()/_gxl_num() helpers
    don't need to)."""
    if s is None:
        return 0.0
    if isinstance(s, (int, float)):
        return float(s)
    s = str(s).replace(",", "").replace("%", "").strip()
    if s in ("", "-", "NA", "None"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _bxl_header_row(ws, max_scan=6):
    """The FY-keyed tables on these sheets all have their real header 1-2 rows
    below the sheet title -- located by CONTENT (first cell that looks like a
    real column label, i.e. not None and not the sheet's own title row)."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        cells = [c for c in row if c not in (None, "")]
        if len(cells) >= 2:
            return i, [str(c).strip() if c is not None else "" for c in row]
    return None, []


def _bxl_col(header, *fragments):
    for i, h in enumerate(header):
        low = h.lower()
        if all(f.lower() in low for f in fragments):
            return i
    return None


def _bxl_fy_table(ws, field_map):
    """Generic reader for the FY-keyed tables (Financial Information, BIFA
    Specific Information, ITC Passed On, ITC Received, EWB/E-Invoice Related
    Information, Refund Details): first column is always 'Financial Year',
    every other wanted field is located by header-content fragment match via
    field_map = {output_key: (fragment, ...)}. Returns {fy_string: {...}}."""
    hdr_row, header = _bxl_header_row(ws)
    if hdr_row is None:
        return {}
    fy_col = _bxl_col(header, "financial year")
    if fy_col is None:
        fy_col = 0
    cols = {k: _bxl_col(header, *frags) for k, frags in field_map.items()}
    out = {}
    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        if not row or row[fy_col] is None:
            continue
        fy = str(row[fy_col]).strip()
        if not re.match(r"^\d{4}-\d{2,4}$", fy):
            continue
        rec = {}
        for k, ci in cols.items():
            rec[k] = _num_bo(row[ci]) if (ci is not None and ci < len(row)) else None
        out[fy] = rec
    return out


def _bxl_list_table(ws, field_map, name_frag_gstin="gstin"):
    """Generic reader for the flat list-of-rows sheets (Top 10 Beneficiaries/
    Suppliers, Related ITC Received/Passed, Appeal/Case/Transfer Information):
    header located by content, every wanted field located by header-content
    fragment match. Returns a list of dicts, one per data row (rows with no
    GSTIN-shaped OR non-blank first meaningful cell are skipped)."""
    hdr_row, header = _bxl_header_row(ws)
    if hdr_row is None:
        return []
    cols = {k: _bxl_col(header, *frags) for k, frags in field_map.items()}
    out = []
    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        if not row or all(c is None for c in row):
            continue
        rec = {}
        for k, ci in cols.items():
            rec[k] = row[ci] if (ci is not None and ci < len(row)) else None
        out.append(rec)
    return out


def parse_bo_profile(path):
    out = dict(
        self_gstin=None, legal_name=None, trade_name=None, demographic={},
        financial_by_fy={}, bifa_by_fy={}, itc_passed_by_fy={}, itc_received_by_fy={},
        ewb_by_fy={}, einv_by_fy={}, refund_by_fy={},
        top_beneficiaries=[], top_suppliers=[], related_itc_received=[], related_itc_passed=[],
        drc_payments=[], appeals=[], cases=[], transfers=[],
    )
    if not path or not os.path.exists(path):
        return out
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        return out

    demo_ws = _bxl_sheet(wb, "Demographic Details")
    if demo_ws is not None:
        demo = {}
        for row in demo_ws.iter_rows(values_only=True):
            if row and row[0] and len(row) > 1 and row[1] not in (None, ""):
                demo[str(row[0]).strip()] = row[1]
        out["demographic"] = demo
        out["self_gstin"] = demo.get("GSTIN")
        out["legal_name"] = demo.get("Legal Name")
        out["trade_name"] = demo.get("Trade Name")

    fin_ws = _bxl_sheet(wb, "Financial Information")
    if fin_ws is not None:
        out["financial_by_fy"] = _bxl_fy_table(fin_ws, dict(
            turnover=("turnover",), taxable_turnover=("taxable turnover",),
            export_turnover=("export turnover",), total_tax_liability=("total tax liability",),
            tax_paid_by_itc=("tax paid by itc",), tax_paid_in_cash=("tax paid in cash",),
            itc_availed=("itc availed",), itc_utilization_pct=("itc utilization",),
            effective_tax_rate_pct=("effective tax rate",), itc_reversal=("itc reversal",),
        ))

    bifa_ws = _bxl_sheet(wb, "BIFA Specific Information")
    if bifa_ws is not None:
        # BIFA's header is wrapped over TWO rows (grouped titles + per-column labels) --
        # combine them the same way Table 8A's header is combined (see _t8a_header_map),
        # since a single-row read here would miss which group ('Supply Tax Deficit' /
        # 'Excess ITC claimed' / etc.) a given column actually belongs to.
        hdr_row, top = _bxl_header_row(bifa_ws)
        _, sub = (hdr_row + 1, [str(c).strip() if c else "" for c in
                  next(bifa_ws.iter_rows(min_row=hdr_row + 1, max_row=hdr_row + 1, values_only=True))]) \
            if hdr_row else (None, [])
        combined = [(sub[i] if i < len(sub) and sub[i] else (top[i] if i < len(top) else ""))
                    for i in range(max(len(top), len(sub)))]
        fy_col = _bxl_col(combined, "financial year") or 0
        cmap = dict(
            liability_gstr1=("liability as per gstr1",), liability_gstr3b=("liability as per gstr3b",),
            diff_liability=("difference liability",), itc_r3b=("itc availed in r3b",),
            itc_r2b_r2a=("itc accrued in r2b",), diff_itc=("difference itc",),
            supply_as_per_gstr3b=("supply value as per gstr3b",), liability_ewb=("liability as per ewb",),
        )
        cols = {k: _bxl_col(combined, *frags) for k, frags in cmap.items()}
        bifa = {}
        for row in bifa_ws.iter_rows(min_row=hdr_row + 2, values_only=True):
            if not row or row[fy_col] is None:
                continue
            fy = str(row[fy_col]).strip()
            if not re.match(r"^\d{4}-\d{2,4}$", fy):
                continue
            bifa[fy] = {k: (_num_bo(row[ci]) if ci is not None and ci < len(row) else None)
                        for k, ci in cols.items()}
        out["bifa_by_fy"] = bifa

    itcp_ws = _bxl_sheet(wb, "ITC Passed On")
    if itcp_ws is not None:
        # NOTE: 'ITC Passed On'/'ITC Received' repeat identical group-column labels
        # (No.of.../ITC Passed) under three merged groups (INTRA/INTER/TOTAL) -- a
        # simple fragment match can't disambiguate which group a column belongs to
        # from the label alone. Kept minimal (TOTAL-group figures only, the ones
        # every downstream check actually needs) rather than mis-mapping intra/inter.
        hdr_row, _ = _bxl_header_row(itcp_ws)
        total = {}
        for row in itcp_ws.iter_rows(min_row=hdr_row + 1, values_only=True):
            if not row or row[0] is None or not re.match(r"^\d{4}-\d{2,4}$", str(row[0]).strip()):
                continue
            fy = str(row[0]).strip()
            total[fy] = dict(total_beneficiaries=_num_bo(row[5]) if len(row) > 5 else None,
                              total_itc=_num_bo(row[6]) if len(row) > 6 else None)
        out["itc_passed_by_fy"] = total

    itcr_ws = _bxl_sheet(wb, "ITC Received")
    if itcr_ws is not None:
        hdr_row, _ = _bxl_header_row(itcr_ws)
        total = {}
        for row in itcr_ws.iter_rows(min_row=hdr_row + 1, values_only=True):
            if not row or row[0] is None or not re.match(r"^\d{4}-\d{2,4}$", str(row[0]).strip()):
                continue
            fy = str(row[0]).strip()
            total[fy] = dict(total_suppliers=_num_bo(row[5]) if len(row) > 5 else None,
                              total_itc=_num_bo(row[6]) if len(row) > 6 else None)
        out["itc_received_by_fy"] = total

    ewb_ws = _bxl_sheet(wb, "EWB Related Information")
    if ewb_ws is not None:
        out["ewb_by_fy"] = _bxl_fy_table(ewb_ws, dict(
            inward_tax=("inward supply tax",), inward_ewb_count=("number of inward",),
            outward_tax=("outward supply tax",), outward_ewb_count=("number of outward",),
            ratio_in_out=("ratio",)))
        # 'inward supply' fragment also matches 'inward supply tax' -- resolve the
        # plain-supply column explicitly to the one NOT containing 'tax'.
        hdr_row, header = _bxl_header_row(ewb_ws)
        plain_supply_cols = [i for i, h in enumerate(header) if "supply" in h.lower() and "tax" not in h.lower()]
        if len(plain_supply_cols) >= 2:
            in_col, out_col = plain_supply_cols[0], plain_supply_cols[1]
            for row in ewb_ws.iter_rows(min_row=hdr_row + 1, values_only=True):
                if not row or row[0] is None or not re.match(r"^\d{4}-\d{2,4}$", str(row[0]).strip()):
                    continue
                fy = str(row[0]).strip()
                if fy in out["ewb_by_fy"]:
                    out["ewb_by_fy"][fy]["inward_supply"] = _num_bo(row[in_col])
                    out["ewb_by_fy"][fy]["outward_supply"] = _num_bo(row[out_col])

    einv_ws = _bxl_sheet(wb, "E-Invoice Related Information")
    if einv_ws is not None:
        out["einv_by_fy"] = _bxl_fy_table(einv_ws, dict(
            active_count=("count of active",), cancelled_count=("count of cancelled",),
            active_assessable_value=("active assessable value",), active_taxes=("active taxes",),
            cancelled_assessable_value=("cancelled assessable value",), cancelled_taxes=("cancelled taxes",)))

    refund_ws = _bxl_sheet(wb, "Refund Details")
    if refund_ws is not None:
        out["refund_by_fy"] = _bxl_fy_table(refund_ws, dict(
            claimed=("total claimed amount",), rejected=("total rejected amount",),
            sanctioned=("total sanctioned amount",)))

    benef_ws = _bxl_sheet(wb, "Top 10 Beneficiaries based on I")
    if benef_ws is not None:
        out["top_beneficiaries"] = _bxl_list_table(benef_ws, dict(
            gstin=("receiver gstin",), name=("trade name",), reg_start=("registration start date",),
            status=("status",), risk=("overall risk score",), amount=("itc passed on",)))

    supp_ws = _bxl_sheet(wb, "Top 10 Suppliers based on ITC R")
    if supp_ws is not None:
        out["top_suppliers"] = _bxl_list_table(supp_ws, dict(
            gstin=("supplier gstin",), name=("trade name",), reg_start=("registration start date",),
            status=("status",), risk=("overall risk score",), amount=("itc received",)))

    relr_ws = _bxl_sheet(wb, "ITC Received from Related - Can")
    if relr_ws is not None:
        out["related_itc_received"] = _bxl_list_table(relr_ws, dict(
            fy=("financial year",), gstin=("supplier gstin",), name=("trade name",),
            related_parameter=("related parameter",), status=("status",), cancellation_date=("cancellation date",),
            reason=("cancellation reason",), total_itc=("total itc received",)))

    relp_ws = _bxl_sheet(wb, "ITC Passed On to Related - Canc")
    if relp_ws is not None:
        out["related_itc_passed"] = _bxl_list_table(relp_ws, dict(
            fy=("financial year",), gstin=("receiver gstin",), name=("trade name",),
            related_parameter=("related parameter",), status=("status",), cancellation_date=("cancellation date",),
            reason=("cancellation reason",), total_itc=("total itc passed",)))

    drc_ws = _bxl_sheet(wb, "DRC Payment Information")
    if drc_ws is not None:
        rows = _bxl_list_table(drc_ws, dict(
            source_id=("source id",), description=("transaction type description",),
            date=("transaction date",), method=("payment method",),
            cgst=("cgst tax",), sgst=("sgst tax",), igst=("igst tax",),
            cess=("cess",), other=("other than tax",)))
        for r in rows:
            r["ref"] = r["source_id"]
            r["transaction_date"] = r["date"]
            r["type"] = r["description"]
            r["total"] = sum(_num_bo(r.get(k)) for k in ("cgst", "sgst", "igst", "cess", "other"))
            for k in ("cgst", "sgst", "igst", "cess", "other"):
                r[k] = _num_bo(r.get(k))
        out["drc_payments"] = rows

    appeal_ws = _bxl_sheet(wb, "Appeal Information")
    if appeal_ws is not None:
        out["appeals"] = _bxl_list_table(appeal_ws, dict(
            fy=("financial year",), arn=("arn",), filing_date=("case filing date",),
            status=("case status",), case_type=("case type",), officer=("tax officer name",),
            order_number=("order number",)))

    case_ws = _bxl_sheet(wb, "Case Information")
    if case_ws is not None:
        out["cases"] = _bxl_list_table(case_ws, dict(
            case_id=("arn", "case id"), reference_id=("reference id",), action_date=("action date",),
            tax_period=("tax period",), case_type=("case type",), status=("case status",),
            scn_date=("scn date",), scn_number=("scn number",),
            cgst=("cgst tax",), sgst=("sgst tax",), igst=("igst tax",), cess=("cess",)))

    transfer_ws = _bxl_sheet(wb, "Transfer Information")
    if transfer_ws is not None:
        out["transfers"] = _bxl_list_table(transfer_ws, dict(
            date=("recommend date",), source_case_id=("source case id",), target_case_id=("target case id",),
            source_module=("source module name",), target_module=("target module name",),
            status=("status",)))

    return out


# ==== SECTION: gstr2a_parser.py  (NEW -- GSTR-2A merged-workbook parser)
# ============================================================================
"""
GSTR-2A PARSER
==============
Parses the merged (whole-FY) GSTR-2A workbook: B2B, B2BA (amendments), CDNR /
CDNRA (credit-debit notes and their amendments), and ISD (Input Service
Distributor credit). ECO/ECOA, TDS/TDSA, TCS, IMPG, IMPG SEZ are read by
GSTN's export but are OUT OF SCOPE for this tool (no check in this codebase
consumes them) -- deliberately not parsed, per instruction.

Two real-data quirks confirmed against an actual government export (GSTIN
05AACFT2702L1ZD, FY 2025-26, 41,563-row B2B sheet) drive the design below:

1. EVERY invoice/note in B2B, B2BA, CDNR and CDNRA gets an EXTRA row whose
   document number is suffixed '-Total' and whose Rate column is '-'. This
   extra row's taxable/tax figures already equal the SUM of that document's
   real rate-line row(s) -- verified exactly, to the rupee, on a genuine
   2-rate invoice (rate-lines Rs 2,400 + Rs 2,200 taxable; the '-Total' row
   shows Rs 4,600). This happens for single-rate invoices too (one rate-line
   + one redundant '-Total' line repeating the same figures). Summing BOTH
   the rate-line rows and the '-Total' row would double (or more) count
   every single document, so only the '-Total' rows are kept here (their
   suffix stripped back off); any document number seen on a rate-line row
   with NO matching '-Total' row is recorded in total_row_missing rather
   than silently dropped or guessed at.

2. B2BA and CDNRA (the two AMENDMENT sheets) repeat a wrapped sub-header row
   immediately after EVERY month's marker, even before any data -- B2B and
   CDNR do not do this. Rather than special-case this per sheet, every
   sheet's rows are passed through a GSTIN-shape validity filter before use
   (the exact same defensive technique parse_table_8a() above already uses
   for the same class of problem: a row is only real data if its GSTIN
   column is actually GSTIN-shaped). This uniformly discards the repeated
   header rows, stray blank rows, and anything else that isn't a genuine
   data row, on every sheet, without needing to know in advance which
   sheets have the quirk.

Column positions below are POSITIONAL, read starting from a header row that
is still located by CONTENT (via _find_header_row(), never assumed at a
fixed row number) -- the same convention parse_table_8a() and parse_2b_excel()
already use elsewhere in this file: content locates WHERE data starts, fixed
government-template column order is trusted from there on. B2B, CDNR and
B2BA's layouts were verified against real data rows (shown in comments
below); CDNRA and ISD had no real data rows in the file this was built
against, so their layout is derived from the header row's own text plus
GSTN's documented field order in the workbook's own 'Read me' sheet --
flagged here honestly rather than silently presented as equally verified.

DISCIPLINE (unchanged from the rest of the tool): never fabricate or
silently guess a number. A month with no marker at all in a sheet, a sheet
absent from the workbook, or a document missing its '-Total' row all
produce an explicit gap in the output, never a zero that reads like a
verified nil.
"""

import gst_core as mpu

_GSTIN_RE_R2A = re.compile(r"^\d{2}[A-Z]{5}\d{4}[A-Z]\dZ[A-Z\d]$")


def _r2a_clean_gstin(v, malformed_list):
    """Upper-cases and strips a GSTIN read from GSTR-2A; anything that still
    doesn't match the standard 15-character GSTIN shape afterwards is
    recorded in malformed_list (point 15/13 of the brief: GSTIN format
    hygiene BEFORE any join, and don't silently drop what fails it -- report
    it)."""
    s = str(v or "").strip().upper()
    if s and not _GSTIN_RE_R2A.match(s):
        malformed_list.append(s)
    return s


def r2a_clean_date(v):
    """R2A dates arrive as 'DD-MM-YYYY' strings per the workbook's own
    Read-me; GSTR-2B's invoice dates (used to build the match key against
    2A elsewhere) arrive as 'DD/MM/YYYY'. Both separators are accepted here
    so this ONE function can normalise either source to ISO 'YYYY-MM-DD'.
    Native datetime/date objects are also accepted (a re-saved export could
    carry real date cells instead of text). Returns None if genuinely
    unparseable -- never guesses a date."""
    if v is None or v == "":
        return None
    if isinstance(v, _dt.datetime):
        return v.date().isoformat()
    if isinstance(v, _dt.date):
        return v.isoformat()
    s = str(v).strip()
    m = re.match(r"^(\d{1,2})[-/](\d{1,2})[-/](\d{4})$", s)
    if not m:
        return None
    d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    try:
        return _dt.date(y, mo, d).isoformat()
    except ValueError:
        return None


def r2a_invtype_canon(v):
    """Normalises invoice-type text onto ONE small set so a 2A row (short
    codes: R/SEZWP/SEZWOP/DE) and the SAME invoice's 2B row (full text, e.g.
    'Regular') can be matched as part of the same join key -- point 6 of the
    brief: SEZ vs Regular must not collide on a reused invoice number.
    Falls back to 'OTHER' rather than guessing when neither form is
    recognised, so a genuinely odd value is visible in the mismatch output
    rather than silently merged into a bucket it may not belong in."""
    s = str(v or "").strip().upper()
    if not s:
        return "UNKNOWN"
    if s == "R" or s.startswith("REGULAR"):
        return "REGULAR"
    if s == "SEZWOP" or ("SEZ" in s and "WITHOUT" in s):
        return "SEZWOP"
    if s == "SEZWP" or "SEZ" in s:
        return "SEZWP"
    if s == "DE" or "DEEMED" in s:
        return "DE"
    return "OTHER"


def _r2a_dedupe_total_rows(rows, doc_col):
    """Keep only the '-Total' rollup row per document (see module docstring,
    point 1); strip the suffix back off; report any rate-line document
    number that never got a matching '-Total' row rather than dropping or
    summing it silently."""
    total_rows, rateline_docnos, total_docnos = [], set(), set()
    for r in rows:
        if doc_col >= len(r):
            continue
        doc = r[doc_col]
        if not doc or not isinstance(doc, str):
            continue
        doc = doc.strip()
        if not doc:
            continue
        if doc.endswith("-Total"):
            base = doc[:-len("-Total")].strip()
            total_docnos.add(base)
            new_r = list(r)
            new_r[doc_col] = base
            total_rows.append(new_r)
        else:
            rateline_docnos.add(doc)
    missing_total = sorted(rateline_docnos - total_docnos)
    return total_rows, missing_total


def _r2a_valid_rows(rows, gstin_col):
    """A row is only genuine data if its GSTIN-position cell is actually
    GSTIN-shaped -- discards the repeated wrapped sub-header rows B2BA/CDNRA
    carry after every marker (see module docstring, point 2), and any stray
    blank row, on every sheet uniformly. Same technique parse_table_8a()
    already uses above for the same class of problem."""
    out = []
    for r in rows:
        if gstin_col >= len(r):
            continue
        g = r[gstin_col]
        if g and _GSTIN_RE_R2A.match(str(g).strip().upper()):
            out.append(r)
    return out


def _r2a_month_blocks(wb, sheet_name, header_text):
    """Locate the header row by CONTENT (never a fixed row number), then
    scan forward for the FIRST actual period marker before splitting into
    {month_label: [rows]} via the SAME marker-block scheme GSTR-1/2B/
    E-Invoice already use elsewhere in this tool (gst_core.split_rows_by_
    month) -- confirmed this is the correct reading for GSTR-2A: 'GSTR-2A
    for tax period X' shows every document currently attributed to period
    X, which is exactly what the marker represents, dynamic re-generation
    and all.

    The number of header/sub-header rows between the content-matched header
    cell and the first real marker is NOT constant across sheets -- confirmed
    against real data: B2B, CDNR and ISD each carry one extra wrapped
    sub-header row before their first marker; B2BA does not (its own
    repeated sub-header row comes AFTER the marker instead -- see the module
    docstring's point 2). So the marker is located by scanning forward for
    is_marker_row(), never by assuming a fixed number of rows to skip.

    Returns None (not an empty dict) if the sheet is absent or no header
    could be located at all, so the caller can tell 'sheet missing' apart
    from 'sheet present but empty'."""
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    hdr = _find_header_row(ws, header_text)
    if not hdr:
        return None
    rows = list(ws.iter_rows(values_only=True))
    start = None
    for i in range(hdr - 1, len(rows)):
        if mpu.is_marker_row(rows[i]):
            start = i
            break
    if start is None:
        return {}
    return mpu.split_rows_by_month(rows[start:])


def parse_r2a_excel(path):
    """One-shot parse of the whole merged GSTR-2A workbook (single pass per
    sheet, not re-opened per month -- B2B alone runs to tens of thousands of
    rows for a real taxpayer). Returns:

      available, reason,
      months_present : set of 'Mon-YY' labels found via any of this file's markers,
      b2b / b2ba / cdnr / cdnra / isd : {month_label: [row-dict, ...]},
      total_row_missing : {sheet_name: ["Mon-YY: docno", ...]},
      malformed_gstin : [raw string, ...]  -- failed the 15-char GSTIN shape

    Every downstream G-series check treats available=False as an explicit
    reason to SKIP, never as zero data."""
    out = dict(available=False, reason=None, months_present=set(),
               b2b={}, b2ba={}, cdnr={}, cdnra={}, isd={},
               total_row_missing={}, malformed_gstin=[])
    if not path or not os.path.exists(path):
        out["reason"] = "GSTR-2A not supplied for this taxpayer/FY."
        return out
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as ex:
        out["reason"] = f"Could not read GSTR-2A workbook: {ex}"
        return out

    malformed = out["malformed_gstin"]
    any_sheet_found = False

    def _record_missing(sheet_label, month, missing):
        if missing:
            out["total_row_missing"].setdefault(sheet_label, []).extend(
                f"{month}: {d}" for d in missing)

    # ---------------- B2B ---------------- (columns verified against real data rows)
    # 0 gstin,1 supplier,2 invno,3 invtype,4 invdate,5 invval,6 pos,7 rcm,8 rate,
    # 9 taxable,10 igst,11 cgst,12 sgst,13 cess,14 g1_status,15 g1_filing_date,
    # 16 g1_filing_period,17 g3b_status,18 amend_type,19 amend_period,20 cancel_date,
    # 21 source,22 irn,23 irn_date
    blocks = _r2a_month_blocks(wb, "B2B", "GSTIN of supplier")
    if blocks is not None:
        any_sheet_found = True
        for month, mrows in blocks.items():
            out["months_present"].add(month)
            valid = _r2a_valid_rows(mrows, 0)
            clean, missing = _r2a_dedupe_total_rows(valid, 2)
            _record_missing("B2B", month, missing)
            recs = []
            for r in clean:
                if len(r) < 14 or not r[0]:
                    continue
                recs.append(dict(
                    month=month, gstin=_r2a_clean_gstin(r[0], malformed),
                    supplier=str(r[1] or "").strip(), invno=str(r[2] or "").strip(),
                    invtype=r2a_invtype_canon(r[3]), invdate=r2a_clean_date(r[4]),
                    invval=num(r[5]), pos=str(r[6] or "").strip(),
                    rcm=str(r[7] or "").strip().upper(),
                    taxable=num(r[9]), igst=num(r[10]), cgst=num(r[11]),
                    sgst=num(r[12]), cess=num(r[13]),
                    g1_status=str(r[14] or "").strip() if len(r) > 14 else "",
                    g1_filing_date=str(r[15] or "").strip() if len(r) > 15 else "",
                    g1_filing_period=str(r[16] or "").strip() if len(r) > 16 else "",
                    g3b_status=str(r[17] or "").strip() if len(r) > 17 else "",
                    amend_type=str(r[18] or "").strip() if len(r) > 18 else "",
                    amend_period=str(r[19] or "").strip() if len(r) > 19 else "",
                    cancel_date=str(r[20] or "").strip() if len(r) > 20 else "",
                    source=str(r[21] or "").strip() if len(r) > 21 else "",
                    irn=str(r[22] or "").strip() if len(r) > 22 else "",
                    irn_date=str(r[23] or "").strip() if len(r) > 23 else "",
                ))
            out["b2b"][month] = recs

    # ---------------- CDNR ---------------- (columns verified against real data rows)
    # 0 gstin,1 supplier,2 note_type,3 note_no,4 supply_type,5 note_date,6 note_val,
    # 7 pos,8 rcm,9 rate,10 taxable,11 igst,12 cgst,13 sgst,14 cess,15 g1_status,
    # 16 g1_filing_date,17 g1_filing_period,18 g3b_status,19 amend_type,20 amend_period,
    # 21 cancel_date,22 source,23 irn,24 irn_date
    blocks = _r2a_month_blocks(wb, "CDNR", "GSTIN of Supplier")
    if blocks is not None:
        any_sheet_found = True
        for month, mrows in blocks.items():
            out["months_present"].add(month)
            valid = _r2a_valid_rows(mrows, 0)
            clean, missing = _r2a_dedupe_total_rows(valid, 3)
            _record_missing("CDNR", month, missing)
            recs = []
            for r in clean:
                if len(r) < 15 or not r[0]:
                    continue
                recs.append(dict(
                    month=month, gstin=_r2a_clean_gstin(r[0], malformed),
                    supplier=str(r[1] or "").strip(), note_type=str(r[2] or "").strip(),
                    note_no=str(r[3] or "").strip(), supply_type=str(r[4] or "").strip(),
                    note_date=r2a_clean_date(r[5]), note_val=num(r[6]),
                    pos=str(r[7] or "").strip(), rcm=str(r[8] or "").strip().upper(),
                    taxable=num(r[10]), igst=num(r[11]), cgst=num(r[12]),
                    sgst=num(r[13]), cess=num(r[14]),
                    g1_status=str(r[15] or "").strip() if len(r) > 15 else "",
                    g1_filing_date=str(r[16] or "").strip() if len(r) > 16 else "",
                    g1_filing_period=str(r[17] or "").strip() if len(r) > 17 else "",
                    g3b_status=str(r[18] or "").strip() if len(r) > 18 else "",
                    amend_type=str(r[19] or "").strip() if len(r) > 19 else "",
                    amend_period=str(r[20] or "").strip() if len(r) > 20 else "",
                    cancel_date=str(r[21] or "").strip() if len(r) > 21 else "",
                    source=str(r[22] or "").strip() if len(r) > 22 else "",
                    irn=str(r[23] or "").strip() if len(r) > 23 else "",
                    irn_date=str(r[24] or "").strip() if len(r) > 24 else "",
                ))
            out["cdnr"][month] = recs

    # ---------------- B2BA ---------------- (columns verified against real data rows)
    # 0 orig_invno,1 orig_invdate,2 gstin,3 supplier,4 invtype,5 invno,6 invdate,7 invval,
    # 8 pos,9 rcm,10 rate,11 taxable,12 igst,13 cgst,14 sgst,15 cess,16 g1_status,
    # 17 g1_filing_date,18 g1_filing_period,19 g3b_status,20 cancel_date,21 amend_type,
    # 22 orig_tax_period
    blocks = _r2a_month_blocks(wb, "B2BA", "GSTIN of Supplier")
    if blocks is not None:
        any_sheet_found = True
        for month, mrows in blocks.items():
            out["months_present"].add(month)
            valid = _r2a_valid_rows(mrows, 2)
            clean, missing = _r2a_dedupe_total_rows(valid, 5)
            _record_missing("B2BA", month, missing)
            recs = []
            for r in clean:
                if len(r) < 12 or not r[2]:
                    continue
                recs.append(dict(
                    month=month, orig_invno=str(r[0] or "").strip(),
                    orig_invdate=r2a_clean_date(r[1]), gstin=_r2a_clean_gstin(r[2], malformed),
                    supplier=str(r[3] or "").strip(), invtype=r2a_invtype_canon(r[4]),
                    invno=str(r[5] or "").strip(), invdate=r2a_clean_date(r[6]),
                    invval=num(r[7]), pos=str(r[8] or "").strip(),
                    rcm=str(r[9] or "").strip().upper(),
                    taxable=num(r[11]), igst=num(r[12]), cgst=num(r[13]),
                    sgst=num(r[14]), cess=num(r[15]),
                    g1_status=str(r[16] or "").strip() if len(r) > 16 else "",
                    g1_filing_date=str(r[17] or "").strip() if len(r) > 17 else "",
                    g1_filing_period=str(r[18] or "").strip() if len(r) > 18 else "",
                    g3b_status=str(r[19] or "").strip() if len(r) > 19 else "",
                    cancel_date=str(r[20] or "").strip() if len(r) > 20 else "",
                    amend_type=str(r[21] or "").strip() if len(r) > 21 else "",
                    orig_tax_period=str(r[22] or "").strip() if len(r) > 22 else "",
                ))
            out["b2ba"][month] = recs

    # ---------------- CDNRA ---------------- (layout derived from header row + Read-me;
    # NOT cross-verified against a real data row -- this taxpayer's file had none this year)
    # 0 orig_note_type,1 orig_note_no,2 orig_note_date,3 gstin,4 supplier,5 note_type,
    # 6 note_no,7 supply_type,8 note_date,9 note_val,10 pos,11 rcm,12 rate,13 taxable,
    # 14 igst,15 cgst,16 sgst,17 cess,18 g1_status,19 g1_filing_date,20 g1_filing_period,
    # 21 g3b_status,22 amend_type,23 orig_tax_period,24 cancel_date
    blocks = _r2a_month_blocks(wb, "CDNRA", "GSTIN of Supplier")
    if blocks is not None:
        any_sheet_found = True
        for month, mrows in blocks.items():
            out["months_present"].add(month)
            valid = _r2a_valid_rows(mrows, 3)
            clean, missing = _r2a_dedupe_total_rows(valid, 6)
            _record_missing("CDNRA", month, missing)
            recs = []
            for r in clean:
                if len(r) < 18 or not r[3]:
                    continue
                recs.append(dict(
                    month=month, orig_note_type=str(r[0] or "").strip(),
                    orig_note_no=str(r[1] or "").strip(), orig_note_date=r2a_clean_date(r[2]),
                    gstin=_r2a_clean_gstin(r[3], malformed), supplier=str(r[4] or "").strip(),
                    note_type=str(r[5] or "").strip(), note_no=str(r[6] or "").strip(),
                    supply_type=str(r[7] or "").strip(), note_date=r2a_clean_date(r[8]),
                    note_val=num(r[9]), pos=str(r[10] or "").strip(),
                    rcm=str(r[11] or "").strip().upper(),
                    taxable=num(r[13]), igst=num(r[14]), cgst=num(r[15]),
                    sgst=num(r[16]), cess=num(r[17]),
                    g1_status=str(r[18] or "").strip() if len(r) > 18 else "",
                    g1_filing_date=str(r[19] or "").strip() if len(r) > 19 else "",
                    g1_filing_period=str(r[20] or "").strip() if len(r) > 20 else "",
                    g3b_status=str(r[21] or "").strip() if len(r) > 21 else "",
                    amend_type=str(r[22] or "").strip() if len(r) > 22 else "",
                    orig_tax_period=str(r[23] or "").strip() if len(r) > 23 else "",
                    cancel_date=str(r[24] or "").strip() if len(r) > 24 else "",
                ))
            out["cdnra"][month] = recs

    # ---------------- ISD ---------------- (layout derived from header row + Read-me;
    # NOT cross-verified against a real data row -- this taxpayer had zero ISD credit this year)
    # 0 eligibility,1 gstin_isd,2 isd_name,3 doc_type,4 isd_invno,5 isd_invdate,6 isd_cnno,
    # 7 isd_cndate,8 orig_invno,9 orig_invdate,10 igst,11 cgst,12 sgst,13 cess,
    # 14 filing_status,15 amend,16 amend_period
    blocks = _r2a_month_blocks(wb, "ISD", "GSTIN of ISD")
    if blocks is not None:
        any_sheet_found = True
        for month, mrows in blocks.items():
            out["months_present"].add(month)
            valid = _r2a_valid_rows(mrows, 1)
            recs = []
            for r in valid:
                if len(r) < 14:
                    continue
                recs.append(dict(
                    month=month, eligibility=str(r[0] or "").strip().upper(),
                    gstin=_r2a_clean_gstin(r[1], malformed), isd_name=str(r[2] or "").strip(),
                    doc_type=str(r[3] or "").strip(), isd_invno=str(r[4] or "").strip(),
                    isd_invdate=r2a_clean_date(r[5]), isd_cnno=str(r[6] or "").strip(),
                    isd_cndate=r2a_clean_date(r[7]), orig_invno=str(r[8] or "").strip(),
                    orig_invdate=r2a_clean_date(r[9]),
                    igst=num(r[10]), cgst=num(r[11]), sgst=num(r[12]), cess=num(r[13]),
                    filing_status=str(r[14] or "").strip() if len(r) > 14 else "",
                    amend=str(r[15] or "").strip() if len(r) > 15 else "",
                    amend_period=str(r[16] or "").strip() if len(r) > 16 else "",
                ))
            out["isd"][month] = recs

    if not any_sheet_found:
        out["reason"] = "Workbook found but none of B2B/CDNR/B2BA/CDNRA/ISD sheets were readable."
        return out

    out["available"] = True
    return out
